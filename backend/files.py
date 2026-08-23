import asyncio
from contextlib import contextmanager
from dataclasses import dataclass
import ctypes
import errno
import fcntl
import hashlib
import heapq
import os
import json
import re
import secrets
import shutil
import stat
import sys
import threading
import time
from collections import OrderedDict
from pathlib import Path
from fastapi import (
    APIRouter, Depends, File, Form, HTTPException, Query, UploadFile,
)
from fastapi.responses import FileResponse, HTMLResponse, PlainTextResponse
from pydantic import BaseModel
from .auth import require_token, require_token_query
from .capability_tickets import tickets
from .private_storage import (
    UnsafePrivatePath,
    ensure_private_directory,
    ensure_private_regular_file,
    private_path_kind,
)
from .settings import ROOT, atomic_write_text, env_int
from .workspaces import (
    registry as workspace_registry,
    resolve_workspace_root as _workspace_root,
)

router = APIRouter(prefix="/api/files", tags=["files"])

# ============================================================
# Trash / dustbin — soft delete moves to <ROOT>/.muselab-dustbin/ instead
# of unlink. Restore + permanent-purge are separate endpoints. The dir is
# always excluded from file tree listings, search, and grep (it has its
# own dedicated UI surface in the frontend).
#
# Layout per deletion:
#   <ROOT>/.muselab-dustbin/<trash_id>.json   ← manifest (original path,
#                                                deletion time, kind, size)
#   <ROOT>/.muselab-dustbin/<trash_id>        ← payload (file OR dir, the
#                                                inode rename'd in place)
# trash_id = "<unix_ts>_<8-hex>" — sortable, collision-resistant, opaque
# to the client.
# ============================================================
TRASH_DIR_NAME = ".muselab-dustbin"
INTERNAL_DIR_NAME = ".muselab"

# Serializes every "check destination then rename into place" sequence
# (upload finalize / rename / trash-restore). Each of those does an
# exists() probe followed by a rename() — two concurrent requests against
# the same destination could both pass the probe and the later rename
# silently clobbers the earlier file (TOCTOU). Locks are per workspace so an
# unrelated root never waits behind another root's metadata transaction.
_ROOT_WRITE_LOCKS_GUARD = threading.Lock()
_ROOT_WRITE_LOCKS: dict[str, threading.RLock] = {}


def _root_or_default(root: Path | None) -> Path:
    return ROOT if root is None else root


def _trash_dir(root: Path | None = None) -> Path:
    return _root_or_default(root) / TRASH_DIR_NAME


def _guard_not_trash(target: Path, root: Path | None = None) -> None:
    """Refuse write/upload/rename/copy operations that target the dustbin.

    /delete already blocks the dustbin (writes there have dedicated
    /trash/* endpoints with a different mental model), but the write
    endpoints used to let callers freely create / overwrite / rename
    files inside .muselab-dustbin, corrupting the soft-delete bookkeeping
    (orphan payloads, manifest mismatch). Apply the same guard everywhere
    for consistency."""
    roots = {_root_or_default(root), *workspace_registry.paths()}
    for workspace in roots:
        try:
            trash_root = (workspace / TRASH_DIR_NAME).resolve()
        except (OSError, RuntimeError):
            continue
        if target == trash_root or trash_root in target.parents:
            raise HTTPException(
                status_code=400,
                detail="cannot write inside the dustbin — use /trash/* endpoints",
            )


def _ensure_trash_dir(root: Path | None = None) -> Path:
    d = _trash_dir(root)
    before_kind = private_path_kind(d)
    before_mode = None
    if before_kind == "directory":
        before_mode = stat.S_IMODE(d.lstat().st_mode)
    ensure_private_directory(d)
    if before_kind == "missing" or before_mode != 0o700:
        # Persist the inode and, on first creation, its workspace-root entry.
        _fsync_directory(d)
        if before_kind == "missing":
            _fsync_directory(d.parent)
    return d


_TRASH_ID_RE = re.compile(r"^\d{1,20}_[0-9a-f]{8}$")
_TRASH_SCHEMA_VERSION = 2
_TRASH_STATE_KEY = "transaction_state"
_TRASH_DELETE_PREPARED = "delete_prepared"
_TRASHED = "trashed"
_TRASH_RESTORE_PREPARED = "restore_prepared"
_TRASH_RESTORED = "restored"
_TRASH_PURGED = "purged"
_TRASH_STATES = frozenset({
    _TRASH_DELETE_PREPARED,
    _TRASHED,
    _TRASH_RESTORE_PREPARED,
    _TRASH_RESTORED,
})
_TRASH_LOCK_NAME = ".transaction.lock"
_TRASH_RECEIPT_SUFFIX = ".restored-receipt"
_TRASH_LOCK_CONTEXT = threading.local()


class _RenameDurabilityError(OSError):
    """The rename committed, but one of its parent fsync barriers failed."""


class _TombstoneCleanupError(OSError):
    """The logical delete committed, but physical cleanup is deferred."""


_PURGE_RECEIPT_SUFFIX = ".purged-receipt"
_PURGE_TOMBSTONE_PREFIX = ".purging-"
_PERMANENT_TOMBSTONE_PREFIX = ".permanent-"


@dataclass
class _TrashAnchor:
    """Stable workspace and trash directory handles for one transaction."""

    root: Path
    root_fd: int
    root_identity: dict
    trash: Path
    trash_fd: int
    trash_identity: dict
    lock_fd: int = -1


def _root_write_lock(root: Path) -> threading.RLock:
    key = str(Path(root).absolute())
    with _ROOT_WRITE_LOCKS_GUARD:
        return _ROOT_WRITE_LOCKS.setdefault(key, threading.RLock())


def _stat_identity(info: os.stat_result) -> dict:
    if stat.S_ISREG(info.st_mode):
        kind = "file"
    elif stat.S_ISDIR(info.st_mode):
        kind = "directory"
    elif stat.S_ISLNK(info.st_mode):
        kind = "symlink"
    else:
        kind = "other"
    return {
        "device": int(info.st_dev),
        "inode": int(info.st_ino),
        "kind": kind,
    }


def _fd_identity(fd: int) -> dict:
    return _stat_identity(os.fstat(fd))


def _active_trash_anchors() -> tuple[_TrashAnchor, ...]:
    active = getattr(_TRASH_LOCK_CONTEXT, "active", None)
    if not isinstance(active, dict):
        return ()
    return tuple(
        anchor for anchor in active.values()
        if isinstance(anchor, _TrashAnchor)
    )


def _anchor_for_path(path: Path) -> tuple[_TrashAnchor, tuple[str, ...]] | None:
    candidate = Path(path)
    for anchor in _active_trash_anchors():
        try:
            relative = candidate.relative_to(anchor.root)
        except ValueError:
            continue
        return anchor, tuple(relative.parts)
    return None


def _open_anchor_directory(
    anchor: _TrashAnchor,
    parts: tuple[str, ...],
    *,
    reuse_trash: bool = True,
) -> int:
    """Open a root-relative directory chain without following any symlink."""
    remaining = parts
    if reuse_trash and remaining[:1] == (TRASH_DIR_NAME,):
        current_fd = os.dup(anchor.trash_fd)
        remaining = remaining[1:]
    else:
        current_fd = os.dup(anchor.root_fd)
    try:
        for part in remaining:
            next_fd = os.open(
                part,
                _directory_open_flags(),
                dir_fd=current_fd,
            )
            os.close(current_fd)
            current_fd = next_fd
        return current_fd
    except BaseException:
        os.close(current_fd)
        raise


def _active_path_stat(path: Path) -> tuple[bool, os.stat_result | None]:
    anchored = _anchor_for_path(path)
    if anchored is None:
        return False, None
    anchor, parts = anchored
    if not parts:
        return True, os.fstat(anchor.root_fd)
    try:
        parent_fd = _open_anchor_directory(anchor, parts[:-1])
    except FileNotFoundError:
        return True, None
    try:
        try:
            return True, os.stat(
                parts[-1],
                dir_fd=parent_fd,
                follow_symlinks=False,
            )
        except FileNotFoundError:
            return True, None
    finally:
        os.close(parent_fd)


def _secure_path_kind(path: Path) -> str:
    anchored, info = _active_path_stat(path)
    if not anchored:
        return private_path_kind(path)
    if info is None:
        return "missing"
    return _stat_identity(info)["kind"]


def _open_active_path(path: Path, flags: int) -> int | None:
    anchored = _anchor_for_path(path)
    if anchored is None:
        return None
    anchor, parts = anchored
    if not parts:
        return os.dup(anchor.root_fd)
    parent_fd = _open_anchor_directory(anchor, parts[:-1])
    try:
        return os.open(parts[-1], flags, dir_fd=parent_fd)
    finally:
        os.close(parent_fd)


def _open_workspace_parent(
    anchor: _TrashAnchor,
    relative: str,
    *,
    create: bool = False,
) -> tuple[int, tuple[str, ...], str]:
    logical = _logical_relative_path(relative)
    if INTERNAL_DIR_NAME in logical.parts:
        raise HTTPException(
            status_code=403,
            detail="muselab internal state is not accessible",
        )
    parts = tuple(part for part in logical.parts if part not in {"", "."})
    if not parts:
        raise HTTPException(
            status_code=403,
            detail="workspace roots cannot be deleted or replaced",
        )
    parent_parts = parts[:-1]
    current_fd = os.dup(anchor.root_fd)
    current_path = anchor.root
    try:
        for part in parent_parts:
            try:
                next_fd = os.open(
                    part,
                    _directory_open_flags(),
                    dir_fd=current_fd,
                )
            except FileNotFoundError:
                if not create:
                    raise
                os.mkdir(part, mode=0o777, dir_fd=current_fd)
                next_fd = os.open(
                    part,
                    _directory_open_flags(),
                    dir_fd=current_fd,
                )
                next_path = current_path / part
                _fsync_open_directory(next_fd, next_path)
                _fsync_open_directory(current_fd, current_path)
            else:
                next_path = current_path / part
            os.close(current_fd)
            current_fd = next_fd
            current_path = next_path
        return current_fd, parent_parts, parts[-1]
    except BaseException:
        os.close(current_fd)
        raise


def _directory_is_root_reachable(
    anchor: _TrashAnchor,
    parts: tuple[str, ...],
    held_fd: int,
) -> bool:
    try:
        reachable_fd = _open_anchor_directory(
            anchor,
            parts,
            reuse_trash=False,
        )
    except OSError:
        return False
    try:
        reachable = os.fstat(reachable_fd)
        held = os.fstat(held_fd)
        return (
            int(reachable.st_dev),
            int(reachable.st_ino),
        ) == (
            int(held.st_dev),
            int(held.st_ino),
        )
    finally:
        os.close(reachable_fd)



def _child_directory_reachable(
    parent_fd: int,
    name: str,
    held_fd: int,
) -> bool:
    try:
        reachable_fd = os.open(
            name,
            _directory_open_flags(),
            dir_fd=parent_fd,
        )
    except OSError:
        return False
    try:
        reachable = os.fstat(reachable_fd)
        held = os.fstat(held_fd)
        return (
            int(reachable.st_dev),
            int(reachable.st_ino),
        ) == (
            int(held.st_dev),
            int(held.st_ino),
        )
    finally:
        os.close(reachable_fd)
def _entry_identity_at(parent_fd: int, name: str) -> dict | None:
    try:
        info = os.stat(name, dir_fd=parent_fd, follow_symlinks=False)
    except FileNotFoundError:
        return None
    return _stat_identity(info)


def _entry_kind_at(parent_fd: int, name: str) -> str:
    identity = _entry_identity_at(parent_fd, name)
    return "missing" if identity is None else str(identity["kind"])



def _directory_size_fd(directory_fd: int) -> int:
    total = 0
    with os.scandir(directory_fd) as entries:
        for entry in entries:
            try:
                info = entry.stat(follow_symlinks=False)
                if stat.S_ISREG(info.st_mode):
                    total += int(info.st_size)
                elif stat.S_ISDIR(info.st_mode):
                    child_fd = os.open(
                        entry.name,
                        _directory_open_flags(),
                        dir_fd=directory_fd,
                    )
                    try:
                        total += _directory_size_fd(child_fd)
                    finally:
                        os.close(child_fd)
            except OSError:
                continue
    return total


def _directory_size_at(parent_fd: int, name: str) -> int:
    directory_fd = os.open(
        name,
        _directory_open_flags(),
        dir_fd=parent_fd,
    )
    try:
        return _directory_size_fd(directory_fd)
    finally:
        os.close(directory_fd)
def _restore_mode_at(
    parent_fd: int,
    name: str,
    identity: dict,
    mode: int,
) -> None:
    if not _valid_identity(identity) or not 0 <= mode <= 0o7777:
        raise UnsafePrivatePath("restore mode metadata is invalid")
    flags = os.O_RDONLY | getattr(os, "O_CLOEXEC", 0)
    flags |= getattr(os, "O_NOFOLLOW", 0)
    flags |= getattr(os, "O_NONBLOCK", 0)
    if identity["kind"] == "directory":
        flags |= getattr(os, "O_DIRECTORY", 0)
    try:
        fd = os.open(name, flags, dir_fd=parent_fd)
    except PermissionError:
        private_mode = 0o700 if identity["kind"] == "directory" else 0o600
        if _entry_identity_at(parent_fd, name) != identity:
            raise UnsafePrivatePath(
                "restored inode changed before reopen") from None
        os.chmod(
            name,
            private_mode,
            dir_fd=parent_fd,
            follow_symlinks=False,
        )
        fd = os.open(name, flags, dir_fd=parent_fd)
    try:
        if _stat_identity(os.fstat(fd)) != identity:
            raise UnsafePrivatePath("restored inode changed while opening")
        os.fchmod(fd, mode)
        os.fsync(fd)
    finally:
        os.close(fd)


def _prepare_directory_at(
    parent_fd: int,
    name: str,
    identity: dict,
) -> None:
    if identity.get("kind") != "directory":
        return
    current = os.stat(name, dir_fd=parent_fd, follow_symlinks=False)
    if _stat_identity(current) != identity:
        raise UnsafePrivatePath("trash source changed before chmod")
    if stat.S_IMODE(current.st_mode) != 0o700:
        _restore_mode_at(parent_fd, name, identity, 0o700)


def _open_workspace_root(root: Path) -> tuple[int, dict]:
    root = Path(root)
    expected = _path_identity_unanchored(root)
    if (not _valid_identity(expected)
            or expected["kind"] != "directory"):
        raise UnsafePrivatePath("workspace root is unsafe")
    fd = os.open(root, _directory_open_flags())
    try:
        opened = _fd_identity(fd)
        if opened != expected:
            raise UnsafePrivatePath("workspace root changed while opening")
    except BaseException:
        os.close(fd)
        raise
    return fd, expected


def _open_or_create_trash(
    root: Path,
    root_fd: int,
) -> tuple[int, dict]:
    before = _entry_identity_at(root_fd, TRASH_DIR_NAME)
    created = False
    if before is None:
        try:
            os.mkdir(TRASH_DIR_NAME, mode=0o700, dir_fd=root_fd)
            created = True
        except FileExistsError:
            pass
    elif before.get("kind") != "directory":
        raise UnsafePrivatePath("private storage directory is unsafe")
    trash_fd = os.open(
        TRASH_DIR_NAME,
        _directory_open_flags(),
        dir_fd=root_fd,
    )
    try:
        opened = _fd_identity(trash_fd)
        if opened.get("kind") != "directory":
            raise UnsafePrivatePath("private storage directory is unsafe")
        old_mode = stat.S_IMODE(os.fstat(trash_fd).st_mode)
        if old_mode != 0o700:
            os.fchmod(trash_fd, 0o700)
        if created or old_mode != 0o700:
            _fsync_open_directory(trash_fd, root / TRASH_DIR_NAME)
            _fsync_open_directory(root_fd, root)
    except BaseException:
        os.close(trash_fd)
        raise
    return trash_fd, opened

def _gen_trash_id() -> str:
    return f"{int(time.time())}_{secrets.token_hex(4)}"


def _valid_trash_id(tid: str) -> bool:
    """trash_id format check — `^\\d+_[0-9a-f]{8}$` (matches _gen_trash_id).

    Defense in depth: the trash_id flows from the user back through
    /trash/restore + /trash/purge endpoints, where it's used to build
    paths under <ROOT>/.muselab-dustbin/. Without validation, a payload
    like ``"../../etc/passwd"`` would resolve outside the trash dir and
    trash_purge would rmtree arbitrary directories. The auth token is
    the primary defense, but a strict format check costs nothing and
    blocks the exploit class entirely.
    """
    return bool(tid) and bool(_TRASH_ID_RE.fullmatch(tid))


def _trash_state(data: dict) -> str | None:
    """Return the durable state; legacy manifests are committed trash."""
    state = data.get(_TRASH_STATE_KEY)
    if state is None:
        # Only a versionless v1 record lacks state by design.
        return _TRASHED if data.get("schema_version") is None else None
    return state if state in _TRASH_STATES else None


def _path_identity_unanchored(path: Path) -> dict | None:
    try:
        identity = _stat_identity(Path(path).lstat())
    except FileNotFoundError:
        return None
    return identity if identity["kind"] in {"file", "directory"} else None


def _path_identity(path: Path) -> dict | None:
    """Capture an identity through the active root dirfd when available."""
    anchored, info = _active_path_stat(path)
    if not anchored:
        return _path_identity_unanchored(path)
    if info is None:
        return None
    identity = _stat_identity(info)
    return identity if identity["kind"] in {"file", "directory"} else None


def _valid_identity(identity: object) -> bool:
    if not isinstance(identity, dict):
        return False
    return (
        type(identity.get("device")) is int
        and type(identity.get("inode")) is int
        and identity.get("device", -1) >= 0
        and identity.get("inode", -1) >= 0
        and identity.get("kind") in {"file", "directory"}
    )


def _identity_matches(path: Path, identity: object) -> bool:
    return _valid_identity(identity) and _path_identity(path) == identity


def _fsync_directory(path: Path) -> None:
    """Persist a directory through the active root anchor when available."""
    anchored = _anchor_for_path(path)
    if anchored is not None:
        anchor, parts = anchored
        fd = _open_anchor_directory(anchor, parts)
        try:
            _fsync_open_directory(fd, Path(path))
        finally:
            os.close(fd)
        return
    flags = os.O_RDONLY | getattr(os, "O_CLOEXEC", 0)
    flags |= getattr(os, "O_DIRECTORY", 0)
    flags |= getattr(os, "O_NOFOLLOW", 0)
    fd = os.open(Path(path), flags)
    try:
        os.fsync(fd)
    finally:
        os.close(fd)


def _fsync_open_directory(fd: int, path: Path) -> None:
    """Fsync an already-anchored directory fd; path is for fault attribution."""
    del path
    os.fsync(fd)


def _fsync_path(path: Path) -> None:
    """Persist content/metadata without following a final symlink."""
    kind = _secure_path_kind(path)
    if kind not in {"directory", "file"}:
        raise UnsafePrivatePath("trash payload is unsafe")
    identity = _path_identity(path)
    if identity is None:
        raise UnsafePrivatePath("trash payload identity is unavailable")
    flags = os.O_RDONLY | getattr(os, "O_CLOEXEC", 0)
    flags |= getattr(os, "O_NOFOLLOW", 0)
    flags |= getattr(os, "O_NONBLOCK", 0)
    if kind == "directory":
        flags |= getattr(os, "O_DIRECTORY", 0)
    active_fd = _open_active_path(path, flags)
    fd = os.open(path, flags) if active_fd is None else active_fd
    try:
        info = os.fstat(fd)
        opened_mode = info.st_mode
        if ((kind == "file" and not stat.S_ISREG(opened_mode))
                or (kind == "directory" and not stat.S_ISDIR(opened_mode))):
            raise UnsafePrivatePath("trash payload changed while opening")
        opened_identity = {
            "device": int(info.st_dev),
            "inode": int(info.st_ino),
            "kind": kind,
        }
        if opened_identity != identity:
            raise UnsafePrivatePath("trash payload identity changed while opening")
        os.fsync(fd)
    finally:
        os.close(fd)


def _restore_original_mode(path: Path, identity: dict, mode: int) -> None:
    """Restore mode through an identity-checked fd, including unreadable modes."""
    anchored = _anchor_for_path(path)
    if anchored is not None:
        anchor, parts = anchored
        if not parts:
            raise UnsafePrivatePath("workspace root mode cannot be changed")
        parent_fd = _open_anchor_directory(anchor, parts[:-1])
        try:
            _restore_mode_at(parent_fd, parts[-1], identity, mode)
        finally:
            os.close(parent_fd)
        return
    if not _valid_identity(identity) or not 0 <= mode <= 0o7777:
        raise UnsafePrivatePath("restore mode metadata is invalid")
    if not _identity_matches(path, identity):
        raise UnsafePrivatePath("restored inode changed before chmod")
    flags = os.O_RDONLY | getattr(os, "O_CLOEXEC", 0)
    flags |= getattr(os, "O_NOFOLLOW", 0)
    flags |= getattr(os, "O_NONBLOCK", 0)
    if identity["kind"] == "directory":
        flags |= getattr(os, "O_DIRECTORY", 0)
    try:
        fd = os.open(path, flags)
    except PermissionError:
        # A previous process may have completed fchmod(000/0200) and crashed
        # before terminal-state persistence. Temporarily regain owner access,
        # then anchor and revalidate the inode before restoring the exact mode.
        private_mode = 0o700 if identity["kind"] == "directory" else 0o600
        if not _identity_matches(path, identity):
            raise UnsafePrivatePath("restored inode changed before reopen") from None
        os.chmod(path, private_mode, follow_symlinks=False)
        fd = os.open(path, flags)
    try:
        info = os.fstat(fd)
        opened_kind = (
            "file" if stat.S_ISREG(info.st_mode)
            else "directory" if stat.S_ISDIR(info.st_mode)
            else "other"
        )
        opened_identity = {
            "device": int(info.st_dev),
            "inode": int(info.st_ino),
            "kind": opened_kind,
        }
        if opened_identity != identity:
            raise UnsafePrivatePath("restored inode changed while opening")
        os.fchmod(fd, mode)
        os.fsync(fd)
    finally:
        os.close(fd)


def _prepare_directory_for_rename(path: Path, identity: dict) -> None:
    """Make a directory renameable after its durable intent is published.

    Linux requires write permission on a moved directory because rename must
    update its ``..`` entry. A mode-000 directory therefore needs a temporary
    private mode. The caller's WAL contains the original mode before this
    helper runs, so recovery can roll this metadata change back if the rename
    never commits.
    """
    anchored = _anchor_for_path(path)
    if anchored is not None:
        anchor, parts = anchored
        if not parts:
            raise UnsafePrivatePath("workspace root cannot be renamed")
        parent_fd = _open_anchor_directory(anchor, parts[:-1])
        try:
            _prepare_directory_at(parent_fd, parts[-1], identity)
        finally:
            os.close(parent_fd)
        return
    if identity.get("kind") != "directory":
        return
    if not _identity_matches(path, identity):
        raise UnsafePrivatePath("trash source changed before chmod")
    current_mode = stat.S_IMODE(path.lstat().st_mode)
    if current_mode == 0o700:
        return
    _restore_original_mode(path, identity, 0o700)


@contextmanager
def _trash_transaction(root: Path | None = None):
    """Hold stable root/trash dirfds plus a per-root process/flock lock."""
    root_path = Path(_root_or_default(root)).absolute()
    key = str(root_path)
    with _root_write_lock(root_path):
        active = getattr(_TRASH_LOCK_CONTEXT, "active", None)
        if active is None:
            active = {}
            _TRASH_LOCK_CONTEXT.active = active
        nested = active.get(key)
        if isinstance(nested, _TrashAnchor):
            yield nested.trash
            return

        root_fd = -1
        trash_fd = -1
        lock_fd = -1
        locked = False
        anchor: _TrashAnchor | None = None
        try:
            root_fd, root_identity = _open_workspace_root(root_path)
            trash_fd, trash_identity = _open_or_create_trash(
                root_path,
                root_fd,
            )
            before_lock = _entry_identity_at(trash_fd, _TRASH_LOCK_NAME)
            flags = os.O_RDWR | os.O_CREAT | getattr(os, "O_CLOEXEC", 0)
            flags |= getattr(os, "O_NOFOLLOW", 0)
            flags |= getattr(os, "O_NONBLOCK", 0)
            lock_fd = os.open(
                _TRASH_LOCK_NAME,
                flags,
                0o600,
                dir_fd=trash_fd,
            )
            lock_info = os.fstat(lock_fd)
            if not stat.S_ISREG(lock_info.st_mode):
                raise UnsafePrivatePath("trash transaction lock is unsafe")
            old_mode = stat.S_IMODE(lock_info.st_mode)
            if old_mode != 0o600:
                os.fchmod(lock_fd, 0o600)
            if before_lock is None or old_mode != 0o600:
                os.fsync(lock_fd)
                _fsync_open_directory(
                    trash_fd,
                    root_path / TRASH_DIR_NAME,
                )
            fcntl.flock(lock_fd, fcntl.LOCK_EX)
            locked = True
            anchor = _TrashAnchor(
                root=root_path,
                root_fd=root_fd,
                root_identity=root_identity,
                trash=root_path / TRASH_DIR_NAME,
                trash_fd=trash_fd,
                trash_identity=trash_identity,
                lock_fd=lock_fd,
            )
            active[key] = anchor
            yield anchor.trash
        finally:
            if anchor is not None and active.get(key) is anchor:
                active.pop(key, None)
            if locked:
                fcntl.flock(lock_fd, fcntl.LOCK_UN)
            if lock_fd >= 0:
                os.close(lock_fd)
            if trash_fd >= 0:
                os.close(trash_fd)
            if root_fd >= 0:
                os.close(root_fd)


def _fsync_rename(source_parent: Path, destination_parent: Path) -> None:
    """Persist both sides of a rename; one fsync is enough for one parent."""
    seen: set[Path] = set()
    for parent in (Path(source_parent), Path(destination_parent)):
        if parent in seen:
            continue
        seen.add(parent)
        _fsync_directory(parent)


def _directory_open_flags() -> int:
    flags = os.O_RDONLY | getattr(os, "O_CLOEXEC", 0)
    flags |= getattr(os, "O_DIRECTORY", 0)
    flags |= getattr(os, "O_NOFOLLOW", 0)
    flags |= getattr(os, "O_NONBLOCK", 0)
    return flags


def _directory_creation_anchor(path: Path) -> tuple[Path, dict]:
    """Capture the nearest existing real directory before any mkdir occurs."""
    cursor = Path(path)
    while _secure_path_kind(cursor) == "missing":
        if cursor == cursor.parent:
            break
        cursor = cursor.parent
    identity = _path_identity(cursor)
    if not _valid_identity(identity) or identity["kind"] != "directory":
        raise UnsafePrivatePath("directory creation anchor is unsafe")
    return cursor, identity


def _open_directory_anchor(path: Path, identity: dict | None = None) -> int:
    """Open one parent directory and prove it is the expected inode."""
    expected = identity if identity is not None else _path_identity(path)
    if not _valid_identity(expected) or expected["kind"] != "directory":
        raise UnsafePrivatePath("rename parent identity is invalid")
    flags = _directory_open_flags()
    anchored = _anchor_for_path(path)
    if anchored is None:
        fd = os.open(path, flags)
    else:
        anchor, parts = anchored
        fd = _open_anchor_directory(anchor, parts)
    try:
        info = os.fstat(fd)
        opened = {
            "device": int(info.st_dev),
            "inode": int(info.st_ino),
            "kind": "directory" if stat.S_ISDIR(info.st_mode) else "other",
        }
        if opened != expected:
            raise UnsafePrivatePath("rename parent changed while opening")
    except BaseException:
        os.close(fd)
        raise
    return fd


def _mkdir_durable(
    path: Path,
    *,
    anchor_path: Path | None = None,
    anchor_identity: dict | None = None,
) -> dict:
    """Create a directory chain via mkdirat and fsync every new entry."""
    path = Path(path)
    if anchor_path is None or anchor_identity is None:
        anchor_path, anchor_identity = _directory_creation_anchor(path)
    anchor_path = Path(anchor_path)
    if not _valid_identity(anchor_identity):
        raise UnsafePrivatePath("directory creation anchor is invalid")
    try:
        relative = path.relative_to(anchor_path)
    except ValueError:
        raise UnsafePrivatePath(
            "directory target escaped its creation anchor") from None

    current_path = anchor_path
    current_fd = _open_directory_anchor(anchor_path, anchor_identity)
    try:
        for part in relative.parts:
            created = False
            try:
                next_fd = os.open(
                    part, _directory_open_flags(), dir_fd=current_fd)
            except FileNotFoundError:
                try:
                    os.mkdir(part, mode=0o777, dir_fd=current_fd)
                    created = True
                except FileExistsError:
                    pass
                next_fd = os.open(
                    part, _directory_open_flags(), dir_fd=current_fd)
            next_path = current_path / part
            try:
                if created:
                    _fsync_open_directory(next_fd, next_path)
                    _fsync_open_directory(current_fd, current_path)
            except BaseException:
                os.close(next_fd)
                raise
            os.close(current_fd)
            current_fd = next_fd
            current_path = next_path
        info = os.fstat(current_fd)
        return {
            "device": int(info.st_dev),
            "inode": int(info.st_ino),
            "kind": "directory",
        }
    finally:
        os.close(current_fd)


def _fsync_rename_fds(
    source_fd: int,
    source_parent: Path,
    destination_fd: int,
    destination_parent: Path,
) -> None:
    """Persist a rename using the exact parent inodes used by the syscall."""
    _fsync_open_directory(source_fd, source_parent)
    source_info = os.fstat(source_fd)
    destination_info = os.fstat(destination_fd)
    if (source_info.st_dev, source_info.st_ino) != (
            destination_info.st_dev, destination_info.st_ino):
        _fsync_open_directory(destination_fd, destination_parent)


def _fsync_anchored_parents(
    source_parent: Path,
    destination_parent: Path,
    *,
    source_parent_identity: dict,
    destination_parent_identity: dict,
) -> None:
    """Replay both rename barriers against previously captured parent inodes."""
    source_fd = _open_directory_anchor(
        source_parent, source_parent_identity)
    destination_fd = -1
    try:
        destination_fd = _open_directory_anchor(
            destination_parent, destination_parent_identity)
        _fsync_rename_fds(
            source_fd,
            source_parent,
            destination_fd,
            destination_parent,
        )
    finally:
        if destination_fd >= 0:
            os.close(destination_fd)
        os.close(source_fd)


def _rename_noreplace(
    source: Path,
    destination: Path,
    *,
    source_parent_identity: dict | None = None,
    destination_parent_identity: dict | None = None,
    source_parent_fd: int | None = None,
    destination_parent_fd: int | None = None,
) -> None:
    """Rename without replacement through identity-checked parent dirfds."""
    source = Path(source)
    destination = Path(destination)
    own_source_fd = source_parent_fd is None
    own_destination_fd = destination_parent_fd is None
    source_fd = (
        _open_directory_anchor(source.parent, source_parent_identity)
        if source_parent_fd is None
        else source_parent_fd
    )
    destination_fd = (
        -1 if destination_parent_fd is None else destination_parent_fd
    )
    try:
        if destination_parent_fd is None:
            destination_fd = _open_directory_anchor(
                destination.parent, destination_parent_identity)
        if (source_parent_identity is not None
                and _fd_identity(source_fd) != source_parent_identity):
            raise UnsafePrivatePath("rename source parent changed")
        if (destination_parent_identity is not None
                and _fd_identity(destination_fd)
                != destination_parent_identity):
            raise UnsafePrivatePath("rename destination parent changed")
        source_name = os.fsencode(source.name)
        destination_name = os.fsencode(destination.name)
        if sys.platform.startswith("linux"):
            libc = ctypes.CDLL(None, use_errno=True)
            try:
                renameat2 = libc.renameat2
            except AttributeError:
                raise OSError(
                    errno.ENOTSUP,
                    "atomic no-replace rename is unavailable",
                    str(destination),
                ) from None
            renameat2.argtypes = (
                ctypes.c_int,
                ctypes.c_char_p,
                ctypes.c_int,
                ctypes.c_char_p,
                ctypes.c_uint,
            )
            renameat2.restype = ctypes.c_int
            ctypes.set_errno(0)
            result = renameat2(
                source_fd, source_name,
                destination_fd, destination_name,
                1,
            )
        elif sys.platform == "darwin":
            libc = ctypes.CDLL(None, use_errno=True)
            try:
                renameatx_np = libc.renameatx_np
            except AttributeError:
                raise OSError(
                    errno.ENOTSUP,
                    "atomic no-replace rename is unavailable",
                    str(destination),
                ) from None
            renameatx_np.argtypes = (
                ctypes.c_int,
                ctypes.c_char_p,
                ctypes.c_int,
                ctypes.c_char_p,
                ctypes.c_uint,
            )
            renameatx_np.restype = ctypes.c_int
            ctypes.set_errno(0)
            result = renameatx_np(
                source_fd, source_name,
                destination_fd, destination_name,
                0x4,
            )
        else:
            raise OSError(
                errno.ENOTSUP,
                "atomic no-replace rename is unavailable",
                str(destination),
            )
        if result != 0:
            error = ctypes.get_errno() or errno.EIO
            raise OSError(error, os.strerror(error), str(destination))
        try:
            _fsync_rename_fds(
                source_fd,
                source.parent,
                destination_fd,
                destination.parent,
            )
        except OSError as exc:
            raise _RenameDurabilityError(
                f"rename committed but parent fsync failed: {exc}"
            ) from exc
    finally:
        if own_destination_fd and destination_fd >= 0:
            os.close(destination_fd)
        if own_source_fd:
            os.close(source_fd)


def _active_trash_child(path: Path) -> tuple[_TrashAnchor, str] | None:
    anchored = _anchor_for_path(path)
    if anchored is None:
        return None
    anchor, parts = anchored
    if len(parts) != 2 or parts[0] != TRASH_DIR_NAME:
        return None
    return anchor, parts[1]


def _unlink_trash_manifest(path: Path, *, missing_ok: bool = False) -> bool:
    kind = _secure_path_kind(path)
    if kind == "missing" and missing_ok:
        return False
    if kind != "file":
        raise UnsafePrivatePath("trash manifest is unsafe")
    active_child = _active_trash_child(path)
    if active_child is None:
        path.unlink()
        _fsync_directory(path.parent)
        return True
    anchor, name = active_child
    os.unlink(name, dir_fd=anchor.trash_fd)
    _fsync_open_directory(anchor.trash_fd, anchor.trash)
    return True


def _manifest_json(data: dict) -> str:
    return json.dumps(data, ensure_ascii=False, separators=(",", ":"))


def _write_manifest_fd(fd: int, data: dict) -> None:
    """Write, chmod and fsync one already-exclusive manifest inode."""
    os.fchmod(fd, 0o600)
    with os.fdopen(fd, "w", encoding="utf-8", closefd=False) as stream:
        stream.write(_manifest_json(data))
        stream.flush()
        os.fsync(stream.fileno())


def _manifest_open_flags() -> int:
    flags = os.O_WRONLY | os.O_CREAT | os.O_EXCL
    flags |= getattr(os, "O_CLOEXEC", 0)
    flags |= getattr(os, "O_NOFOLLOW", 0)
    return flags

def _create_manifest_at(
    anchor: _TrashAnchor,
    name: str,
    data: dict,
) -> None:
    fd = -1
    created_identity: dict | None = None
    try:
        fd = os.open(
            name,
            _manifest_open_flags(),
            0o600,
            dir_fd=anchor.trash_fd,
        )
        created_identity = _stat_identity(os.fstat(fd))
        _write_manifest_fd(fd, data)
        os.close(fd)
        fd = -1
        _fsync_open_directory(anchor.trash_fd, anchor.trash)
    except BaseException:
        if fd >= 0:
            os.close(fd)
        try:
            if (_entry_identity_at(anchor.trash_fd, name)
                    == created_identity):
                os.unlink(name, dir_fd=anchor.trash_fd)
                _fsync_open_directory(anchor.trash_fd, anchor.trash)
        except OSError:
            pass
        raise


def _write_manifest_at(
    anchor: _TrashAnchor,
    name: str,
    data: dict,
) -> None:
    if _entry_kind_at(anchor.trash_fd, name) != "file":
        raise FileNotFoundError(anchor.trash / name)
    tmp_name = (
        f".{name}.txn.{os.getpid()}.{secrets.token_hex(4)}"
    )
    fd = -1
    try:
        fd = os.open(
            tmp_name,
            _manifest_open_flags(),
            0o600,
            dir_fd=anchor.trash_fd,
        )
        _write_manifest_fd(fd, data)
        os.close(fd)
        fd = -1
        if _entry_kind_at(anchor.trash_fd, name) != "file":
            raise UnsafePrivatePath(
                "trash manifest changed during transition")
        os.replace(
            tmp_name,
            name,
            src_dir_fd=anchor.trash_fd,
            dst_dir_fd=anchor.trash_fd,
        )
        _fsync_open_directory(anchor.trash_fd, anchor.trash)
    finally:
        if fd >= 0:
            os.close(fd)
        if _entry_kind_at(anchor.trash_fd, tmp_name) == "file":
            try:
                os.unlink(tmp_name, dir_fd=anchor.trash_fd)
            except OSError:
                pass


def _read_manifest_at(
    anchor: _TrashAnchor,
    name: str,
    *,
    strict_io: bool = False,
) -> dict | None:
    flags = os.O_RDONLY | getattr(os, "O_CLOEXEC", 0)
    flags |= getattr(os, "O_NOFOLLOW", 0)
    flags |= getattr(os, "O_NONBLOCK", 0)
    fd = -1
    try:
        fd = os.open(name, flags, dir_fd=anchor.trash_fd)
        if not stat.S_ISREG(os.fstat(fd).st_mode):
            return None
        os.fchmod(fd, 0o600)
        with os.fdopen(fd, "r", encoding="utf-8") as stream:
            fd = -1
            data = json.load(stream)
    except json.JSONDecodeError:
        return None
    except OSError:
        if strict_io:
            raise
        return None
    finally:
        if fd >= 0:
            os.close(fd)
    return data if isinstance(data, dict) else None


def _create_trash_manifest(path: Path, data: dict) -> None:
    """Create and fsync a 0600 prepare record without replacing anything."""
    active_child = _active_trash_child(path)
    if active_child is not None:
        _create_manifest_at(*active_child, data)
        return
    ensure_private_directory(path.parent)
    fd = -1
    created_identity: dict | None = None
    try:
        fd = os.open(path, _manifest_open_flags(), 0o600)
        info = os.fstat(fd)
        created_identity = {
            "device": int(info.st_dev),
            "inode": int(info.st_ino),
            "kind": "file",
        }
        _write_manifest_fd(fd, data)
        os.close(fd)
        fd = -1
        _fsync_directory(path.parent)
    except BaseException:
        if fd >= 0:
            os.close(fd)
        # O_EXCL can fail because another transaction already owns this name.
        # Only remove the inode this call actually created; never unlink a
        # collided manifest or a path swapped in by an external writer.
        try:
            if _identity_matches(path, created_identity):
                _unlink_trash_manifest(path, missing_ok=True)
        except (OSError, UnsafePrivatePath):
            pass
        raise


def _write_trash_manifest(path: Path, data: dict) -> None:
    """Atomically and durably transition an existing transaction record."""
    active_child = _active_trash_child(path)
    if active_child is not None:
        _write_manifest_at(*active_child, data)
        return
    if not ensure_private_regular_file(path):
        raise FileNotFoundError(path)
    tmp = path.with_name(
        f".{path.name}.txn.{os.getpid()}.{secrets.token_hex(4)}")
    fd = -1
    try:
        fd = os.open(tmp, _manifest_open_flags(), 0o600)
        _write_manifest_fd(fd, data)
        os.close(fd)
        fd = -1
        if private_path_kind(path) != "file":
            raise UnsafePrivatePath("trash manifest changed during transition")
        os.replace(tmp, path)
        _fsync_directory(path.parent)
    finally:
        if fd >= 0:
            os.close(fd)
        if private_path_kind(tmp) == "file":
            try:
                tmp.unlink()
            except OSError:
                pass


def _read_private_manifest_file(
    path: Path,
    *,
    strict_io: bool = False,
) -> dict | None:
    """Read one real 0600 manifest without following a symlink."""
    active_child = _active_trash_child(path)
    if active_child is not None:
        return _read_manifest_at(
            *active_child,
            strict_io=strict_io,
        )
    flags = os.O_RDONLY | getattr(os, "O_CLOEXEC", 0)
    flags |= getattr(os, "O_NOFOLLOW", 0)
    flags |= getattr(os, "O_NONBLOCK", 0)
    fd = -1
    try:
        fd = os.open(path, flags)
        if not stat.S_ISREG(os.fstat(fd).st_mode):
            return None
        os.fchmod(fd, 0o600)
        with os.fdopen(fd, "r", encoding="utf-8") as stream:
            fd = -1
            data = json.load(stream)
    except (UnsafePrivatePath, json.JSONDecodeError):
        return None
    except OSError:
        if strict_io:
            raise
        return None
    finally:
        if fd >= 0:
            os.close(fd)
    return data if isinstance(data, dict) else None


def _restore_receipt_path(d: Path, tid: str) -> Path:
    return d / f"{tid}{_TRASH_RECEIPT_SUFFIX}"

def _purge_receipt_path(d: Path, tid: str) -> Path:
    return d / f"{tid}{_PURGE_RECEIPT_SUFFIX}"


def _purge_tombstone_name(tid: str) -> str:
    return f"{_PURGE_TOMBSTONE_PREFIX}{tid}"


def _permanent_tombstone_name() -> str:
    return (
        f"{_PERMANENT_TOMBSTONE_PREFIX}"
        f"{int(time.time())}-{secrets.token_hex(8)}"
    )


def _same_restore_transaction(left: dict, right: dict) -> bool:
    """Match immutable transaction identity, never just a reused trash id."""
    left_nonce = left.get("transaction_nonce")
    right_nonce = right.get("transaction_nonce")
    if (isinstance(left_nonce, str) and left_nonce
            and isinstance(right_nonce, str) and right_nonce):
        return secrets.compare_digest(left_nonce, right_nonce)
    return all(left.get(field) == right.get(field) for field in (
        "trash_id", "original_path", "payload_identity", "deleted_at",
    ))


def _read_restore_receipt(d: Path, tid: str) -> dict | None:
    data = _read_private_manifest_file(_restore_receipt_path(d, tid))
    if data is None:
        return None
    if (
        data.get("schema_version") != _TRASH_SCHEMA_VERSION
        or data.get(_TRASH_STATE_KEY) != _TRASH_RESTORED
        or data.get("trash_id") != tid
        or not _valid_identity(data.get("payload_identity"))
        or not isinstance(data.get("original_path"), str)
        or not data["original_path"]
    ):
        return None
    # Visibility after a failed directory fsync is not durability. Replaying
    # both barriers makes an existing receipt safe to use as a 200 response.
    _fsync_path(_restore_receipt_path(d, tid))
    _fsync_directory(d)
    return data



def _read_purge_receipt(d: Path, tid: str) -> dict | None:
    receipt_path = _purge_receipt_path(d, tid)
    data = _read_private_manifest_file(receipt_path)
    if data is None:
        return None
    if (
        data.get("schema_version") != _TRASH_SCHEMA_VERSION
        or data.get(_TRASH_STATE_KEY) != _TRASH_PURGED
        or data.get("trash_id") != tid
    ):
        return None
    _fsync_path(receipt_path)
    _fsync_directory(d)
    return data

def _finalize_restore_manifest(manifest_path: Path, data: dict) -> None:
    """Durably preserve completion before removing the active manifest."""
    tid = str(data.get("trash_id") or "")
    if not _valid_trash_id(tid) or manifest_path.name != f"{tid}.json":
        raise UnsafePrivatePath("restore manifest identity is invalid")
    restored = {
        **data,
        "schema_version": _TRASH_SCHEMA_VERSION,
        _TRASH_STATE_KEY: _TRASH_RESTORED,
        "restored_at": data.get("restored_at") or time.time(),
    }
    receipt_path = _restore_receipt_path(manifest_path.parent, tid)
    receipt_kind = _secure_path_kind(receipt_path)
    if receipt_kind == "missing":
        _create_trash_manifest(receipt_path, restored)
    elif receipt_kind == "file":
        receipt = _read_restore_receipt(manifest_path.parent, tid)
        if (receipt is None
                or not _same_restore_transaction(receipt, restored)):
            raise UnsafePrivatePath("restore receipt is invalid")
    else:
        raise UnsafePrivatePath("restore receipt is unsafe")
    # If unlink succeeds but directory fsync fails, the durable receipt still
    # makes both immediate and post-crash retries idempotent.
    _unlink_trash_manifest(manifest_path, missing_ok=True)


def _persist_restore_completion(manifest_path: Path, data: dict) -> dict:
    """Publish a durable terminal fact before restore may return success."""
    restored = {
        **data,
        "schema_version": _TRASH_SCHEMA_VERSION,
        _TRASH_STATE_KEY: _TRASH_RESTORED,
        "restored_at": data.get("restored_at") or time.time(),
    }
    terminal_durable = False
    if _trash_state(data) == _TRASH_RESTORED:
        try:
            _fsync_path(manifest_path)
            _fsync_directory(manifest_path.parent)
            terminal_durable = True
        except (OSError, UnsafePrivatePath):
            pass
    if not terminal_durable:
        try:
            _write_trash_manifest(manifest_path, restored)
            terminal_durable = True
        except (OSError, UnsafePrivatePath) as exc:
            sys.stderr.write(
                f"[files] restore terminal state deferred "
                f"tid={restored.get('trash_id')} ({type(exc).__name__})\n"
            )
    try:
        _finalize_restore_manifest(manifest_path, restored)
        return restored
    except (OSError, UnsafePrivatePath):
        receipt = _read_restore_receipt(
            manifest_path.parent, str(restored.get("trash_id") or ""))
        if (receipt is not None
                and _same_restore_transaction(receipt, restored)):
            return restored
        if terminal_durable:
            return restored
        raise


def _open_manifest_original_parent(
    anchor: _TrashAnchor,
    data: dict,
) -> tuple[int, tuple[str, ...], str, Path] | None:
    """Open a manifest destination only through the held workspace root."""
    original_rel = data.get("original_path")
    if not isinstance(original_rel, str) or not original_rel:
        return None
    try:
        logical = _logical_relative_path(original_rel)
    except HTTPException:
        return None
    if (
        INTERNAL_DIR_NAME in logical.parts
        or TRASH_DIR_NAME in logical.parts
    ):
        return None
    try:
        parent_fd, parent_parts, name = _open_workspace_parent(
            anchor,
            logical.as_posix(),
        )
    except (OSError, HTTPException):
        return None
    return parent_fd, parent_parts, name, logical


def _repair_trash_item(
    d: Path,
    manifest_path: Path,
    *,
    strict_io: bool = False,
) -> dict | None:
    """Reconcile one durable transaction through stable root/trash dirfds."""
    anchored = _anchor_for_path(d)
    if anchored is None:
        raise UnsafePrivatePath("trash transaction anchor is unavailable")
    anchor = anchored[0]
    data = _read_private_manifest_file(
        manifest_path,
        strict_io=strict_io,
    )
    if data is None:
        return None
    tid = str(data.get("trash_id") or "")
    if not _valid_trash_id(tid) or manifest_path.name != f"{tid}.json":
        return None
    if data.get("schema_version") not in {None, _TRASH_SCHEMA_VERSION}:
        return None
    state = _trash_state(data)
    if state is None:
        return None

    payload_kind = _entry_kind_at(anchor.trash_fd, tid)
    identity = data.get("payload_identity")
    if payload_kind in {"directory", "file"}:
        if state == _TRASH_RESTORED:
            return None
        payload_info = os.stat(
            tid,
            dir_fd=anchor.trash_fd,
            follow_symlinks=False,
        )
        captured_identity = _stat_identity(payload_info)
        if (
            _valid_identity(identity)
            and captured_identity != identity
        ):
            return None
        original_mode = data.get("original_mode")
        if (
            type(original_mode) is not int
            or not 0 <= original_mode <= 0o7777
        ):
            original_mode = stat.S_IMODE(payload_info.st_mode)
        transaction_nonce = data.get("transaction_nonce")
        if not isinstance(transaction_nonce, str) or not transaction_nonce:
            transaction_nonce = secrets.token_hex(16)
        repaired = {
            **data,
            "schema_version": _TRASH_SCHEMA_VERSION,
            _TRASH_STATE_KEY: state,
            "payload_identity": captured_identity,
            "original_mode": original_mode,
            "transaction_nonce": transaction_nonce,
        }
        # Publish legacy migration/original mode before tightening permissions.
        if repaired != data:
            try:
                _write_trash_manifest(manifest_path, repaired)
            except (OSError, UnsafePrivatePath):
                return data

        desired_mode = (
            0o700 if payload_kind == "directory" else 0o600
        )
        current_mode = stat.S_IMODE(payload_info.st_mode)
        if (
            state == _TRASH_DELETE_PREPARED
            or current_mode != desired_mode
        ):
            try:
                _restore_mode_at(
                    anchor.trash_fd,
                    tid,
                    captured_identity,
                    desired_mode,
                )
            except (OSError, UnsafePrivatePath):
                return repaired

        if state == _TRASH_DELETE_PREPARED:
            opened = _open_manifest_original_parent(anchor, repaired)
            if opened is None:
                return repaired
            parent_fd, parent_parts, _name, logical = opened
            try:
                source_parent_identity = repaired.get(
                    "source_parent_identity"
                )
                if not _valid_identity(source_parent_identity):
                    source_parent_identity = _fd_identity(parent_fd)
                trash_parent_identity = repaired.get(
                    "trash_parent_identity"
                )
                if not _valid_identity(trash_parent_identity):
                    trash_parent_identity = _fd_identity(anchor.trash_fd)
                if (
                    _fd_identity(parent_fd) != source_parent_identity
                    or _fd_identity(anchor.trash_fd)
                    != trash_parent_identity
                    or not _directory_is_root_reachable(
                        anchor,
                        parent_parts,
                        parent_fd,
                    )
                    or not _directory_is_root_reachable(
                        anchor,
                        (TRASH_DIR_NAME,),
                        anchor.trash_fd,
                    )
                ):
                    return repaired
                try:
                    _fsync_rename_fds(
                        parent_fd,
                        anchor.root / logical.parent,
                        anchor.trash_fd,
                        d,
                    )
                except OSError:
                    return repaired
            finally:
                os.close(parent_fd)

            committed = {
                **repaired,
                _TRASH_STATE_KEY: _TRASHED,
            }
            try:
                _write_trash_manifest(manifest_path, committed)
            except (OSError, UnsafePrivatePath):
                return repaired
            return committed
        return repaired

    if payload_kind != "missing" or not _valid_identity(identity):
        return None

    # RESTORED is terminal even if users later edit or remove the target.
    if state == _TRASH_RESTORED:
        return data

    opened = _open_manifest_original_parent(anchor, data)
    if state == _TRASH_DELETE_PREPARED:
        # A crash before rename leaves the source inode in its held-root path.
        if opened is None:
            return None
        parent_fd, parent_parts, name, _logical = opened
        try:
            source_parent_identity = data.get(
                "source_parent_identity"
            )
            if (
                _valid_identity(source_parent_identity)
                and _fd_identity(parent_fd) != source_parent_identity
            ):
                return None
            if not _directory_is_root_reachable(
                anchor,
                parent_parts,
                parent_fd,
            ):
                return None
            if _entry_identity_at(parent_fd, name) != identity:
                return None
            try:
                original_mode = data.get("original_mode")
                if (
                    type(original_mode) is int
                    and 0 <= original_mode <= 0o7777
                ):
                    _restore_mode_at(
                        parent_fd,
                        name,
                        identity,
                        original_mode,
                    )
                _unlink_trash_manifest(manifest_path)
            except (OSError, UnsafePrivatePath):
                pass
            return None
        finally:
            os.close(parent_fd)

    if state != _TRASH_RESTORE_PREPARED:
        if opened is not None:
            os.close(opened[0])
        return None
    if opened is None:
        return data

    parent_fd, parent_parts, name, logical = opened
    try:
        restore_parent_identity = data.get(
            "restore_parent_identity"
        )
        if not _valid_identity(restore_parent_identity):
            restore_parent_identity = _fd_identity(parent_fd)
        trash_parent_identity = data.get("trash_parent_identity")
        if not _valid_identity(trash_parent_identity):
            trash_parent_identity = _fd_identity(anchor.trash_fd)
        if (
            _fd_identity(parent_fd) != restore_parent_identity
            or _fd_identity(anchor.trash_fd)
            != trash_parent_identity
            or _entry_identity_at(parent_fd, name) != identity
            or not _directory_is_root_reachable(
                anchor,
                parent_parts,
                parent_fd,
            )
            or not _directory_is_root_reachable(
                anchor,
                (TRASH_DIR_NAME,),
                anchor.trash_fd,
            )
        ):
            return data
        try:
            _fsync_rename_fds(
                anchor.trash_fd,
                d,
                parent_fd,
                anchor.root / logical.parent,
            )
            original_mode = data.get("original_mode")
            if (
                type(original_mode) is int
                and 0 <= original_mode <= 0o7777
            ):
                _restore_mode_at(
                    parent_fd,
                    name,
                    identity,
                    original_mode,
                )
        except (OSError, UnsafePrivatePath):
            return data
    finally:
        os.close(parent_fd)

    try:
        return _persist_restore_completion(manifest_path, data)
    except (OSError, UnsafePrivatePath):
        return data



def ensure_private_trash_storage(
    root: Path | None = None,
    *,
    create: bool = False,
) -> int:
    """Repair private storage through its stable trash directory handle."""
    d = _trash_dir(root)
    if not ensure_private_directory(d, create=create):
        return 0
    with _trash_transaction(root) as d:
        anchored = _anchor_for_path(d)
        if anchored is None:
            raise UnsafePrivatePath(
                "trash transaction anchor is unavailable")
        anchor = anchored[0]
        repaired = 1
        for name in os.listdir(anchor.trash_fd):
            if not name.endswith(".json"):
                continue
            data = _repair_trash_item(d, d / name)
            if data is None:
                continue
            repaired += 1
            if _entry_kind_at(
                anchor.trash_fd,
                str(data.get("trash_id")),
            ) in {"directory", "file"}:
                repaired += 1
    try:
        _gc_trash_auxiliary(root)
    except (OSError, UnsafePrivatePath) as exc:
        sys.stderr.write(
            "[files] trash auxiliary GC deferred "
            f"({type(exc).__name__})\n"
        )
    return repaired


def _dir_size(p: Path) -> int:
    """Sum of file sizes (best-effort; OSError on individual files skipped)."""
    total = 0
    try:
        for sub in p.rglob("*"):
            try:
                if private_path_kind(sub) == "file":
                    total += sub.lstat().st_size
            except OSError:
                continue
    except OSError:
        pass
    return total


def _move_to_trash(
    target: Path,
    root: Path | None = None,
    original_rel: str | None = None,
) -> dict:
    """Durably prepare, atomically move, then commit one soft delete.

    Caller is responsible for ensuring `target` exists + is inside ROOT.
    The prepare manifest is fsynced before rename, so a crash never creates
    an undiscoverable payload. Recovery commits a prepare record whose payload
    exists, or removes one whose original inode never moved."""
    root = Path(_root_or_default(root)).absolute()
    if original_rel is None:
        original_rel = _logical_relative_path(
            str(Path(target).relative_to(root))
        ).as_posix()
    else:
        original_rel = _logical_relative_path(original_rel).as_posix()
    target = root / _logical_relative_path(original_rel)

    with _trash_transaction(root) as trash:
        anchored = _anchor_for_path(trash)
        if anchored is None:
            raise UnsafePrivatePath("trash transaction anchor is unavailable")
        anchor = anchored[0]
        try:
            source_parent_fd, source_parent_parts, source_name = (
                _open_workspace_parent(anchor, original_rel))
        except OSError as exc:
            raise UnsafePrivatePath(
                "trash source path contains an unsafe ancestor") from exc
        try:
            source_info = os.stat(
                source_name,
                dir_fd=source_parent_fd,
                follow_symlinks=False,
            )
            identity = _stat_identity(source_info)
            target_kind = str(identity["kind"])
            if target_kind not in {"directory", "file"}:
                raise UnsafePrivatePath(
                    "trash payload must be a real file or directory")
            source_parent_identity = _fd_identity(source_parent_fd)
            trash_parent_identity = _fd_identity(anchor.trash_fd)
            original_mode = stat.S_IMODE(source_info.st_mode)
            try:
                size = (
                    _directory_size_at(source_parent_fd, source_name)
                    if target_kind == "directory"
                    else int(source_info.st_size)
                )
            except OSError:
                size = 0

            for _attempt in range(32):
                tid = _gen_trash_id()
                payload = trash / tid
                manifest_path = trash / f"{tid}.json"
                receipt_path = _restore_receipt_path(trash, tid)
                purge_receipt_path = _purge_receipt_path(trash, tid)
                tombstone_name = _purge_tombstone_name(tid)
                if (
                    _secure_path_kind(payload) != "missing"
                    or _secure_path_kind(receipt_path) != "missing"
                    or _secure_path_kind(purge_receipt_path) != "missing"
                    or _entry_kind_at(
                        anchor.trash_fd, tombstone_name) != "missing"
                ):
                    continue
                manifest = {
                    "schema_version": _TRASH_SCHEMA_VERSION,
                    _TRASH_STATE_KEY: _TRASH_DELETE_PREPARED,
                    "transaction_nonce": secrets.token_hex(16),
                    "trash_id": tid,
                    "original_path": original_rel,
                    "original_name": source_name,
                    "deleted_at": time.time(),
                    "kind": (
                        "dir" if target_kind == "directory" else "file"
                    ),
                    "size": size,
                    "original_mode": original_mode,
                    "payload_identity": identity,
                    "source_parent_identity": source_parent_identity,
                    "trash_parent_identity": trash_parent_identity,
                }
                try:
                    _create_trash_manifest(manifest_path, manifest)
                except FileExistsError:
                    continue
                break
            else:
                raise OSError("unable to reserve a unique trash transaction")

            if _entry_identity_at(source_parent_fd, source_name) != identity:
                _unlink_trash_manifest(manifest_path)
                raise UnsafePrivatePath("trash source changed after prepare")
            try:
                _prepare_directory_at(
                    source_parent_fd,
                    source_name,
                    identity,
                )
                _rename_noreplace(
                    target,
                    payload,
                    source_parent_identity=source_parent_identity,
                    destination_parent_identity=trash_parent_identity,
                    source_parent_fd=source_parent_fd,
                    destination_parent_fd=anchor.trash_fd,
                )
            except OSError as exc:
                if isinstance(exc, _RenameDurabilityError):
                    raise
                if _entry_identity_at(anchor.trash_fd, tid) == identity:
                    try:
                        _fsync_rename_fds(
                            source_parent_fd,
                            target.parent,
                            anchor.trash_fd,
                            trash,
                        )
                    except OSError as barrier_exc:
                        raise _RenameDurabilityError(
                            "rename committed but parent fsync replay failed: "
                            f"{barrier_exc}"
                        ) from barrier_exc
                    sys.stderr.write(
                        f"[files] trash rename outcome recovered tid={tid} "
                        f"({type(exc).__name__})\n"
                    )
                elif (
                    _entry_identity_at(source_parent_fd, source_name)
                    == identity
                    and _entry_kind_at(anchor.trash_fd, tid) == "missing"
                ):
                    try:
                        _restore_mode_at(
                            source_parent_fd,
                            source_name,
                            identity,
                            original_mode,
                        )
                        _unlink_trash_manifest(manifest_path)
                    except (OSError, UnsafePrivatePath):
                        pass
                    raise
                else:
                    raise OSError(
                        "trash rename outcome is uncertain") from exc

            parents_reachable = (
                _directory_is_root_reachable(
                    anchor,
                    source_parent_parts,
                    source_parent_fd,
                )
                and _directory_is_root_reachable(
                    anchor,
                    (TRASH_DIR_NAME,),
                    anchor.trash_fd,
                )
            )
            if not parents_reachable:
                try:
                    _rename_noreplace(
                        payload,
                        target,
                        source_parent_identity=trash_parent_identity,
                        destination_parent_identity=source_parent_identity,
                        source_parent_fd=anchor.trash_fd,
                        destination_parent_fd=source_parent_fd,
                    )
                    _restore_mode_at(
                        source_parent_fd,
                        source_name,
                        identity,
                        original_mode,
                    )
                    _unlink_trash_manifest(
                        manifest_path,
                        missing_ok=True,
                    )
                except (OSError, UnsafePrivatePath):
                    pass
                raise UnsafePrivatePath(
                    "trash parent is no longer reachable from workspace root")

            if _entry_identity_at(anchor.trash_fd, tid) != identity:
                raise UnsafePrivatePath(
                    "trash payload identity changed after rename")
            private_mode = (
                0o700 if target_kind == "directory" else 0o600
            )
            _restore_mode_at(
                anchor.trash_fd,
                tid,
                identity,
                private_mode,
            )
            committed = {**manifest, _TRASH_STATE_KEY: _TRASHED}
            try:
                _write_trash_manifest(manifest_path, committed)
            except (OSError, UnsafePrivatePath) as exc:
                sys.stderr.write(
                    f"[files] trash commit deferred tid={tid} "
                    f"({type(exc).__name__})\n"
                )
            return committed
        finally:
            os.close(source_parent_fd)


def _read_manifest(tid: str, root: Path | None = None) -> dict | None:
    if not _valid_trash_id(tid):
        return None
    d = _trash_dir(root)
    if not ensure_private_directory(d, create=False):
        return None
    with _trash_transaction(root) as d:
        return _repair_trash_item(d, d / f"{tid}.json")


def _list_trash(root: Path | None = None) -> list[dict]:
    """Reconcile trash or propagate I/O failure to the API as degraded."""
    items: list[dict] = []
    with _trash_transaction(root) as d:
        anchored = _anchor_for_path(d)
        if anchored is None:
            raise UnsafePrivatePath(
                "trash transaction anchor is unavailable")
        anchor = anchored[0]
        for name in os.listdir(anchor.trash_fd):
            if not name.endswith(".json"):
                continue
            tid = name[:-5]
            if not _valid_trash_id(tid):
                continue
            if _entry_kind_at(
                anchor.trash_fd,
                name,
            ) != "file":
                continue
            data = _repair_trash_item(
                d,
                d / name,
                strict_io=True,
            )
            if data is None:
                continue
            if _entry_kind_at(anchor.trash_fd, tid) in {
                "directory",
                "file",
            }:
                items.append(data)
    items.sort(key=lambda x: x.get("deleted_at", 0), reverse=True)
    return items


# Auto-expire trash items older than this many days. Tunable via env so
# users on tiny SSDs can be more aggressive; 0 = never auto-purge.
# Default 30 days mirrors macOS Finder / GNOME's "permanently delete after
# 30 days" behaviour — long enough for "wait, I needed that", short enough
# that the dustbin doesn't silently eat the disk.
#
# env_int handles non-numeric input (typo / "30 days" / etc.) by falling
# back to the default + logging to stderr — a config mistake leaves the
# feature disabled with a clear reason instead of bricking backend import.
_TRASH_TTL_DAYS = env_int("MUSELAB_TRASH_TTL_DAYS", 30, min_value=0)

_TRASH_AUX_TTL_DAYS = env_int(
    "MUSELAB_TRASH_AUX_TTL_DAYS",
    7,
    min_value=1,
)
_TRASH_AUX_TTL_SECONDS = _TRASH_AUX_TTL_DAYS * 86400


def auto_purge_expired_trash(root: Path | None = None) -> int:
    """Purge trash items whose `deleted_at` is older than _TRASH_TTL_DAYS.
    Returns the count purged. Called once at startup (see backend/main.py)
    and ignores any per-item errors so a single corrupt manifest can't
    block the cleanup of healthy entries. Returns 0 when disabled."""
    if _TRASH_TTL_DAYS <= 0:
        return 0
    cutoff = time.time() - (_TRASH_TTL_DAYS * 86400)
    purged = 0
    try:
        items = _list_trash(root)
    except (OSError, UnsafePrivatePath) as exc:
        sys.stderr.write(
            "[files] expired trash scan failed "
            f"({type(exc).__name__})\n"
        )
        return 0
    for data in items:
        if (data.get("deleted_at") or 0) >= cutoff:
            continue
        tid = str(data.get("trash_id") or "")
        try:
            outcome, _cleanup_deferred = _purge_one(tid, root)
            if outcome == "purged":
                purged += 1
        except (OSError, UnsafePrivatePath) as exc:
            # Keep both the payload remainder and manifest discoverable for a
            # later retry. Startup cleanup is best-effort, but never reports a
            # failed deletion as purged.
            sys.stderr.write(
                f"[files] expired trash purge failed tid={tid} "
                f"({type(exc).__name__})\n"
            )
    return purged


def _path_lexists(path: Path) -> bool:
    """Like lexists(): a broken symlink is still a deletable payload."""
    return path.exists() or path.is_symlink()


def _remove_path_strict(path: Path) -> bool:
    """Remove one path and verify it is gone; never swallow an I/O failure."""
    if not _path_lexists(path):
        return False
    if path.is_dir() and not path.is_symlink():
        shutil.rmtree(path)
    else:
        path.unlink()
    if _path_lexists(path):
        raise OSError("permanent delete did not remove the target")
    return True


def _purge_one(
    tid: str,
    root: Path | None = None,
) -> tuple[str, bool]:
    """Linearize purge under the root lock, then remove its tombstone outside."""
    if not _valid_trash_id(tid):
        return "missing", False
    root = Path(_root_or_default(root)).absolute()
    cleanup_fd = -1
    tombstone_name = _purge_tombstone_name(tid)
    outcome = "missing"

    with _trash_transaction(root) as d:
        anchored = _anchor_for_path(d)
        if anchored is None:
            raise UnsafePrivatePath("trash transaction anchor is unavailable")
        anchor = anchored[0]

        if _read_restore_receipt(d, tid) is not None:
            return "restored", False
        manifest_path = d / f"{tid}.json"
        data = _repair_trash_item(d, manifest_path)
        payload_kind = _entry_kind_at(anchor.trash_fd, tid)
        if (
            data is not None
            and (
                _trash_state(data) == _TRASH_RESTORED
                or (
                    _trash_state(data) == _TRASH_RESTORE_PREPARED
                    and payload_kind == "missing"
                )
            )
        ):
            return "restored", False

        purge_receipt = _read_purge_receipt(d, tid)
        tombstone_kind = _entry_kind_at(
            anchor.trash_fd,
            tombstone_name,
        )
        if purge_receipt is not None:
            outcome = "already_purged"
            if tombstone_kind == "directory":
                cleanup_fd = os.dup(anchor.trash_fd)
        else:
            manifest_kind = _entry_kind_at(
                anchor.trash_fd,
                f"{tid}.json",
            )
            if (
                tombstone_kind == "missing"
                and payload_kind == "missing"
                and manifest_kind == "missing"
            ):
                return "missing", False
            if tombstone_kind == "missing":
                os.mkdir(
                    tombstone_name,
                    mode=0o700,
                    dir_fd=anchor.trash_fd,
                )
                tombstone_fd = os.open(
                    tombstone_name,
                    _directory_open_flags(),
                    dir_fd=anchor.trash_fd,
                )
                os.fchmod(tombstone_fd, 0o700)
                _fsync_open_directory(
                    tombstone_fd,
                    d / tombstone_name,
                )
                _fsync_open_directory(anchor.trash_fd, d)
            elif tombstone_kind == "directory":
                tombstone_fd = os.open(
                    tombstone_name,
                    _directory_open_flags(),
                    dir_fd=anchor.trash_fd,
                )
                os.fchmod(tombstone_fd, 0o700)
            else:
                raise UnsafePrivatePath("purge tombstone is unsafe")

            try:
                if payload_kind in {"directory", "file"}:
                    if _entry_kind_at(tombstone_fd, "payload") != "missing":
                        raise UnsafePrivatePath(
                            "purge payload tombstone is occupied")
                    _rename_noreplace(
                        d / tid,
                        d / tombstone_name / "payload",
                        source_parent_identity=_fd_identity(anchor.trash_fd),
                        destination_parent_identity=_fd_identity(tombstone_fd),
                        source_parent_fd=anchor.trash_fd,
                        destination_parent_fd=tombstone_fd,
                    )
                elif payload_kind != "missing":
                    raise UnsafePrivatePath("trash payload is unsafe")

                manifest_kind = _entry_kind_at(
                    anchor.trash_fd,
                    f"{tid}.json",
                )
                if manifest_kind == "file":
                    if _entry_kind_at(tombstone_fd, "manifest") != "missing":
                        raise UnsafePrivatePath(
                            "purge manifest tombstone is occupied")
                    _rename_noreplace(
                        manifest_path,
                        d / tombstone_name / "manifest",
                        source_parent_identity=_fd_identity(anchor.trash_fd),
                        destination_parent_identity=_fd_identity(tombstone_fd),
                        source_parent_fd=anchor.trash_fd,
                        destination_parent_fd=tombstone_fd,
                    )
                elif manifest_kind != "missing":
                    raise UnsafePrivatePath("trash manifest is unsafe")

                if (
                    not _child_directory_reachable(
                        anchor.trash_fd,
                        tombstone_name,
                        tombstone_fd,
                    )
                    or not _directory_is_root_reachable(
                        anchor,
                        (TRASH_DIR_NAME,),
                        anchor.trash_fd,
                    )
                ):
                    try:
                        if (
                            _entry_kind_at(tombstone_fd, "payload")
                            in {"directory", "file"}
                            and _entry_kind_at(
                                anchor.trash_fd, tid) == "missing"
                        ):
                            _rename_noreplace(
                                d / tombstone_name / "payload",
                                d / tid,
                                source_parent_identity=_fd_identity(
                                    tombstone_fd),
                                destination_parent_identity=_fd_identity(
                                    anchor.trash_fd),
                                source_parent_fd=tombstone_fd,
                                destination_parent_fd=anchor.trash_fd,
                            )
                        if (
                            _entry_kind_at(tombstone_fd, "manifest")
                            == "file"
                            and _entry_kind_at(
                                anchor.trash_fd,
                                f"{tid}.json",
                            )
                            == "missing"
                        ):
                            _rename_noreplace(
                                d / tombstone_name / "manifest",
                                manifest_path,
                                source_parent_identity=_fd_identity(
                                    tombstone_fd),
                                destination_parent_identity=_fd_identity(
                                    anchor.trash_fd),
                                source_parent_fd=tombstone_fd,
                                destination_parent_fd=anchor.trash_fd,
                            )
                    except OSError:
                        pass
                    raise UnsafePrivatePath(
                        "purge tombstone is no longer root-reachable")

                purge_record = {
                    "schema_version": _TRASH_SCHEMA_VERSION,
                    _TRASH_STATE_KEY: _TRASH_PURGED,
                    "trash_id": tid,
                    "purged_at": time.time(),
                }
                _create_trash_manifest(
                    _purge_receipt_path(d, tid),
                    purge_record,
                )
                outcome = "purged"
                cleanup_fd = os.dup(anchor.trash_fd)
            finally:
                os.close(tombstone_fd)

    cleanup_deferred = False
    if cleanup_fd >= 0:
        try:
            _remove_tombstone_at(cleanup_fd, tombstone_name)
        except (OSError, UnsafePrivatePath) as exc:
            cleanup_deferred = True
            sys.stderr.write(
                f"[files] purge cleanup deferred tid={tid} "
                f"({type(exc).__name__})\n"
            )
        finally:
            os.close(cleanup_fd)
    return outcome, cleanup_deferred


def _gc_trash_auxiliary(root: Path | None = None) -> int:
    """Bound receipts, crashed temp files and staged tombstones by TTL."""
    root = Path(_root_or_default(root)).absolute()
    cutoff = time.time() - _TRASH_AUX_TTL_SECONDS
    tombstones: list[str] = []
    cleanup_fd = -1
    removed = 0
    with _trash_transaction(root) as d:
        anchored = _anchor_for_path(d)
        if anchored is None:
            raise UnsafePrivatePath("trash transaction anchor is unavailable")
        anchor = anchored[0]
        metadata_changed = False
        for name in os.listdir(anchor.trash_fd):
            try:
                info = os.stat(
                    name,
                    dir_fd=anchor.trash_fd,
                    follow_symlinks=False,
                )
            except FileNotFoundError:
                continue
            if info.st_mtime > cutoff:
                continue
            is_receipt = (
                name.endswith(_TRASH_RECEIPT_SUFFIX)
                or name.endswith(_PURGE_RECEIPT_SUFFIX)
            )
            is_temp = name.startswith(".") and ".txn." in name
            is_tombstone = (
                name.startswith(_PURGE_TOMBSTONE_PREFIX)
                or name.startswith(_PERMANENT_TOMBSTONE_PREFIX)
            )
            kind = _stat_identity(info)["kind"]
            if (is_receipt or is_temp) and kind in {"file", "symlink"}:
                os.unlink(name, dir_fd=anchor.trash_fd)
                metadata_changed = True
                removed += 1
            elif is_tombstone and kind in {"directory", "symlink"}:
                tombstones.append(name)
        if metadata_changed:
            _fsync_open_directory(anchor.trash_fd, d)
        if tombstones:
            cleanup_fd = os.dup(anchor.trash_fd)

    try:
        for name in tombstones:
            _remove_tombstone_at(cleanup_fd, name)
            removed += 1
    finally:
        if cleanup_fd >= 0:
            os.close(cleanup_fd)
    return removed

def _delete_failure(
    exc: OSError,
    *,
    target: Path,
    directory_payload: bool,
    detail: str,
) -> HTTPException:
    """Map destructive I/O failures without exposing a filesystem path."""
    if isinstance(exc, PermissionError):
        return HTTPException(
            status_code=403,
            detail=detail,
            headers={"X-MuseLab-Error-Code": "permission_denied"},
        )
    error_code = (
        "partial_delete"
        if directory_payload and _path_lexists(target)
        else "io_error"
    )
    return HTTPException(
        status_code=500,
        detail=detail,
        headers={"X-MuseLab-Error-Code": error_code},
    )

# Filenames without extensions that are commonly text (Dockerfile, Makefile, etc.).
# Compared case-insensitively against the full name.
# Known-binary extensions — fast reject, don't even try to sniff.
# Everything NOT in this set + not containing NUL bytes in the sniff window
# is treated as text-previewable. This lets .tmpl / .vue.bak / random custom
# extensions all preview without us maintaining a whitelist.
BINARY_EXT = {
    # archives / packages
    ".zip", ".tar", ".gz", ".bz2", ".xz", ".7z", ".rar", ".tgz", ".tbz",
    ".whl", ".jar", ".war", ".ear", ".deb", ".rpm", ".pkg", ".dmg", ".iso",
    ".apk", ".ipa", ".xpi", ".crx",
    # images (have their own img preview)
    ".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp", ".ico", ".tiff", ".tif",
    ".heic", ".heif", ".raw", ".psd", ".ai", ".sketch", ".fig",
    # audio / video
    ".mp3", ".m4a", ".wav", ".flac", ".ogg", ".opus", ".aac", ".wma",
    ".mp4", ".webm", ".mkv", ".avi", ".mov", ".wmv", ".flv", ".m4v",
    # binary docs (PDF has its own preview; office formats need conversion)
    ".pdf", ".doc", ".docx", ".xls", ".xlsx", ".ppt", ".pptx", ".odt",
    ".ods", ".odp", ".rtf", ".epub", ".mobi",
    # executables / libs
    ".exe", ".dll", ".so", ".dylib", ".bin", ".o", ".a", ".lib", ".obj",
    ".class", ".pyc", ".pyo", ".elc", ".wasm",
    # fonts
    ".ttf", ".otf", ".woff", ".woff2", ".eot",
    # databases
    ".db", ".sqlite", ".sqlite3", ".mdb",
}
MAX_TEXT_SIZE = 2 * 1024 * 1024  # 2 MB — bigger files refuse with 413
SNIFF_BYTES = 4096                # how much we read to detect NUL bytes


def _looks_binary(p: Path) -> bool:
    """Heuristic: read up to 4 KB, presence of NUL byte → binary. Otherwise
    decode with `errors="replace"` and check how many bytes turned into the
    Unicode replacement character (U+FFFD). High ratio → binary / garbage.

    Important: we must NOT use plain `decode("utf-8")` here. The sniff window
    cuts at a fixed byte offset, which routinely splits a multi-byte UTF-8
    character (CJK chars are 3 bytes). A clean text file would then raise
    UnicodeDecodeError purely because of the chunk boundary — wrongly tagged
    binary. `errors="replace"` decodes whatever can be decoded and only the
    truly invalid bytes become U+FFFD."""
    try:
        with p.open("rb") as f:
            chunk = f.read(SNIFF_BYTES)
    except OSError:
        return True
    if b"\x00" in chunk:
        return True
    # decode with replacement; count how many chars are the replacement marker
    decoded = chunk.decode("utf-8", errors="replace")
    if not decoded:
        return False
    bad = decoded.count("�")
    # >5% replacement chars across a 4 KB window strongly suggests non-UTF-8
    # binary. A clean text file boundary-split mid-char contributes ≤1 bad char.
    return (bad / len(decoded)) > 0.05

# Files whose contents should never be served or overwritten through this API,
# regardless of extension. Matches against the basename (case-insensitive).
SENSITIVE_NAMES = {
    ".env", ".env.local", ".env.production", ".env.development",
    ".netrc", ".pgpass", ".npmrc", ".pypirc", ".dockercfg",
    ".htpasswd", ".htaccess",
    "credentials", "credentials.json", "service-account.json",
    "id_rsa", "id_dsa", "id_ecdsa", "id_ed25519",
    "authorized_keys", "known_hosts",
    # Shell / language history files — frequently contain pasted tokens,
    # one-off commands with secrets. Added when MUSELAB_ROOT=$HOME became
    # supported (2026-05-17) so a token leak doesn't expose them.
    ".bash_history", ".zsh_history", ".python_history", ".node_repl_history",
    ".sqlite_history", ".lesshst", ".viminfo", ".wget-hsts",
    ".npm-debug.log", ".yarn-error.log",
}
# Extension suffixes treated as sensitive — private keys, cert bundles, and
# `.env`-style files regardless of basename (prod.env, staging.env, etc.).
SENSITIVE_SUFFIX = {".pem", ".key", ".p12", ".pfx", ".keystore", ".jks", ".env"}


def _is_sensitive(p: Path) -> bool:
    name = p.name.lower()
    if name in SENSITIVE_NAMES:
        return True
    if name.startswith(".env."):  # .env.* variants like .env.local
        return True
    if p.suffix.lower() in SENSITIVE_SUFFIX:
        return True
    return False


def _logical_relative_path(rel: str) -> Path:
    """Return a normalized workspace-relative path without following links.

    A resolved target may legitimately leave the selected workspace through a
    symlink, so containment cannot distinguish that case from ``../`` after
    resolution. Reject parent traversal lexically first; with no ``..``
    component, a later escape can only have happened through a symlink.
    """
    logical = Path((rel or "").lstrip("/"))
    if any(part == ".." for part in logical.parts):
        raise HTTPException(status_code=400, detail="path escapes root")
    return logical


def _inside(path: Path, root: Path) -> bool:
    return path == root or root in path.parents


def _inside_registered_workspace(target: Path) -> bool:
    for workspace in workspace_registry.paths():
        try:
            if _inside(target, workspace.resolve()):
                return True
        except (OSError, RuntimeError):
            continue
    return False


def _inside_internal_state(target: Path) -> bool:
    for workspace in workspace_registry.paths():
        try:
            internal_root = (
                workspace.resolve() / INTERNAL_DIR_NAME
            ).resolve()
            if _inside(target, internal_root):
                return True
        except (OSError, RuntimeError):
            continue
    return False


def safe_resolve(
    rel: str,
    allow_sensitive: bool = False,
    root: Path | None = None,
) -> Path:
    """Resolve a path relative to ROOT, blocking traversal outside ROOT and,
    by default, blocking access to credential-shaped filenames.

    Defends against:
      - `../../etc/passwd` style traversal (resolve() + ROOT-prefix check)
      - **Symlink escape**: links may target another registered workspace, but
        never an unregistered filesystem area such as ``/etc``.
      - `..` traversal, including attempts to jump directly into another
        registered workspace instead of following an actual symlink.
      - `.env`, `id_rsa`, `*.pem` etc. (SENSITIVE_SUFFIX / SENSITIVE_NAMES)."""
    root = _root_or_default(root)
    logical = _logical_relative_path(rel)
    if INTERNAL_DIR_NAME in logical.parts:
        raise HTTPException(
            status_code=403,
            detail="muselab internal state is not accessible",
        )
    # NUL byte in a path raises ValueError from (root / rel) and FastAPI
    # converts that to a 500 with a traceback that leaks internal module
    # paths. Reject early as 400. Same for any string that Python's path
    # layer refuses (control chars trip OS-level checks downstream).
    if "\x00" in rel:
        raise HTTPException(status_code=400, detail="invalid path")
    # First-pass resolve (follows symlinks → catches symlink escape):
    try:
        target = (root / logical).resolve()
    except (ValueError, OSError):
        raise HTTPException(status_code=400, detail="invalid path") from None
    # The selected root itself might be a symlink target; compare real paths.
    root_real = root.resolve()
    if not _inside(target, root_real) and not _inside_registered_workspace(target):
        raise HTTPException(status_code=400, detail="path escapes root")
    # Logical-path checks alone are insufficient: a workspace symlink can point
    # at `.muselab`. Protect the resolved target across every registered root.
    if _inside_internal_state(target):
        raise HTTPException(
            status_code=403,
            detail="muselab internal state is not accessible",
        )
    # Block by name regardless of whether the file already exists, so the API
    # can neither read nor write `.env` / private-key shaped paths.
    if not allow_sensitive and not target.is_dir() and _is_sensitive(target):
        raise HTTPException(status_code=403, detail="sensitive file blocked")
    return target


class Entry(BaseModel):
    name: str
    path: str  # relative to ROOT
    is_dir: bool
    size: int
    mtime: float


MAX_LIST_ENTRIES = 500  # safety cap so huge dirs (.git/objects) don't freeze the UI


@router.get("/list", dependencies=[Depends(require_token)])
def list_dir(
    path: str = "",
    show_hidden: bool = False,
    root: Path = Depends(_workspace_root),
) -> dict:
    target = safe_resolve(path, root=root)
    if not target.exists():
        raise HTTPException(status_code=404, detail="not found")
    if not target.is_dir():
        raise HTTPException(status_code=400, detail="not a directory")
    entries: list[dict] = []
    # Keep the caller's logical path when `path` traverses a directory
    # symlink. `target` is the resolved real directory (required for the
    # containment check), but DirEntry.path therefore points at that real
    # location too. Returning it made a tree row such as `vendor -> shared`
    # yield children named `shared/x` instead of `vendor/x`; Alpine then
    # filtered them as duplicates when the real directory was also visible,
    # making the symlink look impossible to expand.
    logical_parent = _logical_relative_path(path)
    # Trash dir is always hidden from the file tree (even when
    # show_hidden=true) — it has its own dedicated UI surface; mixing it
    # back into the tree would surface deleted files in a confusing
    # context. Only relevant at the root level since trash dir lives there.
    is_root_listing = (target == root)
    # DirEntry caches d_type/stat data supplied by the OS. The Path-based
    # implementation called is_dir() during sort and then two more times per
    # rendered row, which made large network-mounted directories needlessly
    # syscall-heavy.
    try:
        with os.scandir(target) as scan:
            # Keep only the first UI page while scanning. The previous code
            # accumulated and sorted every DirEntry before slicing to 500, so a
            # dump directory with tens of thousands of files caused a large
            # allocation and O(N log N) sort even though the browser could never
            # receive more than MAX_LIST_ENTRIES rows. nsmallest still preserves
            # deterministic directory-first ordering while bounding memory and
            # sorting work to O(MAX_LIST_ENTRIES).
            def candidates():
                for child in scan:
                    if is_root_listing and child.name in {
                        TRASH_DIR_NAME,
                        INTERNAL_DIR_NAME,
                    }:
                        continue
                    if not show_hidden and child.name.startswith("."):
                        continue
                    try:
                        is_dir = child.is_dir()
                    except OSError:
                        continue
                    yield (
                        (not is_dir, child.name.lower(), child.name),
                        child,
                        is_dir,
                    )

            selected = heapq.nsmallest(
                MAX_LIST_ENTRIES + 1,
                candidates(),
                key=lambda item: item[0],
            )
    except OSError as e:
        raise HTTPException(status_code=500, detail=f"failed to list: {e}")
    truncated = len(selected) > MAX_LIST_ENTRIES
    for _, child, is_dir in selected[:MAX_LIST_ENTRIES]:
        try:
            stat = child.stat()
        except OSError:
            continue
        entries.append(Entry(
            name=child.name,
            path=(logical_parent / child.name).as_posix(),
            is_dir=is_dir,
            size=stat.st_size if not is_dir else 0,
            mtime=stat.st_mtime,
        ).model_dump())
    return {"root": str(root), "path": path, "entries": entries, "truncated": truncated}


# xlsx preview caps. Read-only mode + capped per-sheet rows/cols so a
# 1M-cell spreadsheet doesn't OOM the SSE event loop or blow up the JSON
# payload over the wire. Truncation is signaled to the FE so it can hint
# the user instead of silently dropping data.
XLSX_MAX_SHEETS = 20
XLSX_MAX_ROWS = 500
XLSX_MAX_COLS = 50
XLSX_CELL_MAX_CHARS = 500   # one obnoxious cell shouldn't blow the page


@router.get("/xlsx", dependencies=[Depends(require_token)])
def xlsx_preview(path: str, root: Path = Depends(_workspace_root)) -> dict:
    """Read-only xlsx preview as structured JSON.

    Returns each sheet's first XLSX_MAX_ROWS×XLSX_MAX_COLS cells as
    strings. Formulas are NOT evaluated — `data_only=True` returns the
    cached value the spreadsheet app last wrote. If a file was created
    programmatically without ever being opened in Excel/LibreOffice,
    formula cells will be null and surface as empty strings.
    """
    target = safe_resolve(path, root=root)
    if not target.exists() or not target.is_file():
        raise HTTPException(status_code=404, detail="not a file")
    if target.suffix.lower() not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        raise HTTPException(status_code=415, detail="not an xlsx-family file")
    try:
        import openpyxl  # local import — openpyxl is only loaded on demand
    except ImportError:
        raise HTTPException(status_code=500,
                            detail="openpyxl not installed — run `uv sync`")
    try:
        wb = openpyxl.load_workbook(target, read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=422,
                            detail=f"failed to parse xlsx: {type(e).__name__}: {e}")
    try:
        sheets: list[dict] = []
        sheet_names = wb.sheetnames
        sheets_truncated = len(sheet_names) > XLSX_MAX_SHEETS
        for sheet_name in sheet_names[:XLSX_MAX_SHEETS]:
            ws = wb[sheet_name]
            rows: list[list[str]] = []
            rows_truncated = False
            cols_truncated = False
            for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
                if r_idx >= XLSX_MAX_ROWS:
                    rows_truncated = True
                    break
                cells: list[str] = []
                for c_idx, val in enumerate(row):
                    if c_idx >= XLSX_MAX_COLS:
                        cols_truncated = True
                        break
                    if val is None:
                        cells.append("")
                    else:
                        s = str(val)
                        if len(s) > XLSX_CELL_MAX_CHARS:
                            s = s[:XLSX_CELL_MAX_CHARS] + "…"
                        cells.append(s)
                rows.append(cells)
            sheets.append({
                "name": sheet_name,
                "rows": rows,
                "rows_truncated": rows_truncated,
                "cols_truncated": cols_truncated,
            })
        return {
            "path": path,
            "sheets": sheets,
            "sheets_truncated": sheets_truncated,
            "limits": {"max_rows": XLSX_MAX_ROWS, "max_cols": XLSX_MAX_COLS,
                       "max_sheets": XLSX_MAX_SHEETS},
        }
    finally:
        wb.close()


# CSV preview caps. Paginated by design — CSV files in the wild can be
# millions of rows, so we never load the whole file into memory. Each
# request returns one window; the UI calls back with offset += limit when
# the user pages forward.
CSV_DEFAULT_LIMIT = 200       # default page size
CSV_MAX_LIMIT = 1000          # hard ceiling the client can request
CSV_MAX_COLS = 50             # per-row column cap
CSV_CELL_MAX_CHARS = 500
CSV_SNIFF_BYTES = 8192        # sample size for delimiter / header detection

# Counting every row is necessarily O(file size), but the old endpoint paid
# that cost on *every* page. Cache only the total (never cell content), keyed by
# the file's stat signature; writes invalidate naturally via mtime/size. The
# small LRU is protected because FastAPI sync handlers run in a thread pool.
CSV_TOTAL_CACHE_MAX = 64
_CSV_TOTAL_CACHE: OrderedDict[str, tuple[int, int, int]] = OrderedDict()
_CSV_TOTAL_CACHE_LOCK = threading.Lock()


def _csv_total_cache_get(target: Path, mtime_ns: int, size: int) -> int | None:
    key = str(target)
    with _CSV_TOTAL_CACHE_LOCK:
        value = _CSV_TOTAL_CACHE.get(key)
        if value is None or value[:2] != (mtime_ns, size):
            if value is not None:
                _CSV_TOTAL_CACHE.pop(key, None)
            return None
        _CSV_TOTAL_CACHE.move_to_end(key)
        return value[2]


def _csv_total_cache_set(target: Path, mtime_ns: int, size: int, total: int) -> None:
    key = str(target)
    with _CSV_TOTAL_CACHE_LOCK:
        _CSV_TOTAL_CACHE[key] = (mtime_ns, size, total)
        _CSV_TOTAL_CACHE.move_to_end(key)
        while len(_CSV_TOTAL_CACHE) > CSV_TOTAL_CACHE_MAX:
            _CSV_TOTAL_CACHE.popitem(last=False)


@router.get("/csv", dependencies=[Depends(require_token)])
def csv_preview(
    path: str,
    offset: int = 0,
    limit: int = CSV_DEFAULT_LIMIT,
    root: Path = Depends(_workspace_root),
) -> dict:
    """Read-only paginated CSV / TSV preview as structured JSON.

    Returns rows[offset : offset+limit] from the file, plus the sniffed
    delimiter and a `total_rows` count so the UI can show pagination.
    Header row (if csv.Sniffer flags one) is returned separately.

    Designed to never load more than a window into memory: the file is
    iterated row-by-row, skipping rows below offset and breaking once
    `limit` is filled. The trailing total-rows count is the only full
    scan, and it just discards each row.
    """
    import csv as _csv  # local import — csv is stdlib, but keep import local
                       # so import overhead stays out of every other route.
    target = safe_resolve(path, root=root)
    if not target.exists() or not target.is_file():
        raise HTTPException(status_code=404, detail="not a file")
    if target.suffix.lower() not in {".csv", ".tsv"}:
        raise HTTPException(status_code=415, detail="not a csv/tsv file")
    try:
        stat = target.stat()
    except OSError:
        raise HTTPException(status_code=404, detail="not a file") from None
    cached_total = _csv_total_cache_get(target, stat.st_mtime_ns, stat.st_size)
    if limit < 1:
        limit = CSV_DEFAULT_LIMIT
    if limit > CSV_MAX_LIMIT:
        limit = CSV_MAX_LIMIT
    if offset < 0:
        offset = 0
    # Sniff delimiter + header from a small head sample. Defaults to
    # excel-style comma if Sniffer can't tell (e.g. one-column file).
    try:
        with target.open("r", encoding="utf-8", errors="replace", newline="") as f:
            sample = f.read(CSV_SNIFF_BYTES)
        try:
            dialect = _csv.Sniffer().sniff(sample, delimiters=",\t;|")
            has_header = _csv.Sniffer().has_header(sample)
        except _csv.Error:
            dialect = _csv.excel
            has_header = False
        # Override sniff for explicit .tsv — Sniffer sometimes guesses comma
        # on tab-separated files when the first row has no tabs.
        if target.suffix.lower() == ".tsv":
            dialect = _csv.excel_tab
    except OSError as e:
        raise HTTPException(status_code=500, detail=f"failed to read: {e}")

    header: list[str] = []
    rows: list[list[str]] = []
    cols_truncated = False
    total_rows = 0
    try:
        with target.open("r", encoding="utf-8", errors="replace", newline="") as f:
            reader = _csv.reader(f, dialect=dialect)
            # Pull header before any data offset is applied. The user paging
            # to offset=200 still wants column titles at the top of the page.
            if has_header:
                try:
                    header_row = next(reader)
                    header = [_clip_cell(c) for c in header_row[:CSV_MAX_COLS]]
                    if len(header_row) > CSV_MAX_COLS:
                        cols_truncated = True
                except StopIteration:
                    pass
            row_idx = 0
            # A cached total means pagination only needs to walk through the
            # requested window; it can stop immediately afterwards. Without a
            # cached total (first request or changed file), continue to EOF once
            # so future pages are cheap.
            if cached_total is None or offset < cached_total:
                for raw in reader:
                    if row_idx < offset:
                        row_idx += 1
                        continue
                    if len(rows) < limit:
                        cells = [_clip_cell(c) for c in raw[:CSV_MAX_COLS]]
                        if len(raw) > CSV_MAX_COLS:
                            cols_truncated = True
                        rows.append(cells)
                    row_idx += 1
                    if cached_total is not None and len(rows) >= limit:
                        break
            total_rows = cached_total if cached_total is not None else row_idx
    except OSError as e:
        raise HTTPException(status_code=500, detail=f"failed to read: {e}")

    if cached_total is None:
        # Do not cache a count under an obsolete signature if an external
        # writer changed the file during our scan.
        try:
            end_stat = target.stat()
        except OSError:
            end_stat = None
        if (end_stat is not None
                and (end_stat.st_mtime_ns, end_stat.st_size)
                == (stat.st_mtime_ns, stat.st_size)):
            _csv_total_cache_set(target, stat.st_mtime_ns, stat.st_size, total_rows)

    return {
        "path": path,
        "header": header,
        "rows": rows,
        "offset": offset,
        "limit": limit,
        "total_rows": total_rows,
        "has_header": has_header,
        "delimiter": dialect.delimiter,
        "cols_truncated": cols_truncated,
        "limits": {"max_cols": CSV_MAX_COLS, "max_limit": CSV_MAX_LIMIT},
    }


def _clip_cell(value: str) -> str:
    """Cap a single CSV cell so one runaway value can't blow up the page."""
    s = "" if value is None else str(value)
    if len(s) > CSV_CELL_MAX_CHARS:
        s = s[:CSV_CELL_MAX_CHARS] + "…"
    return s


@router.get("/read", dependencies=[Depends(require_token)])
def read_file(
    path: str,
    root: Path = Depends(_workspace_root),
) -> PlainTextResponse:
    target = safe_resolve(path, root=root)
    if not target.exists() or not target.is_file():
        raise HTTPException(status_code=404, detail="not a file")
    suffix = target.suffix.lower()
    # Fast reject for known binary extensions.
    if suffix in BINARY_EXT:
        raise HTTPException(status_code=415, detail="binary file — not previewable as text")
    # Single stat() reused for both the size gate and the empty-file check.
    # The previous code called target.stat() twice; if the file vanished
    # between the two calls (TOCTOU) the second stat raised
    # FileNotFoundError → 500 instead of a clean 404.
    try:
        st_size = target.stat().st_size
    except OSError:
        raise HTTPException(status_code=404, detail="not a file") from None
    if st_size > MAX_TEXT_SIZE:
        raise HTTPException(status_code=413, detail="file too large for preview")
    # Empty extension + not a known text name? Sniff content. Empty files OK.
    # This is the path that picks up .tmpl, .conf.j2, .env.staging, etc.
    if st_size > 0 and _looks_binary(target):
        raise HTTPException(status_code=415, detail="binary content — not previewable as text")
    content = target.read_text(encoding="utf-8", errors="replace")
    if len(content) > MAX_TEXT_SIZE:
        raise HTTPException(
            status_code=413,
            detail=f"File too large to read as text ({len(content)} bytes > {MAX_TEXT_SIZE})",
        )
    return PlainTextResponse(
        content,
        headers={
            "Content-Disposition": "inline",
            "X-Content-Type-Options": "nosniff",
        },
    )


@router.get("/stat", dependencies=[Depends(require_token)])
def stat_file(path: str, root: Path = Depends(_workspace_root)) -> dict:
    """Lightweight metadata for a single path — name, size, mtime, is_dir.

    Powers the preview header's "real path + last-modified" strip: the
    frontend only knows a tab's path string, not its on-disk mtime (the
    tree-list carries mtime but a file opened via chat-link / search is
    never in the visible tree). One cheap stat() fills that gap without
    re-reading the whole file. 404 when the path is gone — same contract
    as /read, so a stale/phantom tab surfaces honestly instead of showing
    a path that no longer exists."""
    target = safe_resolve(path, root=root)
    try:
        st = target.stat()
    except OSError:
        raise HTTPException(status_code=404, detail="not found") from None
    is_dir = target.is_dir()
    return {
        "path": _logical_relative_path(path).as_posix(),
        "name": target.name,
        "is_dir": is_dir,
        "size": 0 if is_dir else st.st_size,
        "mtime": st.st_mtime,
    }


# Types we serve inline (images / PDF / media render natively in browser).
INLINE_OK_SUFFIX = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp", ".ico",
                    ".pdf", ".mp4", ".webm", ".mp3", ".ogg", ".wav"}
# Types we serve inline INSIDE A SANDBOXED IFRAME (HTML / SVG can render but
# the strong CSP + sandbox attribute on the iframe blocks JS execution and
# same-origin token theft).
SANDBOXED_INLINE_SUFFIX = {".html", ".htm", ".svg"}

# HTML-preview bridge. The iframe has an opaque sandbox origin, so the parent
# cannot inspect its document scroll position or intercept image clicks. This
# script is injected only for preview=1 and uses postMessage for both jobs;
# neither feature requires relaxing the sandbox. The parent validates
# event.source against its own preview iframe, while this child accepts restore
# messages only from its parent. CSP permits this inline bridge.
_PREVIEW_HTML_BRIDGE = (
    "<script>(function(){"
    "var pending=false,behaviorRoot=null,behaviorValue='',behaviorPriority='',behaviorFrame=0;"
    "function sendScroll(){try{parent.postMessage({__muselab:'preview-scroll',"
    "top:Math.max(0,window.scrollY||document.documentElement.scrollTop||0),"
    "left:Math.max(0,window.scrollX||document.documentElement.scrollLeft||0)},'*');}"
    "catch(_e){}}"
    "function jump(left,top){var root=document.scrollingElement||document.documentElement;"
    "left=Math.max(0,Number(left)||0);top=Math.max(0,Number(top)||0);"
    "if(root&&!behaviorFrame){behaviorRoot=root;"
    "behaviorValue=root.style.getPropertyValue('scroll-behavior');"
    "behaviorPriority=root.style.getPropertyPriority('scroll-behavior');}"
    "if(root)root.style.setProperty('scroll-behavior','auto','important');"
    "try{scrollTo({left:left,top:top,behavior:'instant'});}catch(_e){scrollTo(left,top);}"
    "if(root){if(behaviorFrame)cancelAnimationFrame(behaviorFrame);"
    "behaviorFrame=requestAnimationFrame(function(){"
    "if(behaviorValue)behaviorRoot.style.setProperty('scroll-behavior',behaviorValue,behaviorPriority);"
    "else behaviorRoot.style.removeProperty('scroll-behavior');behaviorFrame=0;});}}"
    "addEventListener('scroll',function(){if(pending)return;pending=true;"
    "requestAnimationFrame(function(){pending=false;sendScroll();});},{passive:true});"
    "addEventListener('message',function(e){var d=e.data;"
    "if(e.source!==parent||!d||d.__muselab!=='preview-scroll-restore')return;"
    "jump(d.left,d.top);"
    "requestAnimationFrame(sendScroll);});"
    "function ready(){try{parent.postMessage({__muselab:'preview-ready'},'*');}"
    "catch(_e){}sendScroll();}"
    "if(document.readyState==='loading'){document.addEventListener('DOMContentLoaded',ready,{once:true});}"
    "else{ready();}"
    "document.addEventListener('click',function(e){"
    "var t=e.target;var img=t&&t.closest?t.closest('img'):null;"
    "if(!img||img.closest('a'))return;var src=img.currentSrc||img.src;"
    "if(!src)return;e.preventDefault();"
    "try{parent.postMessage({__muselab:'preview-img',src:src,alt:img.alt||''},'*');}"
    "catch(_e){}},true);})();</script>"
)
# Cap the in-memory read used for injection. Bigger HTML (e.g. reports with
# megabytes of base64 images) falls back to streaming untouched — it won't
# have click-to-zoom or reading-position restoration, an acceptable degradation.
_PREVIEW_INJECT_MAX_BYTES = 12 * 1024 * 1024

# Browser-native preview surfaces (notably a sandboxed HTML iframe) cannot add
# X-Auth-Token.  Never put the long-lived API token in their URL: the previewed
# document can read its own location.search even though the sandbox blocks
# cookies and parent storage.  These tickets are short-lived and bound to one
# exact resolved file + workspace.  They are reusable during the TTL because
# PDF viewers and conditional/range requests may fetch the same URL more than
# once; replay cannot reach any other file or API.
_PREVIEW_TICKET_TTL_S = max(
    30, min(env_int("MUSELAB_PREVIEW_TICKET_TTL_S", 600, min_value=1), 3600))
_PREVIEW_TICKET_MAX = 2048
_preview_tickets: OrderedDict[str, tuple[str, str, float]] = OrderedDict()
_preview_ticket_lock = threading.Lock()


class PreviewTicketReq(BaseModel):
    path: str


def _prune_preview_tickets(now: float) -> None:
    while _preview_tickets:
        digest, row = next(iter(_preview_tickets.items()))
        if row[2] >= now and len(_preview_tickets) <= _PREVIEW_TICKET_MAX:
            break
        _preview_tickets.pop(digest, None)


def _preview_ticket_ok(ticket: str, path: str, root: Path) -> bool:
    if not ticket.startswith("preview."):
        return False
    digest = hashlib.sha256(ticket[8:].encode("utf-8")).hexdigest()
    now = time.monotonic()
    try:
        target = safe_resolve(path, root=root)
    except HTTPException:
        return False
    with _preview_ticket_lock:
        _prune_preview_tickets(now)
        row = _preview_tickets.get(digest)
        if row is None or row[2] < now:
            return False
    return row[0] == str(target) and row[1] == str(root.resolve())


@router.post("/preview-ticket", dependencies=[Depends(require_token)])
def mint_preview_ticket(
    req: PreviewTicketReq,
    root: Path = Depends(_workspace_root),
) -> dict:
    target = safe_resolve(req.path, root=root)
    if not target.exists() or not target.is_file():
        raise HTTPException(status_code=404, detail="not a file")
    raw = secrets.token_urlsafe(32)
    digest = hashlib.sha256(raw.encode("utf-8")).hexdigest()
    now = time.monotonic()
    with _preview_ticket_lock:
        _prune_preview_tickets(now)
        _preview_tickets[digest] = (
            str(target),
            str(root.resolve()),
            now + _PREVIEW_TICKET_TTL_S,
        )
        _prune_preview_tickets(now)
    return {
        "ticket": "preview." + raw,
        "expires_in": _PREVIEW_TICKET_TTL_S,
    }


_DOWNLOAD_TICKET_TTL_S = 60


@router.post("/download-ticket", dependencies=[Depends(require_token)])
def mint_download_ticket(
    req: PreviewTicketReq,
    root: Path = Depends(_workspace_root),
) -> dict:
    target = safe_resolve(req.path, root=root)
    if not target.exists() or not target.is_file():
        raise HTTPException(status_code=404, detail="not a file")
    ticket = tickets.mint(
        "download",
        (str(target), str(root.resolve())),
        ttl=_DOWNLOAD_TICKET_TTL_S,
        single_use=True,
    )
    return {"ticket": ticket, "expires_in": _DOWNLOAD_TICKET_TTL_S}


async def _require_raw_access(
    path: str = Query(...),
    ticket: str = Query(""),
    token: str | None = Query(default=None),
    root: Path = Depends(_workspace_root),
) -> None:
    if ticket and _preview_ticket_ok(ticket, path, root):
        return
    # Backward compatibility for old clients, copied download links and image
    # URLs.  The first-party HTML preview no longer uses this long-lived token.
    await require_token_query(token)


def _inject_preview_html_bridge(target: Path) -> str | None:
    """Return HTML with the preview bridge, or None for untouched streaming."""
    try:
        if target.stat().st_size > _PREVIEW_INJECT_MAX_BYTES:
            return None
        html = target.read_text(encoding="utf-8")
    except (OSError, UnicodeDecodeError):
        return None
    lower = html.lower()
    idx = lower.rfind("</body>")
    if idx == -1:
        idx = lower.rfind("</html>")
    if idx == -1:
        return html + _PREVIEW_HTML_BRIDGE
    return html[:idx] + _PREVIEW_HTML_BRIDGE + html[idx:]


@router.get("/raw", dependencies=[Depends(_require_raw_access)])
def raw_file(
    path: str = Query(...),
    preview: bool = Query(False),
    root: Path = Depends(_workspace_root),
):
    """Stream a raw file using a path-bound preview ticket or legacy token.

    Everything outside the whitelists is forced to download as octet-stream.
    """
    target = safe_resolve(path, root=root)
    if not target.exists() or not target.is_file():
        raise HTTPException(status_code=404, detail="not a file")
    suffix = target.suffix.lower()
    # `no-cache` (NOT no-store) — let browsers cache but force a conditional
    # GET (If-None-Match / If-Modified-Since) every time. FileResponse still
    # sends ETag + Last-Modified, so unchanged files return 304 cheaply; the
    # moment mtime changes, the etag flips and the browser pulls the new
    # body. Without this, browsers happily served the disk-cached version on
    # every page reload (URLs identical) and edits never showed until users
    # hit the manual reload button — see 2026-05-18 dark-mode HTML report bug.
    base_headers = {
        "X-Content-Type-Options": "nosniff",
        "Cache-Control": "no-cache",
    }
    # RFC 5987: non-ASCII filenames must be URL-encoded in filename* attribute.
    # HTTP headers are latin-1 only; Chinese / emoji filenames break encode().
    from urllib.parse import quote
    disp_filename = f'filename="file{suffix}"; filename*=UTF-8\'\'{quote(target.name)}'

    if suffix in INLINE_OK_SUFFIX:
        return FileResponse(target, headers={
            **base_headers,
            "Content-Disposition": f"inline; {disp_filename}",
        })
    if suffix in SANDBOXED_INLINE_SUFFIX:
        # CSP relaxed enough for academic HTML reports (MathJax / KaTeX / highlight.js
        # from CDN, inline <script>window.MathJax = {...}</script> config blocks).
        # The server-side `sandbox allow-scripts` DIRECTIVE puts JS in a unique
        # opaque origin even when the file is opened TOP-LEVEL (URL pasted into
        # the address bar): scripts still run, but cannot act as our origin —
        # /api/* fetches become cross-origin (CORS-blocked), cookies/storage
        # are unavailable, so the query token can't be replayed against the API.
        # Previously only the frontend iframe's sandbox attribute provided this
        # isolation, which a top-level open silently bypassed.
        sandbox_headers = {
            **base_headers,
            "Content-Disposition": f"inline; {disp_filename}",
            "Content-Security-Policy": (
                "sandbox allow-scripts; "
                "default-src 'none'; "
                "script-src 'self' 'unsafe-inline' https:; "
                "style-src 'self' 'unsafe-inline' https:; "
                "img-src 'self' data: https:; "
                "font-src https: data:; "
                "connect-src 'self'; "
                "base-uri 'none'; form-action 'none'"
            ),
        }
        # Preview pane requests ?preview=1 for HTML so we can inject the image
        # and scroll-position bridge. SVG and top-level/download paths never
        # get it (no preview flag) and stream untouched.
        if preview and suffix in (".html", ".htm"):
            injected = _inject_preview_html_bridge(target)
            if injected is not None:
                return HTMLResponse(content=injected, headers=sandbox_headers)
        return FileResponse(target, headers=sandbox_headers)
    # FileResponse(filename=) sets Content-Disposition itself; use our safe one.
    return FileResponse(target, media_type="application/octet-stream", headers={
        **base_headers,
        "Content-Disposition": f"attachment; {disp_filename}",
    })


def _require_download_ticket(
    path: str = Query(...),
    ticket: str = Query(""),
    root: Path = Depends(_workspace_root),
) -> None:
    try:
        target = safe_resolve(path, root=root)
    except HTTPException:
        raise HTTPException(status_code=401, detail="invalid download ticket") from None
    if not tickets.validate(
        ticket,
        "download",
        (str(target), str(root.resolve())),
    ):
        raise HTTPException(status_code=401, detail="invalid or expired download ticket")


@router.get("/download", dependencies=[Depends(_require_download_ticket)])
def download_file(
    path: str = Query(...),
    root: Path = Depends(_workspace_root),
) -> FileResponse:
    target = safe_resolve(path, root=root)
    if not target.exists() or not target.is_file():
        raise HTTPException(status_code=404, detail="not a file")
    from urllib.parse import quote
    suffix = target.suffix.lower()
    disp = f'attachment; filename="file{suffix}"; filename*=UTF-8\'\'{quote(target.name)}'
    return FileResponse(target, media_type="application/octet-stream",
                        headers={"Content-Disposition": disp})


class WriteReq(BaseModel):
    path: str
    content: str


# Upper bound on a single editor-save payload. Generous enough for real-world
# documents (a 10 MB Markdown file is ~3 million words) but stops a runaway
# script from filling the disk via this endpoint. Matches the spirit of
# MAX_TEXT_SIZE on the read path.
MAX_WRITE_BYTES = 10 * 1024 * 1024


@router.put("/write", dependencies=[Depends(require_token)])
def write_file(req: WriteReq, root: Path = Depends(_workspace_root)) -> dict:
    """Overwrite a file at `path` with `content`. Atomic (tmpfile + rename),
    so a crash mid-write leaves the previous content intact instead of a
    truncated half-file. Capped at MAX_WRITE_BYTES to prevent the editor
    from accidentally serving as an unbounded ingest path."""
    target = safe_resolve(req.path, root=root)
    _guard_not_trash(target, root)
    if target.exists() and target.is_dir():
        raise HTTPException(status_code=400, detail="path is a directory")
    # Two-stage size gate. Each char is 1-4 UTF-8 bytes, so the upper
    # bound `len(s) * 4` is cheap (no encoding); if it already exceeds
    # the limit we reject without materializing the encoded bytes at
    # all. Only the borderline case (str length close to limit) needs
    # the precise encode. Saves ~10 MB transient RSS on a max-size
    # payload that was previously rejected anyway.
    char_len = len(req.content)
    if char_len * 4 > MAX_WRITE_BYTES:
        if char_len > MAX_WRITE_BYTES \
                or len(req.content.encode("utf-8")) > MAX_WRITE_BYTES:
            raise HTTPException(
                status_code=413,
                detail=f"content exceeds {MAX_WRITE_BYTES // (1024 * 1024)} MB limit",
            )
    with _trash_transaction(root):
        target.parent.mkdir(parents=True, exist_ok=True)
        atomic_write_text(target, req.content)
        return {"ok": True, "size": target.stat().st_size}


# Default 100 MB cap per uploaded file. Override via MUSELAB_MAX_UPLOAD_MB.
MAX_UPLOAD_BYTES = env_int("MUSELAB_MAX_UPLOAD_MB", 100, min_value=1) * 1024 * 1024
# Filename extensions that are likely to be hostile or pointless to host in
# a local workspace. Block at upload (cleaner than after-the-fact cleanup).
UPLOAD_BLOCKED_SUFFIX = {
    ".exe", ".dll", ".so", ".dylib", ".scr", ".com", ".bat", ".cmd",
    ".ps1",  # PowerShell scripts — block by default; allow via .env override later
    ".msi", ".app",
}


@router.post("/upload", dependencies=[Depends(require_token)])
async def upload(
    path: str = Form(""),
    file: UploadFile = File(...),
    root: Path = Depends(_workspace_root),
) -> dict:
    target_dir = safe_resolve(path, root=root)
    _guard_not_trash(target_dir, root)
    if not target_dir.exists() or not target_dir.is_dir():
        raise HTTPException(status_code=400, detail="target dir invalid")
    safe_name = Path(file.filename or "upload.bin").name
    # Path("." ).name and Path("..").name are both "" — those filenames
    # produced an empty safe_name → `target_dir / ""` == target_dir, and
    # the directory checks below raised 500 instead of a clean 400.
    if not safe_name or safe_name in {".", ".."}:
        raise HTTPException(status_code=400, detail="invalid filename")
    # Block dangerous extensions early.
    suffix = Path(safe_name).suffix.lower()
    if suffix in UPLOAD_BLOCKED_SUFFIX:
        raise HTTPException(status_code=400,
                             detail=f"upload blocked by extension: {suffix}")
    # Also block uploads with sensitive filenames (.env, id_rsa etc.).
    if _is_sensitive(Path(safe_name)):
        raise HTTPException(status_code=403,
                             detail="sensitive filename blocked")
    dest = target_dir / safe_name
    # Stream + enforce size cap. Write to a temporary file first, then
    # atomically rename to dest so a crash or size-exceeded abort never
    # leaves a partial file at the intended path.
    import uuid as _uuid
    tmp_path = dest.parent / f".~{dest.name}.{_uuid.uuid4().hex[:8]}.uploading"
    written = 0
    try:
        with tmp_path.open("wb") as f:
            while chunk := await file.read(1024 * 1024):
                written += len(chunk)
                if written > MAX_UPLOAD_BYTES:
                    f.close()
                    tmp_path.unlink(missing_ok=True)
                    raise HTTPException(
                        status_code=413,
                        detail=f"upload exceeds {MAX_UPLOAD_BYTES // (1024*1024)} MB cap",
                    )
                # Off-load the blocking disk write so a large multi-MB upload
                # doesn't stall the event loop chunk-by-chunk. (perf: RED —
                # files.py upload sync write)
                await asyncio.to_thread(f.write, chunk)
        # Overwrite protection: a same-name upload used to silently clobber the
        # existing file via rename() — no 409, no trash, no undo — which
        # contradicts /rename's 409 guard and the whole soft-delete design.
        # Move the old file to trash first (recoverable) only AFTER the new
        # upload fully streamed to tmp, so a failed/aborted upload never
        # destroys the existing file. A name colliding with a directory can't
        # be auto-resolved → 409.
        # The check + trash + rename runs as one cross-process transaction in
        # a thread (the lock is sync; never hold it on the event loop). Two
        # concurrent same-name uploads previously both passed exists() and
        # the later rename clobbered the earlier file with no trash entry.
        def _finalize():
            with _trash_transaction(root):
                trashed = None
                if dest.exists():
                    if dest.is_dir():
                        raise HTTPException(
                            status_code=409,
                            detail=f"a directory named {safe_name!r} already exists here",
                        )
                    logical_dest = (
                        _logical_relative_path(path) / safe_name
                    ).as_posix()
                    trashed = _move_to_trash(
                        dest,
                        root,
                        original_rel=logical_dest,
                    )
                _rename_noreplace(tmp_path, dest)
                _fsync_rename(tmp_path.parent, dest.parent)
                return trashed
        trashed = await asyncio.to_thread(_finalize)
    except HTTPException:
        tmp_path.unlink(missing_ok=True)
        raise
    except Exception:
        tmp_path.unlink(missing_ok=True)
        raise
    return {
        "ok": True,
        "path": (_logical_relative_path(path) / safe_name).as_posix(),
        "size": dest.stat().st_size,
        # Non-null when an existing same-name file was moved to trash so the
        # frontend can surface "replaced (old version in trash)".
        "replaced_trash_id": (trashed or {}).get("trash_id"),
    }



def _reject_workspace_root_mutation(root: Path, relative: str) -> Path:
    logical = _logical_relative_path(relative)
    parts = tuple(part for part in logical.parts if part not in {"", "."})
    if not parts:
        raise HTTPException(
            status_code=400,
            detail="cannot delete a workspace root",
        )
    if TRASH_DIR_NAME in parts:
        raise HTTPException(
            status_code=400,
            detail="cannot mutate the dustbin through the file endpoint",
        )
    candidate = Path(root).absolute().joinpath(*parts)
    registered = {
        Path(workspace).absolute()
        for workspace in workspace_registry.paths()
    }
    if any(
        workspace == candidate or candidate in workspace.parents
        for workspace in registered
    ):
        raise HTTPException(
            status_code=400,
            detail="cannot delete a workspace root",
        )
    return Path(*parts)


def _remove_tombstone_at(trash_fd: int, name: str) -> None:
    """Delete a staged tree outside the transaction lock, without following it."""
    kind = _entry_kind_at(trash_fd, name)
    if kind == "missing":
        return
    if kind == "directory":
        shutil.rmtree(name, dir_fd=trash_fd)
    elif kind in {"file", "symlink"}:
        os.unlink(name, dir_fd=trash_fd)
    else:
        raise UnsafePrivatePath("purge tombstone is unsafe")
    if _entry_kind_at(trash_fd, name) != "missing":
        raise OSError("purge tombstone was not removed")
    os.fsync(trash_fd)


def _permanent_delete_anchored(relative: str, root: Path) -> bool:
    """Unlink a leaf or stage a directory, never following the final symlink."""
    logical = _reject_workspace_root_mutation(root, relative)
    cleanup_fd = -1
    tombstone_name = ""
    with _trash_transaction(root) as trash:
        anchored = _anchor_for_path(trash)
        if anchored is None:
            raise UnsafePrivatePath("trash transaction anchor is unavailable")
        anchor = anchored[0]
        try:
            parent_fd, parent_parts, name = _open_workspace_parent(
                anchor,
                logical.as_posix(),
            )
        except OSError as exc:
            raise UnsafePrivatePath(
                "delete path contains an unsafe ancestor") from exc
        try:
            identity = _entry_identity_at(parent_fd, name)
            if identity is None:
                return False
            kind = str(identity["kind"])
            if kind in {"file", "symlink"}:
                os.unlink(name, dir_fd=parent_fd)
                _fsync_open_directory(
                    parent_fd,
                    root / logical.parent,
                )
                return True
            if kind != "directory":
                raise UnsafePrivatePath(
                    "permanent delete target is unsafe")

            directory_info = os.stat(
                name,
                dir_fd=parent_fd,
                follow_symlinks=False,
            )
            if _stat_identity(directory_info) != identity:
                raise UnsafePrivatePath(
                    "permanent delete target changed")
            original_mode = stat.S_IMODE(directory_info.st_mode)

            tombstone_name = _permanent_tombstone_name()
            source_parent_identity = _fd_identity(parent_fd)
            trash_parent_identity = _fd_identity(anchor.trash_fd)
            _rename_noreplace(
                root / logical,
                trash / tombstone_name,
                source_parent_identity=source_parent_identity,
                destination_parent_identity=trash_parent_identity,
                source_parent_fd=parent_fd,
                destination_parent_fd=anchor.trash_fd,
            )
            if (
                not _directory_is_root_reachable(
                    anchor,
                    parent_parts,
                    parent_fd,
                )
                or not _directory_is_root_reachable(
                    anchor,
                    (TRASH_DIR_NAME,),
                    anchor.trash_fd,
                )
            ):
                try:
                    _rename_noreplace(
                        trash / tombstone_name,
                        root / logical,
                        source_parent_identity=trash_parent_identity,
                        destination_parent_identity=source_parent_identity,
                        source_parent_fd=anchor.trash_fd,
                        destination_parent_fd=parent_fd,
                    )
                except OSError:
                    pass
                raise UnsafePrivatePath(
                    "delete parent is no longer reachable "
                    "from workspace root"
                )
            try:
                _restore_mode_at(
                    anchor.trash_fd,
                    tombstone_name,
                    identity,
                    0o700,
                )
            except (OSError, UnsafePrivatePath):
                try:
                    _rename_noreplace(
                        trash / tombstone_name,
                        root / logical,
                        source_parent_identity=trash_parent_identity,
                        destination_parent_identity=source_parent_identity,
                        source_parent_fd=anchor.trash_fd,
                        destination_parent_fd=parent_fd,
                    )
                    _restore_mode_at(
                        parent_fd,
                        name,
                        identity,
                        original_mode,
                    )
                except (OSError, UnsafePrivatePath):
                    pass
                raise
            cleanup_fd = os.dup(anchor.trash_fd)
        finally:
            os.close(parent_fd)

    try:
        try:
            _remove_tombstone_at(cleanup_fd, tombstone_name)
        except PermissionError:
            raise
        except OSError as exc:
            raise _TombstoneCleanupError(
                "permanent delete cleanup is deferred") from exc
    finally:
        if cleanup_fd >= 0:
            os.close(cleanup_fd)

    return True


class DeleteReq(BaseModel):
    path: str


@router.delete("/delete", dependencies=[Depends(require_token)])
def delete(
    req: DeleteReq,
    permanent: bool = Query(default=False),
    root: Path = Depends(_workspace_root),
) -> dict:
    """Delete one root-relative entry without following mutation-time links."""
    root = Path(root).absolute()
    logical = _reject_workspace_root_mutation(root, req.path)
    target = root / logical

    if permanent:
        try:
            existed = _permanent_delete_anchored(
                logical.as_posix(),
                root,
            )
        except PermissionError:
            raise HTTPException(
                status_code=403,
                detail="permanent delete failed",
                headers={
                    "X-MuseLab-Error-Code": "permission_denied",
                },
            ) from None
        except _TombstoneCleanupError:
            raise HTTPException(
                status_code=500,
                detail=(
                    "permanent delete committed; "
                    "physical cleanup is deferred"
                ),
                headers={
                    "X-MuseLab-Error-Code": "partial_delete",
                },
            ) from None
        except (OSError, UnsafePrivatePath):
            raise HTTPException(
                status_code=500,
                detail="permanent delete failed; the target may still exist",
                headers={
                    "X-MuseLab-Error-Code": "io_error",
                },
            ) from None
        if not existed:
            raise HTTPException(status_code=404, detail="not found")
        return {"ok": True, "permanent": True}

    try:
        manifest = _move_to_trash(
            target,
            root,
            original_rel=logical.as_posix(),
        )
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail="not found") from None
    except UnsafePrivatePath as exc:
        raise HTTPException(
            status_code=400,
            detail=str(exc),
        ) from None
    return {
        "ok": True,
        "permanent": False,
        "trash_id": manifest["trash_id"],
        "manifest": manifest,
    }

# ============================================================
# Trash management endpoints
# ============================================================
@router.get("/trash/list", dependencies=[Depends(require_token)])
def trash_list(root: Path = Depends(_workspace_root)) -> dict:
    """List trash explicitly; storage faults are reported as degraded."""
    try:
        _gc_trash_auxiliary(root)
        items = _list_trash(root)
    except (OSError, UnsafePrivatePath):
        raise HTTPException(
            status_code=503,
            detail="trash storage is temporarily degraded",
            headers={
                "X-MuseLab-Error-Code": "trash_degraded",
            },
        ) from None
    return {
        "items": items,
        "total_size": sum(int(i.get("size") or 0) for i in items),
        "ttl_days": _TRASH_TTL_DAYS,
        "degraded": False,
    }

class TrashIdReq(BaseModel):
    trash_id: str


@router.post("/trash/restore", dependencies=[Depends(require_token)])
def trash_restore(req: TrashIdReq, root: Path = Depends(_workspace_root)) -> dict:
    """Resume or start an idempotent, crash-recoverable restore."""
    if not _valid_trash_id(req.trash_id):
        raise HTTPException(status_code=400, detail="invalid trash_id")

    root = Path(root).absolute()
    with _trash_transaction(root) as d:
        anchored = _anchor_for_path(d)
        if anchored is None:
            raise HTTPException(
                status_code=500,
                detail="trash transaction anchor is unavailable",
            )
        anchor = anchored[0]
        manifest_path = d / f"{req.trash_id}.json"
        data = _repair_trash_item(d, manifest_path)
        if data is None:
            receipt = _read_restore_receipt(d, req.trash_id)
            if receipt is None:
                raise HTTPException(
                    status_code=404,
                    detail="trash item not found",
                )
            return {
                "ok": True,
                "restored_path": receipt["original_path"],
            }

        state = _trash_state(data)
        identity = data.get("payload_identity")
        if state is None or not _valid_identity(identity):
            raise HTTPException(
                status_code=500,
                detail="trash manifest is invalid",
            )
        orig_rel = data.get("original_path")
        if not isinstance(orig_rel, str) or not orig_rel:
            raise HTTPException(
                status_code=500,
                detail="manifest missing original_path",
            )
        try:
            logical = _logical_relative_path(orig_rel)
        except HTTPException:
            raise HTTPException(
                status_code=500,
                detail="trash manifest path is invalid",
            ) from None
        if INTERNAL_DIR_NAME in logical.parts:
            raise HTTPException(
                status_code=500,
                detail="trash manifest path is invalid",
            )
        orig = root / logical

        if state == _TRASH_RESTORED:
            try:
                _persist_restore_completion(manifest_path, data)
            except (OSError, UnsafePrivatePath):
                raise HTTPException(
                    status_code=500,
                    detail="restore completion could not be persisted",
                ) from None
            return {"ok": True, "restored_path": orig_rel}

        payload_kind = _entry_kind_at(anchor.trash_fd, req.trash_id)
        original_mode = data.get("original_mode")
        if (
            type(original_mode) is not int
            or not 0 <= original_mode <= 0o7777
        ):
            original_mode = None

        if payload_kind == "missing" and state == _TRASH_RESTORE_PREPARED:
            try:
                parent_fd, parent_parts, original_name = (
                    _open_workspace_parent(anchor, orig_rel))
            except (OSError, HTTPException):
                raise HTTPException(
                    status_code=500,
                    detail="restore transaction requires recovery",
                ) from None
            try:
                if _entry_identity_at(parent_fd, original_name) != identity:
                    raise HTTPException(
                        status_code=500,
                        detail="restore transaction requires recovery",
                    )
                restore_parent_identity = data.get(
                    "restore_parent_identity")
                trash_parent_identity = data.get("trash_parent_identity")
                if (
                    _fd_identity(parent_fd) != restore_parent_identity
                    or _fd_identity(anchor.trash_fd)
                    != trash_parent_identity
                    or not _directory_is_root_reachable(
                        anchor,
                        parent_parts,
                        parent_fd,
                    )
                    or not _directory_is_root_reachable(
                        anchor,
                        (TRASH_DIR_NAME,),
                        anchor.trash_fd,
                    )
                ):
                    raise HTTPException(
                        status_code=500,
                        detail="restore transaction requires recovery",
                    )
                try:
                    _fsync_rename_fds(
                        anchor.trash_fd,
                        d,
                        parent_fd,
                        orig.parent,
                    )
                    if original_mode is not None:
                        _restore_mode_at(
                            parent_fd,
                            original_name,
                            identity,
                            original_mode,
                        )
                except (OSError, UnsafePrivatePath):
                    raise HTTPException(
                        status_code=500,
                        detail=(
                            "restore durability barrier could not be persisted"
                        ),
                    ) from None
                try:
                    _persist_restore_completion(manifest_path, data)
                except (OSError, UnsafePrivatePath):
                    raise HTTPException(
                        status_code=500,
                        detail="restore completion could not be persisted",
                    ) from None
                return {"ok": True, "restored_path": orig_rel}
            finally:
                os.close(parent_fd)

        if payload_kind not in {"directory", "file"}:
            raise HTTPException(
                status_code=404,
                detail="trash payload missing",
            )
        if _entry_identity_at(anchor.trash_fd, req.trash_id) != identity:
            raise HTTPException(
                status_code=500,
                detail="trash payload changed",
            )

        try:
            parent_fd, parent_parts, original_name = _open_workspace_parent(
                anchor,
                orig_rel,
                create=True,
            )
        except OSError:
            raise HTTPException(
                status_code=400,
                detail="restore path contains an unsafe ancestor",
            ) from None
        try:
            if _entry_kind_at(parent_fd, original_name) != "missing":
                if state == _TRASH_RESTORE_PREPARED:
                    try:
                        _write_trash_manifest(
                            manifest_path,
                            {**data, _TRASH_STATE_KEY: _TRASHED},
                        )
                    except (OSError, UnsafePrivatePath):
                        pass
                raise HTTPException(
                    status_code=409,
                    detail=(
                        "original path is occupied; "
                        "rename or clear it first"
                    ),
                )

            restore_parent_identity = _fd_identity(parent_fd)
            trash_parent_identity = _fd_identity(anchor.trash_fd)
            prepared = {
                **data,
                _TRASH_STATE_KEY: _TRASH_RESTORE_PREPARED,
                "restore_parent_identity": restore_parent_identity,
                "trash_parent_identity": trash_parent_identity,
            }
            if prepared != data:
                _write_trash_manifest(manifest_path, prepared)
                data = prepared

            try:
                _rename_noreplace(
                    d / req.trash_id,
                    orig,
                    source_parent_identity=trash_parent_identity,
                    destination_parent_identity=restore_parent_identity,
                    source_parent_fd=anchor.trash_fd,
                    destination_parent_fd=parent_fd,
                )
            except FileExistsError:
                if (
                    _entry_kind_at(anchor.trash_fd, req.trash_id)
                    == "missing"
                    and _entry_identity_at(parent_fd, original_name)
                    == identity
                ):
                    pass
                else:
                    if (
                        _entry_identity_at(
                            anchor.trash_fd,
                            req.trash_id,
                        )
                        == identity
                    ):
                        try:
                            _write_trash_manifest(
                                manifest_path,
                                {**data, _TRASH_STATE_KEY: _TRASHED},
                            )
                        except (OSError, UnsafePrivatePath):
                            pass
                    raise HTTPException(
                        status_code=409,
                        detail=(
                            "original path is occupied; "
                            "rename or clear it first"
                        ),
                    ) from None

            parents_reachable = (
                _directory_is_root_reachable(
                    anchor,
                    parent_parts,
                    parent_fd,
                )
                and _directory_is_root_reachable(
                    anchor,
                    (TRASH_DIR_NAME,),
                    anchor.trash_fd,
                )
            )
            if not parents_reachable:
                try:
                    _rename_noreplace(
                        orig,
                        d / req.trash_id,
                        source_parent_identity=restore_parent_identity,
                        destination_parent_identity=trash_parent_identity,
                        source_parent_fd=parent_fd,
                        destination_parent_fd=anchor.trash_fd,
                    )
                    _write_trash_manifest(
                        manifest_path,
                        {**data, _TRASH_STATE_KEY: _TRASHED},
                    )
                except (OSError, UnsafePrivatePath):
                    pass
                raise HTTPException(
                    status_code=500,
                    detail=(
                        "restore parent is no longer reachable "
                        "from workspace root"
                    ),
                )

            if _entry_identity_at(parent_fd, original_name) != identity:
                raise HTTPException(
                    status_code=500,
                    detail="restored payload changed",
                )
            if original_mode is not None:
                _restore_mode_at(
                    parent_fd,
                    original_name,
                    identity,
                    original_mode,
                )
            else:
                payload_fd = os.open(
                    original_name,
                    os.O_RDONLY
                    | getattr(os, "O_CLOEXEC", 0)
                    | getattr(os, "O_NOFOLLOW", 0)
                    | getattr(os, "O_NONBLOCK", 0),
                    dir_fd=parent_fd,
                )
                try:
                    os.fsync(payload_fd)
                finally:
                    os.close(payload_fd)
            try:
                _persist_restore_completion(manifest_path, data)
            except (OSError, UnsafePrivatePath):
                raise HTTPException(
                    status_code=500,
                    detail="restore completion could not be persisted",
                ) from None
            return {"ok": True, "restored_path": orig_rel}
        finally:
            os.close(parent_fd)

@router.delete("/trash/purge", dependencies=[Depends(require_token)])
def trash_purge(req: TrashIdReq, root: Path = Depends(_workspace_root)) -> dict:
    """Permanently delete one trash item with restore/purge linearization."""
    if not _valid_trash_id(req.trash_id):
        raise HTTPException(status_code=400, detail="invalid trash_id")
    try:
        outcome, cleanup_deferred = _purge_one(req.trash_id, root)
    except PermissionError:
        raise HTTPException(
            status_code=403,
            detail="trash purge failed",
            headers={"X-MuseLab-Error-Code": "permission_denied"},
        ) from None
    except (OSError, UnsafePrivatePath):
        raise HTTPException(
            status_code=500,
            detail="trash purge could not be committed",
            headers={"X-MuseLab-Error-Code": "io_error"},
        ) from None

    if outcome == "missing":
        raise HTTPException(
            status_code=404,
            detail="trash item not found",
        )
    if outcome == "restored":
        raise HTTPException(
            status_code=409,
            detail="trash item was already restored",
        )
    return {
        "ok": True,
        "idempotent": outcome == "already_purged",
        "cleanup_deferred": cleanup_deferred,
    }

@router.delete("/trash/empty", dependencies=[Depends(require_token)])
def trash_empty(root: Path = Depends(_workspace_root)) -> dict:
    """Purge the current linearized trash snapshot."""
    try:
        items = _list_trash(root)
    except (OSError, UnsafePrivatePath):
        raise HTTPException(
            status_code=503,
            detail="trash storage is temporarily degraded",
            headers={"X-MuseLab-Error-Code": "trash_degraded"},
        ) from None
    count = 0
    failed = 0
    cleanup_deferred = 0
    for item in items:
        tid = str(item.get("trash_id") or "")
        try:
            outcome, deferred = _purge_one(tid, root)
            if outcome in {"purged", "already_purged"}:
                count += 1
                cleanup_deferred += int(deferred)
        except (OSError, UnsafePrivatePath):
            failed += 1
    if failed:
        raise HTTPException(
            status_code=500,
            detail={
                "message": "trash cleanup was only partially completed",
                "purged": count,
                "failed": failed,
                "cleanup_deferred": cleanup_deferred,
            },
            headers={"X-MuseLab-Error-Code": "partial_delete"},
        )
    return {
        "ok": True,
        "purged": count,
        "failed": 0,
        "cleanup_deferred": cleanup_deferred,
    }

class MkdirReq(BaseModel):
    path: str


@router.post("/mkdir", dependencies=[Depends(require_token)])
def mkdir(req: MkdirReq, root: Path = Depends(_workspace_root)) -> dict:
    target = safe_resolve(req.path, root=root)
    _guard_not_trash(target, root)
    with _trash_transaction(root):
        _mkdir_durable(target)
    return {"ok": True, "path": _logical_relative_path(req.path).as_posix()}


class RenameReq(BaseModel):
    src: str
    dst: str   # relative to ROOT


@router.post("/rename", dependencies=[Depends(require_token)])
def rename(req: RenameReq, root: Path = Depends(_workspace_root)) -> dict:
    src = safe_resolve(req.src, root=root)
    dst = safe_resolve(req.dst, root=root)
    _guard_not_trash(src, root)
    _guard_not_trash(dst, root)
    # Atomic check+rename under the workspace lock — without it, two
    # concurrent renames onto the same dst both pass the exists() probe
    # and the later rename silently replaces the earlier file (TOCTOU).
    with _trash_transaction(root):
        if not src.exists():
            raise HTTPException(status_code=404, detail="source not found")
        if dst.exists():
            raise HTTPException(status_code=409, detail="destination already exists")
        dst.parent.mkdir(parents=True, exist_ok=True)
        _rename_noreplace(src, dst)
        _fsync_rename(src.parent, dst.parent)
    return {"ok": True, "path": _logical_relative_path(req.dst).as_posix()}


# ============================================================
# Copy as .bak — the only "copy" we expose. Frontend supports both a
# Ctrl+C / Ctrl+V flow and a "Copy as .bak" context-menu item; both
# land here. Files only (directories rejected with 400) and the new
# name is server-side derived so the API can't be tricked into
# clobbering anything: <stem><suffix>.bak, .bak.2, .bak.3 … picking
# the first non-existing name in the target directory.
# ============================================================
class CopyBakReq(BaseModel):
    src: str
    # Where to drop the .bak. Empty / omitted = same directory as src
    # (covers the "Ctrl+D / context-menu duplicate" path). Frontend
    # passes the currently-selected directory for cross-dir paste.
    dst_dir: str = ""


def _next_bak_name(parent: Path, original_name: str) -> str:
    """Pick the first non-existing <original_name>.bak[.N] under parent.

    Always appends `.bak`, even if original already ends in `.bak` (so
    `foo.txt.bak` → `foo.txt.bak.bak`). Increments via `.bak.2`, `.bak.3`,
    … This keeps the rule mechanical and predictable instead of trying
    to be clever about "already a backup".
    """
    base = f"{original_name}.bak"
    if not (parent / base).exists():
        return base
    # .bak exists → try .bak.2, .bak.3, … Cap at a sane upper bound so a
    # pathological directory full of .bak.N siblings can't hang the call.
    for i in range(2, 1000):
        cand = f"{original_name}.bak.{i}"
        if not (parent / cand).exists():
            return cand
    raise HTTPException(status_code=409, detail="too many .bak siblings")


@router.post("/copy-bak", dependencies=[Depends(require_token)])
def copy_bak(req: CopyBakReq, root: Path = Depends(_workspace_root)) -> dict:
    src = safe_resolve(req.src, root=root)
    _guard_not_trash(src, root)
    if not src.exists():
        raise HTTPException(status_code=404, detail="source not found")
    # Files only. Directory copy is a different beast (shutil.copytree,
    # permission edge cases, can be slow on big trees) — out of scope for
    # the .bak shortcut.
    if src.is_dir():
        raise HTTPException(status_code=400, detail="directories not supported")
    if req.dst_dir:
        parent = safe_resolve(req.dst_dir, root=root)
        if not parent.exists() or not parent.is_dir():
            raise HTTPException(status_code=404, detail="dst_dir not found")
    else:
        parent = src.parent
    logical_parent = (
        _logical_relative_path(req.dst_dir)
        if req.dst_dir
        else _logical_relative_path(req.src).parent
    )

    # Copy potentially large content outside the global destination lock. The
    # temporary lives beside the final file, so the later hard-link commit is
    # same-filesystem, atomic, and refuses to overwrite an external writer.
    while True:
        tmp = parent / f".~{src.name}.{secrets.token_hex(8)}.copying"
        try:
            with tmp.open("xb"):
                pass
            break
        except FileExistsError:
            continue
    try:
        shutil.copy2(src, tmp)
        with _trash_transaction(root):
            while True:
                new_name = _next_bak_name(parent, src.name)
                dst = parent / new_name
                _guard_not_trash(dst, root)
                # safe_resolve the final path so the anti-traversal guard fires
                # for the destination too.
                dst_rel = (logical_parent / new_name).as_posix()
                safe_resolve(dst_rel, allow_sensitive=True, root=root)
                # The appended `.bak[.N]` suffix means `_is_sensitive`
                # (exact-name / suffix match) never fires on the destination —
                # `secrets.env.bak` wouldn't match `.env.*`. Strip the trailing
                # suffix chain and re-check the underlying name.
                underlying = new_name
                while True:
                    match = re.match(r"^(.*)\.bak(?:\.\d+)?$", underlying)
                    if not match:
                        break
                    underlying = match.group(1)
                if _is_sensitive(Path(underlying)):
                    raise HTTPException(
                        status_code=403,
                        detail="sensitive file blocked",
                    )
                try:
                    os.link(tmp, dst)
                except FileExistsError:
                    # An external writer can race the in-process lock. Retry the
                    # server-derived name; hard-link never replaces its file.
                    continue
                break
    finally:
        tmp.unlink(missing_ok=True)
    return {"ok": True, "path": dst_rel, "name": new_name}


SEARCH_IGNORE = {".git", "node_modules", "__pycache__", ".venv", "venv",
                 ".cache", ".pytest_cache", ".mypy_cache", "dist", "build",
                 INTERNAL_DIR_NAME,
                 # Trash always excluded from search/grep regardless of
                 # show_hidden — otherwise a search for "foo" surfaces every
                 # version of foo.md the user has ever deleted, which the
                 # trash UI is purpose-built to present separately.
                 TRASH_DIR_NAME}

GREP_EXTS = {".md", ".markdown", ".txt", ".html", ".htm", ".json", ".yaml", ".yml",
             ".py", ".js", ".ts", ".css", ".sh", ".toml", ".ini", ".csv", ".sql",
             ".log", ".xml", ".rst", ".tex",
             ".c", ".cc", ".cpp", ".cxx", ".h", ".hh", ".hpp", ".hxx"}


MAX_GREP_FILE_SIZE = 1_000_000   # 1MB per file — skip large files
MAX_GREP_TIME_SEC = 8            # soft time budget


# ============================================================
# Directory-listing cache for search / grep (2026-05-28).
#
# Both endpoints used to call `os.walk(ROOT)` from scratch on every
# request — for a 3000-file archive this is ~200-500 ms of pure
# scandir + per-entry stat() overhead even before any file content is
# touched. Most directories don't change between calls, so we cache
# `{dir_path: (mtime, [(name, is_dir), ...])}` keyed by directory
# mtime; a hit skips the scandir entirely. Filesystem mtime semantics
# guarantee a directory's mtime updates iff its entry list changes
# (add / remove / rename), which is exactly the cache invalidation
# trigger we need. File CONTENT changes do NOT bump parent mtime on
# ext4 / btrfs, but we deliberately don't cache file size / mtime —
# callers that need those `stat()` per file independently (fast).
#
# Thread-safe because FastAPI runs sync route handlers in a thread
# pool; two concurrent /api/files/search calls would race on the dict
# without _DIR_CACHE_LOCK.
# ============================================================
_DIR_CACHE: dict[str, tuple[float, list[tuple[str, bool]]]] = {}
_DIR_CACHE_LOCK = threading.Lock()
# Bound the cache so a pathological archive (millions of dirs) can't
# OOM the process. Typical local workspaces have 50-500 dirs total,
# so this rarely matters. On overflow we drop the oldest insertion.
_DIR_CACHE_MAX = 5000


def _cached_walk(root: Path, ignore: set[str], show_hidden: bool):
    """Generator that mimics `os.walk(root)` but caches each directory's
    entry list by mtime, and applies the `ignore` / `show_hidden`
    filters in one pass.

    Yields `(dirpath: Path, dirnames: list[str], filenames: list[str])`.
    Callers that previously mutated `dirnames[:]` to filter no longer
    need to — this function pre-filters."""
    # Explicit stack so we control descent order + can interleave the
    # cache hit/miss path cleanly. DFS by `.pop()` matches os.walk's
    # top-down behavior for callers that bail on a time budget.
    stack: list[Path] = [root]
    while stack:
        dp = stack.pop()
        try:
            dir_mtime = dp.stat().st_mtime
        except OSError:
            continue
        key = str(dp)
        with _DIR_CACHE_LOCK:
            cached = _DIR_CACHE.get(key)
        entries: list[tuple[str, bool]]
        if cached is not None and cached[0] == dir_mtime:
            entries = cached[1]
        else:
            entries = []
            try:
                with os.scandir(dp) as it:
                    for de in it:
                        try:
                            entries.append(
                                (de.name, de.is_dir(follow_symlinks=False)))
                        except OSError:
                            continue
            except OSError:
                continue
            with _DIR_CACHE_LOCK:
                # Bound the cache via FIFO eviction (insertion-ordered
                # dict). Not strict LRU but cheap and good enough.
                if len(_DIR_CACHE) > _DIR_CACHE_MAX:
                    try:
                        _DIR_CACHE.pop(next(iter(_DIR_CACHE)))
                    except StopIteration:
                        pass
                _DIR_CACHE[key] = (dir_mtime, entries)
        dirnames: list[str] = []
        filenames: list[str] = []
        for name, is_dir in entries:
            if name in ignore:
                continue
            if not show_hidden and name.startswith("."):
                continue
            if is_dir:
                dirnames.append(name)
                stack.append(dp / name)
            else:
                filenames.append(name)
        yield dp, dirnames, filenames


# Concurrency gate: each grep can burn up to MAX_GREP_TIME_SEC of CPU while
# holding a threadpool thread. Without a cap, a few rapid keystrokes (or two
# devices searching at once) stack full-archive scans and starve the pool —
# every other endpoint (chat, sessions) stalls behind them. Two concurrent
# scans is plenty for interactive use; excess requests fail fast with 429
# rather than queueing (the UI debounces and just issues a fresh search).
_GREP_GATE = threading.BoundedSemaphore(2)


@router.get("/grep", dependencies=[Depends(require_token)])
def grep(
    q: str,
    limit: int = 50,
    show_hidden: bool = False,
    root: Path = Depends(_workspace_root),
) -> dict:
    """Cross-platform full-text search (pure Python, no grep dependency).
    Uses `_cached_walk` so the directory-listing phase is O(changed-dirs)
    instead of O(all-dirs) — repeat searches on a quiet archive only stat
    file contents, not the directory structure itself."""
    if not _GREP_GATE.acquire(blocking=False):
        raise HTTPException(429, "search busy — try again")
    try:
        return _grep_impl(q, limit, show_hidden, root)
    finally:
        _GREP_GATE.release()


def _grep_impl(
    q: str,
    limit: int,
    show_hidden: bool,
    root: Path | None = None,
) -> dict:
    root = _root_or_default(root)
    q_lower = q.strip().lower()
    # Minimum query length: a single character matches nearly every file
    # and always runs the full archive scan to the 8s time budget while
    # holding a threadpool thread hostage. Short queries early-return empty
    # (the UI debounces and only the user typing 1 char hits this).
    if len(q_lower) < 2:
        return {"hits": []}
    hits: list[dict] = []
    started = time.monotonic()
    timed_out = False
    # ROOT.resolve() is a loop invariant — hoist it out of the per-file loop
    # so the symlink-escape guard doesn't re-resolve ROOT once per candidate
    # file (was N stat-resolves per search).
    root_real = root.resolve()
    for dirpath, _dirnames, filenames in _cached_walk(
            root, SEARCH_IGNORE, show_hidden):
        if time.monotonic() - started > MAX_GREP_TIME_SEC:
            timed_out = True
            break
        for fname in filenames:
            # 隐藏文件即使没扩展名也允许 grep（用户主动开了 show_hidden 说明想看）
            if Path(fname).suffix.lower() not in GREP_EXTS and not (show_hidden and fname.startswith(".")):
                continue
            full = Path(dirpath) / fname
            try:
                # Symlink escape guard: resolve() follows symlinks, so a file
                # named `notes.md` inside ROOT pointing at /etc/passwd would
                # otherwise be opened and grepped. Confirm the real target is
                # still under ROOT before reading. Also run _is_sensitive on the
                # RESOLVED path so a symlink masking a `.env`/`*.pem` target is
                # caught (name-only check misses that).
                try:
                    resolved = full.resolve()
                except (OSError, ValueError):
                    continue
                if root_real != resolved and root_real not in resolved.parents:
                    continue
                if _is_sensitive(full) or _is_sensitive(resolved):
                    continue
                # File-level stat IS NOT cached (file content changes
                # don't bump parent dir mtime on ext4/btrfs, so a cached
                # size would lie). One stat per candidate is sub-µs.
                if full.stat().st_size > MAX_GREP_FILE_SIZE:
                    continue
                with full.open("r", encoding="utf-8", errors="replace") as f:
                    for i, line in enumerate(f, 1):
                        if q_lower in line.lower():
                            try:
                                rel = str(full.relative_to(root))
                            except ValueError:
                                continue
                            hits.append({
                                "path": rel,
                                "name": fname,
                                "line": i,
                                "snippet": line.strip()[:200],
                            })
                            if len(hits) >= limit:
                                return {"hits": hits, "truncated": True}
            except OSError:
                continue
            if time.monotonic() - started > MAX_GREP_TIME_SEC:
                timed_out = True
                break
        if timed_out:
            break
    return {"hits": hits, "truncated": timed_out}


@router.get("/search", dependencies=[Depends(require_token)])
def search(
    q: str,
    limit: int = 100,
    show_hidden: bool = False,
    exact: bool = False,
    root: Path = Depends(_workspace_root),
) -> dict:
    """Filename / dirname search, optionally matching the whole name.

    Uses the same `_cached_walk` win as grep — bigger here in relative
    terms because search only reads names (no file content), so the
    directory-listing IS the entire cost. Repeat searches over a quiet
    archive drop from ~200 ms to ~20 ms on a 3000-file tree.
    """
    q_lower = q.strip().lower()
    if not q_lower:
        return {"entries": []}
    hits: list[dict] = []
    for dirpath, dirnames, filenames in _cached_walk(
            root, SEARCH_IGNORE, show_hidden):
        for name in dirnames + filenames:
            name_lower = name.lower()
            matches_query = (
                name_lower == q_lower if exact else q_lower in name_lower
            )
            if matches_query:
                full = Path(dirpath) / name
                try:
                    stat = full.stat()
                except OSError:
                    continue
                # is_dir() was called twice per hit (once for the flag, once
                # for the size branch) — each is a syscall. Compute once.
                is_dir = full.is_dir()
                hits.append({
                    "name": name,
                    "path": str(full.relative_to(root)),
                    "is_dir": is_dir,
                    "size": stat.st_size if not is_dir else 0,
                    "mtime": stat.st_mtime,
                })
                if len(hits) >= limit:
                    return {"entries": hits, "truncated": True}
    return {"entries": hits, "truncated": False}
