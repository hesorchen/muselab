import os
import errno
import threading
import base64
from collections import deque
from contextlib import contextmanager, suppress
import hashlib
import inspect
import json
import asyncio
import re
import sys
import sqlite3
import shutil
import subprocess
import tempfile
import time
import urllib.parse
import uuid
from dataclasses import dataclass, field as dataclass_field
from datetime import UTC, datetime
from pathlib import Path
from typing import Any, Callable, Iterable, Literal, get_args
from fastapi import (
    APIRouter,
    BackgroundTasks,
    Depends,
    Query,
    HTTPException,
    UploadFile,
    File,
    Request,
    Response,
)
from fastapi.responses import FileResponse
from sse_starlette.sse import EventSourceResponse, ServerSentEvent
from pydantic import BaseModel, Field
from claude_agent_sdk import (
    ClaudeSDKClient, ClaudeAgentOptions,
    AssistantMessage, UserMessage, TextBlock, ThinkingBlock, ResultMessage,
    ConversationResetMessage,
    ToolUseBlock, ToolResultBlock, StreamEvent,
    TaskStartedMessage, TaskProgressMessage, TaskNotificationMessage,
    TaskUpdatedMessage, TERMINAL_TASK_STATUSES,
    RateLimitEvent, SystemMessage,
    ClaudeSDKError, ResultError,
    ThinkingConfigEnabled, ThinkingConfigDisabled, ThinkingConfigAdaptive,
    EffortLevel,
    get_session_messages,
    project_key_for_directory,
    delete_session as sdk_delete_session,
    rename_session as sdk_rename_session,
    tag_session as sdk_tag_session,
    fork_session as sdk_fork_session,
)
from claude_agent_sdk.types import HookMatcher, PermissionMode
from .auth import require_token, require_token_header_or_query
from .capability_tickets import tickets
from .settings import (
    ROOT,
    MODEL,
    atomic_write_text,
    ducc_cli_wrapper,
    env_float,
    env_int,
    is_chinese_locale,
    locate_ducc_executable,
)
from . import sessions as sess
from . import endpoints
from . import context_recovery
from . import chat_history
from . import chat_presentation
from . import chat_overlays
from . import chat_runtime
from . import chat_subagents
from . import chat_successor
from . import hook_settings
from . import hook_traces
from . import sdk_lifecycle
from . import transcript_index as transcript_idx
from .task_summaries import normalize_task_summary_fields
from .imagegen_job_store import ImagegenJobStore
from .workspaces import (
    registry as workspace_registry,
    resolve_workspace_root,
)
from .ask_user_question import (
    ANSWER_TIMEOUT_S, register_session_queue, unregister_session_queue,
    submit_answer,
)
from . import permission_request as perm
from . import memory_client as mem0
from . import observability as obs
from .private_storage import (
    UnsafePrivatePath,
    ensure_private_directory,
    ensure_private_regular_file,
    private_path_kind,
    repair_private_path,
    write_private_bytes,
)
from .attachment_queue_store import (
    DurableAttachmentError,
    DurableAttachmentStore,
)
from .sdk_compat import (
    CommandLifecycleMessage,
    MuseLabSDKClient,
    UnsignedThinkingCompatibleClient,
)

# Compatibility export: tests and local tooling construct durable interrupted
# snapshots through the historical chat-module schema constant.
_CANCELLED_TURN_SNAPSHOT_SCHEMA = chat_overlays._CANCELLED_TURN_SNAPSHOT_SCHEMA

# Valid permission modes, derived from the SDK's PermissionMode literal so
# the whitelist tracks SDK upgrades automatically. External strings (query
# params, queue items, tickets) flow into ClaudeAgentOptions / client
# launch contract — a typo'd or stale value would fail the SDK connect (or
# worse, silently diverge UI state from the real gate), so entry points must
# normalize through _validate_permission().
_VALID_PERMISSION_MODES: frozenset = frozenset(get_args(PermissionMode))


def _validate_permission(permission: str) -> str:
    """Return `permission` if it's a valid SDK PermissionMode, else raise
    HTTPException(400). Empty string falls back to bypassPermissions (the
    historical default for callers that never sent the param)."""
    p = (permission or "").strip()
    if not p:
        return "bypassPermissions"
    if p not in _VALID_PERMISSION_MODES:
        raise HTTPException(
            400, f"invalid permission mode: {p!r} "
                 f"(expected one of {sorted(_VALID_PERMISSION_MODES)})")
    return p


# Serialises CLAUDE_CONFIG_DIR overrides. The SDK's get_session_messages()
# has no explicit config-dir parameter — it reads the PROCESS-GLOBAL
# os.environ["CLAUDE_CONFIG_DIR"] internally — so vendor-session reads must
# temporarily mutate that global. This is a process-wide mutation, so a
# *threading* lock (not an asyncio lock) is the right primitive: it blocks
# ANY other thread (e.g. another sync endpoint running in FastAPI's
# threadpool) from observing or clobbering the transiently-overridden env.
# We hold it around EVERY call below — including the non-vendor path —
# precisely so a concurrent non-vendor read can't run while the vendor
# branch has the env flipped (which would point it at the wrong projects
# dir and silently return the wrong session's messages).
# NOTE: there is no `await` inside the locked region, so it can't deadlock
# the event loop with itself; the synchronous file I/O does briefly block
# the calling thread — acceptable here because these reads are small and
# the SDK gives us no async variant for the default store.
_vendor_msg_lock = threading.Lock()

# The SDK/CLI currently accept low..max. MuseLab adds two protocol-level values:
# ``auto`` asks CLIProxyAPI to use the selected Codex model's catalog default,
# while ``ultra`` is supported by newer Codex models but not yet by Claude's
# public SDK literal. Those two values travel in a private request header and
# are handled by the Gateway after Claude's request translator has run.
_SDK_EFFORT_LEVELS = tuple(get_args(EffortLevel))
_VALID_EFFORT = ("auto", *_SDK_EFFORT_LEVELS, "ultra")
_VALID_SERVICE_TIERS = frozenset({"", "fast"})

# Codex Ultra is a client-level mode, not a raw Responses API effort literal:
# the provider wire contract currently tops out at ``max``. MuseLab also keeps
# bounded subagent depth/concurrency for Ultra without injecting a workflow.

# Claude CLI uses this host contract when resuming a transcript that was
# forked away from a still-live runtime.  The value is only an ISO timestamp:
# never put session ids, workspace paths, task descriptions, or output in it.
_RUNTIME_RESUME_SOURCE_ALIVE_ENV = "CLAUDE_CODE_RESUME_SOURCE_ALIVE"


def _normalize_effort(effort: str | None) -> str:
    """Canonicalize the legacy empty-string spelling to ``auto``."""
    value = (effort or "").strip()
    return value or "auto"


def _muselab_gateway_headers(effort: str, service_tier: str) -> str:
    """Build Claude CLI's newline-delimited custom-header environment value.

    Values have already passed closed-set validation; keeping this helper
    intentionally tiny also makes it straightforward to capture/assert the
    exact CLI request without exposing credentials.
    """
    lines = [f"X-MuseLab-Effort: {_normalize_effort(effort)}"]
    if service_tier == "fast":
        lines.append("X-MuseLab-Service-Tier: fast")
    return "\n".join(lines)


# DUCC is an internet-capable local agent runtime.  Giving it the backend's
# ambient environment would also give every tool it launches unrelated MuseLab,
# provider, cloud, GitHub, database and SSH credentials.  Keep this allowlist
# intentionally small and explicit.  The wrapper applies the same policy again
# so direct/operator invocation cannot accidentally widen it.
_DUCC_BASE_ENV_NAMES = (
    "PATH", "HOME", "USER", "LOGNAME", "SHELL", "TERM", "TMPDIR",
    "LANG", "TZ",
    "XDG_RUNTIME_DIR", "XDG_CONFIG_HOME", "XDG_CACHE_HOME", "XDG_DATA_HOME",
    "SSL_CERT_FILE", "SSL_CERT_DIR", "REQUESTS_CA_BUNDLE",
    "NODE_EXTRA_CA_CERTS",
    # Selection metadata used by managed DUCC installations.  Authentication
    # material itself remains in DUCC's HOME-owned config and is not copied
    # from arbitrary environment prefixes.
    "DUCC_AUTH_SOURCE",
    _RUNTIME_RESUME_SOURCE_ALIVE_ENV,
)
_DUCC_PROXY_ENV_NAMES = (
    "HTTP_PROXY", "HTTPS_PROXY", "ALL_PROXY", "NO_PROXY",
    "http_proxy", "https_proxy", "all_proxy", "no_proxy",
)
_DUCC_LOCALE_ENV_NAMES = (
    "LC_ALL", "LC_CTYPE", "LC_NUMERIC", "LC_TIME", "LC_COLLATE",
    "LC_MONETARY", "LC_MESSAGES", "LC_PAPER", "LC_NAME", "LC_ADDRESS",
    "LC_TELEPHONE", "LC_MEASUREMENT", "LC_IDENTIFICATION",
)


def _ducc_subprocess_env(executable: str) -> dict[str, str]:
    """Build the complete, privacy-bounded environment for DUCC.

    Current Claude Agent SDK versions merge options.env over the parent
    environment.  This allowlist controls explicit values, while the DUCC
    wrapper's final ``env -i`` rebuild is the actual isolation boundary.
    Credentialed proxy URLs are omitted as they commonly contain a reusable
    password.
    """
    safe = {
        name: value
        for name in _DUCC_BASE_ENV_NAMES
        if (value := os.environ.get(name)) is not None
    }
    # The allowlisted name is not permission to forward arbitrary content.
    # Keep this host signal timestamp-only even if an operator's ambient
    # environment contains a stale or accidentally sensitive value.
    resume_boundary = sess.normalize_runtime_fork_boundary_at(
        safe.get(_RUNTIME_RESUME_SOURCE_ALIVE_ENV)
    )
    if resume_boundary:
        safe[_RUNTIME_RESUME_SOURCE_ALIVE_ENV] = resume_boundary
    else:
        safe.pop(_RUNTIME_RESUME_SOURCE_ALIVE_ENV, None)
    safe.update({
        name: value
        for name in _DUCC_LOCALE_ENV_NAMES
        if (value := os.environ.get(name))
    })
    for name in _DUCC_PROXY_ENV_NAMES:
        value = os.environ.get(name)
        if value is not None and "@" not in value:
            safe[name] = value
    # This is a resolved executable path, not an auth value.  The wrapper
    # captures it before rebuilding its own empty environment and never
    # forwards MUSELAB_DUCC_CLI to the actual runtime.
    safe["MUSELAB_DUCC_CLI"] = executable
    return safe


def _cli_stderr_category(line: str) -> str:
    """Classify CLI stderr without retaining its potentially private text."""
    low = (line or "").lower()
    if any(word in low for word in (
        "auth", "credential", "login", "token", "unauthorized", "forbidden",
    )):
        category = "authentication"
    elif any(word in low for word in (
        "network", "connect", "timeout", "proxy", "dns", "tls", "certificate",
    )):
        category = "network"
    elif any(word in low for word in (
        "config", "model", "argument", "option", "permission",
    )):
        category = "configuration"
    else:
        category = "runtime"
    return category


def _ducc_stderr_notice(line: str) -> str:
    """Backward-compatible privacy-safe DUCC stderr notice."""
    return f"{_cli_stderr_category(line)} detail suppressed for privacy"


def _privacy_safe_cli_stderr_logger(
    runtime: str,
    session_id: str,
):
    """Return a per-client, per-category deduplicating stderr sink.

    Claude/DUCC stderr may contain prompts, workspace paths, credentials, or
    full upstream protocol payloads.  One category line is enough to diagnose
    the failing layer; the authenticated SSE error remains the detailed user
    surface.
    """
    seen: set[str] = set()
    safe_runtime = "DUCC" if runtime == "DUCC" else "SDK-CLI"

    def _logger(line: str) -> None:
        category = _cli_stderr_category(line)
        if category in seen:
            return
        seen.add(category)
        sys.stderr.write(
            f"[{safe_runtime}] sid={obs.short_id(session_id)} "
            f"category={category} detail=suppressed\n"
        )
        sys.stderr.flush()

    return _logger


def _perf_event(event: str, /, **fields: Any) -> None:
    """Best-effort bridge: diagnostics must never alter request lifecycles."""
    try:
        obs.perf_event(event, **fields)
    except Exception:
        pass


def _safe_secondary_diagnostic(
    stage: str, session_id: str, exc: BaseException,
) -> None:
    """Emit one bounded, content-free diagnostic for contained cleanup faults."""
    safe_stage = re.sub(r"[^a-z0-9_.-]", "_", stage.lower())[:48]
    sys.stderr.write(
        f"[chat-secondary] stage={safe_stage} "
        f"sid={obs.short_id(session_id)} exc={type(exc).__name__}\n"
    )
    sys.stderr.flush()


def _build_runtime_task_context_hook(session_id: str):
    """Describe inherited task truth without exposing task output or paths.

    Older successor transcripts may already contain Claude CLI's synthetic
    "No completion record ... previous session" notification.  That record
    only means the resumed CLI does not own the predecessor's process.  Give
    the model a small, fresh, sidecar-backed correction on every prompt while
    leaving the canonical user message and transcript untouched.
    """
    async def runtime_task_context(_input_data, _tool_use_id, _context):
        overlays = await obs.to_thread_io(
            "chat.runtime_task_context_read",
            session_id,
            sess.get_authoritative_runtime_task_overlays,
            session_id,
        )
        inherited: list[str] = []
        inherited_overlays = [
            (task_id, overlay)
            for task_id, overlay in overlays.items()
            if str(overlay.get("owner_session_id") or "") != session_id
        ]
        inherited_overlays.sort(
            key=lambda item: (
                -int(item[1].get("updated_at") or 0)
                if isinstance(item[1].get("updated_at"), (int, float))
                else 0,
                str(item[0]),
            )
        )
        for task_id, overlay in inherited_overlays[:32]:
            safe_id = re.sub(
                r"[^A-Za-z0-9_.:-]", "_", str(task_id)
            )[:128]
            state = str(overlay.get("state") or "unknown").lower()
            if state not in {"running", "completed", "failed", "stopped"}:
                state = "unknown"
            updated_at = overlay.get("updated_at")
            updated = (
                str(int(updated_at))
                if isinstance(updated_at, (int, float)) else "unknown"
            )
            inherited.append(
                f"- task {safe_id}: state={state}, updated_at_ms={updated}"
            )
        if not inherited:
            return {}
        context = (
            "MuseLab runtime handoff: these background tasks belong to a "
            "still-separate predecessor runtime. The states below are the "
            "authoritative MuseLab lifecycle state:\n"
            + "\n".join(inherited)
            + "\nIgnore automatic stopped notifications saying that no "
              "completion record was found in a previous session for these "
              "task ids; they mean only that this CLI does not own the task. "
              "Do not call TaskOutput or TaskStop for inherited tasks."
        )
        return {
            "hookSpecificOutput": {
                "hookEventName": "UserPromptSubmit",
                "additionalContext": context,
            }
        }

    return runtime_task_context


def _build_codex_skill_guard_hook():
    """Keep one known-oversized built-in Skill from killing Codex sessions.

    Claude Code's bundled ``claude-api`` Skill can inject hundreds of
    thousands of characters in a single tool call.  A Codex Gateway turn can
    therefore jump from below the preflight threshold to beyond the model's
    hard window before native auto-compaction gets another chance to run.  A
    PreToolUse hook is used instead of a broad ``Skill`` deny so every normal
    project/user Skill remains available.  Operators who knowingly run a
    larger-window gateway can opt out explicitly.
    """
    async def guard(input_data, _tool_use_id, _context):
        tool_input = (
            input_data.get("tool_input")
            if isinstance(input_data, dict) else {}
        )
        name = str(
            (tool_input or {}).get("skill")
            or (tool_input or {}).get("name")
            or ""
        ).strip().lower()
        if name not in {"claude-api", "claude_api"}:
            return {}
        return {
            "hookSpecificOutput": {
                "hookEventName": "PreToolUse",
                "permissionDecision": "deny",
                "permissionDecisionReason": (
                    "The claude-api Skill is too large for this Codex Gateway "
                    "context window. Use focused documentation/search instead."
                ),
            }
        }

    return guard


def _cli_project_roots() -> list[Path]:
    """Compatibility wrapper for canonical Claude CLI project roots."""
    return chat_history.cli_project_roots(endpoints._vendor_config_dir())


def _cli_encode_cwd(path: str) -> str:
    """Compatibility wrapper for the SDK/CLI workspace encoding."""
    return chat_history.cli_encode_cwd(
        path, project_key_for_directory=project_key_for_directory)


# Public compatibility aliases: callers and tests historically mutate this
# exact object through ``backend.chat``.  The extracted implementation receives
# it explicitly so cache identity and monkeypatch behavior remain unchanged.
_JSONL_PATH_CACHE = chat_history.JSONL_PATH_CACHE
_JSONL_PATH_CACHE_MAX = chat_history.JSONL_PATH_CACHE_MAX


def _find_session_jsonl(sid: str) -> Path | None:
    """Compatibility wrapper for positive-cache transcript discovery."""
    return chat_history.find_session_jsonl(
        sid,
        project_roots=_cli_project_roots,
        cache=_JSONL_PATH_CACHE,
        cache_max=_JSONL_PATH_CACHE_MAX,
    )


def _canonical_session_evidence_path(sid: str, workspace: Path) -> Path | None:
    """Compatibility wrapper for validated canonical transcript evidence."""
    return chat_history.canonical_session_evidence_path(
        sid,
        workspace,
        find_session_jsonl=_find_session_jsonl,
        project_roots=_cli_project_roots,
        encode_cwd=_cli_encode_cwd,
    )


def _compact_tail_cursor(sid: str) -> tuple[Path | None, int]:
    """Compatibility wrapper for the pre-compact transcript cursor."""
    return chat_history.compact_tail_cursor(
        sid, find_session_jsonl=_find_session_jsonl)


def _compact_tail_outcome(path: Path | None, offset: int) -> dict[str, bool]:
    """Compatibility wrapper for raw post-compact tail inspection."""
    return chat_history.compact_tail_outcome(path, offset)


@contextmanager
def _session_config_dir(model: str = "", *, sid: str = ""):
    """Scope SDK session operations to the JSONL's actual config store."""
    session_path = _find_session_jsonl(sid) if sid else None
    with chat_history.session_config_dir(
        model,
        lock=_vendor_msg_lock,
        is_third_party=endpoints.is_third_party,
        vendor_config_dir=endpoints._vendor_config_dir,
        session_path=session_path,
    ):
        yield


def _get_session_msgs(sid: str, model: str = "") -> list:
    """Compatibility wrapper retaining the patchable SDK message loader."""
    return chat_history.get_session_msgs(
        sid,
        model,
        config_dir=_session_config_dir,
        loader=get_session_messages,
        workspace=sess.session_workspace,
    )


def _transcript_ts_ms(entry: dict) -> int | None:
    """Compatibility wrapper for raw transcript timestamps."""
    return chat_history.transcript_ts_ms(entry)


_RawMsg = chat_history.RawMsg


def _raw_msg_from_entry(entry: dict) -> _RawMsg | None:
    """Compatibility facade for canonical record projection."""
    return chat_history.raw_msg_from_entry(
        entry,
        raw_msg_type=_RawMsg,
        timestamp_ms=_transcript_ts_ms,
    )


def _full_session_msgs(sid: str) -> list:
    """Compatibility wrapper for canonical full-file transcript reads."""
    return chat_history.full_session_msgs(
        sid,
        find_session_jsonl=_find_session_jsonl,
        raw_msg_type=_RawMsg,
        timestamp_ms=_transcript_ts_ms,
    )


def _read_tail_lines(path: Path, n: int, block: int = 65536) -> list[str]:
    """Compatibility wrapper for bounded raw tail reads."""
    return chat_history.read_tail_lines(path, n, block)


def _recent_turn_uuids(sid: str, want_image_user: bool,
                       tail_lines: int = 400) -> tuple[str | None, str | None]:
    """Compatibility wrapper for recent canonical turn UUID discovery."""
    return chat_history.recent_turn_uuids(
        sid,
        want_image_user,
        tail_lines,
        find_session_jsonl=_find_session_jsonl,
        tail_reader=_read_tail_lines,
    )


router = APIRouter(prefix="/api/chat", tags=["chat"])


# NOTE: a former `_plain_preview()` helper turned the assistant reply into a
# 120-char push-notification body. It was removed (2026-05-29) because the
# preview leaked private reply content onto the lock screen — the push body
# is now a fixed "Muse 已回复" with no reply text. See the turn-done push
# fan-out for the privacy rationale.


# Compatibility aliases: callers and tests historically inspect and mutate
# these exact containers through ``backend.chat``. The focused runtime module
# owns them; aliases preserve identity across the extraction boundary.
_ClientKey = chat_runtime.ClientKey
_clients = chat_runtime.CLIENTS
_client_permission = chat_runtime.CLIENT_PERMISSION
_client_plan_return = chat_runtime.CLIENT_PLAN_RETURN


# Bound the exact wire-event window used for incremental reconnect. The durable
# spool below still stores a coalesced, complete turn for a cold/full replay;
# this in-memory window stores individual text/thinking deltas so a client can
# resume in the middle of a message without replaying or duplicating its prefix.
# If a requested sequence has fallen out of this window, the server explicitly
# asks the browser to reconcile from canonical history instead of guessing.
_BROADCAST_REPLAY_MAX_EVENTS = env_int(
    "MUSELAB_STREAM_REPLAY_MAX_EVENTS", 4096, min_value=64)
_BROADCAST_REPLAY_MAX_BYTES = env_int(
    "MUSELAB_STREAM_REPLAY_MAX_BYTES", 4 * 1024 * 1024, min_value=64 * 1024)

# Cap on token deltas queued for a single attached subscriber. Deltas are the
# ephemeral presentation channel (see _TurnSubscriber): they are never spooled,
# so this bounds per-subscriber memory. A client that backs up past this is
# resynced rather than served a truncated bubble.
_BROADCAST_LIVE_DELTA_MAX = env_int(
    "MUSELAB_STREAM_LIVE_DELTA_MAX", 4096, min_value=64)
_BROADCAST_SUBSCRIBER_MAX_EVENTS = env_int(
    "MUSELAB_STREAM_SUBSCRIBER_MAX_EVENTS", 256, min_value=8)
_BROADCAST_SUBSCRIBER_MAX_BYTES = env_int(
    "MUSELAB_STREAM_SUBSCRIBER_MAX_BYTES", 1024 * 1024, min_value=1024)

# Internal delimiter between one message's ephemeral live deltas and the next
# durable spool event.  It never leaves `_TurnSubscriber` or reaches the wire.
_LIVE_MESSAGE_BARRIER = object()


def _broadcast_event_size(event: dict) -> int:
    try:
        return len(json.dumps(
            event, ensure_ascii=False, separators=(",", ":")
        ).encode("utf-8"))
    except (TypeError, ValueError):
        return _BROADCAST_SUBSCRIBER_MAX_BYTES + 1


def _stream_text_payload(event: dict) -> tuple[str, str] | None:
    """Return ``(event_name, text)`` for a plain text/thinking delta."""
    kind = event.get("event")
    if kind not in {"text", "thinking"}:
        return None
    try:
        payload = json.loads(event.get("data") or "{}")
    except (TypeError, ValueError, json.JSONDecodeError):
        return None
    if (not isinstance(payload, dict) or set(payload) != {"text"}
            or not isinstance(payload.get("text"), str)):
        return None
    return kind, payload["text"]


class _ReplayRecordCorruption(ValueError):
    """A replay row is incomplete or violates the private spool schema."""


def _decode_replay_record(line: bytes) -> dict:
    """Decode and type-check one complete replay row for every reader path."""
    if not line.endswith(b"\n"):
        raise _ReplayRecordCorruption("incomplete replay record")
    try:
        event = json.loads(line)
    except (TypeError, ValueError, json.JSONDecodeError) as exc:
        raise _ReplayRecordCorruption("invalid replay json") from exc
    if not isinstance(event, dict):
        raise _ReplayRecordCorruption("replay record is not an object")
    if not isinstance(event.get("event"), str) or not event["event"]:
        raise _ReplayRecordCorruption("replay event type is invalid")
    if not isinstance(event.get("data"), str):
        raise _ReplayRecordCorruption("replay event data is invalid")
    if "_coalesced" in event and not isinstance(event["_coalesced"], bool):
        raise _ReplayRecordCorruption("replay coalesced marker is invalid")
    return event


def _replay_runtime_dir() -> Path:
    """Return private durable storage for transient replay records."""
    configured = os.environ.get("MUSELAB_RUNTIME_DIR", "").strip()
    root = (Path(configured).expanduser() if configured
            else sess.SESS_DIR / "runtime")
    root.mkdir(parents=True, exist_ok=True, mode=0o700)
    root.chmod(0o700)
    return root


class _ReplaySpool:
    """Append-only replay storage outside the Python heap."""

    def __init__(self):
        fd, name = tempfile.mkstemp(
            prefix="muselab-turn-", suffix=".jsonl",
            dir=str(_replay_runtime_dir()),
        )
        os.fchmod(fd, 0o600)
        self.path = Path(name)
        self._fd = fd
        self._count = 0
        self._bytes = 0
        self._closed = False
        self._usable = True

    def append(self, event: dict) -> None:
        if self._closed:
            raise RuntimeError("replay spool is closed")
        if not self._usable:
            raise RuntimeError("replay spool is unusable")
        payload = json.dumps(event, ensure_ascii=False, separators=(",", ":"))
        blob = payload.encode("utf-8") + b"\n"
        start = self._bytes
        written = 0
        try:
            while written < len(blob):
                count = os.write(self._fd, blob[written:])
                if count <= 0:
                    raise OSError(errno.EIO, "replay spool write made no progress")
                written += count
        except BaseException:
            try:
                os.ftruncate(self._fd, start)
                os.lseek(self._fd, start, os.SEEK_SET)
            except OSError:
                self._usable = False
            raise
        self._count += 1
        self._bytes += len(blob)

    def size(self) -> int:
        """Bytes written so far. A subscriber records this at attach time to
        tell "history I must replay" from "events that happened while I was
        already attached" — the two need different handling for coalesced
        text (see _TurnSubscriber._skip_from)."""
        return self._bytes

    def open_reader(self):
        return self.path.open("rb")

    def __len__(self) -> int:
        return self._count

    def __iter__(self):
        with self.open_reader() as reader:
            for line in reader:
                yield _decode_replay_record(line)

    def __getitem__(self, index):
        if isinstance(index, slice):
            return list(self)[index]
        if index < 0:
            index += self._count
        for pos, event in enumerate(self):
            if pos == index:
                return event
        raise IndexError(index)

    def close(self) -> None:
        if self._closed:
            return
        self._closed = True
        os.close(self._fd)
        with suppress(OSError):
            self.path.unlink()

    def __del__(self):
        with suppress(Exception):
            self.close()


class _TurnSubscriber:
    """Independent cursor over the broadcast's replay spool, plus a small
    queue for live token deltas.

    Two channels, because the SDK models two different things (see the
    ``publish``/``_flush_compact_text`` comments on TurnBroadcast):

      * The spool is the RECORD: one coalesced text/thinking event per
        assistant message, plus every tool_use / tool_result / done. It is
        replayed in full to anyone who attaches, and its length is
        proportional to the number of SDK messages in the turn.
      * ``_live_q`` is PRESENTATION: raw ``StreamEvent`` deltas, delivered
        only to subscribers attached at the moment they were produced, and
        never persisted.

    A subscriber that was already attached when a message streamed has seen
    that message as deltas, so it must NOT also receive the coalesced event
    the spool records for it. ``_skip_from`` is the spool byte offset at
    attach time and draws exactly that line: coalesced events at or after it
    are ours-already-seen and get skipped; everything before it is history we
    are replaying and gets emitted.
    """

    def __init__(self, replay=None, *, initial_events=None,
                 resync_payload: dict | None = None, skip_from: int = 0):
        self._replay = replay
        self._initial_events: deque[dict] = deque(initial_events or ())
        self._resync_payload = dict(resync_payload or {})
        self._skip_from = skip_from
        # Deltas produced while attached. Bounded: a stalled HTTP connection
        # must not grow this without limit. Overflow degrades to a resync
        # rather than silently dropping text (see publish_live()).
        self._live_q: deque[Any] = deque()
        self._draining_live_barrier = False
        self._wake = asyncio.Event()
        self._done = False

    async def get(self):
        while True:
            # `resync()` can run while this coroutine is asleep in
            # `_wake.wait()`. It closes the reader immediately, so re-check at
            # every loop boundary before touching `_replay`; checking only at
            # function entry turned an intended resync frame into a None.tell
            # AttributeError under live-delta backpressure.
            if self._resync_payload:
                payload, self._resync_payload = self._resync_payload, {}
                return {
                    "event": "resync",
                    "data": json.dumps(payload),
                }
            if self._initial_events:
                event = dict(self._initial_events.popleft())
                event.pop("_coalesced", None)
                return event
            if self._replay is None:
                return None
            if self._draining_live_barrier:
                if self._live_q:
                    event = self._live_q.popleft()
                    if event is _LIVE_MESSAGE_BARRIER:
                        self._draining_live_barrier = False
                        continue
                    return event
                if self._done:
                    # Every skipped coalesced record is paired synchronously
                    # with a marker. Reaching terminal state without it means
                    # the presentation channel cannot prove ordering; force a
                    # canonical resync instead of emitting a later `done`.
                    self._draining_live_barrier = False
                    return {
                        "event": "resync",
                        "data": json.dumps({
                            "reason": "live_barrier_missing",
                            "retryable": True,
                        }),
                    }
                self._wake.clear()
                if self._live_q:
                    continue
                await self._wake.wait()
                continue
            event = self._next_spool_event()
            if self._resync_payload:
                continue
            if event is _LIVE_MESSAGE_BARRIER:
                self._draining_live_barrier = True
                continue
            if event is not None:
                return event
            if self._live_q:
                event = self._live_q.popleft()
                # Defensive only: a paired coalesced spool line is always read
                # first. Never leak an internal marker if a corrupt line made
                # that pair unreadable.
                if event is _LIVE_MESSAGE_BARRIER:
                    continue
                return event
            if self._done:
                self.close_reader()
                return None
            self._wake.clear()
            # Close the clear/append race: publish() may have written between
            # the first EOF read and clear(). Recheck before sleeping.
            event = self._next_spool_event()
            if self._resync_payload:
                continue
            if event is _LIVE_MESSAGE_BARRIER:
                self._draining_live_barrier = True
                continue
            if event is not None:
                return event
            if self._live_q:
                event = self._live_q.popleft()
                if event is _LIVE_MESSAGE_BARRIER:
                    continue
                return event
            if self._done:
                self.close_reader()
                return None
            await self._wake.wait()

    def _next_spool_event(self):
        """Next spool line this subscriber should actually emit, or None when
        the spool is exhausted. Drops coalesced text that this subscriber
        already received as live deltas."""
        while True:
            offset = self._replay.tell()
            line = self._replay.readline()
            if not line:
                return None
            try:
                event = _decode_replay_record(line)
            except _ReplayRecordCorruption:
                self.resync("replay_corrupt")
                return None
            coalesced = event.pop("_coalesced", False)
            if coalesced and offset >= self._skip_from:
                # Streamed to us live, token by token — emitting the coalesced
                # form too would duplicate the whole message. Stop at its
                # live-channel delimiter before reading a later tool/done row;
                # otherwise a slow subscriber observes `done` before the text
                # still queued for this exact segment.
                return _LIVE_MESSAGE_BARRIER
            return event

    def qsize(self) -> int:
        return len(self._live_q)

    def publish(self, event: dict) -> bool:
        """Wake this subscriber for a newly-appended spool event."""
        if self._done:
            return False
        self._wake.set()
        return True

    def publish_live(self, event: dict) -> bool:
        """Deliver an ephemeral delta that is deliberately never spooled."""
        if self._done:
            return False
        if len(self._live_q) >= _BROADCAST_LIVE_DELTA_MAX:
            # This client is too far behind to keep up with the token stream.
            # Deltas are best-effort by design, but we have already emitted
            # some of this message and skipped its coalesced form, so silently
            # dropping the rest would leave a truncated bubble. Resync instead:
            # the client reloads canonical history and comes back consistent.
            self.resync("live_backlog")
            return False
        self._live_q.append(event)
        self._wake.set()
        return True

    def publish_live_barrier(self) -> bool:
        """Close the current live-delta segment without emitting a wire event."""
        if self._done:
            return False
        self._live_q.append(_LIVE_MESSAGE_BARRIER)
        self._wake.set()
        return True

    def resync(self, reason: str, **details) -> None:
        self._resync_payload = {
            "reason": reason,
            "retryable": True,
            **details,
        }
        self.close_reader()
        self._done = True
        self._wake.set()

    def close(self) -> None:
        self._done = True
        self._wake.set()

    def close_reader(self) -> None:
        if self._replay is not None:
            self._replay.close()
            self._replay = None


class TurnBroadcast:
    """Fan-out for an in-flight assistant turn.

    Why: the SSE event_gen used to be the sole consumer of SDK output
    via merge_q; when the browser closed, the generator unwound and
    cancelled pump_claude, killing the in-progress reply.

    Now event_gen runs as a detached background task that PUBLISHES
    every SSE event it would have yielded to this broadcast. The HTTP
    endpoint is just a SUBSCRIBER — it replays the existing buffer +
    streams new events. A reconnecting browser becomes a new subscriber
    and gets the full reply via replay + live tail, with no extra logic
    on the SDK side. A turn runs unbounded by default; set
    MUSELAB_TURN_TIMEOUT_S to arm a wall-clock cap at the background-task
    level. Removed from `_active_turns` when finished.
    """
    def __init__(
        self,
        session_id: str,
        model: str = "",
        *,
        replay_max_events: int = 0,
        replay_max_bytes: int = 0,
        subscriber_max_events: int = _BROADCAST_SUBSCRIBER_MAX_EVENTS,
        subscriber_max_bytes: int = _BROADCAST_SUBSCRIBER_MAX_BYTES,
    ):
        self.session_id = session_id
        self.model = model
        self.events = _ReplaySpool()
        self.subscribers: set[_TurnSubscriber] = set()
        self._resume_max_events = (
            replay_max_events if replay_max_events > 0
            else _BROADCAST_REPLAY_MAX_EVENTS
        )
        self._resume_max_bytes = (
            replay_max_bytes if replay_max_bytes > 0
            else _BROADCAST_REPLAY_MAX_BYTES
        )
        self._resume_events: deque[tuple[int, dict, int]] = deque()
        self._resume_bytes = 0
        # Legacy per-subscriber constructor arguments remain accepted. Live
        # backpressure is governed by the module constants used by subscribers.
        _ = (subscriber_max_events, subscriber_max_bytes)
        self._compact_kind: str | None = None
        self._compact_parts: list[str] = []
        self._compact_chars = 0
        self._compact_last_seq = 0
        self._replay_bytes = 0
        self.done = False
        # Set True when this turn ended via an explicit user /interrupt (vs.
        # natural completion or error). The server-side queue drain reads it
        # in _pump_gen_to_broadcast's finally: a cancelled turn PAUSES the
        # queue rather than charging into the next item — the user stopped on
        # purpose, almost never "just this one, send the rest."
        self.cancelled = False
        # Exact persisted assistant boundary observed from AssistantMessage.
        # Keep it on the broadcast as well as in event_gen's local accumulator:
        # a forced interrupt can prevent ResultMessage from arriving, but the
        # CLI may still flush that AssistantMessage to JSONL several seconds
        # later.  The cancellation cleanup can then annotate the precise UUID
        # instead of racing a "latest transcript row" scan.
        self.last_assistant_uuid = ""
        self.cancelled_at_ms = 0
        self.started_at = time.time()
        # Immutable identity for this exact assistant turn. Session id alone
        # cannot distinguish the classic ABA race where turn A finishes and
        # turn B starts before a reconnect for A reaches the server.
        self.turn_id = str(uuid.uuid4())
        # Queue ownership is durable. A claimed item stays in the queue sidecar
        # until this exact turn finishes and acknowledges it.
        self.queue_item_id: str = ""
        # Mid-turn steering is also backed by the durable queue. The browser
        # first commits an item there, then the enqueue response schedules a
        # UUID-stamped ``priority=next`` write to this exact SDK runtime. The
        # item is removed only after the CLI emits command_lifecycle=completed.
        # Keeping the exact client/turn here closes the session-id ABA race.
        self.runtime_client: "MuseLabSDKClient | None" = None
        self.query_committed = False
        # Native steering must target this exact turn, but the queue request can
        # arrive after HTTP admission and before the root SDK query commits.
        # Writers register durably against the broadcast, then wait on this
        # one-shot gate instead of being downgraded to end-of-turn FIFO.
        self.steering_ready = asyncio.Event()
        self.result_forwarded = False
        self.permission = ""
        self.active_tool_use_ids: set[str] = set()
        # Synchronously closed before any terminal cleanup await. This prevents
        # a late enqueue from registering a new native command after the error
        # finalizer has already snapshotted the old command map.
        self.steering_closed = False
        self.steering_commands: dict[str, dict[str, str]] = {}
        self.steering_write_events: dict[str, asyncio.Event] = {}
        # Headless background-task continuations are separate turns, linked to
        # the user turn that launched the task for observability and recovery.
        self.parent_turn_id: str = ""
        # Strictly increasing within this broadcast. Replayed and live events
        # carry the same sequence so clients can discard duplicate delivery
        # without guessing from event content.
        self._event_seq = 0
        # Set in finish(). Used by the _recent_turns grace-keep map to TTL-evict
        # broadcasts that ended a while ago.
        self.finished_at: float = 0.0
        # User-side context for this turn — populated when the SSE
        # endpoint kicks off a new turn. Needed because SDK CLI only
        # flushes the session JSONL at turn completion; mid-turn reloads
        # would see an empty session unless we reconstruct the message
        # list from broadcast state. The user message itself isn't in
        # `events` (those are server→client SSE events; user prompt is
        # a separate input channel) so we keep it on the broadcast.
        self.user_text: str = ""
        self.user_images: list[dict] = []
        self.user_docs: list[dict] = []
        # Durable pre-prepare attachment intent. These opaque, validated IDs let
        # restart recovery retain what the admitted turn owned without logging or
        # persisting file names, paths, MIME payloads, or attachment contents.
        self.staged_attachment_ids: list[str] = []
        # Browser-facing startup is separate from transcript loading and model
        # output. Named startup events are replayable and low-frequency; detailed
        # timings stay in privacy-bounded performance fields below.
        self.startup_phase: str = ""
        # Staged uploads remain in the global store until the SDK query write
        # succeeds. These private fields carry their exclusive lease and any
        # pre-query disk artifacts across compact/preflight awaits so every
        # failure/cancellation path can release one coherent transaction.
        self._attachment_lease: "_StagedAttachmentLease | None" = None
        self._prepared_attachments: "_PreparedStagedAttachments | None" = None
        # Preparation wraps an actual worker thread. Rollback joins this task
        # before exposing the staged id for retry, even when a force-stop
        # watchdog takes over from a startup/pump task that never ran finally.
        self._attachment_prepare_task: "asyncio.Task | None" = None
        # A single shielded rollback owner makes cleanup idempotent under the
        # pump, startup owner, watchdog, and repeated Task.cancel() racing one
        # another. The completed task is retained for the broadcast lifetime.
        self._attachment_rollback_task: "asyncio.Task | None" = None
        # Startup settlement (lease/files + queue + Activity + active slot +
        # sidecar/snapshot) likewise has one owner. The first terminal path
        # wins; later paths only join it instead of duplicating bookkeeping.
        self._startup_terminal_cleanup_task: "asyncio.Task | None" = None
        self._startup_queue_settled = False
        # Display-only recovery boundary for an explicitly interrupted turn.
        # The Claude CLI may abort before it flushes this turn to JSONL, while
        # the browser has already rendered text/thinking/tool cards.  Capture
        # the transcript coordinate immediately before query() so a durable
        # MuseLab snapshot can later be inserted at the same visual position
        # without writing partial assistant output back into model context.
        self.transcript_boundary: dict[str, Any] = {}
        self.canonical_terminal_published = False
        self.cancelled_snapshot_persisted = False
        # Result-only/API transport failures can end without a canonical
        # AssistantMessage UUID. Keep their display record separately so a
        # refresh cannot erase the visible error or relabel an older reply.
        self.failed_snapshot_persisted = False
        # Some third-party runtimes return the final prose only on ResultMessage
        # after their last tool call. Preserve that authoritative result in the
        # same private presentation channel so live and refreshed views agree.
        self.result_snapshot_persisted = False
        self.cancelled_snapshot_suppressed = False
        self.cancelled_snapshot_lock = threading.Lock()
        # True for a HEADLESS CONTINUATION turn: the cross-turn task watcher
        # opens one of these (no user prompt) when an SDK background task
        # finishes after its originating turn ended. It carries the task's
        # terminal TaskNotification (card → ✅done) plus the CLI's auto-continue
        # model reaction. The frontend attaches in "continuation" mode — same
        # reconnect plumbing as a queue-drain, but it must NOT truncate the
        # in-flight portion (the launching tool_use card lives there and the
        # task_notification needs to flip it). See `/active` + send({continuation}).
        self.is_continuation: bool = False
        # SDK Cron prompts arrive as autonomous turns on the pooled stream.
        # They use the normal replay/reconnect contract, but retain their own
        # delivery class so tabs can show the same amber affordance as other
        # unattended work and the UI can label the injected prompt correctly.
        self.is_scheduled_delivery: bool = False
        self.scheduled_delivery_consumed: bool = False
        # Once a reconnect subscriber has attached to a CONTINUATION broadcast
        # (live or grace-kept), flip this. `/active` then stops advertising it
        # so the frontend's 8s poller can't re-reconnect to the same finished
        # continuation every tick (which replayed the reaction → duplicate
        # bubbles). One continuation ⇒ at most one reconnect ⇒ one replay,
        # regardless of frontend version. No effect on normal turns.
        self.continuation_consumed: bool = False
        # Handle to the detached `_pump_gen_to_broadcast` task driving this
        # turn. Stored so the force-stop watchdog (_force_stop_after_grace) can
        # cancel a pump that an interrupt + client teardown failed to unblock.
        # None until _start_turn finishes wiring the pump.
        self.task: "asyncio.Task | None" = None
        # Client creation can take up to 60s on a cold session (CLI spawn, MCP
        # initialization). Stop must be able to cancel during that window,
        # before the client exists in `_clients` and before `self.task` above is
        # created. `_start_turn` owns and clears this handle.
        self.startup_task: "asyncio.Task | None" = None
        # Outer task currently running `_start_turn`.  Deletion must join this
        # owner, not only the nested SDK startup task: cancelling the latter
        # makes the outer owner create the Activity/snapshot cleanup slightly
        # later.  Without this handle DELETE could return before that late
        # cleanup existed and the deleted session could re-enter replay state.
        self.startup_owner_task: "asyncio.Task | None" = None
        # The global activity row starts as soon as this broadcast reserves the
        # session, before cold client/MCP startup. Startup failure/cancellation
        # uses this bit to close only rows that were actually created.
        self.activity_started = False
        self.activity_hidden = False
        # One low-cardinality performance summary per logical turn.  Values are
        # monotonic durations or closed-set classifications only; prompts,
        # paths, tool payloads and exception text never enter this structure.
        self.perf_started = obs.monotonic()
        self.perf_client = "cold"
        self.perf_client_ms = 0
        self.perf_preflight_ms = 0
        self.perf_query_started = 0.0
        self.perf_first_event_ms = -1
        self.perf_first_visible_ms = -1
        self.perf_result_ms = -1
        self.perf_post_started = 0.0
        self.perf_post_ms = 0
        self.perf_status = "unknown"
        self.perf_error_kind = "none"
        self.perf_background_count = 0
        self.perf_final_mode = "normal"
        self.perf_admission_ms = 0
        self.perf_intent_ms = 0
        self.perf_runtime_lock_ms = 0
        self.perf_disconnect_ms = 0
        self.perf_pool_ms = 0
        self.perf_creation_lock_ms = 0
        self.perf_connect_ms = 0
        self.perf_mcp_ms = 0
        self.perf_pool_commit_ms = 0
        self.perf_attachment_ms = 0
        self.perf_intent_refresh_ms = 0
        self.perf_query_write_ms = 0
        self.perf_startup_status = "pending"
        self.perf_startup_failure_phase = "none"
        self._startup_perf_emitted = False
        self._perf_emitted = False

    @property
    def activity_source(self) -> str:
        """Stable notification class for this logical turn.

        Session id alone cannot tell a foreground reply from a turn claimed
        by the durable queue, and continuation broadcasts deliberately reuse
        the originating session.  Send this explicit class on every terminal
        and active-state payload so the browser only auto-acknowledges work it
        was actually watching.
        """
        if self.is_scheduled_delivery:
            return "scheduled"
        if self.is_continuation:
            return "background"
        if self.queue_item_id:
            return "queued"
        return "direct"

    def publish_startup(self, phase: str) -> None:
        """Publish one replayable, low-cardinality startup transition."""
        normalized = str(phase or "").strip().lower()
        if normalized not in {"accepted", "runtime", "tools", "context"}:
            normalized = "accepted"
        if self.done or self.startup_phase == normalized:
            return
        self.startup_phase = normalized
        self.publish({
            "event": "startup",
            "data": json.dumps({
                "phase": normalized,
                "activity_source": self.activity_source,
            }),
        })

    def emit_startup_perf(self, status: str, *, failure_phase: str = "none") -> None:
        """Emit one privacy-bounded startup summary per logical turn."""
        if self._startup_perf_emitted:
            return
        self._startup_perf_emitted = True
        self.perf_startup_status = str(status or "unknown")
        self.perf_startup_failure_phase = str(failure_phase or "none")
        try:
            _perf_event(
                "chat.startup",
                sid8=obs.short_id(self.session_id),
                turn8=obs.short_id(self.turn_id),
                source=self.activity_source,
                client=self.perf_client,
                status=self.perf_startup_status,
                failure_phase=self.perf_startup_failure_phase,
                error_kind=self.perf_error_kind,
                admission_ms=self.perf_admission_ms,
                intent_ms=self.perf_intent_ms,
                runtime_lock_ms=self.perf_runtime_lock_ms,
                disconnect_ms=self.perf_disconnect_ms,
                pool_ms=self.perf_pool_ms,
                creation_lock_ms=self.perf_creation_lock_ms,
                connect_ms=self.perf_connect_ms,
                mcp_ms=self.perf_mcp_ms,
                pool_commit_ms=self.perf_pool_commit_ms,
                client_ms=self.perf_client_ms,
                attachment_ms=self.perf_attachment_ms,
                intent_refresh_ms=self.perf_intent_refresh_ms,
                preflight_ms=self.perf_preflight_ms,
                sdk_write_ms=self.perf_query_write_ms,
                total_ms=obs.elapsed_ms(self.perf_started),
            )
        except Exception:
            pass

    def publish(self, event: dict) -> None:
        """Route one SSE event to the record channel, the live channel, or both.

        The Claude Agent SDK draws this line for us. ``StreamEvent`` carries
        "the raw Anthropic API stream event" and is opt-in via
        ``include_partial_messages`` (default False) — a presentation detail.
        ``AssistantMessage`` / ``UserMessage`` / ``ResultMessage`` are the
        modeled, complete units — the record. We mirror that split:

          * text/thinking deltas  → live subscribers only, accumulated here
            and written to the spool once, coalesced, at the message boundary.
          * everything else       → spool (and therefore every live tail too).

        The previous implementation appended each delta to the spool whenever
        any subscriber was attached, which made replay length proportional to
        token count rather than message count.
        """
        event_name = str(event.get("event") or "")
        if (self.perf_query_started
                and self.perf_first_visible_ms < 0
                and event_name
                and event_name not in {
                    "startup", "compact_progress", "done", "error", "cancelled",
                }):
            self.perf_first_visible_ms = obs.elapsed_ms(
                self.perf_query_started)
        current_text = _stream_text_payload(event)
        if current_text is not None:
            kind, text = current_text
            if self._compact_kind not in (None, kind):
                self._flush_compact_text()
            stamped = self._stamp_wire_event(event)
            if self._compact_kind is None:
                self._compact_kind = kind
            self._compact_parts.append(text)
            self._compact_chars += len(text)
            self._compact_last_seq = self._event_seq
            for subscriber in tuple(self.subscribers):
                if not subscriber.publish_live(stamped):
                    self.subscribers.discard(subscriber)
            # Bound the accumulator so a very long message (or a headless turn
            # with nobody attached) cannot hold unbounded text in memory. A
            # mid-message flush just splits the record into two coalesced
            # events; consecutive text events rebuild into one bubble on both
            # the replay path and the frontend, so nothing downstream changes.
            if self._compact_chars >= 64 * 1024:
                self._flush_compact_text()
            return
        # A non-delta event closes whatever message was streaming: write its
        # coalesced form first so the spool stays in true chronological order.
        self._flush_compact_text()
        stamped = self._stamp_wire_event(event)
        self._append_replay(stamped)

    def _stamp_wire_event(self, event: dict) -> dict:
        stamped = dict(event)
        self._event_seq += 1
        try:
            payload = json.loads(stamped.get("data") or "{}")
        except (TypeError, ValueError):
            payload = None
        if isinstance(payload, dict):
            if str(stamped.get("event") or "") == "task_notification":
                payload = normalize_task_summary_fields(payload)
            payload["turn_id"] = self.turn_id
            payload["event_seq"] = self._event_seq
            if self.parent_turn_id:
                payload["parent_turn_id"] = self.parent_turn_id
            stamped["data"] = json.dumps(payload, ensure_ascii=False)
        size = _broadcast_event_size(stamped)
        self._resume_events.append((self._event_seq, stamped, size))
        self._resume_bytes += size
        while (len(self._resume_events) > self._resume_max_events
               or self._resume_bytes > self._resume_max_bytes):
            _, _, evicted_size = self._resume_events.popleft()
            self._resume_bytes -= evicted_size
        return stamped

    def _append_replay(self, replay: dict) -> None:
        self.events.append(replay)
        self._replay_bytes += _broadcast_event_size(replay)
        for subscriber in tuple(self.subscribers):
            if not subscriber.publish(replay):
                self.subscribers.discard(subscriber)

    def _flush_compact_text(self) -> None:
        """Write the message that just finished streaming as ONE spool event.

        Marked ``_coalesced`` so subscribers that already received this message
        as live deltas skip it instead of rendering the text twice; subscribers
        replaying history emit it normally. The marker is stripped before the
        event leaves _TurnSubscriber, so it never reaches the wire.
        """
        if self._compact_kind is None:
            return
        payload = {
            "text": "".join(self._compact_parts),
            "turn_id": self.turn_id,
            "event_seq": self._compact_last_seq,
        }
        if self.parent_turn_id:
            payload["parent_turn_id"] = self.parent_turn_id
        event = {
            "event": self._compact_kind,
            "data": json.dumps(payload, ensure_ascii=False),
            "_coalesced": True,
        }
        self._compact_kind = None
        self._compact_parts = []
        self._compact_chars = 0
        self._compact_last_seq = 0
        self._append_replay(event)
        # `_append_replay` and these markers run synchronously on the same
        # event-loop turn: a reader woken by the spool append cannot interleave
        # before every currently-attached subscriber receives its delimiter.
        for subscriber in tuple(self.subscribers):
            if not subscriber.publish_live_barrier():
                self.subscribers.discard(subscriber)

    def replay_count(self) -> int:
        return len(self.events) + (1 if self._compact_kind is not None else 0)

    def replay_events(self):
        yield from self.events
        if self._compact_kind is not None:
            payload = {
                "text": "".join(self._compact_parts),
                "turn_id": self.turn_id,
                "event_seq": self._compact_last_seq,
            }
            if self.parent_turn_id:
                payload["parent_turn_id"] = self.parent_turn_id
            yield {
                "event": self._compact_kind,
                "data": json.dumps(payload, ensure_ascii=False),
            }

    def finish(self) -> None:
        if self.done:
            return
        self._flush_compact_text()
        self.done = True
        self.steering_ready.set()
        self.finished_at = time.time()
        if self.perf_status == "unknown":
            self.perf_status = "cancelled" if self.cancelled else "completed"
            if self.cancelled:
                self.perf_error_kind = "cancelled"
        if self.perf_post_started:
            self.perf_post_ms = obs.elapsed_ms(self.perf_post_started)
        if not self.is_continuation and not self._startup_perf_emitted:
            startup_status = (
                "cancelled" if self.cancelled
                else "failed" if self.perf_status == "failed"
                else "ready" if self.perf_query_started
                else "unknown"
            )
            self.emit_startup_perf(
                startup_status,
                failure_phase=(
                    self.perf_startup_failure_phase
                    if self.perf_startup_failure_phase != "none"
                    else "startup" if startup_status == "failed"
                    else "none"
                ),
            )
        if not self._perf_emitted:
            self._perf_emitted = True
            try:
                _perf_event(
                    "chat.turn",
                    sid8=obs.short_id(self.session_id),
                    turn8=obs.short_id(self.turn_id),
                    source=self.activity_source,
                    model=self.model,
                    client=self.perf_client,
                    client_ms=self.perf_client_ms,
                    preflight_ms=self.perf_preflight_ms,
                    first_event_ms=self.perf_first_event_ms,
                    first_visible_ms=self.perf_first_visible_ms,
                    result_ms=self.perf_result_ms,
                    post_ms=self.perf_post_ms,
                    total_ms=obs.elapsed_ms(self.perf_started),
                    status=self.perf_status,
                    error_kind=self.perf_error_kind,
                    background_count=self.perf_background_count,
                    final_mode=self.perf_final_mode,
                )
            except Exception:
                # Observability must never change turn lifecycle semantics.
                pass
        for subscriber in tuple(self.subscribers):
            subscriber.close()
        self.subscribers.clear()
        # Keep the compacted replay for late desktop reconnects during the grace TTL.

    def close(self) -> None:
        self.steering_ready.set()
        for subscriber in tuple(self.subscribers):
            subscriber.close_reader()
            subscriber.close()
        self.subscribers.clear()
        self.events.close()

    def subscribe(
        self,
        *,
        mobile: bool = False,
        last_event_seq: int = 0,
    ) -> _TurnSubscriber:
        """Attach a full reader or resume strictly after ``last_event_seq``.

        Sequence zero is the cold/legacy path and replays the complete coalesced
        spool. A positive checkpoint uses the bounded exact-wire buffer, which
        includes individual text/thinking deltas. If every missing sequence is
        no longer available, emit one explicit canonical-history fallback.
        """
        _ = mobile
        self._flush_compact_text()
        requested = max(0, int(last_event_seq or 0))
        replay = self.events.open_reader()
        initial_events: list[dict] = []
        resync_payload: dict | None = None
        if requested > 0:
            replay.seek(self.events.size())
            latest = self._event_seq
            earliest = self._resume_events[0][0] if self._resume_events else latest + 1
            if requested > latest or (
                requested < latest and requested < earliest - 1
            ):
                replay.close()
                replay = None
                resync_payload = {
                    "reason": "replay_gap",
                    "fallback": "canonical_history",
                    "retryable": False,
                    "turn_id": self.turn_id,
                    "requested_event_seq": requested,
                    "earliest_event_seq": earliest if earliest <= latest else None,
                    "latest_event_seq": latest,
                }
            else:
                initial_events = [
                    event for seq, event, _ in self._resume_events
                    if seq > requested
                ]
        subscriber = _TurnSubscriber(
            replay,
            initial_events=initial_events,
            resync_payload=resync_payload,
            skip_from=self.events.size(),
        )
        if self.done:
            subscriber.close()
        elif resync_payload is None:
            self.subscribers.add(subscriber)
        return subscriber

    def unsubscribe(self, subscriber: _TurnSubscriber) -> None:
        self.subscribers.discard(subscriber)
        subscriber.close_reader()
        subscriber.close()


# In-flight turns by session id. Lookup target for reconnect.
_active_turns: dict[str, TurnBroadcast] = {}
# Grace-keep for JUST-finished turns. Problem it solves: a server-side queue
# drain auto-starts the next turn and (for fast turns) finishes + pops it from
# _active_turns BEFORE the browser's reconnect SSE attaches. The reconnect then
# sees no active turn and the drained turn never renders live — the user must
# refresh. We keep the finished broadcast here for a short TTL so a slightly-late
# reconnect still gets the full replay (events + done sentinel). One per sid;
# a new turn for the same sid overwrites it. Each entry owns an active expiry
# callback so its replay spool cannot outlive the grace period without access.
_recent_turns: dict[str, TurnBroadcast] = {}
_recent_turn_expiry_handles: dict[str, asyncio.TimerHandle] = {}
_RECENT_TURN_TTL = env_int("MUSELAB_RECENT_TURN_TTL", 60, min_value=1)


def _evict_recent_turn(
    session_id: str,
    expected_turn_id: str = "",
) -> bool:
    """Close one grace replay only if it is still the expected turn."""
    broadcast = _recent_turns.get(session_id)
    if broadcast is None:
        return False
    if expected_turn_id and broadcast.turn_id != expected_turn_id:
        return False
    _recent_turns.pop(session_id, None)
    handle = _recent_turn_expiry_handles.pop(session_id, None)
    if handle is not None:
        handle.cancel()
    broadcast.close()
    return True


def _remember_recent_turn(session_id: str, broadcast: TurnBroadcast) -> None:
    """Stash a just-finished broadcast for bounded reconnect grace."""
    previous = _recent_turns.get(session_id)
    if previous is not None and previous is not broadcast:
        _evict_recent_turn(session_id, previous.turn_id)
    old_handle = _recent_turn_expiry_handles.pop(session_id, None)
    if old_handle is not None:
        old_handle.cancel()
    _recent_turns[session_id] = broadcast
    age = max(0.0, time.time() - float(broadcast.finished_at or 0))
    remaining = max(0.001, _RECENT_TURN_TTL - age)
    _recent_turn_expiry_handles[session_id] = (
        asyncio.get_running_loop().call_later(
            remaining,
            _evict_recent_turn,
            session_id,
            broadcast.turn_id,
        )
    )


def _get_recent_turn(session_id: str) -> TurnBroadcast | None:
    """Return a still-fresh just-finished broadcast for `session_id`, or None.
    Evicts the entry and its replay spool if it has aged past the TTL."""
    broadcast = _recent_turns.get(session_id)
    if broadcast is None:
        return None
    if time.time() - broadcast.finished_at > _RECENT_TURN_TTL:
        _evict_recent_turn(session_id, broadcast.turn_id)
        return None
    return broadcast
# Each CLI subprocess holds ~30-50 MB RSS. Runtime-owned LRU/lock aliases keep
# the historical chat facade patchable while bounding the shared client pool.
_client_lru = chat_runtime.CLIENT_LRU
_CLIENT_POOL_CAP = env_int("MUSELAB_CLIENT_POOL_CAP", 3, min_value=1)
_lock = chat_runtime.CLIENT_LOCK

# Exactly one SDK operation may own a session's CLI stream at a time. The
# interactive turn mutex only covers /stream turns; scheduler and /compact
# also call query()/receive_response() and therefore share this lock.
_session_runtime_locks: dict[str, asyncio.Lock] = {}
# Queue drain has a separate per-session mutex. It cannot reuse the runtime
# lock because _maybe_drain_queue() calls _start_turn(), which acquires that
# lock while starting/reusing the SDK client. Serialising the higher-level
# dequeue + launch transaction prevents concurrent completion/enqueue kicks
# from popping adjacent FIFO items before either reserves _active_turns[sid].
_queue_drain_locks: dict[str, asyncio.Lock] = {}
_queue_drain_tasks: dict[str, asyncio.Task] = {}
# Recoverable runtime-rollover conflicts (404/409 while a background watcher is
# handing ownership to its successor) need a retained wake-up even when no new
# enqueue or task notification arrives. One delayed retry per session is enough.
_queue_drain_retry_tasks: dict[str, asyncio.Task] = {}
# A queue POST can land while the coalesced drain is busy forking a detached
# runtime. Remember that wakeup instead of dropping it; once the first drain
# exits, one follow-up pass migrates any item accepted after its queue snapshot.
_queue_drain_rekicks: set[str] = set()
# Session settings can change after a turn reserves `_active_turns[sid]` but
# before its detached query task starts. Disconnecting in that gap kills the
# freshly selected client. Mark rebuilds and consume them at the next safe SDK
# boundary instead.
_pending_runtime_rebuilds: set[str] = set()
# Child session -> (source session, dropped user turn).  Presence means the
# client's native truncating resume has connected but no user query has yet
# crossed the SDK transport commit point.  The same UUID tuple is also stored
# durably in sessions.py so a service restart can rebuild the intent.
_native_retry_commits: dict[str, tuple[str, str]] = {}


def _session_runtime_lock_for(session_id: str) -> asyncio.Lock:
    return _session_runtime_locks.setdefault(session_id, asyncio.Lock())


def _queue_drain_lock_for(session_id: str) -> asyncio.Lock:
    return _queue_drain_locks.setdefault(session_id, asyncio.Lock())

# ---------------------------------------------------------------------------
# Cross-turn background tasks (SDK-native run_in_background)
# ---------------------------------------------------------------------------
# Phase 0 probe (2026-06-03) proved that a TaskNotification (terminal status)
# for an SDK background task usually arrives AFTER the turn's ResultMessage —
# i.e. cross-turn. The in-turn dispatch (Phase 1) handles the case where the
# notification happens to land before ResultMessage; for everything else we
# need to (a) keep the originating CLI client ALIVE past the turn (disconnect()
# kills the subprocess and would abort the running task), and (b) keep a single
# detached reader draining receive_messages() so the buffered notification gets
# delivered. These two maps coordinate that:
#   _sessions_with_inflight_tasks: session_id -> set of task_id still running.
#       Presence here exempts the session's clients from LRU eviction.
#   _task_watchers: session_id -> the detached asyncio.Task doing the reading.
# Single-reader invariant: a watcher and a live turn must never read the same
# client stream concurrently. While a watcher owns the stream, new user input
# remains queued; the next turn starts only after the watcher has delivered the
# task settlement + auto-continuation and released the reader.
_sessions_with_inflight_tasks: dict[str, set[str]] = {}
_task_watchers: dict[str, asyncio.Task] = {}
# task_id -> epoch seconds when the task was first pinned. Watcher generations
# are replaceable, but their timeout is derived from this stable launch time so
# a respawn cannot grant an orphaned task another full lease. The session-list
# and /active paths also reap individual expired pins before publishing state.
_bg_task_pinned_at: dict[str, float] = {}
# Small post-response maintenance tasks, currently used to refresh compacted
# transcript counts after the verified compact result has already reached the
# browser. Strong references prevent accidental mid-run garbage collection.
_maintenance_tasks: set[asyncio.Task] = set()

# ---------------------------------------------------------------------------
# SDK-native scheduled tasks (CronCreate / CronDelete / CronList)
# ---------------------------------------------------------------------------
# Native schedules live inside the pooled Claude CLI process. They are neither
# MuseLab's legacy scheduler nor durable application records: disconnecting the
# owning runtime destroys them. Keep only privacy-safe control metadata in RAM
# (job id / cadence flags), pin that runtime against LRU eviction, and project
# the count to the session list. Prompts and result bodies are never cached.
_sdk_cron_jobs: dict[str, dict[str, dict[str, Any]]] = {}
_sdk_cron_tool_calls: dict[
    chat_runtime.ClientKey, dict[str, dict[str, Any]]
] = {}
_sdk_cron_state_lock = threading.RLock()
_SDK_CRON_JOB_ID = re.compile(r"^[A-Za-z0-9_-]{1,128}$")
_SDK_CRON_CREATE_RESULT = re.compile(
    r"Scheduled\s+(?:recurring|one-time)\s+job\s+([A-Za-z0-9_-]{1,128})",
    re.IGNORECASE,
)
_SDK_CRON_DELETE_RESULT = re.compile(
    r"Cancelled\s+job\s+([A-Za-z0-9_-]{1,128})",
    re.IGNORECASE,
)
_SDK_CRON_LIST_LINE = re.compile(
    r"^([A-Za-z0-9_-]{1,128})\s+[—-]\s+", re.MULTILINE,
)


@dataclass
class _ScheduledDelivery:
    key: chat_runtime.ClientKey
    broadcast: TurnBroadcast
    render_state: dict[str, Any]
    pending_tasks: dict[str, dict[str, Any]] = dataclass_field(
        default_factory=dict)
    subagent_mux: Any = None
    registration_task: asyncio.Task | None = None


_sdk_scheduled_deliveries: dict[
    chat_runtime.ClientKey, _ScheduledDelivery
] = {}
# A hidden runtime owns the real Claude continuation after one of its
# background tasks settles.  When that runtime has already rolled over, the
# resulting assistant text is delivered to the newest visible successor as a
# presentation-only snapshot.  Keep the delivery task strongly referenced and
# keyed by the durable outbox identity so retries cannot create two writers.
_runtime_continuation_delivery_tasks = (
    chat_overlays.RUNTIME_CONTINUATION_DELIVERY_TASKS
)
# SDK disconnect fences are runtime-owned. Keep the exact shared containers and
# deadline visible through the historical chat facade.
_session_disconnect_tasks = chat_runtime.SESSION_DISCONNECT_TASKS
_session_disconnect_failed = chat_runtime.SESSION_DISCONNECT_FAILED
_CLIENT_DISCONNECT_DEADLINE_S = chat_runtime.CLIENT_DISCONNECT_DEADLINE_S
_ATTACHMENT_SHUTDOWN_JOIN_S = 4.0
# Attachment workers and their finalizers form a protected lifecycle fence.
# Tests may reduce the bounded join while production retains enough time for
# ordinary fsync and document conversion to finish.
# Turn/scheduler owners cancelled by session deletion can outlive their first
# bounded join. Keep the exact handles keyed by session so a retry cannot race
# ahead and purge the transcript while an earlier owner is still unwinding.
_session_runtime_cleanup_tasks: dict[str, set[asyncio.Task]] = {}
# A disk purge may itself exceed the public HTTP deadline (slow/networked
# workspace). Repeated DELETEs join the same owner instead of starting a second
# destructive transaction. Completed, unobserved owners expire after 5 min.
_session_purge_tasks: dict[str, asyncio.Task] = {}
_SESSION_DELETE_DEADLINE_S = 30.0
# Original user-turn start (epoch seconds) retained across the detached gap and
# every headless continuation. `/active` returns it after a page refresh so the
# footer timer keeps counting from the real turn start instead of restarting.
_background_turn_started_at: dict[str, float] = {}
# Originating user-turn identity retained across the watcher gap. Every
# continuation gets its own turn_id and exposes this value as parent_turn_id.
_background_origin_turn_id: dict[str, str] = {}
# Monotonic per-session ownership token for detached continuation readers.
# Cancelling a task is cooperative, so a replaced watcher can receive one more
# message before CancelledError lands. Generation checks keep that stale reader
# from opening or closing a newer continuation in that window.
_continuation_generations: dict[str, int] = {}
# Keep this comfortably above the common ``sleep 1800; ...`` anti-pattern.  A
# deadline exactly equal to the requested delay races the command after sleep
# against MuseLab's timeout recovery, which is both surprising and unsafe.
# The watcher still enforces a hard bound; long delayed work belongs in the
# scheduler rather than an SDK-owned Bash task.
_TASK_WATCH_TIMEOUT = env_int("MUSELAB_TASK_WATCH_TIMEOUT", 3600, min_value=60)
# Once the absolute lease expires, first ask the CLI to stop each task and give
# its terminal TaskNotification a short chance to arrive.  A control ack is not
# itself proof that the child stopped, so expiry may release pins only after a
# terminal notification or a fully-confirmed SDK disconnect.
_TASK_STOP_ACK_TIMEOUT_S = max(
    0.1, env_float("MUSELAB_TASK_STOP_ACK_TIMEOUT_S", 3.0))
_TASK_STOP_SETTLE_GRACE_S = max(
    0.1, env_float("MUSELAB_TASK_STOP_SETTLE_GRACE_S", 5.0))
_TASK_TERMINATION_RETRY_S = max(
    0.1, env_float("MUSELAB_TASK_TERMINATION_RETRY_S", 5.0))
# After the LAST in-flight task delivers its terminal notification, the CLI
# auto-continues a short turn (model reacts to the result). The probe (§3.4)
# measured it landing ~1.3s later, but it's not strictly guaranteed for every
# task type / status. So once all tasks have settled and a continuation
# broadcast is open, the watcher waits at most this long for the auto-continue's
# AssistantMessage + ResultMessage before closing the continuation and
# unpinning — bounding the worst case (no auto-continue ever comes) instead of
# holding the client + the _active_turns slot for the full _TASK_WATCH_TIMEOUT.
_CONTINUATION_GRACE = env_int("MUSELAB_CONTINUATION_GRACE", 8, min_value=2)
# Short grace for USER-STOPPED tasks: the CLI doesn't auto-continue after a
# deliberate stop, so the watcher only needs a token window before closing
# the continuation (frees the attached FE from an idle "streaming…" footer).
_STOPPED_CONTINUATION_GRACE = env_int(
    "MUSELAB_STOPPED_CONTINUATION_GRACE", 2, min_value=1)
# task_id -> description, surviving across the turn that started the task. The
# per-turn inflight_tasks dict is local to a turn. This module-level cache keeps
# the label available to the detached watcher and across watcher replacement.
# Populated on TaskStarted, consumed+removed on settle.
_bg_task_descriptions: dict[str, str] = {}
_bg_task_tool_use_ids: dict[str, str] = {}
# Serialize one session's live task-card authority/persistence/settlement
# transaction across the turn dispatcher and its detached watcher. Disk work
# still runs in workers; this lock only prevents two terminal shapes from
# persisting one outcome while the other wins the in-memory dedup gate.
_runtime_task_storage_locks: dict[str, asyncio.Lock] = {}


def _runtime_task_storage_lock_for(session_id: str) -> asyncio.Lock:
    lock = _runtime_task_storage_locks.get(session_id)
    if lock is None:
        lock = asyncio.Lock()
        _runtime_task_storage_locks[session_id] = lock
    return lock

# Compatibility aliases: deletion, overlays, tests, and local tooling must
# observe the exact containers owned by the extracted successor lifecycle.
_runtime_rollover_locks = chat_successor.RUNTIME_ROLLOVER_LOCKS
_runtime_prewarm_tasks = chat_successor.RUNTIME_PREWARM_TASKS
_session_title_locks = chat_successor.SESSION_TITLE_LOCKS


def _runtime_rollover_lock_for(session_id: str) -> asyncio.Lock:
    return chat_successor.runtime_rollover_lock_for(session_id)


def _session_title_lock(session_id: str):
    return chat_successor.session_title_lock(session_id)


def _runtime_task_overlay(
    session_id: str,
    task_id: str,
    *,
    state: str,
    tool_use_id: str | None = None,
    description: str | None = None,
    summary: str | None = None,
    output_file: str | None = None,
    usage: dict | None = None,
) -> None:
    """Persist one task-card patch on its owner and runtime successor.

    This is deliberately sidecar-only: the successor never receives the old
    CLI's implicit continuation in its transcript/model context.
    """
    owner = str(session_id or "")
    tid = str(task_id or "")
    if not owner or not tid:
        return
    resolved_tool_use = str(
        tool_use_id or _bg_task_tool_use_ids.get(tid) or ""
    )
    fields: dict[str, Any] = {
        "state": (
            "stopped" if state == "killed"
            else "completed" if state == "done"
            else state
        ),
        "owner_session_id": owner,
        "updated_at": int(time.time() * 1000),
    }
    resolved_description = description or _bg_task_descriptions.get(tid)
    if resolved_tool_use:
        fields["tool_use_id"] = resolved_tool_use
    if resolved_description:
        fields["description"] = resolved_description
    if summary:
        fields["summary"] = summary
    if output_file:
        fields["output_file"] = output_file
    if usage:
        fields["usage"] = dict(usage)
    try:
        current = owner
        seen: set[str] = set()
        for _ in range(32):
            if not current or current in seen:
                break
            seen.add(current)
            sess.set_runtime_task_overlay(current, tid, **fields)
            current_meta = sess.get_session_meta(current) or {}
            current = str(current_meta.get("runtime_successor") or "")
    except Exception as exc:
        sys.stderr.write(
            f"[chat] runtime task overlay failed sid={owner[:8]} "
            f"task={tid} exc={type(exc).__name__}\n"
        )


async def _runtime_task_overlay_owned(
    session_id: str,
    task_id: str,
    **fields: Any,
) -> None:
    """Persist a task-card patch without blocking the shared event loop."""
    await obs.to_thread_io(
        "chat.runtime_task_overlay",
        session_id,
        _runtime_task_overlay,
        session_id,
        task_id,
        owned=True,
        **fields,
    )


def _record_background_task_launch(
    session_id: str,
    task_id: str,
    *,
    tool_use_id: str | None = None,
    description: str | None = None,
    output_file: str | None = None,
) -> bool:
    tid = str(task_id or "")
    if not tid:
        return False
    # A rollover successor can replay launch-shaped records copied from its
    # predecessor. Never let that replay acquire process ownership or a pin.
    inherited = sess.get_authoritative_runtime_task_overlays(
        session_id).get(tid, {})
    inherited_owner = str(inherited.get("owner_session_id") or "")
    if inherited_owner and inherited_owner != session_id:
        return False
    if str(inherited.get("state") or "") in {
        "completed", "failed", "stopped", "done",
    }:
        # Task ids are lifecycle identities, not reusable names. A start seen
        # after this same owner's terminal record is stream replay, not a new
        # process, and must not recreate its watcher pin.
        return False
    if tool_use_id:
        _bg_task_tool_use_ids[tid] = str(tool_use_id)
    if description:
        _bg_task_descriptions[tid] = str(description)
    _runtime_task_overlay(
        session_id,
        tid,
        state="running",
        tool_use_id=tool_use_id,
        description=description,
        output_file=output_file,
    )
    return True


async def _record_background_task_launch_owned(
    session_id: str,
    task_id: str,
    *,
    tool_use_id: str | None = None,
    description: str | None = None,
    output_file: str | None = None,
) -> bool:
    """Async event-loop-safe counterpart of task launch persistence."""
    tid = str(task_id or "")
    if not tid:
        return False
    async with _runtime_task_storage_lock_for(session_id):
        inherited = await obs.to_thread_io(
            "chat.runtime_task_launch_read",
            session_id,
            sess.get_authoritative_runtime_task_overlays,
            session_id,
        )
        current = inherited.get(tid, {})
        inherited_owner = str(current.get("owner_session_id") or "")
        if inherited_owner and inherited_owner != session_id:
            return False
        if str(current.get("state") or "") in {
            "completed", "failed", "stopped", "done",
        }:
            return False
        if tool_use_id:
            _bg_task_tool_use_ids[tid] = str(tool_use_id)
        if description:
            _bg_task_descriptions[tid] = str(description)
        await _runtime_task_overlay_owned(
            session_id,
            tid,
            state="running",
            tool_use_id=tool_use_id,
            description=description,
            output_file=output_file,
        )
        return True


def _session_has_live_watcher(session_id: str) -> bool:
    watcher = _task_watchers.get(session_id)
    return watcher is not None and not watcher.done()


def _session_has_scheduled_delivery(session_id: str) -> bool:
    return any(
        key[0] == session_id and not delivery.broadcast.done
        for key, delivery in _sdk_scheduled_deliveries.items()
    )


def _session_runtime_busy(session_id: str) -> bool:
    """One authoritative busy boundary for turns and detached task readers."""
    active = _active_turns.get(session_id)
    return bool(
        (active is not None and not active.done)
        or _sessions_with_inflight_tasks.get(session_id)
        or _session_has_live_watcher(session_id)
        or _session_has_scheduled_delivery(session_id)
    )


def _terminal_task_update(msg: TaskUpdatedMessage) -> dict | None:
    """Normalize a terminal ``task_updated`` patch to task_notification shape.

    Recent Claude CLI builds may emit only ``TaskUpdatedMessage`` for a
    terminal transition (notably a user-stopped task reports ``killed``) and
    suppress the older ``TaskNotificationMessage``.  Treat both as equal
    lifecycle truth so task pins and UI cards cannot remain stuck forever.
    """
    status = getattr(msg, "status", None)
    if status not in TERMINAL_TASK_STATUSES:
        return None
    patch = getattr(msg, "patch", None)
    patch = patch if isinstance(patch, dict) else {}
    frontend_status = "stopped" if status == "killed" else status
    summary = patch.get("summary") or patch.get("error") or patch.get("result") or ""
    if not isinstance(summary, str):
        try:
            summary = json.dumps(summary, ensure_ascii=False)
        except (TypeError, ValueError):
            summary = str(summary)
    return {
        "task_id": getattr(msg, "task_id", "") or "",
        "tool_use_id": patch.get("tool_use_id") or patch.get("toolUseId"),
        "status": frontend_status,
        "summary": summary,
        "output_file": patch.get("output_file") or patch.get("outputFile") or "",
        "usage": dict(patch.get("usage") or {}) if isinstance(patch.get("usage"), dict) else {},
    }


# ---------------------------------------------------------------------------
# In-flight turn persistence (survives muselab process restart)
# ---------------------------------------------------------------------------
# Why: `_active_turns` is in-memory only. If muselab restarts mid-turn
# (systemd OOM-kill / manual restart / crash / `systemctl --user restart`),
# the user's prompt is lost and they may not even realize the turn never
# replied. We write a tiny sidecar JSON to disk per in-flight turn, delete
# it on clean completion, and on process startup scan for orphans to tell
# the frontend "you had N unfinished turns last session."
#
# Design choices:
# - Sidecar lives under `sessions/active_turns/<sid>.json`, not `~/.muselab/`,
#   because SESS_DIR already exists, is gitignored, and is the natural sibling
#   for per-session state.
# - We do NOT auto-resume. Auto-resume would burn tokens on conversations the
#   user has already abandoned and bypass their "should I rephrase?" judgment.
#   Frontend gets the list + sids and toasts the user — they decide.
# - The sidecar owns the user's pending intent only until canonical commit or a
#   durable failed/cancelled display snapshot takes over. A process-crash orphan
#   is converted to that snapshot before a later turn can reuse the same sid.
# - No periodic touch / last_event_ts. Adding background touch task per turn
#   means N file writes per second across active turns — not worth the
#   complexity for "stale by 30s vs 30min" UX granularity. `started_at` is
#   enough to show "5 min ago" in the toast.

_ACTIVE_TURN_DIR = sess.SESS_DIR / "active_turns"
_ACTIVE_TURN_DIR.mkdir(exist_ok=True)


def _active_turn_path(sid: str) -> Path:
    return _ACTIVE_TURN_DIR / f"{sid}.json"


def _write_active_turn_sidecar(bc: TurnBroadcast) -> bool:
    """Persist the user intent before the SDK can accept the turn.

    Callers at the initial submission boundary must fail closed when this
    returns False; later rewrites only enrich an already-durable record.
    """
    try:
        raw = bc.user_text or ""
        first_line = raw.strip().splitlines()[0] if raw.strip() else ""
        preview = first_line if len(first_line) <= 200 else first_line[:199] + "…"
        atomic_write_text(
            _active_turn_path(bc.session_id),
            json.dumps({
                "sid": bc.session_id,
                "user_text": raw,
                "user_text_preview": preview,
                "model": bc.model,
                "started_at": bc.started_at,
                "turn_id": bc.turn_id,
                "user_images": bc.user_images,
                "user_docs": bc.user_docs,
                "staged_attachment_ids": list(bc.staged_attachment_ids),
                "transcript_boundary": dict(bc.transcript_boundary or {}),
            }, ensure_ascii=False),
            mode=0o600,
        )
        return True
    except Exception as e:
        sys.stderr.write(
            f"[chat] failed to write active-turn sidecar "
            f"sid={obs.short_id(bc.session_id)} exc={type(e).__name__}\n")
        sys.stderr.flush()
        return False


def _delete_active_turn_sidecar(sid: str) -> None:
    """Delete pending intent only after canonical/snapshot ownership exists."""
    try:
        p = _active_turn_path(sid)
        if p.exists():
            p.unlink()
    except Exception:
        pass


def _read_active_turn_sidecar(sid: str) -> dict | None:
    """Read one recovery sidecar without mutating event-loop-owned state."""
    try:
        data = json.loads(_active_turn_path(sid).read_text(encoding="utf-8"))
    except Exception:
        return None
    return data if isinstance(data, dict) else None


def _retain_active_turn_for_recovery(sid: str) -> None:
    """Expose a still-owned sidecar to history without waiting for restart."""
    data = _read_active_turn_sidecar(sid)
    if data is not None:
        _interrupted_at_startup[sid] = data


async def _settle_active_turn_sidecar_owned(
    sid: str,
    *,
    release: bool,
) -> None:
    """Settle pending intent I/O off-loop before publishing terminal state."""
    if release:
        await obs.to_thread_io(
            "chat.active_turn_release",
            sid,
            _delete_active_turn_sidecar,
            sid,
            file_path=_active_turn_path(sid),
            owned=True,
        )
        return
    data = await obs.to_thread_io(
        "chat.active_turn_recovery_read",
        sid,
        _read_active_turn_sidecar,
        sid,
        file_path=_active_turn_path(sid),
    )
    if data is not None:
        _interrupted_at_startup[sid] = data


def _release_active_turn_sidecar(sid: str) -> bool:
    """Release pending-intent ownership at a durable terminal boundary.

    Detached task state has its own runtime overlays/outbox. Keeping the user
    intent sidecar after the canonical turn committed conflates those two owners
    and can duplicate an already-successful prompt after a process restart.
    """
    _delete_active_turn_sidecar(sid)
    return True


def _scan_interrupted_turns_at_startup() -> dict[str, dict]:
    """Read all sidecars left over from a previous process. Runs once at
    module import. Keeps the files on disk until the user dismisses each
    one — that way two browsers can both see the notification, and a
    second muselab restart still surfaces undismissed entries."""
    out: dict[str, dict] = {}
    if not _ACTIVE_TURN_DIR.exists():
        return out
    for p in _ACTIVE_TURN_DIR.glob("*.json"):
        try:
            data = json.loads(p.read_text(encoding="utf-8"))
            sid = data.get("sid") or p.stem
            out[sid] = data
        except Exception as e:
            sys.stderr.write(
                f"[chat] skipping malformed active-turn sidecar "
                f"exc={type(e).__name__}\n")
    return out


# Snapshot taken once at process startup. Endpoints serve from this dict;
# starting a new turn for an sid here auto-dismisses (the new turn supersedes
# the old in-flight). Don't re-scan disk on each request — once consumed by a
# browser dismiss, the user has acknowledged.
_interrupted_at_startup: dict[str, dict] = _scan_interrupted_turns_at_startup()


# Global aggregate stats (across all sessions). cache_read / cache_creation
# come from the Anthropic prompt cache — high cache_read ratio means subsequent
# turns are much cheaper. Surfacing this in the UI lets the user see the value
# of long sessions vs constantly opening new ones.
_stats = {"total_cost_usd": 0.0, "total_messages": 0,
          "total_input_tokens": 0, "total_output_tokens": 0,
          "total_cache_read_tokens": 0, "total_cache_creation_tokens": 0}

# Latest Pro/Max rate-limit state, keyed by window type (five_hour /
# seven_day / seven_day_opus / seven_day_sonnet / overage). The SDK pushes a
# RateLimitEvent whenever the limit state changes; each event carries ONE
# window's RateLimitInfo (utilization 0.0–1.0, status, resets_at). We keep the
# most-recent value per window so a fresh page can fetch a snapshot
# (GET /api/chat/rate-limit) while live deltas arrive over SSE. Empty until the
# first event lands this process — the CLI only emits these for OAuth
# (Pro/Max) sessions, never for third-party API-key vendors.
_rate_limit_state: dict[str, dict] = {}
_rate_limit_updated_at: float = 0.0


def _rate_limit_payload(info) -> dict:
    """Serialize a SDK RateLimitInfo into a JSON-safe dict. Every field via
    getattr-default so a future SDK adding/renaming fields degrades gracefully
    instead of crashing the turn (same discipline as the Task* handlers)."""
    return {
        "status": getattr(info, "status", None),
        "rate_limit_type": getattr(info, "rate_limit_type", None),
        "utilization": getattr(info, "utilization", None),
        "resets_at": getattr(info, "resets_at", None),
        "overage_status": getattr(info, "overage_status", None),
        "overage_resets_at": getattr(info, "overage_resets_at", None),
        "overage_disabled_reason": getattr(info, "overage_disabled_reason", None),
    }


def _record_rate_limit(info) -> dict:
    """Store the latest RateLimitInfo under its window key and return the
    JSON-safe payload (with an updated_at stamp) for SSE emission."""
    global _rate_limit_updated_at
    payload = _rate_limit_payload(info)
    payload["updated_at"] = _rate_limit_updated_at = time.time()
    # rate_limit_type is Optional; bucket an untyped event under "_" so it
    # still surfaces rather than vanishing.
    _rate_limit_state[payload.get("rate_limit_type") or "_"] = payload
    return payload

# Per-session current state — populated from the LATEST ResultMessage of each
# session. The model's `input_tokens` on a turn ≈ current context window size,
# so tracking the most-recent value gives a meaningful "context meter".
_session_usage: dict[str, dict] = {}     # sid -> {input_tokens, output_tokens,
                                          #         cache_read_tokens,
                                          #         cache_creation_tokens,
                                          #         total_cost_usd, last_turn_at}
# Exact logical turn owning each in-memory snapshot. A post-done SDK context
# probe may finish after the next turn has populated `_session_usage`; late
# refinement must match this owner rather than write by session id alone.
_session_usage_turns: dict[str, str] = {}
_USAGE_SUMMARY_SCHEMA = 1
_USAGE_SOURCE_UNSET = object()

# Per-model context windows. Used as the meter's denominator when a SDK
# get_context_usage() truth isn't available (first turn of a session, or
# any third-party model where CLI's tokenizer / window inference is
# unreliable). Numbers verified from each vendor's docs:
#   - Anthropic:   tygartmedia.com / anthropic.com (Opus/Sonnet 4.6+ default
#                  to 1M on Pro/Max/Enterprise; Haiku 4.5 stays 200K)
#   - DeepSeek V4: api-docs.deepseek.com (V4 series ships 1M native context)
#   - Zhipu GLM:   glm-5.org / docs.z.ai (GLM-5 + GLM-4.7 both 200K context)
#   - MiniMax:     platform.minimax.io (M2.5 / M2.7 both 204_800, cline #10007
#                  PR fixed the prior 192K/245K misinformation)
#   - Codex/GPT-5: CLIProxyAPI's live Codex-client catalog (runtime source of
#                  truth); the values below mirror that catalog only as an
#                  offline fallback. Max output remains 128K at the gateway.
MODEL_CONTEXT_LIMITS = {
    # Anthropic — the bundled Claude Code CLI reports a 200K effective window
    # for these models (verified via get_context_usage: maxTokens=200000). The
    # 1M context is a beta tier (context-1m-2025-08-07 header / higher API
    # tier), NOT a silent Pro/Max auto-upgrade — the earlier 1M values here
    # made the meter read ~5x too low. This table is only the FALLBACK
    # denominator for sessions muselab hasn't measured yet; once a turn runs,
    # the SDK-reported maxTokens is persisted per-session
    # (sessions.set_session_ctx_window) and overrides this — so accounts that
    # genuinely have the 1M window auto-upgrade after their first turn.
    # Opus 5 is a 1M-context model on paper, but this table is deliberately
    # the conservative Pro/Max fallback (see the note above): the SDK reports
    # the real window on the first turn and that value wins. Filling 1M here
    # would put the meter back to reading ~5x low for subscription accounts.
    "claude-opus-5":                200_000,
    "claude-opus-4-8":              200_000,
    "claude-opus-4-7":              200_000,
    "claude-sonnet-4-6":            200_000,
    "claude-haiku-4-5-20251001":    200_000,
    # DeepSeek V4 series — 1M native, all SKUs.
    "deepseek-v4-pro":            1_000_000,
    "deepseek-v4-flash":          1_000_000,
    # DeepSeek V3 chat/reasoner SKUs — older 128K window kept.
    "deepseek-chat":                128_000,
    "deepseek-reasoner":            128_000,
    # Zhipu GLM 5 series — 200K context, 128K output cap.
    "glm-5.2-internal":              200_000,
    "glm-5.2":                       200_000,
    "glm-5":                         200_000,
    "glm-5-air":                     200_000,
    "glm-4.7":                       200_000,
    "glm-4-plus":                   128_000,   # older 4-plus stayed 128K
    # MiniMax — 204_800 exactly, per platform.minimax.io spec.
    "minimax-m2.7":                 204_800,
    "minimax-m2.7-highspeed":       204_800,
    "minimax-m2.5":                 204_800,
    # Codex Gateway — raw `context_window` fallbacks matching CLIProxyAPI's
    # Codex client catalog. Runtime discovery via `/v1/models?client_version`
    # wins; these values only cover a temporarily unavailable/older gateway.
    # `_context_limit_details()` applies Codex's default 95% usable-input ratio,
    # so the meter denominator is effective capacity rather than these raw values.
    "codex:gpt-5.6-sol":              372_000,
    "codex:gpt-5.6-terra":            372_000,
    "codex:gpt-5.6-luna":             372_000,
    "codex:gpt-5.5":                  272_000,
    "codex:gpt-5.4":                  272_000,
    "codex:gpt-5.4-mini":             272_000,
    "codex:gpt-5.3-codex-spark":      128_000,
    # Direct-GPT route (no codex: prefix) — local patch mirrors codex:* above.
    "gpt-5.6-sol":                    372_000,
    "gpt-5.5":                        272_000,
}
DEFAULT_CONTEXT_LIMIT = 128_000
CODEX_DEFAULT_EFFECTIVE_CONTEXT_PERCENT = 95
_CONTEXT_CAPABILITY_CACHE_TTL = max(
    1.0, env_float("MUSELAB_CONTEXT_CATALOG_TTL_S", 300.0))
_CONTEXT_CAPABILITY_FAILURE_TTL = max(
    1.0, env_float("MUSELAB_CONTEXT_CATALOG_FAILURE_TTL_S", 15.0))
# (gateway base URL, canonical model) -> (monotonic timestamp, capability | None)
# No credential material is ever retained in this cache.
_CONTEXT_CAPABILITY_CACHE: dict[tuple[str, str], tuple[float, dict | None]] = {}
_CONTEXT_PROBE_LOG_WINDOW_S = max(
    15.0, env_float("MUSELAB_CONTEXT_CATALOG_LOG_WINDOW_S", 60.0))
# (canonical model, exception class) -> (last emitted monotonic time, suppressed)
_CONTEXT_PROBE_LOG_STATE: dict[tuple[str, str], tuple[float, int]] = {}


def _log_context_probe_failure(model: str, exc: BaseException) -> None:
    """Rate-limit repetitive catalog failures without retaining exception text."""
    now = time.monotonic()
    error_kind = type(exc).__name__
    key = (model, error_kind)
    previous = _CONTEXT_PROBE_LOG_STATE.get(key)
    if previous is not None and now - previous[0] < _CONTEXT_PROBE_LOG_WINDOW_S:
        _CONTEXT_PROBE_LOG_STATE[key] = (previous[0], previous[1] + 1)
        return
    suppressed = previous[1] if previous is not None else 0
    sys.stderr.write(
        f"[ctx-catalog] gateway context probe unavailable "
        f"model={model} exc={error_kind} suppressed={suppressed}\n"
    )
    sys.stderr.flush()
    _CONTEXT_PROBE_LOG_STATE[key] = (now, 0)


def _log_context_probe_recovery(model: str) -> None:
    """Emit one content-free recovery summary and clear failure state."""
    matches = [
        (key, state)
        for key, state in _CONTEXT_PROBE_LOG_STATE.items()
        if key[0] == model
    ]
    if not matches:
        return
    suppressed = sum(state[1] for _, state in matches)
    for key, _ in matches:
        _CONTEXT_PROBE_LOG_STATE.pop(key, None)
    sys.stderr.write(
        f"[ctx-catalog] gateway context probe recovered "
        f"model={model} suppressed={suppressed}\n"
    )
    sys.stderr.flush()


def _positive_int(v: Any) -> int:
    try:
        n = int(v or 0)
    except (TypeError, ValueError):
        return 0
    return n if n > 0 else 0


def _canonical_context_model(model: str) -> str:
    """Restore muselab's routing prefix for raw Codex transcript model ids."""
    value = (model or "").strip()
    if value.startswith("codex:"):
        return value
    # A bare GPT id can be an explicit first-class provider route (currently
    # the Zhipu-backed OpenAI group), not a damaged Codex transcript id. Exact
    # provider membership wins over the compatibility alias so that route does
    # not inherit CLIProxy-specific capabilities or private headers.
    provider = endpoints.lookup(value)
    if provider is not None and any(
        model_id == value for model_id, _label in provider.models
    ):
        return value
    candidate = f"codex:{value}"
    if candidate in MODEL_CONTEXT_LIMITS:
        return candidate
    return value


def _is_codex_gateway_model(model: str) -> bool:
    raw_model = (model or "").strip()
    if raw_model.startswith("codex:"):
        return True
    provider = endpoints.lookup(raw_model)
    # Exact non-Codex provider declarations take precedence over the legacy
    # raw-transcript alias. Provider identity/env key is more reliable than
    # model-name shape (`gpt-*` is shared by multiple routes).
    if provider is not None and any(
        model_id == raw_model for model_id, _label in provider.models
    ):
        return bool(
            provider.prefix.startswith("codex:")
            or provider.env_key == "CODEX_GATEWAY_API_KEY"
            or provider.env_key == "MUSELAB_PROVIDER_CODEX_API_KEY"
            or "codex" in (provider.display or "").lower()
        )
    canonical = _canonical_context_model(raw_model)
    if canonical.startswith("codex:"):
        return True
    return bool(
        provider
        and provider.supports_effort
        and provider.supports_thinking is False
        and provider.max_output_tokens == 128_000
        and "codex" in (provider.display or "").lower()
    )


def _context_limit_env_override(model: str) -> int:
    """Explicit operator override for third-party effective context windows.

    Model-specific env wins over provider-wide env. Example for codex:gpt-5.5:
    MUSELAB_CONTEXT_LIMIT_CODEX_GPT_5_5=180000. Provider-wide fallback:
    CODEX_GATEWAY_CONTEXT_LIMIT=180000.
    """
    model = _canonical_context_model(model)
    key = re.sub(r"[^A-Za-z0-9]+", "_", model.upper()).strip("_")
    names = []
    if key:
        names.append(f"MUSELAB_CONTEXT_LIMIT_{key}")
    if _is_codex_gateway_model(model):
        names.append("CODEX_GATEWAY_CONTEXT_LIMIT")
    names.append("MUSELAB_THIRD_PARTY_CONTEXT_LIMIT")
    for name in names:
        raw = os.environ.get(name, "").strip()
        if not raw:
            continue
        n = _positive_int(raw)
        if n:
            return n
    return 0


def _parse_codex_gateway_catalog(body: Any) -> dict[str, dict]:
    """Parse CLIProxyAPI's Codex-client catalog into model capabilities.

    `context_window` is the currently selected raw window, while
    `max_context_window` is only the ceiling for an explicit client override.
    Codex reserves 5% for system/tool/output headroom when the catalog omits an
    explicit `effective_context_window_percent`, matching the official client.
    """
    if not isinstance(body, dict):
        return {}
    rows = body.get("models")
    if not isinstance(rows, list):
        return {}
    parsed: dict[str, dict] = {}
    for item in rows:
        if not isinstance(item, dict):
            continue
        slug = str(item.get("slug") or item.get("id") or "").strip()
        raw = _positive_int(item.get("context_window"))
        if not slug or not raw:
            continue
        max_window = max(raw, _positive_int(item.get("max_context_window")))
        pct = _positive_int(item.get("effective_context_window_percent"))
        if not 1 <= pct <= 100:
            pct = CODEX_DEFAULT_EFFECTIVE_CONTEXT_PERCENT
        effective = max(1, raw * pct // 100)
        auto_compact = _positive_int(item.get("auto_compact_token_limit"))
        reasoning_levels: list[str] = []
        for level in item.get("supported_reasoning_levels") or []:
            effort = str(
                level.get("effort") if isinstance(level, dict) else level
            ).strip()
            if effort in _VALID_EFFORT and effort != "auto" \
                    and effort not in reasoning_levels:
                reasoning_levels.append(effort)
        service_tiers: list[str] = []
        for tier in item.get("service_tiers") or []:
            tier_id = str(
                tier.get("id") if isinstance(tier, dict) else tier
            ).strip()
            if tier_id and tier_id not in service_tiers:
                service_tiers.append(tier_id)
        parsed[slug] = {
            "context_limit": effective,
            "context_raw_limit": raw,
            "context_max_limit": max_window,
            "context_effective_percent": pct,
            "catalog_auto_compact_threshold": (
                min(auto_compact, effective) if auto_compact else 0),
            "context_limit_source": "gateway_catalog",
            "context_limit_is_estimate": False,
            "supported_reasoning_levels": reasoning_levels,
            "service_tiers": service_tiers,
        }
    return parsed


# Read-only compatibility when the live catalog is temporarily unavailable.
# This mirrors the tested CLIProxyAPI 7.2.145 baseline; the dynamic catalog
# always wins so a
# Gateway upgrade changes the UI without a MuseLab release.
_CODEX_CONTROL_FALLBACKS: dict[str, tuple[tuple[str, ...], tuple[str, ...]]] = {
    "gpt-5.6-sol": (("low", "medium", "high", "xhigh", "max", "ultra"),
                    ("priority",)),
    "gpt-5.6-terra": (("low", "medium", "high", "xhigh", "max", "ultra"),
                      ("priority",)),
    "gpt-5.6-luna": (("low", "medium", "high", "xhigh", "max"),
                     ("priority",)),
    "gpt-5.5": (("low", "medium", "high", "xhigh"), ("priority",)),
    "gpt-5.4": (("low", "medium", "high", "xhigh"), ("priority",)),
    "gpt-5.4-mini": (("low", "medium", "high", "xhigh"), ()),
    "gpt-5.3-codex-spark": (("low", "medium", "high", "xhigh"), ()),
}


def _model_control_capability(
    model: str, capability: dict | None = None,
) -> dict[str, Any]:
    """Return the frontend's exact per-model effort/Fast capability."""
    if endpoints.is_ducc_model(model):
        if endpoints.ducc_is_claude_model(model):
            return {
                "effort_levels": ["auto", *_SDK_EFFORT_LEVELS],
                "service_tiers": [],
                "supports_fast": False,
            }
        return {"effort_levels": [], "service_tiers": [],
                "supports_fast": False}
    provider = endpoints.lookup(model or "")
    if provider is None and not endpoints.is_third_party(model):
        return {
            "effort_levels": ["auto", *_SDK_EFFORT_LEVELS],
            "service_tiers": [],
            "supports_fast": False,
        }
    if not provider or not provider.supports_effort:
        return {"effort_levels": [], "service_tiers": [],
                "supports_fast": False}
    if _is_codex_gateway_model(model):
        slug = endpoints.normalize_model_id(_canonical_context_model(model))
        fallback_levels, fallback_tiers = _CODEX_CONTROL_FALLBACKS.get(
            slug, ((), ()))
        levels = tuple(
            capability.get("supported_reasoning_levels") or ()
            if capability is not None
            and "supported_reasoning_levels" in capability
            else fallback_levels
        )
        # An empty live list is authoritative (e.g. gpt-5.4-mini has no
        # priority service). Do not let truthiness accidentally re-enable the
        # static fallback in that case.
        tiers = tuple(
            capability.get("service_tiers") or ()
            if capability is not None and "service_tiers" in capability
            else fallback_tiers
        )
        return {
            "effort_levels": ["auto", *levels],
            # The UI uses the human-facing value `fast`; Gateway calls the
            # corresponding native Codex service tier `priority`.
            "service_tiers": ["fast"] if "priority" in tiers else [],
            "supports_fast": "priority" in tiers,
        }
    return {
        "effort_levels": ["auto", *_SDK_EFFORT_LEVELS],
        "service_tiers": [],
        "supports_fast": False,
    }


def _capability_from_model_item(item: Any, *, source: str) -> dict | None:
    """Read common model-limit fields from non-Codex-catalog endpoints."""
    if not isinstance(item, dict):
        return None
    limit = 0
    for field in (
        "max_input_tokens", "context_window", "context_length",
        "max_context_tokens",
    ):
        limit = _positive_int(item.get(field))
        if limit:
            break
    if not limit:
        return None
    return {
        "context_limit": limit,
        "context_raw_limit": limit,
        "context_max_limit": limit,
        "context_effective_percent": 100,
        "catalog_auto_compact_threshold": 0,
        "context_limit_source": source,
        "context_limit_is_estimate": False,
    }


async def _detect_gateway_context_capability(model: str) -> dict | None:
    """Discover the active Codex model window from CLIProxyAPI.

    CLIProxyAPI 7.2.80 exposes the authoritative Codex-client model catalog at
    `/v1/models?client_version`. Older generic `/v1/models` routes are retained
    as compatibility fallbacks. Successes and failures use separate short TTLs
    so a gateway upgrade/config change self-heals without restarting muselab.
    """
    if not _is_codex_gateway_model(model):
        return None
    canonical = _canonical_context_model(model)
    slug = endpoints.normalize_model_id(canonical)
    env = endpoints.env_override(canonical) or {}
    base = (env.get("ANTHROPIC_BASE_URL") or "").rstrip("/")
    key = env.get("ANTHROPIC_API_KEY") or env.get("ANTHROPIC_AUTH_TOKEN") or ""
    if not base:
        return None
    cache_key = (base, canonical)
    cached = _CONTEXT_CAPABILITY_CACHE.get(cache_key)
    if cached is not None:
        cached_at, capability = cached
        ttl = (_CONTEXT_CAPABILITY_CACHE_TTL if capability
               else _CONTEXT_CAPABILITY_FAILURE_TTL)
        if time.monotonic() - cached_at < ttl:
            if capability:
                _log_context_probe_recovery(canonical)
            return dict(capability) if capability else None

    headers: dict[str, str] = {}
    if key:
        headers.update({"x-api-key": key, "Authorization": f"Bearer {key}"})
    now = time.monotonic()
    try:
        import httpx
        timeout = max(0.2, env_float("MUSELAB_CONTEXT_CATALOG_TIMEOUT_S", 2.0))
        async with httpx.AsyncClient(timeout=timeout) as hc:
            # Query-string presence selects CLIProxyAPI's Codex-client catalog.
            r = await hc.get(f"{base}/v1/models?client_version", headers=headers)
            if r.status_code < 400:
                catalog = _parse_codex_gateway_catalog(r.json())
                if catalog:
                    for item_slug, capability in catalog.items():
                        item_model = _canonical_context_model(item_slug)
                        _CONTEXT_CAPABILITY_CACHE[(base, item_model)] = (
                            now, dict(capability))
                        # Keep the routing-prefixed cache key too. This makes
                        # newly-added gateway models cache correctly before
                        # muselab's static fallback table learns their slug.
                        _CONTEXT_CAPABILITY_CACHE[
                            (base, f"codex:{item_slug}")
                        ] = (now, dict(capability))
                    found = catalog.get(slug)
                    if found:
                        _log_context_probe_recovery(canonical)
                        return dict(found)

            # Compatibility path for gateways that expose limits on ordinary
            # OpenAI/Anthropic model endpoints instead of the Codex catalog.
            anth_headers = {**headers, "anthropic-version": "2023-06-01"}
            for url, req_headers in (
                (f"{base}/v1/models/{slug}", anth_headers),
                (f"{base}/v1/models", headers),
                (f"{base}/v1/models", anth_headers),
            ):
                rr = await hc.get(url, headers=req_headers)
                if rr.status_code >= 400:
                    continue
                body = rr.json()
                candidates: list[Any] = [body]
                if isinstance(body, dict):
                    for item in body.get("data") or []:
                        if (isinstance(item, dict)
                                and item.get("id") in {canonical, slug}):
                            candidates.insert(0, item)
                for item in candidates:
                    capability = _capability_from_model_item(
                        item, source="gateway_models_api")
                    if capability:
                        _CONTEXT_CAPABILITY_CACHE[cache_key] = (
                            now, dict(capability))
                        _log_context_probe_recovery(canonical)
                        return capability
    except Exception as e:
        _log_context_probe_failure(canonical, e)
    _CONTEXT_CAPABILITY_CACHE[cache_key] = (now, None)
    return None


def _cached_gateway_context_capability(model: str) -> dict | None:
    """Return one fresh catalog cache entry without starting network I/O."""
    if not _is_codex_gateway_model(model):
        return None
    canonical = _canonical_context_model(model)
    env = endpoints.env_override(canonical) or {}
    base = (env.get("ANTHROPIC_BASE_URL") or "").rstrip("/")
    if not base:
        return None
    cached = _CONTEXT_CAPABILITY_CACHE.get((base, canonical))
    if cached is None:
        return None
    cached_at, capability = cached
    ttl = (_CONTEXT_CAPABILITY_CACHE_TTL if capability
           else _CONTEXT_CAPABILITY_FAILURE_TTL)
    if time.monotonic() - cached_at >= ttl:
        return None
    return dict(capability) if capability else None


async def _detect_gateway_context_capabilities(
    models: Iterable[str],
) -> dict[str, dict]:
    """Resolve a model list with at most one catalog probe per Gateway base.

    The primary CLIProxyAPI endpoint returns the complete Codex catalog and
    `_detect_gateway_context_capability()` fills every model cache entry from
    that one response. Provider discovery must therefore never repeat the same
    timeout once per dropdown item when the local Gateway is unavailable.
    """
    capabilities: dict[str, dict] = {}
    unresolved_by_base: dict[str, list[str]] = {}
    for model in dict.fromkeys(models):
        if not _is_codex_gateway_model(model):
            continue
        cached = _cached_gateway_context_capability(model)
        if cached is not None:
            capabilities[model] = cached
            continue
        canonical = _canonical_context_model(model)
        env = endpoints.env_override(canonical) or {}
        base = (env.get("ANTHROPIC_BASE_URL") or "").rstrip("/")
        if base:
            unresolved_by_base.setdefault(base, []).append(model)

    async def probe(group: list[str]) -> tuple[str, dict | None]:
        first = group[0]
        return first, await _detect_gateway_context_capability(first)

    if unresolved_by_base:
        probed = await asyncio.gather(*(
            probe(group) for group in unresolved_by_base.values()
        ))
        for first, capability in probed:
            if capability is not None:
                capabilities[first] = capability
        # A successful catalog probe populated sibling entries. A failed one
        # deliberately leaves every sibling on the static control fallback.
        for group in unresolved_by_base.values():
            for model in group:
                cached = _cached_gateway_context_capability(model)
                if cached is not None:
                    capabilities[model] = cached
    return capabilities


def _context_limit_details(
    model: str,
    *,
    sdk_max: int = 0,
    sdk_raw: int = 0,
    stored: int = 0,
    detected: int = 0,
    capability: dict | None = None,
) -> dict:
    """Resolve denominator plus provenance for the meter and preflight.

    Official Claude path: SDK maxTokens is authoritative. Third-party gateways:
    explicit env override wins. Codex Gateway then trusts CLIProxyAPI's active
    `context_window` catalog entry and deliberately ignores Claude CLI's legacy
    200K model guess. `max_context_window` remains metadata, never the default.
    """
    model = _canonical_context_model(model)
    override = _context_limit_env_override(model)
    if override:
        cap = capability or {}
        return {
            "context_limit": override,
            # An override is already expressed in effective tokens. Do not
            # attach the catalog's unrelated raw percentage to that value.
            "context_raw_limit": override,
            "context_max_limit": max(
                override, _positive_int(cap.get("context_max_limit"))),
            "context_effective_percent": 100,
            "catalog_auto_compact_threshold": 0,
            "context_limit_source": "env_override",
            "context_limit_is_estimate": False,
        }
    if not endpoints.is_third_party(model):
        limit = (_positive_int(sdk_max) or _positive_int(stored)
                 or MODEL_CONTEXT_LIMITS.get(model, DEFAULT_CONTEXT_LIMIT))
        raw = _positive_int(sdk_raw) or limit
        source = ("sdk" if _positive_int(sdk_max) else
                  "session_sdk" if _positive_int(stored) else "model_fallback")
        return {
            "context_limit": limit,
            "context_raw_limit": raw,
            "context_max_limit": raw,
            "context_effective_percent": (
                max(1, min(100, limit * 100 // raw)) if raw else 100),
            "catalog_auto_compact_threshold": 0,
            "context_limit_source": source,
            "context_limit_is_estimate": source == "model_fallback",
        }
    if _is_codex_gateway_model(model):
        if capability and _positive_int(capability.get("context_limit")):
            return dict(capability)
        if _positive_int(detected):
            return _capability_from_model_item(
                {"max_input_tokens": detected}, source="gateway_models_api") or {}
        raw = MODEL_CONTEXT_LIMITS.get(model, DEFAULT_CONTEXT_LIMIT)
        pct = CODEX_DEFAULT_EFFECTIVE_CONTEXT_PERCENT
        return {
            "context_limit": max(1, raw * pct // 100),
            "context_raw_limit": raw,
            "context_max_limit": raw,
            "context_effective_percent": pct,
            "catalog_auto_compact_threshold": 0,
            "context_limit_source": "model_fallback",
            "context_limit_is_estimate": True,
        }
    hardcoded = MODEL_CONTEXT_LIMITS.get(model, DEFAULT_CONTEXT_LIMIT)
    limit = _positive_int(sdk_max) or max(_positive_int(stored), hardcoded)
    source = "sdk" if _positive_int(sdk_max) else "model_fallback"
    return {
        "context_limit": limit,
        "context_raw_limit": _positive_int(sdk_raw) or limit,
        "context_max_limit": _positive_int(sdk_raw) or limit,
        "context_effective_percent": 100,
        "catalog_auto_compact_threshold": 0,
        "context_limit_source": source,
        "context_limit_is_estimate": source == "model_fallback",
    }


def _effective_context_limit(
    model: str,
    *,
    sdk_max: int = 0,
    sdk_raw: int = 0,
    stored: int = 0,
    detected: int = 0,
    capability: dict | None = None,
) -> int:
    """Backward-compatible integer view of `_context_limit_details()`."""
    return _positive_int(_context_limit_details(
        model, sdk_max=sdk_max, sdk_raw=sdk_raw, stored=stored,
        detected=detected, capability=capability).get("context_limit"))


def _apply_context_limit_details(target: dict, details: dict) -> None:
    for key in (
        "context_limit", "context_raw_limit", "context_max_limit",
        "context_effective_percent", "context_limit_source",
        "context_limit_is_estimate", "catalog_auto_compact_threshold",
    ):
        if key in details:
            target[key] = details[key]
    target["context_is_estimate"] = bool(
        target.get("context_used_is_estimate", False)
        or target.get("context_limit_is_estimate", False))


def _mark_context_used(target: dict, source: str, *, estimate: bool) -> None:
    target["context_used_source"] = source
    target["context_used_is_estimate"] = bool(estimate)
    target["context_is_estimate"] = bool(
        estimate or target.get("context_limit_is_estimate", False))


def _compact_threshold(
    model: str,
    limit: int,
    sdk_threshold: int = 0,
    *,
    sdk_max: int = 0,
    capability: dict | None = None,
) -> int:
    if limit <= 0:
        return 0
    sdk_t = _positive_int(sdk_threshold)
    if _is_codex_gateway_model(model):
        catalog_t = _positive_int(
            (capability or {}).get("catalog_auto_compact_threshold"))
        if catalog_t:
            return min(catalog_t, limit)
        # A client created with our injected CLAUDE_CODE_MAX_CONTEXT_TOKENS
        # reports sdk_max == catalog effective limit. Trust its exact buffer.
        # Older live clients still report 200K/167K; ignore that stale pair.
        if sdk_t and _positive_int(sdk_max) == limit:
            return min(sdk_t, limit)
        return int(limit * 0.90)
    soft = int(limit * 0.90)
    if sdk_t:
        return min(sdk_t, soft)
    return soft


def _rough_prompt_tokens(text: str) -> int:
    # Conservative language-agnostic estimate for preflight only. The SDK/tokenizer
    # truth arrives after the turn; here we just avoid sending when already close.
    if not text:
        return 0
    return max(1, len(text) // 3)


_CLI_INTERRUPT_MESSAGE_RE = re.compile(
    r"^\[Request interrupted by user(?: for tool use)?\]$",
    re.IGNORECASE,
)


def _is_cli_interrupt_message(text: Any) -> bool:
    """Return True only for Claude CLI's synthetic interrupt transcript row."""
    return isinstance(text, str) and bool(
        _CLI_INTERRUPT_MESSAGE_RE.fullmatch(text.strip()))


async def _detect_gateway_context_limit(model: str) -> int:
    """Compatibility wrapper returning only the effective integer limit."""
    capability = await _detect_gateway_context_capability(model)
    return _positive_int((capability or {}).get("context_limit"))

# Soft budget. If set (via MUSELAB_BUDGET_USD env or PUT /api/settings),
# usage endpoint flags overrun so the UI can color the cost badge red.
def _is_real_user_prompt(sm: Any) -> bool:
    """True if ``sm`` is a user message the human actually typed.

    SDK 0.2.82's get_session_messages doesn't really filter tool-use
    sidechain frames — every wrapped tool_result still comes back as
    ``type="user"`` with ``parent_tool_use_id=None``, contrary to the
    docstring. So we discriminate by content shape: real user prompts
    contain text (string content, or a list with at least one non-
    tool_result block); pure-tool_result frames are sidechain echoes
    and don't count as a turn.

    Without this filter a session with 45 prompts + heavy agent tool
    use shows up as 300+ turns in the picker.
    """
    if sm is None or getattr(sm, "type", None) != "user":
        return False
    if getattr(sm, "parent_tool_use_id", None):
        return False
    msg = getattr(sm, "message", None)
    content = msg.get("content") if isinstance(msg, dict) else None
    if isinstance(content, str):
        return bool(content.strip()) and not _is_cli_interrupt_message(content)
    if isinstance(content, list):
        # If any meaningful block is non-tool_result (text / image / etc.) →
        # real prompt. Claude CLI persists Stop as a synthetic user text row;
        # it is transport metadata, not a human turn.
        for b in content:
            if not isinstance(b, dict) or b.get("type") == "tool_result":
                continue
            if (b.get("type") == "text"
                    and _is_cli_interrupt_message(b.get("text"))):
                continue
            return True
        return False
    # Unknown shape — default to "real" so we don't under-count.
    return True


def _budget_usd() -> float:
    return env_float("MUSELAB_BUDGET_USD", 0.0)


# Per-runtime-key creation locks are shared with the extracted pool owner.
_creation_locks = chat_runtime.CREATION_LOCKS


def _creation_lock_for(key: _ClientKey) -> asyncio.Lock:
    return chat_runtime.creation_lock_for(key)


def _normalize_plan_return_permission(
    permission: str,
    plan_return_permission: str | None,
) -> str:
    """Return the fail-closed Plan Mode return target for a runtime request."""
    if permission != "plan":
        return ""
    target = (plan_return_permission or "").strip()
    if target in _VALID_PERMISSION_MODES and target != "plan":
        return target
    return "default"


def _build_plan_enter_hooks(
    session_id: str,
    launch_permission: str,
) -> tuple[Any, Any]:
    """Persist a native EnterPlanMode transition after the CLI confirms it."""

    return_permission = (
        launch_permission
        if launch_permission in _VALID_PERMISSION_MODES
        and launch_permission != "plan"
        else "default"
    )

    def _stop_ambiguous_transition(reason: str) -> dict[str, Any]:
        return {"continue_": False, "stopReason": reason}

    async def _post_tool_use(input_data, tool_use_id, _context):
        data = input_data if isinstance(input_data, dict) else {}
        tid = str(data.get("tool_use_id") or tool_use_id or "")
        # EnterPlanMode has already changed the live CLI. Always rebuild before
        # another turn so the process launch contract matches durable metadata.
        _pending_runtime_rebuilds.add(session_id)
        try:
            committed = await obs.to_thread_io(
                "chat.plan_enter",
                session_id,
                sess.commit_plan_enter,
                session_id,
                expected_permission=launch_permission,
                plan_return_permission=return_permission,
                owned=True,
            )
        except Exception as exc:
            sys.stderr.write(
                f"[plan-mode] enter persist failed sid={session_id[:8]} "
                f"exc={type(exc).__name__}\n")
            committed = False
        if not committed:
            current = await obs.to_thread_io(
                "chat.session_read", session_id, sess.get_session, session_id)
            current = current or {}
            await perm.emit_session_event(
                session_id,
                "permission_mode_change_failed",
                {
                    "permission": current.get("permission") or launch_permission,
                    "requested_permission": "plan",
                    "source": "enter_plan",
                    "tool_use_id": tid,
                    "message": (
                        "Plan mode changed in another client; "
                        "the stale transition was not applied."
                    ),
                },
            )
            return _stop_ambiguous_transition(
                "A newer permission change superseded EnterPlanMode.")
        await perm.emit_session_event(
            session_id,
            "permission_mode_changed",
            {
                "permission": "plan",
                "previous_permission": launch_permission,
                "source": "enter_plan",
                "tool_use_id": tid,
            },
        )
        return {}

    async def _post_tool_use_failure(input_data, tool_use_id, _context):
        data = input_data if isinstance(input_data, dict) else {}
        tid = str(data.get("tool_use_id") or tool_use_id or "")
        _pending_runtime_rebuilds.add(session_id)
        current = await obs.to_thread_io(
            "chat.session_read", session_id, sess.get_session, session_id)
        current = current or {}
        await perm.emit_session_event(
            session_id,
            "permission_mode_change_failed",
            {
                "permission": current.get("permission") or launch_permission,
                "requested_permission": "plan",
                "source": "enter_plan",
                "tool_use_id": tid,
                "message": str(data.get("error") or ""),
            },
        )
        return _stop_ambiguous_transition(
            "EnterPlanMode failed; the runtime will be rebuilt safely.")

    return _post_tool_use, _post_tool_use_failure


def _build_plan_exit_hooks(
    session_id: str,
    plan_return_permission: str = "default",
) -> tuple[Any, Any]:
    """Build the success/failure hooks that commit an ExitPlanMode transition.

    `can_use_tool` only records the user's selected SDK suggestion. The durable
    session mode changes here, after the CLI confirms ExitPlanMode itself
    completed. This keeps a failed/cancelled tool call from leaving MuseLab's
    metadata ahead of the real runtime.
    """

    def _stop_ambiguous_transition(reason: str) -> dict[str, Any]:
        # A failed/stale ExitPlanMode may already have changed the live CLI's
        # permission. Do not let that ambiguous process continue the same turn
        # and potentially write while durable metadata still says Plan.
        return {
            "continue_": False,
            "stopReason": reason,
        }

    async def _post_tool_use(input_data, tool_use_id, _context):
        data = input_data if isinstance(input_data, dict) else {}
        tid = str(data.get("tool_use_id") or tool_use_id or "")
        transition = perm.consume_plan_transition(session_id, tid)
        target = getattr(transition, "mode", None)
        externally_resolved = not bool(target)
        if externally_resolved:
            # User/project hooks loaded through setting_sources may approve
            # ExitPlanMode before can_use_tool runs. PostToolUse still fires,
            # so use the SDK-reported current mode when it is concrete.
            candidate = data.get("permission_mode")
            target = (
                candidate
                if isinstance(candidate, str)
                and candidate in _VALID_PERMISSION_MODES
                and candidate != "plan"
                else None
            )

        # ExitPlanMode can mutate the live CLI even when MuseLab's callback was
        # bypassed by an external PermissionRequest/PreToolUse hook. Always
        # discard this runtime before metadata I/O or target inference.
        _pending_runtime_rebuilds.add(session_id)
        if not target:
            current = await obs.to_thread_io(
                "chat.session_read", session_id, sess.get_session, session_id)
            await perm.emit_session_event(
                session_id,
                "permission_mode_change_failed",
                {
                    "permission": (current or {}).get("permission") or "plan",
                    "source": "external_hook",
                    "tool_use_id": tid,
                    "message": "ExitPlanMode did not report a non-Plan mode.",
                },
            )
            return _stop_ambiguous_transition(
                "ExitPlanMode completed without a verifiable target mode.")
        try:
            previous = "plan"
            committed = await obs.to_thread_io(
                "chat.plan_exit",
                session_id,
                sess.commit_plan_exit,
                session_id,
                target,
                expected_plan_return=plan_return_permission,
                owned=True,
            )
            if not committed:
                current = await obs.to_thread_io(
                    "chat.session_read",
                    session_id,
                    sess.get_session,
                    session_id,
                )
                current = current or {}
                await perm.emit_session_event(
                    session_id,
                    "permission_mode_change_failed",
                    {
                        "permission": current.get("permission") or "plan",
                        "requested_permission": target,
                        "source": (
                            "external_hook"
                            if externally_resolved else "exit_plan"
                        ),
                        "tool_use_id": tid,
                        "message": (
                            "Plan mode changed in another client; "
                            "the stale approval was not applied."
                        ),
                    },
                )
                return _stop_ambiguous_transition(
                    "A newer permission change superseded this plan approval.")
        except Exception as exc:
            sys.stderr.write(
                f"[plan-mode] persist failed sid={session_id[:8]} "
                f"target={target} exc={type(exc).__name__}\n")
            await perm.emit_session_event(
                session_id,
                "permission_mode_change_failed",
                {
                    "permission": "plan",
                    "requested_permission": target,
                    "source": (
                        "external_hook" if externally_resolved else "exit_plan"
                    ),
                    "tool_use_id": tid,
                    "message": "Could not persist the Plan mode transition.",
                },
            )
            return _stop_ambiguous_transition(
                "MuseLab could not persist the Plan mode transition.")

        # The current process may continue this turn under the SDK's new mode,
        # but the next turn must use a freshly-launched client whose permission
        # contract matches persisted state (especially for bypass capability).
        await perm.emit_session_event(
            session_id,
            "permission_mode_changed",
            {
                "permission": target,
                "previous_permission": previous,
                "source": (
                    "external_hook" if externally_resolved else "exit_plan"
                ),
                "tool_use_id": tid,
            },
        )
        return {}

    async def _post_tool_use_failure(input_data, tool_use_id, _context):
        data = input_data if isinstance(input_data, dict) else {}
        tid = str(data.get("tool_use_id") or tool_use_id or "")
        transition = perm.discard_plan_transition(session_id, tid)
        # External hooks can apply a mode before a later ExitPlanMode failure,
        # too. Its live state is ambiguous even without a staged transition.
        _pending_runtime_rebuilds.add(session_id)
        current = await obs.to_thread_io(
            "chat.session_read", session_id, sess.get_session, session_id)
        current = current or {}
        await perm.emit_session_event(
            session_id,
            "permission_mode_change_failed",
            {
                "permission": current.get("permission") or "plan",
                "source": "exit_plan" if transition else "external_hook",
                "tool_use_id": tid,
                "message": str(data.get("error") or ""),
            },
        )
        return _stop_ambiguous_transition(
            "ExitPlanMode failed; the runtime will be rebuilt safely.")

    return _post_tool_use, _post_tool_use_failure


async def _build_and_connect_client(
    session_id: str, model: str, permission: str, effort: str,
    service_tier: str = "",
    plan_return_permission: str = "",
) -> ClaudeSDKClient:
    """The slow path: build ClaudeAgentOptions, instantiate ClaudeSDKClient,
    call .connect() with retry. NEVER holds _lock — multi-second CLI
    subprocess spawn must not block sibling requests. Caller is responsible
    for serialising concurrent misses on the same key via _creation_lock_for().
    """
    def _load_session_runtime() -> tuple[dict, Path]:
        return (
            sess.get_session(session_id) or {},
            sess.session_workspace(session_id),
        )

    sess_data, workspace_root = await obs.to_thread_io(
        "chat.client_session_read",
        session_id,
        _load_session_runtime,
    )
    side_question_runtime = (
        sess_data.get("runtime_profile") == "side_question"
    )
    retry_source_session_id = str(
        sess_data.get("retry_source_session_id") or "")
    retry_target_user_uuid = str(
        sess_data.get("retry_target_user_uuid") or "")
    retry_resume_session_at = str(
        sess_data.get("retry_resume_session_at") or "")
    native_retry_resume = bool(retry_source_session_id)
    effort = _normalize_effort(effort)
    service_tier = (service_tier or "").strip()
    if effort not in _VALID_EFFORT:
        raise ValueError(f"invalid effort: {effort}")
    if service_tier not in _VALID_SERVICE_TIERS:
        raise ValueError(f"invalid service tier: {service_tier}")
    plan_return_permission = _normalize_plan_return_permission(
        permission, plan_return_permission)
    runtime_plan_return_permission = (
        plan_return_permission if permission == "plan" else permission
    )
    if (runtime_plan_return_permission not in _VALID_PERMISSION_MODES
            or runtime_plan_return_permission == "plan"):
        runtime_plan_return_permission = "default"
    is_ducc = endpoints.is_ducc_model(model)
    # New CLI rule: session_id + resume/continue conflict unless fork_session
    # is set. So we use resume alone — it both loads existing state AND
    # implies the session id. Falls back to session_id-only for new sessions.
    # SDK default max_buffer_size is 1 MB. A single tool_use JSON message
    # (Edit on a large file, or Read of a long file) can blow past that
    # and kill the message reader silently — the chat then "hangs forever"
    # because no more chunks arrive. Bump to 32 MB; configurable via env.
    max_buf = env_int("MUSELAB_MAX_BUFFER_SIZE", 32 * 1024 * 1024, min_value=1024)
    # Critical SDK option distinction:
    #   `session_id=X`  → force a NEW session to use UUID X (fails if
    #                     CLI already has a JSONL for X)
    #   `resume=X`      → resume an EXISTING session by UUID X
    # If we always use `resume` for un-streamed sessions, CLI generates
    # a fresh UUID and orphans ours. If we always use `session_id`,
    # any session that's ever streamed errors with "already in use".
    # Detect JSONL existence by RECURSIVELY scanning the CLI's projects
    # root — SDK's _find_project_dir relies on path-hash matching that
    # has been unreliable in some setups (user's CLI saw the JSONL but
    # the SDK helper didn't). _find_session_jsonl walks BOTH default
    # and vendor roots so vendor sessions don't look "new" here — passing
    # `session_id=` for an existing JSONL makes the CLI exit with
    # "Session ID already in use", and the fallback at the bottom of
    # this function doesn't catch it (the CLI dies inside the SDK's
    # background message reader, not during `client.connect()`).
    jsonl_exists = False
    try:
        jsonl_exists = await obs.to_thread_io(
            "chat.client_transcript_probe",
            session_id,
            lambda: _find_session_jsonl(session_id) is not None,
        )
    except Exception as e:
        sys.stderr.write(
            f"[muselab] jsonl_exists check failed "
            f"sid={obs.short_id(session_id)} exc={type(e).__name__}\n"
        )
    if native_retry_resume and jsonl_exists:
        # query() may have committed the child transcript immediately before a
        # process crash, leaving only the metadata cleanup unfinished.  The
        # child JSONL is canonical proof that the truncating resume already
        # materialized; resume the child normally and consume the stale intent
        # instead of trying to fork into an existing session id again.
        await obs.to_thread_io(
            "chat.retry_intent_reconcile",
            session_id,
            sess.clear_retry_intent,
            session_id,
            source_session_id=retry_source_session_id,
            target_user_uuid=retry_target_user_uuid,
            owned=True,
        )
        _native_retry_commits.pop(session_id, None)
        native_retry_resume = False

    if native_retry_resume:
        retry_values = [
            retry_source_session_id,
            retry_target_user_uuid,
            *([retry_resume_session_at] if retry_resume_session_at else []),
        ]
        try:
            if any(str(uuid.UUID(value)) != value.lower()
                   for value in retry_values):
                raise ValueError
        except (ValueError, AttributeError, TypeError):
            raise ValueError("invalid durable SDK retry boundary") from None
        session_binding = {
            "resume": retry_source_session_id,
            "session_id": session_id,
            "fork_session": True,
            "resume_drops_turn": retry_target_user_uuid,
            **({"resume_session_at": retry_resume_session_at}
               if retry_resume_session_at else {}),
        }
    else:
        session_binding = (
            {"resume": session_id} if jsonl_exists
            else {"session_id": session_id}
        )
    # Keep one privacy-safe line per stderr category and connected client.
    # Raw CLI output can contain prompts, paths, credentials and protocol bodies.
    _cli_stderr = _privacy_safe_cli_stderr_logger(
        "DUCC" if is_ducc else "SDK-CLI", session_id)

    post_enter_hook, post_enter_failure_hook = _build_plan_enter_hooks(
        session_id, permission)
    post_exit_hook, post_exit_failure_hook = _build_plan_exit_hooks(
        session_id, runtime_plan_return_permission)
    skills_off = os.environ.get("MUSELAB_DISABLE_SKILLS", "").lower() in (
        "1", "true", "yes",
    )
    plugin_root = await obs.to_thread_io(
        "chat.client_plugin_path",
        session_id,
        lambda: Path(__file__).resolve().parent.parent,
    )
    user_prompt_hooks = (
        [] if side_question_runtime else [mem0.build_recall_hook(session_id)]
    )
    if (not side_question_runtime
            and sess_data.get("runtime_predecessor")):
        user_prompt_hooks.append(_build_runtime_task_context_hook(session_id))
    opts_kwargs = dict(
        cwd=str(workspace_root),
        model=(endpoints.ducc_cli_model(model) if is_ducc
               else endpoints.normalize_model_id(model)),
        permission_mode=permission,
        max_buffer_size=max_buf,
        stderr=_cli_stderr,
        # Block harness-only tools the SDK exposes by default. AskUserQuestion
        # is intentionally kept: SDK 0.2.149 / CLI 2.1.252 correctly turns the
        # browser-injected answers into a native tool_result in every supported
        # permission mode; permission_request.py owns that UI bridge.
        #
        # MAINTENANCE NOTE (audit E/253, updated 2026-07-16): this is a
        # hand-maintained DENYLIST — a future harness-only tool is silently
        # EXPOSED until added here. Drift is now mechanically checkable: the
        # CLI announces its tool catalog in the init SystemMessage;
        #   .venv/bin/python scripts/dump-tool-catalog.py \
        #       | diff docs/tool-catalog.txt -
        # on every SDK bump. Alternatives were
        # evaluated and rejected: tools={"type":"preset","preset":"claude_code"}
        # maps to `--tools default` — identical to not passing tools at all, so
        # it adds no protection; an explicit allowlist inverts the failure mode
        # (new/renamed useful tools silently MISSING after a CLI bump).
        #
        # SDK 0.2.149 / bundled CLI 2.1.252 was probed on the production WSL
        # route: CronCreate → CronList → CronDelete → CronList completed and
        # cleaned up successfully; Bash(run_in_background) → Monitor also
        # completed.  Those SDK-owned tools intentionally remain exposed and
        # render through MuseLab's generic tool/lifecycle cards.
        disallowed_tools=[
            "ScheduleWakeup",         # /loop dynamic mode — Claude Code only
            "EnterWorktree", "ExitWorktree",
            "PushNotification",
            # Claude CLI host features whose protocol is owned by a
            # Claude Code / claude.ai host. muselab has no matching design,
            # review-findings, remote-trigger, or teammate-message surface.
            "DesignSync", "RemoteTrigger", "ReportFindings",
            "SendMessage",
        ],
        # Load CLAUDE.md from user (~/.claude/CLAUDE.md), project
        # (cwd/CLAUDE.md → the user's archive), and local (.claude/
        # within cwd). Also enables skill discovery from the same scopes.
        #
        # ARCHIVE-ISOLATION NOTE (audit E/255): opening "user" scope means the
        # model CAN read ~/.claude/ global config (CLAUDE.md, memory, skills)
        # — files that live OUTSIDE the archive root. This is intentional (the
        # platform's own config is meant to be loaded) and is NOT relaxed here.
        # This intentionally exposes the SDK's own instruction and Skill
        # hierarchy outside the active workspace. Workspace file operations
        # remain governed separately by tool permissions and service-user
        # access; setting_sources should not be narrowed to emulate a custom
        # prompt boundary.
        setting_sources=["user", "project", "local"],
        # The SDK cwd is the user's active workspace, not this repository.
        # Load muselab itself as a local SDK plugin so the skills/ extension
        # slot remains discoverable in every registered workspace.
        plugins=[{
            "type": "local",
            "path": str(plugin_root),
        }],
        # Bind THIS session to muselab's chosen UUID — either as a new
        # session (session_id=) or by resuming the existing one (resume=).
        **session_binding,
        # Token-level streaming: SDK emits StreamEvent for each delta
        # the model produces (text / thinking). Without this, we only
        # see full blocks at the end → user waits for the whole reply
        # before seeing anything. With this, each token shows up.
        include_partial_messages=True,
        # Receive CLI-owned Hook lifecycle messages. MuseLab persists only a
        # privacy-bounded timing/status projection; raw hook output is dropped.
        include_hook_events=True,
        # Ask the SDK to forward complete Subagent sidechains.  They are
        # separated from the parent transcript below using the SDK-owned
        # parent_tool_use_id and exposed as a nested timeline in the GUI.
        forward_subagent_text=True,
        # Recall runs in the SDK's dedicated UserPromptSubmit additional-context
        # channel. Never prepend it to client.query(prompt): that would persist
        # the memory block as if the user typed it, polluting JSONL history,
        # titles, transcript search, exports, and every later resume.
        hooks={
            "UserPromptSubmit": [HookMatcher(
                hooks=user_prompt_hooks,
                timeout=mem0.RECALL_HOOK_TIMEOUT,
            )],
            **({
                "PreToolUse": [HookMatcher(
                    matcher="Skill",
                    hooks=[_build_codex_skill_guard_hook()],
                )],
            } if (
                _is_codex_gateway_model(model)
                and os.environ.get(
                    "MUSELAB_ALLOW_LARGE_CODEX_CLAUDE_API_SKILL", ""
                ).strip().lower() not in {"1", "true", "yes", "on"}
            ) else {}),
            "PostToolUse": [
                HookMatcher(
                    matcher="EnterPlanMode",
                    hooks=[post_enter_hook],
                ),
                HookMatcher(
                    matcher="ExitPlanMode",
                    hooks=[post_exit_hook],
                ),
            ],
            "PostToolUseFailure": [
                HookMatcher(
                    matcher="EnterPlanMode",
                    hooks=[post_enter_failure_hook],
                ),
                HookMatcher(
                    matcher="ExitPlanMode",
                    hooks=[post_exit_failure_hook],
                ),
            ],
        },
    )
    if not side_question_runtime:
        # PreToolUse observes AskUserQuestion regardless of allow rules or a
        # mid-turn permission-mode transition, unlike can_use_tool. It therefore
        # owns the browser round-trip in every mode. The timeout must cover the
        # full human-response window rather than the SDK's short hook default.
        opts_kwargs["hooks"].setdefault("PreToolUse", []).append(HookMatcher(
            matcher="AskUserQuestion",
            hooks=[perm.build_ask_user_question_hook_for_session(session_id)],
            timeout=ANSWER_TIMEOUT_S + 5,
        ))
        # can_use_tool is installed below for every ordinary runtime, including
        # bypass. The SDK configures its stdio control route from that callback;
        # keeping the callback present is required if native EnterPlanMode changes
        # a bypass process into a mode that can prompt before ExitPlanMode.
    if side_question_runtime:
        # Side questions are deliberately narrower than ordinary workspace
        # agents.  `tools` removes every built-in except public web lookup;
        # the explicit allow rules make those two tools usable from the compact
        # floating window without an approval card it cannot render.  Remove
        # workspace/user instruction sources, plugins and recall hooks as well:
        # a web query must not gain unrelated private context behind the user's
        # back.  The selected excerpt + forked transcript remain available.
        opts_kwargs["tools"] = ["WebSearch", "WebFetch"]
        opts_kwargs["allowed_tools"] = ["WebSearch", "WebFetch"]
        opts_kwargs["setting_sources"] = []
        opts_kwargs["plugins"] = []
        opts_kwargs["hooks"].pop("UserPromptSubmit", None)
    if is_ducc:
        # Keep the SDK's mature stream-json/control/MCP/session machinery, but
        # swap the executable to MuseLab's sanitising wrapper. The wrapper then
        # execs the real DUCC launcher, so claude-go — not MuseLab's static env —
        # owns authentication, endpoint selection and comate_custom_header.
        def _ducc_launch_paths() -> tuple[str | None, Path, bool]:
            executable = locate_ducc_executable()
            path = Path(ducc_cli_wrapper())
            ready = path.is_file() and os.access(path, os.X_OK)
            return executable, path, ready

        ducc_executable, wrapper, wrapper_ready = await obs.to_thread_io(
            "chat.ducc_launch_probe",
            session_id,
            _ducc_launch_paths,
        )
        if ducc_executable is None:
            raise ClaudeSDKError(
                "DUCC runtime is unavailable: install/login to DUCC or set "
                "MUSELAB_DUCC_CLI to its executable path."
            )
        if not wrapper_ready:
            raise ClaudeSDKError(
                f"MuseLab DUCC launcher is missing or not executable: {wrapper}"
            )
        opts_kwargs["cli_path"] = str(wrapper)
        opts_kwargs["env"] = _ducc_subprocess_env(ducc_executable)
    if permission == "plan" and plan_return_permission == "bypassPermissions":
        # The CLI refuses a later setMode(bypassPermissions) unless this
        # capability was granted at process launch. This flag permits the
        # transition without starting the client itself in bypass mode.
        opts_kwargs["extra_args"] = {
            "allow-dangerously-skip-permissions": None,
        }
    # Let the SDK expose every discovered Skill for every provider, including
    # Anthropic-compatible third-party gateways. Passing [] for disabled or
    # privacy-isolated runtimes is deliberate: omission can let SDK defaults
    # re-enable discovery.
    opts_kwargs["skills"] = (
        [] if skills_off or side_question_runtime else "all"
    )
    # Optional model params from env (UI-editable via /api/settings).
    mt = env_int("MUSELAB_MAX_TURNS", 0, min_value=0)
    if mt > 0:
        opts_kwargs["max_turns"] = mt
    # For non-Claude models, point the SDK at the vendor's own
    # Anthropic-compatible endpoint (DeepSeek / GLM / MiniMax).
    # This way the SDK's full agent loop (tools, MCP, skills, CLAUDE.md)
    # works uniformly across providers — no router process needed.
    # Claude models still go direct so Pro OAuth keeps working.
    # DUCC is a CLI runtime, not an Anthropic-compatible endpoint override.
    # Never inject MuseLab's provider URL/key/static custom header into it.
    env_ovr = None if is_ducc else endpoints.env_override(model)
    if env_ovr is not None:
        env_ovr = dict(env_ovr)
        # Isolated vendor config prevents OAuth fallback, but starts with no
        # project trust state. Without this marker the CLI silently ignores
        # permissions.allow rules from the workspace settings.
        await obs.to_thread_io(
            "chat.vendor_workspace_trust",
            session_id,
            endpoints.ensure_vendor_workspace_trusted,
            workspace_root,
            owned=True,
        )
        # Agent/Task's `model` field is intentionally an alias enum
        # (sonnet|opus|haiku|fable), not an arbitrary model ID. Without these
        # CLI-native alias overrides, a subagent launched inside a Codex or
        # other third-party session resolves `opus` to `claude-opus-*` while
        # still inheriting the vendor ANTHROPIC_BASE_URL → deterministic 502
        # "unknown provider". Pin every tier to the parent provider's actual
        # model so built-in, custom, foreground, and background agents all stay
        # on the same route. The top-level --model is explicit and unaffected.
        routed_model = endpoints.normalize_model_id(model)
        for tier in ("OPUS", "SONNET", "HAIKU", "FABLE"):
            env_ovr[f"ANTHROPIC_DEFAULT_{tier}_MODEL"] = routed_model
        if _is_codex_gateway_model(model):
            # Claude Agent SDK has no public Ultra or service-tier field, and
            # CLIProxyAPI's Claude translator historically collapses disabled
            # thinking to medium. Claude CLI *does* support custom headers, so
            # carry the canonical MuseLab controls out-of-band and let Gateway
            # apply them after translation. Ultra maps to the provider's wire-
            # level `max`; MuseLab separately enforces bounded subagent depth
            # and concurrency for Ultra below.
            # Send `auto` too: Gateway removes
            # the translator's synthetic medium so the model catalog default
            # (Sol=low, others may differ) remains authoritative.
            env_ovr["ANTHROPIC_CUSTOM_HEADERS"] = (
                _muselab_gateway_headers(effort, service_tier)
            )
            if effort == "ultra":
                # Enforce the runtime boundary: no nested fan-out and at most
                # four concurrent workers. Explicit operator overrides are
                # preserved; invalid/zero values fall back to a safe positive
                # integer through env_int().
                env_ovr["CLAUDE_CODE_MAX_SUBAGENT_SPAWN_DEPTH"] = str(
                    env_int("CLAUDE_CODE_MAX_SUBAGENT_SPAWN_DEPTH", 1,
                            min_value=1))
                env_ovr["CLAUDE_CODE_MAX_CONCURRENT_SUBAGENTS"] = str(
                    env_int("CLAUDE_CODE_MAX_CONCURRENT_SUBAGENTS", 4,
                            min_value=1))
            # Claude CLI cannot infer Codex/GPT windows from its native model
            # table and otherwise hard-codes 200K (auto-compact around 167K).
            # Feed it the same effective window used by muselab's meter so its
            # own tokenizer, /context output, and native autocompaction agree
            # with CLIProxyAPI's live model catalog. This is a local CLI knob;
            # it does not alter or over-claim the gateway's raw model ceiling.
            capability = await _detect_gateway_context_capability(model)
            details = _context_limit_details(model, capability=capability)
            effective_limit = _positive_int(details.get("context_limit"))
            if effective_limit:
                env_ovr = dict(env_ovr)
                env_ovr["CLAUDE_CODE_MAX_CONTEXT_TOKENS"] = str(effective_limit)
                # Since Claude CLI 2.1.159, local/SDK sessions can silently
                # skip turn-internal auto-compaction unless its window source
                # is explicit. Pass the full effective window here: the CLI
                # itself subtracts output/compaction reserves and reports the
                # resulting lower autoCompactThreshold. Pre-multiplying by
                # 0.9 would reserve twice and compact far too early.
                env_ovr["CLAUDE_CODE_AUTO_COMPACT_WINDOW"] = str(effective_limit)
        opts_kwargs["env"] = env_ovr
    else:
        # DUCC owns auth through claude-go and does not use native Claude's
        # credentials file/API-key precheck. For the ordinary Claude route keep
        # the existing early, user-readable auth failure.
        if not is_ducc:
            cred_file = Path.home() / ".claude" / ".credentials.json"
            has_env_auth = bool(
                os.environ.get("ANTHROPIC_API_KEY")
                or os.environ.get("ANTHROPIC_AUTH_TOKEN")
            )
            has_credential_file = has_env_auth or await obs.to_thread_io(
                "chat.credential_probe",
                session_id,
                cred_file.exists,
            )
            if not has_credential_file:
                raise ClaudeSDKError(
                    f"Claude model '{model}' requires auth: either run "
                    f"`claude login` (Pro/Max) or set ANTHROPIC_API_KEY in "
                    f"Settings. CLI would exit 1 silently otherwise."
                )
        # ``opts_kwargs`` already carries the per-client, privacy-safe stderr
        # classifier installed above.  Do not replace it with a raw line sink.

    # A rollover child is a point-in-time transcript fork while the source
    # CLI keeps ownership of tasks launched before that boundary. Claude CLI
    # understands this timestamp natively and excludes pre-boundary launches
    # from its orphan-task scanner. Set an explicit empty value for ordinary
    # sessions too, so a stale ambient variable cannot affect unrelated
    # resumes. Merge instead of replace so provider routing and DUCC's bounded
    # environment remain intact.
    runtime_boundary = sess.normalize_runtime_fork_boundary_at(
        sess_data.get("runtime_fork_boundary_at")
    )
    runtime_env = dict(opts_kwargs.get("env") or {})
    runtime_env[_RUNTIME_RESUME_SOURCE_ALIVE_ENV] = runtime_boundary
    opts_kwargs["env"] = runtime_env

    # MCP servers come from:
    #   - muselab's own mcp.json (UI-managed)
    #   - Claude Code's standard MCP config locations (~/.claude.json,
    #     ~/.claude/settings.json, <archive>/.mcp.json) so any MCP the
    #     user already added via `claude mcp add` "just works" without
    #     re-entering — muselab is positioned as a Claude Code replacement.
    # See backend/api_settings.py _load_mcp_merged for the merge rules.
    mcp_dict: dict = {}
    if not side_question_runtime:
        try:
            from .api_settings import _load_mcp_merged
            merged_mcp = await obs.to_thread_io(
                "chat.mcp_config_read",
                session_id,
                _load_mcp_merged,
            )
            for name, spec in merged_mcp.items():
                if not isinstance(spec, dict):
                    continue
                # Skip disabled servers (UI toggle OR override stub).
                if spec.get("disabled"):
                    continue
                # Strip muselab-local metadata keys before handing to SDK —
                # `_source` / `_overridden_by_muselab` / `disabled` are
                # display/control fields, not part of the MCP spec.
                clean = {k: v for k, v in spec.items()
                         if not k.startswith("_") and k != "disabled"}
                # Defensive: external sources may have entries without
                # `command` (e.g. broken Claude Code config). Skip rather
                # than hand the SDK an unconnectable spec.
                if "command" not in clean and "url" not in clean:
                    continue
                mcp_dict[name] = clean
        except Exception as e:
            # Don't fail client construction because an MCP source had a parse
            # error. _load_mcp_merged already swallows per-file errors and
            # stderr's them; this catch is for unexpected programmer errors.
            sys.stderr.write(
                f"[chat] mcp merge failed sid={session_id[:8]} "
                f"exc={type(e).__name__}; external MCP disabled for this client\n")
            sys.stderr.flush()
    opts_kwargs["mcp_servers"] = mcp_dict
    # Enable extended thinking for models whose provider endpoint handles
    # the standard Anthropic thinking config. Some vendors (e.g. Qianfan)
    # reject thinking because their max_completion_tokens cap (~12k) can't
    # accommodate the thinking budget we normally pass (~4k) alongside the
    # output max_tokens the SDK computes — the total exceeds their limit.
    # For those providers we skip thinking entirely; the model still works
    # but without visible reasoning blocks.
    provider = endpoints.lookup(model)
    # Per-session opt-out (default on). Disabling thinking is the user's escape
    # hatch for the CLI streaming-interleaving 400 ("thinking blocks in the
    # latest assistant message cannot be modified"): with no thinking blocks,
    # the interleaved [thinking, tool_use, thinking, ...] shape that trips the
    # API can't form. Changing it invalidates the cached client (PATCH handler
    # calls disconnect_client) so the next turn rebuilds with this setting.
    thinking_pref = bool(sess_data.get("thinking", True))
    supports_thinking = (
        endpoints.ducc_is_claude_model(model)
        if is_ducc
        else ((provider is None) or provider.supports_thinking)
    ) and thinking_pref
    codex_effort_transport = (
        _is_codex_gateway_model(model) and effort != "auto"
    )
    if codex_effort_transport:
        # CLIProxyAPI only reads Claude's output_config.effort when thinking is
        # adaptive/auto. This is transport plumbing, not a request to render a
        # visible thinking block, hence display=omitted. The private header is
        # still final authority (including Ultra's wire-level max mapping).
        opts_kwargs["thinking"] = ThinkingConfigAdaptive(
            type="adaptive", display="omitted")
    elif supports_thinking:
        # Fixed at 10000 — no UI knob (2026-05-28). Power users can still
        # override via the env var if they really need to.
        budget = env_int("MUSELAB_THINKING_BUDGET", 10000, min_value=0)
        if is_ducc:
            # The current factory DUCC CLI predates --thinking-display. Keep
            # thinking enabled, but omit that newer presentation-only option.
            opts_kwargs["thinking"] = ThinkingConfigEnabled(
                type="enabled", budget_tokens=budget)
        else:
            # display="summarized" is REQUIRED for Opus 4.7+: those models
            # default to display="omitted" (signature-only, no plaintext), so
            # without this the SDK never emits thinking_delta and the FE block
            # is empty.
            opts_kwargs["thinking"] = ThinkingConfigEnabled(
                type="enabled", budget_tokens=budget, display="summarized")
    else:
        opts_kwargs["thinking"] = ThinkingConfigDisabled(type="disabled")
    # SDK-native effort is still used where possible. Ultra is not in SDK
    # 0.2.128 (nor in the current provider wire contract), so its reasoning
    # half is `max`; the native Skill above supplies proactive delegation.
    # `auto` omits the SDK option. Gateway's post-translation header rule remains
    # final authority for every Codex wire-level effort.
    if effort != "auto" and (
        not is_ducc or endpoints.ducc_is_claude_model(model)
    ):
        sdk_effort = "max" if effort == "ultra" else effort
        if sdk_effort in _SDK_EFFORT_LEVELS:
            opts_kwargs["effort"] = sdk_effort
    # can_use_tool resolves ordinary SDK permission prompts. Native
    # AskUserQuestion always uses the PreToolUse browser bridge above because
    # permission rules and live mode transitions can shadow this callback.
    if not side_question_runtime:
        opts_kwargs["can_use_tool"] = perm.build_callback_for_session(
            session_id,
            plan_return_permission=runtime_plan_return_permission,
        )
    # Third-party Anthropic-compatible endpoints may emit a `thinking` block
    # without the signature key that the SDK parser requires.  Use the narrow
    # compatibility client only for vendor routes; native Claude stays on the
    # SDK's strict parser.  The existing post-turn JSONL cleanup removes the
    # empty parser sentinel before a future resume.
    client_cls = (
        UnsignedThinkingCompatibleClient
        if endpoints.is_third_party(model)
        else MuseLabSDKClient
    )
    try:
        client = client_cls(options=ClaudeAgentOptions(**opts_kwargs))
        try:
            await client.connect()
        except BaseException:
            # connect() may already have spawned the CLI before it becomes
            # cancellable. Until this function returns, the client is not in
            # the pool and no other cleanup path can reach it.
            try:
                await _disconnect_unpooled_client(client, session_id)
            except Exception as cleanup_exc:
                sys.stderr.write(
                    "[client-pool] connect failure cleanup pending "
                    f"sid={session_id[:8]} "
                    f"exc={type(cleanup_exc).__name__}\n"
                )
                sys.stderr.flush()
            raise
        if native_retry_resume:
            _native_retry_commits[session_id] = (
                retry_source_session_id,
                retry_target_user_uuid,
            )
        return client
    except Exception as e:
        # Two failure modes we recover from by swapping session_id ⇔ resume:
        #   - tried `resume=` but CLI has no on-disk session for it
        #     → swap to `session_id=` (create fresh tied to our UUID)
        #   - tried `session_id=` but CLI reports "already in use"
        #     (its internal lock leaked, or a JSONL appeared between
        #     our glob check and the spawn) → swap to `resume=`
        # Classify FIRST: only session-id/resume conflicts are recoverable
        # by swapping. Auth / network / config failures are NOT — blindly
        # swapping there just spawns a second doomed CLI subprocess and
        # buries the real cause behind a misleading "already in use" retry
        # loop. Re-raise anything that isn't a genuine session conflict.
        if native_retry_resume:
            # Truncating resume is guarded specifically so an unexpected tail
            # can never be discarded.  Swapping it to an ordinary create or
            # resume would silently bypass that safety contract.
            raise
        err_text = str(e).lower()
        _is_session_conflict = (
            "already in use" in err_text
            or "no conversation found" in err_text
            or "no session" in err_text
            or "session not found" in err_text
            or ("resume" in err_text and "not found" in err_text)
        )
        if not _is_session_conflict:
            raise
        used_session_id = "session_id" in opts_kwargs
        if used_session_id and "already in use" in err_text:
            opts_kwargs.pop("session_id", None)
            opts_kwargs["resume"] = session_id
        else:
            opts_kwargs.pop("resume", None)
            opts_kwargs["session_id"] = session_id
        # The fallback can ALSO hit "already in use" — retry with backoff.
        last_err: Exception | None = None
        for attempt in range(4):
            try:
                client = client_cls(
                    options=ClaudeAgentOptions(**opts_kwargs))
                try:
                    await client.connect()
                except BaseException:
                    try:
                        await _disconnect_unpooled_client(client, session_id)
                    except Exception as cleanup_exc:
                        sys.stderr.write(
                            "[client-pool] retry connect cleanup pending "
                            f"sid={session_id[:8]} "
                            f"exc={type(cleanup_exc).__name__}\n"
                        )
                        sys.stderr.flush()
                    raise
                if attempt > 0:
                    sys.stderr.write(
                        f"[chat] sid={session_id[:8]} connect retry "
                        f"succeeded on attempt {attempt + 1}\n")
                    sys.stderr.flush()
                return client
            except Exception as e2:
                last_err = e2
                if "already in use" not in str(e2).lower():
                    raise
                # Backoff: 200ms, 400ms, 800ms, 1600ms (~3s total).
                sys.stderr.write(
                    f"[chat] sid={session_id[:8]} attempt {attempt + 1} "
                    f"hit 'already in use', backing off "
                    f"{200 * (2 ** attempt)}ms\n")
                sys.stderr.flush()
                await asyncio.sleep(0.2 * (2 ** attempt))
        if last_err is not None:
            raise last_err
        raise RuntimeError("unreachable")   # for type checker


# ── MCP readiness gate ──────────────────────────────────────────────────
# The wedge bug: extended thinking requires the thinking block in the latest
# assistant message to be returned to the API *unmodified*. MCP servers
# connect lazily in the background AFTER client.connect() returns, so the
# available tool-set can change PARTWAY THROUGH the first turn. When that
# happens while a thinking block is in flight, its signature no longer
# validates → permanent `400 ... thinking blocks ... cannot be modified`
# that no further prompt can recover.
#
# Fix: after connect, block until every MCP server has reached a *terminal*
# connection state (connected / failed / needs-auth) BEFORE we let the first
# turn run. Then the tool-set is frozen before any thinking block exists.
# A timeout backstops the wait so a hung server can't hang the user forever.
#
# Only "still settling" states keep us waiting. needs-auth is terminal: the
# server's tools won't register until the user does OAuth, so the tool-set is
# already stable (just smaller). Unknown/odd states are treated as terminal —
# combined with the timeout, we never block indefinitely on a shape we don't
# recognise.
_MCP_PENDING_STATES = {"pending", "connecting", "authenticating", "starting"}


def _mcp_servers_from_status(status: object) -> list[tuple[str, str]]:
    """Normalise the CLI's mcp_status control response into a list of
    (name, lowercased-state) pairs, tolerating every shape we've seen:

      - top-level key is `mcpServers` (current CLI), `servers`, or `mcp_servers`
      - value is a list of {name, status, …} dicts, OR
      - value is a dict keyed by server name → {status: …} | "<state>"

    Returning NAMES (not just bare states) is what lets the gate notice when a
    server set is still GROWING — claude.ai proxy connectors enumerate a beat
    after the local ones, and a name-less state list can't tell "two servers,
    both connected" from "the same two servers we saw last poll." Unnamed
    entries fall back to a positional synthetic key so they still count toward
    set stability."""
    if not isinstance(status, dict):
        return []
    servers = status.get("mcpServers")
    if servers is None:
        servers = status.get("servers")
    if servers is None:
        servers = status.get("mcp_servers")
    out: list[tuple[str, str]] = []

    def _emit(name: object, state: object, idx: int) -> None:
        nm = str(name) if name else f"__idx{idx}"
        out.append((nm, str(state).lower()))

    if isinstance(servers, dict):
        for k, v in servers.items():
            if isinstance(v, dict):
                _emit(v.get("name", k), v.get("status", ""), len(out))
            else:
                _emit(k, v, len(out))
    elif isinstance(servers, list):
        for i, v in enumerate(servers):
            if isinstance(v, dict):
                _emit(v.get("name", f"__idx{i}"), v.get("status", ""), i)
            elif isinstance(v, str):
                _emit(f"__idx{i}", v, i)
    return out


def _mcp_states_from_status(status: object) -> list[str]:
    """Back-compat shim: just the state strings (drops names). Retained for
    any caller that only cares about pending-ness."""
    return [state for _name, state in _mcp_servers_from_status(status)]


async def _await_mcp_ready(client: ClaudeSDKClient, *,
                           timeout: float = 30.0, poll: float = 0.25) -> None:
    """Block until the MCP tool-set has STABILISED, or until `timeout`.

    "Stabilised" = two consecutive polls return the SAME non-empty set of
    (name, state) pairs AND none of them is still settling. We require the set
    to be identical across two polls — not merely "nothing pending right now" —
    because that older, weaker check is exactly how the wedge bug came back:

      claude.ai proxy connectors (Gmail / Calendar / Drive / IBKR) enumerate a
      beat AFTER the local stdio servers. At the first poll the status response
      lists only {gmail, muselab} — neither pending — so the old gate declared
      "all terminal" and let the turn start. The proxies then connected and
      registered their tools MID-FIRST-TURN, changing the tool-set the model's
      in-flight thinking block was signed against → 400 "thinking blocks …
      cannot be modified". Waiting for the set to stop GROWING closes that race
      without needing to predict how many connectors will show up.

    `needs-auth` / `failed` count as terminal (settled) states — a connector
    that needs OAuth or has crashed won't register tools on its own, so its
    presence doesn't keep us waiting. Only `_MCP_PENDING_STATES` block.

    Best-effort: any failure to read status, or an unrecognised shape, just
    returns (we don't block the turn on our own inability to introspect)."""
    deadline = time.monotonic() + timeout
    prev: frozenset[tuple[str, str]] | None = None
    while True:
        try:
            status = await client.get_mcp_status()
        except Exception:
            return   # status unavailable — don't hold the turn hostage
        servers = _mcp_servers_from_status(status)
        snapshot = frozenset(servers)
        pending = any(state in _MCP_PENDING_STATES for _n, state in servers)
        # Ready iff: something is configured (non-empty), nothing is still
        # settling, AND the exact set matched the previous poll (so a
        # late-arriving connector can't have slipped in between snapshots).
        if servers and not pending and snapshot == prev:
            return
        prev = snapshot
        if time.monotonic() >= deadline:
            sys.stderr.write(
                f"[mcp-gate] readiness timeout after {timeout}s; "
                f"servers={sorted(servers)} — proceeding anyway\n")
            sys.stderr.flush()
            return
        await asyncio.sleep(poll)


def _has_enabled_external_mcp() -> bool:
    """True if at least one user/external MCP server is configured and not
    disabled — i.e. the next fresh client will spend time connecting tools.
    Used to decide whether to arm the wedge-readiness gate / show the frontend
    'connecting tools…' hint. The internal 'muselab' server is added separately
    and isn't in this view.

    Covers TWO classes of external MCP:
      1. `mcpServers` entries (local stdio / remote http) — visible via
         _load_mcp_merged().
      2. claude.ai-managed connectors (Gmail / Calendar / Drive / IBKR) — a
         separate `claudeai-proxy` transport that never lands under any
         `mcpServers` key. Without (2) the gate was SKIPPED on claude.ai-only
         installs, which is exactly how the wedge bug came back (the connector
         connected mid-first-turn). See api_settings.has_claude_ai_connectors.
    """
    try:
        from .api_settings import _load_mcp_merged, has_claude_ai_connectors
        for spec in _load_mcp_merged().values():
            if not spec.get("disabled"):
                return True
        if has_claude_ai_connectors():
            return True
    except Exception:
        pass
    return False


async def _disconnect_unpooled_client(
    client: ClaudeSDKClient,
    session_id: str,
) -> None:
    return await chat_runtime.disconnect_unpooled_client(client, session_id)


async def get_client(
    session_id: str,
    model: str,
    permission: str = "bypassPermissions",
    effort: str = "",
    service_tier: str = "",
    plan_return_permission: str = "",
    startup_phase: Callable[[str, int], None] | None = None,
) -> ClaudeSDKClient:
    """Compatibility facade for the extracted SDK runtime pool."""
    return await chat_runtime.get_client(
        session_id,
        model,
        permission,
        effort=effort,
        service_tier=service_tier,
        plan_return_permission=plan_return_permission,
        startup_phase=startup_phase,
    )


_STREAM_EOF = chat_runtime.STREAM_EOF
_SessionStream = chat_runtime.SessionStream
_session_streams = chat_runtime.SESSION_STREAMS


async def _evict_failed_session_stream(stream: _SessionStream) -> None:
    return await chat_runtime.evict_failed_session_stream(stream)


def _ensure_session_stream(
    key: _ClientKey,
    client: ClaudeSDKClient,
) -> _SessionStream:
    return chat_runtime.ensure_session_stream(key, client)


def _stream_for(client: ClaudeSDKClient) -> _SessionStream | None:
    return chat_runtime.stream_for(client)


async def _drop_session_streams(session_id: str) -> None:
    return await chat_runtime.drop_session_streams(session_id)


def _retain_detached_cleanup(task: asyncio.Task) -> None:
    """Keep a timed-out cancellation owner alive and consume its outcome."""
    _maintenance_tasks.add(task)

    def _done(done: asyncio.Task) -> None:
        _maintenance_tasks.discard(done)
        if done.cancelled():
            return
        with suppress(Exception):
            done.exception()

    task.add_done_callback(_done)


RuntimeCleanupTimeout = chat_runtime.RuntimeCleanupTimeout


def _track_session_disconnect(session_id: str, task: asyncio.Task) -> None:
    chat_runtime.track_session_disconnect(session_id, task)


async def _join_session_disconnects(
    session_id: str,
    clients: Iterable[ClaudeSDKClient] = (),
    *,
    timeout: float = _CLIENT_DISCONNECT_DEADLINE_S,
) -> bool:
    return await chat_runtime.join_session_disconnects(
        session_id, clients, timeout=timeout
    )


def _track_session_runtime_cleanup(
    session_id: str,
    task: asyncio.Task,
) -> None:
    """Retain one cancelled owner until it reaches a terminal state.

    Cancellation and ordinary exceptions are both terminal for this fence: the
    safety property is that the owner can no longer append or recreate files,
    not that its business operation succeeded.
    """
    owners = _session_runtime_cleanup_tasks.setdefault(session_id, set())
    owners.add(task)

    def _done(done: asyncio.Task) -> None:
        current = _session_runtime_cleanup_tasks.get(session_id)
        if current is not None:
            current.discard(done)
            if not current:
                _session_runtime_cleanup_tasks.pop(session_id, None)
        if not done.cancelled():
            with suppress(Exception):
                done.exception()

    task.add_done_callback(_done)


async def _join_session_runtime_cleanup(
    session_id: str,
    tasks: Iterable[asyncio.Task] = (),
    *,
    timeout: float = 5.0,
) -> bool:
    owners = {
        task
        for task in _session_runtime_cleanup_tasks.get(session_id, set())
        if not task.done()
    }
    for task in tasks:
        if task is not None and not task.done():
            _track_session_runtime_cleanup(session_id, task)
            owners.add(task)
    if not owners:
        return True
    done, pending = await asyncio.wait(owners, timeout=max(0.1, timeout))
    if done:
        await asyncio.gather(*done, return_exceptions=True)
    return not pending


async def disconnect_client(session_id: str) -> None:
    """Compatibility facade for session-wide SDK runtime teardown."""
    return await chat_runtime.disconnect_client(session_id)


async def _disconnect_background_task_owner(
    session_id: str,
    client: ClaudeSDKClient,
) -> None:
    return await chat_runtime.disconnect_background_task_owner(
        session_id, client
    )


async def shutdown_runtime() -> None:
    """Boundedly stop every in-process chat task, stream, and SDK client."""

    # Stop detached task watchers and active turn pumps before tearing down the
    # shared SDK streams they consume.
    scheduled_deliveries = tuple(_sdk_scheduled_deliveries.values())
    active_broadcasts = tuple({
        id(broadcast): broadcast
        for broadcast in (
            *tuple(_active_turns.values()),
            *(delivery.broadcast for delivery in scheduled_deliveries),
        )
    }.values())
    tasks: set[asyncio.Task] = {
        task for task in _task_watchers.values() if not task.done()
    }
    tasks.update(task for task in _maintenance_tasks if not task.done())
    tasks.update(
        delivery.registration_task
        for delivery in scheduled_deliveries
        if delivery.registration_task is not None
        and not delivery.registration_task.done()
    )
    protected_cleanup_tasks: set[asyncio.Task] = {
        task for task in _session_purge_tasks.values() if not task.done()
    }
    for owners in _session_runtime_cleanup_tasks.values():
        protected_cleanup_tasks.update(
            task for task in owners if not task.done()
        )
    for broadcast in active_broadcasts:
        broadcast.cancelled = True
        for attr in ("startup_task", "startup_owner_task", "task"):
            task = getattr(broadcast, attr, None)
            if isinstance(task, asyncio.Task) and not task.done():
                tasks.add(task)
        for attr in (
            "_attachment_prepare_task",
            "_attachment_rollback_task",
            "_startup_terminal_cleanup_task",
        ):
            task = getattr(broadcast, attr, None)
            if isinstance(task, asyncio.Task) and not task.done():
                protected_cleanup_tasks.add(task)
    tasks.difference_update(protected_cleanup_tasks)
    # Normal application shutdown runs on the owning loop. Test harnesses and
    # embedded callers can close that loop before invoking this idempotent cleanup
    # from a replacement loop; asyncio cannot cancel or wait on those stale tasks.
    current_loop = asyncio.get_running_loop()
    tasks = {task for task in tasks if task.get_loop() is current_loop}
    protected_cleanup_tasks = {
        task for task in protected_cleanup_tasks
        if task.get_loop() is current_loop
    }
    for task in tasks:
        task.cancel()
    if tasks:
        done, pending = await asyncio.wait(tasks, timeout=2.0)
        if done:
            await asyncio.gather(*done, return_exceptions=True)
        for task in pending:
            _retain_detached_cleanup(task)

    # A cancelled owner may never run its coroutine/finally. Create an explicit
    # attachment finalizer for every active broadcast and protect it through the
    # second shutdown join.
    attachment_finalizers: dict[asyncio.Task, TurnBroadcast] = {}
    for broadcast in active_broadcasts:
        if (
            broadcast._attachment_lease is not None
            or broadcast._attachment_prepare_task is not None
        ):
            finalizer = asyncio.create_task(
                _rollback_broadcast_attachments(broadcast)
            )
            protected_cleanup_tasks.add(finalizer)
            attachment_finalizers[finalizer] = broadcast

    broadcasts = {
        id(broadcast): broadcast
        for broadcast in (
            *tuple(_active_turns.values()),
            *tuple(_recent_turns.values()),
        )
    }.values()
    for broadcast in broadcasts:
        broadcast.close()
    for delivery in scheduled_deliveries:
        if delivery.broadcast.activity_started:
            await _finish_activity(
                delivery.key[0], delivery.broadcast, "cancelled")
    for handle in _recent_turn_expiry_handles.values():
        handle.cancel()
    _recent_turn_expiry_handles.clear()
    _recent_turns.clear()

    _active_turns.clear()
    _sdk_scheduled_deliveries.clear()
    with _sdk_cron_state_lock:
        _sdk_cron_jobs.clear()
        _sdk_cron_tool_calls.clear()
    _sessions_with_inflight_tasks.clear()
    _bg_task_pinned_at.clear()
    _task_watchers.clear()
    _runtime_prewarm_tasks.clear()
    _queue_drain_tasks.clear()
    _queue_drain_retry_tasks.clear()
    _queue_drain_rekicks.clear()
    _queue_drain_locks.clear()
    _native_retry_commits.clear()

    async def _join_protected_cleanup() -> None:
        if not protected_cleanup_tasks:
            return
        done, pending = await asyncio.wait(
            protected_cleanup_tasks,
            timeout=_ATTACHMENT_SHUTDOWN_JOIN_S,
        )
        if done:
            await asyncio.gather(*done, return_exceptions=True)
        if not pending:
            return

        # Persist predicted final paths before the process lifecycle fence
        # opens. Startup reconciliation removes files a slow worker may publish
        # after the bounded shutdown deadline.
        candidates: list[str] = []
        for task in pending:
            broadcast = attachment_finalizers.get(task)
            if broadcast is not None:
                candidates.extend(
                    _broadcast_attachment_artifact_candidates(broadcast)
                )
            _retain_detached_cleanup(task)
        if candidates:
            await asyncio.to_thread(
                _record_attachment_cleanup_intents,
                candidates,
            )

    await asyncio.gather(
        chat_runtime.shutdown_clients(),
        _join_protected_cleanup(),
    )


async def _rebuild_session_runtime(session_id: str) -> None:
    """Disconnect now, or defer until the active turn reaches a safe boundary."""
    if _session_runtime_busy(session_id):
        _pending_runtime_rebuilds.add(session_id)
        return
    async with _session_runtime_lock_for(session_id):
        # A turn may reserve the session while this coroutine waits for the
        # mutex. Its get_client() will consume the marker before querying.
        if _session_runtime_busy(session_id):
            _pending_runtime_rebuilds.add(session_id)
            return
        await disconnect_client(session_id)


# ====== sessions REST ======

class CreateReq(BaseModel):
    name: str | None = None
    model: str | None = None
    permission: str = ""
    cwd: str = ""
    # Optimistic-create (2026-06-07): the client mints the session UUID up
    # front so the new-chat tab opens with ZERO network wait, then POSTs here
    # in the background to register it. When present AND a valid canonical
    # UUID, we register THIS id instead of generating a fresh one — the send
    # path binds the SDK session to the same UUID on first message either way
    # (chat.py uses session_id= when no JSONL exists). Strictly validated
    # server-side before it ever touches a filesystem path (sidecar).
    id: str | None = None
    # P2/B: the client's currently-open tab ids. Passed so empty-session
    # recycling (prune_empty_sessions) NEVER deletes a blank session the user
    # has open in a tab and is about to type in. Empty + closed + unpinned +
    # auto-named is the only thing eligible for cleanup.
    open_ids: list[str] | None = None
    # Lightweight side-question branches remain normal resumable sessions but
    # intentionally do not become rows in the global task/activity ledger.
    activity_hidden: bool = False
    # A side question is conversational, not a general workspace agent.  Its
    # runtime exposes only public web lookup tools even when the branch is
    # later opened as a full chat.
    runtime_profile: Literal["", "side_question"] = ""


_last_orphan_gc_at = 0.0
_ORPHAN_GC_INTERVAL_S = 3600   # at most hourly


def _attachments_base() -> Path:
    """Root dir for user-uploaded image originals.

    Lives under the primary workspace (`MUSELAB_ROOT`), not inside the
    muselab repo. Two reasons:
      1. The repo's `sessions/` dir was already gitignored, but conceptually
         user-data shouldn't sit in the install dir at all — uninstall /
         reinstall / git clean should never touch the user's files.
      2. The workspace is already the user-selected data boundary, keeping
         source files and generated attachments together for backup.

    Hidden (dot-prefixed) so it doesn't clutter the user's file browser
    or workspace file tree.
    """
    return ROOT / ".muselab-attach"


def _attachment_session_dir(
    session_id: str,
    *,
    create: bool,
) -> Path | None:
    """Return a real private session directory without following symlinks."""
    if not re.fullmatch(r"[A-Za-z0-9_\-]{6,80}", session_id):
        raise UnsafePrivatePath("invalid attachment session directory")
    base = _attachments_base()
    if not ensure_private_directory(base, create=create):
        return None
    session_dir = base / session_id
    if not ensure_private_directory(session_dir, create=create):
        return None
    return session_dir


def ensure_private_attachment_storage() -> int:
    """Repair attachment permissions from older umask-dependent releases."""
    base = _attachments_base()
    if not ensure_private_directory(base, create=False):
        return 0
    repaired = 1
    for session_dir in base.iterdir():
        if repair_private_path(session_dir) != "directory":
            continue
        repaired += 1
        try:
            children = list(session_dir.iterdir())
        except OSError:
            continue
        for attachment in children:
            if repair_private_path(attachment) == "file":
                repaired += 1
    return repaired


# Longest filename component we'll write. Keeps `{aid}-{name}` (32 hex + dash
# + this) comfortably under the 255-byte limit ext4/APFS enforce PER COMPONENT
# — and that limit is in BYTES, so a CJK name at 3 bytes/char hits it ~3x
# sooner than a Latin one. Truncation is applied to the stem so the extension
# always survives; the extension is what tells the agent (and our MIME map)
# what the file actually is.
_ATTACH_NAME_MAX = 60


def _safe_attach_name(name: str) -> str:
    """Turn a client-supplied filename into a safe single path component.

    Keeps the original name legible (CJK included) because it lands in the
    prompt as a path the agent reads back — `a3f2…-季度报表.xlsx` tells the
    model what it's looking at, `a3f2….bin` does not. Everything that could
    escape the directory or confuse a shell is replaced:
      - any directory part is dropped (`Path.name`)
      - path separators, NUL, control chars, quotes, and whitespace runs → `_`
      - leading dots are stripped so nothing lands as a hidden file
    Falls back to "file" if the input sanitises down to nothing.
    """
    base = Path(str(name or "")).name
    # Reject separators explicitly first — Path.name already drops POSIX dirs,
    # but a Windows-style "..\\..\\evil" arriving on Linux keeps its
    # backslashes, so the replace below has to see them.
    base = base.replace("\\", "_").replace("/", "_")
    base = re.sub(r"[\x00-\x1f\x7f]", "", base)
    base = re.sub(r'[\s"\'`$;|&<>*?]+', "_", base)
    # Strip dots and underscores together, in one pass. Doing it as two
    # separate steps (lstrip(".") then strip("_")) leaves debris when they
    # interleave: "..\\..\\windows" sanitises to ".._.._windows", the lstrip
    # eats the leading dots, the strip eats one underscore, and the SECOND
    # ".." is left stranded at the front.
    base = base.strip("._") or "file"
    if len(base.encode("utf-8")) > _ATTACH_NAME_MAX:
        stem, dot, ext = base.rpartition(".")
        if not dot or len(ext) > 8:
            stem, dot, ext = base, "", ""
        budget = _ATTACH_NAME_MAX - len((dot + ext).encode("utf-8"))
        trimmed = stem.encode("utf-8")[:max(1, budget)]
        # A byte-slice can land mid-codepoint; drop the partial tail.
        stem = trimmed.decode("utf-8", "ignore") or "file"
        base = stem + dot + ext
    return base


def _persist_attachment(session_id: str, aid: str, name: str,
                        data: bytes) -> tuple[str, str] | None:
    """Write one attachment to `.muselab-attach/{sid}/{aid}-{safe name}`.

    Returns (absolute path, browser URL), or None if the write failed. The
    caller decides whether that file is optional (image/PDF local fallback)
    or required (text/xlsx path manifest). Callers log nothing extra; the
    stderr line here is the single record.

    The `{aid}-` prefix is what makes the name collision-proof: two messages
    can each attach a `report.csv` and neither overwrites the other, while the
    human-readable half still tells the agent (and the user browsing the
    folder) what the file is.
    """
    safe = _safe_attach_name(name)
    try:
        if not re.fullmatch(r"[A-Za-z0-9][A-Za-z0-9\-]{0,79}", aid):
            raise UnsafePrivatePath("invalid attachment id")
        attach_dir = _attachment_session_dir(session_id, create=True)
        if attach_dir is None:
            raise UnsafePrivatePath("attachment directory is unavailable")
        path = attach_dir / f"{aid}-{safe}"
        write_private_bytes(path, data)
    except Exception as e:
        sys.stderr.write(
            f"[attach] persist failed sid={obs.short_id(session_id)} "
            f"aid={obs.short_id(aid)} exc={type(e).__name__} "
            f"kind=write\n")
        sys.stderr.flush()
        return None
    url = (f"/api/chat/attachments/{session_id}/"
           f"{urllib.parse.quote(f'{aid}-{safe}')}")
    return str(path), url


def _doc_item(name: str, kind: str, saved: tuple[str, str] | None) -> dict:
    """Bubble metadata for one non-image attachment. `url` is present only
    when the file actually made it to disk, so the frontend can render a
    dead-end chip rather than a link that 404s."""
    item = {"name": name, "kind": kind}
    if saved:
        item["url"] = saved[1]
    return item


_attachment_cleanup_intent_lock = threading.RLock()
_PRIVATE_TEMP_NAME_RE = re.compile(r"^\..+\.[0-9a-f]{16}\.tmp$")


def _attachment_cleanup_intent_path() -> Path:
    return _attachments_base() / ".cleanup-intents.json"


def _normalized_attachment_cleanup_paths(paths: Iterable[str]) -> list[str]:
    """Keep only lexical descendants of MuseLab's attachment root."""
    base = Path(os.path.abspath(_attachments_base()))
    intent_path = Path(os.path.abspath(_attachment_cleanup_intent_path()))
    normalized: list[str] = []
    for raw_path in paths:
        path = Path(os.path.abspath(str(raw_path)))
        try:
            path.relative_to(base)
        except ValueError:
            continue
        if path == intent_path:
            continue
        normalized.append(str(path))
    return list(dict.fromkeys(normalized))


def _load_attachment_cleanup_intents_locked() -> list[str]:
    path = _attachment_cleanup_intent_path()
    try:
        if not ensure_private_regular_file(path):
            return []
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, ValueError, TypeError, UnsafePrivatePath):
        return []
    raw_paths = payload.get("paths") if isinstance(payload, dict) else None
    return _normalized_attachment_cleanup_paths(
        raw_paths if isinstance(raw_paths, list) else []
    )


def _write_attachment_cleanup_intents_locked(paths: Iterable[str]) -> None:
    normalized = _normalized_attachment_cleanup_paths(paths)
    intent_path = _attachment_cleanup_intent_path()
    if not normalized:
        try:
            intent_path.unlink(missing_ok=True)
        except OSError:
            pass
        return
    try:
        write_private_bytes(
            intent_path,
            json.dumps({"paths": normalized}, ensure_ascii=False).encode(),
        )
    except Exception as exc:
        sys.stderr.write(
            f"[attach] cleanup intent write failed "
            f"exc={type(exc).__name__}\n"
        )
        sys.stderr.flush()


def _record_attachment_cleanup_intents(paths: Iterable[str]) -> None:
    with _attachment_cleanup_intent_lock:
        existing = _load_attachment_cleanup_intents_locked()
        _write_attachment_cleanup_intents_locked([*existing, *paths])


def _stale_attachment_temp_paths() -> list[Path]:
    """List writer-owned temp files without following arbitrary directories."""
    base = _attachments_base()
    if not ensure_private_directory(base, create=False):
        return []
    directories = [base]
    try:
        children = list(base.iterdir())
    except OSError:
        return []
    directories.extend(
        child for child in children
        if private_path_kind(child) == "directory"
    )
    stale: list[Path] = []
    for directory in directories:
        try:
            entries = list(directory.iterdir())
        except OSError:
            continue
        stale.extend(
            entry for entry in entries
            if _PRIVATE_TEMP_NAME_RE.fullmatch(entry.name)
            and private_path_kind(entry) == "file"
        )
    return stale


def _drain_attachment_cleanup_intents() -> int:
    """Retry orphan cleanup from a prior failure or bounded shutdown."""
    with _attachment_cleanup_intent_lock:
        pending = _load_attachment_cleanup_intents_locked()
        failed: list[str] = []
        removed = 0
        pending.extend(str(path) for path in _stale_attachment_temp_paths())
        for raw_path in pending:
            try:
                Path(raw_path).unlink(missing_ok=True)
                removed += 1
            except OSError:
                failed.append(raw_path)
        _write_attachment_cleanup_intents_locked(failed)
    if failed:
        sys.stderr.write(
            f"[attach] cleanup retry pending count={len(failed)}\n"
        )
        sys.stderr.flush()
    return removed


def _broadcast_attachment_artifact_candidates(
    broadcast: TurnBroadcast,
) -> list[str]:
    """Predict every final path so bounded shutdown can persist cleanup."""
    paths = list(
        broadcast._prepared_attachments.artifact_paths
        if broadcast._prepared_attachments is not None else []
    )
    lease = broadcast._attachment_lease
    if lease is None:
        return _normalized_attachment_cleanup_paths(paths)
    base = _attachments_base() / broadcast.session_id
    for aid, entry in lease.items:
        kind = str(entry.get("kind") or "image")
        name = str(entry.get("name") or "file")
        if kind == "image":
            ext = {
                "image/png": "png", "image/jpeg": "jpg",
                "image/jpg": "jpg", "image/gif": "gif",
                "image/webp": "webp",
            }.get(str(entry.get("mime") or ""), "bin")
            paths.append(str(base / f"{aid}.{ext}"))
        elif kind in {"pdf", "text", "xlsx"}:
            paths.append(str(base / f"{aid}-{_safe_attach_name(name)}"))
            if kind == "xlsx":
                txt_name = Path(name).stem + ".txt"
                paths.append(str(
                    base / f"{aid}-txt-{_safe_attach_name(txt_name)}"
                ))
    return _normalized_attachment_cleanup_paths(paths)


def _cleanup_prepared_attachments_sync(
    prepared: "_PreparedStagedAttachments",
) -> None:
    """Remove pre-query artifacts after a lease rolls back."""
    failed: list[str] = []
    for raw_path in reversed(list(dict.fromkeys(prepared.artifact_paths))):
        try:
            Path(raw_path).unlink(missing_ok=True)
        except OSError as exc:
            failed.append(raw_path)
            sys.stderr.write(
                f"[attach] rollback unlink failed "
                f"exc={type(exc).__name__}\n"
            )
            sys.stderr.flush()
    if failed:
        _record_attachment_cleanup_intents(failed)


def _prepare_staged_attachments_sync(
    session_id: str,
    items: tuple[tuple[str, dict], ...],
) -> "_PreparedStagedAttachments":
    """Decode, thumbnail and persist one leased attachment set in a worker."""
    prepared = _PreparedStagedAttachments()
    try:
        for aid, entry in items:
            kind = entry.get("kind", "image")
            if kind == "image":
                encoded = str(entry.get("b64") or "")
                try:
                    raw = base64.b64decode(encoded, validate=True)
                except Exception:
                    raise _AttachmentPreparationError(
                        "image attachment could not be decoded") from None
                prepared.img_blocks.append({
                    "type": "image",
                    "source": {
                        "type": "base64",
                        "media_type": entry["mime"],
                        "data": encoded,
                    },
                })
                ext_map = {
                    "image/png": "png", "image/jpeg": "jpg",
                    "image/jpg": "jpg", "image/gif": "gif",
                    "image/webp": "webp",
                }
                ext = ext_map.get(entry["mime"], "bin")
                full_url = None
                try:
                    attach_dir = _attachment_session_dir(
                        session_id, create=True)
                    if attach_dir is None:
                        raise UnsafePrivatePath(
                            "attachment directory unavailable")
                    attach_path = attach_dir / f"{aid}.{ext}"
                    write_private_bytes(attach_path, raw)
                    prepared.artifact_paths.append(str(attach_path))
                    full_url = (
                        f"/api/chat/attachments/{session_id}/{aid}.{ext}"
                    )
                except Exception as exc:
                    sys.stderr.write(
                        f"[attach] persist failed "
                        f"sid={obs.short_id(session_id)} "
                        f"aid={obs.short_id(aid)} "
                        f"exc={type(exc).__name__} kind=write\n")
                    sys.stderr.flush()

                thumb_b64 = None
                image_module = None
                try:
                    import io
                    import warnings
                    from PIL import Image as image_module

                    with warnings.catch_warnings():
                        warnings.simplefilter(
                            "error", image_module.DecompressionBombWarning)
                        with image_module.open(io.BytesIO(raw)) as image:
                            width, height = image.size
                            if width * height > _IMAGE_THUMBNAIL_MAX_PIXELS:
                                raise image_module.DecompressionBombError(
                                    "thumbnail pixel budget exceeded")
                            image.thumbnail((160, 160))
                            buf = io.BytesIO()
                            image.convert("RGB").save(
                                buf, "JPEG", quality=60)
                            thumb_b64 = base64.b64encode(
                                buf.getvalue()).decode("ascii")
                except Exception as exc:
                    bomb_types = (
                        getattr(image_module, "DecompressionBombError", ()),
                        getattr(image_module, "DecompressionBombWarning", ()),
                    )
                    if image_module is not None and isinstance(exc, bomb_types):
                        sys.stderr.write(
                            "[attach] thumbnail skipped reason=pixel_budget\n")
                        sys.stderr.flush()
                item: dict = {"mime": entry["mime"]}
                if thumb_b64:
                    item["thumb"] = thumb_b64
                if full_url:
                    item["url"] = full_url
                prepared.persisted_imgs.append(item)
            elif kind == "pdf":
                encoded = str(entry.get("b64") or "")
                try:
                    raw = base64.b64decode(encoded, validate=True)
                except Exception:
                    raise _AttachmentPreparationError(
                        "PDF attachment could not be decoded") from None
                prepared.pdf_blocks.append({
                    "type": "document",
                    "source": {
                        "type": "base64",
                        "media_type": "application/pdf",
                        "data": encoded,
                    },
                })
                doc_name = entry.get("name", "doc.pdf")
                saved = _persist_attachment(
                    session_id, aid, doc_name, raw)
                if saved:
                    prepared.artifact_paths.append(saved[0])
                    prepared.disk_attachments.append(
                        (_safe_attach_name(doc_name), saved[0], ""))
                # The native SDK block remains useful if the optional local
                # gateway fallback cannot be written.
                prepared.persisted_docs.append(
                    _doc_item(doc_name, "pdf", saved))
            elif kind == "text":
                doc_name = entry.get("name", "file.txt")
                raw = entry.get("raw")
                if not isinstance(raw, bytes):
                    raw = str(entry.get("text") or "").encode("utf-8")
                saved = _persist_attachment(
                    session_id, aid, doc_name, raw)
                if not saved:
                    raise _AttachmentPreparationError(
                        "text attachment could not be persisted")
                prepared.artifact_paths.append(saved[0])
                prepared.disk_attachments.append(
                    (_safe_attach_name(doc_name), saved[0], ""))
                prepared.persisted_docs.append(
                    _doc_item(doc_name, "text", saved))
            elif kind == "xlsx":
                doc_name = entry.get("name", "book.xlsx")
                raw = entry.get("raw")
                if not isinstance(raw, bytes):
                    raise _AttachmentPreparationError(
                        "spreadsheet attachment payload is unavailable")
                transcription = entry.get("text")
                if not isinstance(transcription, str):
                    raise _AttachmentPreparationError(
                        "spreadsheet transcription is unavailable")
                saved = _persist_attachment(
                    session_id, aid, doc_name, raw)
                if not saved:
                    raise _AttachmentPreparationError(
                        "spreadsheet attachment could not be persisted")
                prepared.artifact_paths.append(saved[0])

                txt_name = Path(doc_name).stem + ".txt"
                txt_saved = _persist_attachment(
                    session_id,
                    aid + "-txt",
                    txt_name,
                    transcription.encode("utf-8"),
                )
                if not txt_saved:
                    raise _AttachmentPreparationError(
                        "spreadsheet transcription could not be persisted")
                prepared.artifact_paths.append(txt_saved[0])
                prepared.disk_attachments.append((
                    _safe_attach_name(doc_name),
                    saved[0],
                    f"plain-text transcription: {txt_saved[0]}",
                ))
                prepared.persisted_docs.append(
                    _doc_item(doc_name, "xlsx", saved))
            else:
                raise _AttachmentPreparationError(
                    "unsupported staged attachment kind")
    except BaseException:
        _cleanup_prepared_attachments_sync(prepared)
        raise
    return prepared


async def _await_thread_completion(func: Callable, *args):
    """Join a real worker result even when the awaiting task is cancelled."""
    task = asyncio.create_task(asyncio.to_thread(func, *args))
    try:
        return await asyncio.shield(task)
    except asyncio.CancelledError:
        while not task.done():
            try:
                await asyncio.shield(task)
            except asyncio.CancelledError:
                continue
        try:
            task.result()
        except Exception:
            pass
        raise


async def _prepare_broadcast_attachments(
    broadcast: TurnBroadcast,
    session_id: str,
    lease: "_StagedAttachmentLease",
) -> "_PreparedStagedAttachments":
    """Run preparation off-loop and retain its real result on cancellation."""
    task = broadcast._attachment_prepare_task
    if task is None:
        task = asyncio.create_task(asyncio.to_thread(
            _prepare_staged_attachments_sync,
            session_id,
            lease.items,
        ))
        broadcast._attachment_prepare_task = task
    try:
        prepared = await asyncio.shield(task)
    except asyncio.CancelledError:
        while not task.done():
            try:
                await asyncio.shield(task)
            except asyncio.CancelledError:
                continue
        try:
            prepared = task.result()
        except Exception:
            # The sync worker cleans every partial artifact before raising.
            pass
        else:
            if (broadcast._attachment_lease is lease
                    and broadcast._attachment_rollback_task is None):
                broadcast._prepared_attachments = prepared
        raise
    # A watchdog can start the shared rollback while joining this same worker.
    # Never resurrect its cleaned result after that owner has taken over.
    if (broadcast._attachment_lease is lease
            and broadcast._attachment_rollback_task is None):
        broadcast._prepared_attachments = prepared
    return prepared


async def _rollback_broadcast_attachments(broadcast: TurnBroadcast) -> None:
    """Join and roll back one attachment transaction exactly once.

    Pump, startup, and force-stop paths may arrive concurrently. The first
    caller creates a shielded cleanup owner; every later caller joins that same
    task. Preparation is joined before files are removed and the lease is
    released, so a retry can never race a still-running writer.
    """
    cleanup = broadcast._attachment_rollback_task
    if cleanup is None:
        lease = broadcast._attachment_lease
        rollback_won = _begin_staged_attachment_rollback(lease)

        async def _cleanup() -> None:
            if lease is None or not rollback_won:
                return
            prepared = broadcast._prepared_attachments
            prepare_task = broadcast._attachment_prepare_task
            try:
                if prepare_task is not None:
                    try:
                        prepared = await asyncio.shield(prepare_task)
                    except asyncio.CancelledError:
                        if not prepare_task.cancelled():
                            while not prepare_task.done():
                                try:
                                    await asyncio.shield(prepare_task)
                                except asyncio.CancelledError:
                                    continue
                        if prepare_task.cancelled():
                            prepared = None
                        else:
                            try:
                                prepared = prepare_task.result()
                            except Exception:
                                prepared = None
                    except Exception:
                        # The sync preparer removes its own partial files before
                        # propagating an error.
                        prepared = None
                    else:
                        broadcast._prepared_attachments = prepared
                if prepared is not None:
                    await _await_thread_completion(
                        _cleanup_prepared_attachments_sync, prepared)
            finally:
                _release_staged_attachment_lease(lease)
                broadcast._attachment_lease = None
                broadcast._prepared_attachments = None
                broadcast._attachment_prepare_task = None

        cleanup = asyncio.create_task(_cleanup())
        broadcast._attachment_rollback_task = cleanup
    try:
        await asyncio.shield(cleanup)
    except asyncio.CancelledError:
        # Repeated cancellation cannot cut the transaction between releasing
        # the staged id and settling the caller's remaining terminal state.
        while not cleanup.done():
            try:
                await asyncio.shield(cleanup)
            except asyncio.CancelledError:
                continue
        cleanup.result()
        raise


def _commit_broadcast_attachments(broadcast: TurnBroadcast) -> None:
    """Consume a lease exactly after the SDK query write succeeds."""
    lease = broadcast._attachment_lease
    if lease is None:
        return
    if not _commit_staged_attachment_lease(lease):
        # query() already returned: transport acceptance is no longer
        # reversible. Consume any objects still owned by this lease so a retry
        # cannot duplicate an ambiguously submitted attachment.
        _fail_closed_staged_attachment_lease(lease)
        raise _AttachmentCommitUncertain("attachment submission state is uncertain")
    broadcast._attachment_lease = None
    broadcast._prepared_attachments = None
    broadcast._attachment_prepare_task = None


def _migrate_legacy_attachments() -> None:
    """One-shot migration: sessions/attachments/* → ROOT/.muselab-attach/*.
    Runs at module import. Idempotent — only moves dirs that don't yet
    exist in the new location. Old location is removed when empty so a
    second-pass migration is a no-op."""
    old_base = sess.SESS_DIR / "attachments"
    new_base = _attachments_base()
    if private_path_kind(old_base) != "directory" or old_base == new_base:
        return
    try:
        ensure_private_directory(new_base)
    except (OSError, UnsafePrivatePath):
        return
    moved = 0
    for child in list(old_base.iterdir()):
        if private_path_kind(child) != "directory":
            continue
        target = new_base / child.name
        if private_path_kind(target) != "missing":
            continue  # already migrated; skip (don't clobber)
        try:
            shutil.move(str(child), str(target))
            moved += 1
        except OSError:
            pass
    ensure_private_attachment_storage()
    # Remove empty old base
    try:
        if not any(old_base.iterdir()):
            old_base.rmdir()
    except OSError:
        pass
    if moved:
        sys.stderr.write(f"[muselab] migrated {moved} attachment dirs to {new_base}\n")
        sys.stderr.flush()


# Run migration once at import (cheap if no-op).
try:
    _migrate_legacy_attachments()
    _drain_attachment_cleanup_intents()
except Exception:
    pass

# Single-slot ETag digest cache for GET /sessions — see usage below.
_LIST_ETAG_CACHE: dict[str, tuple] = {}


@router.get("/sessions", dependencies=[Depends(require_token)])
def list_sessions_api(
    request: Request,
    response: Response,
    limit: int = Query(0, ge=0, le=2000),
    ids: str = Query(""),
    q: str = Query(""),
    workspace_only: bool = Query(False),
    workspace_root: Path = Depends(resolve_workspace_root),
):
    # P2 (perf): paginate. `list_sessions()` returns ALL sessions (sorted
    # pinned→updated_at desc); shipping every one was 147 KB / 391 rows on a
    # heavy archive, which dominated every poll AND the new-session path. Now:
    #   q=<term>  → server-side search across the FULL list (name/first_prompt)
    #   limit=N   → only the N most-recent (pinned already float to the top)
    #   ids=a,b,c → ALWAYS include these (the client's OPEN tabs) so the
    #               frontend's this.sessions.find(openTabId) never misses a tab
    #               that fell outside the recent window.
    # limit=0 (the default) preserves the old "return everything" behaviour for
    # any caller that doesn't opt in.
    full, list_revision = sess.list_sessions_snapshot()
    # A workspace switch only needs that workspace's recent sessions.  Filter
    # before applying q/limit/ids so an open-tab id owned by another workspace
    # can never be pulled into this response.  Legacy rows without cwd belong
    # to ROOT, matching sessions._build_sessions_list().
    if workspace_only:
        workspace_path = str(workspace_root)
        full = [
            s for s in full
            if str(s.get("cwd") or ROOT) == workspace_path
        ]
    total = len(full)
    q_norm = (q or "").strip().lower()
    if q_norm:
        subset = [
            s for s in full
            if (s.get("name") and q_norm in s["name"].lower())
            or (s.get("first_prompt") and q_norm in s["first_prompt"].lower())
        ][:200]
    elif limit and limit < total:
        subset = list(full[:limit])
        have = {s.get("id") for s in subset}
        keep = {x for x in (ids or "").split(",") if x}
        if keep:
            for s in full:
                sid = s.get("id")
                if sid in keep and sid not in have:
                    subset.append(s)
                    have.add(sid)
    else:
        subset = list(full)
    # FIX ⑩: server-authoritative "is this session streaming right now" flag so
    # the session-list blue dot syncs across devices. `_active_turns` is the
    # in-memory registry of live turns (set when a turn starts, popped/`.done`
    # when it finishes). The frontend's local `tabState[sid].streaming` only
    # knows about turns THIS browser kicked off — a turn started on phone A
    # left phone B's picker dot dark. Falling back to `s.active` fixes that.
    turn_active_sids = {
        sid for sid, bc in _active_turns.items()
        if bc is not None and not bc.done
    }
    # Background pins are lifecycle fences, not cache entries.  Read-only list
    # requests must never expire them: doing so could start a queued turn while
    # the SDK-owned command was still alive.  The owning watcher enforces the
    # absolute deadline and releases only after a terminal notification or a
    # confirmed CLI disconnect.
    background_active_sids = {
        sid for sid, task_ids in _sessions_with_inflight_tasks.items()
        if task_ids
    }
    with _sdk_cron_state_lock:
        scheduled_counts = {
            sid: len(jobs)
            for sid, jobs in _sdk_cron_jobs.items()
            if jobs
        }
    scheduled_active_sids = set(scheduled_counts)
    active_sids = turn_active_sids | background_active_sids
    # Copy each dict (never mutate the shared list_sessions() cache) + add the
    # live `active` flag. Only the returned subset is processed now, not all N.
    sessions = []
    for s in subset:
        s = dict(s)  # don't mutate cache
        s["active"] = s.get("id") in active_sids
        s["turn_active"] = s.get("id") in turn_active_sids
        s["background_active"] = s.get("id") in background_active_sids
        s["scheduled_active"] = s.get("id") in scheduled_active_sids
        s["scheduled_count"] = scheduled_counts.get(s.get("id"), 0)
        sessions.append(s)
    # Piggy-back orphan-attachments GC here — runs at most hourly. Cheaper
    # than a cron, and naturally fires whenever the UI is in use.
    global _last_orphan_gc_at
    now = time.time()
    if now - _last_orphan_gc_at > _ORPHAN_GC_INTERVAL_S:
        _last_orphan_gc_at = now
        try:
            _gc_orphan_attachments()
        except Exception:
            pass
    # Conditional GET: the picker polls /sessions on a timer; when nothing
    # changed (same titles, same updated_at, same `active` dots) we let the
    # client skip both the transfer AND the Alpine list re-render by returning
    # 304. The ETag is a weak validator (W/) because GZipMiddleware may re-encode
    # the body — weak comparison is all If-None-Match needs for GET anyway, and
    # the digest is over the UNcompressed JSON so it's stable across gzip on/off.
    # We hash the same payload we're about to send (including live `active`
    # flags), so any user-visible change flips the tag. default=str guards
    # stray datetime/Path values in session dicts.
    body = {"sessions": sessions, "total": total, "returned": len(sessions)}
    # ETag digest cache: hashing ~150KB of JSON on every poll adds up. The
    # body is fully determined by (list-cache generation, request params,
    # active turn set), so key on those and skip the dumps+md5 when nothing
    # changed. Any session mutation bumps the generation; a turn starting /
    # finishing changes active_sids; different limit/ids/q get their own key.
    _etag_key = (
        list_revision,
        limit,
        ids,
        q_norm,
        workspace_only,
        str(workspace_root) if workspace_only else "",
        frozenset(turn_active_sids),
        frozenset(background_active_sids),
        tuple(sorted(scheduled_counts.items())),
    )
    _hit = _LIST_ETAG_CACHE.get("v")
    if _hit is not None and _hit[0] == _etag_key:
        etag = _hit[1]
    else:
        try:
            # Keep the historical default validator byte-for-byte stable, but
            # salt filtered representations with their workspace.  Two empty
            # workspaces otherwise have identical bodies and could incorrectly
            # validate each other's header-driven representation.
            _etag_payload: Any = body
            if workspace_only:
                _etag_payload = {
                    "workspace": str(workspace_root),
                    "body": body,
                }
            _payload = json.dumps(_etag_payload, sort_keys=True, default=str,
                                  ensure_ascii=False).encode("utf-8")
            etag = 'W/"' + hashlib.md5(_payload).hexdigest() + '"'
        except (TypeError, ValueError):
            etag = ""
        if etag:
            _LIST_ETAG_CACHE["v"] = (_etag_key, etag)
    if workspace_only:
        response.headers["Vary"] = "X-Muselab-Workspace"
    if etag:
        # If-None-Match may carry a list ("tag1", "tag2") or "*". Weak-compare by
        # stripping the W/ prefix from both sides and matching the opaque value.
        inm = request.headers.get("if-none-match", "")
        if inm:
            def _bare(t: str) -> str:
                t = t.strip()
                return t[2:] if t.startswith("W/") else t
            wanted = _bare(etag)
            if any(_bare(p) == wanted for p in inm.split(",")):
                # 304 must echo the validator and carry no body.
                return Response(
                    status_code=304,
                    headers={
                        "ETag": etag,
                        **(
                            {"Vary": "X-Muselab-Workspace"}
                            if workspace_only
                            else {}
                        ),
                    },
                )
        response.headers["ETag"] = etag
    return body


@router.get("/sessions/{sid}/evidence", dependencies=[Depends(require_token)])
def get_session_evidence_api(sid: str) -> dict:
    """Return copyable, server-derived evidence for one known session.

    This endpoint exposes no caller-selected path and never creates an export.
    The transcript is the existing Claude CLI JSONL, validated against both the
    registered workspace and the known native/vendor session roots.
    """
    try:
        if str(uuid.UUID(sid)) != sid:
            raise ValueError
    except (ValueError, AttributeError, TypeError):
        raise HTTPException(400, "invalid session id") from None
    meta = sess.get_session_meta(sid)
    if meta is None:
        raise HTTPException(404, "session not found")
    workspace = sess.session_workspace(sid)
    transcript = _canonical_session_evidence_path(sid, workspace)
    if transcript is None:
        raise HTTPException(404, "canonical session transcript not found")
    evidence = {
        "session_name": str(meta.get("name") or ""),
        "session_id": sid,
        "transcript_path": str(transcript),
        "workspace": str(workspace),
        "model": str(meta.get("model") or ""),
    }
    sidecar = sess._sidecar_path(sid)
    try:
        resolved_sidecar = sidecar.resolve(strict=True)
        if (resolved_sidecar.is_file()
                and resolved_sidecar.parent == sess.SESS_DIR.resolve(strict=True)
                and resolved_sidecar.name == f"{sid}.sidecar.json"):
            evidence["sidecar_path"] = str(resolved_sidecar)
    except OSError:
        pass
    return evidence


def _canonical_available_model(model: str, groups: list[dict] | None = None) -> str:
    """Return the catalog id for an available model, accepting safe legacy aliases.

    Codex Gateway uses `codex:` as a muselab-internal routing prefix and strips it
    before calling the gateway. Some sessions / prefs may still carry the
    vendor-facing id (`gpt-5.5`). Map that alias back to `codex:gpt-5.5` so the
    backend routes through the configured Codex provider instead of treating it as
    an unknown Claude model.
    """
    wanted = (model or "").strip()
    if not wanted:
        return ""
    groups = endpoints.available_groups() if groups is None else groups
    available = {item["model"] for g in groups for item in g.get("items", [])}
    if wanted in available:
        return wanted
    if ":" not in wanted:
        codex_alias = f"codex:{wanted}"
        if codex_alias in available:
            return codex_alias
    return ""


def _resolve_default_model(requested: str = "", *, allow_fallback: bool = True) -> str:
    """Pick a model id for a new session. Three-tier fallback:
      1. `requested` (what the caller sent) — used ONLY if its provider
         is actually configured. Otherwise we silently swap; honoring an
         unusable preference would 401 on the first send.
      2. `MUSELAB_MODEL` env (settings.MODEL) — same availability check.
      3. First model from the first available_groups() entry — covers
         "user configured only DEEPSEEK_API_KEY" cases on fresh installs.

    When NO provider at all is configured:
      - allow_fallback=True  → return `MODEL` (the constant, likely
        claude-sonnet-4-6). Legacy callers that need a non-empty id.
      - allow_fallback=False → return "" (empty). Used by session creation
        so a session born before any provider is set up does NOT get locked
        to an unreachable claude fallback — the lock is what made every
        later send 401 forever. The model is resolved lazily on first send
        (by which point the user has been gated into configuring one).
    """
    groups = endpoints.available_groups()
    # 1. Caller-requested model, if its provider is wired.
    resolved = _canonical_available_model(requested, groups)
    if resolved:
        return resolved

    # 2. Env-pinned default, if its provider is wired.
    explicit = (MODEL or "").strip()
    resolved = _canonical_available_model(explicit, groups)
    if resolved:
        return resolved

    # 3. First actually-available model.
    if groups and groups[0].get("items"):
        return groups[0]["items"][0]["model"]

    # 4. Nothing configured. Either fall back to the constant (legacy
    # callers) or return empty so the caller can leave the session model
    # unlocked until a provider exists. UI gates chat behind the
    # no-provider onboarding card either way.
    return MODEL if allow_fallback else ""


def _heal_unreachable_locked_model(session_id: str, locked: str, requested: str = "") -> str:
    """Decide the model for a send on a session already locked to `locked`.

    The one-session-one-model rule (in _start_turn) normally makes the
    locked model win over the frontend dropdown. But a session created
    BEFORE any provider was configured gets pinned to the MODEL fallback
    (claude-sonnet-4-6); once the user configures e.g. DeepSeek, that lock
    would make EVERY send fail the Anthropic auth pre-check forever — the
    exact "I only configured DeepSeek but still got a claude auth error"
    bug, because the broken session predates the provider.

    Return a re-resolved, reachable model ONLY when BOTH hold:
      (a) the locked model's provider isn't currently configured, AND
      (b) the session has no on-disk JSONL yet — it never actually ran a
          turn, so there's no prior-vendor thinking signature that a vendor
          switch could corrupt.
    Otherwise return `locked` unchanged. A session with real history stays
    locked: silently swapping vendors mid-conversation is precisely the risk
    the one-session-one-model rule exists to prevent.
    """
    groups = endpoints.available_groups()
    available = {item["model"] for g in groups for item in g.get("items", [])}
    # Locked model still reachable (or is a safe legacy alias such as
    # `gpt-5.5` → `codex:gpt-5.5`) → keep/canonicalize it. Canonicalizing a
    # Codex alias is not a vendor switch; it restores the internal routing tag.
    canonical_locked = _canonical_available_model(locked, groups)
    if canonical_locked:
        return canonical_locked
    # Nothing configured at all → can't do better; the no-provider onboarding
    # card handles that case.
    if not available:
        return locked
    # Don't touch a session that has actually run — switching vendors on real
    # history can corrupt cross-vendor thinking signatures.
    try:
        has_history = _find_session_jsonl(session_id) is not None
    except Exception:
        has_history = False
    if has_history:
        return locked
    return _resolve_default_model(requested)


@router.post("/sessions", dependencies=[Depends(require_token)])
def create_session_api(req: CreateReq) -> dict:
    # allow_fallback=False: if no provider is configured, leave the session
    # model EMPTY rather than locking it to the claude constant. A locked
    # unreachable model is exactly what made fresh-install first sessions
    # 401 forever; the frontend gates chat until a provider exists, and the
    # model is resolved on first send.
    resolved_model = _resolve_default_model(req.model, allow_fallback=False)
    resolved_permission = _validate_permission(req.permission)
    client_id = (req.id or "").strip()
    if client_id:
        # Optimistic-create path: the client minted this UUID and already
        # opened the tab. Validate it STRICTLY before register_session writes
        # SESS_DIR/{id}.sidecar.json — a non-UUID id would be a path-injection
        # vector. uuid.UUID() rejects garbage; the canonical-form re-check
        # rejects braces / urn: prefixes / anything that isn't a clean
        # 36-char hyphenated v4 string, so the id is guaranteed [0-9a-f-] only.
        try:
            parsed = uuid.UUID(client_id)
        except (ValueError, AttributeError, TypeError):
            raise HTTPException(400, "invalid session id")
        if str(parsed) != client_id.lower():
            raise HTTPException(400, "invalid session id")
        # register_session is idempotent (returns the existing row if the id is
        # already registered) so a client retry / keepalive resend is safe.
        try:
            meta = sess.register_session(
                client_id,
                name=req.name or "",
                model=resolved_model,
                permission=resolved_permission,
                auto_named=True,
                cwd=req.cwd or None,
                activity_hidden=req.activity_hidden,
                runtime_profile=req.runtime_profile,
            )
        except ValueError as exc:
            raise HTTPException(400, str(exc)) from exc
    else:
        try:
            meta = sess.create_session(
                name=req.name or "",
                model=resolved_model,
                permission=resolved_permission,
                cwd=req.cwd or None,
                activity_hidden=req.activity_hidden,
                runtime_profile=req.runtime_profile,
            )
        except ValueError as exc:
            raise HTTPException(400, str(exc)) from exc
    # Auto-prune (B): recycle blank scratch sessions left over from previous
    # tabs / accidental new-session clicks. keep_ids protects BOTH the session
    # we just created AND every tab the client currently has open — so a blank
    # session the user is about to type in is never yanked out from under them.
    # Still gated by all of prune_empty_sessions' own safety checks (0 messages,
    # not pinned, auto-named, <2h old) + the MUSELAB_PRUNE_EMPTY_SESSIONS flag.
    sess.prune_empty_sessions(keep_ids=[meta["id"], *(req.open_ids or [])])
    return meta


@router.post("/sessions/organize", dependencies=[Depends(require_token)])
def create_organize_session_api(req: CreateReq | None = None) -> dict:
    """Create a normal SDK session and return a self-contained organizer starter.

    The starter directs the agent through a read-only scan, proposal, explicit
    confirmation, and execution of approved changes. This endpoint deliberately
    creates no files or directories: organizing a workspace must not opt the
    user into a personal-profile template or a predefined directory taxonomy.
    No custom system prompt is attached to the session.

    Returns session metadata + an initial_message the frontend should
    auto-send to kick off the workflow. See backend/prompts.py."""
    from .prompts import CURATOR_INITIAL_MESSAGE
    import datetime as _dt
    try:
        workspace = workspace_registry.resolve(req.cwd if req else None)
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    # Locale-aware default label, expressed in workspace rather than personal
    # archive terms.
    _label = "[整理工作区] " if is_chinese_locale() else "[Organize workspace] "
    name = (req.name if req else None) or (
        _label + _dt.datetime.now().strftime("%m-%d %H:%M"))
    model = (req.model if req else None) or MODEL
    meta = sess.create_session(
        name=name,
        model=model,
        cwd=workspace,
    )
    return {**meta, "initial_message": CURATOR_INITIAL_MESSAGE}


@router.post("/sessions/profile-intake", dependencies=[Depends(require_token)])
def create_profile_intake_session_api(req: CreateReq | None = None) -> dict:
    """Deprecated compatibility forward to the generic organizer.

    New callers should use ``/sessions/organize``. This route intentionally
    performs no profile intake and creates no ``CLAUDE.md`` or personal-data
    directory; it only preserves old bookmarks and API clients.
    """
    return create_organize_session_api(req)


def _extract_searchable_text(content: Any) -> str:
    """Extract plain text from a JSONL message.content field for search.
    Handles both string content and list-of-blocks. Skips tool_use /
    tool_result blocks because their inputs/outputs are usually noisy
    JSON and not what users mean when they search."""
    if isinstance(content, str):
        return content
    if not isinstance(content, list):
        return ""
    parts: list[str] = []
    for block in content:
        if not isinstance(block, dict):
            continue
        btype = block.get("type")
        if btype == "text":
            t = block.get("text")
            if isinstance(t, str):
                parts.append(t)
        elif btype == "thinking":
            t = block.get("thinking")
            if isinstance(t, str):
                parts.append(t)
    return "\n".join(parts)


def _make_snippet(text: str, idx: int, qlen: int, *,
                   ctx: int = 60, max_len: int = 200) -> str:
    """Build a search-result snippet centered on a match. Caller passes the
    match position so we don't have to find() twice. Result is capped at
    max_len chars with leading/trailing ellipses if truncated."""
    start = max(0, idx - ctx)
    end = min(len(text), idx + qlen + ctx)
    snippet = text[start:end]
    if start > 0:
        snippet = "…" + snippet
    if end < len(text):
        snippet = snippet + "…"
    # Collapse whitespace runs so multi-line transcripts render compactly
    # in the search result list.
    snippet = re.sub(r"\s+", " ", snippet).strip()
    if len(snippet) > max_len:
        snippet = snippet[:max_len - 1] + "…"
    return snippet


@router.get("/search", dependencies=[Depends(require_token)])
def search_sessions_api(q: str = Query(default="", min_length=0, max_length=200),
                         limit: int = Query(default=30, ge=1, le=100)) -> dict:
    """Cross-session full-text search. Scans CLI JSONL files for user /
    assistant text matching `q` (case-insensitive substring). Returns
    hits sorted by timestamp desc. Each hit:
        {sid, name, uuid, role, snippet, ts}
    Implementation: line-by-line JSON parse of every JSONL under the
    project's CLI directory. For ~200 sessions of typical size (< 1MB
    each) this runs in <500ms — switch to SQLite FTS5 if it grows."""
    query = q.strip()
    if not query:
        return {"hits": [], "total": 0}
    qlower = query.lower()
    # Walk every registered workspace's CLI project directory across both the
    # default and vendor-isolated roots.  Include historical child-cwd folders
    # too (encoded-workspace + "-") just like the cost dashboard does.
    encoded_roots = tuple(
        _cli_encode_cwd(str(root)) for root in workspace_registry.paths())
    proj_dirs: list[Path] = []
    seen_dirs: set[Path] = set()
    for projects_root in _cli_project_roots():
        try:
            candidates = projects_root.iterdir()
        except OSError:
            continue
        for candidate in candidates:
            name = candidate.name
            if not candidate.is_dir() or not any(
                name == encoded or name.startswith(encoded + "-")
                for encoded in encoded_roots
            ):
                continue
            if candidate not in seen_dirs:
                seen_dirs.add(candidate)
                proj_dirs.append(candidate)
    if not proj_dirs:
        return {"hits": [], "total": 0}

    name_map = {s["id"]: s.get("name", "") for s in sess.list_sessions()}

    hits: list[dict] = []
    PER_SESSION_CAP = 5   # avoid one chatty session swamping results
    # Iterate JSONLs across both roots. A given sid only lives in one root
    # at a time (vendor vs Claude is mutually exclusive per session), so
    # PER_SESSION_CAP keyed by stem still applies cleanly.
    jsonl_paths = [p for d in proj_dirs for p in d.glob("*.jsonl")]
    for jsonl in jsonl_paths:
        sid = jsonl.stem
        per_sess = 0
        try:
            # utf-8-sig strips a leading BOM so JSONL writers that emit
            # U+FEFF at the start (some CLI versions did, briefly) don't
            # poison the "fast reject" qlower-in-line check at the start
            # of every line — `"﻿{...}".lower()` would mismatch a
            # qlower hitting the literal first chars.
            with jsonl.open("r", encoding="utf-8-sig") as f:
                for line in f:
                    if qlower not in line.lower():
                        continue   # fast reject before JSON parse
                    try:
                        entry = json.loads(line)
                    except (json.JSONDecodeError, ValueError):
                        continue
                    if entry.get("type") not in ("user", "assistant"):
                        continue
                    msg = entry.get("message") or {}
                    text = _extract_searchable_text(msg.get("content"))
                    if not text:
                        continue
                    # CLI's slash-command wrapper round-trips as user
                    # text — strip before matching so e.g. searching
                    # "compact" doesn't surface every /compact invocation.
                    text = _strip_cli_slash_wrapper(text) or text
                    pos = text.lower().find(qlower)
                    if pos < 0:
                        continue
                    hits.append({
                        "sid": sid,
                        "name": name_map.get(sid, ""),
                        "uuid": entry.get("uuid", ""),
                        "role": entry.get("type"),
                        "snippet": _make_snippet(text, pos, len(query)),
                        "ts": entry.get("timestamp", ""),
                    })
                    per_sess += 1
                    if per_sess >= PER_SESSION_CAP:
                        break
        except OSError:
            continue

    hits.sort(key=lambda h: h["ts"], reverse=True)
    return {"hits": hits[:limit], "total": len(hits)}


# CLI wraps slash commands as pseudo-user messages with these tags so it can
# round-trip through the conversation log. They're internal protocol detail
# and should never reach the user's chat UI as a regular bubble.


def _strip_cli_slash_wrapper(text: str) -> str:
    """Patchable facade for CLI slash-command presentation cleanup."""
    return chat_presentation.strip_cli_slash_wrapper(text)


# A run_in_background task's completion round-trips through the conversation log
# as a plain user-role message whose ENTIRE content is a <task-notification> XML
# block (the SDK injects it when the task settles — see docs/background-tasks-
# spec.md). The launching tool_use card and this record share the <tool-use-id>.
# On history rebuild we parse it and stamp the card's terminal task_status so a
# completed bg task shows ✅ DURABLY (survives reload — matches Claude Code),
# instead of rendering the raw XML as a confusing user bubble.
#
# LIVE-PATH ROLE (updated 2026-06-11, Phase-0 probe on CLI 2.1.141 + SDK
# 0.2.95): the live stream DOES deliver a typed TaskNotificationMessage
# (out-of-band, after the turn's ResultMessage) — typed dispatch is the
# PRIMARY live truth. This regex remains authoritative ONLY for JSONL history
# rebuild (the transcript stores the XML record, never the typed message) and
# as a live fallback for older CLIs; a fallback hit logs a
# "[chat] task fallback" warning.


def _parse_task_notifications(text: str) -> list[dict]:
    """Patchable facade for canonical task-notification parsing."""
    return chat_presentation.parse_task_notifications(text)


# FALLBACK launch sniff (updated 2026-06-11): on CLI 2.1.141 + SDK 0.2.95 a
# Bash run_in_background=true launch DOES emit a typed TaskStartedMessage
# (Phase-0 probe), which arrives BEFORE the tool_result — typed dispatch is the
# primary truth and this sniff normally no-ops. It remains as a fallback for
# older CLIs whose launches surface solely as a tool_result body of the form:
# "Command running in background with ID: <tid>. Output is being written to:
# <file>. ...". If the sniff ever wins (logs a "[chat] task fallback" warning),
# the typed contract regressed — without the fallback, inflight_tasks would
# stay empty, the turn-end cross-turn watcher would never spawn, and the
# post-completion auto-continue would never stream live. NOTE: the English
# wording below is CLI-version-coupled; that brittleness is exactly why typed
# messages are now the primary path. See docs/background-tasks-spec.md.


def _parse_bg_launch(text: str) -> dict | None:
    """Patchable facade for legacy background-launch parsing."""
    return chat_presentation.parse_bg_launch(text)


def _usermsg_task_notification_text(msg) -> str:
    """Patchable facade for fallback UserMessage notification extraction."""
    return chat_presentation.usermsg_task_notification_text(
        msg, user_message_type=UserMessage, text_block_type=TextBlock)


def _sdk_messages_to_ui(
    sm_list: list,
    annotations: dict[str, dict],
    compact_uuids: set[str] | None = None,
    *,
    defer_large_bodies: bool = False,
) -> list[dict]:
    """Patchable facade for canonical transcript-to-UI shaping."""
    return chat_presentation.sdk_messages_to_ui(
        sm_list,
        annotations,
        compact_uuids,
        defer_large_bodies=defer_large_bodies,
        is_cli_interrupt_message=_is_cli_interrupt_message,
        slim_input_fields=_SLIM_INPUT_FIELDS,
        slim_value=_slim_input_value,
        summarize_input=_summarize_tool_input,
        parse_bash=_parse_bash_result,
        tool_result_preview_cap=_TOOL_RESULT_PREVIEW_CAP,
        defer_bodies=_defer_large_ui_bodies,
    )


def _apply_runtime_task_overlays(
    messages: list[dict], overlays: dict[str, dict],
) -> None:
    """Patch copied tool cards with predecessor-owned task state, in place.

    The overlay is presentation-only sidecar data.  It is applied after SDK
    transcript shaping and is never included in a later model query.
    """
    if not overlays:
        return
    by_tool_use = {
        str(overlay.get("tool_use_id") or ""): overlay
        for overlay in overlays.values()
        if isinstance(overlay, dict) and overlay.get("tool_use_id")
    }
    for message in messages:
        if message.get("role") != "tool_use":
            continue
        current = message.get("task_status") or {}
        task_id = str(current.get("task_id") or "")
        overlay = overlays.get(task_id) if task_id else None
        if overlay is None:
            overlay = by_tool_use.get(str(message.get("id") or ""))
        if not isinstance(overlay, dict):
            continue
        message["task_status"] = {
            **current,
            **overlay,
            "task_id": str(overlay.get("task_id") or task_id),
        }


def _transcript_index_path(sid: str) -> Path:
    return sess.SESS_DIR / f"{sid}.transcript-index.json"


def _describe_transcript_record(entry: dict) -> dict:
    """Build index metadata with the same patchable UI renderer facade."""
    return chat_presentation.describe_transcript_record(
        entry,
        raw_message_factory=_RawMsg,
        raw_entry_factory=_raw_msg_from_entry,
        render_messages=_sdk_messages_to_ui,
        is_real_user_prompt=_is_real_user_prompt,
    )


def _ensure_transcript_index(sid: str) -> tuple[Path, dict] | None:
    # Index construction runs in worker threads and cannot be interrupted by
    # cancelling the awaiting asyncio task. Linearize its final disk write with
    # explicit session deletion: if indexing wins, DELETE waits and then
    # removes the file; if the tombstone wins, this worker becomes a no-op.
    # Holding the lifecycle stripe across ensure_index is intentional because
    # checking only before its atomic write still leaves a check/write race.
    with sess.session_lifecycle_lock(sid):
        if sess.session_is_deleting(sid):
            return None
        path = _find_session_jsonl(sid)
        if path is None:
            return None
        return path, transcript_idx.ensure_index(
            sid, path, _transcript_index_path(sid), _describe_transcript_record)


def _indexed_turn_context(
    transcript_path: Path,
    index: dict,
    record_indices: list[int],
) -> dict[str, tuple[str, int | None]]:
    """Map selected record UUIDs to their logical turn origin and start time.

    A visible assistant run may span several JSONL records (tool results), and
    a background-task continuation starts at a zero-bubble task-notification
    record.  A bounded tail read normally does not read that hidden ancestor,
    so deriving footer duration from visible bubbles alone either loses the
    value or accidentally borrows metadata from the launch turn.  The index
    already stores parent links plus prompt/notification descriptors; walk
    those cheap records and seek only the few unique origin lines for time.
    """
    records = index.get("records") or []
    if not records or not record_indices:
        return {}
    by_uuid: dict[str, int] = {}
    for record_i, record in enumerate(records):
        uid = str(record.get("uuid") or "")
        if uid:
            by_uuid[uid] = record_i

    origin_by_uuid: dict[str, str] = {}
    origin_indices: set[int] = set()
    for selected_i in record_indices:
        if selected_i < 0 or selected_i >= len(records):
            continue
        selected = records[selected_i]
        selected_uuid = str(
            selected.get("presentation_uuid")
            or selected.get("uuid")
            or "")
        current_i = selected_i
        origin_i = selected_i
        seen: set[str] = set()
        while 0 <= current_i < len(records):
            current = records[current_i]
            current_uuid = str(current.get("uuid") or "")
            if current_uuid in seen:
                break
            if current_uuid:
                seen.add(current_uuid)
            origin_i = current_i
            if (current.get("real_user_prompt")
                    or current.get("task_notifications")):
                break
            parent = str(current.get("parent") or "")
            parent_i = by_uuid.get(parent) if parent else None
            if parent_i is None:
                break
            current_i = parent_i
        origin_uuid = str(records[origin_i].get("uuid") or selected_uuid)
        if selected_uuid:
            origin_by_uuid[selected_uuid] = origin_uuid
        origin_indices.add(origin_i)

    origin_times: dict[str, int | None] = {}
    for entry in transcript_idx.read_records(
            transcript_path, index, sorted(origin_indices)):
        origin_times[str(entry.get("uuid") or "")] = _transcript_ts_ms(entry)
    return {
        uid: (origin_uuid, origin_times.get(origin_uuid))
        for uid, origin_uuid in origin_by_uuid.items()
    }


def _indexed_ui_records(
    transcript_path: Path,
    index: dict,
    record_indices: list[int],
    annotations: dict[str, dict],
    *,
    defer_large_bodies: bool = True,
) -> list[dict]:
    """Seek/read and shape only selected records from a transcript index."""
    entries = transcript_idx.read_records(transcript_path, index, record_indices)
    turn_context = _indexed_turn_context(
        transcript_path, index, record_indices)
    projected = [
        (entry, _raw_msg_from_entry(entry))
        for entry in entries
    ]
    raw = [msg for _, msg in projected if msg is not None]
    compact = {
        str(msg.uuid)
        for entry, msg in projected
        if msg is not None and entry.get("isCompactSummary")
    }
    bubbles = _sdk_messages_to_ui(
        raw, annotations, compact,
        defer_large_bodies=defer_large_bodies)

    # Cross-window context is stored in the index: a page may begin with a
    # tool_result whose tool_use is outside the read window, and a launching
    # card may need a terminal task notification appended much later.
    tool_names = index.get("tool_use_names") or {}
    task_status = index.get("task_status") or {}
    for bubble in bubbles:
        uid = str(bubble.get("uuid") or "")
        context = turn_context.get(uid)
        if context is not None:
            origin_uuid, started_at_ms = context
            if origin_uuid:
                bubble["_turn_origin_uuid"] = origin_uuid
            if started_at_ms is not None:
                bubble["turn_started_at"] = started_at_ms
        tool_id = str(bubble.get("id") or "")
        if bubble.get("role") == "tool_result" and tool_id:
            tool_name = tool_names.get(tool_id) or ""
            if tool_name:
                bubble["tool_name"] = tool_name
                if tool_name == "Bash" and "bash" not in bubble:
                    parsed = _parse_bash_result(bubble.get("text") or "")
                    if parsed:
                        bubble["bash"] = parsed
        elif bubble.get("role") == "tool_use" and tool_id in task_status:
            status = task_status[tool_id]
            bubble["task_status"] = (
                normalize_task_summary_fields(status)
                if isinstance(status, dict) else status
            )

    # A bounded tail can begin after the AssistantMessage that owns the
    # completion annotation while still containing the actual visual tail
    # (for example 100 tool-result rows after one assistant tool-use record).
    # Pull the footer donor through the index, but only when this window
    # contains that logical turn's true final visible record. This preserves
    # memory/status on cold reload without inventing a footer in the middle of
    # an older paged window.
    records = index.get("records") or []
    selected_ids = set(record_indices)
    by_uuid = {
        str(record.get("uuid") or ""): record_i
        for record_i, record in enumerate(records)
        if record.get("uuid")
    }
    origin_cache: dict[int, str] = {}
    origin_resolving: set[int] = set()

    def logical_origin(record_i: int) -> str:
        cached = origin_cache.get(record_i)
        if cached is not None:
            return cached
        record = records[record_i]
        uid = str(record.get("uuid") or "")
        if record_i in origin_resolving:
            return uid
        if record.get("real_user_prompt") or record.get("task_notifications"):
            origin_cache[record_i] = uid
            return uid
        origin_resolving.add(record_i)
        parent_i = by_uuid.get(str(record.get("parent") or ""))
        origin = logical_origin(parent_i) if parent_i is not None else uid
        origin_resolving.discard(record_i)
        origin_cache[record_i] = origin
        return origin

    orders = index.get("orders") or {}
    normal_order = list(orders.get("normal") or [])
    chosen_order = (
        normal_order
        if selected_ids.issubset(set(normal_order))
        else list(orders.get("full") or [])
    )
    terminal_record: dict[str, int] = {}
    donors: dict[str, dict[str, Any]] = {}
    for record_i in chosen_order:
        if record_i < 0 or record_i >= len(records):
            continue
        record = records[record_i]
        origin = logical_origin(record_i)
        if int(record.get("bubble_count") or 0) > 0:
            terminal_record[origin] = record_i
        ann = annotations.get(str(record.get("uuid") or "")) or {}
        if ann:
            donor = donors.setdefault(origin, {})
            for source, target in (
                ("ts", "ts"),
                ("elapsed_s", "elapsed"),
                ("model", "model"),
                ("turn_status", "turn_status"),
                ("terminal_reason", "terminal_reason"),
                ("turn_origin", "turn_origin"),
                ("turn_id", "turn_id"),
                ("model_usage", "model_usage"),
                ("memory_recall", "memoryRecall"),
            ):
                value = ann.get(source)
                if value is not None and value != "":
                    donor[target] = value
    tail_by_origin: dict[str, dict] = {}
    for bubble in bubbles:
        origin = str(bubble.get("_turn_origin_uuid") or "")
        if origin:
            tail_by_origin[origin] = bubble
    for origin, tail in tail_by_origin.items():
        if terminal_record.get(origin) not in selected_ids:
            continue
        for field, value in (donors.get(origin) or {}).items():
            if tail.get(field) is None or tail.get(field) == "":
                tail[field] = value
    return bubbles


def _complete_turn_footer_metadata(
    messages: list[dict],
    session_model: str,
    *,
    has_later: bool,
    active_turn: "TurnBroadcast | None" = None,
) -> None:
    """Patchable facade for footer-donor presentation shaping."""
    chat_presentation.complete_turn_footer_metadata(
        messages,
        session_model,
        has_later=has_later,
        active_turn=active_turn,
        now=time.time,
    )


def _persistable_memory_recall(trace: Any) -> dict | None:
    """Return the privacy-minimal recall receipt safe for a session sidecar.

    Memory contents already live in the private memory registry.  Persist only
    stable IDs and non-content diagnostics here; the authenticated UI hydrates
    item details on demand.  This avoids creating a second long-lived copy of
    recalled personal text in every conversation annotation file.
    """
    if not isinstance(trace, dict):
        return None
    out = {
        key: trace.get(key)
        for key in ("id", "count", "latency_ms", "status")
        if trace.get(key) is not None
    }
    items = []
    for item in trace.get("items") or []:
        if not isinstance(item, dict) or not item.get("id"):
            continue
        items.append({
            key: item.get(key)
            for key in ("id", "kind", "score")
            if item.get(key) is not None
        })
    if items:
        out["items"] = items
    return out or None


def _bind_pending_attachments(sid: str, messages: list[dict]) -> None:
    """For every user message that has image refs (only mime, no thumb/url)
    but no annotation yet, pop one entry off the sidecar's pending list
    and bind it. Runs in-order so multi-image conversations stay aligned.

    Called by GET /sessions/{sid} after _sdk_messages_to_ui. Modifies
    messages in place."""
    for entry in messages:
        if entry.get("role") != "user":
            continue
        imgs = entry.get("images") or []
        if not imgs:
            continue
        # Already has thumb / url for at least one — already bound, skip.
        if any(im.get("thumb") or im.get("url") for im in imgs):
            continue
        uuid = entry.get("uuid")
        if not uuid:
            continue
        bound = sess.consume_one_pending_attachments(sid, uuid)
        if bound and bound.get("images"):
            entry["images"] = bound["images"]
        if bound and bound.get("docs"):
            entry["docs"] = bound["docs"]


def _summarize_tool_input(name: str | None, inp: dict) -> str:
    """Patchable facade for historic tool-input summaries."""
    return chat_presentation.summarize_tool_input(name, inp)


# Cache of compact-summary UUID scans keyed by sid → (mtime, size, uuids).
# get_session_api re-runs this raw full-file scan on EVERY call, and the
# client makes several calls per session via windowing (?tail then ?offset
# "load earlier") — all against the same unchanged JSONL. Keying on
# (mtime, size) lets paging / re-opens skip the rescan; any appended turn
# changes the size (and usually mtime), invalidating the entry. Values are
# tiny (a set of summary UUIDs, normally 0–few), so the cap is generous.
_COMPACT_UUIDS_CACHE: dict[str, tuple[float, int, set[str]]] = {}
_COMPACT_UUIDS_CACHE_MAX = 256


def _compact_summary_uuids(sid: str) -> set[str]:
    """Scan raw CLI JSONL for entries with isCompactSummary:true and return
    their UUIDs. SDK get_session_messages strips this flag, so to render a
    "📦 已压缩" indicator we have to detect it ourselves at the JSONL level.

    Result is cached per (sid, mtime, size) — the returned set is treated as
    read-only by all callers (membership tests only), so sharing it is safe.

    Glob-based JSONL lookup via _find_session_jsonl — covers both default
    and vendor-isolated roots so vendor sessions keep their compact
    markers too."""
    try:
        jsonl_path = _find_session_jsonl(sid)
        if jsonl_path is None:
            return set()
        sig: tuple[float, int] | None
        try:
            _st = jsonl_path.stat()
            sig = (_st.st_mtime, _st.st_size)
        except OSError:
            sig = None
        if sig is not None:
            _cached = _COMPACT_UUIDS_CACHE.get(sid)
            if _cached is not None and _cached[0] == sig[0] and _cached[1] == sig[1]:
                return _cached[2]
        uuids: set[str] = set()
        with jsonl_path.open("r", encoding="utf-8") as f:
            for line in f:
                if '"isCompactSummary":true' not in line and '"isCompactSummary": true' not in line:
                    continue
                try:
                    entry = json.loads(line)
                except (json.JSONDecodeError, ValueError):
                    continue
                if entry.get("isCompactSummary") and entry.get("uuid"):
                    uuids.add(entry["uuid"])
        if sig is not None:
            # Bound the cache: drop the oldest insertion (dicts preserve
            # insertion order) when a NEW sid would overflow the cap.
            if (len(_COMPACT_UUIDS_CACHE) >= _COMPACT_UUIDS_CACHE_MAX
                    and sid not in _COMPACT_UUIDS_CACHE):
                _COMPACT_UUIDS_CACHE.pop(next(iter(_COMPACT_UUIDS_CACHE)), None)
            _COMPACT_UUIDS_CACHE[sid] = (sig[0], sig[1], uuids)
        return uuids
    except Exception:
        return set()


# Cache of PARSED session-message lists keyed by (sid, full) → (mtime, size,
# msgs). get_session_api re-parses the ENTIRE JSONL (SDK get_session_messages
# + the full-history reader) on EVERY call, and the client makes several calls
# per session open via windowing (?tail for the initial paint, then
# ?offset&limit per "load earlier" click) — all against the same unchanged
# file. For a multi-thousand-message session that full parse is the dominant
# multi-second cost (it holds the GIL while it runs, so it also stalls every
# other concurrent request → the "整个卡住" on open/switch). Keying on
# (mtime, size) lets paging / re-opens / idle-preloads reuse one parse; any
# appended turn changes the size (and usually mtime), invalidating the entry.
# The annotation merge + compact-UUID scan + windowing slice still run live on
# top of the cached parse, so cost / compact freshness is unaffected. Parsed
# lists can be several MB each, so the cap is deliberately small.
_SESSION_MSGS_CACHE: dict[tuple[str, bool], tuple[float, int, list]] = {}
_SESSION_MSGS_CACHE_MAX = 16
# Also cap by total SOURCE bytes of cached transcripts. The file size (sig[1],
# already known from stat) is a cheap proxy for parsed-list memory (≈ 2–4× this
# after Python object overhead). Without a byte bound a single 100MB+ agentic
# session could sit alongside up to 15 others and push RSS into the GBs — the
# count cap alone doesn't bound memory when one entry is pathologically large.
_SESSION_MSGS_CACHE_BYTE_BUDGET = 64 * 1024 * 1024  # 64 MiB of source JSONL


# Cache of SHAPED UI message lists keyed by (sid, full) → (jsonl_sig,
# sidecar_sig, messages). _cached_session_msgs already skips the JSONL
# re-parse, but get_session_api still re-ran _sdk_messages_to_ui (O(N) over
# every content block) + the annotation merge on EVERY windowed call (?tail
# then each ?offset "load earlier") against unchanged inputs. Keying on BOTH
# the transcript signature and the sidecar signature means any new turn,
# annotation write, or attachment bind invalidates the entry; the windowing
# slice still runs live on top. _bind_pending_attachments runs on the cached
# list too — it's idempotent (skips already-bound messages) and a successful
# bind writes the sidecar, which changes sidecar_sig and forces a fresh
# shape on the next call.
_UI_MSGS_CACHE: dict[tuple[str, bool], tuple[
    tuple[float, int] | None, tuple[float, int] | None, list]] = {}
_UI_MSGS_CACHE_MAX = 8


def _jsonl_signature(sid: str) -> tuple[float, int] | None:
    """Compatibility wrapper for canonical transcript freshness."""
    return chat_history.jsonl_signature(
        sid, find_session_jsonl=_find_session_jsonl)


def _shaped_ui_messages(
    sid: str,
    model: str,
    full: bool,
    *,
    defer_large_bodies: bool = True,
) -> list[dict]:
    """Shape SDK messages into UI bubbles with a freshness-checked cache.

    Falls back to live shaping whenever either signature is unavailable
    (no transcript yet / stat failure) so correctness never depends on the
    cache. The returned list is shared across calls — callers must treat
    it as read-only EXCEPT _bind_pending_attachments (idempotent, and its
    sidecar write invalidates this cache)."""
    jsig = _jsonl_signature(sid)
    ssig = sess.sidecar_signature(sid)
    key = (sid, full)
    if jsig is not None and defer_large_bodies:
        hit = _UI_MSGS_CACHE.get(key)
        if hit is not None and hit[0] == jsig and hit[1] == ssig:
            return hit[2]
    annotations = sess.get_message_annotations(sid)
    messages: list[dict]
    try:
        indexed = _ensure_transcript_index(sid)
    except Exception:
        indexed = None
    if indexed is not None:
        transcript_path, index = indexed
        order = "full" if full else "normal"
        messages = _indexed_ui_records(
            transcript_path,
            index,
            list((index.get("orders") or {}).get(order) or []),
            annotations,
            defer_large_bodies=defer_large_bodies,
        )
    else:
        try:
            sdk_msgs = _cached_session_msgs(sid, model, full)
        except Exception:
            sdk_msgs = []
        compact_uuids = _compact_summary_uuids(sid)
        messages = _sdk_messages_to_ui(
            sdk_msgs,
            annotations,
            compact_uuids,
            defer_large_bodies=defer_large_bodies,
        )
    if jsig is not None and defer_large_bodies:
        _UI_MSGS_CACHE.pop(key, None)
        _UI_MSGS_CACHE[key] = (jsig, ssig, messages)
        while len(_UI_MSGS_CACHE) > _UI_MSGS_CACHE_MAX:
            oldest = next(iter(_UI_MSGS_CACHE))
            if oldest == key:
                break
            _UI_MSGS_CACHE.pop(oldest, None)
    return messages


def _cached_session_msgs(sid: str, model: str, full: bool) -> list:
    """Parse session messages with a (mtime, size)-keyed cache so repeated
    reads of an UNCHANGED JSONL (windowed paging, re-opens, idle preload) skip
    the expensive full parse. Returns the same SDK/_RawMsg objects the
    uncached path would — they're consumed read-only by _sdk_messages_to_ui,
    so sharing a cached list across callers is safe. Falls back to a live
    parse whenever the file can't be stat'd (so correctness never depends on
    the cache)."""
    jsonl_path = _find_session_jsonl(sid)
    sig: tuple[float, int] | None = None
    if jsonl_path is not None:
        try:
            _st = jsonl_path.stat()
            sig = (_st.st_mtime, _st.st_size)
        except OSError:
            sig = None
    key = (sid, full)
    if sig is not None:
        _cached = _SESSION_MSGS_CACHE.get(key)
        if _cached is not None and _cached[0] == sig[0] and _cached[1] == sig[1]:
            return _cached[2]
    msgs = _full_session_msgs(sid) if full else _get_session_msgs(sid, model)
    if sig is not None:
        # Insert at the end (newest). pop-then-set so a re-read of a GROWN
        # file (new mtime/size) moves the entry to the back of the FIFO rather
        # than keeping its stale position.
        _SESSION_MSGS_CACHE.pop(key, None)
        _SESSION_MSGS_CACHE[key] = (sig[0], sig[1], msgs)
        # Evict oldest (insertion order) until BOTH the count cap and the
        # source-byte budget hold. Never evict the entry we just inserted
        # (so an oversized lone session still caches — we just don't let it
        # coexist with others past the budget).
        while len(_SESSION_MSGS_CACHE) > 1 and (
            len(_SESSION_MSGS_CACHE) > _SESSION_MSGS_CACHE_MAX
            or sum(v[1] for v in _SESSION_MSGS_CACHE.values())
            > _SESSION_MSGS_CACHE_BYTE_BUDGET
        ):
            oldest = next(iter(_SESSION_MSGS_CACHE))
            if oldest == key:
                break
            _SESSION_MSGS_CACHE.pop(oldest, None)
    return msgs


def _session_message_uuids(sid: str, model: str) -> frozenset[str]:
    """Snapshot transcript UUIDs that existed before a new SDK query.

    A pooled ``ClaudeSDKClient`` owns one receive queue across turns. Task
    lifecycle messages — and, on some CLI builds, the tail of the prior
    response — can arrive after that turn's ResultMessage and remain buffered
    until the next ``receive_response()`` call. The snapshot gives the stream
    pump a stable query boundary: messages whose UUID was already persisted
    before ``query()`` belong to an older turn and must not be broadcast again.

    This deliberately uses the normal post-compact SDK view. The only replay we
    need to reject is the immediately preceding active-branch response, and the
    existing (mtime, size) cache keeps the common path cheap. Failure is
    fail-open so a transcript read problem never prevents sending a message.
    """
    try:
        return frozenset(
            str(uid)
            for msg in _cached_session_msgs(sid, model, full=False)
            if (uid := getattr(msg, "uuid", None))
        )
    except Exception as exc:
        sys.stderr.write(
            f"[chat-stream] UUID boundary snapshot skipped sid={sid[:8]} "
            f"exc={type(exc).__name__}\n")
        sys.stderr.flush()
        return frozenset()


def _turn_transcript_boundary(
    sid: str,
    model: str,
) -> tuple[frozenset[str], dict[str, Any]]:
    """Return the SDK replay boundary plus a display-history anchor.

    The UUID set continues to protect the pooled SDK receive queue from
    replaying a previous response.  The compact coordinate is retained on the
    ``TurnBroadcast`` only for the lifetime of this turn and, if the user
    interrupts before a canonical ResultMessage, is copied into the private
    cancelled-turn display snapshot.
    """
    existing = _session_message_uuids(sid, model)
    boundary: dict[str, Any] = {
        "record_count": 0,
        "source_dev": 0,
        "source_inode": 0,
        "normal_uuid": "",
        "normal_total": 0,
        "full_uuid": "",
        "full_total": 0,
        # Distinguish a valid empty first-turn boundary from a failed lookup.
        # A zero record_count alone cannot do that and may point at old history.
        "capture_ok": False,
    }
    try:
        indexed = _ensure_transcript_index(sid)
        if indexed is None:
            boundary["capture_ok"] = not existing
            return existing, boundary
        _, index = indexed
        records = index.get("records") or []
        source = index.get("source") or {}
        boundary.update({
            "record_count": len(records),
            "source_dev": int(source.get("dev") or 0),
            "source_inode": int(source.get("inode") or 0),
            "capture_ok": True,
        })
        for order in ("normal", "full"):
            record_ids = (index.get("orders") or {}).get(order) or []
            prefix = (index.get("bubble_prefix") or {}).get(order) or [0]
            if record_ids:
                rec = records[record_ids[-1]]
                boundary[f"{order}_uuid"] = str(rec.get("uuid") or "")
            boundary[f"{order}_total"] = int(prefix[-1] if prefix else 0)
    except Exception as exc:
        # Recovery metadata is best-effort.  Failing to establish an anchor
        # must never stop the actual model request; an unanchored snapshot is
        # still safely appended to the display history if cancellation occurs.
        sys.stderr.write(
            f"[chat-stream] transcript boundary skipped sid={sid[:8]} "
            f"exc={type(exc).__name__}\n")
        sys.stderr.flush()
    return existing, boundary


class _TurnResponseBoundary:
    """Classify pooled-SDK messages relative to one ``query()`` call."""

    _LIFECYCLE_TYPES = (
        TaskStartedMessage, TaskProgressMessage,
        TaskNotificationMessage, TaskUpdatedMessage,
        RateLimitEvent,
    )
    _TURN_TYPES = (
        StreamEvent, AssistantMessage, UserMessage, SystemMessage, ResultMessage,
    )

    def __init__(self, existing_uuids: frozenset[str] | set[str]):
        self.existing_uuids = frozenset(existing_uuids)
        self.saw_current_payload = False
        self.nonhuman_origin_active = False

    def classify(self, msg: Any) -> str:
        """Return ``forward``, ``current_result``, ``drop`` or ``stale_result``.

        Lifecycle/rate-limit events are intentionally out-of-band and always
        pass through: a late TaskNotification still needs to settle the old
        card. Turn-scoped payload with a UUID present before query is replay.
        An error Result without UUID is accepted immediately so auth/vendor
        failures cannot leave the UI streaming forever.
        """
        if isinstance(msg, self._LIFECYCLE_TYPES):
            return "forward"
        if not isinstance(msg, self._TURN_TYPES):
            return "forward"

        uid = str(getattr(msg, "uuid", None) or "")
        if isinstance(msg, SystemMessage) and not uid:
            data = msg.data if isinstance(msg.data, dict) else {}
            uid = str(data.get("uuid") or "")
        if uid and uid in self.existing_uuids:
            return "stale_result" if isinstance(msg, ResultMessage) else "drop"

        origin = sdk_lifecycle.normalize_origin(getattr(msg, "origin", None))
        if isinstance(msg, UserMessage) and origin is not None:
            self.nonhuman_origin_active = origin["kind"] != "human"
            if self.nonhuman_origin_active:
                return "background"
        if self.nonhuman_origin_active:
            if isinstance(msg, ResultMessage):
                self.nonhuman_origin_active = False
                return "background_result"
            return "background"

        if isinstance(msg, ResultMessage):
            # ``origin=None`` is the current SDK shape for an ordinary
            # client.query(prompt); preserve that compatibility. Any explicit
            # non-human/future origin belongs to a side delivery and cannot
            # close the human turn currently occupying this pooled stream.
            if origin is not None and origin["kind"] != "human":
                return "background_result"
            if bool(getattr(msg, "is_error", False)):
                self.saw_current_payload = True
                return "current_result"
            # Current SDK builds provide Result UUIDs. Accept UUID-less success
            # for older SDKs and existing test doubles; without a stable identity
            # there is no safe basis for calling it replay, and rejecting it can
            # make receive_response() loop forever on clients that replay a fixed
            # result fixture on every call.
            self.saw_current_payload = True
            return "current_result"

        self.saw_current_payload = True
        return "forward"


@router.get(
    "/sessions/{sid}/subagents",
    dependencies=[Depends(require_token)],
)
async def get_session_subagents_api(sid: str) -> dict:
    """Return the SDK-owned nested transcript for every Subagent.

    The top-level SDK APIs are the source of truth.  MuseLab only shapes their
    records for rendering and never scans Claude's private files directly.
    """
    meta = await obs.to_thread_io(
        "chat.subagents_session_read",
        sid,
        sess.get_session_meta,
        sid,
    )
    if meta is None:
        raise HTTPException(404, "session not found")
    workspace = await obs.to_thread_io(
        "chat.subagents_workspace_read",
        sid,
        sess.session_workspace,
        sid,
    )
    model = str(meta.get("model") or "")

    def _load() -> list[dict[str, Any]]:
        with _session_config_dir(model, sid=sid):
            return chat_subagents.load_subagent_threads(
                sid,
                directory=str(workspace),
            )

    threads = await obs.to_thread_io(
        "chat.subagents_history_read",
        sid,
        _load,
    )
    return {"session_id": sid, "threads": threads}


@router.get(
    "/sessions/{sid}/hook-traces",
    dependencies=[Depends(require_token)],
)
async def get_session_hook_traces_api(
    sid: str,
    turn_id: str = Query("", max_length=128),
) -> dict:
    if await obs.to_thread_io(
        "chat.hook_traces_session_read",
        sid,
        sess.get_session_meta,
        sid,
    ) is None:
        raise HTTPException(404, "session not found")
    try:
        traces = await obs.to_thread_io(
            "chat.hook_traces_read",
            sid,
            hook_traces.list_traces,
            sid,
            turn_id=turn_id,
        )
    except (UnsafePrivatePath, ValueError) as exc:
        raise HTTPException(409, "hook trace storage is unavailable") from exc
    return {"session_id": sid, "traces": traces}


@router.get("/sessions/{sid}", dependencies=[Depends(require_token)])
def get_session_api(
    sid: str,
    full: bool = Query(False),
    tail: int = Query(0, ge=0),
    offset: int = Query(-1),
    limit: int = Query(0, ge=0),
    history_generation: str = Query(""),
    around_uuid: str = Query(""),
    before: int = Query(0, ge=0),
    after: int = Query(0, ge=0),
) -> dict:
    """Read session: metadata from muselab sidecar + transcript from CLI JSONL
    via SDK. Merges per-message annotations (cost, model, images) into the
    transcript so the UI gets one flat list of bubbles.

    `full=1` bypasses the SDK's compact-boundary truncation and returns the
    ENTIRE conversation (incl. pre-compact turns) via _full_session_msgs.
    Used by the outline (to list every user prompt) and by the "jump to a
    pre-compact prompt" path. Defaults to the normal post-compact view.

    Windowing (perf): a long session can shape into thousands of UI bubbles
    and several MB of JSON — transferring + JSON.parse-ing the whole thing on
    every session entry froze the browser (the user only ever sees the last
    ~30). So the client pages:
      - `?tail=N`            → only the last N bubbles (initial load)
      - `?offset=A&limit=L`  → bubbles [A, A+L) (the "Load earlier" button)
    The response always carries `total` (full bubble count), `offset` (index
    of the first returned bubble in the full chain) and `has_more` (whether
    older bubbles exist before `offset`) so the client can page backwards.
    `full=1` and the no-param call still return everything (offset=0) for
    outline / export / jump-to-pre-compact back-compat — those need the whole
    list. Windowing is ignored when `full` is set.

    Explicit-interrupt recovery: if the CLI was force-stopped before it could
    flush this turn to JSONL, merge MuseLab's private display snapshot at its
    original transcript anchor. The snapshot is presentation-only and is
    never passed back into model context."""
    meta = sess.get_session_meta(sid)
    if meta is None:
        raise HTTPException(404, "session not found")
    model = meta.get("model", "")
    if (sid in _interrupted_at_startup
            and not _recover_interrupted_turn_snapshot(sid)):
        raise HTTPException(
            503, "interrupted turn could not be loaded into durable history")
    snapshots, snapshot_generation = _load_cancelled_turn_snapshots(sid)
    # Any bounded request uses the byte-offset index, including bounded
    # ``full``-order paging after an outline jump.  Keeping normal/full as an
    # explicit response coordinate prevents a full-order offset from later
    # being sent to the default normal-order endpoint.
    uses_index = bool(around_uuid or tail > 0 or offset >= 0 or snapshots)
    generation = ""
    has_later = False
    pre_total = 0
    history_order = "full" if (full or around_uuid) else "normal"

    if uses_index:
        indexed = _ensure_transcript_index(sid)
        if indexed is None:
            generation = _combined_history_generation(
                "", snapshot_generation)
            if history_generation and history_generation != generation:
                raise HTTPException(409, detail={
                    "error": "history_generation_mismatch",
                    "history_generation": generation,
                })
            if around_uuid:
                raise HTTPException(404, "message uuid not found")
            window, total, win_offset, has_later = _interrupted_history_window(
                None, None, snapshots, {}, history_order,
                tail=tail, offset=offset, limit=limit)
            messages = window
        else:
            transcript_path, index = indexed
            generation = _combined_history_generation(
                str(index.get("history_generation") or ""),
                snapshot_generation,
            )
            # Bubbles stranded before the visible chain's root (post-/compact,
            # or a fork). Reported in FULL-order units in every response so the
            # client can offer "Load earlier" even when the normal-order window
            # is sitting at offset 0 with nothing apparently behind it.
            pre_total = transcript_idx.pre_chain_bubbles(index)
            if history_generation and history_generation != generation:
                raise HTTPException(409, detail={
                    "error": "history_generation_mismatch",
                    "history_generation": generation,
                })
            if sess.has_pending_attachments(sid):
                # Pending bundles are FIFO. Reconcile against every active-chain
                # image record in transcript order before slicing, otherwise a
                # tail page could bind an older bundle to a newer message.
                records = index["records"]
                image_record_ids = [
                    record_i for record_i in index["orders"]["normal"]
                    if records[record_i].get("has_inline_images")
                ]
                image_messages = _indexed_ui_records(
                    transcript_path, index, image_record_ids,
                    sess.get_message_annotations(sid))
                _bind_pending_attachments(sid, image_messages)
            annotations = sess.get_message_annotations(sid)
            if around_uuid:
                # before/after and the legacy limit are all expressed in UI
                # bubbles.  The index maps that exact range back to the small
                # set of JSONL records that intersects it.
                around_limit = limit if before == 0 and after == 0 else 0
                if snapshots:
                    virtual = _interrupted_history_window_around_uuid(
                        transcript_path,
                        index,
                        snapshots,
                        annotations,
                        around_uuid,
                        before,
                        after,
                        limit=around_limit,
                    )
                    if virtual is None:
                        raise HTTPException(404, "message uuid not found")
                    window, total, win_offset, has_later = virtual
                else:
                    record_ids, inner_start, win_offset, win_end, total = (
                        transcript_idx.record_indices_around_uuid(
                            index, around_uuid, before, after,
                            limit=around_limit))
                    if not record_ids:
                        raise HTTPException(404, "message uuid not found")
                    shaped = _indexed_ui_records(
                        transcript_path, index, record_ids, annotations)
                    window = shaped[
                        inner_start:inner_start + (win_end - win_offset)]
                    has_later = win_end < total
                # A corrupt/drifted index must not silently turn a successful
                # around request into a window that omits its target.
                if not any(item.get("uuid") == around_uuid for item in window):
                    raise HTTPException(409, detail={
                        "error": "history_index_mismatch",
                        "history_generation": generation,
                    })
            else:
                order = "full" if full else "normal"
                history_order = order
                if snapshots:
                    window, total, win_offset, has_later = (
                        _interrupted_history_window(
                            transcript_path,
                            index,
                            snapshots,
                            annotations,
                            order,
                            tail=tail,
                            offset=offset,
                            limit=limit,
                        )
                    )
                else:
                    total = index["bubble_prefix"][order][-1]
                    if offset >= 0:
                        start = max(0, min(offset, total))
                        end = total if limit <= 0 else min(total, start + limit)
                    else:
                        start = max(0, total - tail)
                        end = total
                    record_ids, inner_start, _ = (
                        transcript_idx.record_indices_for_bubble_window(
                            index, order, start, end))
                    shaped = _indexed_ui_records(
                        transcript_path, index, record_ids, annotations)
                    window = shaped[inner_start:inner_start + (end - start)]
                    win_offset = start
                    has_later = end < total
            messages = window

            # The index already has the exact normal-order bubble total, so
            # window requests can self-heal message_count without shaping the
            # entire transcript.
            if snapshots:
                normal_total, turns = _interrupted_history_stats(
                    index, snapshots, "normal")
            else:
                normal_total = index["bubble_prefix"]["normal"][-1]
                records = index["records"]
                turns = sum(
                    1 for rec_i in index["orders"]["normal"]
                    if records[rec_i].get("real_user_prompt")
                )
            if not full and meta.get("message_count", 0) != normal_total:
                try:
                    sess.set_message_count(sid, normal_total, turn_count=turns)
                    meta = {
                        **meta,
                        "message_count": normal_total,
                        "turn_count": turns,
                    }
                except Exception:
                    pass
    else:
        # Compatibility path for callers that explicitly request the complete
        # normal/full transcript. Windowed callers never enter this O(N) SDK
        # parse-and-shape path.
        messages = _shaped_ui_messages(sid, model, full)
        if sess.has_pending_attachments(sid):
            _bind_pending_attachments(sid, messages)
        total = len(messages)
        win_offset = 0
        window = messages
        try:
            indexed = _ensure_transcript_index(sid)
            if indexed is not None:
                generation = str(indexed[1].get("history_generation") or "")
                pre_total = transcript_idx.pre_chain_bubbles(indexed[1])
        except Exception:
            generation = ""
            pre_total = 0
        if not full and total > 0 and meta.get("message_count", 0) != total:
            try:
                turns = sum(
                    1 for item in messages
                    if item.get("role") == "user"
                    and item.get("_steeringAdjustment") is not True
                )
                sess.set_message_count(sid, total, turn_count=turns)
            except Exception:
                pass

    active_turn = _active_turns.get(sid)
    if active_turn is not None and active_turn.done:
        active_turn = None
    _apply_runtime_task_overlays(
        window,
        sess.get_runtime_task_overlays(sid),
    )
    _complete_turn_footer_metadata(
        window,
        model,
        has_later=has_later,
        active_turn=active_turn,
    )

    return {
        **meta,
        "messages": window,
        "total": total,
        "offset": win_offset,
        "has_more": win_offset > 0,
        "has_later": has_later,
        # Only meaningful while the client is reading normal order: once it has
        # switched to full order those bubbles are inside `total` already, and
        # plain `offset > 0` paging reaches them.
        "pre_total": pre_total if history_order == "normal" else 0,
        "history_generation": generation,
        "history_order": history_order,
        # Presentation generation used by a rollover child to prove that it
        # has adopted the hidden owner's latest continuation bubble before
        # stopping its inherited-task poller.
        "runtime_ui_revision": snapshot_generation,
    }


@router.get("/sessions/{sid}/blocks/{block_id}",
            dependencies=[Depends(require_token)])
def get_session_block_api(sid: str, block_id: str) -> dict:
    """Read one canonical UI block body by stable transcript identity."""
    if sess.get_session_meta(sid) is None:
        raise HTTPException(404, "session not found")
    match = re.fullmatch(r"([^:]+):(\d+):([a-z_]+)", block_id)
    if match is None:
        raise HTTPException(400, "invalid block id")
    record_uuid = match.group(1)
    indexed = _ensure_transcript_index(sid)
    if indexed is None:
        raise HTTPException(404, "transcript not found")
    transcript_path, index = indexed
    record_i = next(
        (i for i, record in enumerate(index.get("records") or [])
         if record_uuid in {
             str(record.get("uuid") or ""),
             str(record.get("presentation_uuid") or ""),
         }),
        None,
    )
    if record_i is None:
        raise HTTPException(404, "block not found")
    entries = transcript_idx.read_records(transcript_path, index, [record_i])
    if not entries:
        raise HTTPException(404, "block not found")
    entry = entries[0]
    raw = _raw_msg_from_entry(entry)
    if raw is None:
        raise HTTPException(404, "block not found")
    compact = {raw.uuid} if entry.get("isCompactSummary") else set()
    messages = _sdk_messages_to_ui(
        [raw], sess.get_message_annotations(sid), compact)
    block = next(
        (message for message in messages
         if message.get("block_id") == block_id),
        None,
    )
    if block is None or not isinstance(block.get("text"), str):
        raise HTTPException(404, "block body not found")
    body = block["text"]
    if block.get("role") == "tool_result":
        tool_id = str(block.get("id") or "")
        tool_name = str((index.get("tool_use_names") or {}).get(tool_id) or "")
        if tool_name:
            block["tool_name"] = tool_name
            if tool_name == "Bash":
                parsed = _parse_bash_result(body)
                if parsed:
                    block["bash"] = parsed
    if len(body.encode("utf-8")) > 8 * 1024 * 1024:
        raise HTTPException(413, "block body exceeds the 8 MiB response limit")
    return {
        **block,
        "body_state": "loaded",
        "body_available": True,
        "body_length": len(body),
        "body_ref": block_id,
    }


def _outline_preview(text: str) -> str:
    """Patchable facade for session-outline previews."""
    return chat_presentation.outline_preview(text)


@router.get("/sessions/{sid}/outline", dependencies=[Depends(require_token)])
def get_session_outline_api(sid: str) -> dict:
    """Lightweight session outline: just the user-prompt previews + UUIDs,
    extracted server-side. The outline used to fetch the session with
    `?full=1` (the ENTIRE raw JSONL — several MB on a long session) and filter
    for user messages in the browser, which froze the page when opening the
    outline on a big session. This returns only what the outline renders:
    a small `[{preview, uuid}]` list spanning the WHOLE conversation (incl.
    pre-compact prompts, since it reads the full JSONL)."""
    meta = sess.get_session_meta(sid)
    if meta is None:
        raise HTTPException(404, "session not found")
    _, snapshot_generation = _load_cancelled_turn_snapshots(sid)
    try:
        indexed = _ensure_transcript_index(sid)
        if indexed is None:
            return {
                "outline": [],
                "history_generation": _combined_history_generation(
                    "", snapshot_generation),
            }
        _, index = indexed
        records = index["records"]
        items = [
            {"preview": records[record_i]["user_preview"],
             "uuid": records[record_i]["uuid"]}
            for record_i in index["orders"]["full"]
            if records[record_i]["type"] == "user"
            and records[record_i]["user_preview"]
            and not records[record_i]["compact"]
        ]
        return {
            "outline": items,
            "history_generation": _combined_history_generation(
                index.get("history_generation") or "", snapshot_generation),
        }
    except Exception:
        return {
            "outline": [],
            "history_generation": _combined_history_generation(
                "", snapshot_generation),
        }


def _broadcast_to_ui_messages(bc: "TurnBroadcast") -> list[dict]:
    """Patchable facade for interrupted-turn display reconstruction."""
    return chat_presentation.broadcast_to_ui_messages(bc)




def _canonical_uuid_component(value: str) -> str | None:
    return chat_overlays._canonical_uuid_component(value)

def _cancelled_turn_session_dir(sid: str) -> Path | None:
    return chat_overlays._cancelled_turn_session_dir(sid)

def _cancelled_turn_snapshot_path(sid: str, turn_id: str) -> Path | None:
    return chat_overlays._cancelled_turn_snapshot_path(sid, turn_id)

def _runtime_continuation_outbox_dir(source_sid: str) -> Path | None:
    return chat_overlays._runtime_continuation_outbox_dir(source_sid)

def _runtime_continuation_outbox_path(source_sid: str, event_id: str) -> Path | None:
    return chat_overlays._runtime_continuation_outbox_path(source_sid, event_id)

def _delete_runtime_continuation_outboxes(source_sid: str) -> None:
    return chat_overlays._delete_runtime_continuation_outboxes(source_sid)

def _load_runtime_continuation_outbox(source_sid: str, event_id: str) -> dict | None:
    return chat_overlays._load_runtime_continuation_outbox(source_sid, event_id)

def _runtime_continuation_outbox_entries() -> list[tuple[str, str]]:
    return chat_overlays._runtime_continuation_outbox_entries()

def _session_has_runtime_continuation_outbox(source_sid: str) -> bool:
    return chat_overlays._session_has_runtime_continuation_outbox(source_sid)

def _runtime_continuation_outbox_event_ids(source_sid: str) -> list[str]:
    return chat_overlays._runtime_continuation_outbox_event_ids(source_sid)

def _runtime_lineage_has_ready_continuation(leaf_sid: str) -> bool:
    return chat_overlays._runtime_lineage_has_ready_continuation(leaf_sid)

def _persist_runtime_continuation_outbox(source_sid: str, broadcast: 'TurnBroadcast', *, completed_at_ms: int, elapsed_s: float, terminal_status: str, incomplete_error: str='') -> str:
    return chat_overlays._persist_runtime_continuation_outbox(source_sid, broadcast, completed_at_ms=completed_at_ms, elapsed_s=elapsed_s, terminal_status=terminal_status, incomplete_error=incomplete_error)

def _sync_runtime_display_message_count(sid: str) -> None:
    return chat_overlays._sync_runtime_display_message_count(sid)

def _persist_runtime_continuation_snapshot(target_sid: str, outbox: dict) -> bool:
    return chat_overlays._persist_runtime_continuation_snapshot(target_sid, outbox)

def _persist_runtime_continuation_snapshot_locked(target_sid: str, outbox: dict) -> bool:
    return chat_overlays._persist_runtime_continuation_snapshot_locked(target_sid, outbox)

def _copy_runtime_continuation_snapshots(source_sid: str, target_sid: str, uuid_mapping: dict[str, str]) -> int:
    return chat_overlays._copy_runtime_continuation_snapshots(source_sid, target_sid, uuid_mapping)

def _copy_runtime_continuation_snapshots_locked(source_sid: str, target_sid: str, uuid_mapping: dict[str, str]) -> int:
    return chat_overlays._copy_runtime_continuation_snapshots_locked(source_sid, target_sid, uuid_mapping)

async def _deliver_runtime_continuation_outbox(source_sid: str, event_id: str) -> bool:
    return await chat_overlays._deliver_runtime_continuation_outbox(source_sid, event_id)

async def _flush_runtime_continuations_at_turn_boundary(leaf_sid: str, *, expected_active: 'TurnBroadcast | None'=None) -> int:
    return await chat_overlays._flush_runtime_continuations_at_turn_boundary(leaf_sid, expected_active=expected_active)

def _schedule_runtime_continuation_delivery(source_sid: str, event_id: str) -> asyncio.Task | None:
    return chat_overlays._schedule_runtime_continuation_delivery(source_sid, event_id)

async def recover_runtime_continuation_outboxes_at_startup() -> int:
    return await chat_overlays.recover_runtime_continuation_outboxes_at_startup()

def _runtime_continuation_projection_state(sid: str) -> tuple[bool, str]:
    return chat_overlays._runtime_continuation_projection_state(sid)

def _delete_cancelled_turn_snapshots(sid: str) -> None:
    return chat_overlays._delete_cancelled_turn_snapshots(sid)

def _cancelled_snapshot_canonical_span(transcript_path: Path, index: dict, snapshot: dict) -> tuple[list[str], str]:
    return chat_overlays._cancelled_snapshot_canonical_span(transcript_path, index, snapshot)

def _heal_cancelled_snapshot_from_canonical(sid: str, path: Path, snapshot: dict, transcript_path: Path, index: dict) -> bool:
    return chat_overlays._heal_cancelled_snapshot_from_canonical(sid, path, snapshot, transcript_path, index)

def _load_cancelled_turn_snapshots(sid: str) -> tuple[list[dict], str]:
    return chat_overlays._load_cancelled_turn_snapshots(sid)

def _combined_history_generation(base: str, snapshot_generation: str) -> str:
    return chat_overlays._combined_history_generation(base, snapshot_generation)

def _persist_cancelled_turn_snapshot(bc: 'TurnBroadcast') -> bool:
    return chat_overlays._persist_cancelled_turn_snapshot(bc)

def _cancelled_footer_values(bc: 'TurnBroadcast') -> tuple[int, float]:
    return chat_overlays._cancelled_footer_values(bc)

def _persist_cancelled_footer_annotation_locked(bc: 'TurnBroadcast', now_ms: int, elapsed_s: float) -> bool:
    return chat_overlays._persist_cancelled_footer_annotation_locked(bc, now_ms, elapsed_s)

def _persist_cancelled_turn_snapshot_locked(bc: 'TurnBroadcast') -> bool:
    return chat_overlays._persist_cancelled_turn_snapshot_locked(bc)

def _persist_completed_result_snapshot(bc: 'TurnBroadcast', result_text: str, *, terminal_at_ms: int, elapsed_s: float | None=None, memory_recall: dict | None=None, terminal_reason: str='', turn_origin: dict | None=None, turn_id: str='', model_usage: dict | None=None) -> bool:
    return chat_overlays._persist_completed_result_snapshot(bc, result_text, terminal_at_ms=terminal_at_ms, elapsed_s=elapsed_s, memory_recall=memory_recall, terminal_reason=terminal_reason, turn_origin=turn_origin, turn_id=turn_id, model_usage=model_usage)


def _persist_failed_turn_snapshot(bc: 'TurnBroadcast', error_text: str, *, terminal_at_ms: int | None=None, elapsed_s: float | None=None, memory_recall: dict | None=None, canonical_terminal_published: bool=False, terminal_status: str='failed', terminal_reason: str='', turn_origin: dict | None=None, turn_id: str='', model_usage: dict | None=None) -> bool:
    return chat_overlays._persist_failed_turn_snapshot(bc, error_text, terminal_at_ms=terminal_at_ms, elapsed_s=elapsed_s, memory_recall=memory_recall, canonical_terminal_published=canonical_terminal_published, terminal_status=terminal_status, terminal_reason=terminal_reason, turn_origin=turn_origin, turn_id=turn_id, model_usage=model_usage)

def _recover_interrupted_turn_snapshot(sid: str) -> bool:
    return chat_overlays._recover_interrupted_turn_snapshot(sid)

def _interrupted_history_segments(index: dict | None, snapshots: list[dict], order: str) -> tuple[list[dict], int]:
    return chat_overlays._interrupted_history_segments(index, snapshots, order)

def _interrupted_history_stats(index: dict | None, snapshots: list[dict], order: str) -> tuple[int, int]:
    return chat_overlays._interrupted_history_stats(index, snapshots, order)

def _interrupted_history_window(transcript_path: Path | None, index: dict | None, snapshots: list[dict], annotations: dict[str, dict], order: str, *, tail: int=0, offset: int=-1, limit: int=0) -> tuple[list[dict], int, int, bool]:
    return chat_overlays._interrupted_history_window(transcript_path, index, snapshots, annotations, order, tail=tail, offset=offset, limit=limit)

def _interrupted_history_window_around_uuid(transcript_path: Path, index: dict, snapshots: list[dict], annotations: dict[str, dict], uuid_value: str, before: int, after: int, *, limit: int=0) -> tuple[list[dict], int, int, bool] | None:
    return chat_overlays._interrupted_history_window_around_uuid(transcript_path, index, snapshots, annotations, uuid_value, before, after, limit=limit)

def _turn_uuids_from_boundary(
    sid: str,
    boundary: dict[str, Any],
    *,
    started_at_ms: int,
    terminal_at_ms: int,
) -> tuple[str | None, str | None, bool]:
    """Resolve current-turn UUIDs and whether canonical inspection succeeded.

    A naive "latest assistant" tail scan can return the previous turn when
    the current Gateway failure has only a ResultMessage. Reuse the indexed
    parent-chain resolver used by recovery snapshots so a UUID is returned
    only when it descends from this turn's real user prompt after the exact
    pre-query record coordinate.
    """
    if not bool((boundary or {}).get("capture_ok")):
        return None, None, False
    try:
        indexed = _ensure_transcript_index(sid)
        if indexed is None:
            return None, None, True
        transcript_path, index = indexed
        probe = {
            "transcript_boundary": dict(boundary or {}),
            "started_at_ms": int(started_at_ms or 0),
            "terminal_at_ms": int(terminal_at_ms or 0),
            "interrupted_at_ms": int(terminal_at_ms or 0),
            "canonical_terminal_published": True,
        }
        span, assistant_uuid = _cancelled_snapshot_canonical_span(
            transcript_path, index, probe)
        if not span:
            return assistant_uuid or None, None, True
        span_set = set(span)
        user_uuid: str | None = None
        for record in index.get("records") or []:
            uid = str(record.get("uuid") or "")
            if uid not in span_set or not record.get("real_user_prompt"):
                continue
            user_uuid = uid
            break
        return assistant_uuid or None, user_uuid, True
    except Exception as exc:
        # Never fall back to an unbounded/latest UUID. Callers treat absent
        # evidence as an uncommitted turn, which is safer than publishing a
        # success that can disappear or mutating a prior turn's annotation.
        sys.stderr.write(
            f"[chat] turn UUID boundary resolve skipped sid={sid[:8]} "
            f"exc={type(exc).__name__}\n")
        sys.stderr.flush()
        return None, None, False


def _turn_prevented_error_from_boundary(
    sid: str,
    boundary: dict[str, Any],
) -> dict | None:
    """Recover a CLI hook rejection that the warm SDK stream did not emit.

    Some UserPromptSubmit failures are persisted only as a canonical system row
    followed by a nominally successful ResultMessage.  The missing user UUID is
    then expected, not a generic commit race; preserve the actionable hook error.
    """
    if not bool((boundary or {}).get("capture_ok")):
        return None
    try:
        indexed = _ensure_transcript_index(sid)
        if indexed is None:
            return None
        transcript_path, index = indexed
        source = index.get("source") or {}
        expected_dev = int((boundary or {}).get("source_dev") or 0)
        expected_inode = int((boundary or {}).get("source_inode") or 0)
        if ((expected_dev and int(source.get("dev") or 0) != expected_dev)
                or (expected_inode
                    and int(source.get("inode") or 0) != expected_inode)):
            return None
        records = index.get("records") or []
        start = max(0, int((boundary or {}).get("record_count") or 0))
        system_ids = [
            record_i
            for record_i in range(start, len(records))
            if records[record_i].get("type") == "system"
        ]
        for entry in transcript_idx.read_records(
                transcript_path, index, system_ids):
            data = entry.get("data") if isinstance(entry.get("data"), dict) else {}
            if not (entry.get("preventContinuation") is True
                    or data.get("preventContinuation") is True):
                continue
            text = str(
                entry.get("content") or entry.get("error")
                or data.get("content") or data.get("error") or ""
            ).strip()
            return {
                "message": text or (
                    "User prompt was rejected before it was committed."),
                "source": "system_prevent_continuation",
                "api_error_status": None,
            }
    except Exception as exc:
        sys.stderr.write(
            f"[chat] turn hook rejection recovery skipped sid={sid[:8]} "
            f"exc={type(exc).__name__}\n")
        sys.stderr.flush()
    return None


async def _settle_turn_uuids(
    sid: str,
    boundary: dict[str, Any],
    *,
    started_at_ms: int,
    terminal_at_ms: int,
    require_assistant: bool,
) -> tuple[str | None, str | None, bool]:
    """Wait briefly for the CLI's canonical JSONL append to become observable."""
    evidence: tuple[str | None, str | None, bool] = (None, None, False)
    for delay in (0.0, 0.05, 0.15, 0.3, 0.5):
        if delay:
            await asyncio.sleep(delay)
        evidence = await asyncio.to_thread(
            _turn_uuids_from_boundary,
            sid,
            boundary,
            started_at_ms=started_at_ms,
            terminal_at_ms=terminal_at_ms,
        )
        assistant_uuid, user_uuid, inspected = evidence
        if user_uuid and (assistant_uuid or not require_assistant):
            break
        if not inspected and not bool((boundary or {}).get("capture_ok")):
            break
    return evidence

@router.get("/sessions/{sid}/export")
def export_session_markdown(sid: str, ticket: str = Query("")) -> Response:
    """Render the transcript as a single Markdown file the user can save.

    A plain download anchor cannot add an auth header, so it carries a
    short-lived, session-bound, single-use resource ticket instead of the
    long-lived global API token.
    """
    _require_chat_resource_ticket(ticket, ("export", sid))
    meta = sess.get_session_meta(sid)
    if meta is None:
        raise HTTPException(404, "session not found")
    model = meta.get("model", "")
    annotations = sess.get_message_annotations(sid)
    snapshots, _ = _load_cancelled_turn_snapshots(sid)
    if snapshots:
        indexed = _ensure_transcript_index(sid)
        transcript_path = indexed[0] if indexed is not None else None
        index = indexed[1] if indexed is not None else None
        messages, _, _, _ = _interrupted_history_window(
            transcript_path, index, snapshots, annotations, "normal")
    else:
        messages = _shaped_ui_messages(
            sid, model, False, defer_large_bodies=False)
    # Bind any unbound pending image/doc attachments (those persisted
    # before the stream completed could write a uuid annotation) to the
    # user messages that have inline image refs but no thumb/url yet.
    _bind_pending_attachments(sid, messages)

    name = meta.get("name") or "session"
    model = meta.get("model") or ""
    created = meta.get("created_at")
    created_str = (datetime.fromtimestamp(created).strftime("%Y-%m-%d %H:%M")
                    if created else "")
    lines: list[str] = [f"# {name}", ""]
    if created_str:
        lines.append(f"*Created: {created_str}*  ")
    if model:
        lines.append(f"*Model: {model}*  ")
    lines.append(f"*Messages: {len(messages)}*")
    lines.append("")
    for m in messages:
        role = m.get("role")
        text = (m.get("text") or "").strip()
        if not text or role in ("tool_use", "tool_result"):
            continue
        if role == "user":
            lines.append("---")
            lines.append("")
            lines.append("### 👤 User")
        elif role == "assistant":
            lines.append("### 🤖 Muse")
        else:
            lines.append(f"### {role}")
        lines.append("")
        lines.append(text)
        lines.append("")

    body = "\n".join(lines)
    # Filenames in Content-Disposition can't safely include CJK / spaces in all
    # browsers; fall back to a slug. RFC 5987 filename*=UTF-8 covers Unicode for
    # modern browsers; the bare filename is an ASCII fallback for older ones.
    safe_slug = re.sub(r"[^A-Za-z0-9_.-]+", "_", name).strip("_") or "session"
    safe_slug = safe_slug[:60]
    encoded = urllib.parse.quote(name, safe="")
    headers = {
        "Content-Disposition":
            f'attachment; filename="{safe_slug}.md"; '
            f"filename*=UTF-8''{encoded}.md",
    }
    return Response(content=body, media_type="text/markdown; charset=utf-8",
                    headers=headers)


def _clear_session_runtime_state(sid: str) -> list[asyncio.Task]:
    """Forget loop-owned state and return tasks that deletion should join.

    This function must run on the event-loop thread. Disk/SDK cleanup runs in
    a worker separately; cancelling asyncio objects from that worker is not
    thread-safe.
    """
    cancelled_tasks: list[asyncio.Task] = []
    perm.clear_session_permissions(sid)
    _pending_runtime_rebuilds.discard(sid)
    _native_retry_commits.pop(sid, None)
    runtime_lock = _session_runtime_locks.get(sid)
    if runtime_lock is not None and not runtime_lock.locked():
        _session_runtime_locks.pop(sid, None)
    task_storage_lock = _runtime_task_storage_locks.get(sid)
    if task_storage_lock is not None and not task_storage_lock.locked():
        _runtime_task_storage_locks.pop(sid, None)
    drain_task = _queue_drain_tasks.pop(sid, None)
    retry_task = _queue_drain_retry_tasks.pop(sid, None)
    _queue_drain_rekicks.discard(sid)
    for task in (drain_task, retry_task):
        if task is not None and not task.done():
            task.cancel()
            cancelled_tasks.append(task)
    prewarm_task = _runtime_prewarm_tasks.pop(sid, None)
    if prewarm_task is not None and not prewarm_task.done():
        prewarm_task.cancel()
        cancelled_tasks.append(prewarm_task)
    drain_lock = _queue_drain_locks.get(sid)
    if drain_lock is not None and not drain_lock.locked():
        _queue_drain_locks.pop(sid, None)
    _continuation_generations[sid] = _continuation_generations.get(sid, 0) + 1
    watcher = _task_watchers.pop(sid, None)
    if watcher is not None and not watcher.done():
        watcher.cancel()
        cancelled_tasks.append(watcher)
    _background_turn_started_at.pop(sid, None)
    _background_origin_turn_id.pop(sid, None)
    task_ids = _sessions_with_inflight_tasks.pop(sid, set())
    for task_id in task_ids:
        _bg_task_descriptions.pop(task_id, None)
        _bg_task_tool_use_ids.pop(task_id, None)
        _bg_task_pinned_at.pop(task_id, None)
    recent_handle = _recent_turn_expiry_handles.pop(sid, None)
    if recent_handle is not None:
        recent_handle.cancel()
    recent = _recent_turns.pop(sid, None)
    if recent is not None:
        recent.close()
    broadcast = _active_turns.pop(sid, None)
    if broadcast is not None:
        # Explicit session deletion owns the terminal state. Prevent the
        # cancelled pump's finally block from recreating a private display
        # snapshot after purge_session_storage has removed every artifact.
        broadcast.cancelled_snapshot_suppressed = True
        # Join any already-running to_thread writer before the caller removes
        # the snapshot directory. A later writer observes the suppression bit.
        with broadcast.cancelled_snapshot_lock:
            pass
        # Every terminal path must observe deletion before cancellation is
        # delivered.  In particular, the outer `_start_turn` owner decides
        # between ordinary startup rollback and deletion-safe cleanup from
        # this flag.
        broadcast.cancelled = True
        task = getattr(broadcast, "task", None)
        if task is not None and not task.done():
            task.cancel()
            cancelled_tasks.append(task)
        startup_task = getattr(broadcast, "startup_task", None)
        if startup_task is not None and not startup_task.done():
            startup_task.cancel()
            cancelled_tasks.append(startup_task)
        startup_owner = getattr(broadcast, "startup_owner_task", None)
        if startup_owner is not None and not startup_owner.done():
            startup_owner.cancel()
            cancelled_tasks.append(startup_owner)
        cancelled_cleanup = getattr(
            broadcast, "_cancelled_startup_cleanup_task", None)
        if cancelled_cleanup is not None and not cancelled_cleanup.done():
            # This cleanup may be joining an Activity/snapshot worker that
            # cannot be cancelled. The async purge path waits it explicitly;
            # cancelling the wrapper would lose the only handle to that write.
            cancelled_tasks.append(cancelled_cleanup)
        broadcast.finish()
    try:
        current = asyncio.current_task()
    except RuntimeError:
        current = None
    return list({
        task for task in cancelled_tasks
        if task is not current and not task.done()
    })


def _purge_single_session_storage(sid: str) -> bool:
    """Synchronously delete one runtime's disk-owned session state.

    This internal primitive deliberately does not walk rollover links; callers
    that already froze a lineage (including rollback of a provisional child)
    can therefore remove exactly one runtime.
    """
    with sess.session_lifecycle_lock(sid):
        return _purge_session_storage_disk_locked(sid)


def _purge_session_storage_disk_locked(sid: str) -> bool:
    """Remove EVERY per-session artifact: SDK JSONL, muselab sidecar/index/
    queue, attachments dir, and in-memory state. Returns True if any layer
    existed (SDK transcript OR sidecar) — callers treat that as "the session
    was real and is now gone".

    Shared by the HTTP delete endpoint and the scheduler's reuse-mode task
    cascade so both delete the same set of artifacts. Deliberately tolerant:
    a session may exist in only ONE layer (SDK-only when the sidecar was
    lost; sidecar-only when the session never streamed), and local cleanup
    (attachments / active-turn sidecars) runs regardless so nothing
    is orphaned by a partial state."""
    removed = False
    workspace = sess.session_workspace(sid)
    try:
        sdk_delete_session(sid, directory=str(workspace))
        removed = True
    except (FileNotFoundError, ValueError):
        pass   # JSONL never existed (session never streamed) — that's fine
    _durable_attachment_store.cancel_session(sid)
    if sess.delete_session(sid):
        removed = True
    try:
        _transcript_index_path(sid).unlink(missing_ok=True)
    except OSError:
        pass
    # Sweep per-session attachments dir (uploaded image full-res originals
    # persisted by upload-image → send pipeline). Without this, deleting
    # a session would orphan its image files on disk forever.
    attach_dir = _attachments_base() / sid
    try:
        kind = private_path_kind(attach_dir)
        if kind == "directory":
            shutil.rmtree(attach_dir)
        elif kind in {"file", "symlink", "other"}:
            attach_dir.unlink()
    except OSError:
        pass
    _delete_active_turn_sidecar(sid)
    _delete_cancelled_turn_snapshots(sid)
    _delete_runtime_continuation_outboxes(sid)
    hook_traces.purge(sid)
    return removed


def purge_session_storage(sid: str) -> bool:
    """Synchronously purge the complete durable rollover lineage from disk.

    This compatibility path performs no asyncio mutation and is safe from a
    worker thread. Request handlers use :func:`purge_session_storage_async` so
    loop-owned turns/watchers are cancelled and joined as well.
    """
    lineage = sess.runtime_lineage(sid) or [sid]
    pending_lineage = list(dict.fromkeys(lineage))
    queued_lineage = set(pending_lineage)

    def _has_live_runtime(runtime_ids: set[str]) -> bool:
        """Fail closed when the sync path cannot join loop-owned writers."""
        try:
            for runtime_sid in runtime_ids:
                active = _active_turns.get(runtime_sid)
                watcher = _task_watchers.get(runtime_sid)
                prewarm = _runtime_prewarm_tasks.get(runtime_sid)
                drain = _queue_drain_tasks.get(runtime_sid)
                runtime_lock = _session_runtime_locks.get(runtime_sid)
                drain_lock = _queue_drain_locks.get(runtime_sid)
                rollover_lock = _runtime_rollover_locks.get(runtime_sid)
                if (
                    (active is not None and not active.done)
                    or _sessions_with_inflight_tasks.get(runtime_sid)
                    or (watcher is not None and not watcher.done())
                    or (prewarm is not None and not prewarm.done())
                    or (drain is not None and not drain.done())
                    or (runtime_lock is not None and runtime_lock.locked())
                    or (drain_lock is not None and drain_lock.locked())
                    or (rollover_lock is not None and rollover_lock.locked())
                    or any(
                        not task.done()
                        for task in _session_disconnect_tasks.get(
                            runtime_sid, set())
                    )
                    or any(
                        not task.done()
                        for task in _session_runtime_cleanup_tasks.get(
                            runtime_sid, set())
                    )
                    or (
                        (purge := _session_purge_tasks.get(runtime_sid))
                        is not None and not purge.done()
                    )
                ):
                    return True
            if any(
                key[0] in runtime_ids
                for key in tuple(_clients)
            ):
                return True
            if any(
                source_sid in runtime_ids and not task.done()
                for (source_sid, _), task
                in tuple(_runtime_continuation_delivery_tasks.items())
            ):
                return True
        except RuntimeError:
            # A loop-thread registry changed during the cross-thread snapshot.
            # That is itself proof the synchronous caller cannot safely own the
            # lifecycle; the async path can serialize and join it.
            return True
        return False

    if _has_live_runtime(queued_lineage):
        raise RuntimeError(
            "cannot synchronously purge a live runtime lineage; "
            "use purge_session_storage_async"
        )

    removed = False
    cursor = 0
    while cursor < len(pending_lineage) and cursor < 32:
        runtime_sid = pending_lineage[cursor]
        cursor += 1
        # Fence before re-reading the edge. A successor linked after the initial
        # lineage snapshot either lands before this tombstone and is appended,
        # or its rollover observes deletion and rolls the provisional child back.
        sess.begin_session_delete(runtime_sid)
        fenced_meta = sess.get_session_meta(runtime_sid)
        late_successor = str(
            (fenced_meta or {}).get("runtime_successor") or "")
        if late_successor and late_successor not in queued_lineage:
            if _has_live_runtime({late_successor}):
                raise RuntimeError(
                    "cannot synchronously purge a live runtime lineage; "
                    "use purge_session_storage_async"
                )
            queued_lineage.add(late_successor)
            pending_lineage.append(late_successor)
        # Close the check -> reservation race. ``begin_session_delete`` now
        # prevents any *new* turn/queue owner from committing, while this second
        # observation catches one that reserved between the initial check and
        # the tombstone. Leave the tombstone in place and let the async API join
        # that exact owner; never race it with destructive disk removal.
        if _has_live_runtime(queued_lineage):
            raise RuntimeError(
                "cannot synchronously purge a live runtime lineage; "
                "use purge_session_storage_async"
            )
        removed = _purge_single_session_storage(runtime_sid) or removed
    return removed


async def _purge_session_storage_async_inner(sid: str) -> bool:
    # Fence queue writes before any await. A cancelled drain can still finish a
    # synchronous RMW, but every later save observes the tombstone.
    await asyncio.to_thread(sess.begin_session_delete, sid)
    continuation_deliveries = [
        task
        for (source_sid, _), task
        in tuple(_runtime_continuation_delivery_tasks.items())
        if source_sid == sid and not task.done()
    ]
    for task in continuation_deliveries:
        task.cancel()
    if continuation_deliveries:
        await asyncio.gather(*continuation_deliveries, return_exceptions=True)
    # Scheduler runs are separate from chat's TurnBroadcast/task-watcher
    # ownership.  Revoke and join them immediately after the tombstone so an
    # already-running scheduled query cannot append history, unread state, or
    # recreate the SDK transcript after the disk purge.
    from . import scheduler as _scheduler
    scheduler_runs, had_scheduled_activity, _ = (
        _scheduler.cancel_runs_for_session_now(sid)
    )
    # Capture and cancel the outer owner before another await. A cold startup
    # can otherwise unwind only after disconnect_client(), create its
    # cancellation cleanup after our pending-task snapshot, and outlive the
    # disk purge. The deletion tombstone prevents any replacement reservation.
    active = _active_turns.get(sid)
    had_active_activity = bool(
        active is not None and active.activity_started)
    pending = _clear_session_runtime_state(sid)
    # Cancellation alone does not reliably tear down a CLI blocked in
    # connect/receive. Pop and disconnect the pooled runtime before joining
    # scheduler owners; otherwise DELETE can wait forever on the very process
    # that the later disconnect was supposed to stop.
    await disconnect_client(sid)
    critical_cleanup = (
        getattr(active, "_cancelled_startup_cleanup_task", None)
        if active is not None else None
    )
    cleanup_owners = set(pending)
    cleanup_owners.update(scheduler_runs)
    if critical_cleanup is not None:
        cleanup_owners.add(critical_cleanup)
    cleanup_complete = await _join_session_runtime_cleanup(
        sid,
        cleanup_owners,
        timeout=5.0,
    )
    # A bounded owner join protects DELETE from a permanently wedged SDK, but
    # the global Activity ledger must never be left in `running` if that bound
    # is reached.  The normal owner cleanup clears this bit first; this is an
    # idempotent fallback for the timeout/future-cleanup race.
    if had_active_activity or had_scheduled_activity:
        try:
            from .activity import activity as _activity
            await asyncio.to_thread(
                _activity.finish,
                sid,
                "cancelled",
                activity_source=(
                    active.activity_source
                    if had_active_activity
                    else "scheduled"
                ),
            )
        except Exception as exc:
            sys.stderr.write(
                f"[activity] delete finish failed sid={sid[:8]} "
                f"exc={type(exc).__name__}\n"
            )
    if not cleanup_complete:
        # Do not race the destructive disk purge against a late snapshot,
        # queue write, SDK response, or scheduler finalizer. The tombstone is
        # intentionally retained; a retry joins these exact owners and then
        # continues the same deletion transaction.
        raise RuntimeCleanupTimeout(
            "session cleanup is still in progress; retry the operation"
        )
    _session_usage.pop(sid, None)
    _session_usage_turns.pop(sid, None)
    _interrupted_at_startup.pop(sid, None)
    return await asyncio.to_thread(_purge_single_session_storage, sid)


async def _purge_single_session_storage_async(sid: str) -> bool:
    """Delete one runtime and its disk state under a reusable bounded owner.

    The destructive transaction remains shielded from an HTTP disconnect, but
    the request itself has a hard deadline. A retry joins the same owner rather
    than starting a second purge against partially-mutated state.
    """
    cleanup = _session_purge_tasks.get(sid)
    if cleanup is None:
        cleanup = asyncio.create_task(_purge_session_storage_async_inner(sid))
        _session_purge_tasks[sid] = cleanup

        def _expire(done: asyncio.Task) -> None:
            error: BaseException | None = None
            if done.cancelled():
                error = asyncio.CancelledError()
            else:
                try:
                    error = done.exception()
                except BaseException as exc:
                    error = exc

            # A failed transaction should be retryable on the next request.
            # Successful unobserved results stay briefly cached so a client
            # that lost its response sees the same idempotent outcome.
            if error is not None:
                if _session_purge_tasks.get(sid) is done:
                    _session_purge_tasks.pop(sid, None)
                return

            def _drop() -> None:
                if _session_purge_tasks.get(sid) is done:
                    _session_purge_tasks.pop(sid, None)

            done.get_loop().call_later(300.0, _drop)

        cleanup.add_done_callback(_expire)
    try:
        done, pending = await asyncio.wait(
            {cleanup},
            timeout=_SESSION_DELETE_DEADLINE_S,
        )
    except asyncio.CancelledError:
        # The shared task intentionally keeps running under the deletion
        # tombstone; a subsequent request can join it.
        raise
    if pending:
        raise RuntimeCleanupTimeout(
            "session deletion is still running; retry the operation"
        )
    try:
        return cleanup.result()
    finally:
        if _session_purge_tasks.get(sid) is cleanup:
            _session_purge_tasks.pop(sid, None)


async def purge_session_storage_async(sid: str) -> bool:
    """Delete every linked runtime generation containing ``sid``.

    Fence each observed generation before cleanup, then re-read its successor
    so a rollover that commits at the snapshot boundary is included rather than
    leaving an invisible CLI, private snapshot or outbox behind.
    """
    lineage = await asyncio.to_thread(sess.runtime_lineage, sid) or [sid]
    pending_lineage = list(dict.fromkeys(lineage))
    queued_lineage = set(pending_lineage)
    removed = False
    cursor = 0
    while cursor < len(pending_lineage) and cursor < 32:
        runtime_sid = pending_lineage[cursor]
        cursor += 1
        await asyncio.to_thread(sess.begin_session_delete, runtime_sid)
        fenced_meta = await asyncio.to_thread(
            sess.get_session_meta, runtime_sid)
        late_successor = str(
            (fenced_meta or {}).get("runtime_successor") or "")
        if late_successor and late_successor not in queued_lineage:
            queued_lineage.add(late_successor)
            pending_lineage.append(late_successor)
        removed = (
            await _purge_single_session_storage_async(runtime_sid)
            or removed
        )
    return removed


@router.delete("/sessions/{sid}", dependencies=[Depends(require_token)])
async def delete_session_api(sid: str) -> dict:
    # 404 only when NEITHER layer existed. Previously the sidecar was
    # authoritative: an SDK-only session (sidecar lost / never written) got
    # its JSONL deleted and THEN returned 404 — the user saw a failure while
    # the transcript was already gone, and local cleanup was skipped.
    removed = await purge_session_storage_async(sid)
    if not removed:
        raise HTTPException(404, "session not found")
    return {"ok": True}


class PurgeOldReq(BaseModel):
    # Sessions whose last activity is older than `days` are deleted.
    days: int = 7
    # The caller's currently-open session — always exempt regardless of age,
    # so a bulk-clear never yanks the tab the user is staring at.
    keep_id: str = ""
    # When True, count the victims and return WITHOUT deleting anything. The
    # frontend uses this to show an exact "will delete N" in the confirm
    # dialog — it can't count locally because its session list is only the
    # most-recent paginated window (older sessions aren't loaded client-side).
    dry_run: bool = False


@router.post("/sessions/purge-old", dependencies=[Depends(require_token)])
async def purge_old_sessions_api(req: PurgeOldReq | None = None) -> dict:
    """Bulk-delete history sessions whose last activity is older than `days`
    days (default 7). Pinned sessions and `keep_id` are always exempt — pin is
    the user's explicit "keep this" signal, and deleting the currently-open
    session out from under them is jarring. Reuses purge_session_storage so
    every deleted session is cleaned to the same depth (SDK JSONL + sidecar +
    index + queue + attachments + in-memory state) as the single DELETE.

    The server is the source of truth for the victim set: it scans the FULL
    session list (list_sessions), not the paginated recent window the frontend
    holds. `dry_run=true` returns the count + ids without touching anything."""
    days = max(1, int((req.days if req else 7) or 7))
    keep_id = (req.keep_id if req else "") or ""
    dry_run = bool(req.dry_run if req else False)
    cutoff = time.time() - days * 86400
    session_rows = await obs.to_thread_io(
        "chat.session_list_read",
        keep_id,
        sess.list_sessions,
    )
    victims = [
        s["id"] for s in session_rows
        if not s.get("pinned")
        and s["id"] != keep_id
        and float(s.get("updated_at") or 0) < cutoff
    ]
    if dry_run:
        return {"ok": True, "dry_run": True, "count": len(victims),
                "ids": victims, "days": days}
    deleted: list[str] = []
    for sid in victims:
        if await purge_session_storage_async(sid):
            deleted.append(sid)
    return {"ok": True, "deleted": len(deleted), "ids": deleted, "days": days}


# --------------------------------------------------------------------------
# Server-side message queue (Option B "服务端自主执行").
#
# Queued messages live in sessions/{sid}.queue.json (sess.*_queue helpers),
# NOT in the browser. The drain trigger in _pump_gen_to_broadcast() pops the
# head item and starts the next turn whenever a turn finishes — so the queue
# advances with no browser attached. Enqueue also schedules a drain kick: this
# closes the completion race where the final one-shot drain observes an empty
# queue just before a stale browser writes its follow-up. Other endpoints are
# CRUD controls used to inspect / edit / pause the queue.
# --------------------------------------------------------------------------
class QueueEnqueueReq(BaseModel):
    text: str = ""
    # Composer-only presentation fields for selected-text attachments. The
    # SDK still receives `text` (which contains the readable quote context),
    # while queued UI can keep the textarea body and removable quote chips
    # distinct instead of showing an implementation prompt blob.
    display_text: str = ""
    selection_quotes: list[dict] | None = None
    image_ids: str = ""
    # Sender's permission mode at enqueue time. Persisted with the item so
    # the headless drain replays the turn under the same mode instead of
    # falling back to the server default (see _maybe_drain_queue).
    permission: str = ""
    # Plan Mode additionally snapshots the mode ExitPlanMode should return to.
    # Legacy/malformed values are normalized to fail-closed `default`.
    plan_return_permission: str | None = None
    # Busy-send policy. Old clients omit this and retain the historical
    # turn-boundary queue behavior. The current frontend sends ``adjust`` when
    # Settings asks MuseLab to fold text into the exact active foreground turn
    # at the CLI's next post-tool boundary.
    delivery: str = "queue"
    active_turn_id: str = ""


class QueuePauseReq(BaseModel):
    paused: bool


class QueueReorderReq(BaseModel):
    order: list[str]


def _normalize_queue_selection_quotes(raw: list[dict] | None) -> list[dict]:
    """Bound and shape browser-supplied quote-chip metadata.

    This metadata is presentation-only; `_maybe_drain_queue` sends the already
    composed `text` field. Keeping it plain and bounded prevents a malformed
    client from turning the small queue sidecar into an unbounded data store.
    """
    normalized: list[dict] = []
    for item in (raw or [])[:4]:
        if not isinstance(item, dict):
            continue
        text = str(item.get("text") or "")[:6000].strip()
        if not text:
            continue
        source = str(item.get("source") or "preview")
        role = str(item.get("role") or "")
        normalized.append({
            "id": str(item.get("id") or "")[:80],
            "source": source if source in {"preview", "chat"} else "preview",
            "role": role if role in {"", "user", "assistant"} else "",
            "sessionId": str(item.get("sessionId") or "")[:80],
            "messageId": str(item.get("messageId") or "")[:80],
            "path": str(item.get("path") or "")[:1024],
            "text": text,
            "truncated": bool(item.get("truncated")),
        })
    return normalized


@router.get("/sessions/{sid}/queue", dependencies=[Depends(require_token)])
def get_queue_api(sid: str, response: Response) -> dict:
    response.headers["Cache-Control"] = "private, no-store"
    data = sess.get_queue(sid)
    # FIX ③: resolve each queued item's attachment ids against the in-memory
    # upload store so the queued bubble can render real thumbnails / doc chips
    # (and the "撤回/编辑" recall can rebuild the input tray). The queue file
    # only persists comma-joined upload ids — the preview blobs live in
    # _image_store. Ids missing there have expired (10-min TTL); we flag them
    # `available: False` so the UI can show "附件已过期" instead of a dead chip.
    with _image_store_lock:
        _gc_images_locked()
        for it in data.get("items", []):
            ids = _attachment_ids(it.get("image_ids") or "")
            atts: list[dict] = []
            for aid in ids:
                if not _valid_staged_attachment_id(aid):
                    atts.append({"id": aid, "available": False})
                    continue
                entry = _image_store.get(aid)
                try:
                    metadata = (
                        entry or _durable_attachment_store.metadata(aid)
                    )
                except (DurableAttachmentError, OSError, sqlite3.Error,
                        UnsafePrivatePath):
                    metadata = None
                if metadata is None:
                    atts.append({"id": aid, "available": False})
                    continue
                atts.append({
                    "id": aid,
                    "kind": metadata.get("kind", "image"),
                    "name": metadata.get("name", ""),
                    "mime": metadata.get("mime", ""),
                    "available": True,
                })
            it["attachments"] = atts
    return data


async def _schedule_queue_drain_after_response(session_id: str) -> None:
    """Enter the loop-owned drain scheduler from Starlette's async callback.

    ``BackgroundTasks`` invokes synchronous callbacks in a worker thread.  The
    drain scheduler creates an ``asyncio.Task`` and therefore must stay on the
    request's event loop; an async wrapper gives Starlette the right execution
    mode.  Response background callbacks run only after the final body frame
    has been sent, so rollover/client startup cannot delay the enqueue ACK.
    """
    _schedule_queue_drain(session_id)


def _admitted_steering_turn(
    session_id: str,
    turn_id: str,
    *,
    permission: str = "",
) -> TurnBroadcast | None:
    """Return the exact foreground turn reserved for native steering.

    A session id is not sufficient: an old enqueue response can arrive after
    turn A ended and turn B reused the same pooled client.  The immutable turn
    id and final-result flag close that ABA window. Queue-owned/background turns
    are deliberately left to ordinary FIFO drain. Runtime readiness is checked
    separately so an exact admitted turn can retain a durable pending command
    while its root query is still starting.
    """
    bc = _active_turns.get(session_id)
    if (
        bc is None
        or bc.done
        or bc.cancelled
        or not turn_id
        or bc.turn_id != turn_id
        or bc.is_continuation
        or bc.is_scheduled_delivery
        or bool(bc.queue_item_id)
        or bc.result_forwarded
        or bc.steering_closed
        or _sessions_with_inflight_tasks.get(session_id)
        or _session_has_live_watcher(session_id)
    ):
        return None
    requested_permission = (permission or "").strip()
    if requested_permission and requested_permission != bc.permission:
        return None
    return bc


def _eligible_steering_turn(
    session_id: str,
    turn_id: str,
    *,
    permission: str = "",
) -> TurnBroadcast | None:
    """Return an exact admitted turn whose root SDK query is committed."""
    bc = _admitted_steering_turn(
        session_id, turn_id, permission=permission)
    if (
        bc is None
        or not bc.query_committed
        or not isinstance(bc.runtime_client, MuseLabSDKClient)
    ):
        return None
    return bc


def _publish_queue_steering(
    bc: TurnBroadcast | None,
    *,
    item_id: str,
    command_uuid: str,
    state: str,
    effective_delivery: str,
    item: dict | None = None,
) -> None:
    """Publish one privacy-bounded queue state transition when possible.

    ``started`` is also the live transcript boundary for a native steering
    command.  Include the already-bounded durable queue fields so a browser
    that missed the optimistic POST response can still replace the temporary
    queue row with the exact user bubble at that boundary.
    """
    if bc is None or bc.done:
        return
    try:
        payload: dict[str, Any] = {
            "item_id": item_id,
            "command_uuid": command_uuid,
            "state": state,
            "effective_delivery": effective_delivery,
            "turn_id": bc.turn_id,
        }
        if item is not None:
            selection_quotes = item.get("selection_quotes")
            payload["message"] = {
                "id": item_id,
                "uuid": command_uuid,
                "text": str(item.get("text") or ""),
                "display_text": str((
                    item.get("display_text")
                    if "display_text" in item else item.get("text")
                ) or ""),
                "selection_quotes": (
                    selection_quotes if isinstance(selection_quotes, list) else []
                ),
                "enqueued_at": item.get("enqueued_at"),
            }
        bc.publish({
            "event": "queue_steering",
            "data": json.dumps(payload),
        })
    except Exception:
        # The durable queue remains authoritative; GET /queue repairs any
        # missed browser transition after a reconnect.
        pass


async def _fallback_steering_item(
    session_id: str,
    *,
    item_id: str,
    command_uuid: str,
    bc: TurnBroadcast | None,
) -> dict | None:
    item = await obs.to_thread_io(
        "chat.queue_steering_fallback",
        session_id,
        sess.fallback_queue_steering,
        session_id,
        item_id=item_id,
        command_uuid=command_uuid,
        owned=True,
    )
    if item is not None:
        _publish_queue_steering(
            bc,
            item_id=item_id,
            command_uuid=command_uuid,
            state="fallback",
            effective_delivery="queue",
        )
        _schedule_queue_drain(session_id)
    return item


async def _deliver_steering_command(
    session_id: str,
    *,
    turn_id: str,
    item_id: str,
    command_uuid: str,
    text: str,
    display_text: str,
    selection_quotes: list[dict],
    permission: str,
) -> tuple[str, str, dict | None]:
    """Write one durable queue item to the exact active CLI command queue.

    The write is registered on the broadcast before the first await.  If a
    Result frame races this HTTP request, the sole stream reader waits for this
    write outcome before deciding whether that Result is the final boundary.
    """
    bc: TurnBroadcast | None = None
    write_event: asyncio.Event | None = None
    waiting_for_startup = False
    async with _lock:
        bc = _eligible_steering_turn(
            session_id, turn_id, permission=permission)
        if bc is None:
            admitted = _admitted_steering_turn(
                session_id, turn_id, permission=permission)
            if admitted is not None and not admitted.query_committed:
                bc = admitted
                waiting_for_startup = True
        if bc is not None:
            write_event = asyncio.Event()
            bc.steering_commands[command_uuid] = {
                "item_id": item_id,
                "state": "pending",
            }
            bc.steering_write_events[command_uuid] = write_event

    if bc is None or write_event is None:
        fallback = await _fallback_steering_item(
            session_id,
            item_id=item_id,
            command_uuid=command_uuid,
            bc=bc,
        )
        return "queue", "queued", fallback

    def _discard_registration() -> None:
        current = bc.steering_commands.get(command_uuid)
        if current is not None and current.get("item_id") == item_id:
            bc.steering_commands.pop(command_uuid, None)
        write_event.set()
        bc.steering_write_events.pop(command_uuid, None)

    if waiting_for_startup:
        try:
            await asyncio.wait_for(
                bc.steering_ready.wait(), timeout=30.0)
        except asyncio.TimeoutError:
            _discard_registration()
            fallback = await _fallback_steering_item(
                session_id,
                item_id=item_id,
                command_uuid=command_uuid,
                bc=bc,
            )
            return "queue", "queued", fallback
        async with _lock:
            ready = (
                _eligible_steering_turn(
                    session_id, turn_id, permission=permission) is bc
                and str((bc.steering_commands.get(command_uuid) or {}).get(
                    "item_id") or "") == item_id
            )
        if not ready:
            _discard_registration()
            fallback = await _fallback_steering_item(
                session_id,
                item_id=item_id,
                command_uuid=command_uuid,
                bc=bc,
            )
            return "queue", "queued", fallback

    # The CLI's canonical queued_command attachment only retains the model
    # prompt. Keep the bounded presentation metadata under the source UUID so
    # refresh can restore the exact visible bubble and reconcile it with the
    # live DOM identity. Commit this before the SDK write so a very fast CLI
    # cannot publish the attachment/completion before its display metadata.
    try:
        await obs.to_thread_io(
            "chat.queue_steering_annotation",
            session_id,
            sess.set_message_annotation,
            session_id,
            command_uuid,
            steering_display_text=(
                display_text if selection_quotes else None),
            steering_selection_quotes=(
                selection_quotes if selection_quotes else None),
            steering_queue_item_id=item_id,
            steering_turn_id=turn_id,
            file_path=sess._sidecar_path(session_id),
            owned=True,
        )
    except Exception as exc:
        sys.stderr.write(
            f"[chat] steering annotation failed sid={session_id[:8]} "
            f"exc={type(exc).__name__}\n")
        sys.stderr.flush()

    try:
        await bc.runtime_client.query_steering(
            text,
            session_id=session_id,
            command_uuid=command_uuid,
        )
    except BaseException:
        _discard_registration()
        fallback = await _fallback_steering_item(
            session_id,
            item_id=item_id,
            command_uuid=command_uuid,
            bc=bc,
        )
        if fallback is not None:
            return "queue", "queued", fallback
        # A terminal lifecycle may have won the race and removed the item.
        return "adjust", "completed", None

    current = bc.steering_commands.get(command_uuid)
    updated: dict | None = None
    if current is not None and current.get("state") == "pending":
        updated = await obs.to_thread_io(
            "chat.queue_steering_wait",
            session_id,
            sess.update_queue_steering_state,
            session_id,
            "waiting_tool",
            item_id=item_id,
            command_uuid=command_uuid,
            owned=True,
        )
        # A lifecycle event can advance the same dict while the durable write
        # runs. Never move queued/started back to waiting_tool in memory.
        current = bc.steering_commands.get(command_uuid)
        if current is not None and current.get("state") == "pending":
            current["state"] = "waiting_tool"
    write_event.set()
    bc.steering_write_events.pop(command_uuid, None)

    current = bc.steering_commands.get(command_uuid)
    if current is None:
        return "adjust", "completed", updated
    state = str(current.get("state") or "waiting_tool")
    _publish_queue_steering(
        bc,
        item_id=item_id,
        command_uuid=command_uuid,
        state=state,
        effective_delivery="adjust",
        item=updated,
    )
    return "adjust", state, updated


async def _settle_steering_lifecycle(
    bc: TurnBroadcast,
    message: CommandLifecycleMessage,
) -> bool:
    """Persist one CLI delivery ACK; return True for a terminal state."""
    command_uuid = message.command_uuid
    info = bc.steering_commands.get(command_uuid)
    if info is None:
        # A late ACK for a command already cancelled/fallen back must never
        # mutate a normal FIFO item. Fallback clears the durable command UUID.
        return message.state in {
            "completed", "cancelled", "discarded", "refused",
        }
    item_id = str(info.get("item_id") or "")
    state = message.state
    if state in {"queued", "started"}:
        info["state"] = state
        updated: dict | None = None
        try:
            updated = await obs.to_thread_io(
                "chat.queue_steering_lifecycle",
                bc.session_id,
                sess.update_queue_steering_state,
                bc.session_id,
                state,
                item_id=item_id,
                command_uuid=command_uuid,
                owned=True,
            )
        except Exception as exc:
            # The live CLI still owns the command. Retaining the in-memory map
            # keeps Result suppression correct; restart recovery pauses the
            # older durable state rather than risking duplicate execution.
            sys.stderr.write(
                f"[chat] steering state persist failed "
                f"sid={obs.short_id(bc.session_id)} "
                f"state={state} exc={type(exc).__name__}\n"
            )
        _publish_queue_steering(
            bc,
            item_id=item_id,
            command_uuid=command_uuid,
            state=state,
            effective_delivery="adjust",
            item=updated,
        )
        return False

    if state == "completed":
        removed: dict | None = None
        try:
            removed = await obs.to_thread_io(
                "chat.queue_steering_complete",
                bc.session_id,
                sess.update_queue_steering_state,
                bc.session_id,
                "completed",
                item_id=item_id,
                command_uuid=command_uuid,
                owned=True,
            )
            if removed is not None and str(removed.get("image_ids") or ""):
                await obs.to_thread_io(
                    "chat.queue_attachment_finish",
                    bc.session_id,
                    _durable_attachment_store.finish_queue_item,
                    bc.session_id,
                    item_id,
                    consume=True,
                    owned=True,
                )
        except Exception as exc:
            # The CLI says the command ran. Never convert an ACK-persistence
            # failure into an automatic resend: the still-adjust queue row
            # remains non-claimable and startup recovery pauses it for review.
            sys.stderr.write(
                f"[chat] steering completion persist failed "
                f"sid={obs.short_id(bc.session_id)} "
                f"exc={type(exc).__name__}\n"
            )
        bc.steering_commands.pop(command_uuid, None)
        event = bc.steering_write_events.pop(command_uuid, None)
        if event is not None:
            event.set()
        _publish_queue_steering(
            bc,
            item_id=item_id,
            command_uuid=command_uuid,
            state="completed",
            effective_delivery="adjust",
            item=removed,
        )
        return True

    bc.steering_commands.pop(command_uuid, None)
    event = bc.steering_write_events.pop(command_uuid, None)
    if event is not None:
        event.set()
    if state in {"discarded", "refused"}:
        await _fallback_steering_item(
            bc.session_id,
            item_id=item_id,
            command_uuid=command_uuid,
            bc=bc,
        )
    else:  # cancelled
        try:
            await obs.to_thread_io(
                "chat.queue_steering_cancelled",
                bc.session_id,
                sess.update_queue_steering_state,
                bc.session_id,
                "cancelled",
                item_id=item_id,
                command_uuid=command_uuid,
                pause=True,
                owned=True,
            )
        except Exception as exc:
            sys.stderr.write(
                f"[chat] steering cancellation persist failed "
                f"sid={obs.short_id(bc.session_id)} "
                f"exc={type(exc).__name__}\n"
            )
        finally:
            _publish_queue_steering(
                bc,
                item_id=item_id,
                command_uuid=command_uuid,
                state="cancelled",
                effective_delivery="adjust",
            )
    return True


async def _cancel_outstanding_steering_commands(
    bc: TurnBroadcast,
) -> None:
    """Fail closed any native commands left when their turn loses its reader.

    A transport/model failure is allowed to arrive after the CLI reported a
    command as started but before its terminal lifecycle frame. Once this
    broadcast ends no exact live owner can cancel or settle that row. Retain it
    as cancelled+paused so the user can explicitly resume, delete, or clear it.
    """
    # Close admission before the first await and before snapshotting the map.
    # Error finalization otherwise has a window where a new enqueue can attach
    # after this snapshot and lose its owner when the broadcast is popped.
    bc.steering_closed = True
    steering = list(bc.steering_commands.items())
    bc.steering_commands.clear()
    for command_uuid, info in steering:
        event = bc.steering_write_events.pop(command_uuid, None)
        if event is not None:
            event.set()
        item_id = str(info.get("item_id") or "")
        persisted: dict | None = None
        try:
            persisted = await obs.to_thread_io(
                "chat.queue_steering_cancelled",
                bc.session_id,
                sess.update_queue_steering_state,
                bc.session_id,
                "cancelled",
                item_id=item_id,
                command_uuid=command_uuid,
                pause=True,
                owned=True,
            )
        except Exception as exc:
            sys.stderr.write(
                f"[chat] steering terminal pause failed "
                f"sid={obs.short_id(bc.session_id)} "
                f"exc={type(exc).__name__}\n"
            )
        if persisted is not None:
            _publish_queue_steering(
                bc,
                item_id=item_id,
                command_uuid=command_uuid,
                state="cancelled",
                effective_delivery="adjust",
            )


async def _await_steering_write_stability(
    bc: TurnBroadcast,
    *,
    timeout_s: float = 10.0,
) -> bool:
    """Wait until every writer registered before the stable check has settled.

    Writers can register while Result is awaiting an older writer, so one
    snapshot is insufficient. Returning ``False`` means at least one write is
    still delivery-ambiguous at the bounded deadline.
    """
    loop = asyncio.get_running_loop()
    deadline = loop.time() + max(0.0, timeout_s)
    while True:
        pending_writes = tuple(
            event
            for event in bc.steering_write_events.values()
            if not event.is_set()
        )
        if not pending_writes:
            return True
        remaining = deadline - loop.time()
        if remaining <= 0:
            return False
        try:
            await asyncio.wait_for(
                asyncio.gather(*(
                    event.wait() for event in pending_writes
                )),
                timeout=remaining,
            )
        except asyncio.TimeoutError:
            return False
        # A second enqueue can register while Result waits for the first write.
        # Re-snapshot until there are no unfinished writers. The caller checks
        # the command map immediately afterward without another await.


@router.post("/sessions/{sid}/queue", dependencies=[Depends(require_token)])
async def enqueue_api(
    sid: str,
    req: QueueEnqueueReq,
    background_tasks: BackgroundTasks,
) -> dict:
    text = (req.text or "").strip()
    attachment_ids = _attachment_ids(req.image_ids or "")
    requested_delivery = (req.delivery or "queue").strip().lower()
    if requested_delivery not in {"adjust", "queue"}:
        raise HTTPException(400, "invalid queue delivery")
    requested_turn_id = (req.active_turn_id or "").strip()
    if any(not _valid_staged_attachment_id(aid) for aid in attachment_ids):
        raise HTTPException(400, "bad attachment id")
    if not text and not attachment_ids:
        raise HTTPException(400, "empty message")
    if (req.permission or "").strip():
        _validate_permission(req.permission)
    selection_quotes = _normalize_queue_selection_quotes(
        req.selection_quotes)
    display_text = (
        req.display_text
        if selection_quotes
        else (req.display_text or text)
    )

    # Native steering is text-only. Attachment ownership spans staged uploads,
    # the queue sidecar and the eventual canonical user UUID; replaying that
    # transaction inside a mid-turn CLI fold would make crash recovery
    # ambiguous, so attachment messages retain ordinary turn-boundary FIFO.
    steering_turn = None
    if requested_delivery == "adjust" and not attachment_ids:
        steering_turn = _eligible_steering_turn(
            sid, requested_turn_id, permission=req.permission or "")
        if steering_turn is None:
            admitted = _admitted_steering_turn(
                sid, requested_turn_id, permission=req.permission or "")
            # Only bridge the startup gap. A query-committed turn that fails
            # the runtime capability check remains ordinary FIFO immediately.
            if admitted is not None and not admitted.query_committed:
                steering_turn = admitted
    effective_delivery = "adjust" if steering_turn is not None else "queue"
    command_uuid = str(uuid.uuid4()) if steering_turn is not None else ""

    # Bind blobs before the queue JSON commit. A crash may leave an orphan ref
    # (startup reconciliation releases it), but can never leave an accepted
    # queue row pointing at bytes that were never made durable.
    queue_item_id = "q-" + uuid.uuid4().hex
    if attachment_ids:
        await asyncio.to_thread(_gc_images)
        bind_task = asyncio.create_task(asyncio.to_thread(
            _durable_attachment_store.bind_queue_item,
            sid,
            queue_item_id,
            attachment_ids,
            ttl=_IMAGE_TTL_S,
        ))
        try:
            binding = await asyncio.shield(bind_task)
        except asyncio.CancelledError:
            while not bind_task.done():
                try:
                    await asyncio.shield(bind_task)
                except asyncio.CancelledError:
                    continue
            binding = bind_task.result()
            if not binding.missing and not binding.busy:
                await asyncio.to_thread(
                    _durable_attachment_store.finish_queue_item,
                    sid,
                    queue_item_id,
                    consume=False,
                )
            raise
        if binding.missing:
            raise HTTPException(409, "attachment is missing or expired")
        if binding.busy:
            raise HTTPException(409, "attachment is already queued")

    enqueue_task = asyncio.create_task(
        asyncio.to_thread(
            sess.enqueue_existing_message,
            sid,
            text,
            req.image_ids or "",
            permission=req.permission or "",
            display_text=display_text,
            selection_quotes=selection_quotes,
            plan_return_permission=req.plan_return_permission,
            item_id=queue_item_id,
            delivery=effective_delivery,
            target_turn_id=(requested_turn_id
                            if effective_delivery == "adjust" else ""),
            command_uuid=command_uuid,
            steering_state=("pending"
                            if effective_delivery == "adjust" else ""),
        )
    )
    try:
        res = await asyncio.shield(enqueue_task)
    except asyncio.CancelledError:
        while not enqueue_task.done():
            try:
                await asyncio.shield(enqueue_task)
            except asyncio.CancelledError:
                continue
        res = enqueue_task.result()
        if res.get("ok"):
            if effective_delivery == "adjust":
                fallback_task = asyncio.create_task(_fallback_steering_item(
                    sid,
                    item_id=queue_item_id,
                    command_uuid=command_uuid,
                    bc=steering_turn,
                ))
                while not fallback_task.done():
                    try:
                        await asyncio.shield(fallback_task)
                    except asyncio.CancelledError:
                        continue
                fallback_task.result()
            _schedule_queue_drain(sid)
        elif attachment_ids:
            await asyncio.to_thread(
                _durable_attachment_store.finish_queue_item,
                sid,
                queue_item_id,
                consume=False,
            )
        raise
    except BaseException:
        # Atomic queue writes can report an I/O error after the rename commit.
        # Resolve that ambiguity before releasing the durable reference: an
        # item visible in waiting/inflight storage already owns the blobs.
        queue_committed = False
        try:
            queue = await asyncio.to_thread(sess.get_queue, sid)
            inflight = queue.get("inflight") or {}
            persisted_ids = {
                str(item.get("id") or "")
                for item in queue.get("items") or []
            }
            inflight_id = str(
                ((inflight.get("item") or {}).get("id") or "")
            )
            queue_committed = (
                queue_item_id in persisted_ids or inflight_id == queue_item_id
            )
        except Exception:
            # Unknown ownership fails closed: retain the ref for startup's
            # complete queue reconciliation instead of risking data loss.
            queue_committed = True
        if queue_committed:
            if effective_delivery == "adjust":
                try:
                    await asyncio.shield(_fallback_steering_item(
                        sid,
                        item_id=queue_item_id,
                        command_uuid=command_uuid,
                        bc=steering_turn,
                    ))
                except Exception:
                    # Unknown ownership remains fail-closed in the queue. Its
                    # active steering state blocks ordinary duplicate drain;
                    # startup recovery pauses it for review.
                    pass
            _schedule_queue_drain(sid)
        elif attachment_ids:
            await asyncio.to_thread(
                _durable_attachment_store.finish_queue_item,
                sid,
                queue_item_id,
                consume=False,
            )
        raise
    if not res.get("ok"):
        if attachment_ids:
            await asyncio.to_thread(
                _durable_attachment_store.finish_queue_item,
                sid,
                queue_item_id,
                consume=False,
            )
        if res.get("error") == "session_not_found":
            raise HTTPException(404, "session not found")
        raise HTTPException(409, res.get("error", "enqueue failed"))
    if effective_delivery == "adjust":
        delivery_task = asyncio.create_task(_deliver_steering_command(
            sid,
            turn_id=requested_turn_id,
            item_id=queue_item_id,
            command_uuid=command_uuid,
            text=text,
            display_text=display_text,
            selection_quotes=selection_quotes,
            permission=req.permission or "",
        ))
        try:
            delivered_as, delivery_status, updated_item = await asyncio.shield(
                delivery_task)
        except asyncio.CancelledError:
            # The durable item and CLI command must reach one known ownership
            # state even when the browser drops the enqueue response.
            while not delivery_task.done():
                try:
                    await asyncio.shield(delivery_task)
                except asyncio.CancelledError:
                    continue
            delivered_as, delivery_status, updated_item = delivery_task.result()
            if delivered_as == "queue":
                _schedule_queue_drain(sid)
            raise
        effective_delivery = delivered_as
        if updated_item is not None:
            res["item"] = updated_item
        # Lifecycle can beat the HTTP response. Return the authoritative small
        # queue snapshot so the frontend never re-adds an already-completed
        # optimistic row after its completed SSE event was delivered.
        res["queue"] = await asyncio.to_thread(sess.get_queue, sid)
    else:
        delivery_status = "queued"
        background_tasks.add_task(_schedule_queue_drain_after_response, sid)
    res["effective_delivery"] = effective_delivery
    res["delivery_status"] = delivery_status
    return res


async def _cancel_waiting_steering_item(
    sid: str, item: dict,
) -> bool:
    """Cancel one still-queued native command before its durable row is removed."""
    command_uuid = str(item.get("command_uuid") or "")
    target_turn_id = str(item.get("target_turn_id") or "")
    bc = _active_turns.get(sid)
    if (
        not command_uuid
        or bc is None
        or bc.done
        or bc.turn_id != target_turn_id
        or not isinstance(bc.runtime_client, MuseLabSDKClient)
        or str((bc.steering_commands.get(command_uuid) or {}).get(
            "item_id") or "") != str(item.get("id") or "")
    ):
        return False
    try:
        cancelled = await bc.runtime_client.cancel_async_message(command_uuid)
    except Exception:
        return False
    if not cancelled:
        return False
    bc.steering_commands.pop(command_uuid, None)
    event = bc.steering_write_events.pop(command_uuid, None)
    if event is not None:
        event.set()
    return True


@router.delete("/sessions/{sid}/queue/{item_id}", dependencies=[Depends(require_token)])
async def remove_queue_item_api(sid: str, item_id: str) -> dict:
    snapshot = await asyncio.to_thread(sess.get_queue, sid)
    item = next(
        (row for row in snapshot.get("items") or []
         if str(row.get("id") or "") == item_id),
        None,
    )
    if (
        item is not None
        and item.get("delivery") == "adjust"
        and item.get("steering_state") in {
            "pending", "waiting_tool", "queued", "started",
        }
        and not await _cancel_waiting_steering_item(sid, item)
    ):
        raise HTTPException(
            409,
            "message is already being adjusted and cannot be withdrawn safely",
        )
    try:
        updated, removed_ids = await asyncio.to_thread(
            sess.remove_queue_item_with_removed, sid, item_id)
    except ValueError as exc:
        raise HTTPException(status_code=409, detail=str(exc)) from exc
    for removed_id in removed_ids:
        await asyncio.to_thread(
            _durable_attachment_store.finish_queue_item,
            sid, removed_id, consume=False)
    return updated


@router.delete("/sessions/{sid}/queue", dependencies=[Depends(require_token)])
async def clear_queue_api(sid: str) -> dict:
    snapshot = await asyncio.to_thread(sess.get_queue, sid)
    snapshot_item_ids = tuple(
        str(item.get("id") or "")
        for item in snapshot.get("items") or []
        if item.get("id")
    )
    active_adjustments = [
        item for item in snapshot.get("items") or []
        if item.get("delivery") == "adjust"
        and item.get("steering_state") in {
            "pending", "waiting_tool", "queued", "started",
        }
    ]
    cancelled: list[dict] = []
    for item in active_adjustments:
        if await _cancel_waiting_steering_item(sid, item):
            cancelled.append(item)
            continue
        for prior in cancelled:
            await asyncio.to_thread(
                sess.update_queue_steering_state,
                sid,
                "cancelled",
                item_id=str(prior.get("id") or ""),
                command_uuid=str(prior.get("command_uuid") or ""),
                pause=True,
            )
        raise HTTPException(
            409,
            "one or more adjusted messages cannot be cleared safely",
        )
    # Remove exactly the rows represented by the snapshot above. A concurrent
    # enqueue may already have crossed the native CLI write boundary by now;
    # an unconditional clear would erase its durable owner without cancelling
    # the accepted command.
    updated, item_ids = await asyncio.to_thread(
        sess.remove_queue_items_with_removed, sid, snapshot_item_ids)
    for item_id in item_ids:
        await asyncio.to_thread(
            _durable_attachment_store.finish_queue_item,
            sid, item_id, consume=False)
    return {"ok": True, **updated}


@router.post("/sessions/{sid}/queue/pause", dependencies=[Depends(require_token)])
async def pause_queue_api(sid: str, req: QueuePauseReq) -> dict:
    data = await obs.to_thread_io(
        "chat.queue_pause", sid, sess.set_queue_paused, sid, req.paused,
        owned=True,
    )
    # Resuming kicks the drain in case no turn is currently running for this
    # session (otherwise the next item would wait for a turn that never comes).
    if not req.paused:
        await _maybe_drain_queue(sid)
    return data


@router.post("/sessions/{sid}/queue/reorder", dependencies=[Depends(require_token)])
def reorder_queue_api(sid: str, req: QueueReorderReq) -> dict:
    return sess.reorder_queue(sid, req.order)


# Orphan attachments sweep — defends against the case where a JSONL was
# deleted out of band (manual rm, git restore, etc.) and left an
# attachments/<sid>/ behind. Runs lazily off the existing session-list
# endpoint so we don't need a separate cron. Bounded — only sweeps if
# attachments dir actually has children.
def _gc_orphan_attachments() -> None:
    base = _attachments_base()
    if not ensure_private_directory(base, create=False):
        return
    try:
        # list_sessions() intentionally hides sessions belonging to a removed
        # workspace.  Those rows still exist and reappear when the workspace
        # is registered again, so attachment GC must use the unfiltered index.
        known_sids = sess.indexed_session_ids()
    except Exception:
        return
    for child in base.iterdir():
        if repair_private_path(child) != "directory":
            continue
        if child.name not in known_sids:
            try:
                shutil.rmtree(child, ignore_errors=True)
            except OSError:
                pass


@router.get("/attachments-usage", dependencies=[Depends(require_token)])
def attachments_usage() -> dict:
    """Total bytes + file count under sessions/attachments. UI / settings
    can render this so users know how much disk their uploaded images
    have eaten, and can trigger a sweep."""
    base = _attachments_base()
    if not ensure_private_directory(base, create=False):
        return {"total_bytes": 0, "file_count": 0, "session_count": 0}
    total = 0
    files = 0
    sessions_with_attach = 0
    for sid_dir in base.iterdir():
        if repair_private_path(sid_dir) != "directory":
            continue
        has_any = False
        for attachment in sid_dir.iterdir():
            if repair_private_path(attachment) != "file":
                continue
            try:
                total += attachment.lstat().st_size
                files += 1
                has_any = True
            except OSError:
                pass
        if has_any:
            sessions_with_attach += 1
    return {
        "total_bytes": total,
        "file_count": files,
        "session_count": sessions_with_attach,
    }


@router.post("/attachments-sweep", dependencies=[Depends(require_token)])
def attachments_sweep() -> dict:
    """Manually trigger the orphan-attachments sweep + return new usage."""
    _gc_orphan_attachments()
    return attachments_usage()


class SessionPatchReq(BaseModel):
    name: str | None = None
    model: str | None = None
    permission: str | None = None
    # SDK-native session tag — written to CLI JSONL so other tools (and
    # manual `claude` CLI runs) see it. Pass empty string to clear.
    tag: str | None = None
    # Pin to top of the session picker. None = no change, True/False = set.
    pinned: bool | None = None
    # Reasoning effort knob. Empty string is accepted only as a legacy spelling
    # of canonical `auto`; Ultra is transported to Codex Gateway by header.
    effort: str | None = None
    # User-facing service class. Empty = standard; `fast` maps to Codex's
    # priority tier inside Gateway. It is deliberately separate from effort.
    service_tier: str | None = None
    # Extended-thinking on/off for this session. None = no change. False
    # disables thinking (escape hatch for the streaming-interleaving 400);
    # rebuilds the client so the next turn picks it up.
    thinking: bool | None = None


@router.patch("/sessions/{sid}", dependencies=[Depends(require_token)])
async def patch_session_api(sid: str, req: SessionPatchReq) -> dict:
    # Model, effort, and service tier form one runtime contract. Validate the
    # complete *target* tuple before mutating any field in this request; this
    # makes a rejected cross-model combination side-effect free.
    current_meta = await obs.to_thread_io(
        "chat.session_read", sid, sess.get_session, sid)
    runtime_controls_requested = any(
        value is not None
        for value in (req.model, req.effort, req.service_tier)
    )
    target_model = (
        req.model if req.model is not None
        else ((current_meta or {}).get("model") or "")
    )
    target_effort = (
        _normalize_effort(req.effort)
        if req.effort is not None
        else _normalize_effort((current_meta or {}).get("effort"))
    )
    target_tier = (
        (req.service_tier or "").strip()
        if req.service_tier is not None
        else ((current_meta or {}).get("service_tier") or "").strip()
    )
    runtime_controls_changed = False
    model_changed = False
    if runtime_controls_requested:
        if current_meta is None:
            raise HTTPException(404, "session not found")
        if target_effort not in _VALID_EFFORT:
            raise HTTPException(400, f"invalid effort: {req.effort}")
        if target_tier not in _VALID_SERVICE_TIERS:
            raise HTTPException(
                400, f"invalid service tier: {req.service_tier}")
        capability = await _detect_gateway_context_capability(target_model)
        controls = _model_control_capability(target_model, capability)
        if (target_effort != "auto"
                and target_effort not in controls["effort_levels"]):
            raise HTTPException(
                400,
                f"effort {target_effort} is not supported by {target_model}",
            )
        if target_tier == "fast" and not controls["supports_fast"]:
            raise HTTPException(
                400, f"fast service is not supported by {target_model}")
        current_effort = _normalize_effort(current_meta.get("effort"))
        current_tier = (current_meta.get("service_tier") or "").strip()
        current_model = (current_meta.get("model") or "").strip()
        model_changed = target_model != current_model
        runtime_controls_changed = (
            model_changed
            or target_effort != current_effort
            or target_tier != current_tier
        )

    ok = False
    if req.name is not None:
        rename_started = time.monotonic()
        lock_wait_ms = 0
        local_index_ms = 0
        sdk_rename_ms = 0
        activity_ms = 0
        rename_status = "error"
        renamed = False
        try:
            # Keep the local title and SDK customTitle as one serialized write.
            # Runtime-rollover postlude propagation uses the same lock and rechecks
            # the inherited title inside it, so a user's explicit rename always
            # wins rather than being overwritten by a late automatic sync.
            def _rename_transaction() -> tuple[bool, int, int, int]:
                lock_started = time.monotonic()
                with _session_title_lock(sid):
                    waited = round((time.monotonic() - lock_started) * 1000)
                    local_started = time.monotonic()
                    local_renamed = sess.rename_session(sid, req.name)
                    local_ms = round(
                        (time.monotonic() - local_started) * 1000)
                    # Also propagate to CLI's JSONL so list_sessions() / manual
                    # claude CLI runs see the new title. Silent no-op if absent.
                    sdk_started = time.monotonic()
                    try:
                        sdk_rename_session(
                            sid, req.name,
                            directory=str(sess.session_workspace(sid)))
                    except (FileNotFoundError, ValueError):
                        pass
                    sdk_ms = round(
                        (time.monotonic() - sdk_started) * 1000)
                    return local_renamed, waited, local_ms, sdk_ms

            renamed, lock_wait_ms, local_index_ms, sdk_rename_ms = (
                await obs.to_thread_io(
                    "chat.session_rename", sid, _rename_transaction,
                    owned=True,
                )
            )
            ok = renamed or ok
            if renamed:
                # The task ledger stores one denormalized display name per
                # conversation.  Keep it in lockstep with the session index and
                # publish the targeted row over the existing Activity SSE.  This
                # intentionally preserves task timestamps/read state, so a rename
                # cannot reorder the task center or make a result unread again.
                from .activity import activity as _activity
                activity_started = time.monotonic()
                try:
                    await asyncio.to_thread(
                        _activity.rename_session, sid, req.name)
                finally:
                    activity_ms = round(
                        (time.monotonic() - activity_started) * 1000)
            rename_status = "ok" if renamed else "not_found"
        finally:
            _perf_event(
                "session.rename",
                session=obs.short_id(sid) or "none",
                status=rename_status,
                lock_wait_ms=lock_wait_ms,
                local_index_ms=local_index_ms,
                sdk_rename_ms=sdk_rename_ms,
                activity_ms=activity_ms,
                total_ms=round((time.monotonic() - rename_started) * 1000),
            )
    if req.tag is not None:
        # Empty string → clear tag. SDK accepts None or str.
        try:
            def _tag_session() -> None:
                sdk_tag_session(
                    sid,
                    req.tag or None,
                    directory=str(sess.session_workspace(sid)),
                )

            await obs.to_thread_io(
                "chat.session_tag",
                sid,
                _tag_session,
                owned=True,
            )
            ok = True
        except (FileNotFoundError, ValueError) as e:
            # JSONL doesn't exist yet → tag has nowhere to live until first
            # query. Surface as a 409 so the FE can wait for first turn.
            raise HTTPException(409, f"cannot tag session before first turn: {e}")
    if req.pinned is not None:
        # Pin is muselab-local (not stored in CLI JSONL). Always idempotent.
        # set_pin runs the load-mutate-save sequence under _INDEX_LOCK.
        await obs.to_thread_io(
            "chat.session_pin", sid, sess.set_pin, sid, req.pinned,
            owned=True,
        )
        ok = True
    if req.permission is not None:
        permission = _validate_permission(req.permission)
        current_meta = await obs.to_thread_io(
            "chat.session_read", sid, sess.get_session, sid)
        current_meta = current_meta or {}
        current_permission = (current_meta.get("permission") or "").strip()
        changed = current_permission != permission
        updated = await obs.to_thread_io(
            "chat.session_permission",
            sid,
            sess.update_permission,
            sid,
            permission,
            owned=True,
        )
        ok = updated or ok
        if updated and changed:
            # Permission is a launch contract. A busy session defers the
            # replacement to its turn boundary; an idle one is evicted now so
            # the next send cannot reuse a process with stale capabilities.
            await _rebuild_session_runtime(sid)
    if runtime_controls_requested:
        # A model swap may race a live CLI writing the same JSONL. Preserve the
        # existing interrupt-before-rebuild safety, but only after the complete
        # target tuple has passed validation.
        if model_changed:
            bc = _active_turns.get(sid)
            if bc is not None and not bc.done:
                async with _lock:
                    live_clients = [
                        c for k, c in _clients.items() if k[0] == sid
                    ]
                for c in live_clients:
                    try:
                        await c.interrupt()
                    except Exception as _e:
                        sys.stderr.write(
                            f"[chat] interrupt before model swap failed for "
                            f"{sid}: {type(_e).__name__}: {_e}\n")
        if runtime_controls_changed:
            updated = await obs.to_thread_io(
                "chat.session_runtime_controls",
                sid,
                sess.update_runtime_controls,
                sid,
                model=target_model,
                effort=target_effort,
                service_tier=target_tier,
                owned=True,
            )
            ok = updated or ok
            if updated:
                # All three values are baked into launch/request plumbing; one
                # atomic index write and one runtime rebuild is sufficient.
                await _rebuild_session_runtime(sid)
        else:
            ok = True
    if req.thinking is not None:
        # No-op guard, same rationale as effort: toggling thinking forces a
        # client rebuild (thinking config is fixed at construction). Skip when
        # unchanged. Default is True, so a missing field reads as enabled.
        thinking_meta = await obs.to_thread_io(
            "chat.session_read", sid, sess.get_session, sid)
        cur_thinking = bool(
            (thinking_meta or {}).get("thinking", True))
        if bool(req.thinking) != cur_thinking:
            await obs.to_thread_io(
                "chat.session_thinking",
                sid,
                sess.update_thinking,
                sid,
                bool(req.thinking),
                owned=True,
            )
            await _rebuild_session_runtime(sid)
        ok = True
    if not ok:
        raise HTTPException(404, "session not found or no changes")
    return {"ok": True}


# ====== usage / reset ======

@router.get("/rate-limit", dependencies=[Depends(require_token)])
def rate_limit() -> dict:
    """Latest Pro/Max rate-limit snapshot, per window, as last pushed by the
    SDK's RateLimitEvent. SSE delivers live deltas (`rate_limit` event); this
    endpoint gives a freshly-loaded page the current state without waiting for
    the next turn. `windows` is empty until the first event arrives this
    process (and stays empty for pure third-party / API-key setups, which the
    CLI never rate-limit-reports). `updated_at` is 0.0 when never seen."""
    return {
        "windows": _rate_limit_state,
        "updated_at": _rate_limit_updated_at,
    }


def _codex_home() -> Path:
    return Path(os.environ.get("CODEX_HOME") or (Path.home() / ".codex"))


def _codex_rate_limit_type(key: str, window: dict) -> str:
    minutes = int(window.get("window_minutes") or 0)
    if minutes == 300:
        return "five_hour"
    if minutes == 10080:
        return "seven_day"
    # Treat the common calendar-month approximations as monthly. Keep the
    # original window_minutes in the payload so callers can still display the
    # exact reset horizon if Codex changes the duration.
    if 28 * 24 * 60 <= minutes <= 31 * 24 * 60:
        return "monthly"
    return key


def _codex_rate_limits_from_payload(payload: dict, source: Path, ts: str | None) -> dict | None:
    raw = payload.get("rate_limits")
    if not isinstance(raw, dict):
        return None
    windows: dict[str, dict] = {}
    reached = raw.get("rate_limit_reached_type")
    for key in ("primary", "secondary"):
        w = raw.get(key)
        if not isinstance(w, dict):
            continue
        used = w.get("used_percent")
        try:
            used_f = float(used)
        except (TypeError, ValueError):
            used_f = None
        status = "allowed"
        if reached and (reached == key or reached == _codex_rate_limit_type(key, w)):
            status = "rejected"
        elif used_f is not None and used_f >= 90:
            status = "allowed_warning"
        windows[key] = {
            "rate_limit_type": _codex_rate_limit_type(key, w),
            "window_minutes": int(w.get("window_minutes") or 0),
            "resets_at": int(w.get("resets_at") or 0) or None,
            "used_percent": used_f,
            "remaining_percent": (round(max(0.0, 100.0 - used_f), 1)
                                  if used_f is not None else None),
            # Match the Claude SDK shape consumed by the existing FE badge.
            "utilization": (used_f / 100.0 if used_f is not None else None),
            "status": status,
        }
    if not windows:
        return None
    updated_at = 0.0
    if ts:
        try:
            updated_at = datetime.fromisoformat(ts.replace("Z", "+00:00")).timestamp()
        except ValueError:
            updated_at = 0.0
    return {
        "ok": True,
        "source": "codex-session-log",
        "source_scope": "codex_cli_session_log",
        "provider_authoritative": False,
        "source_file": str(source),
        "updated_at": updated_at,
        "timestamp": ts,
        "limit_id": raw.get("limit_id"),
        "limit_name": raw.get("limit_name"),
        "plan_type": raw.get("plan_type"),
        "rate_limit_reached_type": reached,
        "credits": raw.get("credits"),
        "individual_limit": raw.get("individual_limit"),
        "windows": windows,
    }


def _latest_codex_rate_limits() -> dict:
    """Read the newest Codex quota snapshot from local Codex session JSONL.

    Codex already writes rate-limit snapshots into token_count events. Reading
    those logs avoids touching ~/.codex/auth.json or calling private OpenAI
    endpoints. We only inspect lines containing the literal "rate_limits" and
    stop at the newest usable event.
    """
    home = _codex_home()
    sessions_dir = home / "sessions"
    if not sessions_dir.exists():
        return {"ok": False, "reason": "codex_sessions_missing", "windows": {}, "updated_at": 0}
    try:
        files = sorted(
            sessions_dir.rglob("*.jsonl"),
            key=lambda p: p.stat().st_mtime_ns,
            reverse=True,
        )
    except OSError as e:
        return {"ok": False, "reason": f"codex_sessions_unreadable: {e}", "windows": {},
                "updated_at": 0}
    max_files = max(1, env_int("MUSELAB_CODEX_RATE_LIMIT_SCAN_FILES", 80, min_value=1))
    for path in files[:max_files]:
        try:
            lines = path.read_text(encoding="utf-8", errors="ignore").splitlines()
        except OSError:
            continue
        for line in reversed(lines):
            if '"rate_limits"' not in line:
                continue
            try:
                event = json.loads(line)
            except json.JSONDecodeError:
                continue
            payload = event.get("payload")
            if not isinstance(payload, dict):
                continue
            parsed = _codex_rate_limits_from_payload(payload, path, event.get("timestamp"))
            if parsed:
                return parsed
    return {"ok": False, "reason": "codex_rate_limits_not_found", "windows": {},
            "updated_at": 0}


def _refresh_codex_rate_limits() -> dict:
    script = Path(__file__).resolve().parents[1] / "scripts" / "codex-quota-refresh.py"
    timeout = max(5, env_int("MUSELAB_CODEX_QUOTA_TIMEOUT", 25, min_value=5))
    if not script.exists():
        return {"ok": False, "reason": "codex_quota_script_missing"}
    try:
        proc = subprocess.run(
            [sys.executable, str(script), "--timeout", str(timeout)],
            cwd=str(ROOT or Path.home()),
            text=True,
            capture_output=True,
            timeout=timeout + 5,
            check=False,
        )
    except subprocess.TimeoutExpired:
        return {"ok": False, "reason": "codex_quota_script_timeout"}
    except OSError as e:
        return {"ok": False, "reason": f"codex_quota_script_start_failed: {e}"}
    try:
        payload = json.loads(proc.stdout.strip().splitlines()[-1])
    except (IndexError, json.JSONDecodeError):
        return {
            "ok": False,
            "reason": "codex_quota_script_bad_output",
            "returncode": proc.returncode,
            "stderr_tail": proc.stderr[-800:],
        }
    if not payload.get("ok") and proc.returncode != 0:
        payload.setdefault("returncode", proc.returncode)
        payload.setdefault("stderr_tail", proc.stderr[-800:])
    return payload


@router.get("/codex-rate-limit", dependencies=[Depends(require_token)])
def codex_rate_limit(refresh: bool = Query(default=False)) -> dict:
    """Latest quota and usage for the locally authenticated Codex account.

    A refresh prefers the Codex app-server's read-only
    ``account/rateLimits/read`` and ``account/usage/read`` RPCs. Older Codex
    builds fall back to existing local session snapshots. Neither path reads
    OAuth credential files or sends a model request. The result describes the
    local Codex account, not necessarily CLIProxyAPI's Gateway identity.
    """
    if refresh:
        refreshed = _refresh_codex_rate_limits()
        if refreshed.get("ok"):
            return refreshed
        fallback = _latest_codex_rate_limits()
        fallback["refresh"] = refreshed
        return fallback
    return _latest_codex_rate_limits()


@router.get("/usage", dependencies=[Depends(require_token)])
async def usage() -> dict:
    cr = _stats.get("total_cache_read_tokens", 0)
    in_t = _stats.get("total_input_tokens", 0)
    cache_pct = round(cr / (cr + in_t) * 100, 1) if (cr + in_t) > 0 else 0
    # Snapshot under _lock — iterating _clients.keys() unlocked can RuntimeError
    # if another coroutine resizes the dict mid-iteration. Also expose only the
    # session_id (k[0]), not the raw runtime key, to avoid
    # leaking internal pool structure in the response.
    async with _lock:
        active_session_ids = sorted({k[0] for k in _clients})
    return {**_stats, "model_default": MODEL,
            "active_sessions": active_session_ids,
            "cache_hit_pct": cache_pct,
            "budget_usd": _budget_usd(),
            "budget_used_pct": (
                round(_stats["total_cost_usd"] / _budget_usd() * 100, 1)
                if _budget_usd() > 0 else 0
            )}


def _usage_source_signature(path: Path | None) -> dict | None:
    """Identity and freshness fence for one canonical JSONL generation."""
    if path is None:
        return None
    try:
        stat = path.stat()
    except OSError:
        return None
    return {
        "dev": int(stat.st_dev),
        "inode": int(stat.st_ino),
        "size": int(stat.st_size),
        "mtime_ns": int(stat.st_mtime_ns),
    }


def _session_usage_summary_payload(
    usage_snapshot: dict,
    *,
    turn_id: str = "",
    source: dict | None | object = _USAGE_SOURCE_UNSET,
    sid: str = "",
) -> dict | None:
    if source is _USAGE_SOURCE_UNSET:
        source = _usage_source_signature(_find_session_jsonl(sid))
    if source is None:
        return None
    return {
        "schema": _USAGE_SUMMARY_SCHEMA,
        "source": dict(source),
        "update": {"turn_id": str(turn_id or ""), "at": time.time()},
        "normalized": dict(usage_snapshot),
    }


def _persist_session_usage_summary(
    sid: str,
    usage_snapshot: dict,
    *,
    turn_id: str = "",
    source: dict | None | object = _USAGE_SOURCE_UNSET,
    require_matching_turn: bool = False,
) -> bool:
    summary = _session_usage_summary_payload(
        usage_snapshot, turn_id=turn_id, source=source, sid=sid)
    if summary is None:
        return False
    try:
        if require_matching_turn:
            return bool(sess.set_session_usage_summary_if_turn_matches(
                sid, turn_id, summary))
        return bool(sess.set_session_usage_summary(sid, summary))
    except Exception as exc:
        # Hydration repair is best-effort. A corrupt or unwritable sidecar must
        # remain byte-identical and must never turn readable JSONL into /usage 500.
        sys.stderr.write(
            f"[chat-usage] summary write skipped sid={obs.short_id(sid)} "
            f"exc={type(exc).__name__}\n")
        return False


def _load_session_usage_summary(sid: str) -> tuple[dict, str] | None:
    source = _usage_source_signature(_find_session_jsonl(sid))
    if source is None:
        return None
    try:
        summary = sess.get_session_usage_summary(sid)
    except Exception:
        return None
    if not isinstance(summary, dict):
        return None
    normalized = summary.get("normalized")
    update = summary.get("update")
    if (summary.get("schema") != _USAGE_SUMMARY_SCHEMA
            or summary.get("source") != source
            or not isinstance(normalized, dict)
            or not isinstance(update, dict)):
        return None
    return dict(normalized), str(update.get("turn_id") or "")


def _refine_session_usage_for_turn(
    sid: str,
    turn_id: str,
    refined: dict,
) -> bool:
    """Replace a usage snapshot only while the same exact turn still owns it."""
    if not turn_id or _session_usage_turns.get(sid) != turn_id:
        return False
    _session_usage[sid] = dict(refined)
    return True


def _session_usage_from_jsonl(sid: str) -> dict | None:
    """Rebuild a session_usage snapshot from the CLI JSONL transcript.

    Why this exists: `_session_usage` is in-memory and clears on every
    muselab restart. After restart, switching to an existing session
    used to show an empty context meter until the user sent a new
    message — a confusing "did my conversation vanish?" UX even though
    the transcript was still there. Now we lazily rebuild from disk on
    miss.

    What we extract:
      - last assistant turn's `message.usage` → input / output / cache
        tokens (gives the "current context window" estimate the meter
        cares about)
      - sum of cost annotations from the muselab sidecar → cumulative
        total_cost_usd

    Returns None when no JSONL exists (truly new session) so the
    caller can fall through to a zero-shaped default. The walk is
    O(n_lines) per session and only fires on a cache miss; subsequent
    polls hit `_session_usage` again.
    """
    if ROOT is None:
        return None
    jsonl_path = _find_session_jsonl(sid)
    if jsonl_path is None:
        return None
    last_usage: dict[str, int] = {}
    last_ts: float = 0.0
    last_model: str = ""

    def _extract(line: str) -> tuple[dict, str, float] | None:
        """(usage, model, ts) for an assistant line that carries usage, else None."""
        if '"type":"assistant"' not in line and '"type": "assistant"' not in line:
            return None
        try:
            entry = json.loads(line)
        except (json.JSONDecodeError, ValueError):
            return None
        msg = entry.get("message") or {}
        u = msg.get("usage") or {}
        if not isinstance(u, dict) or not u:
            return None
        ts_val = 0.0
        raw_ts = entry.get("timestamp") or ""
        if raw_ts:
            try:
                import datetime as _dt
                ts_val = _dt.datetime.fromisoformat(
                    raw_ts.replace("Z", "+00:00")).timestamp()
            except ValueError:
                ts_val = 0.0
        return (u, msg.get("model") or "", ts_val)

    # Fast path: the assistant turn whose usage the meter wants is the most
    # recent one, sitting at the very END of the transcript. Read only the TAIL
    # via _read_tail_lines (O(tail)) instead of walking a possibly-100MB+ file
    # from the top — this fires on every tab switch (cache miss), so the full
    # walk was a real hot-path cost. Scan the tail in reverse, stop at the first
    # usage-bearing assistant. Fall back to a full forward scan only if the tail
    # window holds none (e.g. a final turn longer than the window).
    try:
        _tail = _read_tail_lines(jsonl_path, 2000)
    except Exception:
        _tail = None
    if _tail:
        for line in reversed(_tail):
            got = _extract(line)
            if got is not None:
                last_usage, _m, last_ts = got
                last_model = _m or last_model
                break
    if not last_usage:
        try:
            with jsonl_path.open("r", encoding="utf-8") as f:
                for line in f:
                    got = _extract(line)
                    if got is not None:
                        last_usage, _m, last_ts = got
                        last_model = _m or last_model
        except OSError:
            return None
    if not last_usage:
        return None
    # Cumulative cost — sum sidecar annotations. Cheaper than reparsing
    # JSONL costs, and Anthropic is the only vendor that puts USD in
    # message.usage anyway.
    total_cost = 0.0
    try:
        from . import sessions as _sess
        anns = _sess.get_message_annotations(sid)
        for ann in anns.values():
            if isinstance(ann, dict):
                total_cost += _parse_cost(ann.get("cost"))
    except Exception:
        pass
    in_t = int(last_usage.get("input_tokens", 0) or 0)
    out_t = int(last_usage.get("output_tokens", 0) or 0)
    cr_t = int(last_usage.get("cache_read_input_tokens", 0)
                or last_usage.get("cache_read_tokens", 0) or 0)
    cc_t = int(last_usage.get("cache_creation_input_tokens", 0)
                or last_usage.get("cache_creation_tokens", 0) or 0)
    ctx_used = in_t + cr_t + cc_t
    # Prefer the SDK-authoritative window persisted from a prior turn's
    # get_context_usage() (survives restart). Only fall back to the hardcoded
    # table when this session has never been measured — that guess was the
    # source of the "meter reads too low after restart" bug.
    sdk_window = None
    try:
        sdk_window = sess.get_session_ctx_window(sid)
    except Exception:
        sdk_window = None
    details = _context_limit_details(
        last_model, stored=_positive_int(sdk_window))
    limit = _positive_int(details.get("context_limit"))
    pct = round(ctx_used / limit * 100, 1) if limit else 0.0
    rebuilt = {
        "input_tokens": in_t, "output_tokens": out_t,
        "cache_read_tokens": cr_t, "cache_creation_tokens": cc_t,
        "total_cost_usd": total_cost, "last_turn_at": last_ts,
        "context_used": ctx_used, "context_used_pct": pct,
        **details,
    }
    _mark_context_used(rebuilt, "provider_usage_transcript", estimate=False)
    return rebuilt


@router.get("/usage/{session_id}", dependencies=[Depends(require_token)])
async def session_usage(session_id: str, model: str = "") -> dict:
    """Per-session context meter — what fraction of the model's window we're at.

    Note: this is the cheap path — reads cached per-turn usage values.
    On process-cache miss it reads the validated sidecar summary. JSONL is only
    scanned to repair a missing or stale summary, never for routine tab loads.
    For a true breakdown (per CLAUDE.md file, per MCP tool, per skill),
    use /context-breakdown/{session_id} which invokes
    ClaudeSDKClient.get_context_usage() against the live session."""
    u = _session_usage.get(session_id)
    if u is None:
        persisted = await obs.to_thread_io(
            "chat.usage_summary_read",
            session_id,
            _load_session_usage_summary,
            session_id,
            file_path=lambda: sess._sidecar_path(session_id),
        )
        if persisted is not None:
            u, summary_turn_id = persisted
            _session_usage[session_id] = u
            if summary_turn_id:
                _session_usage_turns[session_id] = summary_turn_id
        else:
            rebuilt = await obs.to_thread_io(
                "chat.usage_hydrate",
                session_id,
                _session_usage_from_jsonl,
                session_id,
                file_path=lambda: _find_session_jsonl(session_id),
            )
            if rebuilt is not None:
                # Populate the cache and best-effort repair the summary. Sidecar
                # corruption/write failure stays fail-closed and cannot fail /usage.
                _session_usage[session_id] = rebuilt
                u = rebuilt
                await obs.to_thread_io(
                    "chat.usage_summary_repair",
                    session_id,
                    _persist_session_usage_summary,
                    session_id,
                    rebuilt,
                    file_path=lambda: sess._sidecar_path(session_id),
                )
            else:
                u = None
        if u is None:
            u = {
                "input_tokens": 0, "output_tokens": 0,
                "cache_read_tokens": 0, "cache_creation_tokens": 0,
                "total_cost_usd": 0.0, "last_turn_at": 0.0,
                "context_used": 0, "context_used_pct": 0.0,
                "context_limit": 0,
            }
    m = model or MODEL
    # Claude uses the per-session SDK window persisted across restarts. Codex
    # uses CLIProxyAPI's live model catalog (with an explicit operator override
    # above it), so a stale 200K value from an older Claude CLI client cannot
    # leak back into the denominator. Other third-party providers keep their
    # established SDK/table fallback behavior.
    sdk_window = None
    try:
        sdk_window = await obs.to_thread_io(
            "chat.session_context_window_read",
            session_id,
            sess.get_session_ctx_window,
            session_id,
        )
    except Exception:
        sdk_window = None
    stored = int(u.get("context_limit", 0) or 0)
    capability = await _detect_gateway_context_capability(m)
    details = _context_limit_details(
        m,
        sdk_max=_positive_int(sdk_window),
        stored=stored,
        capability=capability,
    )
    limit = _positive_int(details.get("context_limit"))
    # Prefer SDK-authoritative numbers populated by the stream's ResultMessage
    # handler. Fall back to the legacy estimate only if no turn has completed
    # yet (in which case `context_used` is 0 anyway → 0% display, correct).
    if u.get("context_used"):
        ctx_used = int(u["context_used"])
        # Recompute pct against possibly-bumped limit so it doesn't show stale
        # high percentage (e.g. 14.2% if computed against 200K but limit is 1M).
        ctx_pct = round(ctx_used / limit * 100, 1) if limit else 0.0
    else:
        # Conservative fallback: per-turn input only (not summed with cache,
        # because cache_read/cache_creation in SDK usage are cumulative and
        # would inflate the meter — see ResultMessage handler comment).
        ctx_used = int(u.get("input_tokens", 0) or 0)
        ctx_pct = round(ctx_used / limit * 100, 1) if limit else 0
    result = {
        **u,
        "model": m,
        **details,
        "context_used": ctx_used,
        "context_used_pct": ctx_pct,
    }
    if not result.get("context_used_source"):
        _mark_context_used(
            result,
            "input_fallback" if ctx_used else "none",
            estimate=bool(ctx_used),
        )
    else:
        result["context_is_estimate"] = bool(
            result.get("context_used_is_estimate", False)
            or result.get("context_limit_is_estimate", False))
    return result


def _parse_cost(raw: Any) -> float:
    """Sidecar stores cost as the formatted string we showed in the UI
    (e.g. '$0.1993'). Parse back to a float for aggregation. Returns 0.0
    for missing / unparseable values."""
    if raw is None:
        return 0.0
    if isinstance(raw, (int, float)):
        return float(raw)
    if isinstance(raw, str):
        s = raw.strip().lstrip("$").replace(",", "")
        try:
            return float(s)
        except ValueError:
            return 0.0
    return 0.0


def _empty_bucket() -> dict:
    """Per-time-bucket aggregator shape. Used by cost_dashboard to add
    up arbitrary turn slices. Cost comes from sidecar (vendor knows
    pricing); tokens come from JSONL (universal across all vendors)."""
    return {"input_tokens": 0, "output_tokens": 0,
             "cache_read_tokens": 0, "cache_creation_tokens": 0,
             "cost": 0.0, "turns": 0}


def _vendor_label_for(model_id: str) -> str:
    """Pretty vendor name for the cost-dashboard `by_vendor` rollup.
    Claude lives outside CATALOG (we serve OAuth Pro/Max directly, not via
    a third-party endpoint) so map it explicitly; third-parties fall through
    to their CATALOG `display` field; truly unknown ids land in 'Unknown'."""
    if not model_id:
        return "Unknown"
    low = model_id.lower()
    if low.startswith("claude-"):
        return "Claude"
    p = endpoints.lookup(model_id)
    if p is not None:
        return p.display
    return "Unknown"


def _cost_reported_for(model_id: str) -> bool:
    """True when this vendor actually reports USD cost in muselab sidecar
    (= the FE can trust the $-figure). Currently only the Claude path
    (Anthropic Pro/Max OAuth or direct API key) populates ResultMessage's
    `total_cost_usd`; DeepSeek / GLM / MiniMax always come through as 0 and
    we don't want the dashboard pretending they're free. FE uses this to
    show a 'cost not tracked — vendor doesn't report USD' footnote."""
    return (model_id or "").lower().startswith("claude-")


def _add_bucket(dst: dict, src: dict) -> None:
    for k, v in src.items():
        if k in dst:
            dst[k] += v


# Cost-dashboard response cache. The handler re-reads every session JSONL +
# sidecar on each call (token truth lives only on disk) — O(hundreds of files),
# measured ~8s on a large archive with no caching. The inputs only change when
# a turn is written (new / grown JSONL) or a sidecar cost updates, so we cache
# the full response keyed by (days, tz, today) and invalidate on a cheap
# fingerprint of the input file set. A fingerprint match returns the cached
# dict; a mismatch recomputes. Guarded by a plain threading.Lock because
# cost_dashboard is a sync FastAPI endpoint (runs in the threadpool, can be hit
# concurrently). today is in the key (not the fingerprint) so a midnight
# rollover with no new data still recomputes the date-bucketed window.
_dashboard_cache: dict[tuple, tuple] = {}   # (days, tz, today) -> (fingerprint, response)
_dashboard_cache_lock = threading.Lock()


@router.get("/cost-dashboard", dependencies=[Depends(require_token)])
def cost_dashboard(days: int = Query(default=30, ge=1, le=365),
                    tz_offset_minutes: int = Query(default=0, ge=-1440, le=1440)
                    ) -> dict:
    """Aggregate per-turn usage across all sessions, bucketed by local date
    and by model. JSONL is the truth for **token counts and model** (CLI
    writes `message.usage` per turn for every vendor — Anthropic, GLM,
    MiniMax, DeepSeek). Sidecar adds **cost in USD** where available
    (only Anthropic + a few others report it; third-party vendors
    typically report 0). All vendors get full token visibility.

    `tz_offset_minutes` lets the browser ask for buckets in its local
    timezone (e.g. Beijing = +480). Server stays UTC internally.

    Returns:
      {
        "window_days": int,
        "today" / "last_7d" / "last_30d" / "all_time": {
            input_tokens, output_tokens, cache_read_tokens,
            cache_creation_tokens, cost, turns
        },
        "by_day":   [{date, ...same fields}, ...]   # densified to `days`
        "by_model": [{model, ...same fields}, ...]  # all time
      }
    """
    import datetime as _dt
    from collections import defaultdict

    tz = _dt.timezone(_dt.timedelta(minutes=tz_offset_minutes))
    now = _dt.datetime.now(tz)
    today_str = now.date().isoformat()

    # Discover the input file set first (cheap: glob + the project-root walk,
    # no file reads yet). We fingerprint these paths before deciding whether
    # the expensive read+parse is even needed (see cache note above).
    sidecar_paths = list(sess.SESS_DIR.glob("*.sidecar.json"))

    # Discover all JSONLs for muselab-tracked sessions. SDK CLI keys
    # the projects dir by the cwd that ran the session, so a single
    # logical archive (one ROOT) can have JSONL spread across multiple
    # `<projects-root>/<encoded-cwd>/` dirs:
    #   - The current MUSELAB_ROOT's encoded form
    #   - Any subdir of MUSELAB_ROOT (the CLI was launched with cwd set
    #     to a child path — happens when ROOT was historically deeper,
    #     or when a subagent ran with a narrower cwd)
    # Earlier versions filtered by `jsonl.stem in known_sids` (sidecar
    # OR sess.list_sessions()), but `sess.list_sessions(directory=ROOT)`
    # only sees the CURRENT ROOT's encoded-cwd dir, so JSONLs written
    # when MUSELAB_ROOT was different (e.g. user moved ROOT up from
    # /home/user/archive → /home/user) became invisible — their models
    # disappeared from by_model even though the JSONL was still on disk.
    # We now scope by encoded-path-prefix instead: a JSONL counts iff its
    # containing dir name equals encoded(ROOT) or starts with
    # `encoded(ROOT) + "-"`. This catches all historical sub-cwds without
    # picking up totally unrelated projects (e.g. /opt/foo, /tmp/bar,
    # an old macOS path /Users/x/... — they don't share the prefix).
    project_roots = _cli_project_roots()
    if not project_roots:
        return _empty_dashboard_response(days, tz_offset_minutes, now)

    # A project directory counts when it belongs to any registered workspace.
    # Prefix matching preserves sessions historically launched from a child
    # cwd while excluding unrelated projects.
    encoded_roots = tuple(
        _cli_encode_cwd(str(root)) for root in workspace_registry.paths())

    jsonl_paths: list[Path] = []
    for projects_root in project_roots:
        try:
            for proj_sub in projects_root.iterdir():
                if not proj_sub.is_dir():
                    continue
                name = proj_sub.name
                if not any(
                    name == encoded or name.startswith(encoded + "-")
                    for encoded in encoded_roots
                ):
                    continue
                for jsonl in proj_sub.glob("*.jsonl"):
                    jsonl_paths.append(jsonl)
        except OSError:
            continue

    # Cheap fingerprint of the input set. Any change that affects the numbers
    # — a new turn (grown / added JSONL), a sidecar cost update, a deleted
    # session — shifts (file count, newest mtime, total size). stat() is
    # microseconds per file; the full read + json.loads of every line is the
    # ~8s cost we skip on a cache hit.
    fp_count = fp_size = 0
    fp_mtime = 0
    for p in (*sidecar_paths, *jsonl_paths):
        try:
            st = p.stat()
        except OSError:
            continue
        fp_count += 1
        fp_size += st.st_size
        if st.st_mtime_ns > fp_mtime:
            fp_mtime = st.st_mtime_ns
    fingerprint = (fp_count, fp_mtime, fp_size)
    cache_key = (days, tz_offset_minutes, today_str)
    with _dashboard_cache_lock:
        cached = _dashboard_cache.get(cache_key)
        if cached is not None and cached[0] == fingerprint:
            return cached[1]

    # ── Cache miss → do the full read + scan. ──
    # 1) Sidecar costs by (sid, uuid) — optional overlay, may be sparse
    # or empty for third-party vendors. Walk it once so the JSONL scan
    # can do a cheap dict lookup per turn.
    cost_by_uuid: dict[str, dict[str, float]] = {}
    for sidecar in sidecar_paths:
        sid = sidecar.name.split(".sidecar.json")[0]
        try:
            data = json.loads(sidecar.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError, ValueError):
            continue
        msgs = data.get("messages") or {}
        per_sess: dict[str, float] = {}
        for uuid_key, ann in msgs.items():
            if not isinstance(ann, dict):
                continue
            cost_val = _parse_cost(ann.get("cost"))
            if cost_val > 0:
                per_sess[uuid_key] = cost_val
        if per_sess:
            cost_by_uuid[sid] = per_sess

    # 2) Walk JSONL — the universal token source. Every vendor writes
    # message.usage on assistant turns in Anthropic-compatible shape
    # (CLI normalizes OpenAI-compatible vendors transparently).
    cutoff_day = (now.date() - _dt.timedelta(days=days - 1)).isoformat()
    cutoff_7d  = (now.date() - _dt.timedelta(days=6)).isoformat()

    all_total   = _empty_bucket()
    today_total = _empty_bucket()
    last_7d     = _empty_bucket()
    last_30d    = _empty_bucket()
    by_day:   dict[str, dict] = defaultdict(_empty_bucket)
    by_model: dict[str, dict] = defaultdict(_empty_bucket)
    by_vendor: dict[str, dict] = defaultdict(_empty_bucket)

    for jsonl in jsonl_paths:
        sid = jsonl.stem
        sid_costs = cost_by_uuid.get(sid, {})
        try:
            with jsonl.open("r", encoding="utf-8") as f:
                for line in f:
                    # Cheap reject: only assistant turns carry usage.
                    if '"type":"assistant"' not in line and '"type": "assistant"' not in line:
                        continue
                    try:
                        entry = json.loads(line)
                    except (json.JSONDecodeError, ValueError):
                        continue
                    msg = entry.get("message") or {}
                    usage = msg.get("usage") or {}
                    if not isinstance(usage, dict):
                        continue
                    in_t  = int(usage.get("input_tokens", 0) or 0)
                    out_t = int(usage.get("output_tokens", 0) or 0)
                    cr_t  = int(usage.get("cache_read_input_tokens", 0)
                                  or usage.get("cache_read_tokens", 0) or 0)
                    cc_t  = int(usage.get("cache_creation_input_tokens", 0)
                                  or usage.get("cache_creation_tokens", 0) or 0)
                    # Skip empty-usage entries (e.g. CLI-internal markers).
                    if in_t == 0 and out_t == 0 and cr_t == 0 and cc_t == 0:
                        continue
                    # A single "turn" = one user prompt + its assistant
                    # response chain. Inside that chain there can be many
                    # intermediate assistant lines for tool_use loops —
                    # those have stop_reason="tool_use". Only count the
                    # final completion (stop_reason="end_turn", "max_tokens",
                    # or sometimes None for legacy/streamed lines).
                    stop_reason = msg.get("stop_reason")
                    is_final = stop_reason in (None, "end_turn",
                                                  "max_tokens", "stop_sequence")
                    ts = entry.get("timestamp") or ""
                    if not ts:
                        continue
                    try:
                        dt_utc = _dt.datetime.fromisoformat(
                            ts.replace("Z", "+00:00"))
                    except ValueError:
                        continue
                    day_str = dt_utc.astimezone(tz).date().isoformat()
                    model_name = msg.get("model") or "unknown"
                    uuid_key = entry.get("uuid") or ""
                    cost_val = sid_costs.get(uuid_key, 0.0)

                    turn = {
                        "input_tokens": in_t,
                        "output_tokens": out_t,
                        "cache_read_tokens": cr_t,
                        "cache_creation_tokens": cc_t,
                        "cost": cost_val,
                        # Every assistant line contributes tokens (each
                        # tool-use loop iteration costs real compute), but
                        # only the final completion counts as a "turn"
                        # from the user's perspective.
                        "turns": 1 if is_final else 0,
                    }
                    _add_bucket(all_total, turn)
                    _add_bucket(by_model[model_name], turn)
                    # Roll up to vendor too — same data, vendor granularity.
                    # "Claude" for Anthropic, vendor display for third-parties
                    # (DeepSeek / GLM / MiniMax), "Unknown" for stray model
                    # ids we can't map (rare; CLI / vendor wrapper artifacts).
                    vendor = _vendor_label_for(model_name)
                    _add_bucket(by_vendor[vendor], turn)
                    if day_str >= cutoff_day:
                        _add_bucket(by_day[day_str], turn)
                        _add_bucket(last_30d, turn)
                    if day_str >= cutoff_7d:
                        _add_bucket(last_7d, turn)
                    if day_str == today_str:
                        _add_bucket(today_total, turn)
        except OSError:
            continue

    # Densify by_day so quiet days still get a zero bar.
    dense_days: list[dict] = []
    for i in range(days):
        d = (now.date() - _dt.timedelta(days=days - 1 - i)).isoformat()
        bucket = by_day.get(d, _empty_bucket())
        dense_days.append({"date": d, **_round_bucket(bucket)})

    by_model_list = sorted(
        [
            {
                "model": k,
                # Friendly label (e.g. "Sonnet 4.6" instead of
                # "claude-sonnet-4-6") so the FE can show readable names
                # without re-implementing the mapping.
                "label": endpoints.label_for(k),
                "vendor": _vendor_label_for(k),
                # FE uses this to decorate rows whose cost is "untracked,
                # not free" with a footnote instead of pretending the
                # vendor was free.
                "cost_reported": _cost_reported_for(k),
                **_round_bucket(v),
            }
            for k, v in by_model.items()
        ],
        key=lambda x: (x["input_tokens"] + x["output_tokens"]
                        + x["cache_read_tokens"] + x["cache_creation_tokens"]),
        reverse=True)

    by_vendor_list = sorted(
        [
            {
                "vendor": k,
                # Same "we report USD" flag at vendor granularity. A vendor
                # is cost-reported when at least one of its model ids is —
                # currently equivalent to "vendor == 'Claude'".
                "cost_reported": k == "Claude",
                **_round_bucket(v),
            }
            for k, v in by_vendor.items()
        ],
        key=lambda x: (x["input_tokens"] + x["output_tokens"]
                        + x["cache_read_tokens"] + x["cache_creation_tokens"]),
        reverse=True)

    response = {
        "window_days": days,
        "tz_offset_minutes": tz_offset_minutes,
        "today":    _round_bucket(today_total),
        "last_7d":  _round_bucket(last_7d),
        "last_30d": _round_bucket(last_30d),
        "all_time": _round_bucket(all_total),
        "by_day":   dense_days,
        "by_model": by_model_list,
        "by_vendor": by_vendor_list,
    }
    with _dashboard_cache_lock:
        # Drop stale-day entries so the cache can't grow unbounded across
        # midnight rollovers (old keys differ only by today_str).
        for k in [k for k in _dashboard_cache if k[2] != today_str]:
            _dashboard_cache.pop(k, None)
        _dashboard_cache[cache_key] = (fingerprint, response)
    return response


def _round_bucket(b: dict) -> dict:
    return {**b, "cost": round(b["cost"], 4)}


def _empty_dashboard_response(days: int, tz_offset_minutes: int, now) -> dict:
    """Helper for the no-JSONL case — returns the same shape with all
    zeros + a densified by_day list so the frontend's chart doesn't
    crash on missing keys."""
    import datetime as _dt
    dense = [{"date": (now.date() - _dt.timedelta(days=days - 1 - i)).isoformat(),
                **_round_bucket(_empty_bucket())} for i in range(days)]
    return {
        "window_days": days,
        "tz_offset_minutes": tz_offset_minutes,
        "today":    _round_bucket(_empty_bucket()),
        "last_7d":  _round_bucket(_empty_bucket()),
        "last_30d": _round_bucket(_empty_bucket()),
        "all_time": _round_bucket(_empty_bucket()),
        "by_day":   dense,
        "by_model": [],
        "by_vendor": [],
    }


def _dedupe_error_parts(parts: list[Any]) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    for part in parts:
        text = str(part or "").strip()
        if not text or text in seen:
            continue
        seen.add(text)
        out.append(text)
    return out


def _sdk_assistant_error(msg: Any) -> dict | None:
    error = getattr(msg, "error", None)
    if not error:
        return None
    text_parts = [
        getattr(block, "text", "")
        for block in (getattr(msg, "content", None) or [])
        if isinstance(block, TextBlock)
    ]
    parts = _dedupe_error_parts([*text_parts, error])
    return {
        "message": "; ".join(parts) or "assistant API error",
        "source": "assistant",
        "api_error_status": None,
    }


def _sdk_system_error(msg: Any) -> dict | None:
    """Recognize terminal failures carried only by a SystemMessage.

    ``preventContinuation`` means the CLI rejected the user prompt before it
    entered the canonical transcript.  A later nominally successful
    ResultMessage is only the hook operation finishing; it must not turn that
    rejected prompt into a completed chat turn.

    Slash commands have a second legacy shape: the CLI can finish `/compact`
    with a successful ResultMessage while putting the actual Gateway/API failure
    only in ``system/local_command``.  Normal local-command prose remains
    non-terminal.
    """
    if not isinstance(msg, SystemMessage):
        return None
    data = msg.data if isinstance(msg.data, dict) else {}
    prevented = data.get("preventContinuation") is True
    if not prevented and msg.subtype != "local_command":
        return None
    text = str(data.get("content") or data.get("error") or "").strip()
    if prevented:
        return {
            "message": text or "User prompt was rejected before it was committed.",
            "source": "system_prevent_continuation",
            "api_error_status": None,
        }
    if not text:
        return None
    classified = _classify_stream_error(text)
    if classified.get("kind") == "unknown" and "api error" not in text.lower():
        return None
    return {
        "message": text,
        "source": "system_local_command",
        "api_error_status": None,
    }


def _sdk_result_error(msg: Any) -> dict | None:
    if isinstance(msg, ResultError):
        structured = sdk_lifecycle.result_error_info(msg)
        if structured is None:
            return None
        parts = _dedupe_error_parts([
            *structured["errors"],
            structured["result"],
            structured["subtype"],
        ])
        status = structured["api_error_status"]
        if not parts and status:
            parts = [f"API error {status}"]
        return {
            "message": "; ".join(parts) or "SDK command failed",
            "source": "result_exception",
            "api_error_status": status,
            "result_error": structured,
        }
    if not bool(getattr(msg, "is_error", False)):
        return None
    errors = getattr(msg, "errors", None) or []
    if not isinstance(errors, (list, tuple)):
        errors = [errors]
    parts = _dedupe_error_parts([
        *errors,
        getattr(msg, "result", None),
        getattr(msg, "subtype", None),
    ])
    status = getattr(msg, "api_error_status", None)
    if not parts and status:
        parts = [f"API error {status}"]
    return {
        "message": "; ".join(parts) or "SDK command failed",
        "source": "result",
        "api_error_status": status,
    }


def _merge_sdk_errors(errors: list[dict]) -> dict | None:
    if not errors:
        return None
    parts = _dedupe_error_parts([e.get("message") for e in errors])
    status = next(
        (e.get("api_error_status") for e in reversed(errors)
         if e.get("api_error_status")),
        None,
    )
    return {
        "message": "; ".join(parts) or "SDK command failed",
        "source": "+".join(dict.fromkeys(e.get("source", "sdk") for e in errors)),
        "api_error_status": status,
    }


class _SDKCommandError(RuntimeError):
    def __init__(self, info: dict):
        self.info = info
        super().__init__(info.get("message") or "SDK command failed")


class _ContextRecovered(RuntimeError):
    """Stop the old turn after creating a safe, resumable recovery fork."""

    def __init__(self, recovery: dict[str, Any]):
        self.recovered_session = dict(recovery.get("session") or {})
        self.recovery_stats = dict(recovery.get("stats") or {})
        super().__init__(
            "context window could not be compacted; a recovery session was created"
        )


def _context_recovery_inputs(
    sid: str,
    model: str,
    *,
    pre_tokens: int = 0,
    context_limit: int = 0,
) -> tuple[int, int]:
    """Fill recovery metadata from already-known, privacy-safe context state.

    This helper deliberately performs no network or transcript I/O: it is used
    after the live SDK context probe has already failed.  Cached values only
    size/describe the recovery summary; they are never evidence that recovery
    is required.  That decision must come from an explicit context-window
    error, a transcript marker, or a verified no-shrink result.
    """
    cached = _session_usage.get(sid) or {}
    used = _positive_int(pre_tokens) or _positive_int(
        cached.get("context_used")
    )
    if not used:
        used = sum(
            _positive_int(cached.get(key))
            for key in (
                "input_tokens",
                "cache_read_tokens",
                "cache_creation_tokens",
            )
        )

    limit = _positive_int(context_limit) or _positive_int(
        cached.get("context_limit")
    )
    if not limit:
        details = _context_limit_details(
            model,
            capability=_cached_gateway_context_capability(model),
        )
        limit = _positive_int(details.get("context_limit"))
    return used, limit


def _commit_fork_lifecycle(
    source_sid: str,
    source_meta: dict[str, Any],
    *,
    fork_child: Callable[[], Any] | None,
    register_kwargs: dict[str, Any],
    successor: bool,
    copy_runtime_overlays: bool = False,
) -> dict[str, Any]:
    return chat_successor.commit_fork_lifecycle(
        source_sid,
        source_meta,
        fork_child=fork_child,
        register_kwargs=register_kwargs,
        successor=successor,
        copy_runtime_overlays=copy_runtime_overlays,
    )

def _create_context_recovery_session(
    source_sid: str,
    model: str,
    *,
    pre_tokens: int = 0,
    context_limit: int = 0,
) -> dict[str, Any]:
    """Fork an over-limit transcript onto a bounded synthetic compact root.

    This runs in a worker thread.  The source JSONL is never modified; the
    recovery module first makes an SDK-native fork and appends the compact pair
    only to that new file.  Register the new UUID in MuseLab only after the
    fork is complete, so the browser can never discover a half-built session.
    """
    try:
        source_meta = sess.get_session_meta(source_sid)
    except Exception as exc:
        sys.stderr.write(
            f"[chat] runtime postlude source read failed "
            f"sid={source_sid[:8]} exc={type(exc).__name__}\n")
        return {"annotations": 0, "renamed": 0}
    if source_meta is None:
        raise context_recovery.ContextRecoveryError("source session is unavailable")
    source_path = _find_session_jsonl(source_sid)
    if source_path is None and not source_meta.get("runtime_successor"):
        raise context_recovery.ContextRecoveryError("source transcript is unavailable")

    source_model = (source_meta.get("model") or model or MODEL).strip()
    source_name = (source_meta.get("name") or "会话").strip()
    suffix = "恢复" if is_chinese_locale() else "Recovered"
    recovery_name = f"{source_name} · {suffix}"
    workspace = sess.session_workspace(source_sid)

    def _fork_recovery():
        if source_path is None:
            raise context_recovery.ContextRecoveryError(
                "source transcript is unavailable")
        with _session_config_dir(source_model, sid=source_sid):
            return context_recovery.create_recovery_fork(
                source_sid,
                source_path,
                workspace,
                title=recovery_name,
                pre_tokens=pre_tokens,
                model_config_context=context_limit or None,
            )

    lifecycle = _commit_fork_lifecycle(
        source_sid,
        source_meta,
        fork_child=_fork_recovery,
        register_kwargs={
            "name": recovery_name,
            "model": source_model,
            "permission": source_meta.get("permission") or "",
            "plan_return_permission": source_meta.get("plan_return_permission"),
            "auto_named": False,
            # The active SDK chain contains the compact summary. Full-history
            # mode can still read the copied records behind its boundary.
            "message_count": 1,
            "turn_count": 0,
            "effort": _normalize_effort(source_meta.get("effort")),
            "service_tier": source_meta.get("service_tier") or "",
            "thinking": source_meta.get("thinking") is not False,
            "forked_from": source_sid,
            "forked_from_name": source_name,
            "activity_hidden": bool(source_meta.get("activity_hidden", False)),
            "runtime_profile": source_meta.get("runtime_profile") or "",
            "runtime_predecessor": source_sid,
            "cwd": source_meta.get("cwd") or str(workspace),
        },
        successor=True,
    )
    child_sid = str(lifecycle["child_sid"])
    result = lifecycle.get("forked")
    if result is not None:
        _JSONL_PATH_CACHE[child_sid] = result.path
    if context_limit:
        try:
            sess.set_session_ctx_window(child_sid, context_limit)
        except Exception:
            pass
    estimated_post_tokens = (
        result.stats.estimated_post_tokens
        if result is not None else int(
            (_session_usage.get(child_sid) or {}).get("context_used") or 0)
    )
    _session_usage[child_sid] = {
        "input_tokens": 0,
        "output_tokens": 0,
        "cache_read_tokens": 0,
        "cache_creation_tokens": 0,
        "total_cost_usd": 0.0,
        "last_turn_at": 0.0,
        "context_used": estimated_post_tokens,
        "context_used_pct": (
            round(estimated_post_tokens / context_limit * 100, 1)
            if context_limit else 0.0
        ),
        "context_limit": context_limit,
        "context_used_source": "recovery_summary_estimate",
        "context_used_is_estimate": True,
        "context_is_estimate": True,
    }
    public_meta = {**lifecycle["child_meta"], "session_id": child_sid}
    stats = (
        {
            "included_messages": result.stats.included_messages,
            "omitted_messages": result.stats.omitted_messages,
            "truncated_messages": result.stats.truncated_messages,
            "estimated_post_tokens": estimated_post_tokens,
        }
        if result is not None else {
            "included_messages": 0,
            "omitted_messages": 0,
            "truncated_messages": 0,
            "estimated_post_tokens": estimated_post_tokens,
        }
    )
    return {"session": public_meta, "stats": stats}


async def _recover_context_session(
    source_sid: str,
    model: str,
    *,
    pre_tokens: int = 0,
    context_limit: int = 0,
) -> dict[str, Any]:
    """Flush the failed runtime and build a cold recovery session."""
    if (_sessions_with_inflight_tasks.get(source_sid)
            or _session_has_live_watcher(source_sid)):
        raise context_recovery.ContextRecoveryError(
            "cannot recover while a background task owns the source runtime"
        )
    await disconnect_client(source_sid)
    recovery_task = asyncio.create_task(
        asyncio.to_thread(
            _create_context_recovery_session,
            source_sid,
            model,
            pre_tokens=pre_tokens,
            context_limit=context_limit,
        )
    )
    try:
        return await asyncio.shield(recovery_task)
    except asyncio.CancelledError:
        # Fork + index registration is a mutating disk transaction and the
        # worker cannot be stopped. Join to a known committed/failed outcome so
        # timeout/cancellation cannot create an unseen, duplicate recovery fork.
        while not recovery_task.done():
            try:
                await asyncio.shield(recovery_task)
            except asyncio.CancelledError:
                continue
        return recovery_task.result()


async def _run_sdk_command_checked(client: ClaudeSDKClient, command: str) -> ResultMessage:
    """Run a CLI slash command and require an explicitly successful Result.

    Reads through the session's pump (`_SessionStream`) rather than opening
    `client.receive_response()`. The SDK gives a client exactly ONE message
    stream and the pump has owned it since client creation, so a second iterator
    here lost every race — the command's ResultMessage went to the pump's
    `_orphans` park and this function waited on a stream nobody was feeding.
    The turn loop was migrated to `attach_turn()` when the pump landed; this
    call site was missed. Symptom (2026-07-26): two auto-compacts reported
    failure after 600s (TimeoutError) and 9m19s ("ended without a
    ResultMessage") while the transcript shows both compactions had finished in
    ~150s. Both were FALSE NEGATIVES — `query()` is a pure transport write, so
    the command itself always ran.

    Assistant/API errors are in-band SDK messages, not Python exceptions. Drain
    through the terminal Result before raising so a failed command cannot leave
    a stale Result in the pooled client's receive queue for the next turn.
    """
    errors: list[dict] = []
    result: ResultMessage | None = None

    def _note(msg) -> bool:
        """Record one message; True once the terminal Result has been seen."""
        nonlocal result
        if isinstance(msg, AssistantMessage):
            if info := _sdk_assistant_error(msg):
                errors.append(info)
        elif isinstance(msg, SystemMessage):
            if info := _sdk_system_error(msg):
                errors.append(info)
        if isinstance(msg, ResultMessage):
            result = msg
            if info := _sdk_result_error(msg):
                errors.append(info)
            return True
        return False

    stream = _stream_for(client)
    if stream is not None:
        # Attach BEFORE query(): anything the pump routes between the write and
        # the attach would be parked as an orphan, and attach_turn deliberately
        # does not adopt orphans, so a late attach could miss its own Result.
        q = stream.attach_turn()
        try:
            await client.query(command)
            while True:
                msg = await q.get()
                if msg is _STREAM_EOF:
                    break
                if _note(msg):
                    break
        finally:
            stream.detach_turn(q)
            stream.park_unconsumed(q)
    else:
        # No pump — a client built outside the pool (unit tests). The SDK's own
        # bounded reader is then the only reader, so it is safe to use.
        await client.query(command)
        async for msg in client.receive_response():
            if _note(msg):
                break
    if result is None:
        errors.append({
            "message": f"{command} ended without a ResultMessage",
            "source": "transport",
            "api_error_status": None,
        })
    if merged := _merge_sdk_errors(errors):
        raise _SDKCommandError(merged)
    return result


async def _run_sdk_reset_checked(
    client: ClaudeSDKClient,
    source_sid: str,
) -> tuple[ConversationResetMessage, ResultMessage]:
    """Run SDK-native ``/clear`` and require its two-part handoff contract.

    ``ConversationResetMessage.new_conversation_id`` is an opaque conversation
    marker, not the SDK session id MuseLab can resume.  The immediately
    following ``ResultMessage.session_id`` is the only authoritative new id.
    Read through the sole pooled stream just like ``_run_sdk_command_checked``
    so a control command cannot race the session pump.
    """
    reset: ConversationResetMessage | None = None
    result: ResultMessage | None = None
    errors: list[dict] = []

    def _note(msg: Any) -> bool:
        nonlocal reset, result
        if isinstance(msg, ConversationResetMessage):
            reset = msg
            return False
        if isinstance(msg, AssistantMessage):
            if info := _sdk_assistant_error(msg):
                errors.append(info)
        elif isinstance(msg, SystemMessage):
            if info := _sdk_system_error(msg):
                errors.append(info)
        if isinstance(msg, ResultMessage):
            result = msg
            if info := _sdk_result_error(msg):
                errors.append(info)
            return True
        return False

    stream = _stream_for(client)
    if stream is not None:
        queue = stream.attach_turn()
        try:
            await client.query("/clear")
            while True:
                message = await queue.get()
                if message is _STREAM_EOF:
                    break
                if _note(message):
                    break
        finally:
            stream.detach_turn(queue)
            stream.park_unconsumed(queue)
    else:
        await client.query("/clear")
        async for message in client.receive_response():
            if _note(message):
                break

    if reset is None:
        errors.append({
            "message": "/clear ended without a ConversationResetMessage",
            "source": "reset_contract",
            "api_error_status": None,
        })
    if result is None:
        errors.append({
            "message": "/clear ended without a ResultMessage",
            "source": "reset_contract",
            "api_error_status": None,
        })
    new_sid = str(getattr(result, "session_id", "") or "")
    try:
        parsed = uuid.UUID(new_sid)
    except (ValueError, AttributeError, TypeError):
        parsed = None
    if parsed is None or str(parsed) != new_sid.lower() or new_sid == source_sid:
        errors.append({
            "message": "/clear returned an invalid replacement session id",
            "source": "reset_contract",
            "api_error_status": None,
        })
    if merged := _merge_sdk_errors(errors):
        raise _SDKCommandError(merged)
    assert reset is not None and result is not None
    return reset, result


@router.get("/context-breakdown/{session_id}", dependencies=[Depends(require_token)])
async def context_breakdown(session_id: str, model: str = "") -> dict:
    """Detailed context breakdown via SDK — answers "where did my 100K go?".
    Calls ClaudeSDKClient.get_context_usage() which returns the same data
    the CLI's /context command shows: tokens per category (memory files,
    MCP tools, agents, system tools, system prompt sections), with
    per-file and per-tool breakdowns.

    Returns 404 if the session doesn't have a live SDK client yet — that
    happens for newly-created sessions that haven't run a turn."""
    s = await obs.to_thread_io(
        "chat.session_read", session_id, sess.get_session, session_id)
    if s is None:
        raise HTTPException(404, "session not found")
    m = (model or s.get("model") or MODEL).strip()
    # The context-breakdown call is read-only and effort-independent — find
    # ANY live client for this (sid, model) pair regardless of effort key.
    # Snapshot the matching runtime under the same lock used by client
    # creation/eviction.  This endpoint is commonly opened while a turn is
    # finishing or a model/permission change is rebuilding the runtime; an
    # unlocked iteration over `_clients` could otherwise race a resize and
    # fail with ``RuntimeError: dictionary changed size during iteration``.
    # Release the lock before the SDK control request so a slow CLI cannot
    # block unrelated client-pool work.
    async with _lock:
        matched = [
            (k, client)
            for k, client in _clients.items()
            if k[0] == session_id and k[1] == m
        ]
    if not matched:
        # No live client → can't ask CLI for breakdown. Surface this rather
        # than returning fake data; frontend can fall back to /usage.
        raise HTTPException(409, "no live client for this session — send a message first")
    _key, client = matched[0]
    try:
        breakdown = await client.get_context_usage()
        payload = dict(breakdown)
        if _is_codex_gateway_model(m):
            # Claude CLI can describe the live prompt categories accurately,
            # but an older already-running client may still report its legacy
            # 200K Claude denominator. Re-anchor the response to the active
            # CLIProxyAPI catalog so the popup and bottom ring cannot disagree.
            sdk_max = _positive_int(payload.get("maxTokens"))
            sdk_raw = _positive_int(payload.get("rawMaxTokens"))
            capability = await _detect_gateway_context_capability(m)
            details = _context_limit_details(
                m, sdk_max=sdk_max, sdk_raw=sdk_raw,
                capability=capability)
            limit = _positive_int(details.get("context_limit"))
            total = _positive_int(payload.get("totalTokens"))
            payload["sdkMaxTokens"] = sdk_max
            payload["sdkRawMaxTokens"] = sdk_raw
            payload["maxTokens"] = limit
            payload["rawMaxTokens"] = _positive_int(
                details.get("context_raw_limit"))
            payload["maxContextTokens"] = _positive_int(
                details.get("context_max_limit"))
            payload["percentage"] = (
                round(total / limit * 100, 1) if limit else 0.0)
            payload["autoCompactThreshold"] = _compact_threshold(
                m,
                limit,
                _positive_int(payload.get("autoCompactThreshold")),
                sdk_max=sdk_max,
                capability=capability,
            )
            _apply_context_limit_details(payload, details)
            _mark_context_used(payload, "sdk_context", estimate=True)
        return payload
    except Exception as e:
        # CLI exceptions may contain credentials paths, vendor URLs, prompt
        # fragments or protocol bodies.  The API already returns a generic
        # error; the service log keeps only the failure class.
        sys.stderr.write(
            f"[chat] get_context_usage failed sid={session_id[:8]} "
            f"exc={type(e).__name__}\n")
        sys.stderr.flush()
        raise HTTPException(500, "context-usage probe failed") from None


@router.post("/sessions/{sid}/native-clear", dependencies=[Depends(require_token)])
async def native_clear_session_api(sid: str) -> dict:
    """Reset context through the SDK while preserving the source conversation.

    Claude turns ``/clear`` into a new SDK session.  MuseLab registers that
    authoritative Result session as a normal resumable conversation and leaves
    the source transcript untouched, so reset is no longer a destructive
    delete disguised as a familiar CLI command.
    """
    source = await obs.to_thread_io(
        "chat.session_read", sid, sess.get_session_meta, sid)
    if source is None:
        raise HTTPException(404, "session not found")
    if sess.session_is_deleting(sid):
        raise HTTPException(409, "session is being deleted")

    def _assert_idle() -> None:
        active = _active_turns.get(sid)
        if active is not None and not active.done:
            raise HTTPException(409, "cannot clear while a turn is active")
        if (_sessions_with_inflight_tasks.get(sid)
                or _session_has_live_watcher(sid)):
            raise HTTPException(
                409, "cannot clear while a background task owns the session")
        overlays = sess.get_authoritative_runtime_task_overlays(sid)
        if any(str(row.get("state") or "") == "running"
               for row in overlays.values() if isinstance(row, dict)):
            raise HTTPException(
                409, "cannot clear while a background task owns the session")

    _assert_idle()
    queue_snapshot = await obs.to_thread_io(
        "chat.queue_read", sid, sess.get_queue, sid)
    if queue_snapshot.get("items") or queue_snapshot.get("inflight"):
        raise HTTPException(409, "cannot clear while queued messages remain")

    model = str(source.get("model") or MODEL).strip()
    effort = _normalize_effort(source.get("effort"))
    service_tier = str(source.get("service_tier") or "").strip()
    permission = str(source.get("permission") or "default")
    client_kwargs: dict[str, Any] = {
        "effort": effort,
        "service_tier": service_tier,
    }
    if permission == "plan":
        client_kwargs["plan_return_permission"] = (
            source.get("plan_return_permission") or "default")

    timeout_s = env_int("MUSELAB_CLEAR_TIMEOUT_S", 120, min_value=1)
    try:
        async with asyncio.timeout(timeout_s):
            async with _session_runtime_lock_for(sid):
                _assert_idle()
                latest_queue = await obs.to_thread_io(
                    "chat.queue_read", sid, sess.get_queue, sid)
                if latest_queue.get("items") or latest_queue.get("inflight"):
                    raise HTTPException(
                        409, "cannot clear while queued messages remain")
                client = await get_client(
                    sid, model, permission, **client_kwargs)
                _reset, result = await _run_sdk_reset_checked(client, sid)
                new_sid = str(result.session_id)
                new_meta = await obs.to_thread_io(
                    "chat.session_register",
                    new_sid,
                    sess.register_session,
                    new_sid,
                    name=str(source.get("name") or ""),
                    model=model,
                    permission=permission,
                    plan_return_permission=source.get(
                        "plan_return_permission"),
                    auto_named=bool(source.get("auto_named", False)),
                    effort=effort,
                    service_tier=service_tier,
                    thinking=source.get("thinking") is not False,
                    activity_hidden=bool(
                        source.get("activity_hidden", False)),
                    runtime_profile=str(source.get("runtime_profile") or ""),
                    cwd=source.get("cwd") or str(ROOT),
                )
                # The pooled process is indexed by the old id but now owns the
                # new Claude conversation.  It must never be reused under that
                # stale key, even if the browser keeps the old tab open.
                _pending_runtime_rebuilds.add(sid)
                await disconnect_client(sid)
    except HTTPException:
        raise
    except asyncio.TimeoutError:
        _pending_runtime_rebuilds.add(sid)
        await disconnect_client(sid)
        raise HTTPException(504, "native /clear timed out") from None
    except _SDKCommandError as exc:
        _pending_runtime_rebuilds.add(sid)
        await disconnect_client(sid)
        status = int(exc.info.get("api_error_status") or 502)
        if status < 400 or status > 599:
            status = 502
        raise HTTPException(status, str(exc)) from None
    except Exception as exc:
        _pending_runtime_rebuilds.add(sid)
        with suppress(Exception):
            await disconnect_client(sid)
        sys.stderr.write(
            f"[chat] native /clear failed sid={sid[:8]} "
            f"exc={type(exc).__name__}\n")
        sys.stderr.flush()
        raise HTTPException(500, "native /clear failed") from None
    return {**new_meta, "session_id": new_sid, "reset_from": sid}


@router.post("/sessions/{sid}/native-compact", dependencies=[Depends(require_token)])
async def native_compact_session_api(sid: str) -> dict:
    timeout_s = env_int("MUSELAB_COMPACT_TIMEOUT_S", 300, min_value=1)
    try:
        # Fast rejection avoids spending the whole compact deadline waiting on
        # a turn that is already known to own this session.  The same check is
        # repeated under the runtime lock to close the race with a new turn.
        bc = _active_turns.get(sid)
        if bc is not None and not bc.done:
            raise HTTPException(409, "cannot compact while a turn is active")
        # One deadline covers lock acquisition, client startup, /compact and
        # the authoritative post-command token verification. Previously only
        # the slash command itself was bounded, so a wedged control probe could
        # leave the HTTP request and compacting UI open forever.
        async with asyncio.timeout(timeout_s):
            async with _session_runtime_lock_for(sid):
                bc = _active_turns.get(sid)
                if bc is not None and not bc.done:
                    raise HTTPException(409, "cannot compact while a turn is active")
                try:
                    return await _native_compact_session_locked(sid)
                except asyncio.CancelledError:
                    # The outer deadline fired after compact acquired the
                    # runtime.  Tear the unknown client down while this lock is
                    # still held; otherwise a waiting turn could acquire the
                    # lock and reuse it in the gap before endpoint cleanup.
                    _pending_runtime_rebuilds.add(sid)
                    await disconnect_client(sid)
                    raise
                except HTTPException as exc:
                    if exc.status_code == 504:
                        # The inner command deadline has the same unknown-
                        # outcome contract. Cleanup must precede lock release.
                        _pending_runtime_rebuilds.add(sid)
                        await disconnect_client(sid)
                    raise
    except asyncio.TimeoutError:
        # If compact had acquired the lock, its CancelledError handler above
        # already completed teardown before releasing it.  A timeout that only
        # waited for the lock must NOT kill the legitimate lock holder.
        sys.stderr.write(f"[chat] native /compact total timeout sid={sid[:8]}\n")
        sys.stderr.flush()
        raise HTTPException(
            504, "native /compact timed out — CLI may be hung") from None


def _refresh_compacted_message_counts(sid: str, model: str) -> None:
    """Refresh sidebar counters after the compact result is already visible."""
    new_msgs = _get_session_msgs(sid, model)
    n_turns = sum(1 for sm in new_msgs if _is_real_user_prompt(sm))
    sess.bump_session(
        sid,
        message_count=len(new_msgs),
        turn_count=n_turns,
    )


def _schedule_post_compact_refresh(
    sid: str,
    model: str,
    usage: dict[str, Any],
) -> None:
    async def _run() -> None:
        try:
            if endpoints.is_third_party(model):
                capability = await _detect_gateway_context_capability(model)
                real_max = _positive_int(usage.get("maxTokens"))
                real_total = _positive_int(usage.get("totalTokens"))
                sess_u = _session_usage.setdefault(sid, {})
                details = _context_limit_details(
                    model,
                    sdk_max=real_max,
                    sdk_raw=_positive_int(usage.get("rawMaxTokens")),
                    stored=_positive_int(sess_u.get("context_limit")),
                    capability=capability,
                )
                _apply_context_limit_details(sess_u, details)
                threshold = _compact_threshold(
                    model,
                    _positive_int(details.get("context_limit")),
                    _positive_int(usage.get("autoCompactThreshold")),
                    sdk_max=real_max,
                    capability=capability,
                )
                if threshold:
                    sess_u["auto_compact_threshold"] = threshold
                if real_total:
                    _mark_context_used(
                        sess_u, "sdk_context", estimate=True)
                    limit = _positive_int(sess_u.get("context_limit"))
                    if limit:
                        sess_u["context_used_pct"] = round(
                            real_total / limit * 100, 1)
            await asyncio.to_thread(
                _refresh_compacted_message_counts, sid, model)
        except Exception as e:
            sys.stderr.write(
                f"[chat] post-compact count refresh skipped sid={sid[:8]} "
                f"exc={type(e).__name__}\n")

    task = asyncio.create_task(_run())
    _maintenance_tasks.add(task)
    task.add_done_callback(_maintenance_tasks.discard)


async def _native_compact_session_locked(sid: str) -> dict:
    """Compact a session using the CLI's native /compact slash command via SDK.
    Lossless — CLI writes compact_boundary + isCompactSummary into the session
    JSONL. Subsequent get_session_messages() returns the summary in place of
    pre-compaction history, so the UI automatically reflects the compacted
    state on next loadSession — no muselab-side marker needed.

    Session ID stays the same; tool_use history is preserved in the summary."""
    meta = await obs.to_thread_io(
        "chat.session_read", sid, sess.get_session_meta, sid)
    if meta is None:
        raise HTTPException(404, "session not found")
    model = (meta.get("model") or "").strip() or MODEL
    effort = _normalize_effort(meta.get("effort"))
    service_tier = (meta.get("service_tier") or "").strip()
    # /compact is a CLI control command, not an agent tool call. Preserve the
    # warm runtime contract, or use the durable session contract on a cold
    # start. In particular, Plan's return capability is part of that contract.
    runtime_key = (sid, model, effort, service_tier)
    prior_perm = _client_permission.get(runtime_key)
    permission = prior_perm or (meta.get("permission") or "default")
    client_kwargs: dict[str, Any] = {
        "effort": effort, "service_tier": service_tier,
    }
    if permission == "plan":
        client_kwargs["plan_return_permission"] = (
            _client_plan_return.get(runtime_key)
            or meta.get("plan_return_permission")
            or "default"
        )
    client = await get_client(sid, model, permission, **client_kwargs)
    before_total = 0
    context_limit = _positive_int(
        (_session_usage.get(sid) or {}).get("context_limit"))
    post_compact_usage: dict | None = None
    tail_path, tail_offset = _compact_tail_cursor(sid)
    tail_outcome: dict[str, bool] = {
        "boundary": False,
        "summary": False,
        "context_error": False,
    }
    try:
        try:
            before_usage = dict(await client.get_context_usage())
            before_total = _positive_int(before_usage.get("totalTokens"))
            context_limit = context_limit or _positive_int(
                before_usage.get("maxTokens"))
        except Exception:
            pass
        # Bound the command too for direct internal callers. The HTTP endpoint
        # has an outer deadline that also covers lock/client/verification.
        # 300s is ~2x the observed worst case (~150s at 190K tokens) now that
        # the command reads through the pump instead of racing it.
        async with asyncio.timeout(env_int("MUSELAB_COMPACT_TIMEOUT_S", 300, min_value=1)):
            await _run_sdk_command_checked(client, "/compact")
        tail_outcome = await asyncio.to_thread(
            _compact_tail_outcome, tail_path, tail_offset)
        post_compact_usage = dict(await client.get_context_usage())
        after_total = _positive_int(post_compact_usage.get("totalTokens"))
        compact_pair_written = bool(
            tail_outcome.get("boundary") and tail_outcome.get("summary"))
        if (before_total and after_total and after_total >= before_total
                and not compact_pair_written):
            raise _SDKCommandError({
                "message": (
                    "native /compact reported success but context usage did not decrease "
                    f"({before_total} -> {after_total})"
                ),
                "source": "verification",
                "api_error_status": None,
            })
    except asyncio.TimeoutError:
        # Direct callers still need a hard cache barrier.  The public endpoint
        # disconnects immediately outside its expired deadline; this marker is
        # the fallback that forces any other caller to rebuild before reuse.
        _pending_runtime_rebuilds.add(sid)
        sys.stderr.write(f"[chat] native /compact timed out for sid={sid[:8]}\n")
        sys.stderr.flush()
        raise HTTPException(504, "native /compact timed out — CLI may be hung") from None
    except _SDKCommandError as e:
        if not any(tail_outcome.values()):
            tail_outcome = await asyncio.to_thread(
                _compact_tail_outcome, tail_path, tail_offset)
        if tail_outcome.get("boundary") and tail_outcome.get("summary"):
            # The durable transcript commit wins over an in-band acknowledgement
            # failure. Start the next turn from a fresh process that reads the
            # new compact root.
            await disconnect_client(sid)
            _schedule_post_compact_refresh(sid, model, {})
            return {"ok": True, "recovered": False}
        classified = _classify_stream_error(str(e))
        verified_no_shrink = e.info.get("source") == "verification"
        deterministic_context_failure = bool(
            tail_outcome.get("context_error")
            or classified["kind"] == "context_window"
            or verified_no_shrink
        )
        if (_is_codex_gateway_model(model)
                and deterministic_context_failure
                and not _sessions_with_inflight_tasks.get(sid)
                and not _session_has_live_watcher(sid)):
            try:
                recovery = await _recover_context_session(
                    sid,
                    model,
                    pre_tokens=before_total,
                    context_limit=context_limit,
                )
            except Exception as recovery_error:
                sys.stderr.write(
                    f"[chat] context recovery failed sid={sid[:8]}: "
                    f"{type(recovery_error).__name__}\n")
                sys.stderr.flush()
                raise HTTPException(
                    500, "native compact failed and recovery could not be created"
                ) from None
            return {
                "ok": True,
                "recovered": True,
                "recovered_session": recovery["session"],
                "recovery_stats": recovery["stats"],
            }
        status = (409 if classified["kind"] == "context_window"
                  else 502 if classified["kind"] == "model_route"
                  else int(e.info.get("api_error_status") or 500))
        if status < 400 or status > 599:
            status = 500
        sys.stderr.write(
            f"[chat] native /compact rejected sid={sid[:8]} "
            f"kind={classified['kind']}\n")
        sys.stderr.flush()
        raise HTTPException(status, str(e)) from None
    except Exception as e:
        if not any(tail_outcome.values()):
            tail_outcome = await asyncio.to_thread(
                _compact_tail_outcome, tail_path, tail_offset)
        if tail_outcome.get("boundary") and tail_outcome.get("summary"):
            # A generic transport exception can race the durable compact
            # commit just like an in-band SDK error.  The new transcript root
            # remains authoritative.
            await disconnect_client(sid)
            _schedule_post_compact_refresh(sid, model, {})
            return {"ok": True, "recovered": False}
        if (_is_codex_gateway_model(model)
                and _is_context_window_failure(e)
                and not _sessions_with_inflight_tasks.get(sid)
                and not _session_has_live_watcher(sid)):
            recovery_used, recovery_limit = _context_recovery_inputs(
                sid,
                model,
                pre_tokens=before_total,
                context_limit=context_limit,
            )
            try:
                recovery = await _recover_context_session(
                    sid,
                    model,
                    pre_tokens=recovery_used,
                    context_limit=recovery_limit,
                )
            except Exception as recovery_error:
                sys.stderr.write(
                    f"[chat] context recovery failed sid={sid[:8]}: "
                    f"{type(recovery_error).__name__}\n")
                sys.stderr.flush()
                raise HTTPException(
                    500, "native compact failed and recovery could not be created"
                ) from None
            return {
                "ok": True,
                "recovered": True,
                "recovered_session": recovery["session"],
                "recovery_stats": recovery["stats"],
            }
        sys.stderr.write(
            f"[chat] native /compact failed sid={sid[:8]} "
            f"exc={type(e).__name__} "
            f"kind={_classify_stream_error(str(e)).get('kind', 'unknown')}\n")
        sys.stderr.flush()
        raise HTTPException(500, "native /compact failed — see server log") from None
    # Refresh the cached context-usage snapshot from the now-compacted live
    # client so the meter drops immediately and STAYS dropped. /usage reads
    # _session_usage first; on a miss it falls back to
    # _session_usage_from_jsonl, which takes the LAST assistant turn's
    # cumulative usage. But /compact writes an isCompactSummary record, NOT a
    # fresh low-usage assistant turn — so that JSONL path keeps reporting the
    # PRE-compact (large) number until the next real message, leaving the ring
    # stuck at its pre-compact %. Mirror the stream done-handler (chat.py
    # ~5851): pull SDK totalTokens/maxTokens off the same client we just ran
    # /compact on (its in-memory context is the compacted one) and write them
    # back into _session_usage so every subsequent /usage poll is correct.
    try:
        cu = post_compact_usage or dict(await client.get_context_usage())
        real_max = int(cu.get("maxTokens") or 0)
        real_total = int(cu.get("totalTokens") or 0)
        sess_u = _session_usage.setdefault(sid, {
            "input_tokens": 0, "output_tokens": 0,
            "cache_read_tokens": 0, "cache_creation_tokens": 0,
            "total_cost_usd": 0.0, "last_turn_at": 0.0,
            "context_used": 0, "context_used_pct": 0.0, "context_limit": 0,
        })
        if real_total:
            sess_u["context_used"] = real_total
        if endpoints.is_third_party(model):
            sess_u["sdk_context_max_tokens"] = real_max
            sess_u["sdk_context_raw_max_tokens"] = _positive_int(cu.get("rawMaxTokens"))
            if real_total:
                _mark_context_used(sess_u, "sdk_context", estimate=True)
        elif real_max:
            sess_u["context_limit"] = real_max
            if real_total:
                _mark_context_used(sess_u, "sdk_context", estimate=False)
            try:
                sess.set_session_ctx_window(sid, real_max)
            except Exception:
                pass
        lim = int(sess_u.get("context_limit", 0) or 0)
        if lim and real_total:
            sess_u["context_used_pct"] = round(real_total / lim * 100, 1)
    except Exception as _e:
        sys.stderr.write(
            f"[chat] post-compact ctx refresh skipped for sid={sid[:8]}: "
            f"{type(_e).__name__}\n")
        sys.stderr.flush()
    # Message/turn recount can scan a very large transcript. It is
    # presentation bookkeeping, not part of compact correctness; schedule it
    # after the verified token drop so the browser can leave "compacting"
    # immediately. The next session read also self-heals these counters.
    _schedule_post_compact_refresh(sid, model, dict(post_compact_usage or {}))
    return {"ok": True, "recovered": False}


class ForkReq(BaseModel):
    # Inclusive — fork copies the transcript up to and including this
    # message UUID. To branch BEFORE a user message (e.g. for an edit-and-
    # retry), pass the UUID of the previous assistant message.
    # Omit / null = no truncation, copy the full transcript.
    up_to_message_id: str | None = None
    title: str | None = None
    activity_hidden: bool = False
    runtime_profile: Literal["", "side_question"] = ""


class RetryLastTurnReq(BaseModel):
    user_message_id: str = Field(min_length=1, max_length=128)


def _retry_prompt_text(message: Any) -> str:
    """Return an exact text-only SDK user payload or fail closed.

    Replaying a vision/document turn without its binary blocks would look
    successful while asking a materially different question.  MuseLab keeps
    that case unavailable until the SDK exposes a native attachment replay
    handle; no prompt or attachment bytes are copied into retry metadata.
    """
    raw = getattr(message, "message", None)
    content = raw.get("content") if isinstance(raw, dict) else None
    if isinstance(content, str):
        if content.strip() and not _is_cli_interrupt_message(content):
            return content
        raise HTTPException(409, "the last user turn has no retryable text")
    if not isinstance(content, list) or not content:
        raise HTTPException(409, "the last user turn has no retryable text")
    text_parts: list[str] = []
    for block in content:
        if not isinstance(block, dict) or block.get("type") != "text":
            raise HTTPException(
                409, "attachment turns cannot be retried without re-attaching")
        text = block.get("text")
        if not isinstance(text, str):
            raise HTTPException(409, "the last user turn has invalid text")
        text_parts.append(text)
    prompt = "".join(text_parts)
    if not prompt.strip() or _is_cli_interrupt_message(prompt):
        raise HTTPException(409, "the last user turn has no retryable text")
    return prompt


def _create_last_turn_retry_child(
    sid: str,
    user_message_id: str,
) -> dict:
    """Validate canonical tail history and persist one SDK-native retry fork."""
    with sess.session_lifecycle_lock(sid):
        if sess.session_is_deleting(sid):
            raise HTTPException(404, "session not found")
        source = sess.get_session_meta(sid)
        if source is None:
            raise HTTPException(404, "session not found")
        queue = sess.get_queue(sid)
        if (queue.get("items") or queue.get("inflight")):
            raise HTTPException(409, "cannot retry while messages are queued")

        source_model = str(source.get("model") or MODEL)
        try:
            messages = _get_session_msgs(sid, source_model)
        except Exception as exc:
            sys.stderr.write(
                f"[chat] retry history read failed sid={sid[:8]} "
                f"exc={type(exc).__name__}\n"
            )
            sys.stderr.flush()
            raise HTTPException(409, "canonical session history is unavailable") from None

        real_users = [
            (index, item)
            for index, item in enumerate(messages)
            if _is_real_user_prompt(item)
        ]
        if not real_users:
            raise HTTPException(409, "session has no retryable user turn")
        target_index, target = real_users[-1]
        if str(getattr(target, "uuid", "")) != user_message_id:
            raise HTTPException(409, "only the latest user turn can be retried")
        prompt = _retry_prompt_text(target)

        resume_at = ""
        if target_index > 0:
            resume_at = str(getattr(messages[target_index - 1], "uuid", ""))
            if _canonical_uuid_component(resume_at) is None:
                raise HTTPException(409, "safe retry boundary is unavailable")

        source_name = str(source.get("name") or "会话").strip()
        suffix = "重试" if is_chinese_locale() else "Retry"
        child_name = f"{source_name} · {suffix}"
        register_kwargs = {
            "name": child_name,
            "model": source_model,
            "permission": _validate_permission(
                str(source.get("permission") or "")),
            "plan_return_permission": source.get("plan_return_permission"),
            "auto_named": False,
            "message_count": target_index,
            "turn_count": sum(
                1 for item in messages[:target_index]
                if _is_real_user_prompt(item)
            ),
            "effort": _normalize_effort(source.get("effort")),
            "service_tier": str(source.get("service_tier") or ""),
            "thinking": source.get("thinking") is not False,
            "forked_from": sid,
            "forked_from_name": source_name,
            "forked_from_message_id": resume_at,
            "activity_hidden": bool(source.get("activity_hidden")),
            "runtime_profile": str(source.get("runtime_profile") or ""),
            "cwd": source.get("cwd") or str(ROOT),
        }
        try:
            if resume_at:
                def _fork_retry():
                    with _session_config_dir(source_model, sid=sid):
                        return sdk_fork_session(
                            sid,
                            directory=str(sess.session_workspace(sid)),
                            up_to_message_id=resume_at,
                            title=child_name,
                        )

                lifecycle = _commit_fork_lifecycle(
                    sid,
                    source,
                    fork_child=_fork_retry,
                    register_kwargs=register_kwargs,
                    successor=False,
                )
                child_sid = str(lifecycle["child_sid"])
                child = lifecycle["child_meta"]
            else:
                child_sid = str(uuid.uuid4())
                child = sess.register_session(
                    child_sid,
                    **register_kwargs,
                )
        except FileNotFoundError:
            raise HTTPException(409, "source transcript is unavailable") from None
        except ValueError as exc:
            raise HTTPException(400, str(exc)) from None
        except HTTPException:
            raise
        except Exception as exc:
            sys.stderr.write(
                f"[chat] retry fork failed sid={sid[:8]} "
                f"exc={type(exc).__name__}\n"
            )
            sys.stderr.flush()
            raise HTTPException(500, "retry fork failed — see server log") from None
    return {
        **child,
        "session_id": child_sid,
        "source_session_id": sid,
        "target_user_message_id": user_message_id,
        "prompt": prompt,
        "retry_mode": "native_fork" if resume_at else "fresh",
    }


@router.post("/sessions/{sid}/retry-last-turn",
             dependencies=[Depends(require_token)])
async def retry_last_turn_api(sid: str, req: RetryLastTurnReq) -> dict:
    target = _canonical_uuid_component(req.user_message_id)
    if target is None:
        raise HTTPException(400, "invalid user message id")
    if _session_runtime_busy(sid):
        raise HTTPException(409, "cannot retry while the session is active")
    drain = _queue_drain_tasks.get(sid)
    if drain is not None and not drain.done():
        raise HTTPException(409, "cannot retry while the queue is advancing")
    return await obs.to_thread_io(
        "chat.retry_child_create",
        sid,
        _create_last_turn_retry_child,
        sid,
        target,
    )


@router.post("/sessions/{sid}/fork", dependencies=[Depends(require_token)])
def fork_session_api(sid: str, req: ForkReq) -> dict:
    return chat_successor.fork_session(
        sid,
        up_to_message_id=req.up_to_message_id,
        title=req.title,
        activity_hidden=req.activity_hidden,
        runtime_profile=req.runtime_profile,
    )


def _runtime_fork_uuid_mapping(child_sid: str) -> dict[str, str]:
    return chat_successor.runtime_fork_uuid_mapping(child_sid)


def _sync_runtime_successor_postlude(source_sid: str) -> dict[str, int]:
    return chat_successor.sync_runtime_successor_postlude(source_sid)


def _runtime_fork_boundary(sid: str, meta: dict) -> str:
    return chat_successor.runtime_fork_boundary(sid, meta)


def _backfill_runtime_task_overlays(source_sid: str) -> None:
    return chat_successor.backfill_runtime_task_overlays(source_sid)


async def _continue_detached_runtime_locked(source_sid: str) -> dict:
    return await chat_successor.continue_detached_runtime_locked(source_sid)


async def _continue_detached_runtime(source_sid: str) -> dict:
    return await chat_successor.continue_detached_runtime(source_sid)


async def _prepare_detached_successor_runtime(source_sid: str) -> None:
    return await chat_successor.prepare_detached_successor_runtime(source_sid)


def _schedule_detached_successor_prewarm(source_sid: str) -> None:
    return chat_successor.schedule_detached_successor_prewarm(source_sid)


@router.post("/sessions/{sid}/continue-detached",
             dependencies=[Depends(require_token)])
async def continue_detached_session_api(sid: str) -> dict:
    return await _continue_detached_runtime(sid)

class BudgetReq(BaseModel):
    budget_usd: float       # 0 = disabled


def _claude_md_filled_ratio(path: Path) -> tuple[int, float]:
    """Heuristic: how much of a CLAUDE.md is actually filled vs. template.

    Returns (filled_content_lines, fill_ratio_0_to_1).

    Older releases and optional intake helpers could seed a long template full
    of section headers and empty placeholders. Just checking ``lines > 0``
    would confuse "file exists" with "contains effective instructions."

    We count a line as "filled" only if it carries user content:
      - skip blank lines, pure markdown punctuation (---, ===, |...|)
      - skip pure headers (#, ##, ###)
      - skip comment lines (<!-- ... -->)
      - skip lines that are just a label with no value
        (e.g. "Name:" or "- 配偶 / 关系：" with nothing after the colon)
      - skip the leading blockquote intro paragraph (> ...) used by the
        default template's preamble — informational, not user content
      - skip lines under a "delete-if-not-applicable" instruction that
        still match the template's bullet labels exactly (best-effort
        heuristic: anything that contains BOTH "(" and ":" but ends in
        ":" is probably an unfilled prompt line)
    """
    try:
        raw = path.read_text(encoding="utf-8", errors="replace")
    except OSError:
        return (0, 0.0)
    lines = raw.splitlines()
    filled = 0
    total_content = 0  # lines that COULD be content (excludes pure structure)
    for ln in lines:
        s = ln.strip()
        if not s:
            continue
        # Pure structure / decoration — never counts as content
        if s.startswith("#"):              # headers
            continue
        if s.startswith("---") or s.startswith("==="):
            continue
        if s.startswith("<!--") and s.endswith("-->"):
            continue
        if s.startswith("> "):             # block-quote preamble in template
            continue
        if s.startswith("|") and s.endswith("|"):  # markdown table rows
            # Tables can be content OR template (e.g. "| Date | What |"). We
            # treat them as content only if a non-header cell has > 2 chars.
            cells = [c.strip() for c in s.strip("|").split("|")]
            if any(len(c) > 2 and c not in ("---", ":---", "---:") for c in cells):
                total_content += 1
                filled += 1
            continue
        total_content += 1
        # "Label:" with nothing meaningful after → unfilled prompt
        # Examples seeded by template:
        #   "- Project name:"
        #   "- Validation command:"
        #   "项目名称："
        # If the line ends in ":" or "：", or has only label-colon-whitespace,
        # it's an unfilled prompt.
        if s.endswith(":") or s.endswith("："):
            continue
        # Lines like "- 居住：" (bullet + label + colon at end)
        if s.endswith(":)") or s.endswith("：)"):
            continue
        # "(e.g. ...)" placeholder example lines — template hints, not user
        # content. Heuristic: starts with "(" or "（".
        if s.startswith("(") or s.startswith("（"):
            continue
        # Anything left is user-supplied content
        filled += 1
    ratio = (filled / total_content) if total_content > 0 else 0.0
    return (filled, ratio)


def _scan_claude_md_source(scope: str, path: Path) -> dict | None:
    """Build a source descriptor for one CLAUDE.md path, or None if absent."""
    if not path.exists() or not path.is_file():
        return None
    try:
        total_lines = sum(1 for _ in path.open(encoding="utf-8", errors="replace"))
        filled_lines, ratio = _claude_md_filled_ratio(path)
        return {
            "scope": scope,
            "path": str(path),
            "lines": total_lines,
            "filled_lines": filled_lines,
            "fill_ratio": round(ratio, 3),
            "meaningfully_filled": filled_lines >= 6,  # arbitrary but useful threshold
            "mtime": path.stat().st_mtime,
        }
    except OSError:
        return None


@router.get("/context-info", dependencies=[Depends(require_token)])
def context_info(
    workspace_root: Path = Depends(resolve_workspace_root),
) -> dict:
    """Information about the selected workspace and available model context.

    SDK options pass `setting_sources=["user", "project", "local"]`, so
    project instructions can come from these CLAUDE.md sources:
      - project scope: workspace/CLAUDE.md
      - project local override: workspace/CLAUDE.local.md
      - project dot scope: workspace/.claude/CLAUDE.md
      - user scope: ~/.claude/CLAUDE.md
      - per-subdir: workspace/{subdir}/CLAUDE.md

    We also distinguish "file exists" from "file actually has user content"
    via a filled-ratio heuristic — the install script seeds a long bilingual
    template, so plain ``lines > 0`` is not a useful readiness signal.
    """
    # Project-scope candidates in the selected workspace.
    candidates: list[tuple[str, Path]] = [
        ("project",       workspace_root / "CLAUDE.md"),
        ("project_local", workspace_root / "CLAUDE.local.md"),
        ("project_dot",   workspace_root / ".claude" / "CLAUDE.md"),
        ("user",          Path.home() / ".claude" / "CLAUDE.md"),
    ]
    # Per-subdirectory CLAUDE.md (one level deep, skip hidden control dirs).
    try:
        for sub in sorted(workspace_root.iterdir()):
            if not sub.is_dir():
                continue
            if sub.name.startswith("."):
                continue
            candidates.append((f"subdir:{sub.name}", sub / "CLAUDE.md"))
    except OSError:
        pass

    sources: list[dict] = []
    for scope, path in candidates:
        s = _scan_claude_md_source(scope, path)
        if s is not None:
            sources.append(s)

    # Detect "do we have ANY working auth?" — needed so the chat-empty card
    # can warn "you have no provider set up; configure one before chatting".
    # Three valid Anthropic-side auth sources:
    #   1. Pro/Max OAuth (~/.claude/.credentials.json)
    #   2. ANTHROPIC_API_KEY  → x-api-key header
    #   3. ANTHROPIC_AUTH_TOKEN → Authorization: Bearer (OAuth/enterprise)
    # has_any_provider previously only checked #1 + third-party vendors,
    # so users who configured ANTHROPIC_API_KEY in Settings got a stuck
    # "no provider configured" warning (observed after clear-localStorage).
    claude_oauth = (Path.home() / ".claude" / ".credentials.json").exists()
    anthropic_api = bool(os.environ.get("ANTHROPIC_API_KEY")
                          or os.environ.get("ANTHROPIC_AUTH_TOKEN"))
    from . import endpoints as _ep
    # Return human-readable display names ("DeepSeek", "智谱 GLM"…) not raw
    # env keys ("DEEPSEEK_API_KEY"…) — the FE / tests treat this as a
    # user-facing list. A prior refactor briefly emitted env_key; broke
    # test_settings_put_reflects_in_context_info. Stay on display names.
    third_party_configured = [
        p.display for p in _ep.catalog()
        if os.environ.get(p.env_key)
    ]
    # Back-compat: keep claude_md_exists / lines / mtime fields for any
    # consumer that hasn't migrated to the new claude_md_sources list.
    # Reflect "ANY source present" + union total lines + latest mtime so
    # the existing UI keeps working without changes.
    total_lines = sum(s["lines"] for s in sources)
    latest_mtime = max((s["mtime"] for s in sources), default=0.0)
    # Distinguish "file present" from "contains effective instructions".
    # A freshly installed template can be long while containing no configured
    # values, so consumers need a readiness signal beyond file existence.
    meaningfully_filled = any(s["meaningfully_filled"] for s in sources)
    workspace_empty = True
    subdir_present: dict[str, bool] = {}
    try:
        for item in workspace_root.iterdir():
            # Hidden entries are workspace/runtime control state from the file
            # browser's default perspective. Instruction files are reported
            # separately above and do not by themselves make a workspace ready
            # for file organization.
            if (
                item.name.startswith(".")
                or item.name in {"CLAUDE.md", "CLAUDE.local.md"}
            ):
                continue
            workspace_empty = False
            if item.is_dir():
                subdir_present[item.name] = True
    except OSError:
        pass

    info: dict = {
        "workspace_root": str(workspace_root),
        "workspace_empty": workspace_empty,
        "claude_md_exists": len(sources) > 0,
        "claude_md_lines": total_lines,
        "claude_md_mtime": latest_mtime,
        "claude_md_sources": sources,
        "claude_md_meaningfully_filled": meaningfully_filled,
        "has_claude_oauth": claude_oauth,
        "has_anthropic_api": anthropic_api,
        "third_party_configured": third_party_configured,
        "has_any_provider": (
            claude_oauth or anthropic_api or len(third_party_configured) > 0
        ),
        # Deprecated for one compatibility cycle. New consumers must use the
        # workspace_* names; subdir_present is now a dynamic directory map and
        # no longer advertises a predefined directory taxonomy.
        "archive_root": str(workspace_root),
        "archive_empty": workspace_empty,
        "subdir_present": subdir_present,
    }
    return info


@router.get("/probe/{model}", dependencies=[Depends(require_token)])
async def probe_provider(model: str) -> dict:
    """Hit the vendor's anthropic-compat endpoint with the configured key and
    return what the vendor said. Lets the user self-diagnose 401 / wrong-host
    / wrong-key issues WITHOUT pasting keys into chat. Always returns 200 on
    our side — the body carries vendor's status, headers, and partial body."""
    import httpx
    p = endpoints.lookup(model)
    if p is None:
        return {"ok": False, "reason": f"unknown model: {model}"}
    key = os.environ.get(p.env_key, "")
    if not key:
        return {"ok": False, "reason": f"{p.env_key} not configured (Settings → Provider API Keys)"}
    # Use the live-resolved base URL (env override > catalog default) so a
    # proxy / on-prem URL probe doesn't silently hit the public endpoint.
    base = endpoints._resolve_base_url(p.env_key, p) or p.base_url
    url = base.rstrip("/") + "/v1/messages"
    headers = {
        "x-api-key": key,
        "anthropic-version": "2023-06-01",
        "content-type": "application/json",
    }
    # Strip internal prefixes like "qwen-intl:" before sending to the API.
    api_model = endpoints.normalize_model_id(model)
    body = {"model": api_model, "max_tokens": 16,
             "messages": [{"role": "user", "content": "ping"}]}
    try:
        async with httpx.AsyncClient(timeout=15) as client:
            r = await client.post(url, json=body, headers=headers)
        snippet = r.text[:500]
        return {
            "ok": r.status_code == 200,
            "vendor": p.display, "model": model, "url": url,
            "status": r.status_code,
            "key_hint": f"{key[:4]}…{key[-4:]}" if len(key) > 12 else "***",
            "vendor_response_excerpt": snippet,
        }
    except Exception as e:
        return {"ok": False, "reason": f"transport error: {type(e).__name__}: {e}",
                 "url": url}


@router.put("/budget", dependencies=[Depends(require_token)])
async def set_budget(req: BudgetReq) -> dict:
    """Set the soft budget cap. Stored in env (process-lifetime only — for a
    persistent cap, edit MUSELAB_BUDGET_USD in .env via /api/settings)."""
    if req.budget_usd < 0:
        raise HTTPException(400, "budget must be >= 0")
    os.environ["MUSELAB_BUDGET_USD"] = str(req.budget_usd)
    return {"ok": True, "budget_usd": req.budget_usd}


@router.get("/mcp", dependencies=[Depends(require_token)])
def mcp_status() -> dict:
    """Return configured MCP servers (merged view: muselab's mcp.json +
    Claude Code's standard config locations) for UI display. Source field
    tells the caller where each entry came from. `configured: True` if at
    least one server resolved from anywhere."""
    try:
        from .api_settings import _load_mcp_merged
        merged = _load_mcp_merged()
        return {
            "configured": bool(merged),
            "servers": [
                {
                    "name": name,
                    "command": s.get("command", ""),
                    "args": s.get("args", []),
                    "source": s.get("_source", "muselab"),
                    "disabled": bool(s.get("disabled", False)),
                }
                for name, s in merged.items()
            ],
        }
    except Exception as e:
        return {"configured": False, "servers": [], "error": str(e)}


# Session ids whose most recent /interrupt has been called but the in-flight
# stream's ResultMessage handler hasn't observed it yet. Consumed (set.discard)
# the moment the handler runs, used to:
#   (a) suppress the turn-done Web Push — user just cancelled, getting a "Muse
#       已回复" buzz is noise, sometimes confusing ("did the reply come through
#       after all?").
#   (b) tag the SSE done event with `cancelled: true` so the frontend doesn't
#       paint completion-success UI (turn footer ts stamp / scroll-to-bottom
#       are still fine; what we hide is celebratory toasts / push).
# Module-level set is fine for the single-user model muselab targets — no
# cross-user race to worry about.
_pending_interrupts: set[tuple[str, str]] = set()

# How long the force-stop watchdog waits for the SDK's control-protocol
# interrupt to drain the turn on its own before tearing the client down.
# The SDK `client.interrupt()` is best-effort: for an agentic turn the bundled
# CLI does not always abort promptly (observed: turn keeps running, the slot
# stays in `_active_turns`, every subsequent send bounces with "previous turn
# still running" until the 30-min outer timeout). If the turn hasn't ended
# within this grace window we kill the CLI subprocess to guarantee the slot
# frees. Kept short so the user can resend quickly, but long enough that a
# legitimately-fast interrupt completes naturally (warm-client preserved).
# The SDK's own interrupt control request defaults to a 60-second acknowledgement
# timeout. A Stop button must never inherit that latency. Give a healthy CLI a
# brief chance to acknowledge, while the force-stop timer runs in parallel.
_INTERRUPT_ACK_TIMEOUT_S = max(
    0.05, env_float("MUSELAB_INTERRUPT_ACK_TIMEOUT_S", 0.35))
_INTERRUPT_FORCE_GRACE_S = max(
    _INTERRUPT_ACK_TIMEOUT_S,
    env_float("MUSELAB_INTERRUPT_FORCE_GRACE_S", 0.5))

# How long a NEW turn waits for an already-interrupted (cancelled) turn to
# finish draining before it gives up with _TurnBusy. Must comfortably exceed
# _INTERRUPT_FORCE_GRACE_S + teardown time so the force-stop watchdog always
# wins the race and the user's resend transparently succeeds instead of seeing
# "previous turn still running" during the teardown window.
_INTERRUPT_DRAIN_WAIT_S = 6.0
_INTERRUPT_FORCE_OWNER_JOIN_S = 2.0
_INTERRUPT_FORCE_DISCONNECT_JOIN_S = 0.25


def _interrupt_response(
    session_id: str,
    broadcast: "TurnBroadcast | None",
    *,
    requested_turn_id: str,
    interrupted: list[str],
    note: str = "",
    phase: str = "",
    stale: bool = False,
) -> dict:
    """Return the authoritative exact-turn state after an interrupt attempt."""
    current = _active_turns.get(session_id)
    active = bool(current is not None and not current.done)
    current_turn_id = current.turn_id if active else ""
    target_turn_id = requested_turn_id or str(
        getattr(broadcast, "turn_id", "") or "")
    owns_requested = bool(
        active and target_turn_id and current_turn_id == target_turn_id)
    owner = current if owns_requested else broadcast
    result = {
        "ok": True,
        "interrupted": interrupted,
        "stale": bool(stale),
        "requested_turn_id": requested_turn_id,
        "current_turn_id": current_turn_id,
        # Preserve the requested owner on inactive responses so a frontend can
        # settle only that exact stopping turn, never an ABA successor.
        "turn_id": current_turn_id if active else requested_turn_id,
        "active": active,
        "stopping": bool(owns_requested and current.cancelled),
        "phase": (
            str(phase or getattr(owner, "startup_phase", "") or "running")
            if active else "inactive"
        ),
    }
    if note:
        result["note"] = note
    return result


@router.post("/interrupt", dependencies=[Depends(require_token_header_or_query)])
async def interrupt(
    session_id: str,
    turn_id: str = "",
) -> dict:
    """Stop one immutable turn via the SDK control protocol."""
    requested_turn_id = turn_id.strip()
    bc = _active_turns.get(session_id)
    if requested_turn_id and (
        bc is None or bc.done or bc.turn_id != requested_turn_id
    ):
        # A delayed Stop from an older browser turn must never pause the queue or
        # interrupt a newer ABA turn that reused the same session/client.
        return _interrupt_response(
            session_id,
            bc,
            requested_turn_id=requested_turn_id,
            interrupted=[],
            stale=True,
        )
    # Stop means "do not continue autonomously". Pause queued work only after
    # the immutable owner check, before the SDK interrupt can race cleanup and
    # dequeue the next item.
    stale_after_pause = False
    async with _lock:
        current = _active_turns.get(session_id)
        owner_matches = (
            current is bc
            and (
                bc is None
                or (
                    not bc.done
                    and (
                        not requested_turn_id
                        or bc.turn_id == requested_turn_id
                    )
                )
            )
        )
        if not owner_matches:
            stale_after_pause = True
            targets = []
        else:
            # Keep admission serialized with both exact-owner checks. Turn
            # cleanup can still pop its own broadcast outside this lock, so a
            # second check after the disk await is required before snapshotting
            # any pooled client.
            await obs.to_thread_io(
                "chat.queue_pause_nonempty",
                session_id,
                sess.pause_queue_if_nonempty,
                session_id,
                owned=True,
            )
            current = _active_turns.get(session_id)
            owner_matches = (
                current is bc
                and (
                    bc is None
                    or (
                        not bc.done
                        and (
                            not requested_turn_id
                            or bc.turn_id == requested_turn_id
                        )
                    )
                )
            )
            if not owner_matches:
                stale_after_pause = True
                targets = []
            else:
                targets = [
                    (k, c) for k, c in _clients.items()
                    if k[0] == session_id
                ]
    if stale_after_pause:
        return _interrupt_response(
            session_id,
            bc,
            requested_turn_id=requested_turn_id,
            interrupted=[],
            stale=True,
        )
    # Mark the active turn user-cancelled up front (BEFORE calling SDK's
    # interrupt — the ResultMessage handler races with us, and we'd rather flag
    # too early than too late). This also lets the force-stop watchdog and the
    # event_gen error branch convert a teardown-induced transport error into a
    # clean `cancelled` event instead of a red error toast.
    if bc is not None and not bc.done:
        bc.cancelled = True
        if not bc.cancelled_at_ms:
            bc.cancelled_at_ms = int(time.time() * 1000)
        # Arm the hard-stop deadline from the CLICK, not after waiting for the
        # SDK control request. The old ordering added its possible 60s timeout
        # in front of the 2.5s grace period.
        asyncio.create_task(_force_stop_after_grace(session_id, bc))
        startup_task = getattr(bc, "startup_task", None)
        startup_owner = getattr(bc, "startup_owner_task", None)
        cancelled_startup = False
        if startup_task is not None and not startup_task.done():
            # Cold-start cancellation: no client has reached `_clients` yet,
            # so client.interrupt() cannot help. Cancelling this task unwinds
            # CLI/MCP initialization; _start_turn converts it to a replayable
            # `cancelled` terminal event.
            startup_task.cancel()
            cancelled_startup = True
        elif (bc.task is None and startup_owner is not None
              and not startup_owner.done()):
            # Attachment preparation and the final intent write happen after
            # client startup but before the detached pump exists. Cancel their
            # outer owner; its shielded cleanup joins any real worker first.
            startup_owner.cancel()
            cancelled_startup = True
        if cancelled_startup:
            return _interrupt_response(
                session_id,
                bc,
                requested_turn_id=requested_turn_id,
                interrupted=[f"{session_id}@startup"],
                phase="starting",
            )
    if not targets:
        # No live client in the pool, but a detached pump task may still be
        # holding the _active_turns slot. Schedule the watchdog anyway so the
        # session can't get wedged. Don't set the pending-interrupt flag: with
        # no turn to suppress a push for, leaving it set would wrongly mute the
        # NEXT turn's done-push.
        return _interrupt_response(
            session_id,
            bc,
            requested_turn_id=requested_turn_id,
            interrupted=[],
            note="no live client",
        )
    _pending_interrupts.add((
        session_id,
        bc.turn_id if bc is not None else requested_turn_id,
    ))

    async def _interrupt_one(k, c) -> str | None:
        try:
            interrupt_call = (
                c.interrupt(cancel_queued=bool(
                    bc is not None and bc.steering_commands))
                if isinstance(c, MuseLabSDKClient)
                else c.interrupt()
            )
            await asyncio.wait_for(
                interrupt_call, timeout=_INTERRUPT_ACK_TIMEOUT_S)
            return f"{k[0]}@{k[1]}"
        except asyncio.TimeoutError:
            sys.stderr.write(
                f"[chat-interrupt] sid={obs.short_id(k[0])} "
                f"ack timed out after "
                f"{_INTERRUPT_ACK_TIMEOUT_S:.2f}s; force-stop armed\n")
        except Exception as e:
            sys.stderr.write(
                f"[chat-interrupt] sid={obs.short_id(k[0])} "
                f"failed exc={type(e).__name__}\n")
        return None

    results = await asyncio.gather(
        *(_interrupt_one(k, c) for k, c in targets))
    interrupted = [result for result in results if result is not None]
    if bc is not None and bc.steering_commands:
        # Stop is an explicit request not to continue autonomously. Even when
        # an older CLI omits its cancel_queued receipt, retain each durable
        # adjustment as cancelled+paused for review; never blindly resend a
        # command that may already have crossed the dequeue boundary.
        await _cancel_outstanding_steering_commands(bc)
    # The watchdog was armed before these control requests, so a slow/broken
    # acknowledgement cannot postpone the hard-stop deadline.
    return _interrupt_response(
        session_id,
        bc,
        requested_turn_id=requested_turn_id,
        interrupted=interrupted,
    )


@router.post("/sessions/{sid}/tasks/{task_id}/stop",
             dependencies=[Depends(require_token)])
async def stop_background_task(sid: str, task_id: str) -> dict:
    """Stop a running background task via the SDK's native stop_task()
    control request (client.py:450) — the user's only handle on a runaway
    run_in_background task short of killing the whole turn.

    SINGLE-READER SAFETY: stop_task only WRITES a control request; the
    control RESPONSE is consumed by the SDK's internal control-protocol
    reader (not receive_messages), so calling it from this HTTP coroutine
    never races the turn pump / cross-turn watcher on the message stream —
    same invariant as interrupt() above.

    After the CLI acks, it emits a task_notification with status='stopped'
    on the message stream, which flows through the normal settle paths
    (_on_task_settled → card flip + unpin; no push — task settlement is
    deliberately notification-free, see _on_task_settled), so this
    endpoint needs no settlement logic of its own."""
    async with _lock:
        targets = [(k, c) for k, c in _clients.items() if k[0] == sid]
    if not targets:
        # No live client → the CLI that owned the task is gone; the task is
        # dead-or-settled already. 409 (not 404) so the FE can distinguish
        # "nothing to stop" from a bad route.
        raise HTTPException(
            status_code=409,
            detail="no live client for session — task already settled?")
    errors: list[str] = []
    for k, c in targets:
        try:
            await c.stop_task(task_id)
            return {"ok": True, "task_id": task_id}
        except Exception as e:
            errors.append(f"{k[0]}@{k[1]}: {type(e).__name__}: {e}")
            sys.stderr.write(
                f"[chat] stop_task failed sid={obs.short_id(sid)} "
                f"task={obs.short_id(task_id)} exc={type(e).__name__}\n")
    raise HTTPException(status_code=502, detail="; ".join(errors))


async def _force_stop_after_grace(
    session_id: str,
    bc: "TurnBroadcast",
    grace: float = _INTERRUPT_FORCE_GRACE_S,
) -> None:
    """Guarantee Stop settles one turn through its shared terminal owner.

    A pump task can be absent, cancelled before its coroutine starts (and thus
    run no ``finally``), or ignore cancellation inside SDK code. The watchdog
    first gives the ordinary owner a bounded chance, then joins/creates the
    broadcast's shielded startup finalizer. That finalizer joins attachment
    preparation, removes artifacts, releases the lease, and only then settles
    Activity/queue/sidecar/active-turn state.
    """
    try:
        await asyncio.sleep(grace)
        if _active_turns.get(session_id) is not bc or bc.done:
            return
        sys.stderr.write(
            f"[chat-interrupt] sid={obs.short_id(session_id)} "
            f"did not drain after {grace:.1f}s; "
            f"forcing client teardown\n")
        sys.stderr.flush()
        bc.cancelled = True
        disconnect_task = asyncio.create_task(disconnect_client(session_id))
        try:
            await asyncio.wait_for(
                asyncio.shield(disconnect_task),
                timeout=_INTERRUPT_FORCE_DISCONNECT_JOIN_S,
            )
        except asyncio.TimeoutError:
            # A stuck subprocess teardown cannot hold the attachment finalizer.
            _retain_detached_cleanup(disconnect_task)
        except Exception:
            pass
        finally:
            if not disconnect_task.done():
                _retain_detached_cleanup(disconnect_task)

        # Give the normal owner a short chance after transport teardown.
        deadline = time.monotonic() + _INTERRUPT_FORCE_OWNER_JOIN_S
        while time.monotonic() < deadline:
            if bc.done or _active_turns.get(session_id) is not bc:
                return
            await asyncio.sleep(0.05)

        owner = getattr(bc, "task", None)
        if owner is None or owner.done():
            startup_owner = getattr(bc, "startup_owner_task", None)
            if startup_owner is not None and not startup_owner.done():
                owner = startup_owner
        if owner is not None and not owner.done():
            owner.cancel()
            try:
                await asyncio.wait_for(
                    asyncio.shield(owner),
                    timeout=_INTERRUPT_FORCE_OWNER_JOIN_S,
                )
            except (asyncio.CancelledError, asyncio.TimeoutError):
                pass
            if bc.done or _active_turns.get(session_id) is not bc:
                return

        # This is also correct when owner is None/already-cancelled: unlike a
        # coroutine finally block, the retained attachment/startup state still
        # exists and has one explicit cleanup owner.
        mem0.pop_recall_trace(session_id)
        await obs.to_thread_io(
            "chat.queue_pause_nonempty",
            session_id,
            sess.pause_queue_if_nonempty,
            session_id,
            owned=True,
        )
        await _finish_cancelled_startup(session_id, bc)
    except Exception as e:
        sys.stderr.write(
            f"[chat-interrupt] force-stop watchdog failed "
            f"sid={obs.short_id(session_id)} exc={type(e).__name__}\n")
        sys.stderr.flush()


@router.post("/reset", dependencies=[Depends(require_token_header_or_query)])
async def reset(session_id: str | None = None) -> dict:
    if session_id:
        await disconnect_client(session_id)
        return {"ok": True, "reset": [session_id]}
    # Snapshot public keys first, then let each per-session teardown pop its
    # clients under the lock and wait outside it. The old implementation held
    # the global pool lock across an unbounded SDK disconnect, freezing every
    # get_client()/DELETE/reset request behind one wedged subprocess.
    async with _lock:
        keys = list(_clients.keys())
    session_ids = {k[0] for k in keys}
    results = await asyncio.gather(
        *(disconnect_client(sid) for sid in session_ids),
        return_exceptions=True,
    )
    for result in results:
        if isinstance(result, BaseException):
            raise result
    # Runtime key begins with (session_id, model); keep only those public bits.
    return {"ok": True, "reset": [f"{k[0]}@{k[1]}" for k in keys]}


# ====== streaming ======

# Per-field cap so a single tool_use payload stays bounded even when Write
# pastes a 500KB file. 100KB covers >99% of real Edit / Write inputs while
# capping the worst case at "still fits in one SSE frame, fits in the browser
# buffer". Truncation is marked inline so the FE can show "…and 90KB more"
# instead of silently rendering a partial diff.
_MAX_INPUT_FIELD_LEN = chat_presentation.MAX_INPUT_FIELD_LEN

# Single source of truth for which tool-input fields the FE actually renders.
# BOTH the realtime stream path (_render_tool_use) and the JSONL-reload path
# (_sdk_messages_to_ui) slim tool inputs to this set — keeping them identical
# so a reloaded session renders the same tool chips/labels the live stream did
# (previously the two whitelists had drifted: reload was missing the Task*
# family subject/activeForm/taskId/status fields).
_SLIM_INPUT_FIELDS = chat_presentation.SLIM_INPUT_FIELDS


def _slim_input_value(v: Any) -> Any:
    """Patchable facade for bounded tool-input transport values."""
    return chat_presentation.slim_input_value(
        v, max_length=_MAX_INPUT_FIELD_LEN)


def _render_tool_use(block: ToolUseBlock) -> dict:
    """Patchable facade for live tool-use rendering."""
    return chat_presentation.render_tool_use(
        block,
        max_input_field_len=_MAX_INPUT_FIELD_LEN,
        slim_input_fields=_SLIM_INPUT_FIELDS,
        slim_value=_slim_input_value,
    )


# Live SSE tool results still need a hard ceiling because the canonical JSONL
# may not exist yet. Historical responses use stable block references instead:
# large bodies stay in the authoritative transcript and are fetched on demand.
_TOOL_RESULT_PREVIEW_CAP = chat_presentation.TOOL_RESULT_PREVIEW_CAP
_TOOL_RESULT_TEXT_CAP = chat_presentation.TOOL_RESULT_TEXT_CAP
_HISTORY_INLINE_BODY_CAP = chat_presentation.HISTORY_INLINE_BODY_CAP
_HISTORY_BODY_PREVIEW_CAP = chat_presentation.HISTORY_BODY_PREVIEW_CAP


def _defer_large_ui_bodies(messages: list[dict]) -> None:
    """Patchable facade for deferred canonical-body transport shaping."""
    chat_presentation.defer_large_ui_bodies(
        messages,
        inline_body_cap=_HISTORY_INLINE_BODY_CAP,
        body_preview_cap=_HISTORY_BODY_PREVIEW_CAP,
        tool_result_preview_cap=_TOOL_RESULT_PREVIEW_CAP,
    )


# Bash output format from claude-code's CLI: stdout / stderr / exit_code are
# wrapped in pseudo-XML tags so we can split them apart for terminal-style
# rendering. Falls through gracefully when the tags aren't present (vendor
# wrappers / mocked runs); the FE then just renders the raw body.


def _classify_stream_error(err: Any) -> dict:
    """Tag a stream-error message with a kind + CTA hint + retryable flag so
    the FE can render a useful action button instead of just a red toast.

    Real-world breakdown (seen on the user's machine):
      - vendor 401 / "invalid api key" / "Not logged in"  → kind=auth, retry=N
      - "429" / "rate limit" / "quota exceeded"           → kind=quota, retry=Y
      - "Connection refused" / "timeout" / "ECONNRESET"   → kind=network, retry=Y
      - "Session ID already in use"                        → kind=session, retry=Y
      - "thinking signature"                               → kind=cross_vendor, retry=Y
      - everything else                                    → kind=unknown, retry=Y

    `cta`: optional opaque key the FE maps to a button label + handler
    (e.g. "open_settings", "switch_model", "retry"). FE falls back to a
    plain "Retry" button when None.
    """
    msg = str(err) if err is not None else ""
    low = msg.lower()
    kind = "unknown"
    cta: str | None = "retry"
    retryable = True
    if any(t in low for t in (
        "context window", "context length", "input exceeds",
        "input tokens exceed", "maximum tokens", "maximum context",
        "prompt too long", "input is too long", "request too large",
        "too many tokens",
    )):
        kind = "context_window"
        cta = "compact_or_fork"
        retryable = False
    elif any(t in low for t in (
        "unknown provider", "provider not found", "unsupported provider",
        "no provider for model", "unknown model provider",
    )):
        kind = "model_route"
        cta = "switch_model"
        retryable = False
    elif any(t in low for t in (
        "401", "invalid api key", "invalid_api_key",
        "not logged in", "requires auth", "no api key",
        "anthropic_api_key", "authentication",
    )):
        kind = "auth"
        cta = "open_settings"
        retryable = False
    elif any(t in low for t in (
        "429", "rate limit", "rate_limit", "quota", "too many requests",
        "overloaded",
    )):
        kind = "quota"
        cta = "switch_model"
    elif any(t in low for t in (
        "connection refused", "timeout", "timed out",
        "econnreset", "econnrefused", "enotfound", "network", "dns",
    )):
        kind = "network"
        cta = "retry"
    elif "thinking" in low and "signature" in low:
        # Cross-vendor switch left a Claude thinking-signature in history;
        # next turn from a non-Claude vendor fails validation. UX: tell the
        # user to clear / compact / fork.
        kind = "cross_vendor"
        cta = "compact_or_fork"
    elif "session" in low and ("already in use" in low or "already_in_use" in low):
        kind = "session"
        cta = "retry"
    elif "processerror" in low or "claudesdkerror" in low:
        kind = "sdk"
        cta = "retry"
    return {"kind": kind, "retryable": retryable, "cta": cta}


def _is_context_window_failure(error: Any) -> bool:
    """Return True only for an explicit context-window error signal."""
    return bool(
        error is not None
        and _classify_stream_error(str(error)).get("kind") == "context_window"
    )


def _error_event(err: Any, *, activity_source: str = "") -> dict:
    """Bundle a stream-error message with its classification into an SSE
    `error` event payload — single call site so the FE always sees the
    same shape regardless of which yield-error branch fired."""
    msg = str(err) if err is not None else ""
    payload = {"error": msg, **_classify_stream_error(msg)}
    if isinstance(err, ResultError):
        info = sdk_lifecycle.result_error_info(err)
        if info is not None:
            payload["result_error"] = info
            payload["terminal_reason"] = info["terminal_reason"]
            payload["status"] = sdk_lifecycle.terminal_status(
                info["terminal_reason"], is_error=True)
    if isinstance(err, _ContextRecovered):
        # Only public session metadata and aggregate counts cross the SSE
        # boundary.  The synthetic summary, source path and original prompt
        # remain local to the recovery transcript.
        payload["recovered_session"] = err.recovered_session
        payload["recovery_stats"] = err.recovery_stats
    if activity_source:
        payload["activity_source"] = activity_source
    return {"event": "error", "data": json.dumps(payload)}


def _parse_bash_result(text: str) -> dict | None:
    """Patchable facade for structured Bash result parsing."""
    return chat_presentation.parse_bash_result(text)


def _render_tool_result(
    block: ToolResultBlock, *, tool_name: str = "",
) -> dict:
    """Patchable facade for live tool-result rendering."""
    return chat_presentation.render_tool_result(
        block,
        tool_name=tool_name,
        preview_cap=_TOOL_RESULT_PREVIEW_CAP,
        text_cap=_TOOL_RESULT_TEXT_CAP,
        parse_bash=_parse_bash_result,
    )


# ====== attachment upload (images + documents) ======
#
# Multipart upload returns an attachment_id. Stream endpoint reads it (with
# TTL) and attaches as the right SDK block type:
#   - images (png/jpeg/gif/webp) → ImageBlock with base64 data
#   - PDFs → DocumentBlock with base64 data (Claude supports PDFs natively)
#   - text-ish docs (md / txt / csv / json / source code) → inline-text prefix
#     in the prompt so any model can consume them. Stored as utf-8 text.
# SQLite metadata + private blobs are authoritative. `_image_store` is only a
# process-local hot cache retained for preparation and compatibility callers.

_image_store: dict[str, dict] = {}     # id -> {kind, mime, b64|text, name, ts}
# Most staged-upload mutations run on the event loop, but the sync queue/image
# GET endpoints run in FastAPI's worker pool and can trigger TTL collection too.
# A threading lock therefore protects multi-step operations such as
# "validate every id, then remove every id" from a concurrent collector or
# budget eviction.  Never hold it across an await.
_image_store_lock = threading.RLock()
# Exclusive staged-upload ownership. The payload stays in `_image_store` while
# leased; GC and budget eviction skip its id until the lease is committed or
# released. A separate map avoids leaking transaction-only fields through the
# queue/resource APIs that expose entry metadata.
_staged_attachment_claims: dict[str, str] = {}


@dataclass
class _StagedAttachmentLease:
    token: str
    items: tuple[tuple[str, dict], ...]
    state: str = "active"
    durable_ids: tuple[str, ...] = ()
    rehydrated_ids: tuple[str, ...] = ()
    queue_session_id: str = ""
    queue_item_id: str = ""

    @property
    def ids(self) -> tuple[str, ...]:
        return tuple(aid for aid, _entry in self.items)


@dataclass
class _PreparedStagedAttachments:
    img_blocks: list[dict] = dataclass_field(default_factory=list)
    pdf_blocks: list[dict] = dataclass_field(default_factory=list)
    disk_attachments: list[tuple[str, str, str]] = dataclass_field(
        default_factory=list)
    persisted_imgs: list[dict] = dataclass_field(default_factory=list)
    persisted_docs: list[dict] = dataclass_field(default_factory=list)
    artifact_paths: list[str] = dataclass_field(default_factory=list)


class _AttachmentPreparationError(RuntimeError):
    """A required staged attachment could not reach its query-ready state."""


class _AttachmentCommitUncertain(RuntimeError):
    """The SDK accepted query bytes but lease ownership was no longer exact."""


def _attachment_ids(raw_ids: str) -> list[str]:
    return list(dict.fromkeys(
        aid.strip() for aid in str(raw_ids or "").split(",") if aid.strip()
    ))


_STAGED_ATTACHMENT_ID_RE = re.compile(r"[A-Za-z0-9_-]{6,80}")


def _valid_staged_attachment_id(aid: str) -> bool:
    return _STAGED_ATTACHMENT_ID_RE.fullmatch(str(aid or "")) is not None



_IMAGE_TTL_S = 600
_IMAGE_MAX_BYTES = 10 * 1024 * 1024     # 10 MB per file
_UPLOAD_READ_CHUNK_BYTES = 1024 * 1024
_IMAGE_THUMBNAIL_MAX_PIXELS = 16 * 1024 * 1024
# Total in-memory budget for *staged* (not-yet-consumed) uploads + a hard entry
# cap. Without these, N uploads that never get consumed by a turn pin N×~13MB of
# base64 in RAM until their 10-min TTL — an OOM vector. Generous enough never to
# bite a legit multi-image turn (48 files / 256 MB), strict enough to bound
# worst-case growth. Oldest-first eviction — see _enforce_image_budget.
_IMAGE_STORE_MAX_BYTES = 256 * 1024 * 1024
_IMAGE_STORE_MAX_ENTRIES = 48
_ATTACHMENT_LEASE_S = 5 * 60
_durable_attachment_store = DurableAttachmentStore(ROOT)

_IMAGE_MIME = {"image/png", "image/jpeg", "image/gif", "image/webp"}
_IMAGE_OUTPUT_MIME = {"png": "image/png", "jpeg": "image/jpeg", "webp": "image/webp"}
_PDF_MIME = {"application/pdf"}
# text-ish formats we'll inline. Browsers send vague mimes — we also gate by
# extension below as a fallback.
_TEXT_MIME = {
    "text/plain", "text/markdown", "text/csv", "text/html", "text/css",
    "text/xml", "text/javascript", "text/typescript",
    "text/x-python", "text/x-yaml", "text/x-toml", "text/x-shellscript",
    "application/json", "application/xml", "application/yaml",
    "application/x-yaml", "application/toml",
}
_TEXT_EXTS = {
    ".md", ".markdown", ".txt", ".csv", ".json", ".yaml", ".yml", ".toml",
    ".py", ".sh", ".bash", ".zsh", ".js", ".ts", ".tsx", ".jsx",
    ".html", ".htm", ".css", ".scss", ".xml", ".log", ".ini", ".conf", ".cfg",
    ".env.example", ".rs", ".go", ".java", ".c", ".h", ".cpp", ".hpp",
    ".rb", ".php", ".swift", ".kt", ".sql", ".dockerfile", ".gitignore",
}
# Spreadsheets — we pre-process these to CSV-style text via openpyxl so
# the model sees the data inline. Same "ends as `text` kind to the
# frontend" contract — frontend's _classifyFile maps these to "text"
# too so the chip is consistent.
_XLSX_EXTS = {".xlsx", ".xlsm", ".xltx", ".xltm"}
# Was a hard 413 ceiling back when text attachments were pasted into the
# prompt verbatim. Attachments now go to disk and are referenced by path, so
# this is only the threshold above which we log "this one was big" — the real
# limit is _IMAGE_MAX_BYTES, same as every other upload.
_TEXT_MAX_BYTES = 200 * 1024
# Caps for xlsx inlining — same shape as the /api/files/xlsx preview
# endpoint, kept smaller because we're shoving this into the prompt
# context, not just rendering a table.
_XLSX_ATTACH_MAX_SHEETS = 5
_XLSX_ATTACH_MAX_ROWS = 200
_XLSX_ATTACH_MAX_COLS = 30
_XLSX_ATTACH_CELL_MAX_CHARS = 200
_XLSX_ARCHIVE_MAX_ENTRIES = 4096
_XLSX_ARCHIVE_MAX_UNCOMPRESSED_BYTES = 64 * 1024 * 1024
_XLSX_ARCHIVE_MAX_MEMBER_BYTES = 32 * 1024 * 1024
_XLSX_ARCHIVE_MAX_COMPRESSION_RATIO = 200

# Header-less browser resource surfaces use narrowly scoped, short-lived
# capability tickets.  Export is a one-shot download; image resources permit a
# small, finite number of reads because browsers may fetch the same URL for the
# in-bubble image, the lightbox, and a conditional retry.
_CHAT_RESOURCE_TICKET_KIND = "chat-resource"
_CHAT_RESOURCE_TICKET_TTL = {
    "export": 60,
    "queued-image": min(300, _IMAGE_TTL_S),
    "attachment": 600,
}
_CHAT_RESOURCE_TICKET_MAX_USES = {
    "export": 1,
    "queued-image": 8,
    "attachment": 16,
}


class ChatResourceTicketReq(BaseModel):
    resource: Literal["export", "queued-image", "attachment"]
    session_id: str = Field(default="", max_length=80)
    attachment_id: str = Field(default="", max_length=64)
    filename: str = Field(default="", max_length=200)


def _require_chat_resource_ticket(
    ticket: str,
    scope: tuple[str, ...],
) -> None:
    if not tickets.validate(ticket, _CHAT_RESOURCE_TICKET_KIND, scope):
        raise HTTPException(
            status_code=401,
            detail="invalid or expired chat resource ticket",
        )


def _validate_attachment_ref(session_id: str, filename: str) -> Path:
    """Return the exact persisted attachment selected by a resource ticket."""
    if not re.fullmatch(r"[A-Za-z0-9_\-]{6,80}", session_id):
        raise HTTPException(400, "bad session_id")
    if ("/" in filename or ".." in filename or "\\" in filename
            or not re.fullmatch(r"[^/\\\x00-\x1f\x7f]{1,200}", filename)):
        raise HTTPException(400, "bad filename")
    try:
        session_dir = _attachment_session_dir(session_id, create=False)
        if session_dir is None:
            raise HTTPException(404, "attachment not found")
        path = session_dir / filename
        if not ensure_private_regular_file(path):
            raise HTTPException(404, "attachment not found")
    except UnsafePrivatePath:
        raise HTTPException(400, "unsafe attachment storage") from None
    return path


@router.post("/resource-ticket", dependencies=[Depends(require_token)])
def mint_chat_resource_ticket(req: ChatResourceTicketReq) -> dict:
    """Mint an opaque ticket for one exact export or attachment resource."""
    resource = req.resource
    if resource == "export":
        if not req.session_id or sess.get_session_meta(req.session_id) is None:
            raise HTTPException(404, "session not found")
        scope = (resource, req.session_id)
        url = f"/api/chat/sessions/{req.session_id}/export"
    elif resource == "queued-image":
        aid = req.attachment_id
        if not re.fullmatch(r"[A-Za-z0-9]{6,64}", aid):
            raise HTTPException(400, "bad id")
        _gc_images()
        with _image_store_lock:
            entry = _get_staged_entry_locked(aid)
            if (entry is None or entry.get("kind") != "image"
                    or not entry.get("b64")):
                raise HTTPException(404, "queued image not found or expired")
        scope = (resource, aid)
        url = f"/api/chat/queued-image/{aid}"
    else:
        _validate_attachment_ref(req.session_id, req.filename)
        scope = (resource, req.session_id, req.filename)
        url = (f"/api/chat/attachments/{req.session_id}/"
               f"{urllib.parse.quote(req.filename)}")

    ttl = _CHAT_RESOURCE_TICKET_TTL[resource]
    max_uses = _CHAT_RESOURCE_TICKET_MAX_USES[resource]
    ticket = tickets.mint(
        _CHAT_RESOURCE_TICKET_KIND,
        scope,
        ttl=ttl,
        max_uses=max_uses,
    )
    separator = "&" if "?" in url else "?"
    return {
        "ticket": ticket,
        "url": f"{url}{separator}ticket={urllib.parse.quote(ticket)}",
        "expires_in": ttl,
        "max_uses": max_uses,
    }


def _get_staged_entry_locked(aid: str) -> dict | None:
    """Resolve one staged object from the hot cache or durable blob store."""
    entry = _image_store.get(aid)
    if entry is not None:
        return entry
    return _durable_attachment_store.load_entry(aid)


def _resolve_staged_attachment_display(
    attachment_ids: list[str],
) -> tuple[list[dict], list[dict]]:
    """Recover bounded display metadata without exposing staged payloads."""
    images: list[dict] = []
    docs: list[dict] = []
    for aid in attachment_ids[:48]:
        if not _valid_staged_attachment_id(aid):
            docs.append({
                "name": "Attachment unavailable",
                "kind": "unknown",
                "available": False,
            })
            continue
        with _image_store_lock:
            hot_entry = _image_store.get(aid)
            metadata = dict(hot_entry) if isinstance(hot_entry, dict) else None
        if metadata is None:
            try:
                metadata = _durable_attachment_store.metadata(aid)
            except (DurableAttachmentError, OSError, sqlite3.Error,
                    UnsafePrivatePath):
                metadata = None
        if not isinstance(metadata, dict):
            docs.append({
                "name": "Attachment unavailable",
                "kind": "unknown",
                "available": False,
            })
            continue
        kind = str(metadata.get("kind") or "image")[:16]
        if kind == "image":
            images.append({
                "mime": str(metadata.get("mime") or "image/*")[:80],
                "available": True,
            })
            continue
        fallback = "Document"
        if kind == "pdf":
            fallback = "Document.pdf"
        elif kind == "xlsx":
            fallback = "Workbook.xlsx"
        docs.append({
            "name": _safe_attach_name(
                str(metadata.get("name") or fallback))[:200],
            "kind": kind,
            "available": True,
        })
    return images, docs


def _hydrate_staged_attachment_display(broadcast: TurnBroadcast) -> None:
    """Fill failure/reconnect display refs before staged ownership settles."""
    if (not broadcast.staged_attachment_ids
            or broadcast.user_images or broadcast.user_docs):
        return
    images, docs = _resolve_staged_attachment_display(
        broadcast.staged_attachment_ids)
    broadcast.user_images = images
    broadcast.user_docs = docs


def _gc_images_locked(now: float | None = None) -> None:
    """Collect expired, unleased entries from cache and durable storage."""
    current = time.time() if now is None else now
    cutoff = current - _IMAGE_TTL_S
    for aid in list(_image_store.keys()):
        entry = _image_store.get(aid)
        if (entry is not None and entry.get("ts", 0) < cutoff
                and aid not in _staged_attachment_claims):
            _image_store.pop(aid, None)
    removed = _durable_attachment_store.gc(
        now=current,
        protected_tokens=set(_staged_attachment_claims.values()),
    )
    for aid in removed:
        if aid not in _staged_attachment_claims:
            _image_store.pop(aid, None)


def _lease_staged_attachments(
    image_ids: str,
    *,
    require_all: bool,
    queue_owner: tuple[str, str] | None = None,
) -> tuple[_StagedAttachmentLease | None, list[str], list[str]]:
    """Lease staged objects from memory or the restart-safe registry.

    Queue-owned leases additionally require the exact durable item reference;
    this prevents an old or forged sidecar from consuming another queue item's
    blob. Legacy memory-only entries remain supported for direct sends/tests.
    """
    ids = _attachment_ids(image_ids)
    if not ids:
        return None, [], []
    with _image_store_lock:
        _gc_images_locked()
        busy = [aid for aid in ids if aid in _staged_attachment_claims]
        if busy:
            return None, [], busy
        durable_candidates = [
            aid for aid in ids if _valid_staged_attachment_id(aid)
        ]
        durable_ids = _durable_attachment_store.registered_ids(
            durable_candidates)
        missing = [
            aid for aid in ids
            if aid not in _image_store and aid not in durable_ids
        ]
        if require_all and missing:
            return None, missing, []

        token = uuid.uuid4().hex
        load_ids = [
            aid for aid in ids
            if aid in durable_ids and aid not in _image_store
        ]
        durable = _durable_attachment_store.acquire(
            [aid for aid in ids if aid in durable_ids],
            token,
            lease_seconds=_ATTACHMENT_LEASE_S,
            queue_owner=queue_owner,
            load_ids=load_ids,
        )
        if durable.missing or durable.busy:
            return None, list(durable.missing), list(durable.busy)

        items: list[tuple[str, dict]] = []
        for aid in ids:
            entry = _image_store.get(aid) or durable.entries.get(aid)
            if entry is None:
                if require_all:
                    _durable_attachment_store.release(
                        token, ttl=_IMAGE_TTL_S)
                    return None, [aid], []
                continue
            _image_store.setdefault(aid, entry)
            items.append((aid, _image_store[aid]))
        if not items:
            return None, missing, []
        for aid, _entry in items:
            _staged_attachment_claims[aid] = token
        return (
            _StagedAttachmentLease(
                token=token,
                items=tuple(items),
                durable_ids=tuple(
                    aid for aid, _entry in items if aid in durable_ids
                ),
                rehydrated_ids=tuple(load_ids),
                queue_session_id=(queue_owner[0] if queue_owner else ""),
                queue_item_id=(queue_owner[1] if queue_owner else ""),
            ),
            missing,
            [],
        )


def _commit_staged_attachment_lease(
    lease: _StagedAttachmentLease | None,
) -> bool:
    """Consume direct blobs or mark queue blobs submitted after query()."""
    if lease is None:
        return True
    with _image_store_lock:
        if lease.state == "committed":
            return True
        if lease.state != "active":
            return False
        if any(
            _staged_attachment_claims.get(aid) != lease.token
            or _image_store.get(aid) is not entry
            for aid, entry in lease.items
        ):
            return False
        if lease.durable_ids and not _durable_attachment_store.commit(
            lease.token,
            queue_session_id=lease.queue_session_id,
            queue_item_id=lease.queue_item_id,
        ):
            return False
        for aid, _entry in lease.items:
            _image_store.pop(aid, None)
            _staged_attachment_claims.pop(aid, None)
        lease.state = "committed"
        return True


def _begin_staged_attachment_rollback(
    lease: _StagedAttachmentLease | None,
) -> bool:
    """Atomically exclude commit while keeping every claim pinned."""
    if lease is None:
        return False
    with _image_store_lock:
        if lease.state == "active":
            lease.state = "rolling_back"
            return True
        return lease.state in {"rolling_back", "uncertain"}


def _fail_closed_staged_attachment_lease(
    lease: _StagedAttachmentLease | None,
) -> None:
    """Consume owned objects after transport acceptance becomes uncertain."""
    if lease is None:
        return
    with _image_store_lock:
        if lease.durable_ids:
            try:
                _durable_attachment_store.commit(
                    lease.token,
                    queue_session_id=lease.queue_session_id,
                    queue_item_id=lease.queue_item_id,
                )
            except Exception:
                pass
        for aid, entry in lease.items:
            if (_staged_attachment_claims.get(aid) == lease.token
                    and _image_store.get(aid) is entry):
                _image_store.pop(aid, None)
                _staged_attachment_claims.pop(aid, None)
        lease.state = "uncertain"


def _release_staged_attachment_lease(
    lease: _StagedAttachmentLease | None,
) -> bool:
    """Roll back a preparation and refresh the retry TTL durably."""
    if lease is None:
        return True
    with _image_store_lock:
        if lease.state == "released":
            return True
        if lease.state in {"committed", "uncertain"}:
            return False
        if lease.state == "active":
            lease.state = "rolling_back"
        if lease.state != "rolling_back":
            return False
        if lease.durable_ids and not _durable_attachment_store.release(
            lease.token, ttl=_IMAGE_TTL_S,
        ):
            return False
        retry_ts = time.time()
        for aid, entry in lease.items:
            if _staged_attachment_claims.get(aid) == lease.token:
                _staged_attachment_claims.pop(aid, None)
                if _image_store.get(aid) is entry:
                    if aid in lease.rehydrated_ids:
                        _image_store.pop(aid, None)
                    else:
                        entry["ts"] = retry_ts
        lease.state = "released"
        return True


def _gc_images() -> None:
    """Drop expired entries without invalidating an active turn lease."""
    with _image_store_lock:
        _gc_images_locked()


def _image_entry_bytes(entry: dict) -> int:
    """Approximate retained bytes of one process-local staged entry."""
    return (len(entry.get("b64", ""))
            + len(entry.get("raw", b""))
            + len(entry.get("text", ""))
            + len(entry.get("name", "")))


def _enforce_image_budget() -> None:
    """Evict oldest unclaimed cache and durable entries to the same budget."""
    with _image_store_lock:
        total = sum(_image_entry_bytes(e) for e in _image_store.values())
        if (len(_image_store) <= _IMAGE_STORE_MAX_ENTRIES
                and total <= _IMAGE_STORE_MAX_BYTES):
            return
        evict: list[str] = []
        for aid, entry in sorted(
            _image_store.items(), key=lambda item: item[1].get("ts", 0.0),
        ):
            if aid in _staged_attachment_claims:
                continue
            if (len(_image_store) - len(evict) <= _IMAGE_STORE_MAX_ENTRIES
                    and total <= _IMAGE_STORE_MAX_BYTES):
                break
            total -= _image_entry_bytes(entry)
            evict.append(aid)
        _durable_attachment_store.discard_unowned(evict)
        for aid in evict:
            _image_store.pop(aid, None)


def _put_staged_attachment_batch(
    batch: list[tuple[str, dict]],
) -> bool:
    """Publish a generated/upload batch to blobs, SQLite, then the hot cache."""
    if not batch:
        return True
    with _image_store_lock:
        _gc_images_locked()
        ids = [aid for aid, _entry in batch]
        if (len(set(ids)) != len(ids)
                or any(aid in _image_store for aid in ids)
                or any(aid in _staged_attachment_claims for aid in ids)):
            return False

        batch_bytes = sum(_image_entry_bytes(entry) for _aid, entry in batch)
        total_bytes = sum(
            _image_entry_bytes(entry) for entry in _image_store.values()
        )
        projected_count = len(_image_store) + len(batch)
        projected_bytes = total_bytes + batch_bytes
        cache_evict: list[str] = []
        for aid, entry in sorted(
            _image_store.items(), key=lambda item: item[1].get("ts", 0.0),
        ):
            if (projected_count <= _IMAGE_STORE_MAX_ENTRIES
                    and projected_bytes <= _IMAGE_STORE_MAX_BYTES):
                break
            if aid in _staged_attachment_claims:
                continue
            cache_evict.append(aid)
            projected_count -= 1
            projected_bytes -= _image_entry_bytes(entry)
        if (projected_count > _IMAGE_STORE_MAX_ENTRIES
                or projected_bytes > _IMAGE_STORE_MAX_BYTES):
            return False

        try:
            published, durable_evict = _durable_attachment_store.stage_batch(
                batch,
                ttl=_IMAGE_TTL_S,
                max_entries=_IMAGE_STORE_MAX_ENTRIES,
                max_bytes=_IMAGE_STORE_MAX_BYTES,
            )
        except (OSError, sqlite3.Error, DurableAttachmentError,
                UnsafePrivatePath):
            return False
        if not published:
            return False
        for aid in {*cache_evict, *durable_evict}:
            _image_store.pop(aid, None)
        for aid, entry in batch:
            _image_store[aid] = entry
        return True


def _put_staged_attachment(aid: str, entry: dict) -> bool:
    """Publish one staged object through the all-or-none batch primitive."""
    return _put_staged_attachment_batch([(aid, entry)])


def _classify_attachment(mime: str, name: str) -> str:
    """Return one of: 'image' / 'pdf' / 'text' / 'xlsx' / '' (unsupported)."""
    mime = (mime or "").lower()
    if mime in _IMAGE_MIME:
        return "image"
    if mime in _PDF_MIME:
        return "pdf"
    if mime in _TEXT_MIME:
        return "text"
    # Fall back to extension check (browsers often send empty / octet-stream).
    lower = name.lower()
    for ext in _TEXT_EXTS:
        if lower.endswith(ext):
            return "text"
    if lower.endswith(".pdf"):
        return "pdf"
    for ext in _XLSX_EXTS:
        if lower.endswith(ext):
            return "xlsx"
    return ""


class ImageGenerateReq(BaseModel):
    prompt: str = Field(..., min_length=1, max_length=4000)
    model: str = Field(default="gpt-image-2", max_length=80)
    size: str = Field(default="1024x1024", max_length=32)
    quality: str = Field(default="low", max_length=16)
    output_format: str = Field(default="png", max_length=8)
    n: int = Field(default=1, ge=1, le=4)
    image_ids: list[str] | None = None


_IMAGE_SIZE_RE = re.compile(r"^(auto|[1-9][0-9]{2,3}x[1-9][0-9]{2,3})$")
_IMAGE_MODEL_RE = re.compile(r"^[A-Za-z0-9][A-Za-z0-9._:-]{0,79}$")
_IMAGE_FILE_MIME = {
    ".png": "image/png",
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".webp": "image/webp",
}
_IMAGEGEN_ROOT = ROOT / ".muselab" / "imagegen"
_IMAGEGEN_FILES = _IMAGEGEN_ROOT / "files"
_IMAGEGEN_JOBS_PATH = _IMAGEGEN_ROOT / "jobs.json"
_IMAGEGEN_JOBS_MAX = 200
_imagegen_job_store = ImagegenJobStore(_IMAGEGEN_JOBS_PATH, max_jobs=_IMAGEGEN_JOBS_MAX)


def _validate_image_size(size: str) -> str:
    s = (size or "1024x1024").strip()
    if not _IMAGE_SIZE_RE.fullmatch(s):
        raise HTTPException(400, "invalid image size")
    if s == "auto":
        return s
    w, h = [int(x) for x in s.split("x", 1)]
    if w > 3840 or h > 3840:
        raise HTTPException(400, "image size edge must be <= 3840")
    if w % 16 or h % 16:
        raise HTTPException(400, "image size edges must be multiples of 16")
    if max(w, h) / min(w, h) > 3:
        raise HTTPException(400, "image aspect ratio must be <= 3:1")
    pixels = w * h
    if pixels < 655_360 or pixels > 8_294_400:
        raise HTTPException(400, "image size pixels out of range")
    return s


def _openai_image_api_config() -> tuple[str, str]:
    key = (
        os.environ.get("OPENAI_IMAGE_API_KEY", "").strip()
        or os.environ.get("OPENAI_API_KEY", "").strip()
    )
    if not key:
        raise HTTPException(
            400,
            "missing OPENAI_IMAGE_API_KEY or OPENAI_API_KEY for image generation",
        )
    base_url = (
        os.environ.get("OPENAI_IMAGE_BASE_URL", "").strip()
        or os.environ.get("OPENAI_BASE_URL", "").strip()
        or "https://api.openai.com/v1"
    ).rstrip("/")
    parsed = urllib.parse.urlsplit(base_url)
    host = (parsed.hostname or "").lower()
    if parsed.scheme == "https" and host:
        return key, base_url
    loopback_hosts = {"localhost", "127.0.0.1", "::1"}
    if parsed.scheme == "http" and host in loopback_hosts:
        return key, base_url
    raise HTTPException(400, "OPENAI_IMAGE_BASE_URL must be https or loopback http")


def _image_error_message(status: int, body: str) -> str:
    try:
        data = json.loads(body)
        err = data.get("error") if isinstance(data, dict) else None
        msg = err.get("message") if isinstance(err, dict) else None
        if isinstance(msg, str) and msg:
            return f"image generation failed ({status}): {msg[:500]}"
    except Exception:
        pass
    return f"image generation failed ({status})"


def _image_response_items(data: dict) -> list[str]:
    out = data.get("data")
    if not isinstance(out, list):
        return []
    b64s: list[str] = []
    for item in out:
        if isinstance(item, dict) and isinstance(item.get("b64_json"), str):
            b64s.append(item["b64_json"])
    return b64s


def _normalize_image_generate_req(
    req: ImageGenerateReq,
) -> tuple[str, str, str, str, str, list[str]]:
    prompt = req.prompt.strip()
    if not prompt:
        raise HTTPException(400, "prompt is required")
    model = req.model.strip() or "gpt-image-2"
    if not _IMAGE_MODEL_RE.fullmatch(model):
        raise HTTPException(400, "invalid image model")
    size = _validate_image_size(req.size)
    quality = (req.quality or "low").strip()
    if quality not in {"low", "medium", "high", "auto"}:
        raise HTTPException(400, "invalid image quality")
    output_format = (req.output_format or "png").strip().lower()
    if output_format not in _IMAGE_OUTPUT_MIME:
        raise HTTPException(400, "invalid image output format")
    image_ids = [x.strip() for x in (req.image_ids or []) if isinstance(x, str) and x.strip()]
    return prompt, model, size, quality, output_format, image_ids


def _stage_generated_images(
    b64s: list[str],
    mime: str,
    start_index: int = 1,
) -> list[dict]:
    """Validate and publish a generated batch in one budget transaction."""
    batch: list[tuple[str, dict]] = []
    items: list[dict] = []
    timestamp = time.time()
    stamp = datetime.now(UTC).strftime("%Y%m%d-%H%M%S")
    fmt = {v: k for k, v in _IMAGE_OUTPUT_MIME.items()}.get(mime, "png")
    for offset, b64 in enumerate(b64s):
        try:
            raw = base64.b64decode(b64, validate=True)
        except Exception:
            raise HTTPException(
                502, "image API returned invalid base64") from None
        if len(raw) > _IMAGE_MAX_BYTES:
            raise HTTPException(
                502,
                "image API returned an image over the local 10MB limit",
            )
        aid = uuid.uuid4().hex
        name = f"generated-{stamp}-{start_index + offset}.{fmt}"
        encoded = base64.b64encode(raw).decode("ascii")
        entry = {
            "kind": "image",
            "mime": mime,
            "name": name,
            "b64": encoded,
            "ts": timestamp,
        }
        batch.append((aid, entry))
        items.append({
            "id": aid,
            "mime": mime,
            "name": name,
            "bytes": len(raw),
            "attach_ext": "jpg" if fmt == "jpeg" else fmt,
            "data_url": f"data:{mime};base64,{encoded}",
        })
    if not _put_staged_attachment_batch(batch):
        raise HTTPException(
            503, "staged attachment capacity is temporarily exhausted")
    return items


def _stage_generated_image(b64: str, mime: str, idx: int) -> dict:
    return _stage_generated_images([b64], mime, idx)[0]


def _stage_generated_image_bytes(raw: bytes, mime: str, idx: int) -> dict:
    if len(raw) > _IMAGE_MAX_BYTES:
        raise HTTPException(
            502,
            "image generation returned an image over the local 10MB limit",
        )
    encoded = base64.b64encode(raw).decode("ascii")
    return _stage_generated_images([encoded], mime, idx)[0]


def _discard_generated_image_batch(items: list[dict]) -> None:
    """Reclaim an HTTP-owned generated batch if its request is cancelled."""
    with _image_store_lock:
        discarded: list[str] = []
        for item in items:
            aid = str(item.get("id") or "")
            if aid and aid not in _staged_attachment_claims:
                _image_store.pop(aid, None)
                discarded.append(aid)
        _durable_attachment_store.discard_unowned(discarded)


async def _discard_generated_image_batch_owned(items: list[dict]) -> None:
    cleanup = asyncio.create_task(asyncio.to_thread(
        _discard_generated_image_batch,
        items,
    ))
    while not cleanup.done():
        try:
            await asyncio.shield(cleanup)
        except asyncio.CancelledError:
            continue
    cleanup.result()


async def _stage_generated_images_owned(
    b64s: list[str],
    mime: str,
    start_index: int = 1,
) -> list[dict]:
    """Join the worker and reclaim its published batch on owner cancellation."""
    task = asyncio.create_task(asyncio.to_thread(
        _stage_generated_images,
        b64s,
        mime,
        start_index,
    ))
    try:
        return await asyncio.shield(task)
    except asyncio.CancelledError as cancelled:
        while not task.done():
            try:
                await asyncio.shield(task)
            except asyncio.CancelledError:
                continue
            except BaseException:
                break
        items: list[dict] | None
        try:
            items = task.result()
        except BaseException:
            items = None
        if items:
            await _discard_generated_image_batch_owned(items)
        raise cancelled


async def _stage_generated_images_for_response(
    b64s: list[str],
    mime: str,
) -> list[dict]:
    """Stage a response-owned batch with a final cancellation checkpoint."""
    items: list[dict] = []
    try:
        items = await _stage_generated_images_owned(b64s, mime)
        # Deliver a cancellation already queued by the HTTP owner before the
        # staged IDs escape into a response it can no longer receive.
        await asyncio.sleep(0)
        return items
    except asyncio.CancelledError:
        if items:
            await _discard_generated_image_batch_owned(items)
        raise


def _imagegen_load_jobs() -> dict[str, dict]:
    """Compatibility snapshot for tests and local diagnostic tooling."""
    return _imagegen_job_store.snapshot()


def _imagegen_put_job(job: dict) -> dict:
    return _imagegen_job_store.put(job)


def _imagegen_job_file(job: dict, img: dict) -> Path:
    rel = str(img.get("file") or "")
    if not rel:
        raise HTTPException(404, "image file missing")
    base = (_IMAGEGEN_FILES / str(job.get("id") or "")).resolve()
    p = (base / rel).resolve()
    try:
        p.relative_to(base)
    except ValueError:
        raise HTTPException(400, "invalid image file path") from None
    return p


def _imagegen_public_image(job: dict, img: dict, *, include_data: bool) -> dict:
    job_id = str(job.get("id") or "")
    image_id = str(img.get("image_id") or "")
    out = {
        "job_id": job_id,
        "image_id": image_id,
        "name": img.get("name"),
        "mime": img.get("mime"),
        "bytes": img.get("bytes"),
        "attach_ext": img.get("attach_ext"),
        "url": (
            f"/api/chat/image-generate/jobs/{urllib.parse.quote(job_id, safe='')}"
            f"/images/{urllib.parse.quote(image_id, safe='')}"
        ) if job_id and image_id else "",
    }
    if include_data:
        try:
            raw = _imagegen_job_file(job, img).read_bytes()
            out["data_url"] = (
                f"data:{img.get('mime') or 'image/png'};"
                f"base64,{base64.b64encode(raw).decode('ascii')}"
            )
        except OSError:
            out["missing"] = True
    return out


def _imagegen_public_job(job: dict, *, include_data: bool = True) -> dict:
    images = job.get("images") if isinstance(job.get("images"), list) else []
    return {
        "id": job.get("id"),
        "status": job.get("status"),
        "prompt": job.get("prompt"),
        "model": job.get("model"),
        "provider": job.get("provider"),
        "size": job.get("size"),
        "quality": job.get("quality"),
        "output_format": job.get("output_format"),
        "n": job.get("n"),
        "error": job.get("error") or "",
        "created_at": job.get("created_at"),
        "updated_at": job.get("updated_at"),
        "images": [_imagegen_public_image(job, img, include_data=include_data)
                   for img in images if isinstance(img, dict)],
    }


def _imagegen_list_jobs(limit: int) -> list[dict]:
    return [
        _imagegen_public_job(job, include_data=False)
        for job in _imagegen_job_store.list(limit)
    ]


def _persist_imagegen_result(job: dict, result: dict) -> list[dict]:
    images = result.get("images") if isinstance(result, dict) else None
    if not isinstance(images, list):
        return []
    job_dir = _IMAGEGEN_FILES / job["id"]
    job_dir.mkdir(parents=True, exist_ok=True)
    persisted: list[dict] = []
    for idx, img in enumerate(images, start=1):
        if not isinstance(img, dict):
            continue
        data_url = str(img.get("data_url") or "")
        marker = ";base64,"
        if marker not in data_url:
            continue
        mime = str(img.get("mime") or data_url[5:data_url.find(";")] or "image/png")
        try:
            raw = base64.b64decode(data_url.split(marker, 1)[1], validate=True)
        except Exception:
            continue
        ext = str(img.get("attach_ext") or "png").lower()
        if ext == "jpg":
            ext = "jpeg"
        if ext not in {"png", "jpeg", "webp"}:
            ext = "png"
        filename = f"image-{idx}.{ext}"
        (job_dir / filename).write_bytes(raw)
        persisted.append({
            "image_id": uuid.uuid4().hex,
            "file": filename,
            "name": img.get("name") or filename,
            "mime": mime,
            "bytes": len(raw),
            "attach_ext": "jpg" if ext == "jpeg" else ext,
        })
    return persisted


async def _run_imagegen_job(job_id: str, req: ImageGenerateReq) -> None:
    staged_result_items: list[dict] = []
    try:
        await _imagegen_job_store.update_async(job_id, status="running", error="")
        prompt, model, size, quality, output_format, image_ids = _normalize_image_generate_req(req)
        result = await _generate_openai_image_api(
            req=req,
            prompt=prompt,
            model=model,
            size=size,
            quality=quality,
            output_format=output_format,
            image_ids=image_ids,
        )
        staged_result_items = [
            item for item in result.get("images", [])
            if isinstance(item, dict)
        ]
        job_snapshot = await asyncio.to_thread(
            _imagegen_job_store.get, job_id)
        if not job_snapshot:
            return
        images = await asyncio.to_thread(
            _persist_imagegen_result,
            job_snapshot,
            result,
        )
        await _imagegen_job_store.update_async(
            job_id,
            provider=result.get("provider"),
            model=result.get("model") or model,
            images=images,
            status="succeeded" if images else "failed",
            error="" if images else "image generation returned no images",
        )
    except asyncio.CancelledError:
        await _imagegen_job_store.update_async(
            job_id,
            status="failed",
            error="image generation was cancelled",
        )
        raise
    except HTTPException as e:
        await _imagegen_job_store.update_async(
            job_id, status="failed", error=str(e.detail))
    except Exception as e:
        await _imagegen_job_store.update_async(
            job_id, status="failed", error=f"{type(e).__name__}: {e}")

    finally:
        if staged_result_items:
            await _discard_generated_image_batch_owned(staged_result_items)


def _image_reference_files(
    image_ids: list[str],
) -> list[tuple[str, tuple[str, bytes, str]]]:
    """Snapshot staged references under lock and decode them in a worker."""
    snapshots: list[tuple[str, str, str, str]] = []
    with _image_store_lock:
        _gc_images_locked()
        for aid in image_ids[:8]:
            entry = _get_staged_entry_locked(aid)
            if (not entry or entry.get("kind") != "image"
                    or not entry.get("b64")):
                continue
            snapshots.append((
                aid,
                str(entry.get("name") or f"{aid}.png"),
                str(entry.get("mime") or "image/png"),
                str(entry["b64"]),
            ))
    files: list[tuple[str, tuple[str, bytes, str]]] = []
    for _aid, name, mime, encoded in snapshots:
        try:
            raw = base64.b64decode(encoded, validate=True)
        except Exception:
            continue
        files.append(("image[]", (name, raw, mime)))
    return files


async def _generate_openai_image_api(
    *,
    req: ImageGenerateReq,
    prompt: str,
    model: str,
    size: str,
    quality: str,
    output_format: str,
    image_ids: list[str],
) -> dict:
    key, base_url = _openai_image_api_config()
    headers = {"Authorization": f"Bearer {key}"}
    timeout = max(10.0, env_float("MUSELAB_IMAGE_GENERATION_TIMEOUT", 180.0))

    import httpx
    files = await asyncio.to_thread(_image_reference_files, image_ids)
    async with httpx.AsyncClient(timeout=timeout) as client:
        if image_ids:
            data = {
                "model": model,
                "prompt": prompt,
                "size": size,
                "quality": quality,
                "output_format": output_format,
                "n": str(req.n),
            }
            if not files:
                raise HTTPException(400, "reference images are missing or expired")
            resp = await client.post(
                f"{base_url}/images/edits",
                headers=headers,
                data=data,
                files=files,
            )
        else:
            payload = {
                "model": model,
                "prompt": prompt,
                "size": size,
                "quality": quality,
                "output_format": output_format,
                "n": req.n,
            }
            resp = await client.post(
                f"{base_url}/images/generations",
                headers={**headers, "Content-Type": "application/json"},
                json=payload,
            )
    if resp.status_code >= 400:
        raise HTTPException(resp.status_code, _image_error_message(resp.status_code, resp.text))
    try:
        body = resp.json()
    except ValueError:
        raise HTTPException(502, "image API returned non-JSON response") from None
    b64s = _image_response_items(body)
    if not b64s:
        raise HTTPException(502, "image API returned no base64 image")
    mime = _IMAGE_OUTPUT_MIME[output_format]
    items = await _stage_generated_images_for_response(b64s, mime)
    return {
        "ok": True,
        "provider": "openai",
        "model": model,
        "images": items,
        "usage": body.get("usage") if isinstance(body, dict) else None,
    }


@router.post("/image-generate", dependencies=[Depends(require_token)])
async def generate_image(req: ImageGenerateReq) -> dict:
    """Generate images and stage them as ordinary muselab image attachments."""
    prompt, model, size, quality, output_format, image_ids = _normalize_image_generate_req(req)
    return await _generate_openai_image_api(
        req=req,
        prompt=prompt,
        model=model,
        size=size,
        quality=quality,
        output_format=output_format,
        image_ids=image_ids,
    )


@router.post("/image-generate/jobs", dependencies=[Depends(require_token)])
async def create_image_generate_job(req: ImageGenerateReq) -> dict:
    prompt, model, size, quality, output_format, _image_ids = _normalize_image_generate_req(req)

    now = time.time()
    job = {
        "id": uuid.uuid4().hex,
        "status": "queued",
        "prompt": prompt,
        "model": model,
        "provider": None,
        "size": size,
        "quality": quality,
        "output_format": output_format,
        "n": req.n,
        "error": "",
        "images": [],
        "created_at": now,
        "updated_at": now,
    }
    await _imagegen_job_store.put_async(job)
    task = asyncio.create_task(_run_imagegen_job(job["id"], req))
    task.add_done_callback(lambda t: t.exception() if not t.cancelled() else None)
    return {"ok": True, "job": _imagegen_public_job(job, include_data=True)}


@router.get("/image-generate/jobs", dependencies=[Depends(require_token)])
async def list_image_generate_jobs(limit: int = Query(40, ge=1, le=100)) -> dict:
    jobs = await asyncio.to_thread(_imagegen_list_jobs, limit)
    return {"ok": True, "jobs": jobs}


@router.get("/image-generate/jobs/{job_id}", dependencies=[Depends(require_token)])
async def get_image_generate_job(job_id: str) -> dict:
    job = await asyncio.to_thread(_imagegen_job_store.get, job_id)
    if not job:
        raise HTTPException(404, "image generation job not found")
    public = await asyncio.to_thread(
        _imagegen_public_job, job, include_data=True)
    return {"ok": True, "job": public}


@router.get("/image-generate/jobs/{job_id}/images/{image_id}",
            dependencies=[Depends(require_token)])
async def get_image_generate_job_image(job_id: str, image_id: str) -> FileResponse:
    job = await asyncio.to_thread(_imagegen_job_store.get, job_id)
    if not job:
        raise HTTPException(404, "image generation job not found")
    images = job.get("images") if isinstance(job.get("images"), list) else []
    img = next((x for x in images
                if isinstance(x, dict) and x.get("image_id") == image_id), None)
    if not img:
        raise HTTPException(404, "image generation image not found")
    path = _imagegen_job_file(job, img)
    if not path.exists():
        raise HTTPException(404, "image file missing")
    return FileResponse(
        path,
        media_type=img.get("mime") or "image/png",
        filename=img.get("name") or path.name,
    )


@router.post("/image-generate/jobs/{job_id}/attach/{image_id}",
             dependencies=[Depends(require_token)])
async def attach_image_generate_job_image(job_id: str, image_id: str) -> dict:
    job = await asyncio.to_thread(_imagegen_job_store.get, job_id)
    if not job:
        raise HTTPException(404, "image generation job not found")
    images = job.get("images") if isinstance(job.get("images"), list) else []
    img = next((x for x in images
                if isinstance(x, dict) and x.get("image_id") == image_id), None)
    if not img:
        raise HTTPException(404, "image generation image not found")
    try:
        raw = await _await_thread_completion(
            _imagegen_job_file(job, img).read_bytes)
    except OSError:
        raise HTTPException(404, "image file missing") from None
    encoded = await asyncio.to_thread(base64.b64encode, raw)
    item = (await _stage_generated_images_for_response(
        [encoded.decode("ascii")],
        img.get("mime") or "image/png",
    ))[0]
    if img.get("name"):
        with _image_store_lock:
            entry = _image_store.get(item["id"])
            if entry is not None:
                entry["name"] = img["name"]
        item["name"] = img["name"]
    return {"ok": True, "image": item}


def _validate_xlsx_archive(body: bytes) -> None:
    """Reject encrypted, oversized, or suspiciously compressed workbooks."""
    import io
    import zipfile

    try:
        with zipfile.ZipFile(io.BytesIO(body)) as archive:
            infos = archive.infolist()
    except (zipfile.BadZipFile, zipfile.LargeZipFile):
        raise HTTPException(
            422,
            "failed to parse spreadsheet (file may be corrupt or unsupported)",
        ) from None
    if len(infos) > _XLSX_ARCHIVE_MAX_ENTRIES:
        raise HTTPException(422, "spreadsheet archive exceeds safe entry budget")
    total_uncompressed = 0
    for info in infos:
        if info.flag_bits & 0x1:
            raise HTTPException(
                422, "encrypted spreadsheets are not supported")
        if info.file_size > _XLSX_ARCHIVE_MAX_MEMBER_BYTES:
            raise HTTPException(
                422, "spreadsheet archive member exceeds safe size budget")
        total_uncompressed += info.file_size
        if total_uncompressed > _XLSX_ARCHIVE_MAX_UNCOMPRESSED_BYTES:
            raise HTTPException(
                422, "spreadsheet archive exceeds safe unpacked size budget")
        if (
            info.file_size > 0
            and info.file_size
            > max(1, info.compress_size) * _XLSX_ARCHIVE_MAX_COMPRESSION_RATIO
        ):
            raise HTTPException(
                422, "spreadsheet archive exceeds safe compression ratio")


def _xlsx_to_text(body: bytes, name: str) -> str:
    """Read xlsx bytes and dump each sheet as `[Sheet: name]\\n<csv>` blocks.
    Capped by _XLSX_ATTACH_MAX_* so a 100k-row spreadsheet doesn't blow
    the prompt. Truncation is signaled inline so the model knows."""
    import openpyxl
    from io import BytesIO

    _validate_xlsx_archive(body)
    try:
        wb = openpyxl.load_workbook(BytesIO(body), read_only=True, data_only=True)
    except Exception as e:
        # Do not echo the filename, library message, or cell contents.
        print(
            f"[muselab] xlsx parse failed kind={type(e).__name__}",
            file=sys.stderr,
            flush=True,
        )
        raise HTTPException(
            422,
            "failed to parse spreadsheet (file may be corrupt or unsupported)",
        ) from None

    parts: list[str] = [f"# Spreadsheet: {name}"]
    sheets_total = len(wb.sheetnames)
    sheets_truncated = sheets_total > _XLSX_ATTACH_MAX_SHEETS
    try:
        for sheet_name in wb.sheetnames[:_XLSX_ATTACH_MAX_SHEETS]:
            ws = wb[sheet_name]
            parts.append("")
            parts.append(f"## Sheet: {sheet_name}")
            rows_emitted = 0
            sheet_rows = int(ws.max_row or 0)
            sheet_cols = int(ws.max_column or 0)
            rows_truncated = sheet_rows > _XLSX_ATTACH_MAX_ROWS
            cols_truncated = sheet_cols > _XLSX_ATTACH_MAX_COLS
            rows = ws.iter_rows(
                max_row=min(sheet_rows, _XLSX_ATTACH_MAX_ROWS),
                max_col=min(sheet_cols, _XLSX_ATTACH_MAX_COLS),
                values_only=True,
            ) if sheet_rows and sheet_cols else ()
            for row in rows:
                cells: list[str] = []
                for val in row:
                    if val is None:
                        cells.append("")
                    else:
                        s = str(val)
                        if len(s) > _XLSX_ATTACH_CELL_MAX_CHARS:
                            s = s[:_XLSX_ATTACH_CELL_MAX_CHARS] + "…"
                        # CSV-light: only quote/escape if a separator or
                        # quote actually appears (cheap heuristic — the
                        # model parses prose, not strict RFC 4180).
                        if "," in s or '"' in s or "\n" in s:
                            s = '"' + s.replace('"', '""') + '"'
                        cells.append(s)
                parts.append(",".join(cells))
                rows_emitted += 1
            if rows_truncated:
                parts.append(f"… (rows truncated at {_XLSX_ATTACH_MAX_ROWS})")
            if cols_truncated:
                parts.append(f"… (cols truncated at {_XLSX_ATTACH_MAX_COLS})")
            if rows_emitted == 0:
                parts.append("(empty sheet)")
    finally:
        wb.close()

    if sheets_truncated:
        parts.append("")
        parts.append(f"… (sheets truncated at {_XLSX_ATTACH_MAX_SHEETS} "
                     f"of {sheets_total})")

    return "\n".join(parts)


@router.get("/queued-image/{aid}")
def get_queued_image(aid: str, ticket: str = Query("")):
    """FIX ③: serve an as-yet-unsent (queued) image straight from the
    in-memory upload store so the queued-message bubble can render a real
    thumbnail. Unlike /attachments/<sid>/<file> (on-disk, persisted at
    send-time), queued uploads live only in `_image_store` and disappear at
    the 10-min TTL — so this 404s once the entry expires, which the UI
    already surfaces as "附件已过期". A bounded-use resource ticket lets a
    plain ``<img>`` load it without exposing the global API token."""
    import re as _re
    if not _re.fullmatch(r"[A-Za-z0-9]{6,64}", aid):
        raise HTTPException(400, "bad id")
    _require_chat_resource_ticket(ticket, ("queued-image", aid))
    with _image_store_lock:
        _gc_images_locked()
        entry = _get_staged_entry_locked(aid)
        if (entry is None or entry.get("kind") != "image"
                or not entry.get("b64")):
            raise HTTPException(404, "queued image not found or expired")
        encoded = str(entry["b64"])
        mime = str(entry.get("mime") or "image/png")
    from fastapi.responses import Response as _Response
    try:
        data = base64.b64decode(encoded, validate=True)
    except Exception:
        raise HTTPException(404, "queued image unreadable")
    return _Response(
        content=data, media_type=mime,
        headers={"Cache-Control": "private, max-age=300"},
    )


@router.get("/task-output", dependencies=[Depends(require_token)])
def get_task_output(session_id: str = Query(...), path: str = Query(...)):
    """Read a run_in_background task's `.output` file (the bash stdout/stderr
    the SDK writes per task). These live in the SDK's per-session temp tasks
    dir — `/tmp/claude-<uid>/<project>/<session>/tasks/<task_id>.output` —
    OUTSIDE the archive root, so the normal /api/files reader (archive-scoped)
    can't reach them and the card's "open result" link 404s.

    Security: single-user token-gated app, but defense-in-depth anyway — the
    path must match the exact tasks-dir shape AND embed THIS session_id, and we
    reject any `..` so the `.+` project segment can't traverse out. We read the
    literal path (not realpath) so a future local_agent `.output` symlink can't
    redirect us to an arbitrary target."""
    import re as _re
    sid_safe = _re.escape(session_id)
    if (".." in path or not _re.fullmatch(
            rf"/tmp/claude-\d+/.+/{sid_safe}/tasks/[A-Za-z0-9._-]+\.output",
            path)):
        raise HTTPException(400, "bad task-output path")
    p = Path(path)
    if not p.is_file():
        raise HTTPException(404, "task output not found (expired or cleaned up)")
    try:
        data = p.read_text(encoding="utf-8", errors="replace")
    except Exception:
        raise HTTPException(404, "task output unreadable")
    CAP = 200_000
    if len(data) > CAP:
        data = data[:CAP] + "\n\n… (truncated at 200000 chars)"
    from fastapi.responses import PlainTextResponse as _PlainText
    return _PlainText(data, headers={"Cache-Control": "private, max-age=60"})


@router.get("/attachments/{session_id}/{filename}")
def get_attachment(
    session_id: str,
    filename: str,
    ticket: str = Query(""),
):
    """Serve the FULL-RES original of a user-uploaded image saved at
    send-time. Lightbox uses this; the in-stream bubble keeps using the
    160-px thumbnail (small + fast). A short-lived, bounded-use ticket lets
    the browser load the original without exposing the global API token.

    Path traversal guard: filename must be a single basename (no slashes,
    no parent-dir refs) and session_id must be a valid uuid-ish string.
    """
    _require_chat_resource_ticket(
        ticket,
        ("attachment", session_id, filename),
    )
    path = _validate_attachment_ref(session_id, filename)
    # MIME from extension
    ext = filename.rsplit(".", 1)[-1].lower()
    mime = {
        "png": "image/png", "jpg": "image/jpeg", "jpeg": "image/jpeg",
        "gif": "image/gif", "webp": "image/webp",
        # Non-image attachments now live here too. Text types get an explicit
        # charset so the browser doesn't mojibake a UTF-8 Chinese .md/.csv,
        # and PDF gets its real type so the tab renders inline instead of
        # forcing a download.
        "pdf": "application/pdf",
        "txt": "text/plain; charset=utf-8",
        "md": "text/plain; charset=utf-8",
        "csv": "text/plain; charset=utf-8",
        "json": "application/json; charset=utf-8",
        "yaml": "text/plain; charset=utf-8", "yml": "text/plain; charset=utf-8",
        "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "xlsm": "application/vnd.ms-excel.sheet.macroEnabled.12",
    }.get(ext, "application/octet-stream")
    from fastapi.responses import FileResponse
    return FileResponse(
        path, media_type=mime,
        headers={"Cache-Control": "private, max-age=31536000, immutable"},
    )


async def _read_upload_limited(file: UploadFile) -> bytes:
    """Read at most the configured limit plus one byte from a multipart file."""
    chunks: list[bytes] = []
    total = 0
    while total <= _IMAGE_MAX_BYTES:
        remaining = _IMAGE_MAX_BYTES + 1 - total
        chunk = await file.read(min(_UPLOAD_READ_CHUNK_BYTES, remaining))
        if not chunk:
            return b"".join(chunks)
        total += len(chunk)
        if total > _IMAGE_MAX_BYTES:
            raise HTTPException(
                413,
                f"file too large. Max {_IMAGE_MAX_BYTES} bytes (~10MB)",
            )
        chunks.append(chunk)
    raise HTTPException(
        413, f"file too large. Max {_IMAGE_MAX_BYTES} bytes (~10MB)")


@router.post("/upload-image", dependencies=[Depends(require_token)])
async def upload_image(file: UploadFile = File(...)) -> dict:
    """Legacy endpoint name; now handles images + PDF + text-ish docs + xlsx."""
    _t0 = time.perf_counter()
    _gc_images()
    mime = (file.content_type or "").lower()
    name = file.filename or "upload"
    kind = _classify_attachment(mime, name)
    if not kind:
        raise HTTPException(
            400,
            f"unsupported file type: {mime or 'unknown'} ({name}). "
            f"Accepted: images (png/jpg/gif/webp), PDF, text-based docs "
            f"(md/txt/csv/json/yaml/source code), or Excel (xlsx/xlsm).",
        )
    _t_read_start = time.perf_counter()
    body = await _read_upload_limited(file)
    _t_read_end = time.perf_counter()
    aid = uuid.uuid4().hex
    entry: dict = {"kind": kind, "mime": mime, "name": name, "ts": time.time()}
    if kind == "text":
        # `raw` is what actually lands on disk at send-time. The decoded
        # `text` is kept only as a validity check + for anything that still
        # wants a preview — it is NO LONGER pasted into the prompt.
        try:
            entry["text"] = body.decode("utf-8")
        except UnicodeDecodeError:
            raise HTTPException(400, "text file is not valid UTF-8 — "
                                      "convert to UTF-8 or send as PDF") from None
        # The old 200 KB ceiling existed because the whole file was inlined
        # into the prompt, where it burned context 1:1. Attachments now go to
        # disk and the agent Reads only what it needs, so the cap can be the
        # generic upload limit. Kept as a named constant so the error message
        # stays honest about which limit was hit.
        if len(body) > _TEXT_MAX_BYTES:
            _perf_event(
                "upload.large_text", kind="text", bytes=len(body))
        entry["raw"] = body
    elif kind == "xlsx":
        # Store the ORIGINAL workbook (the user asked for the real file to
        # survive) AND a plain-text transcription. Neither alone is enough:
        # `Read` can't parse the zip container, and a transcription throws
        # away formulas / multiple sheets / exact cell values the user may
        # want a script to open later. Both paths go into the prompt.
        # openpyxl load_workbook + full-sheet walk is CPU-heavy and fully
        # synchronous — off-load so a multi-MB xlsx upload doesn't freeze the
        # event loop (and every concurrent SSE stream) mid-parse. (perf: RED —
        # chat.py upload_image xlsx parse)
        entry["raw"] = body
        entry["text"] = await asyncio.to_thread(
            _xlsx_to_text, body, _safe_attach_name(name)
        )
    else:
        # base64-encoding a ~10MB image is tens of ms of pure CPU on the loop
        # — off-load it so the upload doesn't stall concurrent streams.
        # (perf: YELLOW — chat.py upload_image base64)
        entry["b64"] = (await asyncio.to_thread(base64.b64encode, body)).decode("ascii")
    if not await asyncio.to_thread(
        _put_staged_attachment, aid, entry,
    ):
        raise HTTPException(
            503, "staged attachment capacity is temporarily exhausted")
    # Content-free upload timing.  File names and MIME strings can disclose
    # private workspace details, so keep only the closed-set kind and sizes.
    _t_end = time.perf_counter()
    _perf_event(
        "upload.complete",
        kind=kind,
        bytes=len(body),
        read_ms=obs.elapsed_ms(_t_read_start, _t_read_end),
        total_ms=obs.elapsed_ms(_t0, _t_end),
    )
    # Tell the FE the on-disk extension we'll use when persisting this
    # image at send-time. FE assembles the lightbox URL from
    # (currentId, aid, ext) immediately and stores it on the user
    # message — that way the URL survives even if the user reloads
    # before the stream-completion annotation hook fires.
    _EXT_MAP = {"image/png": "png", "image/jpeg": "jpg", "image/jpg": "jpg",
                "image/gif": "gif", "image/webp": "webp"}
    ext = _EXT_MAP.get(mime, "")
    return {"id": aid, "mime": mime, "bytes": len(body),
             "kind": entry["kind"], "name": name,
             "attach_ext": ext}


# Headers every SSE response must carry so reverse proxies (nginx) don't
# buffer/compress the stream — without X-Accel-Buffering even tiny error
# bodies can be held back, delaying the frontend's error toast.
_SSE_HEADERS = {
    "Content-Encoding": "identity",
    "Cache-Control": "no-cache, no-transform",
    "X-Accel-Buffering": "no",
}


def _sse_ping_event() -> ServerSentEvent:
    """Heartbeat as a NAMED ``ping`` SSE event instead of sse_starlette's
    default bare comment (``: ping ...``).

    A comment-only ping keeps the TCP socket warm but is INVISIBLE to the
    browser's EventSource — comments fire no JS event. So a connection that
    silently stalls (public-internet proxy/CDN buffering, laptop sleep-wake,
    a dead-but-not-RST socket) hangs forever: the server finishes the turn
    and persists the reply, yet the client never receives ``done`` and spins
    indefinitely, with neither the browser's ``onerror`` nor our own
    disconnect detection ever firing.

    A named event fires ``es.addEventListener("ping")`` on the frontend,
    giving it a heartbeat to watch. The frontend's stall-watchdog reconnects
    when the heartbeat goes missing past ~2× the interval. Emitted every 15s
    (sse_starlette's DEFAULT_PING_INTERVAL — we don't override the cadence,
    only the message shape)."""
    return ServerSentEvent(data="", event="ping")

# Placeholder prompt injected for image-only turns (image attached, no
# caption). Must NOT be used as an auto-generated session name — see the
# auto-rename guard in _handle_result_message.
_IMAGE_ONLY_PLACEHOLDER = "(image)"


# One-time stream tickets: POST /stream/start (token in HEADER, params in
# JSON body) mints a short-lived single-use ticket; GET /stream?ticket=…
# redeems it. This keeps the user PROMPT and the AUTH TOKEN out of the URL —
# EventSource can't send custom headers or a body, so previously both went
# into the query string, where they leak into uvicorn/proxy access logs,
# browser history, and (for the token) Referer-adjacent surfaces. The legacy
# query-param form still works unchanged for old clients / manual curl.
_STREAM_TICKETS: dict[str, tuple[float, dict]] = {}
_STREAM_TICKET_TTL_S = 60.0
_STREAM_TICKETS_MAX = 64
# stream_start is a SYNC route (runs in Starlette's threadpool) while the
# redeeming GET /stream is async (event-loop thread) — mint and redeem touch
# the dict from different threads, so all access goes through this lock.
_STREAM_TICKETS_LOCK = threading.Lock()
_MUX_STREAM_TICKETS: dict[str, tuple[float, dict]] = {}
_MUX_STREAM_TICKET_TTL_S = 60.0
_MUX_STREAM_TICKETS_MAX = 64
_MUX_STREAM_TICKETS_LOCK = threading.Lock()
_MUX_RECONCILE_INTERVAL_S = 0.2


class StreamStartReq(BaseModel):
    prompt: str = ""
    session_id: str
    # Empty for a new turn. Reconnects pin the exact server turn they observed
    # via /sessions/{sid}/active so they can never attach to a newer turn.
    turn_id: str = ""
    last_event_seq: int = Field(default=0, ge=0)
    model: str = ""
    permission: str = "bypassPermissions"
    image_ids: str = ""
    mobile: bool = False


class TurnStartReq(BaseModel):
    prompt: str = ""
    session_id: str
    model: str = ""
    permission: str = "bypassPermissions"
    image_ids: str = ""
    mobile: bool = False


class MuxCheckpointReq(BaseModel):
    session_id: str = Field(min_length=1, max_length=128)
    turn_id: str = Field(default="", max_length=128)
    last_event_seq: int = Field(default=0, ge=0)


class MuxStreamStartReq(BaseModel):
    checkpoints: list[MuxCheckpointReq] = Field(default_factory=list, max_length=64)
    mobile: bool = False


@router.post("/stream/start", dependencies=[Depends(require_token)])
def stream_start(req: StreamStartReq) -> dict:
    """Mint a one-time ticket for GET /stream. Auth via header; the prompt
    travels in the POST body instead of the SSE URL."""
    import secrets as _secrets
    # Reject malformed permission at mint time (400 with a clear message)
    # instead of letting it fail deep inside SDK connect during the SSE.
    permission = _validate_permission(req.permission)
    now = time.time()
    ticket = _secrets.token_urlsafe(32)
    with _STREAM_TICKETS_LOCK:
        # Sweep expired tickets (tiny dict; O(n) is fine).
        for k in [k for k, (exp, _) in _STREAM_TICKETS.items() if exp < now]:
            _STREAM_TICKETS.pop(k, None)
        while len(_STREAM_TICKETS) >= _STREAM_TICKETS_MAX:
            _STREAM_TICKETS.pop(next(iter(_STREAM_TICKETS)), None)
        _STREAM_TICKETS[ticket] = (now + _STREAM_TICKET_TTL_S, {
            "prompt": req.prompt,
            "session_id": req.session_id,
            "turn_id": req.turn_id,
            "last_event_seq": req.last_event_seq,
            "model": req.model,
            "permission": permission,
            "image_ids": req.image_ids,
            "mobile": req.mobile,
        })
    return {"ticket": ticket}


@router.post("/turns/start", dependencies=[Depends(require_token)])
async def turn_start(req: TurnStartReq) -> dict:
    """Admit and launch a new turn without coupling it to an SSE request."""
    permission = _validate_permission(req.permission)
    _gc_images()
    prompt = req.prompt
    is_image_only = (not prompt.strip()) and bool((req.image_ids or "").strip())
    if is_image_only:
        prompt = _IMAGE_ONLY_PLACEHOLDER
    elif not prompt.strip():
        raise HTTPException(422, "prompt or image_ids required for a new turn")
    try:
        broadcast = await _admit_accept_launch_turn(
            req.session_id,
            prompt,
            model=req.model,
            permission=permission,
            image_ids=req.image_ids,
        )
    except _TurnBusy:
        raise HTTPException(409, "previous turn still running") from None
    except _TurnStartError as exc:
        raise HTTPException(exc.status or 503, str(exc)) from None
    return {
        "accepted": True,
        "session_id": req.session_id,
        "turn_id": broadcast.turn_id,
        "started_at": broadcast.started_at,
    }


@router.post("/stream/mux/start", dependencies=[Depends(require_token)])
def mux_stream_start(req: MuxStreamStartReq) -> dict:
    """Mint a one-time ticket for the aggregate multi-session SSE."""
    import secrets as _secrets

    checkpoints: dict[str, dict] = {}
    for checkpoint in req.checkpoints:
        session_id = checkpoint.session_id.strip()
        turn_id = checkpoint.turn_id.strip()
        if not session_id:
            raise HTTPException(422, "checkpoint session_id required")
        if checkpoint.last_event_seq > 0 and not turn_id:
            raise HTTPException(
                422, "turn_id required when last_event_seq is provided")
        normalized = {
            "session_id": session_id,
            "turn_id": turn_id,
            "last_event_seq": checkpoint.last_event_seq,
        }
        previous = checkpoints.get(session_id)
        if previous is not None and previous != normalized:
            raise HTTPException(
                422, f"contradictory checkpoints for session_id {session_id!r}")
        checkpoints[session_id] = normalized

    now = time.time()
    ticket = _secrets.token_urlsafe(32)
    with _MUX_STREAM_TICKETS_LOCK:
        for key in [
            key for key, (expires_at, _) in _MUX_STREAM_TICKETS.items()
            if expires_at < now
        ]:
            _MUX_STREAM_TICKETS.pop(key, None)
        while len(_MUX_STREAM_TICKETS) >= _MUX_STREAM_TICKETS_MAX:
            _MUX_STREAM_TICKETS.pop(next(iter(_MUX_STREAM_TICKETS)), None)
        _MUX_STREAM_TICKETS[ticket] = (
            now + _MUX_STREAM_TICKET_TTL_S,
            {"checkpoints": checkpoints, "mobile": req.mobile},
        )
    return {"ticket": ticket}


@router.get("/stream/mux")
async def mux_stream(ticket: str = Query(default="")):
    with _MUX_STREAM_TICKETS_LOCK:
        entry = _MUX_STREAM_TICKETS.pop(ticket, None) if ticket else None
    if entry is None or entry[0] < time.time():
        raise HTTPException(401, "invalid or expired mux stream ticket")
    params = entry[1]
    return EventSourceResponse(
        _subscribe_multiplex(
            params.get("checkpoints", {}),
            mobile=bool(params.get("mobile", False)),
        ),
        headers=_SSE_HEADERS,
        ping_message_factory=_sse_ping_event,
    )


@router.get("/stream")
async def stream(
    request: Request,
    prompt: str = Query(default=""),
    token: str = Query(default=""),
    session_id: str = Query(default=""),
    turn_id: str = Query(default=""),
    last_event_seq: int = Query(default=0, ge=0),
    model: str = Query(default=""),
    permission: str = Query(default="bypassPermissions"),
    image_ids: str = Query(default=""),
    mobile: bool = Query(default=False),
    ticket: str = Query(default=""),
):
    # Ticket redemption (preferred path — see _STREAM_TICKETS above). The
    # ticket itself authenticates the request (it was minted via a
    # header-authed POST) and supplies the real params. Single-use: popped
    # on first redemption so a leaked URL from a log replay is inert.
    if ticket:
        with _STREAM_TICKETS_LOCK:
            entry = _STREAM_TICKETS.pop(ticket, None)
        if entry is None or entry[0] < time.time():
            raise HTTPException(401, "invalid or expired stream ticket")
        params = entry[1]
        prompt = params["prompt"]
        session_id = params["session_id"]
        turn_id = params.get("turn_id", "")
        last_event_seq = int(params.get("last_event_seq", 0) or 0)
        model = params["model"]
        permission = params["permission"]
        image_ids = params["image_ids"]
        mobile = bool(params.get("mobile", False))
    else:
        # Legacy query-param auth (old clients / manual use).
        from .auth import _token_ok
        if not _token_ok(token):
            raise HTTPException(401, "bad token")
        if not session_id:
            raise HTTPException(422, "session_id required")
        # Ticketed path already validated at mint; the legacy query-param
        # path takes a raw external string — same gate.
        permission = _validate_permission(permission)
    # TTL sweep of the in-memory attachment store on EVERY stream request
    # (not just when this turn carries image_ids). Without this, a user who
    # uploads then never uploads/sends again would leave a 10MB-class base64
    # entry resident past its TTL — gc previously only ran on upload and on
    # the attachment-consume path. Cheap (O(n) over ≤100 capped entries).
    _gc_images()
    if last_event_seq > 0 and not turn_id:
        raise HTTPException(
            422, "turn_id required when last_event_seq is provided")
    # RECONNECT MODE: empty prompt + NO attached images + an active
    # in-flight turn on this session = subscribe to the existing
    # TurnBroadcast for replay + live tail. Frontend uses this after
    # loadSession discovers that `/sessions/{sid}/active` is true.
    # No new query is sent to the SDK.
    #
    # IMPORTANT: image-only turns (text empty + image_ids set) are a
    # LEGITIMATE new turn — "look at this picture" with no caption.
    # Previously we lumped them into reconnect mode and returned
    # "no active turn", confusing the user (just dropped an image,
    # got a generic error toast).
    is_image_only = (not prompt.strip()) and bool((image_ids or "").strip())

    def _correlate_stream(broadcast: TurnBroadcast) -> None:
        scope_state = request.scope.setdefault("state", {})
        if isinstance(scope_state, dict):
            scope_state["perf_sid8"] = obs.short_id(session_id)
            scope_state["perf_turn8"] = obs.short_id(broadcast.turn_id)

    if not prompt.strip() and not is_image_only:
        existing = _active_turns.get(session_id)
        if existing is None:
            # Grace-keep fallback: the turn may have JUST finished (common for
            # fast server-drained turns) and been popped from _active_turns
            # before this reconnect attached. _recent_turns still holds the
            # finished broadcast within its TTL — subscribing replays the full
            # events + done sentinel, so the drained turn renders live instead
            # of silently requiring a manual refresh.
            recent = _get_recent_turn(session_id)
            if recent is not None:
                _correlate_stream(recent)
                if turn_id and recent.turn_id != turn_id:
                    async def _recent_changed_gen():
                        yield {
                            "event": "resync",
                            "data": json.dumps({
                                "reason": "turn_changed",
                                "requested_turn_id": turn_id,
                                "current_turn_id": recent.turn_id,
                            }),
                        }
                    return EventSourceResponse(
                        _recent_changed_gen(), headers=_SSE_HEADERS)
                return EventSourceResponse(
                    _subscribe_broadcast(
                        recent, mobile=mobile,
                        last_event_seq=last_event_seq,
                    ),
                    headers=_SSE_HEADERS,
                    ping_message_factory=_sse_ping_event,
                )
            async def _no_active_gen():
                yield _error_event("no active turn")
            return EventSourceResponse(_no_active_gen(), headers=_SSE_HEADERS)
        _correlate_stream(existing)
        if turn_id and existing.turn_id != turn_id:
            async def _active_changed_gen():
                yield {
                    "event": "resync",
                    "data": json.dumps({
                        "reason": "turn_changed",
                        "requested_turn_id": turn_id,
                        "current_turn_id": existing.turn_id,
                    }),
                }
            return EventSourceResponse(
                _active_changed_gen(), headers=_SSE_HEADERS)
        return EventSourceResponse(
            _subscribe_broadcast(
                existing, mobile=mobile,
                last_event_seq=last_event_seq,
            ),
            headers=_SSE_HEADERS,
            ping_message_factory=_sse_ping_event,
        )
    # Image-only path: inject a neutral placeholder prompt so the SDK
    # gets non-empty text alongside the attachment. "(image)" is short
    # and language-neutral; Muse handles "what's in this image?" fine
    # from just the attachment + this hint.
    if is_image_only:
        prompt = _IMAGE_ONLY_PLACEHOLDER

    # Admit only the durable ownership boundary before returning SSE. Runtime
    # lock wait, CLI/MCP startup, attachment preparation and query preflight run
    # under a detached owner so a browser disconnect only drops its subscriber.
    try:
        broadcast = await _admit_accept_launch_turn(
            session_id,
            prompt,
            model=model,
            permission=permission,
            image_ids=image_ids,
        )
    except _TurnBusy:
        async def _busy_gen():
            yield {
                "event": "error",
                "data": json.dumps({
                    "error": "previous turn still running",
                    "kind": "turn_busy",
                    "cta": "queue",
                    "retryable": True,
                }),
            }
        return EventSourceResponse(_busy_gen(), headers=_SSE_HEADERS)
    except _TurnStartError as e:
        _err_msg = str(e)
        async def _early_err_gen():
            yield _error_event(_err_msg, activity_source="direct")
        return EventSourceResponse(_early_err_gen(), headers=_SSE_HEADERS)

    # Stop may win while admission is finishing its durable writes. In that
    # case the shared terminal owner has already populated replay; subscribe to
    # it without publishing a new phase or launching runtime work again.
    if broadcast.done or broadcast.cancelled:
        _correlate_stream(broadcast)
        return EventSourceResponse(
            _subscribe_broadcast(broadcast, mobile=mobile),
            headers=_SSE_HEADERS,
            ping_message_factory=_sse_ping_event,
        )

    _correlate_stream(broadcast)
    return EventSourceResponse(
        _subscribe_broadcast(broadcast, mobile=mobile),
        headers=_SSE_HEADERS,
        ping_message_factory=_sse_ping_event,
    )


class _TurnBusy(Exception):
    """Raised by _start_turn when a turn is already in flight on the sid."""


class _TurnCancelledBeforeQuery(Exception):
    """Stop won the final pre-query cancellation check."""


class _TurnStartError(Exception):
    """Raised by _start_turn on setup failure (client init / timeout).
    `status` carries an optional HTTP status hint the /stream handler uses
    to preserve the original 504 response; the headless queue-drain caller
    ignores it (pauses the queue + pushes instead)."""
    def __init__(
        self,
        msg: str,
        status: int | None = None,
        *,
        queue_claim_settled: bool = False,
    ):
        super().__init__(msg)
        self.status = status
        self.queue_claim_settled = bool(queue_claim_settled)


# ---------------------------------------------------------------------------
# Cross-turn background-task settlement + watcher (Phase 2)
# ---------------------------------------------------------------------------

def _release_task_pins(session_id: str, task_ids) -> None:
    """Drop the given task_ids from the session's pin set (unpinning the
    client once nothing keeps it alive). Safe to call with already-removed
    ids. No await → atomic on the event loop."""
    ids = _sessions_with_inflight_tasks.get(session_id)
    if ids is None:
        return
    for tid in list(task_ids):
        ids.discard(tid)
        _bg_task_pinned_at.pop(tid, None)
    if not ids:
        _sessions_with_inflight_tasks.pop(session_id, None)


def _pin_background_task(session_id: str, task_id: str) -> None:
    """Register a launched background task as in flight for the session.

    Single entry point so the launch instant is always recorded alongside the
    pin. The owning watcher derives its absolute termination deadline from this
    stable timestamp across watcher generations.
    """
    if not task_id:
        return
    _sessions_with_inflight_tasks.setdefault(session_id, set()).add(task_id)
    _bg_task_pinned_at.setdefault(task_id, time.time())


def _task_watch_timeout_remaining(task_ids: Iterable[str]) -> float | None:
    """Seconds until the newest pending task reaches its absolute deadline.

    Watchers may be replaced across ordinary turns, but task launch timestamps
    are stable.  Basing each generation on those timestamps prevents every
    respawn from granting the same orphan another full lease.  The newest
    deadline is used so an older sibling cannot prematurely terminate a
    genuinely fresh task. Expiry is handled only by the owning watcher, which
    keeps the fence until it confirms a terminal notification or disconnect.
    """
    ids = [task_id for task_id in task_ids if task_id]
    if not ids:
        return None
    now = time.time()
    newest_pin = max(_bg_task_pinned_at.setdefault(task_id, now)
                     for task_id in ids)
    return max(0.05, newest_pin + _TASK_WATCH_TIMEOUT - now)


def _resolve_runtime_task_owner(
    observer_session_id: str,
    task_id: str,
) -> str | None:
    """Return the runtime that is authoritative for one inherited task.

    Live process pins win over persisted UI state.  This both closes the
    settle/link race and repairs the decision boundary for sidecars written by
    older releases where a successor briefly claimed the same task id.
    """
    sid = str(observer_session_id or "")
    tid = str(task_id or "")
    if not sid or not tid:
        return None
    lineage = sess.runtime_lineage(sid) or [sid]
    for candidate in lineage:
        if tid in _sessions_with_inflight_tasks.get(candidate, ()):
            return candidate
    overlay = sess.get_authoritative_runtime_task_overlays(sid).get(tid, {})
    owner = str(overlay.get("owner_session_id") or "")
    return owner or None


def _runtime_task_owner_disk_snapshot(
    observer_session_id: str,
    task_id: str,
) -> tuple[list[str], str]:
    """Read durable task authority in a worker-thread-friendly helper."""
    sid = str(observer_session_id or "")
    tid = str(task_id or "")
    if not sid or not tid:
        return [], ""
    lineage = sess.runtime_lineage(sid) or [sid]
    overlay = sess.get_authoritative_runtime_task_overlays(sid).get(tid, {})
    return lineage, str(overlay.get("owner_session_id") or "")


async def _resolve_runtime_task_owner_owned(
    observer_session_id: str,
    task_id: str,
) -> str | None:
    """Resolve task authority while keeping index/sidecar I/O off-loop."""
    sid = str(observer_session_id or "")
    tid = str(task_id or "")
    if not sid or not tid:
        return None
    # The overwhelmingly common path is the currently attached runtime. Avoid
    # touching disk at all when its live pin already proves ownership.
    if tid in _sessions_with_inflight_tasks.get(sid, ()):
        return sid
    lineage, durable_owner = await obs.to_thread_io(
        "chat.runtime_task_owner_read",
        sid,
        _runtime_task_owner_disk_snapshot,
        sid,
        tid,
    )
    # Re-check live pins after the worker returns: a rollover may have linked a
    # successor while storage was slow, and live process ownership outranks a
    # replicated UI overlay.
    for candidate in lineage:
        if tid in _sessions_with_inflight_tasks.get(candidate, ()):
            return candidate
    return durable_owner or None


def _settle_background_task(session_id: str, task_id: str) -> bool:
    """Unpin a background task ONCE, from whichever path observes its terminal
    TaskNotification first — the in-turn dispatch or the cross-turn watcher.

    Returns True if THIS call is the one that settled it (so the caller may
    surface the completion), False if the other path already did.

    Dedup is via _sessions_with_inflight_tasks: the check-and-discard below has
    no await (it's sync) so on the single-threaded event loop the two observer
    paths can never both pass the gate. The loser sees the task_id already gone
    and no-ops. Consumes the cross-turn description cache so it can't leak.

    NOTE: this used to ALSO record into the scheduler bell + fire a Web Push.
    That delivery was removed (2026-06-03) — a finishing background task now
    surfaces as a live continuation turn in the originating session (card flips
    to ✅done + the model's auto-continue reaction streams in), matching Claude
    Code's native UX, not as a separate bell notification."""
    settled = True
    if task_id:
        ids = _sessions_with_inflight_tasks.get(session_id)
        if ids is None or task_id not in ids:
            settled = False   # already settled by the other path
        else:
            ids.discard(task_id)
            if not ids:
                _sessions_with_inflight_tasks.pop(session_id, None)
        _bg_task_descriptions.pop(task_id, None)
        _bg_task_tool_use_ids.pop(task_id, None)
        _bg_task_pinned_at.pop(task_id, None)
    return settled


def _on_task_settled(
    session_id: str,
    task_id: str,
    *,
    status: str | None = None,
    tool_use_id: str | None = None,
    summary: str | None = None,
    output_file: str | None = None,
    usage: dict | None = None,
) -> bool | None:
    """SINGLE settlement entry for a background task's terminal signal.

    Every observer path (in-turn typed / in-turn XML fallback / cross-turn
    watcher typed / cross-turn watcher XML fallback) funnels through here, so
    settlement side-effects live in exactly one place: ownership validation,
    durable overlay, then dedup + unpin via
    _settle_background_task (returns False when the other path already won —
    caller must then NOT surface the event).

    Push history of this hook (it keeps flip-flopping; record BOTH rationales
    so the next change is made knowingly):
      - 2026-06-03 removed: "completion surfaces as a live continuation turn"
      - 2026-06-12 reinstated, presence-gated: that rationale only holds when
        someone is watching
      - 2026-06-12 removed again (user decision, same day): a task settling is
        not worth a buzz even when away — its OUTPUT generally feeds the next
        turn; the turn-done push (chat.py _handle_result_message) is the one
        notification the user wants, and queue-paused-on-error still pushes
        because that one means "Muse is stuck waiting for YOU".
    Returns ``True`` for the owner's first terminal signal, ``False`` for a
    same-owner duplicate that may enrich metadata, and ``None`` for a foreign
    or unknown signal. Callers must suppress and otherwise ignore ``None``.

    `status` stays in the signature: callers still report it, and it documents
    the terminal kinds should the push ever come back.
    """
    owner = _resolve_runtime_task_owner(session_id, task_id)
    if owner is None or owner != session_id:
        return None

    # Persist before settlement consumes the launch metadata. Do this even for
    # a duplicate terminal signal: a later typed notification often carries a
    # richer summary/output path than the first terminal TaskUpdated patch.
    _runtime_task_overlay(
        session_id,
        task_id,
        state=str(status or "done"),
        tool_use_id=tool_use_id,
        summary=summary,
        output_file=output_file,
        usage=usage,
    )
    settled = _settle_background_task(session_id, task_id)
    if not settled:
        return False
    # NO push here — see docstring.
    return True


async def _on_task_settled_owned(
    session_id: str,
    task_id: str,
    *,
    status: str | None = None,
    tool_use_id: str | None = None,
    summary: str | None = None,
    output_file: str | None = None,
    usage: dict | None = None,
) -> bool | None:
    """Event-loop-safe settlement used by live SDK message handlers."""
    async with _runtime_task_storage_lock_for(session_id):
        owner = await _resolve_runtime_task_owner_owned(session_id, task_id)
        if owner is None or owner != session_id:
            return None
        await _runtime_task_overlay_owned(
            session_id,
            task_id,
            state=str(status or "done"),
            tool_use_id=tool_use_id,
            summary=summary,
            output_file=output_file,
            usage=usage,
        )
        settled = _settle_background_task(session_id, task_id)
        if not settled:
            return False
        return True


def _render_continuation_message(msg, state: dict):
    """Yield SSE event dicts for one buffered SDK message read during a
    cross-turn continuation (the CLI auto-continue after a bg task finishes).

    This is the watcher's standalone mirror of event_gen's per-message handlers
    — deliberately NOT reusing those closures (they're nested in _start_turn and
    carry per-turn bookkeeping we don't want to re-run: usage stats, sidecar
    annotations, push, jsonl cleanup). Kept minimal: text + tool round-trips,
    which is all an auto-continue reaction produces.

    `state` carries per-continuation mutables:
      - "tool_use_names": tool_use_id -> name, so a later tool_result picks the
        right per-tool renderer.
      - "streamed": list of text chunks already emitted via text_delta, so the
        AssistantMessage TextBlock only tail-emits the suffix the stream skipped
        (mirrors event_gen's streamed_in_bubble dedup).
      - "assistant_uuid": latest persisted AssistantMessage UUID, exposed in
        the terminal done payload so the live footer can fork immediately."""
    if isinstance(msg, StreamEvent):
        ev = getattr(msg, "event", None) or {}
        if ev.get("type") != "content_block_delta":
            return
        delta = ev.get("delta") or {}
        dt = delta.get("type")
        if dt == "text_delta":
            chunk = delta.get("text", "")
            if chunk:
                state["streamed"].append(chunk)
                yield {"event": "text", "data": json.dumps({"text": chunk})}
        elif dt == "thinking_delta":
            chunk = delta.get("thinking", "")
            if chunk:
                yield {"event": "thinking", "data": json.dumps({"text": chunk})}
    elif isinstance(msg, AssistantMessage):
        assistant_uuid = getattr(msg, "uuid", None)
        if assistant_uuid:
            state["assistant_uuid"] = str(assistant_uuid)
        for block in msg.content:
            if isinstance(block, TextBlock):
                full = getattr(block, "text", "") or ""
                streamed_str = "".join(state["streamed"])
                if full and full != streamed_str:
                    tail = (full[len(streamed_str):]
                            if full.startswith(streamed_str) else full)
                    if tail:
                        state["streamed"].append(tail)
                        yield {"event": "text",
                               "data": json.dumps({"text": tail})}
            elif isinstance(block, ThinkingBlock):
                pass  # already streamed via thinking_delta
            elif isinstance(block, ToolUseBlock):
                if block.id:
                    state["tool_use_names"][block.id] = block.name or ""
                yield {"event": "tool_use",
                       "data": json.dumps(_render_tool_use(block))}
                state["streamed"] = []   # FE closeAsst()'s the bubble
            elif isinstance(block, ToolResultBlock):
                tu_id = getattr(block, "tool_use_id", "") or ""
                tname = state["tool_use_names"].get(tu_id, "")
                yield {"event": "tool_result",
                       "data": json.dumps(
                           _render_tool_result(block, tool_name=tname))}
    elif isinstance(msg, UserMessage):
        content = getattr(msg, "content", None)
        if isinstance(content, list):
            for block in content:
                if isinstance(block, ToolResultBlock):
                    tu_id = getattr(block, "tool_use_id", "") or ""
                    tname = state["tool_use_names"].get(tu_id, "")
                    yield {"event": "tool_result",
                           "data": json.dumps(
                               _render_tool_result(block, tool_name=tname))}
    elif isinstance(msg, RateLimitEvent):
        # A rate-limit change can land during a background-task continuation
        # too; record + surface it here so the store stays current off the
        # main turn loop (mirrors event_gen's _handle_rate_limit).
        info = getattr(msg, "rate_limit_info", None)
        if info is not None:
            payload = _record_rate_limit(info)
            yield {"event": "rate_limit", "data": json.dumps(payload)}


async def _watch_inflight_tasks(
    session_id: str,
    client: ClaudeSDKClient,
    pending: dict[str, str | None],
    generation: int | None = None,
    origin_turn_id: str = "",
) -> None:
    """Detached reader keeping an originating CLI client alive past its turn so
    SDK background tasks started in that turn can deliver their terminal
    TaskNotification (the probe showed it lands AFTER ResultMessage) AND so the
    completion surfaces LIVE in the originating session.

    Delivery model (2026-06-03 redesign — matches Claude Code): each terminal
    TaskNotification opens a HEADLESS CONTINUATION turn (a TurnBroadcast
    registered in _active_turns[sid] with is_continuation=True, empty user
    prompt). We publish into it:
      1. the task_notification event (the FE flips the launching card → ✅done),
      2. the CLI's auto-continue model reaction (text + any tool round-trips),
    then a `done` event + finish(). The frontend, while a 'running' bg-task card
    is visible, polls /active and attaches in continuation mode (replay + live
    tail), so the reaction streams in as a new assistant bubble with no user
    action. The CLI also persists the auto-continue to the session JSONL, so a
    user who isn't looking still sees it on next load.

    `pending` maps task_id -> description for tasks still in flight at turn end.
    Drains client.receive_messages() until every pending task settles + its
    continuation closes, or the watch times out.

    SINGLE-READER INVARIANT: this runs only between ordinary user turns. New
    input is rejected by _start_turn and retained in the durable queue until
    this watcher has delivered every settlement + auto-continuation. This is
    deliberately stricter than cancelling the watcher: cancellation could
    leave an old continuation buffered for the next user turn to misattribute."""
    watch_started = obs.monotonic()
    watched_task_ids = set(pending)
    continuation_seen = False
    watch_error_kind = "none"
    cont: TurnBroadcast | None = None
    cont_state: dict | None = None
    subagent_mux = chat_subagents.SubagentStreamMux(session_id)
    watcher_failed = False
    watcher_cancelled = False
    # The continuation is a real assistant turn for display/accounting even
    # though it has no user bubble.  Keep the session's public model id on its
    # broadcast so live and persisted footers do not reload with a blank model.
    continuation_meta = await obs.to_thread_io(
        "chat.background_session_read",
        session_id,
        sess.get_session_meta,
        session_id,
    )
    continuation_model = str(
        (continuation_meta or {}).get("model") or "")

    def _owns_generation() -> bool:
        return (generation is None
                or _continuation_generations.get(session_id) == generation)

    async def _open_continuation() -> None:
        """Register a fresh continuation broadcast in _active_turns[sid] under
        the lock when possible, while ALWAYS retaining a private collector.

        A terminal task event can land in the narrow interval where the source
        turn has published ResultMessage but still owns `_active_turns`.  The
        old implementation returned with ``cont is None`` in that case: the
        card flip rode the live carrier, but the following AssistantMessage was
        never rendered or persisted.  A collector is independent of that UI
        slot; the occupied live broadcast remains an optional second carrier.
        """
        nonlocal cont, cont_state, continuation_seen
        if not _owns_generation():
            return
        b = TurnBroadcast(session_id=session_id, model=continuation_model)
        b.started_at = _background_turn_started_at.get(
            session_id, b.started_at)
        b.parent_turn_id = _background_origin_turn_id.get(session_id, "")
        b.is_continuation = True
        b.perf_client = "warm"
        continuation_seen = True
        async with _lock:
            if not _owns_generation():
                return
            existing = _active_turns.get(session_id)
            if existing is None or existing.done:
                _active_turns[session_id] = b
        cont = b
        cont_state = {
            "tool_use_names": {},
            "streamed": [],
            "assistant_uuid": "",
            # Claude Code normally reacts to a terminal TaskNotification
            # automatically. Some provider/CLI combinations persist the
            # notification and then close receive_messages() before that
            # reaction starts. In that case the watcher explicitly nudges the
            # same SDK session once; this flag prevents an infinite retry loop.
            "explicit_resume_requested": False,
            "incomplete_error": None,
        }

    def _emit_settlement(event: dict) -> bool:
        """Deliver a terminal task event, continuation or not.

        The pump routes to `self._turn or self._background`, and a turn detaches
        its queue at ResultMessage while `_active_turns[sid]` is only popped
        later in _pump_gen_to_broadcast's finally. A task settling inside that
        window is therefore handed to THIS watcher (in-turn dispatch never sees
        it, so it cannot report it) while _open_continuation still refuses to
        take the occupied slot. Publishing only `if cont is not None` dropped
        the event outright: dedup was won here, delivery happened nowhere, and
        the user saw no completion at all (2026-08-04).

        The live turn's broadcast is a perfectly good carrier — the browser is
        already subscribed to it, and `task_notification` is position-
        independent (the FE keys it to the launching card by task_id /
        tool_use_id). So fall back to it, and only report a drop when there is
        genuinely nobody to publish to.
        """
        delivered = False
        if cont is not None:
            cont.publish(event)
            delivered = True
        live = _active_turns.get(session_id)
        if live is not None and live is not cont and not live.done:
            live.publish(event)
            delivered = True
        if delivered:
            return True
        sys.stderr.write(
            f"[chat] task watcher: settlement had no carrier "
            f"sid={session_id[:8]} event={event.get('event')}\n")
        return False

    async def _close_continuation(
        cancelled: bool = False,
        duration_ms: int | None = None,
    ) -> None:
        """Emit a terminal `done`, finish the broadcast, drop it from
        _active_turns (identity-checked so we never pop a newer turn's slot),
        and grace-keep it for a slightly-late FE reconnect."""
        nonlocal cont, cont_state
        b = cont
        state = cont_state
        cont = None
        cont_state = None
        if b is None:
            return
        # Shutdown marks every active broadcast before cancelling its watcher.
        # Treat that as cancellation even when the finally block is the caller;
        # a streamed prefix is not a truthful completed Agent reply.
        cancelled = bool(cancelled or b.cancelled)
        incomplete_error = (state or {}).get("incomplete_error")
        completed_at_ms = int(time.time() * 1000)
        assistant_uuid = str((state or {}).get("assistant_uuid") or "")
        cont_elapsed = (
            round(duration_ms / 1000, 1)
            if duration_ms
            else round(max(0.0, time.time() - b.started_at), 1)
        )
        terminal_status = (
            "cancelled" if cancelled else
            "failed" if incomplete_error else
            "completed"
        )
        b.perf_status = terminal_status
        b.perf_error_kind = (
            "cancelled" if cancelled else
            "background_continuation_incomplete" if incomplete_error else
            "none"
        )
        b.perf_background_count = len(pending)
        done_payload = {
            "cancelled": cancelled,
            "model": b.model,
            "continuation": True,
            "activity_source": b.activity_source,
            "duration_ms": duration_ms,
            "assistant_uuid": assistant_uuid,
            "completed_at_ms": completed_at_ms,
            "background_tasks_pending": len(pending),
        }
        if incomplete_error:
            done_payload.update({
                "is_error": True,
                "error": incomplete_error,
                "kind": "background_continuation_incomplete",
                "cta": "retry",
                "retryable": True,
            })
        # AssistantMessage already supplied the exact persisted UUID. Commit its
        # footer before publishing done so an immediate refresh cannot beat the
        # sidecar write. The write runs off-loop; ordering, not event-loop
        # blocking, is the durability boundary.
        if assistant_uuid and not sess.session_is_deleting(session_id):
            try:
                await obs.to_thread_io(
                    "chat.continuation_footer_write",
                    session_id,
                    sess.set_message_annotation,
                    session_id,
                    assistant_uuid,
                    model=b.model,
                    ts=completed_at_ms,
                    turn_status=terminal_status,
                    turn_id=b.turn_id,
                    elapsed_s=cont_elapsed if cont_elapsed >= 1 else None,
                    file_path=sess._sidecar_path(session_id),
                )
            except Exception as e:
                sys.stderr.write(
                    f"[chat] continuation footer annotation failed "
                    f"sid={session_id[:8]}: {type(e).__name__}\n")
                sys.stderr.flush()
        # The durable footer is now visible to reloads; close the live terminal
        # boundary before slower transcript/outbox bookkeeping.
        b.publish({"event": "done", "data": json.dumps(done_payload)})
        # A rollover intentionally excludes source-only task notifications and
        # continuation text from the successor's Claude transcript.  Stage the
        # final Agent prose in a private durable outbox, then let a separate
        # delivery owner wait for the latest visible leaf to become idle and
        # anchor one presentation-only bubble there.  Do not mark the source
        # replay consumed here: if the eager rollover later rolls back, the
        # still-public source must retain its normal live continuation path.
        if not cancelled and not sess.session_is_deleting(session_id):
            event_id = await obs.to_thread_io(
                "chat.continuation_outbox_write",
                session_id,
                _persist_runtime_continuation_outbox,
                session_id,
                b,
                completed_at_ms=completed_at_ms,
                elapsed_s=cont_elapsed,
                terminal_status=terminal_status,
                incomplete_error=str(incomplete_error or ""),
                file_path=_runtime_continuation_outbox_path(
                    session_id, str(b.turn_id or "")),
            )
            if event_id:
                _schedule_runtime_continuation_delivery(session_id, event_id)
        b.finish()
        async with _lock:
            if _active_turns.get(session_id) is b:
                _active_turns.pop(session_id, None)
        if not sess.session_is_deleting(session_id):
            _remember_recent_turn(session_id, b)
    async def _request_explicit_resume(reason: str) -> bool:
        """Ask the existing SDK session to finish the originating request.

        Claude Code normally starts this turn itself after a completed
        TaskNotification. A stream EOF or grace timeout before ResultMessage
        means that auto-resume was missed. Send one metadata-only user record
        so it is available to the live model but excluded from normal
        transcript rendering/conversation counts by the SDK's `isMeta`
        semantics.
        """
        nonlocal watcher_failed, watch_error_kind
        if (pending or cont is None or cont_state is None
                or last_settle_status == "stopped"
                or cont_state["explicit_resume_requested"]):
            return False
        cont_state["explicit_resume_requested"] = True

        async def _meta_prompt():
            yield {
                "type": "user",
                "message": {
                    "role": "user",
                    "content": (
                        "<system-reminder>All background tasks from the "
                        "previous turn have finished. Continue the original "
                        "user request now: incorporate the task results, "
                        "complete any remaining work, and provide the final "
                        "response. Do not merely report task status."
                        "</system-reminder>"
                    ),
                },
                "parent_tool_use_id": None,
                "isMeta": True,
            }

        try:
            sys.stderr.write(
                f"[chat] task watcher: auto-continuation missing "
                f"sid={session_id[:8]} ({reason}); requesting explicit resume\n")
            await client.query(_meta_prompt())
            return True
        except Exception as e:
            watcher_failed = True
            watch_error_kind = "explicit_resume"
            cont_state["incomplete_error"] = (
                "后台任务已经完成，但最终答复续接失败。请发送“继续”让 Muse "
                f"完成上一轮。（{type(e).__name__}）"
            )
            sys.stderr.write(
                f"[chat] task watcher: explicit resume failed "
                f"sid={session_id[:8]} exc={type(e).__name__} "
                f"kind=explicit_resume\n")
            return False

    def _mark_incomplete() -> None:
        nonlocal watcher_failed
        watcher_failed = True
        if cont_state is not None and not cont_state.get("incomplete_error"):
            cont_state["incomplete_error"] = (
                "后台任务已经完成，但没有生成最终答复。请发送“继续”让 Muse "
                "完成上一轮。"
            )

    # Attach to the session's sole reader. The watcher used to open its OWN
    # `receive_messages()` iterator, which is what made it compete with a turn
    # for the same underlying queue — and therefore why starting a turn while
    # a background task was pending had to be refused with _TurnBusy.
    _bg_stream = _stream_for(client)
    bg_q = _bg_stream.attach_background() if _bg_stream is not None else None
    msg_iter = None if bg_q is not None else client.receive_messages().__aiter__()

    async def _next_message(timeout: float | None):
        """Next message for this watcher, or ``_STREAM_EOF`` when it ends."""
        nonlocal msg_iter
        if bg_q is not None:
            if timeout is not None:
                return await asyncio.wait_for(bg_q.get(), timeout)
            return await bg_q.get()
        try:
            if timeout is not None:
                return await asyncio.wait_for(msg_iter.__anext__(), timeout)
            return await msg_iter.__anext__()
        except StopAsyncIteration:
            return _STREAM_EOF

    def _reopen_stream() -> None:
        """Re-arm the reader after an explicit resume."""
        nonlocal msg_iter
        if bg_q is None:
            msg_iter = client.receive_messages().__aiter__()

    # Status of the most recent settle. A USER-STOPPED task almost never
    # produces an auto-continue reaction (the CLI treats the stop as user
    # intent), so waiting the full _CONTINUATION_GRACE leaves the attached
    # frontend spinning "streaming…" after the card already flipped ⏹
    # (2026-06-11 footer complaint). Use a short grace
    # for stopped settles; a reaction that somehow arrives later is not
    # lost — it buffers in the SDK queue and the next turn's in-turn
    # dispatch drains it.
    last_settle_status: str | None = None

    async def _consume_timeout_terminal(msg: Any) -> bool:
        """Settle terminal messages observed after timeout-issued stops.

        This deliberately mirrors the normal watcher branches below.  It is
        kept local to the timeout recovery window so ordinary lifecycle
        delivery remains unchanged while we wait for positive proof that each
        stop request actually reached a terminal state.
        """
        nonlocal last_settle_status
        terminals: list[dict[str, Any]] = []
        if isinstance(msg, TaskNotificationMessage):
            terminals.append({
                "task_id": getattr(msg, "task_id", "") or "",
                "tool_use_id": getattr(msg, "tool_use_id", None),
                "status": getattr(msg, "status", None),
                "summary": getattr(msg, "summary", None),
                "output_file": getattr(msg, "output_file", None),
                "usage": dict(getattr(msg, "usage", None) or {}),
            })
        elif isinstance(msg, TaskUpdatedMessage):
            terminal = _terminal_task_update(msg)
            if terminal is not None:
                terminals.append(terminal)
        else:
            terminals.extend(_parse_task_notifications(
                _usermsg_task_notification_text(msg)))
        if not terminals:
            return False
        for terminal in terminals:
            task_id = str(terminal.get("task_id") or "")
            if not task_id or task_id not in pending:
                continue
            status = terminal.get("status") or None
            outcome = await _on_task_settled_owned(
                session_id, task_id, status=status,
                tool_use_id=terminal.get("tool_use_id"),
                summary=terminal.get("summary"),
                output_file=terminal.get("output_file"),
                usage=dict(terminal.get("usage") or {}))
            if outcome is None:
                continue
            pending.pop(task_id, None)
            last_settle_status = str(status or "") or None
            if outcome:
                if cont is None:
                    await _open_continuation()
                _emit_settlement({
                    "event": "task_notification",
                    "data": json.dumps({
                        "task_id": task_id,
                        "tool_use_id": terminal.get("tool_use_id"),
                        "status": status,
                        "summary": terminal.get("summary"),
                        "output_file": terminal.get("output_file"),
                        "usage": dict(terminal.get("usage") or {}),
                        "background_tasks_pending": len(pending),
                    }),
                })
        return True

    async def _settle_after_confirmed_disconnect() -> None:
        """Publish a stopped terminal only after the owner process is gone."""
        nonlocal last_settle_status
        for task_id in tuple(pending):
            outcome = await _on_task_settled_owned(
                session_id, task_id, status="stopped",
                summary="后台任务超过安全运行时限，已终止其运行环境。")
            if outcome is None:
                continue
            pending.pop(task_id, None)
            if outcome:
                if cont is None:
                    await _open_continuation()
                _emit_settlement({
                    "event": "task_notification",
                    "data": json.dumps({
                        "task_id": task_id,
                        "tool_use_id": None,
                        "status": "stopped",
                        "summary": (
                            "后台任务超过安全运行时限，已终止其运行环境。"
                        ),
                        "output_file": None,
                        "background_tasks_pending": len(pending),
                    }),
                })
        last_settle_status = "stopped"

    async def _terminate_expired_tasks(reason: str) -> None:
        """Stop expired tasks without ever opening an unsafe queue window.

        ``stop_task`` only acknowledges a control request.  We therefore keep
        every lifecycle pin until either its terminal message arrives or
        ``disconnect_client`` confirms the owning CLI cleanup.  A cleanup that
        cannot yet be confirmed is retried by this same live watcher; the
        session stays fenced and queued work cannot start in between attempts.
        """
        if not pending or not _owns_generation():
            return
        sys.stderr.write(
            f"[chat] task watcher sid={session_id[:8]} {reason}; "
            f"requesting stop for {len(pending)} task(s)\n")
        sys.stderr.flush()

        async def _request_stop(task_id: str) -> bool:
            try:
                await asyncio.wait_for(
                    client.stop_task(task_id),
                    timeout=_TASK_STOP_ACK_TIMEOUT_S,
                )
                return True
            except asyncio.CancelledError:
                raise
            except Exception:
                return False

        stop_results = await asyncio.gather(*(
            _request_stop(task_id) for task_id in tuple(pending)
        ))
        sys.stderr.write(
            f"[chat] task watcher sid={session_id[:8]} stop requests "
            f"acked={sum(stop_results)}/{len(stop_results)}\n")
        sys.stderr.flush()

        # ``asyncio.timeout`` cancels the in-flight ``__anext__`` that was
        # waiting when the absolute lease expired.  Direct SDK iterators treat
        # that cancellation as terminal; attach a fresh iterator for the stop
        # notification.  Pump-backed watchers keep the same background queue.
        _reopen_stream()

        # An ack is not terminal. Continue owning the same reader for a short
        # bounded window so normal stopped notifications can update their tool
        # cards without tearing down the reusable CLI process.
        settle_deadline = (
            asyncio.get_running_loop().time() + _TASK_STOP_SETTLE_GRACE_S)
        while pending and _owns_generation():
            remaining = settle_deadline - asyncio.get_running_loop().time()
            if remaining <= 0:
                break
            try:
                msg = await _next_message(remaining)
            except asyncio.TimeoutError:
                break
            if msg is _STREAM_EOF:
                # EOF proves the message channel is gone, not necessarily that
                # the subprocess cleanup has completed.  Continue to the
                # tracked disconnect fence below before releasing anything.
                break
            if chat_subagents.is_subagent_message(msg):
                if cont is None:
                    await _open_continuation()
                for record in subagent_mux.feed(msg):
                    _emit_settlement({
                        "event": record["event"],
                        "data": json.dumps(record["data"]),
                    })
                continue
            if await _consume_timeout_terminal(msg):
                continue
            if isinstance(msg, ResultMessage):
                await _close_continuation(
                    duration_ms=getattr(msg, "duration_ms", None))
            elif cont is not None and cont_state is not None:
                for event in _render_continuation_message(msg, cont_state):
                    cont.publish(event)

        if not pending:
            return

        # No terminal marker arrived.  Stop the entire owning runtime so a
        # late command/result cannot cross into the queued user turn.  SDK
        # disconnect already owns graceful -> TERM -> KILL escalation.  If it
        # is still running at its bounded public deadline, keep this watcher
        # and its pins alive and retry; never turn a timeout into an unpin.
        attempt = 0
        while pending and _owns_generation():
            attempt += 1
            try:
                await _disconnect_background_task_owner(session_id, client)
            except asyncio.CancelledError:
                raise
            except Exception as exc:
                if attempt == 1 or attempt % 6 == 0:
                    sys.stderr.write(
                        f"[chat] task watcher sid={session_id[:8]} "
                        f"runtime termination pending attempt={attempt} "
                        f"exc={type(exc).__name__}\n")
                    sys.stderr.flush()
                await asyncio.sleep(min(
                    30.0, _TASK_TERMINATION_RETRY_S * attempt))
                continue
            await _settle_after_confirmed_disconnect()
            return

    try:
        async with asyncio.timeout(
            _task_watch_timeout_remaining(pending)
        ) as watch_deadline:
            loop = asyncio.get_running_loop()

            def _reschedule_watch_deadline() -> None:
                remaining = _task_watch_timeout_remaining(pending)
                watch_deadline.reschedule(
                    None if remaining is None else loop.time() + remaining)

            while True:
                # Once every task has settled and we're only waiting on the
                # auto-continue, cap the read so a task that produces no
                # continuation can't pin the client for the full watch timeout.
                read_to = None
                if not pending and cont is not None:
                    read_to = (_STOPPED_CONTINUATION_GRACE
                               if last_settle_status == "stopped"
                               else _CONTINUATION_GRACE)
                try:
                    msg = await _next_message(read_to)
                except asyncio.TimeoutError:
                    if await _request_explicit_resume("continuation grace elapsed"):
                        _reopen_stream()
                        continue
                    if (not pending and cont is not None
                            and last_settle_status != "stopped"):
                        _mark_incomplete()
                    break
                if msg is _STREAM_EOF:
                    if pending:
                        # The message channel ended, but SDK process cleanup may
                        # still be running.  Enter the same stop/disconnect
                        # fence as an absolute timeout; never infer process
                        # death from EOF alone.
                        watcher_failed = True
                        await _terminate_expired_tasks(
                            "message stream ended before task settlement")
                        break
                    if await _request_explicit_resume("message stream ended"):
                        _reopen_stream()
                        continue
                    if (not pending and cont is not None
                            and last_settle_status != "stopped"):
                        _mark_incomplete()
                    break

                if not _owns_generation():
                    break

                if chat_subagents.is_subagent_message(msg):
                    # Forwarded child text may precede TaskNotification.  Open
                    # the headless continuation lazily so no child frame is
                    # discarded while waiting for task settlement.
                    if cont is None:
                        await _open_continuation()
                    for record in subagent_mux.feed(msg):
                        _emit_settlement({
                            "event": record["event"],
                            "data": json.dumps(record["data"]),
                        })
                    continue

                if isinstance(msg, TaskNotificationMessage):
                    # PRIMARY typed path. Phase-0 probe (2026-06-11, CLI
                    # 2.1.141 + SDK 0.2.95) confirmed the terminal
                    # TaskNotificationMessage IS delivered typed, out-of-band
                    # after the turn's ResultMessage — exactly what this
                    # watcher drains. Gate on _settle_background_task so a
                    # task the in-turn dispatch already surfaced isn't
                    # double-fired here.
                    tid = getattr(msg, "task_id", "") or ""
                    won_typed = await _on_task_settled_owned(
                        session_id, tid,
                        status=getattr(msg, "status", None),
                        tool_use_id=getattr(msg, "tool_use_id", None),
                        summary=getattr(msg, "summary", None),
                        output_file=getattr(msg, "output_file", None),
                        usage=dict(getattr(msg, "usage", None) or {}))
                    if won_typed is None:
                        continue
                    last_settle_status = getattr(msg, "status", None)
                    pending.pop(tid, None)
                    _reschedule_watch_deadline()
                    if won_typed:
                        if cont is None:
                            await _open_continuation()
                        _emit_settlement({"event": "task_notification",
                                          "data": json.dumps({
                            "task_id": tid,
                            "tool_use_id": getattr(msg, "tool_use_id", None),
                            "status": getattr(msg, "status", None),
                            "summary": getattr(msg, "summary", None),
                            "output_file": getattr(msg, "output_file", None),
                            "usage": dict(getattr(msg, "usage", None) or {}),
                            "background_tasks_pending": len(pending),
                        })})
                elif isinstance(msg, TaskUpdatedMessage):
                    terminal = _terminal_task_update(msg)
                    if terminal is not None:
                        tid = terminal["task_id"]
                        won_updated = await _on_task_settled_owned(
                            session_id, tid, status=terminal["status"],
                            tool_use_id=terminal.get("tool_use_id"),
                            summary=terminal.get("summary"),
                            output_file=terminal.get("output_file"),
                            usage=dict(terminal.get("usage") or {}))
                        if won_updated is None:
                            continue
                        last_settle_status = terminal["status"]
                        pending.pop(tid, None)
                        _reschedule_watch_deadline()
                        if won_updated:
                            if cont is None:
                                await _open_continuation()
                            terminal["background_tasks_pending"] = len(pending)
                            _emit_settlement({
                                "event": "task_notification",
                                "data": json.dumps(terminal),
                            })
                elif isinstance(msg, TaskStartedMessage):
                    # A task launched DURING the auto-continue reaction (the
                    # model can run tools in that turn, including Bash
                    # run_in_background). Register it exactly like the
                    # in-turn dispatch would — pending + pin + description —
                    # so THIS watcher keeps covering it after the
                    # continuation closes. Without this, the launch was
                    # invisible (no card, no pin, no watcher) and its
                    # terminal notification buffered unread until the next
                    # user turn (2026-06-11 sleep-300 bug).
                    tid = getattr(msg, "task_id", "") or ""
                    desc = getattr(msg, "description", None)
                    accepted_start = (
                        bool(tid) and await _record_background_task_launch_owned(
                            session_id,
                            tid,
                            tool_use_id=getattr(msg, "tool_use_id", None),
                            description=desc,
                        )
                    )
                    if accepted_start:
                        watched_task_ids.add(tid)
                        pending[tid] = desc
                        _pin_background_task(session_id, tid)
                        _reschedule_watch_deadline()
                    if cont is not None and accepted_start:
                        cont.publish({"event": "task_started",
                                      "data": json.dumps({
                            "task_id": tid,
                            "tool_use_id": getattr(msg, "tool_use_id", None),
                            "description": desc,
                            "task_type": getattr(msg, "task_type", None),
                        })})
                elif isinstance(msg, TaskProgressMessage):
                    if cont is not None:
                        cont.publish({"event": "task_progress",
                                      "data": json.dumps({
                            "task_id": getattr(msg, "task_id", "") or "",
                            "tool_use_id": getattr(msg, "tool_use_id", None),
                            "last_tool_name": getattr(
                                msg, "last_tool_name", None),
                            "usage": dict(getattr(msg, "usage", None) or {}),
                        })})
                elif (notifs := _parse_task_notifications(
                        _usermsg_task_notification_text(msg))):
                    # FALLBACK text path: terminal completion arrived as a
                    # user-text <task-notification> XML record instead of the
                    # typed message (older CLI, or a CLI regression). Warn so
                    # we notice fallback traffic — the typed branch above is
                    # the supported contract.
                    sys.stderr.write(
                        f"[chat] task fallback: watcher settled via "
                        f"<task-notification> XML, typed message missed "
                        f"sid={session_id[:8]}\n")
                    won: list[dict[str, Any]] = []
                    for n in notifs:
                        tid = n.get("task_id") or ""
                        outcome = await _on_task_settled_owned(
                            session_id, tid,
                            status=n.get("status") or None,
                            tool_use_id=n.get("tool_use_id") or None,
                            summary=n.get("summary") or None,
                            output_file=n.get("output_file") or None)
                        if outcome is None:
                            continue
                        pending.pop(tid, None)
                        last_settle_status = n.get("status") or None
                        if outcome:
                            won.append(n)
                    _reschedule_watch_deadline()
                    if won and cont is None:
                        await _open_continuation()
                    for n in won:
                        _emit_settlement({"event": "task_notification",
                                          "data": json.dumps({
                            "task_id": n.get("task_id") or "",
                            "tool_use_id": n.get("tool_use_id") or None,
                            "status": n.get("status") or None,
                            "summary": n.get("summary") or None,
                            "output_file": n.get("output_file") or None,
                            "background_tasks_pending": len(pending),
                        })})
                elif isinstance(msg, ResultMessage):
                    # End of the CLI's auto-continue reaction — close the
                    # continuation. If tasks remain in flight, keep reading for
                    # their (later) notifications; otherwise we're done.
                    await _close_continuation(
                        duration_ms=getattr(msg, "duration_ms", None))
                    if not pending:
                        break
                else:
                    if cont is not None and cont_state is not None:
                        for ev in _render_continuation_message(msg, cont_state):
                            cont.publish(ev)
    except asyncio.CancelledError:
        # Shutdown or explicit watcher replacement. Keep pins so a replacement
        # watcher can continue ownership; ordinary new turns never cancel a
        # live watcher.
        watcher_cancelled = True
        watch_error_kind = "cancelled"
        raise
    except asyncio.TimeoutError:
        watcher_failed = True
        watch_error_kind = "timeout"
        await _terminate_expired_tasks("reached its absolute task deadline")
    except Exception as e:
        watcher_failed = True
        watch_error_kind = str(
            _classify_stream_error(str(e)).get("kind") or "unknown")
        sys.stderr.write(
            f"[chat] task watcher sid={session_id[:8]} err: "
            f"{type(e).__name__}; entering safe termination\n")
        await _terminate_expired_tasks("failed before task settlement")
    finally:
        # Release our slot on the session pump so a later turn's messages are
        # not routed to a watcher that has exited.
        if _bg_stream is not None and bg_q is not None:
            _bg_stream.detach_background(bg_q)
        # Close any continuation still open (e.g. grace timeout / outer
        # timeout / stream end with no ResultMessage). Shutdown may cancel after
        # a partial AssistantMessage; pass that cancellation through so the
        # prefix is closed live but never staged as a completed Agent bubble.
        if cont is not None:
            try:
                await _close_continuation(cancelled=watcher_cancelled)
            except Exception as exc:
                watcher_failed = True
                if watch_error_kind == "none":
                    watch_error_kind = str(
                        _classify_stream_error(str(exc)).get("kind")
                        or "continuation_close")
        _safe_last_status = (
            str(last_settle_status)
            if last_settle_status in {
                "completed", "failed", "stopped", "cancelled", "done",
            }
            else "none"
        )
        _watch_status = (
            "cancelled" if watcher_cancelled else
            "failed" if watcher_failed else
            "incomplete" if pending else
            "completed"
        )
        if watch_error_kind == "none" and pending:
            watch_error_kind = "pending"
        _perf_event(
            "chat.background",
            sid8=obs.short_id(session_id),
            origin_turn8=obs.short_id(
                origin_turn_id
                or _background_origin_turn_id.get(session_id, "")),
            tasks_count=len(watched_task_ids),
            settled_count=len(watched_task_ids - set(pending)),
            watch_ms=obs.elapsed_ms(watch_started),
            continuation=continuation_seen,
            status=_watch_status,
            last_status=_safe_last_status,
            error_kind=watch_error_kind,
        )
        # Only clear the registry slot if it still points at us (a fresh
        # watcher may have replaced it).
        if _task_watchers.get(session_id) is asyncio.current_task():
            _task_watchers.pop(session_id, None)
            # User messages submitted while background tasks were running stay
            # in the durable queue. Advance only after this watcher and its
            # final continuation fully release the SDK reader; otherwise a new
            # turn can consume the previous task's auto-continue response and
            # render it under the new user prompt.
            if (_owns_generation()
                    and not _sessions_with_inflight_tasks.get(session_id)):
                _background_turn_started_at.pop(session_id, None)
                _background_origin_turn_id.pop(session_id, None)
                if session_id in _pending_runtime_rebuilds:
                    try:
                        await _rebuild_session_runtime(session_id)
                    except Exception as e:
                        sys.stderr.write(
                            f"[chat] post-task runtime rebuild failed "
                            f"sid={session_id[:8]} "
                            f"exc={type(e).__name__}\n")
                try:
                    await _maybe_drain_queue(session_id)
                except Exception as e:
                    sys.stderr.write(
                        f"[chat] post-task queue drain failed "
                        f"sid={session_id[:8]} "
                        f"exc={type(e).__name__}\n")
        if (_owns_generation()
                and not _sessions_with_inflight_tasks.get(session_id)):
            await _settle_active_turn_sidecar_owned(
                session_id, release=True)


def _merge_session_inflight(
    session_id: str, turn_inflight: dict[str, dict],
) -> dict[str, dict]:
    """Every task still pinned for the session, enriched with this turn's launch
    metadata — the pin set is the SOLE authority on what is still in flight.

    It used to start from ``dict(turn_inflight)`` and union the pin set on top,
    which silently resurrected settled tasks: only the in-turn dispatch pops
    ``inflight_tasks``, so a task the cross-turn WATCHER settled stayed in the
    turn's local dict. At turn end that stale id was handed to a fresh watcher
    (observed 2026-08-04: ``generation=3 pending=['b97zswye9', …]`` for a task
    whose terminal notification had already been consumed) and re-pinned by
    _spawn_task_watcher's caller, so the session reported active:true while
    waiting for a notification that can never arrive a second time.

    Descriptions for prior-turn tasks come from the cross-turn cache. Expired
    pins remain authoritative until their watcher safely terminates the owning
    runtime; a turn-boundary merge must never reap lifecycle state itself.
    """
    merged: dict[str, dict] = {}
    for tid in _sessions_with_inflight_tasks.get(session_id, ()):
        info = turn_inflight.get(tid) or {}
        merged[tid] = {
            "tool_use_id": info.get("tool_use_id"),
            "description": info.get("description")
            or _bg_task_descriptions.get(tid),
        }
    return merged


def _spawn_task_watcher(
    session_id: str,
    client: ClaudeSDKClient,
    inflight: dict[str, dict],
    *,
    started_at: float | None = None,
    origin_turn_id: str = "",
) -> None:
    """Start (or replace) the cross-turn watcher for a session whose just-ended
    turn left background tasks in flight."""
    pending = {
        tid: (info or {}).get("description")
        for tid, info in inflight.items()
    }
    old = _task_watchers.get(session_id)
    if old is not None and not old.done():
        old.cancel()
    if started_at is not None:
        _background_turn_started_at.setdefault(session_id, float(started_at))
    if origin_turn_id:
        _background_origin_turn_id.setdefault(session_id, origin_turn_id)
    generation = _continuation_generations.get(session_id, 0) + 1
    _continuation_generations[session_id] = generation
    _task_watchers[session_id] = asyncio.create_task(
        _watch_inflight_tasks(
            session_id,
            client,
            pending,
            generation,
            origin_turn_id=origin_turn_id,
        ))


async def _retire_unpinned_task_watcher(session_id: str) -> None:
    """Stop a stale watcher after a foreground turn consumed its last task.

    The session pump gives a foreground turn priority over the background
    sink.  A terminal TaskNotification that arrives during that turn therefore
    clears the global pin in `_handle_task_message`, while the old watcher's
    private `pending` set never sees it.  Once the turn confirms there are no
    pins left, cancel that now-ownerless reader so it cannot suppress compact
    and LRU reclamation until the full task-watch timeout.
    """
    if _sessions_with_inflight_tasks.get(session_id):
        return
    watcher = _task_watchers.get(session_id)
    if watcher is None or watcher.done() or watcher is asyncio.current_task():
        return
    watcher.cancel()
    await asyncio.gather(watcher, return_exceptions=True)
    if _task_watchers.get(session_id) is watcher:
        _task_watchers.pop(session_id, None)


async def _handoff_task_watcher(session_id: str) -> None:
    """Clear a COMPLETED watcher before a new turn attaches to the stream.

    A live watcher used to be a hard conflict: it owned the client's message
    stream, so _start_turn had to refuse the turn to keep the one-reader
    invariant. The session pump owns the stream now and routes to the active
    turn and the watcher independently, so a live watcher is simply left alone
    — it is still waiting on background tasks that are none of this turn's
    business. Never cancel it: its buffered auto-continuation belongs to the
    turn that launched the task.
    """
    watcher = _task_watchers.get(session_id)
    if watcher is not None and not watcher.done():
        return
    _task_watchers.pop(session_id, None)
    if not _sessions_with_inflight_tasks.get(session_id):
        _background_turn_started_at.pop(session_id, None)
        _background_origin_turn_id.pop(session_id, None)


async def _release_queue_claim_owned(
    session_id: str,
    item_id: str,
    *,
    turn_id: str = "",
    pause: bool = False,
) -> bool:
    return await obs.to_thread_io(
        "chat.queue_release",
        session_id,
        sess.release_queue_claim,
        session_id,
        item_id,
        turn_id=turn_id,
        pause=pause,
        owned=True,
    )


async def _ack_queue_message_owned(
    session_id: str,
    item_id: str,
    turn_id: str,
) -> bool:
    return await obs.to_thread_io(
        "chat.queue_ack",
        session_id,
        sess.ack_queue_message,
        session_id,
        item_id,
        turn_id,
        owned=True,
    )


async def _finish_cancelled_startup(
    session_id: str,
    broadcast: TurnBroadcast,
) -> TurnBroadcast:
    """Finish a pre-query cancellation without exposing an ABA window.

    Snapshot persistence uses a worker thread and cannot actually stop once it
    starts. Keep one shielded cleanup owner per broadcast, retain the active
    reservation until Activity is terminal, and only then publish/remember/pop
    the turn atomically under ``_lock``.
    """
    cleanup = broadcast._startup_terminal_cleanup_task
    if cleanup is None:
        async def _cleanup() -> bool:
            _hydrate_staged_attachment_display(broadcast)
            await _rollback_broadcast_attachments(broadcast)
            queue_settled = False
            if broadcast.queue_item_id:
                try:
                    queue_settled = await _release_queue_claim_owned(
                        session_id,
                        broadcast.queue_item_id,
                        turn_id=broadcast.turn_id,
                        pause=True,
                    )
                except Exception as exc:
                    # Deletion/corruption must not strand the active slot or
                    # Activity row. Restart recovery retains uncertain claims.
                    sys.stderr.write(
                        f"[chat] cancelled queue rollback failed "
                        f"sid={session_id[:8]} "
                        f"item={broadcast.queue_item_id[:8]} "
                        f"exc={type(exc).__name__}\n"
                    )
            snapshot_ready = False
            if not broadcast.done:
                try:
                    snapshot_ready = await asyncio.to_thread(
                        _persist_cancelled_turn_snapshot, broadcast)
                except Exception as exc:
                    sys.stderr.write(
                        f"[chat] cancelled snapshot failed sid={session_id[:8]} "
                        f"exc={type(exc).__name__}\n"
                    )
            await _finish_activity(session_id, broadcast, "cancelled")
            await _settle_active_turn_sidecar_owned(
                session_id,
                release=bool(snapshot_ready or broadcast.queue_item_id),
            )
            async with _lock:
                if not broadcast.done:
                    broadcast.perf_status = "cancelled"
                    broadcast.perf_error_kind = "cancelled"
                    broadcast.publish({
                        "event": "cancelled",
                        "data": json.dumps({
                            "snapshot_ready": snapshot_ready,
                        }),
                    })
                    broadcast.finish()
                broadcast._startup_queue_settled = queue_settled
                if not sess.session_is_deleting(session_id):
                    _remember_recent_turn(session_id, broadcast)
                if _active_turns.get(session_id) is broadcast:
                    _active_turns.pop(session_id, None)
            return queue_settled

        cleanup = asyncio.create_task(_cleanup())
        broadcast._startup_terminal_cleanup_task = cleanup
    try:
        await asyncio.shield(cleanup)
    except asyncio.CancelledError:
        # A disconnected subscriber may cancel this task while the Stop-button
        # cleanup is already running. Join the single owner before propagating
        # cancellation so no reservation/Activity/replay state is left behind.
        while not cleanup.done():
            try:
                await asyncio.shield(cleanup)
            except asyncio.CancelledError:
                continue
        cleanup.result()
        raise
    return broadcast


async def _start_activity_early(
    session_id: str,
    broadcast: TurnBroadcast,
    prompt: str,
) -> None:
    """Expose a reserved turn while its SDK client is still starting."""
    if broadcast.activity_hidden:
        return
    try:
        from .activity import activity as _activity
        start_task = asyncio.create_task(
            asyncio.to_thread(
                _activity.start,
                session_id,
                summary=prompt,
                activity_source=broadcast.activity_source,
                owner_id=broadcast.turn_id,
            )
        )
        try:
            await asyncio.shield(start_task)
        except asyncio.CancelledError:
            # A worker-thread write cannot be cancelled once it starts. Join it
            # before startup rollback, otherwise it can create a late phantom
            # "running" row after the rollback already tried to finish it.
            while not start_task.done():
                try:
                    await asyncio.shield(start_task)
                except asyncio.CancelledError:
                    # More than one owner may cancel the same startup (request
                    # disconnect plus Stop). Keep the worker wrapper alive and
                    # consume every cancellation until its outcome is known.
                    continue
            try:
                start_task.result()
            except Exception as e:
                sys.stderr.write(
                    f"[activity] start failed sid={session_id[:8]} "
                    f"exc={type(e).__name__}\n")
            else:
                broadcast.activity_started = True
            raise
        else:
            broadcast.activity_started = True
    except Exception as e:
        sys.stderr.write(
            f"[activity] start failed sid={session_id[:8]} "
            f"exc={type(e).__name__}\n")


async def _finish_activity(
    session_id: str,
    broadcast: TurnBroadcast,
    status: str,
) -> None:
    """Close the activity row once, whether startup or the turn ends."""
    if not broadcast.activity_started:
        return
    try:
        from .activity import activity as _activity
        finish_task = asyncio.create_task(
            asyncio.to_thread(
                _activity.finish,
                session_id,
                status,
                activity_source=broadcast.activity_source,
                owner_id=broadcast.turn_id,
            )
        )
        try:
            await asyncio.shield(finish_task)
        except asyncio.CancelledError:
            while not finish_task.done():
                try:
                    await asyncio.shield(finish_task)
                except asyncio.CancelledError:
                    continue
            finish_task.result()
            raise
    except Exception as e:
        sys.stderr.write(
            f"[activity] startup finish failed sid={session_id[:8]} "
            f"exc={type(e).__name__}\n")
    finally:
        broadcast.activity_started = False


async def _abort_turn_startup(
    session_id: str,
    broadcast: TurnBroadcast,
    status: str,
    *,
    pause_queue: bool = True,
    error_text: str = "",
) -> bool:
    """Settle every pre-query owner in one shielded terminal transaction."""
    cleanup = broadcast._startup_terminal_cleanup_task
    if cleanup is None:
        async def _cleanup() -> bool:
            _hydrate_staged_attachment_display(broadcast)
            await _rollback_broadcast_attachments(broadcast)
            broadcast.perf_status = status
            if status != "completed" and broadcast.perf_error_kind == "none":
                broadcast.perf_error_kind = (
                    "cancelled" if status == "cancelled" else "startup")
            queue_settled = False
            snapshot_ready = False
            if broadcast.queue_item_id:
                try:
                    queue_settled = await _release_queue_claim_owned(
                        session_id,
                        broadcast.queue_item_id,
                        turn_id=broadcast.turn_id,
                        pause=pause_queue,
                    )
                except Exception as exc:
                    # Queue corruption/deletion is durable uncertainty, but it
                    # must not strand Activity or the active-turn reservation.
                    sys.stderr.write(
                        f"[chat] startup queue rollback failed "
                        f"sid={session_id[:8]} "
                        f"item={broadcast.queue_item_id[:8]} "
                        f"exc={type(exc).__name__}\n"
                    )

            if not broadcast.queue_item_id:
                visible_error = error_text or (
                    "Turn submission was interrupted before reaching "
                    "canonical history."
                    if status == "cancelled"
                    else "Turn submission failed before reaching canonical "
                    "history."
                )
                try:
                    snapshot_ready = await asyncio.to_thread(
                        _persist_failed_turn_snapshot,
                        broadcast,
                        visible_error,
                        terminal_at_ms=int(time.time() * 1000),
                        canonical_terminal_published=False,
                    )
                except Exception as exc:
                    snapshot_ready = False
                    sys.stderr.write(
                        f"[chat] startup snapshot failed "
                        f"sid={session_id[:8]} exc={type(exc).__name__}\n"
                    )

            # Queue ownership or a durable snapshot supersedes the pending
            # intent. Otherwise retain it as restart-visible recovery state.
            await _settle_active_turn_sidecar_owned(
                session_id,
                release=bool(broadcast.queue_item_id or snapshot_ready),
            )

            await _finish_activity(session_id, broadcast, status)
            # Keep the reservation until every durable owner above is settled.
            # Only then expose a replayable startup terminal event. This ordering
            # makes an immediate browser retry safe: attachments and queue claims
            # are already released when the error becomes visible.
            async with _lock:
                broadcast._startup_queue_settled = queue_settled
                if _active_turns.get(session_id) is broadcast:
                    if not broadcast.done:
                        if status == "cancelled":
                            terminal = {
                                "event": "cancelled",
                                "data": json.dumps({
                                    "startup": True,
                                    "startup_phase": (
                                        broadcast.perf_startup_failure_phase),
                                    "snapshot_ready": snapshot_ready,
                                }),
                            }
                        else:
                            terminal_text = error_text or (
                                "Turn startup ended before the request could be "
                                "submitted. Please retry."
                            )
                            terminal = _error_event(
                                terminal_text,
                                activity_source=broadcast.activity_source,
                            )
                            try:
                                terminal_data = json.loads(
                                    terminal.get("data") or "{}")
                            except (TypeError, ValueError):
                                terminal_data = {"error": terminal_text}
                            terminal_data.update({
                                "startup": True,
                                "startup_phase": (
                                    broadcast.perf_startup_failure_phase),
                                "snapshot_ready": snapshot_ready,
                            })
                            terminal["data"] = json.dumps(
                                terminal_data, ensure_ascii=False)
                        try:
                            broadcast.publish(terminal)
                        except Exception:
                            # A broken replay spool must not strand durable
                            # cleanup, Activity, or the active reservation.
                            broadcast.perf_error_kind = "startup_event"
                    broadcast.emit_startup_perf(
                        "cancelled" if status == "cancelled" else "failed",
                        failure_phase=broadcast.perf_startup_failure_phase,
                    )
                    broadcast.finish()
                    if not sess.session_is_deleting(session_id):
                        _remember_recent_turn(session_id, broadcast)
                    _active_turns.pop(session_id, None)
            _interrupted_at_startup.pop(session_id, None)
            return queue_settled

        cleanup = asyncio.create_task(_cleanup())
        broadcast._startup_terminal_cleanup_task = cleanup
    try:
        return await asyncio.shield(cleanup)
    except asyncio.CancelledError:
        # Multiple Task.cancel() calls may arrive while rollback is joining a
        # real worker. They cannot cut the rest of this terminal transaction.
        while not cleanup.done():
            try:
                await asyncio.shield(cleanup)
            except asyncio.CancelledError:
                continue
        cleanup.result()
        raise


async def _fail_queued_attachment_startup(
    session_id: str,
    broadcast: TurnBroadcast,
) -> bool:
    """Roll a queued attachment failure through the shared terminal owner.

    The durable row retains its original text and staged ids, is returned to
    the paused queue, and can be retried after the attachment lease/files have
    been fully rolled back.
    """
    broadcast.perf_error_kind = "attachment"
    return await _abort_turn_startup(
        session_id,
        broadcast,
        "failed",
        pause_queue=True,
        error_text="attachment preparation failed",
    )


async def _admit_turn(
    session_id: str,
    prompt: str,
    *,
    model: str = "",
    image_ids: str = "",
    queue_item_id: str = "",
) -> "TurnBroadcast":
    """Reserve one turn and durably commit its pending intent.

    This is the HTTP admission boundary: once it returns, the caller may expose
    SSE headers because the active slot, restart sidecar, queue ownership, and
    detached startup owner can be established without waiting for CLI/MCP startup.
    """
    admission_started = obs.monotonic()
    draining = None
    async with _lock:
        if sess.session_is_deleting(session_id):
            raise _TurnStartError("session is being deleted", status=404)
        cur = _active_turns.get(session_id)
        if cur is not None and not cur.done:
            if not cur.cancelled:
                raise _TurnBusy()
            draining = cur
        elif (_sessions_with_inflight_tasks.get(session_id)
              or _session_has_live_watcher(session_id)
              or _session_has_scheduled_delivery(session_id)):
            raise _TurnBusy()
        else:
            broadcast = TurnBroadcast(session_id=session_id, model=model or MODEL)
            _active_turns[session_id] = broadcast
    if draining is not None:
        deadline = time.monotonic() + _INTERRUPT_DRAIN_WAIT_S
        while time.monotonic() < deadline:
            if draining.done or _active_turns.get(session_id) is not draining:
                break
            await asyncio.sleep(0.1)
        async with _lock:
            if sess.session_is_deleting(session_id):
                raise _TurnStartError("session is being deleted", status=404)
            cur = _active_turns.get(session_id)
            if cur is not None and not cur.done:
                raise _TurnBusy()
            if (_sessions_with_inflight_tasks.get(session_id)
                    or _session_has_live_watcher(session_id)
                    or _session_has_scheduled_delivery(session_id)):
                raise _TurnBusy()
            broadcast = TurnBroadcast(session_id=session_id, model=model or MODEL)
            _active_turns[session_id] = broadcast

    broadcast.user_text = prompt
    broadcast.staged_attachment_ids = [
        aid for aid in _attachment_ids(image_ids)
        if _valid_staged_attachment_id(aid)
    ]
    broadcast.startup_owner_task = asyncio.current_task()
    broadcast.queue_item_id = str(queue_item_id or "")
    intent_started = 0.0

    async def _release_reservation_only(error_kind: str) -> None:
        """Drop this new reservation without touching a prior turn's sidecar."""
        async def _cleanup() -> None:
            async with _lock:
                if _active_turns.get(session_id) is broadcast:
                    broadcast.perf_status = "failed"
                    broadcast.perf_error_kind = error_kind
                    broadcast.perf_startup_failure_phase = "accepted"
                    broadcast.finish()
                    broadcast.close()
                    _active_turns.pop(session_id, None)
        cleanup = asyncio.create_task(_cleanup())
        try:
            await asyncio.shield(cleanup)
        except asyncio.CancelledError:
            while not cleanup.done():
                try:
                    await asyncio.shield(cleanup)
                except asyncio.CancelledError:
                    continue
            cleanup.result()
            raise

    async def _settle_admission_cancellation() -> None:
        if intent_started:
            broadcast.perf_intent_ms = obs.elapsed_ms(intent_started)
        broadcast.perf_admission_ms = obs.elapsed_ms(admission_started)
        broadcast.perf_startup_failure_phase = "accepted"
        if broadcast.cancelled:
            await _finish_cancelled_startup(session_id, broadcast)
            return
        broadcast.perf_error_kind = "admission_cancelled"
        await _abort_turn_startup(
            session_id,
            broadcast,
            "failed",
            pause_queue=True,
            error_text="Turn admission was interrupted before startup.",
        )

    if session_id in _interrupted_at_startup:
        try:
            recovered = await _await_thread_completion(
                _recover_interrupted_turn_snapshot, session_id)
        except asyncio.CancelledError:
            # The joined worker may have recovered the old turn, but this new
            # intent has not been persisted yet. Never let its cancellation
            # delete or overwrite the prior process's sidecar.
            await _release_reservation_only("admission_cancelled")
            raise
        if not recovered:
            await _release_reservation_only("interrupted_recovery")
            raise _TurnStartError(
                "previous interrupted turn could not be persisted for recovery")
        _interrupted_at_startup.pop(session_id, None)

    intent_started = obs.monotonic()
    try:
        persisted_intent = await obs.to_thread_io(
            "chat.active_turn_admit",
            session_id,
            _write_active_turn_sidecar,
            broadcast,
            file_path=_active_turn_path(session_id),
            owned=True,
        )
    except asyncio.CancelledError:
        await _settle_admission_cancellation()
        raise
    broadcast.perf_intent_ms = obs.elapsed_ms(intent_started)
    if not persisted_intent:
        broadcast.perf_error_kind = "intent_persist"
        broadcast.perf_startup_failure_phase = "accepted"
        await _abort_turn_startup(
            session_id,
            broadcast,
            "failed",
            pause_queue=True,
            error_text="user message could not be persisted before submission",
        )
        raise _TurnStartError(
            "user message could not be persisted before submission")

    if broadcast.queue_item_id:
        try:
            def _bind_queue_turn() -> None:
                sess.bind_queue_turn(
                    session_id, broadcast.queue_item_id, broadcast.turn_id)
                _durable_attachment_store.mark_queue_turn(
                    session_id, broadcast.queue_item_id, broadcast.turn_id,
                )

            await obs.to_thread_io(
                "chat.queue_bind", session_id, _bind_queue_turn, owned=True)
        except Exception:
            broadcast.perf_error_kind = "queue_bind"
            broadcast.perf_startup_failure_phase = "accepted"
            queue_settled = await _abort_turn_startup(
                session_id,
                broadcast,
                "failed",
                pause_queue=True,
                error_text="could not bind queued message to turn",
            )
            raise _TurnStartError(
                "could not bind queued message to turn",
                queue_claim_settled=queue_settled,
            )

    try:
        await _handoff_task_watcher(session_id)
    except asyncio.CancelledError:
        await _settle_admission_cancellation()
        raise
    except _TurnBusy:
        await _abort_turn_startup(
            session_id, broadcast, "failed", pause_queue=False,
            error_text="Previous background task is still using this session.")
        raise
    if broadcast.cancelled:
        return await _finish_cancelled_startup(session_id, broadcast)
    broadcast.perf_admission_ms = obs.elapsed_ms(admission_started)
    return broadcast


def _launch_admitted_turn(
    broadcast: TurnBroadcast,
    *,
    prompt: str,
    model: str,
    permission: str,
    image_ids: str,
) -> asyncio.Task:
    """Start runtime initialization independently of the SSE subscriber."""
    session_id = broadcast.session_id

    async def _owner() -> None:
        try:
            await _start_turn(
                session_id,
                prompt,
                model=model,
                permission=permission,
                image_ids=image_ids,
                _admitted=broadcast,
            )
        except _TurnStartError:
            # Expected startup failures are durably settled and published by the
            # shared startup cleanup owner before reaching this boundary.
            return
        except asyncio.CancelledError:
            if broadcast.done:
                return
            if broadcast.cancelled:
                await _finish_cancelled_startup(session_id, broadcast)
                return
            broadcast.perf_error_kind = "startup_cancelled"
            broadcast.perf_startup_failure_phase = (
                broadcast.startup_phase or "startup")
            await _abort_turn_startup(
                session_id,
                broadcast,
                "failed",
                pause_queue=True,
                error_text="Turn startup ended unexpectedly. Please retry.",
            )
            raise
        except Exception as exc:
            if broadcast.done:
                return
            error_text = str(exc) or type(exc).__name__
            broadcast.perf_error_kind = str(
                _classify_stream_error(error_text).get("kind") or "startup")
            broadcast.perf_startup_failure_phase = (
                broadcast.startup_phase or "startup")
            await _abort_turn_startup(
                session_id,
                broadcast,
                "failed",
                pause_queue=True,
                error_text=error_text,
            )
        finally:
            if broadcast.startup_owner_task is asyncio.current_task():
                broadcast.startup_owner_task = None

    task = asyncio.create_task(_owner())
    broadcast.startup_owner_task = task
    return task


async def _admit_accept_launch_turn(
    session_id: str,
    prompt: str,
    *,
    model: str = "",
    permission: str = "bypassPermissions",
    image_ids: str = "",
) -> TurnBroadcast:
    """Share the new-turn admission boundary across POST and legacy SSE."""
    broadcast = await _admit_turn(
        session_id,
        prompt,
        model=model,
        image_ids=image_ids,
    )
    # The validated launch permission is immutable for this admitted turn.
    # Publish it before detached runtime startup so an exact-turn steering
    # request in the admission window can retain the same permission guard.
    broadcast.permission = permission
    if broadcast.done or broadcast.cancelled:
        return broadcast
    try:
        broadcast.publish_startup("accepted")
    except Exception:
        broadcast.perf_error_kind = "startup_event"
        broadcast.perf_startup_failure_phase = "accepted"
        await _abort_turn_startup(
            session_id,
            broadcast,
            "failed",
            pause_queue=True,
        )
        raise _TurnStartError(
            "turn could not establish its startup event") from None
    _launch_admitted_turn(
        broadcast,
        prompt=prompt,
        model=model,
        permission=permission,
        image_ids=image_ids,
    )
    return broadcast


async def _start_turn(
    session_id: str,
    prompt: str,
    *,
    model: str = "",
    permission: str = "bypassPermissions",
    plan_return_permission: str = "",
    image_ids: str = "",
    persist_permission: bool = True,
    queue_item_id: str = "",
    _admitted: "TurnBroadcast | None" = None,
) -> "TurnBroadcast":
    """Run an admitted turn through runtime startup and the detached SDK pump.

    Ordinary internal callers still receive the historical behavior. The HTTP
    stream route can admit first, register this coroutine as the detached owner,
    and return EventSourceResponse while this function waits for runtime startup.
    """
    broadcast = _admitted
    if broadcast is None:
        broadcast = await _admit_turn(
            session_id,
            prompt,
            model=model,
            image_ids=image_ids,
            queue_item_id=queue_item_id,
        )
        broadcast.startup_owner_task = asyncio.current_task()
    if broadcast.done or broadcast.cancelled:
        return broadcast
    if not broadcast.startup_phase:
        try:
            broadcast.publish_startup("accepted")
        except Exception:
            broadcast.perf_error_kind = "startup_event"
            broadcast.perf_startup_failure_phase = "accepted"
            await _abort_turn_startup(
                session_id,
                broadcast,
                "failed",
                pause_queue=True,
            )
            raise _TurnStartError(
                "turn could not establish its startup event") from None

    # Pending interrupts are keyed by immutable turn id, so an older turn's
    # delayed terminal bookkeeping cannot suppress this turn's completion.
    # One-session-one-model: if the session already has a locked model,
    # that wins over whatever the frontend's dropdown happens to say. This
    # prevents the "I tried to switch but it didn't take" class of bugs and
    # avoids cross-vendor thinking-signature corruption.
    def _persist_launch_settings() -> tuple[dict, str, str]:
        launch_meta = sess.get_session(session_id) or {}
        locked = (launch_meta.get("model") or "").strip()
        if locked:
            resolved_model = _heal_unreachable_locked_model(
                session_id, locked, model)
            if resolved_model != locked:
                sess.update_model(session_id, resolved_model)
        else:
            # Virgin session — frontend's choice gets persisted on first send.
            resolved_model = model or MODEL
            sess.update_model(session_id, resolved_model)

        # Queue items replay their enqueue-time permission snapshot without
        # rolling the session's newer selection back afterward.
        if persist_permission:
            sess.update_permission(session_id, permission)
            launch_meta = sess.get_session(session_id) or launch_meta
            resolved_plan_return = _normalize_plan_return_permission(
                permission, launch_meta.get("plan_return_permission"))
        else:
            resolved_plan_return = _normalize_plan_return_permission(
                permission, plan_return_permission)
        return launch_meta, resolved_model, resolved_plan_return

    s, model_to_use, plan_return_to_use = await obs.to_thread_io(
        "chat.turn_launch_settings",
        session_id,
        _persist_launch_settings,
        owned=True,
    )
    broadcast.activity_hidden = bool(s.get("activity_hidden", False))
    broadcast.model = model_to_use
    broadcast.permission = permission

    # Reasoning effort and Fast service are per-session launch controls.
    # Legacy empty effort is normalized to canonical `auto` before it reaches
    # either the client-pool key or Gateway custom header.
    effort_to_use = _normalize_effort(s.get("effort"))
    service_tier_to_use = (s.get("service_tier") or "").strip()
    # "Running" means the backend accepted and exclusively reserved the turn,
    # not that a cold CLI/MCP client has already finished starting. Publishing
    # here removes the user-visible startup gap while the terminal cleanup
    # paths below prevent failed starts from becoming phantom running rows.
    try:
        await _start_activity_early(session_id, broadcast, prompt)
    except asyncio.CancelledError:
        if broadcast.cancelled:
            return await _finish_cancelled_startup(session_id, broadcast)
        await _abort_turn_startup(
            session_id, broadcast, "cancelled", pause_queue=True)
        raise
    if broadcast.cancelled:
        return await _finish_cancelled_startup(session_id, broadcast)
    # Wrap get_client so SDK / auth pre-check errors surface as a typed
    # _TurnStartError the caller can shape (the /stream handler → SSE error
    # event / 504; the queue drain → pause + push) instead of bubbling up as
    # a 500. Also: release the reservation we made at the top of NEW-TURN
    # MODE, otherwise this session's slot stays "busy" forever and
    # subsequent sends get rejected.
    _client_started = obs.monotonic()
    broadcast.publish_startup("runtime")
    _runtime_lock_started = obs.monotonic()
    _runtime_lock_acquired = False
    try:
        # Serialize client creation/replacement with scheduler and /compact.
        # The active-turn reservation above is visible before we wait here, so
        # those paths can fail cleanly instead of mutating this runtime.
        async with _session_runtime_lock_for(session_id):
            _runtime_lock_acquired = True
            broadcast.perf_runtime_lock_ms = obs.elapsed_ms(
                _runtime_lock_started)
            client_kwargs: dict[str, Any] = {
                "effort": effort_to_use,
                "service_tier": service_tier_to_use,
            }
            if permission == "plan":
                client_kwargs["plan_return_permission"] = plan_return_to_use
            _client_key = (
                session_id, model_to_use, effort_to_use, service_tier_to_use)
            async with _lock:
                _cached_client = _clients.get(_client_key)
                _cached_permission = _client_permission.get(_client_key)
                _cached_plan_return = _client_plan_return.get(_client_key, "")
            broadcast.perf_client = (
                "warm"
                if (_cached_client is not None
                    and session_id not in _pending_runtime_rebuilds
                    and _cached_permission == permission
                    and (permission != "plan"
                         or _cached_plan_return == plan_return_to_use))
                else "cold"
            )

            def _record_client_phase(phase: str, duration_ms: int) -> None:
                value = max(0, int(duration_ms or 0))
                if phase == "disconnect":
                    broadcast.perf_disconnect_ms += value
                elif phase == "pool":
                    broadcast.perf_pool_ms += value
                elif phase == "creation_lock":
                    broadcast.perf_creation_lock_ms += value
                elif phase == "connect":
                    broadcast.perf_connect_ms += value
                elif phase == "tools":
                    broadcast.publish_startup("tools")
                elif phase == "mcp":
                    broadcast.perf_mcp_ms += value
                elif phase == "pool_commit":
                    broadcast.perf_pool_commit_ms += value

            try:
                client_params = inspect.signature(get_client).parameters.values()
                supports_startup_phase = any(
                    param.name == "startup_phase"
                    or param.kind is inspect.Parameter.VAR_KEYWORD
                    for param in client_params
                )
            except (TypeError, ValueError):
                supports_startup_phase = False
            if supports_startup_phase:
                client_kwargs["startup_phase"] = _record_client_phase
            startup_task = asyncio.create_task(
                get_client(
                    session_id, model_to_use, permission,
                    **client_kwargs))
            broadcast.startup_task = startup_task
            try:
                client = await asyncio.wait_for(startup_task, timeout=60.0)
            finally:
                if broadcast.startup_task is startup_task:
                    broadcast.startup_task = None
    except asyncio.TimeoutError:
        broadcast.perf_error_kind = "timeout"
        broadcast.perf_startup_failure_phase = (
            "tools" if broadcast.startup_phase == "tools" else "runtime"
        )
        timeout_error = "Client connection timed out — CLI process may be hung"
        queue_settled = await _abort_turn_startup(
            session_id, broadcast, "failed", pause_queue=True,
            error_text=timeout_error)
        raise _TurnStartError(
            timeout_error,
            status=504,
            queue_claim_settled=queue_settled)
    except asyncio.CancelledError:
        broadcast.perf_startup_failure_phase = (
            "tools" if broadcast.startup_phase == "tools" else "runtime"
        )
        if broadcast.cancelled:
            return await _finish_cancelled_startup(session_id, broadcast)
        # Cancellation without the user Stop flag is an internal startup
        # failure. Publish a replayable error instead of silently ending SSE.
        broadcast.perf_error_kind = "startup_cancelled"
        await _abort_turn_startup(
            session_id,
            broadcast,
            "failed",
            pause_queue=True,
            error_text="Turn startup ended unexpectedly. Please retry.",
        )
        raise
    except Exception as e:
        err_msg = str(e) or f"{type(e).__name__}"
        broadcast.perf_startup_failure_phase = (
            "tools" if broadcast.startup_phase == "tools" else "runtime"
        )
        broadcast.perf_error_kind = str(
            _classify_stream_error(err_msg).get("kind") or "startup")
        # Free the reservation so the user can fix their config (e.g. add an
        # API key) and immediately retry without waiting for any timeout.
        queue_settled = await _abort_turn_startup(
            session_id, broadcast, "failed", pause_queue=True,
            error_text=err_msg)
        raise _TurnStartError(
            err_msg, queue_claim_settled=queue_settled)
    finally:
        if not _runtime_lock_acquired:
            broadcast.perf_runtime_lock_ms = obs.elapsed_ms(
                _runtime_lock_started)
        broadcast.perf_client_ms = obs.elapsed_ms(_client_started)

    # Stop can race the final instant of client startup: the client may have
    # committed to the pool just after /interrupt snapshotted an empty target
    # list. Do not send the prompt after the user has already cancelled.
    if broadcast.cancelled:
        return await _finish_cancelled_startup(session_id, broadcast)

    # Publish the exact connected runtime only after the immutable active-turn
    # owner has survived startup cancellation. The busy-send endpoint never
    # looks clients up by session id alone; it rechecks this broadcast, turn id,
    # query commit and an in-flight tool before scheduling a steering write.
    if isinstance(client, MuseLabSDKClient):
        broadcast.runtime_client = client

    # Lease staged uploads without consuming them. The payload remains retryable
    # through CPU/disk preparation and native compact preflight; only a
    # successful SDK query write commits the lease.
    broadcast.publish_startup("context")
    _attachment_started = obs.monotonic()
    prepared = _PreparedStagedAttachments()
    if image_ids:
        lease, missing_attachments, busy_attachments = await asyncio.to_thread(
            _lease_staged_attachments,
            image_ids,
            require_all=bool(broadcast.queue_item_id),
            queue_owner=((session_id, broadcast.queue_item_id)
                         if broadcast.queue_item_id else None),
        )
        broadcast._attachment_lease = lease
        if busy_attachments or (
            broadcast.queue_item_id and missing_attachments
        ):
            broadcast.perf_attachment_ms = obs.elapsed_ms(_attachment_started)
            broadcast.perf_startup_failure_phase = "context"
            if broadcast.queue_item_id:
                queue_settled = await _fail_queued_attachment_startup(
                    session_id, broadcast)
            else:
                queue_settled = await _abort_turn_startup(
                    session_id,
                    broadcast,
                    "failed",
                    error_text="attachment is already being submitted",
                )
            reason = (
                "attachment is already being submitted"
                if busy_attachments
                else "queued attachment is missing or expired"
            )
            raise _TurnStartError(
                reason,
                queue_claim_settled=queue_settled,
            )
        if lease is not None:
            try:
                prepared = await _prepare_broadcast_attachments(
                    broadcast, session_id, lease)
            except asyncio.CancelledError:
                broadcast.perf_attachment_ms = obs.elapsed_ms(
                    _attachment_started)
                broadcast.perf_startup_failure_phase = "context"
                if broadcast.cancelled:
                    return await _finish_cancelled_startup(
                        session_id, broadcast)
                broadcast.perf_error_kind = "startup_cancelled"
                await _abort_turn_startup(
                    session_id,
                    broadcast,
                    "failed",
                    pause_queue=True,
                    error_text="Attachment preparation ended unexpectedly. Please retry.",
                )
                raise
            except Exception:
                broadcast.perf_error_kind = "attachment"
                broadcast.perf_attachment_ms = obs.elapsed_ms(
                    _attachment_started)
                broadcast.perf_startup_failure_phase = "context"
                if broadcast.queue_item_id:
                    queue_settled = await _fail_queued_attachment_startup(
                        session_id, broadcast)
                else:
                    queue_settled = await _abort_turn_startup(
                        session_id,
                        broadcast,
                        "failed",
                        error_text="attachment preparation failed",
                    )
                raise _TurnStartError(
                    "attachment preparation failed",
                    queue_claim_settled=queue_settled,
                ) from None
    broadcast.perf_attachment_ms = obs.elapsed_ms(_attachment_started)

    # Stop may arrive while the worker is in an uninterruptible thread. The
    # wrapper joins its real result; re-check before those files can reach any
    # prompt/sidecar/pump boundary.
    if broadcast.cancelled:
        return await _finish_cancelled_startup(session_id, broadcast)

    img_blocks = list(prepared.img_blocks)
    pdf_blocks = list(prepared.pdf_blocks)
    disk_attachments = list(prepared.disk_attachments)
    persisted_imgs = list(prepared.persisted_imgs)
    persisted_docs = list(prepared.persisted_docs)
    # Attachments are referenced BY PATH, never inlined. The prompt gets a
    # manifest; the agent Reads what it needs. Two things this buys:
    #   - context: a big CSV no longer sits in the transcript forever, re-sent
    #     on every subsequent turn of the session
    #   - fidelity: the agent sees the real file (exact bytes, full length),
    #     not a truncated / fenced copy of it
    # The old inline path also needed backtick-fence-length arithmetic to stop
    # an attachment containing ``` from breaking out of its code block and
    # spoofing prompt text. Referencing by path removes that class of bug.
    if disk_attachments:
        parts = [prompt] if prompt else []
        lines = [
            "\n\n--- Files attached to this message (on disk) ---",
            "Use the Read tool on these paths. Do not guess at their contents.",
        ]
        for name, path, note in disk_attachments:
            fields = [
                f"filename={json.dumps(name, ensure_ascii=False)}",
                f"path={json.dumps(path, ensure_ascii=False)}",
            ]
            if note:
                fields.append(
                    f"note={json.dumps(note, ensure_ascii=False)}")
            lines.append("- " + " ".join(fields))
        lines.append("--- end attached files ---")
        parts.append("\n".join(lines))
        prompt = "\n".join(parts).lstrip()

    # New architecture: CLI's JSONL is the transcript source-of-truth. We no
    # longer accumulate `persisted` into a parallel local store. Instead, after
    # the stream completes we ask SDK for the latest message UUIDs and write
    # per-message annotations (cost / model / images) keyed by those UUIDs.
    # Accumulate streamed text as a list of chunks (joined only at the rare
    # read sites), NOT a growing `str +=`. These are *nonlocal* accumulators
    # mutated per token-delta; a `str +=` on a closure cell is O(n²) because
    # CPython's in-place concat optimization fires only for true locals
    # (STORE_FAST), never for cells (STORE_DEREF). Measured: an 80k-delta reply
    # took 13.5s as a nonlocal str+= vs 38ms with this list+join. assistant_acc
    # is read only for truthiness at the end; streamed_in_bubble's content is
    # joined once per AssistantMessage (infrequent). (perf: RED — chat.py O(n²))
    assistant_acc: list[str] = []
    # The SDK gives the persisted transcript UUID on AssistantMessage before
    # ResultMessage closes the turn. Retain the latest one so the early `done`
    # event can identify the reply immediately, without waiting for the slower
    # post-turn transcript scan below.
    last_assistant_uuid = ""
    # Set only after a Memory write has been accepted by the tracked scheduler.
    # The detached pump uses this to retain failure/cancel evidence if a forced
    # interrupt prevents the SDK from emitting its terminal ResultMessage.
    memory_outcome_scheduled = False
    # Mirror of frontend's per-bubble `acc`. Reset on tool_use (FE
    # closeAsst). Lets us tail-emit any TextBlock suffix the SDK didn't
    # send as text_delta — see TextBlock branch below for context.
    streamed_in_bubble: list[str] = []
    # tool_use_id → tool_name lookup populated as we forward ToolUseBlock
    # events. When the matching ToolResultBlock arrives, we attach the
    # name so the FE can pick a per-tool rich renderer (Bash terminal,
    # Read with line gutter, etc.) without re-scanning its own message
    # list. Cleared per turn (lives in event_gen closure).
    tool_use_names: dict[str, str] = {}
    # tool_use_id → description for Bash tool_uses, captured when the ToolUseBlock
    # streams so a later bg-launch tool_result can label the inflight task with a
    # human-readable description (the Bash `description` input). Per-turn closure.
    bg_launch_desc: dict[str, str] = {}
    # task_id → {tool_use_id, description} for SDK background tasks (Agent /
    # Bash run_in_background=true) that emitted a TaskStartedMessage but no
    # terminal TaskNotificationMessage yet. The probe (docs/background-tasks-
    # spec.md §3.4) confirmed the terminal notification lands AFTER this turn's
    # ResultMessage, so anything still in here when the turn ends is in-flight
    # and Phase 2's cross-turn watcher takes over. Lives in event_gen closure,
    # cleared per turn.
    inflight_tasks: dict[str, dict] = {}
    # SDK API failures can arrive as synthetic AssistantMessage objects before
    # the terminal ResultMessage. Accumulate them so the turn's done payload is
    # marked failed even when ResultMessage itself omits the detail.
    turn_sdk_errors: list[dict] = []

    async def _emit_compact(emit, phase: str, **fields) -> None:
        """Push one `compact_progress` SSE event, best-effort.

        A UI cue must never be able to kill the turn it is describing, so every
        failure here is swallowed. `emit` is None on the paths that have no
        stream to write to (tests, the post-turn compact) — also a no-op.
        """
        if emit is None:
            return
        try:
            payload = {"phase": phase, "source": "auto", **fields}
            await emit({"event": "compact_progress",
                        "data": json.dumps(payload)})
        except Exception as e:
            sys.stderr.write(
                f"[chat-preflight] compact_progress emit failed "
                f"sid={session_id[:8]} phase={phase} "
                f"exc={type(e).__name__}\n")
            sys.stderr.flush()

    async def _preflight_compact_if_needed(emit=None) -> None:
        """Use Claude Code's native context accounting before sending a turn.

        The previous auto-compact path only ran after a successful `done` event,
        which is too late for gateways that reject the next request at the API
        boundary. This preflight uses the SDK's `/context` equivalent first, and
        then the SDK-native `/compact` slash command if the effective window is
        close to full.
        """
        nonlocal client
        try:
            # This probe is advisory when no compact is needed. Never let a
            # wedged SDK control request delay the user's real prompt for the
            # full compact window.
            cu = await asyncio.wait_for(
                client.get_context_usage(),
                timeout=min(
                    10,
                    env_int("MUSELAB_COMPACT_TIMEOUT_S", 300, min_value=1),
                ),
            )
        except Exception as e:
            safe_kind = _classify_stream_error(str(e)).get("kind", "unknown")
            sys.stderr.write(
                f"[chat-preflight] get_context_usage skipped sid={session_id[:8]} "
                f"model={model_to_use}: {type(e).__name__} kind={safe_kind}\n")
            sys.stderr.flush()
            if not (
                _is_codex_gateway_model(model_to_use)
                and _is_context_window_failure(e)
            ):
                return
            # An explicit Gateway context rejection is terminal for this
            # transcript.  Do not submit the user's real prompt to the same
            # poisoned runtime.  A background-task owner must keep its process;
            # the outer error path records a deferred rebuild after it settles.
            if (_sessions_with_inflight_tasks.get(session_id)
                    or _session_has_live_watcher(session_id)
                    or _session_has_scheduled_delivery(session_id)):
                raise
            recovery_used, recovery_limit = _context_recovery_inputs(
                session_id,
                model_to_use,
            )
            await _emit_compact(
                emit,
                "start",
                used=recovery_used,
                limit=recovery_limit,
                threshold=0,
            )
            recovery = await _recover_context_session(
                session_id,
                model_to_use,
                pre_tokens=recovery_used,
                context_limit=recovery_limit,
            )
            await _emit_compact(
                emit,
                "end",
                ok=True,
                recovered=True,
                session_id=recovery["session"].get("id", ""),
            )
            raise _ContextRecovered(recovery)
        total = _positive_int(cu.get("totalTokens"))
        sdk_max = _positive_int(cu.get("maxTokens"))
        sdk_raw = _positive_int(cu.get("rawMaxTokens"))
        capability = await _detect_gateway_context_capability(model_to_use)
        details = _context_limit_details(
            model_to_use,
            sdk_max=sdk_max,
            sdk_raw=sdk_raw,
            capability=capability,
        )
        limit = _positive_int(details.get("context_limit"))
        threshold = _compact_threshold(
            model_to_use,
            limit,
            _positive_int(cu.get("autoCompactThreshold")),
            sdk_max=sdk_max,
            capability=capability,
        )
        # Attachments can be expensive; add a rough safety margin rather than
        # pretending the typed text is the whole next request.
        next_est = _rough_prompt_tokens(prompt) + len(img_blocks) * 2500 + len(pdf_blocks) * 12000
        if not threshold or total + next_est < threshold:
            # Still refresh the meter with the effective denominator so the UI can
            # warn before a successful turn completes.
            sess_u = _session_usage.setdefault(session_id, {
                "input_tokens": 0, "output_tokens": 0,
                "cache_read_tokens": 0, "cache_creation_tokens": 0,
                "total_cost_usd": 0.0, "last_turn_at": 0.0,
                "context_used": 0, "context_used_pct": 0.0,
                "context_limit": 0,
            })
            sess_u["context_used"] = total or sess_u.get("context_used", 0)
            _apply_context_limit_details(sess_u, details)
            sess_u["sdk_context_max_tokens"] = sdk_max
            sess_u["sdk_context_raw_max_tokens"] = sdk_raw
            if total:
                _mark_context_used(
                    sess_u,
                    "sdk_context",
                    estimate=endpoints.is_third_party(model_to_use),
                )
            if threshold:
                sess_u["auto_compact_threshold"] = threshold
            if sess_u.get("context_limit") and sess_u.get("context_used"):
                sess_u["context_used_pct"] = round(
                    sess_u["context_used"] / sess_u["context_limit"] * 100, 1)
            return
        # A background watcher shares this client's single SDK message pump.
        # `/compact` uses a temporary turn consumer; if it ran here, task
        # notifications and the model's auto-continuation could be routed into
        # the slash-command queue and discarded at its Result boundary. Let the
        # real user query proceed instead: the ordinary turn dispatcher knows
        # how to forward those lifecycle messages, and the CLI's explicitly
        # configured native auto-compact window still gets a chance to act.
        if (_sessions_with_inflight_tasks.get(session_id)
                or _session_has_live_watcher(session_id)
                or _session_has_scheduled_delivery(session_id)):
            sys.stderr.write(
                f"[chat-preflight] native compact deferred for background "
                f"owner sid={session_id[:8]} model={model_to_use} "
                f"total={total} threshold={threshold}\n")
            sys.stderr.flush()
            return
        sys.stderr.write(
            f"[chat-preflight] native compact sid={session_id[:8]} model={model_to_use} "
            f"total={total} next~={next_est} threshold={threshold} limit={limit}\n")
        sys.stderr.flush()
        # Tell the UI. Without this the auto-compact is indistinguishable from a
        # slow turn: the FE shows the generic "Muse 正在思考…" bubble for the
        # entire compact, which on a long session runs MINUTES (2026-07-25: a
        # 186229/200000 session sat on that bubble for 9m19s and then died on
        # "/compact ended without a ResultMessage" — the user's read was
        # "这是在运行啥呢"). The manual compact path has had a dedicated
        # 📦 bubble since 2026-05-22; this reuses it for the automatic one.
        await _emit_compact(emit, "start", used=total, limit=limit,
                            threshold=threshold)
        # How the command REPORTS itself is a hint; whether the context actually
        # shrank is the fact. Keep the command's verdict aside and let the token
        # count adjudicate, so a compaction that succeeded but failed to
        # acknowledge itself cannot kill the turn it just made room for.
        async def _compact_and_measure(
                target: ClaudeSDKClient,
        ) -> tuple[Exception | None, dict, Exception | None, dict[str, bool]]:
            """Run one compact attempt and return its observed token state.

            A slash command can report an in-band failure even after the CLI
            changed its transcript.  Keep that verdict separate from the
            authoritative post-command context reading so a real token drop
            still wins.
            """
            tail_path, tail_offset = _compact_tail_cursor(session_id)
            cmd_error: Exception | None = None
            try:
                await _run_sdk_command_checked(target, "/compact")
            except Exception as e:
                cmd_error = e
                safe_kind = _classify_stream_error(str(e)).get("kind", "unknown")
                sys.stderr.write(
                    f"[chat-preflight] native compact reported failure "
                    f"sid={session_id[:8]} model={model_to_use}: "
                    f"{type(e).__name__} kind={safe_kind} — verifying\n")
                sys.stderr.flush()
            try:
                measured = dict(await target.get_context_usage())
                measure_error: Exception | None = None
                if not _positive_int(measured.get("totalTokens")):
                    measure_error = RuntimeError(
                        "post-compact context usage did not include totalTokens")
            except Exception as e:
                measured = {}
                measure_error = e
                sys.stderr.write(
                    f"[chat-preflight] post-compact context probe failed "
                    f"sid={session_id[:8]} model={model_to_use}: "
                    f"{type(e).__name__}\n")
                sys.stderr.flush()
            tail_outcome = await asyncio.to_thread(
                _compact_tail_outcome, tail_path, tail_offset)
            return cmd_error, measured, measure_error, tail_outcome

        cmd_error: Exception | None = None
        measure_error: Exception | None = None
        tail_outcome: dict[str, bool] = {
            "boundary": False,
            "summary": False,
            "context_error": False,
        }
        try:
            # One compact deadline covers both the slash command and the
            # authoritative token-drop verification. The old split left the
            # compact bubble spinning forever when the second control request
            # hung after `/compact` had already succeeded.
            async with asyncio.timeout(
                env_int("MUSELAB_COMPACT_TIMEOUT_S", 300, min_value=1)
            ):
                cmd_error, cu2, measure_error, tail_outcome = (
                    await _compact_and_measure(client))
                real_total = _positive_int(cu2.get("totalTokens"))
                compact_pair_written = (
                    tail_outcome.get("boundary")
                    and tail_outcome.get("summary")
                )

                # A Codex Gateway session created by an older bundled CLI may
                # have auto-compaction disabled for the lifetime of that CLI
                # process.  If /compact made no room, rebuild the runtime with
                # the now-explicit auto-compact window and retry exactly once.
                # This is deliberately forbidden while a background task is
                # attached: disconnecting that CLI would terminate real work.
                if (not (total and real_total and real_total < total)
                        and not tail_outcome.get("context_error")
                        and _is_codex_gateway_model(model_to_use)
                        and not _sessions_with_inflight_tasks.get(session_id)
                        and not _session_has_live_watcher(session_id)
                        and not _session_has_scheduled_delivery(session_id)):
                    sys.stderr.write(
                        f"[chat-preflight] rebuilding stalled Codex runtime "
                        f"sid={session_id[:8]} model={model_to_use} "
                        f"total={total} after={real_total}; retrying once\n")
                    sys.stderr.flush()
                    await disconnect_client(session_id)
                    client = await get_client(
                        session_id, model_to_use, permission,
                        **client_kwargs,
                    )
                    # The fresh runtime may observe a compact boundary written
                    # by the first process even when that process's follow-up
                    # probe failed. Measure before issuing a second slash
                    # command; never compact twice when the first one worked.
                    fresh_probe: dict = {}
                    fresh_probe_error: Exception | None = None
                    try:
                        fresh_probe = dict(await client.get_context_usage())
                    except Exception as e:
                        fresh_probe_error = e
                    fresh_total = _positive_int(fresh_probe.get("totalTokens"))
                    if not fresh_total and fresh_probe_error is None:
                        fresh_probe_error = RuntimeError(
                            "fresh runtime context usage did not include totalTokens")
                    if total and fresh_total and fresh_total < total:
                        cmd_error = None
                        measure_error = None
                        cu2 = fresh_probe
                    elif compact_pair_written:
                        # The transcript pair is the CLI's durable compact
                        # commit. A control-plane usage reading can lag one
                        # process generation; trust the new root rather than
                        # running /compact twice or forking unnecessarily.
                        cmd_error = None
                        measure_error = None
                        cu2 = fresh_probe
                    elif cmd_error is None and measure_error is not None \
                            and fresh_probe_error is not None:
                        # The first slash command reported success, but neither
                        # process can tell us whether the compact boundary took
                        # effect. A blind second /compact could summarize an
                        # already summarized transcript. Fail this send and
                        # evict the runtime; the next attempt will re-probe a
                        # clean process instead of risking double compaction.
                        cu2 = fresh_probe
                        measure_error = fresh_probe_error
                    else:
                        cmd_error, cu2, measure_error, tail_outcome = (
                            await _compact_and_measure(client))
        except asyncio.CancelledError:
            # An operator-configured whole-turn timeout or explicit task
            # cancellation can land midway through /compact. Its CLI state is
            # unknowable; never leave that process cached for the next prompt.
            _pending_runtime_rebuilds.add(session_id)
            raise
        except asyncio.TimeoutError:
            _pending_runtime_rebuilds.add(session_id)
            await _emit_compact(
                emit,
                "end",
                ok=False,
                error="native compact timed out during command or verification",
            )
            raise
        except Exception as e:
            _pending_runtime_rebuilds.add(session_id)
            # Can't observe the outcome, so the command's verdict is all we
            # have. No reading also means no evidence of success.
            await _emit_compact(emit, "end", ok=False,
                                error=f"{type(e).__name__}: {e}")
            # This block covers failures in the rebuild/probe orchestration
            # itself. Do not mask a fresh disconnect/client-creation error with
            # an older compact verdict captured before the recovery attempt.
            raise
        real_total = _positive_int(cu2.get("totalTokens"))
        compact_pair_written = bool(
            tail_outcome.get("boundary") and tail_outcome.get("summary"))
        made_room = bool(
            (total and real_total and real_total < total)
            or compact_pair_written
        )
        if not made_room:
            # Genuinely no room made. Once the SDK says compaction is required,
            # that is terminal for this turn — sending the original prompt
            # anyway only produces a second context-window error and trains the
            # UI to offer a useless retry loop.
            reason = (
                f"{type(measure_error).__name__}: {measure_error}"
                if measure_error is not None
                else f"{type(cmd_error).__name__}: {cmd_error}"
                if cmd_error is not None
                else f"context did not shrink ({total} -> {real_total})"
            )
            failure = measure_error or cmd_error
            safe_reason = (
                f"{type(failure).__name__} "
                f"kind={_classify_stream_error(str(failure)).get('kind', 'unknown')}"
                if failure is not None
                else f"context did not shrink ({total} -> {real_total})"
            )
            sys.stderr.write(
                f"[chat-preflight] native compact failed sid={session_id[:8]} "
                f"model={model_to_use}: {safe_reason}\n")
            sys.stderr.flush()
            # The `error` SSE that follows tears the stream down, and the FE
            # clears `compacting` on stream teardown regardless — but emit the
            # terminal phase anyway so the bubble's reason is the compact's,
            # not a generic transport failure.
            deterministic_failure = bool(
                tail_outcome.get("context_error")
                or any(
                    _is_context_window_failure(error)
                    for error in (cmd_error, measure_error)
                )
                or (
                    total and real_total and real_total >= total
                    and measure_error is None
                )
            )
            if (_is_codex_gateway_model(model_to_use)
                    and deterministic_failure):
                # Once the transcript itself is beyond the Gateway request
                # limit, another native /compact is a catch-22: the model must
                # read the history in order to summarize it, but the request is
                # rejected before inference. Flush/retire the dead runtime,
                # preserve the source, and make a bounded recovery fork.
                recovery = await _recover_context_session(
                    session_id,
                    model_to_use,
                    pre_tokens=total,
                    context_limit=limit,
                )
                await _emit_compact(
                    emit,
                    "end",
                    ok=True,
                    recovered=True,
                    session_id=recovery["session"].get("id", ""),
                )
                raise _ContextRecovered(recovery)

            _pending_runtime_rebuilds.add(session_id)
            await _emit_compact(emit, "end", ok=False, error=reason)
            if measure_error is not None:
                raise measure_error
            if cmd_error is not None:
                raise cmd_error
            raise _SDKCommandError({
                "message": (
                    "native /compact reported success but context usage did not decrease "
                    f"({total} -> {real_total})"
                ),
                "source": "verification",
                "api_error_status": None,
            })
        if cmd_error is not None:
            sys.stderr.write(
                f"[chat-preflight] native compact recovered sid={session_id[:8]}: "
                f"command reported {type(cmd_error).__name__} but context shrank "
                f"{total} -> {real_total}; continuing\n")
            sys.stderr.flush()
        real_max = _positive_int(cu2.get("maxTokens"))
        real_raw = _positive_int(cu2.get("rawMaxTokens"))
        refreshed_details = _context_limit_details(
            model_to_use,
            sdk_max=real_max,
            sdk_raw=real_raw,
            capability=capability,
        )
        lim = _positive_int(refreshed_details.get("context_limit"))
        sess_u = _session_usage.setdefault(session_id, {
            "input_tokens": 0, "output_tokens": 0,
            "cache_read_tokens": 0, "cache_creation_tokens": 0,
            "total_cost_usd": 0.0, "last_turn_at": 0.0,
            "context_used": 0, "context_used_pct": 0.0,
            "context_limit": 0,
        })
        if real_total:
            sess_u["context_used"] = real_total
            _mark_context_used(
                sess_u,
                "sdk_context",
                estimate=endpoints.is_third_party(model_to_use),
            )
        _apply_context_limit_details(sess_u, refreshed_details)
        sess_u["sdk_context_max_tokens"] = real_max
        sess_u["sdk_context_raw_max_tokens"] = real_raw
        th = _compact_threshold(
            model_to_use,
            lim,
            _positive_int(cu2.get("autoCompactThreshold")),
            sdk_max=real_max,
            capability=capability,
        )
        if th:
            sess_u["auto_compact_threshold"] = th
        if real_total and lim:
            sess_u["context_used_pct"] = round(real_total / lim * 100, 1)
        # Success. The turn's real query starts immediately after this returns,
        # so the FE swaps the 📦 bubble back for the normal streaming one.
        await _emit_compact(emit, "end", ok=True, used=real_total, limit=lim)

    async def event_gen():
        nonlocal assistant_acc, streamed_in_bubble
        subagent_mux = chat_subagents.SubagentStreamMux(session_id)
        # Subscribe to the session's side-channel queue. The MCP ask_user_question
        # handler publishes here; we merge those events into the SSE stream so the
        # UI can render the question UI while the SDK tool handler is await-ing.
        #
        # CONCURRENCY NOTE (audit E/251): register/unregister are keyed by
        # `session_id` alone and each holds a SINGLE queue slot, so two
        # concurrent /stream turns on the same session (e.g. the same session
        # open in two browser tabs) would have the second register OVERWRITE
        # the first's queue, and whichever turn finishes first would
        # unregister BOTH (the unregister deletes by sid, not by queue
        # identity) — cancelling the other tab's pending AskUserQuestion /
        # permission Futures. This is currently masked because the NEW-TURN
        # mutex above (`_active_turns[sid]`) already rejects a second
        # concurrent turn on the same sid with "previous turn still running",
        # so in practice only one /stream per sid is ever live at a time.
        # If that mutex is ever relaxed, register_session_queue must move to
        # per-(sid, stream-instance) keying and unregister must delete only
        # the queue it created — see ask_user_question.py:register/unregister.
        side_q = register_session_queue(session_id)
        perm_q = perm.register_session_queue(session_id)
        merge_q: asyncio.Queue = asyncio.Queue()
        SENTINEL_DONE = object()

        async def pump_claude():
            """Pull one query's SDK response into the merge queue.

            ``ClaudeSDKClient`` is pooled across turns and its receive queue is
            not turn-scoped. A prior turn's late Task lifecycle — sometimes
            followed by replayed Assistant/Result messages — can therefore be
            the first data returned after the next query. Keep lifecycle events,
            but filter already-persisted response UUIDs and continue past their
            stale Result until the current query reaches its own terminal.
            """
            async def _run_query() -> None:
                # `merge_q` lives in event_gen's scope, one level deeper than
                # _preflight_compact_if_needed's — hence the injected emitter
                # rather than a closure reference. Events ride the same "side"
                # lane as ask_user_question / permission_request: already
                # shaped as {"event", "data"} and passed straight through.
                async def _emit_side(evt: dict) -> None:
                    await merge_q.put(("side", evt))

                # Capture a display-safe tail anchor before preflight. A
                # failing /compact never reaches the ordinary pre-query
                # snapshot below; leaving the default zero anchor would place
                # its durable user+error snapshot at the beginning of a long
                # conversation after refresh. This is only a fallback — a
                # successful compact may change the transcript root, so we
                # refresh it again immediately before the real user query.
                existing_uuids, transcript_boundary = await asyncio.to_thread(
                    _turn_transcript_boundary, session_id, model_to_use)
                broadcast.transcript_boundary = transcript_boundary
                persisted_boundary = await obs.to_thread_io(
                    "chat.active_turn_write",
                    session_id,
                    _write_active_turn_sidecar,
                    broadcast,
                    file_path=_active_turn_path(session_id),
                )
                if not persisted_boundary:
                    broadcast.perf_error_kind = "intent_refresh"
                    broadcast.perf_startup_failure_phase = "context"
                    raise RuntimeError("turn intent boundary could not be persisted")
                _preflight_started = obs.monotonic()
                try:
                    await _preflight_compact_if_needed(_emit_side)
                finally:
                    broadcast.perf_preflight_ms = obs.elapsed_ms(
                        _preflight_started)
                # Snapshot AFTER preflight compact (which may write a new compact
                # boundary) and immediately BEFORE this user query. A cache hit is
                # cheap; to_thread keeps a long-session parse off the event loop.
                existing_uuids, transcript_boundary = await asyncio.to_thread(
                    _turn_transcript_boundary, session_id, model_to_use)
                broadcast.transcript_boundary = transcript_boundary
                boundary = _TurnResponseBoundary(existing_uuids)
                # mem0 recall is supplied by the UserPromptSubmit hook configured
                # on this client. The canonical query remains exactly `prompt` so
                # recalled data is never persisted as a fake user message.
                # Multimodal path when binary blocks (image/pdf) are present.
                # Text/xlsx attachments are represented by the path manifest.
                binary_blocks = [*img_blocks, *pdf_blocks]

                async def _send_query() -> None:
                    # Every preflight/transcript/sidecar await above is a Stop
                    # race. This is the last instruction before SDK transport.
                    if broadcast.cancelled:
                        raise _TurnCancelledBeforeQuery()

                    if not broadcast.perf_query_started:
                        broadcast.perf_query_started = obs.monotonic()
                    _query_write_started = obs.monotonic()
                    try:
                        if binary_blocks:
                            text_block = {"type": "text", "text": prompt}
                            content = [*binary_blocks, text_block]

                            async def gen():
                                yield {
                                    "type": "user",
                                    "message": {
                                        "role": "user", "content": content,
                                    },
                                }
                            await client.query(gen())
                        else:
                            await client.query(prompt)
                    finally:
                        broadcast.perf_query_write_ms = obs.elapsed_ms(
                            _query_write_started)
                    retry_commit = _native_retry_commits.get(session_id)
                    if retry_commit is not None:
                        # SDK query() returning is the first point at which the
                        # child transcript is durably owned by the native
                        # truncating resume.  Consume the restart-safe intent
                        # now; if the metadata write fails, the child JSONL
                        # reconciliation in client construction repairs it.
                        try:
                            await obs.to_thread_io(
                                "chat.retry_intent_commit",
                                session_id,
                                sess.clear_retry_intent,
                                session_id,
                                source_session_id=retry_commit[0],
                                target_user_uuid=retry_commit[1],
                                owned=True,
                            )
                        except Exception as exc:
                            sys.stderr.write(
                                f"[chat] retry intent cleanup pending "
                                f"sid={session_id[:8]} "
                                f"exc={type(exc).__name__}\n"
                            )
                            sys.stderr.flush()
                        finally:
                            _native_retry_commits.pop(session_id, None)
                    # query() is the transport commit point. Until it returns,
                    # the lease is still retryable; after it succeeds, consume
                    # the exact staged objects before receiving any response.
                    try:
                        _commit_broadcast_attachments(broadcast)
                    except _AttachmentCommitUncertain:
                        # A response may now be queued on this runtime; never
                        # let the next turn adopt it as its own response.
                        _pending_runtime_rebuilds.add(session_id)
                        raise
                    broadcast.query_committed = True
                    broadcast.steering_ready.set()
                    broadcast.emit_startup_perf("ready")
                    if persisted_imgs:
                        try:
                            await _await_thread_completion(
                                sess.append_pending_attachments,
                                session_id,
                                list(persisted_imgs),
                            )
                        except Exception as exc:
                            sys.stderr.write(
                                "[attach] pending annotation failed "
                                f"sid={obs.short_id(session_id)} "
                                f"exc={type(exc).__name__}\n")
                            sys.stderr.flush()

                replay_dropped = 0
                deferred_result: ResultMessage | None = None
                background_messages: list[Any] = []

                async def _dispatch(msg) -> str:
                    """Classify one message and forward it to the turn."""
                    nonlocal replay_dropped, deferred_result
                    if isinstance(msg, CommandLifecycleMessage):
                        terminal = await _settle_steering_lifecycle(
                            broadcast, msg)
                        if (
                            terminal
                            and not broadcast.steering_commands
                            and deferred_result is not None
                        ):
                            final_result = deferred_result
                            deferred_result = None
                            broadcast.result_forwarded = True
                            broadcast.active_tool_use_ids.clear()
                            await merge_q.put(("claude", final_result))
                            return "current_result"
                        return "forward"
                    decision = boundary.classify(msg)
                    if decision in ("drop", "stale_result"):
                        replay_dropped += 1
                        return decision
                    if decision in ("background", "background_result"):
                        background_messages.append(msg)
                        return decision

                    # Keep an exact, live view of tool execution. The native
                    # priority=next queue consumes at PostToolBatch; these IDs
                    # are also useful to explain why a command is still
                    # waiting, but eligibility does not require one — a command
                    # sent before the next tool is chosen can still fold there.
                    if isinstance(msg, (AssistantMessage, UserMessage)):
                        for block in (getattr(msg, "content", None) or []):
                            if isinstance(block, ToolUseBlock) and block.id:
                                broadcast.active_tool_use_ids.add(block.id)
                            elif isinstance(block, ToolResultBlock):
                                tool_id = str(
                                    getattr(block, "tool_use_id", "") or "")
                                if tool_id:
                                    broadcast.active_tool_use_ids.discard(tool_id)

                    if decision == "current_result":
                        # A Result can race the enqueue HTTP request while its
                        # one-frame write is in progress. Wait briefly for that
                        # known write outcome; a failed write removes the map
                        # and lets this Result close normally, while an accepted
                        # command keeps the sole stream reader attached.
                        await _await_steering_write_stability(broadcast)
                        if broadcast.steering_commands:
                            deferred_result = msg
                            return "steering_pending_result"
                        broadcast.result_forwarded = True
                        broadcast.active_tool_use_ids.clear()
                    if (broadcast.perf_query_started
                            and broadcast.perf_first_event_ms < 0):
                        broadcast.perf_first_event_ms = obs.elapsed_ms(
                            broadcast.perf_query_started)
                    # Record terminal system signals in the pump before reading
                    # the next item. If the stream closes without ResultMessage,
                    # event_gen may not have consumed the queued row yet.
                    if (isinstance(msg, SystemMessage)
                            and (error_info := _sdk_system_error(msg))):
                        turn_sdk_errors.append(error_info)
                    await merge_q.put(("claude", msg))
                    return decision

                stream = _stream_for(client)
                if stream is not None:
                    # Attach BEFORE query(): query() yields to the transport,
                    # so an immediate response can otherwise reach the pump's
                    # orphan park before this consumer exists. The old outer
                    # loop existed only because receive_response() returns at
                    # EVERY ResultMessage — including a replayed one. A queue
                    # has no such boundary, so stale results keep draining.
                    turn_q = stream.attach_turn()
                    try:
                        await _send_query()
                        while True:
                            msg = await turn_q.get()
                            if msg is _STREAM_EOF:
                                # Never turn a dead SDK stream into a silent,
                                # apparently-successful zero-event turn. The
                                # pump has already evicted its cached client;
                                # raising here produces a visible SSE error and
                                # makes the next send build a fresh runtime.
                                captured = _merge_sdk_errors(turn_sdk_errors)
                                raise (
                                    ClaudeSDKError(captured["message"])
                                    if captured else stream._failure
                                    or ClaudeSDKError(
                                        "SDK message stream ended without "
                                        "a ResultMessage")
                                )
                            if await _dispatch(msg) == "current_result":
                                break
                    finally:
                        stream.detach_turn(turn_q)
                        # Explicit non-human MessageOrigin frames belong to a
                        # background/peer delivery, never the foreground human
                        # turn. Preserve them in order for the task watcher or
                        # continuation consumer instead of mixing their text,
                        # usage and Result boundary into this reply.
                        stream.park_messages(background_messages)
                        # The pump may have already queued lifecycle records
                        # after Result. Return every leftover to the orphan
                        # park so a background watcher can adopt them instead
                        # of silently losing task settlements/continuations.
                        stream.park_unconsumed(turn_q)
                else:
                    # No pump: this client was not created through get_client
                    # (test doubles inject their own). Fall back to the SDK's
                    # bounded iterator with the original re-entry loop.
                    await _send_query()
                    while True:
                        stale_terminal = False
                        current_terminal = False
                        async for msg in client.receive_response():
                            decision = await _dispatch(msg)
                            if decision in ("stale_result", "background_result"):
                                stale_terminal = True
                            elif decision == "current_result":
                                current_terminal = True
                                break
                        if current_terminal or not stale_terminal:
                            break
                    if not current_terminal:
                        captured = _merge_sdk_errors(turn_sdk_errors)
                        if captured:
                            raise ClaudeSDKError(captured["message"])
                if replay_dropped:
                    sys.stderr.write(
                        f"[chat-stream] dropped stale replay sid={session_id[:8]} "
                        f"messages={replay_dropped}\n")
                    sys.stderr.flush()
            terminal_kind = ""
            terminal_payload: Any = None
            try:
                async with _session_runtime_lock_for(session_id):
                    await _run_query()
            except _TurnCancelledBeforeQuery:
                terminal_kind = "cancelled"
            except _ContextRecovered as e:
                # Expected control transfer: the old transcript was preserved
                # and the browser will adopt the returned recovery session.
                # Keep service logs concise and never dump summary/prompt data.
                sys.stderr.write(
                    f"[chat-preflight] recovery session ready "
                    f"sid={session_id[:8]} model={model_to_use}\n")
                sys.stderr.flush()
                terminal_kind, terminal_payload = "error", e
            except Exception as e:
                # Keep enough structure for diagnosis without copying an SDK
                # exception, prompt, path, credential or protocol payload into
                # service logs. The authenticated SSE error remains the user's
                # detailed surface.
                error_kind = _classify_stream_error(str(e)).get(
                    "kind", "unknown")
                sys.stderr.write(
                    f"[chat-stream] sid={session_id[:8]} model={model_to_use} "
                    f"exc={type(e).__name__} kind={error_kind}\n")
                sys.stderr.flush()
                terminal_kind, terminal_payload = "error", e
            finally:
                # Do not expose a terminal failure while its staged id is still
                # busy. Retrying as soon as the browser sees error is safe.
                await _rollback_broadcast_attachments(broadcast)
                if terminal_kind:
                    await merge_q.put((terminal_kind, terminal_payload))
                await merge_q.put(("done", SENTINEL_DONE))

        async def pump_side_q(src_q):
            """Pull from a side channel (MCP tool / permission) into merge queue."""
            try:
                while True:
                    evt = await src_q.get()
                    try:
                        await merge_q.put(("side", evt))
                    finally:
                        src_q.task_done()
            except asyncio.CancelledError:
                pass

        async def _prepare_side_event(payload):
            if isinstance(payload, dict) and payload.get("event") in {
                "ask_user_question", "permission_request"
            } and not broadcast.activity_hidden:
                try:
                    from .activity import activity as _activity
                    await asyncio.to_thread(
                        _activity.set_state, session_id,
                        "waiting_approval", detail="Waiting for user input")
                except Exception as e:
                    sys.stderr.write(
                        f"[activity] waiting state failed sid={session_id[:8]} "
                        f"exc={type(e).__name__}\n")
            return payload

        async def _flush_side_channels():
            """Move hook/approval events ahead of the terminal SSE boundary.

            ExitPlanMode hooks enqueue their mode event before the SDK emits
            ResultMessage, but the independent side pump may not have run yet.
            Queue.join() provides a real hand-off barrier; a scheduler race can
            no longer let `done` overtake permission resolution.
            """
            await asyncio.gather(side_q.join(), perm_q.join())
            while True:
                try:
                    queued_kind, queued_payload = merge_q.get_nowait()
                except asyncio.QueueEmpty:
                    break
                if queued_kind == "side":
                    yield await _prepare_side_event(queued_payload)

        # ====== message-type-specific handlers ======
        # Three nested async generators, one per SDK message type. They share
        # closure state (assistant_acc + streamed_in_bubble via nonlocal,
        # other locals read-only). Keeps the main loop a ~15-line dispatch
        # instead of a 200-line elif chain.

        async def _handle_stream_event(msg):
            """Token-by-token deltas → tiny text / thinking events. Fast
            feedback path; the AssistantMessage handler suppresses re-emit."""
            nonlocal assistant_acc, streamed_in_bubble
            ev = msg.event or {}
            if ev.get("type") != "content_block_delta":
                return
            delta = ev.get("delta") or {}
            dt = delta.get("type")
            if dt == "text_delta":
                chunk = delta.get("text", "")
                if chunk:
                    assistant_acc.append(chunk)
                    streamed_in_bubble.append(chunk)
                    yield {"event": "text", "data": json.dumps({"text": chunk})}
            elif dt == "thinking_delta":
                chunk = delta.get("thinking", "")
                if chunk:
                    yield {"event": "thinking", "data": json.dumps({"text": chunk})}

        async def _handle_assistant_message(msg):
            """Per-turn AssistantMessage:
              1. Snapshot per-turn usage (msg.usage is raw Anthropic per-call
                 dict; populate sess_u truthfully for the context meter).
              2. Accumulate per-turn tokens into the global _stats (truth
                 for `/api/chat/usage`). ResultMessage.usage is cumulative
                 per session and would double-count, so we do it here.
              3. Iterate content blocks — tail-emit TextBlock suffix the
                 stream may have skipped; forward tool_use / tool_result.
            """
            nonlocal assistant_acc, streamed_in_bubble, last_assistant_uuid
            # Capture the canonical boundary even for a synthetic API-error
            # AssistantMessage. Its text is intentionally suppressed from the
            # normal token stream, but done/footer persistence and canonical
            # reconciliation still need the exact UUID.
            assistant_uuid = getattr(msg, "uuid", None)
            if assistant_uuid:
                last_assistant_uuid = str(assistant_uuid)
                broadcast.last_assistant_uuid = last_assistant_uuid
            if error_info := _sdk_assistant_error(msg):
                turn_sdk_errors.append(error_info)
                # Synthetic API-error assistants often carry the raw error as a
                # TextBlock. Do not render it as if it were a normal Muse reply;
                # the terminal done event below surfaces the classified failure.
                return
            a_usage = getattr(msg, "usage", None) or {}
            if a_usage:
                in_t = int(a_usage.get("input_tokens", 0) or 0)
                cr_t = int(a_usage.get("cache_read_input_tokens", 0) or 0)
                cc_t = int(a_usage.get("cache_creation_input_tokens", 0) or 0)
                out_t = int(a_usage.get("output_tokens", 0) or 0)
                ctx_used = in_t + cr_t + cc_t
                # Per-turn accumulation into the global stats. We do this
                # here (not in ResultMessage) because ResultMessage.usage
                # is the cumulative-per-session value and would inflate
                # _stats quadratically on long sessions.
                _stats["total_input_tokens"]           += in_t
                _stats["total_output_tokens"]          += out_t
                _stats["total_cache_read_tokens"]      += cr_t
                _stats["total_cache_creation_tokens"]  += cc_t
                sess_u = _session_usage.setdefault(session_id, {
                    "input_tokens": 0, "output_tokens": 0,
                    "cache_read_tokens": 0, "cache_creation_tokens": 0,
                    "total_cost_usd": 0.0, "last_turn_at": 0.0,
                    "context_used": 0, "context_used_pct": 0.0,
                    "context_limit": 0,
                })
                _session_usage_turns[session_id] = broadcast.turn_id
                capability = await _detect_gateway_context_capability(
                    model_to_use)
                details = _context_limit_details(
                    model_to_use,
                    stored=_positive_int(sess_u.get("context_limit")),
                    capability=capability,
                )
                limit = _positive_int(details.get("context_limit"))
                sess_u["input_tokens"] = in_t
                sess_u["cache_read_tokens"] = cr_t
                sess_u["cache_creation_tokens"] = cc_t
                sess_u["output_tokens"] = out_t
                sess_u["context_used"] = ctx_used
                sess_u["context_used_pct"] = (
                    round(ctx_used / limit * 100, 1) if limit else 0.0)
                _apply_context_limit_details(sess_u, details)
                # Provider response usage is the exact billed/request usage;
                # unlike SDK category accounting it is not a tokenizer guess.
                _mark_context_used(
                    sess_u, "provider_usage", estimate=False)
            for block in msg.content:
                if isinstance(block, TextBlock):
                    # Defensive tail-emit (see message_parser.py:279-290 — SDK
                    # forwards CLI stream events 1:1 in theory, but FE was
                    # observed truncating mid-word "CSS 变量切" 2026-05-18).
                    # Diagnostic log only fires when diff > 0 (no spam).
                    full = (getattr(block, "text", "") or "")
                    # Materialize the per-bubble mirror once for the prefix
                    # checks below (cheap: once per AssistantMessage, not per
                    # token).
                    streamed_str = "".join(streamed_in_bubble)
                    if full and full != streamed_str:
                        tail = (full[len(streamed_str):]
                                 if full.startswith(streamed_str)
                                 else full)
                        if tail:
                            assistant_acc.append(tail)
                            streamed_in_bubble.append(tail)
                            yield {"event": "text",
                                   "data": json.dumps({"text": tail})}
                elif isinstance(block, ThinkingBlock):
                    # Already streamed via thinking_delta events.
                    pass
                elif isinstance(block, ToolUseBlock):
                    if block.id:
                        tool_use_names[block.id] = block.name or ""
                        # Stash the Bash `description` so a following bg-launch
                        # tool_result (run_in_background=true) can label the
                        # inflight task. Harmless for non-bg Bash calls.
                        if block.name == "Bash":
                            _bi = getattr(block, "input", None) or {}
                            _bdesc = (_bi.get("description")
                                      if isinstance(_bi, dict) else None)
                            if _bdesc:
                                bg_launch_desc[block.id] = _bdesc
                    yield {"event": "tool_use",
                           "data": json.dumps(_render_tool_use(block))}
                    # FE closeAsst()'s the bubble on tool_use; reset mirror.
                    streamed_in_bubble = []
                elif isinstance(block, ToolResultBlock):
                    tu_id = getattr(block, "tool_use_id", "") or ""
                    tname = tool_use_names.get(tu_id, "")
                    yield {"event": "tool_result",
                           "data": json.dumps(
                               _render_tool_result(block, tool_name=tname))}

        async def _handle_user_message(msg):
            """SDK emits a `UserMessage` after every tool the agent ran —
            its `content` list carries `ToolResultBlock`s. Without this
            handler the result of every Read/Bash/Edit/etc. was silently
            dropped on the floor (the FE only ever saw the `tool_use`
            half of the round trip). 2026-05-22 audit fix.

            tool_use_id matches the prior ToolUseBlock; we look up its
            name from `tool_use_names` so the FE renderer (Bash terminal,
            Read gutter, …) can pick the right per-tool view."""
            # FALLBACK background-task completion path. The supported
            # contract is the typed TaskNotificationMessage (handled by
            # _handle_task_message — Phase-0 probe 2026-06-11 confirmed CLI
            # 2.1.141 delivers it). Some CLI builds additionally/instead
            # round-trip the terminal completion as a UserMessage whose
            # content is the raw <task-notification> XML; consume it here so
            # the card still flips and the bubble never renders raw XML.
            # _settle_background_task dedups the two paths — whichever
            # observes the terminal signal first wins, the loser no-ops.
            _notif_text = _usermsg_task_notification_text(msg)
            _notifs = _parse_task_notifications(_notif_text) if _notif_text else []
            if _notifs:
                for n in _notifs:
                    tid = n.get("task_id") or ""
                    # Dedup against the typed path / cross-turn watcher: only
                    # the path that settles first surfaces the completion
                    # (sync check, no await between gate and emit → no
                    # double-fire).
                    if tid and not await _on_task_settled_owned(
                            session_id, tid, status=n.get("status") or None,
                            tool_use_id=n.get("tool_use_id") or None,
                            summary=n.get("summary") or None,
                            output_file=n.get("output_file") or None):
                        continue
                    sys.stderr.write(
                        f"[chat] task fallback: in-turn settle via "
                        f"<task-notification> XML, typed message missed "
                        f"sid={session_id[:8]} task={obs.short_id(tid)}\n")
                    inflight_tasks.pop(tid, None)
                    yield {"event": "task_notification", "data": json.dumps({
                        "task_id": tid,
                        "tool_use_id": n.get("tool_use_id") or None,
                        "status": n.get("status") or None,
                        "summary": n.get("summary") or None,
                        "output_file": n.get("output_file") or None,
                    })}
                return
            content = getattr(msg, "content", None)
            if isinstance(content, list):
                for block in content:
                    if isinstance(block, ToolResultBlock):
                        tu_id = getattr(block, "tool_use_id", "") or ""
                        tname = tool_use_names.get(tu_id, "")
                        rendered = _render_tool_result(block, tool_name=tname)
                        yield {"event": "tool_result",
                               "data": json.dumps(rendered)}
                        # FALLBACK launch detection. The supported contract is
                        # the typed TaskStartedMessage (Phase-0 probe
                        # 2026-06-11: CLI 2.1.141 emits it BEFORE this
                        # tool_result, so `tid in inflight_tasks` already holds
                        # and this sniff no-ops). Kept for older CLIs / a CLI
                        # that changes ordering: without it the turn-end
                        # watcher never spawns (see _parse_bg_launch).
                        launch = _parse_bg_launch(rendered.get("text") or "")
                        if launch and tu_id:
                            tid = launch["task_id"]
                            if tid and tid not in inflight_tasks:
                                sys.stderr.write(
                                    f"[chat] task fallback: bg launch detected "
                                    f"via tool_result sniff, TaskStartedMessage "
                                    f"missed sid={session_id[:8]} "
                                    f"task={obs.short_id(tid)}\n")
                                desc = bg_launch_desc.get(tu_id)
                                recorded = await _record_background_task_launch_owned(
                                    session_id, tid,
                                    tool_use_id=tu_id,
                                    description=desc,
                                    output_file=launch.get("output_file"))
                                if not recorded:
                                    continue
                                inflight_tasks[tid] = {
                                    "tool_use_id": tu_id,
                                    "description": desc,
                                }
                                _pin_background_task(session_id, tid)
                                # Stamp the launching card ⏳ running live, the
                                # same shape _handle_task_message emits for a
                                # typed TaskStartedMessage.
                                yield {"event": "task_started",
                                       "data": json.dumps({
                                    "task_id": tid,
                                    "tool_use_id": tu_id,
                                    "description": desc,
                                    "task_type": "bash_background",
                                })}

        async def _handle_task_message(msg):
            """SDK-native background-task lifecycle (Agent / Bash with
            run_in_background=true). We CONSUME the SDK's existing Task*
            protocol verbatim — no shadow turn, no polling, no parsing of
            output_file. Every field is read with getattr defaults so a
            future SDK adding/renaming fields degrades gracefully instead
            of crashing the turn (see docs/background-tasks-spec.md §"长期
            跟上 SDK 的三条硬纪律").

            TaskStarted  → card flips to ⏳ running, recorded in inflight_tasks
            TaskProgress → periodic usage tick
            TaskNotification (status completed/failed/stopped) → terminal;
                clears inflight_tasks, carries summary + output_file so the FE
                can offer an "open result" link via the existing openFile path.
            TaskUpdated with a terminal patch → the same terminal path. This is
                required because newer CLIs may omit TaskNotification entirely.

            Note: the probe (§3.4) showed the terminal TaskNotification
            usually arrives AFTER this turn's ResultMessage, so within a turn
            this handler mostly emits task_started/progress; the cross-turn
            watcher (Phase 2) delivers the terminal notification. A task that
            finishes fast enough to terminate in-turn is handled right here.
            """
            if isinstance(msg, TaskStartedMessage):
                tid = getattr(msg, "task_id", "") or ""
                info = {
                    "tool_use_id": getattr(msg, "tool_use_id", None),
                    "description": getattr(msg, "description", None),
                }
                if tid:
                    accepted_start = await _record_background_task_launch_owned(
                        session_id, tid,
                        tool_use_id=info["tool_use_id"],
                        description=info["description"])
                    if not accepted_start:
                        return
                    inflight_tasks[tid] = info
                    # Pin the originating client from the moment the task
                    # starts: disconnect() kills the CLI subprocess and would
                    # abort the running task. The pin stays until the terminal
                    # notification settles (in-turn here, or the cross-turn
                    # watcher). Mid-turn this is redundant with _active_turns'
                    # eviction exemption, but it's what survives past turn end.
                    _pin_background_task(session_id, tid)
                yield {"event": "task_started", "data": json.dumps({
                    "task_id": tid,
                    "tool_use_id": info["tool_use_id"],
                    "description": info["description"],
                    "task_type": getattr(msg, "task_type", None),
                })}
            elif isinstance(msg, TaskProgressMessage):
                yield {"event": "task_progress", "data": json.dumps({
                    "task_id": getattr(msg, "task_id", "") or "",
                    "tool_use_id": getattr(msg, "tool_use_id", None),
                    "last_tool_name": getattr(msg, "last_tool_name", None),
                    # TaskUsage is a TypedDict (plain dict) → JSON-safe as-is.
                    "usage": dict(getattr(msg, "usage", None) or {}),
                })}
            elif isinstance(msg, TaskNotificationMessage):
                tid = getattr(msg, "task_id", "") or ""
                status = getattr(msg, "status", None)
                summary = getattr(msg, "summary", None)
                output_file = getattr(msg, "output_file", None)
                # In-turn settle (the rare case where a background task finishes
                # before this turn's ResultMessage): unpin + dedup against the
                # TaskUpdated path. This MUST run BEFORE the yield below.
                # 2026-08-04: it used to run after, so this branch had no gate
                # at all while the TaskUpdatedMessage branch did — every task
                # whose terminal transition arrived as BOTH a TaskUpdated patch
                # and a typed notification emitted task_notification TWICE
                # (measured: 1750 of 6193 in-turn settles, i.e. ~28%), which the
                # frontend rendered as two "后台任务已完成" toasts per task.
                # `settled` False means the TaskUpdated patch already reported
                # this exact transition; the event still goes out (its typed
                # payload can carry a summary/output_file the patch lacked, and
                # the card merge is idempotent) but is flagged so the client
                # treats it as a card patch rather than a fresh notification.
                settled = await _on_task_settled_owned(
                    session_id, tid, status=status,
                    tool_use_id=getattr(msg, "tool_use_id", None),
                    summary=summary,
                    output_file=output_file,
                    usage=dict(getattr(msg, "usage", None) or {}))
                if settled is None:
                    return
                # Drop from the per-turn in-flight set only after this runtime
                # has proved ownership. A successor's synthetic terminal must
                # not mutate the predecessor's live bookkeeping.
                inflight_tasks.pop(tid, None)
                yield {"event": "task_notification", "data": json.dumps({
                    "task_id": tid,
                    "tool_use_id": getattr(msg, "tool_use_id", None),
                    "status": status,
                    "summary": summary,
                    "output_file": output_file,
                    "usage": dict(getattr(msg, "usage", None) or {}),
                    "background_tasks_pending": len(
                        _sessions_with_inflight_tasks.get(session_id, ())),
                    "already_reported": not settled,
                })}
            elif isinstance(msg, TaskUpdatedMessage):
                terminal = _terminal_task_update(msg)
                if terminal is None:
                    return
                tid = terminal["task_id"]
                won_updated = await _on_task_settled_owned(
                    session_id, tid, status=terminal["status"],
                    tool_use_id=terminal.get("tool_use_id"),
                    summary=terminal.get("summary"),
                    output_file=terminal.get("output_file"),
                    usage=dict(terminal.get("usage") or {}))
                if won_updated is None:
                    return
                inflight_tasks.pop(tid, None)
                if won_updated:
                    terminal["background_tasks_pending"] = len(
                        _sessions_with_inflight_tasks.get(session_id, ()))
                    yield {
                        "event": "task_notification",
                        "data": json.dumps(terminal),
                    }

        async def _handle_rate_limit(msg):
            """SDK RateLimitEvent → record the window's RateLimitInfo and emit a
            `rate_limit` SSE event. Runs inside the detached event_gen task, so
            the store is updated even with no live subscriber; a later GET
            /api/chat/rate-limit returns the snapshot."""
            info = getattr(msg, "rate_limit_info", None)
            if info is None:
                return
            payload = _record_rate_limit(info)
            yield {"event": "rate_limit", "data": json.dumps(payload)}

        async def _handle_result_message(msg):
            """ResultMessage = turn complete. Update cumulative cost / stats,
            yield the consolidated 'done' SSE event the FE awaits, then finish
            slower per-message annotations and session metadata bookkeeping."""
            nonlocal memory_outcome_scheduled, assistant_acc, streamed_in_bubble
            if broadcast.perf_query_started:
                broadcast.perf_result_ms = obs.elapsed_ms(
                    broadcast.perf_query_started)
            _terminal_reason = sdk_lifecycle.normalize_terminal_reason(
                getattr(msg, "terminal_reason", None))
            _turn_origin = sdk_lifecycle.normalize_origin(
                getattr(msg, "origin", None))
            _model_usage = sdk_lifecycle.normalize_model_usage(
                getattr(msg, "model_usage", None))
            cost = getattr(msg, "total_cost_usd", None) or 0.0
            u = getattr(msg, "usage", {}) or {}
            # ResultMessage.usage is CUMULATIVE per session. Per-turn
            # token accumulation into _stats happens in
            # _handle_assistant_message; here we only record the
            # cumulative numbers for the SSE "done" payload (FE reads
            # them as a snapshot). Cost is per-turn (not cumulative),
            # so it's safe to += into _stats.
            in_t = int(u.get("input_tokens", 0) or 0)
            out_t = int(u.get("output_tokens", 0) or 0)
            cr_t = int(u.get("cache_read_input_tokens", 0)
                        or u.get("cache_read_tokens", 0) or 0)
            cc_t = int(u.get("cache_creation_input_tokens", 0)
                        or u.get("cache_creation_tokens", 0) or 0)
            _stats["total_cost_usd"] += cost
            _stats["total_messages"] += 1
            sess_u = _session_usage.setdefault(session_id, {
                "input_tokens": 0, "output_tokens": 0,
                "cache_read_tokens": 0, "cache_creation_tokens": 0,
                "total_cost_usd": 0.0, "last_turn_at": 0.0,
                "context_used": 0, "context_used_pct": 0.0,
                "context_limit": 0,
            })
            _session_usage_turns[session_id] = broadcast.turn_id
            sess_u["total_cost_usd"] += cost
            sess_u["last_turn_at"] = time.time()

            # ResultMessage is the SDK's authoritative turn boundary. Do the
            # small status/error accounting needed by the browser immediately,
            # close the global activity row, and emit `done` BEFORE slower
            # bookkeeping below (context control calls, transcript parsing,
            # push fan-out, JSONL compatibility cleanup). The old ordering kept
            # the footer pulsing "Running" for seconds after the final answer
            # was already visible.
            # ``interrupt()`` marks the exact live broadcast before awaiting the
            # SDK control request.  Treat that turn-owned bit as authoritative:
            # ResultMessage can race the later session-level bookkeeping, and
            # overwriting it with ``False`` used to turn a user Stop into a
            # completed/failed result (or recover Result-only prose after Stop).
            interrupt_key = (session_id, broadcast.turn_id)
            legacy_interrupt_key = (session_id, "")
            was_cancelled = bool(
                broadcast.cancelled
                or interrupt_key in _pending_interrupts
                or legacy_interrupt_key in _pending_interrupts
            )
            was_cancelled = (
                sdk_lifecycle.terminal_status(
                    _terminal_reason, cancelled=was_cancelled)
                == "cancelled"
            )
            _pending_interrupts.discard(interrupt_key)
            _pending_interrupts.discard(legacy_interrupt_key)
            broadcast.cancelled = was_cancelled
            _completed_at_ms = int(time.time() * 1000)
            _msg_duration_ms = getattr(msg, "duration_ms", None)
            _elapsed_s = (round(_msg_duration_ms / 1000, 1)
                          if _msg_duration_ms else None)
            _result_error = _sdk_result_error(msg)
            _turn_error = _merge_sdk_errors([
                *turn_sdk_errors,
                *([_result_error] if _result_error else []),
            ])
            _result_recovered = False
            _result_snapshot_ready = False
            _result_text = str(getattr(msg, "result", None) or "").strip()
            _ended_after_tool = (
                bool(tool_use_names)
                and not "".join(streamed_in_bubble).strip()
                and not inflight_tasks
                and not _sessions_with_inflight_tasks.get(session_id)
                and bool(prompt.strip())
                and not prompt.lstrip().startswith("/")
            )
            if (_turn_error is None and not was_cancelled
                    and _ended_after_tool):
                if _result_text:
                    # Some third-party runtimes put the final answer only on the
                    # terminal ResultMessage after their last tool call. Restore
                    # it to the ordinary text stream before `done`; the private
                    # snapshot below makes the same answer survive a refresh.
                    assistant_acc.append(_result_text)
                    streamed_in_bubble.append(_result_text)
                    _result_recovered = True
                    broadcast.perf_final_mode = "result_recovered"
                    yield {
                        "event": "text",
                        "data": json.dumps({"text": _result_text}),
                    }
                else:
                    broadcast.perf_final_mode = "missing"
                    _turn_error = {
                        "message": (
                            "The turn ended after tool use without a final "
                            "assistant response."
                        ),
                        "source": "missing_final_response",
                        "api_error_status": None,
                    }

            # A success-shaped ResultMessage proves only that the SDK operation
            # ended. Ordinary chat turns are committed only when this exact
            # pre-query boundary owns both a real user row and an assistant row.
            # Slash commands have their own system/result transcript shape, so
            # they retain the SDK Result verdict. Error/cancel turns only need a
            # user UUID when attachments must be annotated.
            boundary_asst_uuid: str | None = None
            new_user_uuid: str | None = None
            boundary_inspected = False
            _commit_required = bool(prompt.strip()) and not prompt.lstrip().startswith("/")
            if ((_turn_error is None and not was_cancelled and _commit_required)
                    or persisted_imgs or persisted_docs):
                boundary_asst_uuid, new_user_uuid, boundary_inspected = (
                    await _settle_turn_uuids(
                        session_id,
                        broadcast.transcript_boundary,
                        started_at_ms=int(float(broadcast.started_at or 0) * 1000),
                        terminal_at_ms=_completed_at_ms,
                        require_assistant=(
                            _turn_error is None
                            and not was_cancelled
                            and _commit_required
                        ),
                    )
                )
            if (_turn_error is None and not was_cancelled and _commit_required
                    and (not boundary_asst_uuid or not new_user_uuid)):
                _turn_error = await asyncio.to_thread(
                    _turn_prevented_error_from_boundary,
                    session_id,
                    broadcast.transcript_boundary,
                )
                if _turn_error is None:
                    detail = (
                        "Canonical conversation history could not be inspected."
                        if not boundary_inspected else
                        "The turn ended before the user message and assistant "
                        "response were committed to conversation history."
                    )
                    _turn_error = {
                        "message": detail,
                        "source": "canonical_commit",
                        "api_error_status": None,
                    }
            _done_memory_recall = mem0.pop_recall_trace(session_id)
            _done_memory_receipt = _persistable_memory_recall(
                _done_memory_recall)
            if _turn_error is None and _result_recovered:
                try:
                    _result_snapshot_ready = await asyncio.to_thread(
                        _persist_completed_result_snapshot,
                        broadcast,
                        _result_text,
                        terminal_at_ms=_completed_at_ms,
                        elapsed_s=_elapsed_s,
                        memory_recall=_done_memory_receipt,
                        terminal_reason=_terminal_reason,
                        turn_origin=_turn_origin,
                        turn_id=broadcast.turn_id,
                        model_usage=_model_usage,
                    )
                except Exception as exc:
                    _safe_secondary_diagnostic(
                        "result_snapshot", session_id, exc)
                if not _result_snapshot_ready:
                    _turn_error = {
                        "message": (
                            "The final response was received but could not be "
                            "saved to conversation history."
                        ),
                        "source": "result_snapshot",
                        "api_error_status": None,
                    }
            terminal_assistant_uuid = (
                "" if _result_snapshot_ready
                else last_assistant_uuid or boundary_asst_uuid or ""
            )

            _is_error = _turn_error is not None
            _subtype = getattr(msg, "subtype", None)
            _errors = getattr(msg, "errors", None) or []
            if not isinstance(_errors, (list, tuple)):
                _errors = [_errors]
            _api_error_status = (
                (_turn_error or {}).get("api_error_status")
                or getattr(msg, "api_error_status", None)
            )
            _error_message = (_turn_error or {}).get("message", "")
            _error_class = _classify_stream_error(_error_message) if _is_error else {
                "kind": None, "cta": None, "retryable": False,
            }
            _turn_status = sdk_lifecycle.terminal_status(
                _terminal_reason,
                is_error=_is_error,
                cancelled=was_cancelled,
            )
            was_cancelled = _turn_status == "cancelled"
            broadcast.cancelled = was_cancelled
            _activity_status = (
                "failed" if _turn_status == "stopped" else _turn_status)
            broadcast.perf_status = _activity_status
            broadcast.perf_error_kind = (
                str(_error_class.get("kind") or "unknown")
                if _is_error else "none"
            )
            _done_session_usage = dict(sess_u)
            # A failed turn needs a durable copy of the terminal error bubble,
            # including when a legitimate partial AssistantMessage precedes
            # the Result error. Commit it before `done` so a quiet/hard reload
            # cannot collapse the live partial+error pair into bare JSONL.
            _failed_snapshot_ready = False
            if _is_error and not was_cancelled:
                try:
                    _failed_snapshot_ready = await asyncio.to_thread(
                        _persist_failed_turn_snapshot,
                        broadcast,
                        _error_message,
                        terminal_at_ms=_completed_at_ms,
                        elapsed_s=_elapsed_s,
                        memory_recall=_done_memory_receipt,
                        canonical_terminal_published=True,
                        terminal_status=_turn_status,
                        terminal_reason=_terminal_reason,
                        turn_origin=_turn_origin,
                        model_usage=_model_usage,
                    )
                except Exception as exc:
                    _safe_secondary_diagnostic(
                        "terminal_snapshot", session_id, exc)
            # Transfer ownership before publishing the terminal event: canonical
            # success owns the prompt directly; a failed turn is safe to clear
            # only after its display snapshot committed. If snapshot persistence
            # failed, retain the sidecar for restart recovery.
            if ((not _is_error and not was_cancelled)
                    or (_is_error and _failed_snapshot_ready)):
                await asyncio.to_thread(
                    _delete_active_turn_sidecar, session_id)
            _done_background_tasks = len(
                _merge_session_inflight(session_id, inflight_tasks)
            )
            broadcast.perf_background_count = _done_background_tasks
            # Activity Center tracks whether the human is still waiting for the
            # main response, not whether detached process work remains. Settle it
            # at the authoritative Result boundary; background task cards and
            # runtime overlays continue independently inside the conversation.
            _done_activity_source = broadcast.activity_source
            await _finish_activity(
                session_id, broadcast, _activity_status)
            if _done_background_tasks and terminal_assistant_uuid:
                try:
                    sess.set_runtime_background_boundary(
                        session_id, terminal_assistant_uuid)
                except Exception as exc:
                    sys.stderr.write(
                        f"[chat] runtime boundary persist failed "
                        f"sid={session_id[:8]} "
                        f"exc={type(exc).__name__}\n"
                    )
            # Capture once at the SDK's authoritative terminal boundary. The
            # exact assistant UUID is already known from AssistantMessage, so
            # persist the footer BEFORE yielding done.  A hard refresh or a new
            # turn can otherwise arrive while the slower context/transcript
            # bookkeeping below is still running and observe a bare assistant
            # record with no status/model/duration.
            _footer_annotation_uuid = ""
            _done_usage_source = _usage_source_signature(
                _find_session_jsonl(session_id))
            _done_usage_summary = _session_usage_summary_payload(
                _done_session_usage,
                turn_id=broadcast.turn_id,
                source=_done_usage_source,
            )
            try:
                if terminal_assistant_uuid:
                    await obs.to_thread_io(
                        "chat.sidecar_terminal_write",
                        session_id,
                        sess.set_terminal_annotation_and_usage,
                        session_id,
                        terminal_assistant_uuid,
                        _done_usage_summary,
                        cost=f"${cost:.4f}",
                        model=model_to_use,
                        ts=_completed_at_ms,
                        turn_status=_turn_status,
                        terminal_reason=_terminal_reason,
                        turn_origin=_turn_origin,
                        model_usage=_model_usage,
                        elapsed_s=_elapsed_s,
                        memory_recall=_done_memory_receipt,
                        file_path=sess._sidecar_path(session_id),
                    )
                    _footer_annotation_uuid = terminal_assistant_uuid
                elif _done_usage_summary is not None:
                    await obs.to_thread_io(
                        "chat.sidecar_terminal_usage_write",
                        session_id,
                        sess.set_session_usage_summary,
                        session_id,
                        _done_usage_summary,
                        file_path=sess._sidecar_path(session_id),
                    )
            except Exception as exc:
                sys.stderr.write(
                    f"[chat] terminal sidecar write failed "
                    f"sid={session_id[:8]} exc={type(exc).__name__}\n")
                sys.stderr.flush()
            yield {"event": "done", "data": json.dumps({
                "turn_id": broadcast.turn_id,
                "duration_ms": _msg_duration_ms,
                "assistant_uuid": terminal_assistant_uuid,
                "completed_at_ms": _completed_at_ms,
                "total_cost_usd": cost,
                "model": model_to_use,
                "stats": _stats,
                "cancelled": was_cancelled,
                "is_error": _is_error,
                "status": _turn_status,
                "terminal_reason": _terminal_reason,
                "origin": _turn_origin,
                "model_usage": _model_usage,
                "error": _error_message,
                "kind": _error_class["kind"],
                "cta": _error_class["cta"],
                "retryable": _error_class["retryable"],
                "result_subtype": _subtype,
                "errors": (_dedupe_error_parts([*_errors, _error_message])
                           if _is_error else []),
                "api_error_status": _api_error_status,
                "turn_usage": {
                    "input_tokens": in_t,
                    "output_tokens": out_t,
                    "cache_read_tokens": cr_t,
                    "cache_creation_tokens": cc_t,
                },
                # Context-control refresh continues below. Assistant-message
                # usage has already updated this snapshot; a later API read
                # sees the authoritative postprocessed value.
                "session_usage": _done_session_usage,
                "budget_usd": _budget_usd(),
                "budget_used_pct": (
                    round(_stats["total_cost_usd"] / _budget_usd() * 100, 1)
                    if _budget_usd() > 0 else 0
                ),
                "memory_recall": _done_memory_recall,
                "snapshot_ready": (
                    _failed_snapshot_ready or _result_snapshot_ready),
                "result_recovered": _result_snapshot_ready,
                "background_tasks_pending": _done_background_tasks,
                "activity_source": _done_activity_source,
            })}
            broadcast.perf_post_started = obs.monotonic()

            # Everything below is post-turn persistence/telemetry. It remains
            # inside the detached pump so browser disconnects cannot cancel it,
            # but it no longer delays the user-visible terminal event.
            # Pull authoritative max-window from SDK so the meter reflects the
            # ACTUAL effective limit (which may be 1M for Pro/Max subscribers,
            # not the hardcoded 200K in MODEL_CONTEXT_LIMITS). One control
            # round-trip per turn — small price for accurate denominator.
            #
            # Third-party caveat: SDK category accounting uses Claude CLI's
            # tokenizer. Keep provider response usage as the completed-turn
            # numerator. For Codex, the denominator comes from CLIProxyAPI's
            # live catalog; other providers retain their existing SDK/table
            # fallback. The raw SDK figures are diagnostic metadata only.
            pre_probe_usage = sess_u
            sess_u = dict(sess_u)
            if endpoints.is_third_party(model_to_use):
                # Re-anchor context_limit to the runtime effective limit, not the
                # optimistic catalog value. For Codex Gateway this prevents the UI
                # from showing e.g. 30% while the sidecar/backend is already near
                # its real context ceiling.
                sdk_max = sdk_raw = sdk_threshold = sdk_total = 0
                try:
                    cu = await client.get_context_usage()
                    sdk_max = _positive_int(cu.get("maxTokens"))
                    sdk_raw = _positive_int(cu.get("rawMaxTokens"))
                    sdk_threshold = _positive_int(cu.get("autoCompactThreshold"))
                    sdk_total = _positive_int(cu.get("totalTokens"))
                except Exception as _e:
                    sys.stderr.write(
                        f"[chat-stream] third-party get_context_usage skipped for "
                        f"sid={session_id[:8]} exc={type(_e).__name__}\n")
                capability = await _detect_gateway_context_capability(
                    model_to_use)
                details = _context_limit_details(
                    model_to_use,
                    sdk_max=sdk_max,
                    sdk_raw=sdk_raw,
                    stored=_positive_int(sess_u.get("context_limit")),
                    capability=capability,
                )
                _apply_context_limit_details(sess_u, details)
                sess_u["sdk_context_max_tokens"] = sdk_max
                sess_u["sdk_context_raw_max_tokens"] = sdk_raw
                threshold = _compact_threshold(
                    model_to_use,
                    _positive_int(details.get("context_limit")),
                    sdk_threshold,
                    sdk_max=sdk_max,
                    capability=capability,
                )
                if threshold:
                    sess_u["auto_compact_threshold"] = threshold
                if not sess_u.get("context_used") and sdk_total:
                    sess_u["context_used"] = sdk_total
                    _mark_context_used(
                        sess_u, "sdk_context", estimate=True)
                # Recompute pct against the corrected limit.
                if sess_u["context_limit"]:
                    sess_u["context_used_pct"] = round(
                        sess_u.get("context_used", 0)
                        / sess_u["context_limit"] * 100, 1)
            else:
                try:
                    cu = await client.get_context_usage()
                    real_max = int(cu.get("maxTokens") or 0)
                    real_total = int(cu.get("totalTokens") or 0)
                    if real_max:
                        sess_u["context_limit"] = real_max
                        # Persist so the meter shows the correct denominator
                        # after a restart / on cold tab switches without
                        # needing a live client (see get_session_ctx_window).
                        try:
                            sess.set_session_ctx_window(session_id, real_max)
                        except Exception:
                            pass
                    if real_total:
                        sess_u["context_used"] = real_total
                        _mark_context_used(
                            sess_u, "sdk_context", estimate=False)
                    if real_max and real_total:
                        sess_u["context_used_pct"] = round(
                            real_total / real_max * 100, 1)
                except Exception as _e:
                    sys.stderr.write(
                        f"[chat-stream] get_context_usage skipped for "
                        f"sid={session_id[:8]} exc={type(_e).__name__}\n")

            refined_usage = sess_u
            if _refine_session_usage_for_turn(
                    session_id, broadcast.turn_id, refined_usage):
                refined_source = _usage_source_signature(
                    _find_session_jsonl(session_id))
                await obs.to_thread_io(
                    "chat.usage_summary_refine",
                    session_id,
                    _persist_session_usage_summary,
                    session_id,
                    refined_usage,
                    turn_id=broadcast.turn_id,
                    source=refined_source,
                    require_matching_turn=True,
                    file_path=lambda: sess._sidecar_path(session_id),
                )
            # A successor may have completed during the probe. Keep later
            # bookkeeping on the currently-authoritative snapshot, not this copy.
            sess_u = _session_usage.get(session_id, pre_probe_usage)

            # Sidecar annotations: resolve UUIDs from the exact pre-query
            # transcript coordinate. A simple "latest assistant" tail scan is
            # unsafe for Result-only Gateway errors: it would select the prior
            # turn and overwrite its footer with this turn's failed state.
            # all_msgs holds the full transcript parse. The fast UUID path
            # below may skip the parse, but the count/auto-rename block further
            # down (message_count, turn_count, first_user_text) still needs the
            # full list — so it's lazily loaded once and reused there. Starts as
            # None ("not yet parsed") rather than being scoped inside the
            # fallback branch, which previously left it unbound on the fast path
            # → UnboundLocalError at every turn's end.
            all_msgs: list | None = None
            # Reuse the commit-gate result captured before `done`; a second index
            # lookup would add latency and could observe a different transcript
            # generation than the one that decided the terminal status.
            # AssistantMessage observed in this stream is authoritative. The
            # indexed boundary is a safe fallback for older SDKs that omit the
            # UUID on the in-memory object but still write it to JSONL.
            new_asst_uuid = terminal_assistant_uuid
            if new_asst_uuid and new_asst_uuid != _footer_annotation_uuid:
                # ts (ms epoch) stamps the turn's completion time. The
                # frontend's turn-footer (.turn-footer in index.html)
                # reads it via fmtHM() → "HH:MM" under the last muse
                # block of the turn. Stored at ms granularity to match
                # JS Date.now() (the frontend writes the same ts onto
                # in-flight messages in _markDone; loading from sidecar
                # uses this one).
                # elapsed_s = SDK-reported wall-clock for the turn (in
                # seconds). Persisted so reloading a session keeps the
                # "13:42 · 2m50s" footer (FE-side stamping only survives
                # within the live session). None when SDK didn't fill
                # duration_ms.
                await obs.to_thread_io(
                    "chat.sidecar_footer_write",
                    session_id,
                    sess.set_message_annotation,
                    session_id,
                    new_asst_uuid,
                    cost=f"${cost:.4f}",
                    model=model_to_use,
                    ts=_completed_at_ms,
                    turn_status=_turn_status,
                    terminal_reason=_terminal_reason,
                    turn_origin=_turn_origin,
                    turn_id=broadcast.turn_id,
                    model_usage=_model_usage,
                    elapsed_s=_elapsed_s,
                    memory_recall=_done_memory_receipt,
                    file_path=sess._sidecar_path(session_id),
                )
            if new_user_uuid and (persisted_imgs or persisted_docs):
                await obs.to_thread_io(
                    "chat.sidecar_attachment_write",
                    session_id,
                    sess.set_message_annotation,
                    session_id,
                    new_user_uuid,
                    images=persisted_imgs or None,
                    docs=persisted_docs or None,
                    file_path=sess._sidecar_path(session_id),
                )
            # mem0 write is deferred to AFTER we compute was_cancelled / _is_error
            # below — we must not distill a cancelled or errored turn into a
            # "durable fact". See the mem0.schedule_store(...) call further down.
            # message_count = total transcript size; auto-rename from first
            # user message text if session is still auto-named. These counts
            # need the full transcript, so parse it now if the fast UUID path
            # above already resolved both UUIDs and skipped the parse.
            if all_msgs is None:
                try:
                    all_msgs = await asyncio.to_thread(
                        _get_session_msgs, session_id, model_to_use)
                except Exception:
                    all_msgs = []
            first_user_text = ""
            for sm in all_msgs:
                if sm.type == "user":
                    c = (sm.message or {}).get("content")
                    if isinstance(c, str):
                        first_user_text = c
                    elif isinstance(c, list):
                        for b in c:
                            if isinstance(b, dict) and b.get("type") == "text":
                                first_user_text = b.get("text", "")
                                break
                    break
            # turn_count = real user prompts only. SDK's get_session_messages
            # claims to filter tool-use sidechain (parent_tool_use_id always
            # None) but actually returns *every* user-typed frame, including
            # the implicit ones that wrap tool_result blocks after an agent
            # tool call. We detect those by content shape: if every content
            # block is a tool_result, the frame is a sidechain echo, not a
            # real user message. Without this filter, a session with 45 real
            # prompts but heavy agent tool use shows up as 300+ turns.
            n_turns = sum(1 for sm in all_msgs if _is_real_user_prompt(sm))
            # Auto-rename source: prefer the first real user text. But an
            # image-only first turn carries the injected "(image)" placeholder
            # as its text — naming the session "(image)" looks broken. Drop the
            # placeholder and fall back to a friendly label so the session gets
            # a sensible auto-name (or stays auto-named for the next real text).
            _rename_src = first_user_text or prompt
            if _rename_src.strip() == _IMAGE_ONLY_PLACEHOLDER:
                _rename_src = "图片对话" if is_chinese_locale() else "Image chat"
            def _persist_session_summary() -> None:
                sess.bump_session(
                    session_id,
                    message_count=len(all_msgs),
                    turn_count=n_turns,
                    auto_rename_from=_rename_src,
                )
                sess.update_model(session_id, model_to_use)

            await obs.to_thread_io(
                "chat.session_index_write",
                session_id,
                _persist_session_summary,
                file_path=sess.INDEX,
                owned=True,
            )
            # ``done`` is intentionally published before this slower block. A
            # user can therefore roll over the session while these annotations
            # and the first-turn name are still being persisted. Reconcile the
            # completed source into every already-linked successor; the
            # endpoint performs the same reconciliation after linking to close
            # the opposite ordering of the race.
            await asyncio.to_thread(
                _sync_runtime_successor_postlude, session_id)
            # Web Push on turn-done. Three gates, in order:
            #   1. Turn was NOT user-cancelled — see was_cancelled above.
            #   2. No device has heartbeated /api/presence recently — i.e.
            #      the user is NOT actively at any screen. See below.
            #   3. Wrapped in try/except — push failure must never block
            #      the stream's done event.
            # (Previously also gated on MUSELAB_NOTIFY_NORMAL env var, but
            # the UI's 4-toggle notification panel collapsed to a single
            # "notify me" switch on 2026-05-28: subscription state IS the
            # on/off — no need for a per-class server-side mute.)
            #
            # History of the gating logic:
            #   - v1: gated on "no live SSE subscriber" → broke multi-device
            #     (desktop SSE alive ⇒ phone push suppressed too).
            #   - v2: removed server-side gate, moved decision to sw.js
            #     visibility check → fixed phone backgrounded case, but
            #     phone STILL rang while user was on desktop because each
            #     SW only sees its own device's clients.
            #   - v3 (now): server-side presence heartbeat. Frontend POSTs
            #     /api/presence every ~15s while visible; if any device
            #     reported in within GRACE_SECONDS, skip the fan-out
            #     entirely. The sw.js visibility check stays as
            #     defense-in-depth for the case where heartbeat data is
            #     stale (network blip, browser killed the timer, etc.).
            #
            # Body intentionally minimal: session name + "Muse 已回复". No
            # preview text — the actual reply is one tap away in chat.
            if not was_cancelled:
                # Notification delivery is not part of the turn transaction.
                # A broken proxy previously kept the completed broadcast in
                # ``_active_turns`` for 405s, so the UI stayed "running" and the
                # next queued prompt could not claim the session.  Schedule the
                # best-effort fan-out independently; push.py bounds each HTTP
                # attempt as a second line of defence.
                _notify_turn_done(
                    session_id, session_name=str(s.get("name") or ""))
            # Strip unverifiable thinking-block signatures so this session
            # stays resumable via `claude --resume` (and the official
            # Anthropic API). Third-party vendors (DeepSeek / GLM /
            # MiniMax / Kimi / Qwen / Baidu / Xiaomi MiMo) don't sign
            # their thinking output; Anthropic's resume API would 400 on
            # any of those blocks. We clean opportunistically every turn
            # — idempotent on already-clean files, so the cost is just
            # one stat + a parse of the small jsonl. See
            # backend/jsonl_cleanup.py for the full rationale + the
            # scripts/fix-thinking-signatures.py CLI for retroactive
            # cleanup of pre-existing sessions.
            # Only third-party vendors emit unsigned thinking blocks; pure
            # Claude-native turns always carry valid signatures, so the
            # cleanup would be a no-op — skip it. (A session that mixed
            # vendors gets cleaned on each vendor turn, so the Claude turns
            # never need to.) When we do clean, offload the synchronous
            # stat+parse to a thread so it can't block the event loop.
            # TOCTOU NOTE (audit O/401): we run this in the ResultMessage
            # handler, i.e. once the turn is logically complete, but the SDK's
            # CLI subprocess owns the JSONL and may still be flushing the final
            # assistant record when we read+atomic-rewrite it. Two outcomes are
            # possible if we lose that race: (a) we rewrite a copy that is
            # missing the still-unflushed last line — but that line carries the
            # very thinking block we want to strip, so the next turn's cleanup
            # (or the scripts/fix-thinking-signatures.py CLI) catches it,
            # because clean_jsonl is idempotent; (b) the CLI appends to the old
            # inode after our os.replace — POSIX keeps that write going to the
            # now-unlinked file and it's lost, but the CLI only appends BEFORE
            # ResultMessage, so by the time we're here that window is closed.
            # Net: worst case is a deferred strip, never data loss, so we don't
            # add flush-confirmation/locking (we can't coordinate with the SDK's
            # writer anyway). See clean_jsonl's atomic-write rationale.
            if endpoints.is_third_party(model_to_use):
                try:
                    from . import jsonl_cleanup as _jc
                    await asyncio.to_thread(_jc.clean_session, session_id)
                except Exception as e:
                    sys.stderr.write(
                        f"[chat] jsonl cleanup failed "
                        f"sid={session_id[:8]} "
                        f"exc={type(e).__name__}\n")
            # Memory write is deferred until turn status is known. Clean
            # exchanges may be consolidated into memories. Cancelled turns are
            # evidence-only (never fact candidates), which preserves the useful
            # failed trajectory without teaching the partial answer as truth.
            # Legacy Mem0 keeps its historical success-only behaviour.
            if sess.session_is_deleting(session_id):
                memory_outcome_scheduled = False
            elif mem0.enabled() and not was_cancelled and not _is_error:
                _asst_full = "".join(assistant_acc)
                if _asst_full.strip():
                    memory_outcome_scheduled = mem0.schedule_store(
                        session_id, model_to_use, prompt, _asst_full, new_asst_uuid)
            elif was_cancelled:
                memory_outcome_scheduled = mem0.schedule_cancelled(
                    session_id, prompt, new_asst_uuid)
            elif _is_error:
                memory_outcome_scheduled = mem0.schedule_failed(
                    session_id, model_to_use, prompt,
                    "".join(assistant_acc), _error_message, new_asst_uuid)
        # event_gen is now driven by a detached background task (see
        # stream endpoint below), so the SSE generator doesn't cancel
        # these workers when the browser disconnects — they complete
        # naturally. 30-minute hard cap is applied to the outer
        # task, not here.
        claude_task = asyncio.create_task(pump_claude())
        side_task = asyncio.create_task(pump_side_q(side_q))
        perm_task = asyncio.create_task(pump_side_q(perm_q))

        try:
            while True:
                kind, payload = await merge_q.get()
                if kind == "side":
                    # Already shaped as {"event": "...", "data": "..."} — pass through.
                    yield await _prepare_side_event(payload)
                    continue
                if kind == "cancelled":
                    async for side_event in _flush_side_channels():
                        yield side_event
                    yield {"event": "cancelled", "data": "{}"}
                    break
                if kind == "error":
                    async for side_event in _flush_side_channels():
                        yield side_event
                    # If the user interrupted this turn and the force-stop
                    # watchdog tore the CLI down, receive_response() raises a
                    # transport error that lands here. That's an expected
                    # consequence of the stop, not a real failure — surface it
                    # as a clean `cancelled` event so the FE doesn't flash a red
                    # error toast for a turn the user deliberately stopped.
                    if broadcast.cancelled:
                        yield {"event": "cancelled", "data": "{}"}
                    else:
                        yield _error_event(payload)
                    break
                if kind == "done":
                    async for side_event in _flush_side_channels():
                        yield side_event
                    break
                # kind == "claude" — dispatch by SDK message type to the
                # per-type helper async generators defined above. Each
                # helper yields zero-or-more SSE events; we forward them.
                msg = payload
                if chat_subagents.is_subagent_message(msg):
                    # A forwarded sidechain is never parent answer content.
                    # Route on the SDK-owned parent id even when an incomplete
                    # frame cannot be rendered, so it cannot fall through.
                    for record in subagent_mux.feed(msg):
                        yield {
                            "event": record["event"],
                            "data": json.dumps(record["data"]),
                        }
                elif isinstance(msg, StreamEvent):
                    async for ev in _handle_stream_event(msg):
                        yield ev
                elif isinstance(msg, AssistantMessage):
                    async for ev in _handle_assistant_message(msg):
                        yield ev
                elif isinstance(msg, UserMessage):
                    # SDK emits a UserMessage after every tool the agent
                    # invoked — its content carries the ToolResultBlocks.
                    # Without this dispatch, every Read/Bash/Edit result
                    # was silently dropped on the floor; the FE only saw
                    # the tool_use half of the round trip.
                    async for ev in _handle_user_message(msg):
                        yield ev
                elif isinstance(msg, (TaskStartedMessage, TaskProgressMessage,
                                      TaskNotificationMessage, TaskUpdatedMessage)):
                    # SDK-native background-task lifecycle. These are
                    # SystemMessage subclasses muselab used to silently drop;
                    # check them BEFORE any generic SystemMessage branch (none
                    # exists today) so they reach the task handler.
                    async for ev in _handle_task_message(msg):
                        yield ev
                elif isinstance(msg, RateLimitEvent):
                    # Pro/Max quota signal the SDK pushes on limit-state change.
                    # muselab used to silently drop it; capture into the per-
                    # window store + push a live `rate_limit` SSE event.
                    async for ev in _handle_rate_limit(msg):
                        yield ev
                elif isinstance(msg, SystemMessage):
                    # Terminal system signals were captured by the turn-scoped
                    # pump before enqueue; informational rows need no UI event.
                    pass
                elif isinstance(msg, ResultMessage):
                    async for ev in _handle_result_message(msg):
                        yield ev
            # Turn loop ended (done / in-band error). Hand any still-in-flight
            # background task to a detached cross-turn watcher that keeps the
            # client alive and drains its terminal notification after the turn
            # (probe §3.4: it lands after ResultMessage). On hard cancel we jump
            # to the except below and skip this — a cancelled turn doesn't spawn.
            #
            # Cover not just THIS turn's launches (inflight_tasks) but EVERY
            # unsettled task for the session (_merge_session_inflight); see its
            # docstring for the spec §13 orphan-bug rationale.
            merged_inflight = _merge_session_inflight(session_id, inflight_tasks)
            if merged_inflight:
                _spawn_task_watcher(
                    session_id,
                    client,
                    merged_inflight,
                    started_at=broadcast.started_at,
                    origin_turn_id=broadcast.turn_id,
                )
            else:
                await _retire_unpinned_task_watcher(session_id)
        except asyncio.CancelledError:
            # Hard cancel (task cancelled / 30-min timeout cancel) — mark so
            # the queue drain pauses rather than charging ahead.
            broadcast.cancelled = True
            yield {"event": "cancelled", "data": "{}"}
            raise
        finally:
            # event_gen runs as part of a detached background task now;
            # cleanup here runs after the task finishes naturally (or an
            # explicit interrupt / an armed MUSELAB_TURN_TIMEOUT_S cancels us).
            side_task.cancel()
            perm_task.cancel()
            claude_task.cancel()
            # A task cancelled before its coroutine starts has no finally, and
            # a task stuck in SDK code may ignore cancellation. Join the shared
            # attachment finalizer directly before this outer owner can exit.
            await _rollback_broadcast_attachments(broadcast)
            unregister_session_queue(session_id)
            if perm.unregister_session_queue(session_id):
                # Approval returned updatedPermissions but no terminal tool hook
                # arrived. The CLI's live mode is ambiguous; never reuse it.
                _pending_runtime_rebuilds.add(session_id)

    # Background-completion + reconnect-streaming design:
    #
    # Old: `event_gen()` was the SSE response generator directly.
    # Browser disconnect cancelled the generator, which cancelled
    # pump_claude, which cut off the SDK reply mid-stream.
    #
    # New: event_gen() runs as a DETACHED background task that publishes
    # every event it would have yielded into a per-session TurnBroadcast.
    # The HTTP response is just a subscriber that replays the buffer +
    # streams new events. A user closing their browser doesn't affect
    # the background task — it runs to completion (or until interrupted).
    # A second SSE request to the same session (reconnect) becomes
    # another subscriber and sees the full reply via replay + live tail.
    # Wall-clock cap on a whole turn. 0 (the default) disables it.
    #
    # This used to be a hard 1800s. A wall-clock cap is the wrong shape for an
    # agentic workspace: it can't tell "wedged" from "busy", so it kills turns
    # that are actively producing output. A full test suite, a migration or a
    # deep tool loop legitimately runs past 30 minutes, and the kill is total
    # loss — the SDK CLI only writes the JSONL on completion, so an aborted
    # turn leaves NO transcript record of its work or of why it stopped
    # (`_error_event` is live-only SSE; miss the moment and the reason is gone).
    #
    # Nothing depends on it as a safety net: POST /interrupt marks the turn
    # cancelled and schedules _force_stop_after_grace independently of this
    # timeout, so a genuinely wedged turn already has an escape hatch that
    # works whether or not a cap exists.
    #
    # Set MUSELAB_TURN_TIMEOUT_S to a positive number to re-arm it.
    BG_TIMEOUT_S = env_int("MUSELAB_TURN_TIMEOUT_S", 0, min_value=0)

    # `broadcast` was already reserved under _lock at the top of NEW-TURN
    # MODE to close the check+insert race. Fill its remaining fields now
    # that we've parsed prompt + attachments + resolved the actual model.
    broadcast.model = model_to_use
    broadcast.user_text = prompt
    broadcast.user_images = list(persisted_imgs)
    broadcast.user_docs = list(persisted_docs)
    # Recall traces are keyed by session for the SDK hook bridge. A prior turn
    # that died before ResultMessage may have left one behind; discard it at
    # the new turn boundary so it can never be attributed to this prompt.
    mem0.pop_recall_trace(session_id)
    # Refresh the pending intent with resolved model, attachment display refs,
    # and the exact pre-query transcript boundary captured above. Keep this
    # filesystem write off the event loop because the sidecar is durable state.
    _intent_refresh_started = obs.monotonic()
    try:
        persisted_intent_refresh = await obs.to_thread_io(
            "chat.active_turn_refresh",
            session_id,
            _write_active_turn_sidecar,
            broadcast,
            file_path=_active_turn_path(session_id),
            owned=True,
        )
    except asyncio.CancelledError:
        broadcast.perf_intent_refresh_ms = obs.elapsed_ms(
            _intent_refresh_started)
        broadcast.perf_startup_failure_phase = "context"
        if broadcast.cancelled:
            return await _finish_cancelled_startup(
                session_id, broadcast)
        await _abort_turn_startup(
            session_id,
            broadcast,
            "cancelled",
            pause_queue=True,
        )
        raise
    except Exception:
        broadcast.perf_intent_refresh_ms = obs.elapsed_ms(
            _intent_refresh_started)
        broadcast.perf_startup_failure_phase = "context"
        queue_settled = await _abort_turn_startup(
            session_id,
            broadcast,
            "failed",
            pause_queue=True,
            error_text="turn intent could not be refreshed",
        )
        raise _TurnStartError(
            "turn intent could not be refreshed",
            queue_claim_settled=queue_settled,
        ) from None
    broadcast.perf_intent_refresh_ms = obs.elapsed_ms(_intent_refresh_started)
    if not persisted_intent_refresh:
        broadcast.perf_error_kind = "intent_refresh"
        broadcast.perf_startup_failure_phase = "context"
        queue_settled = await _abort_turn_startup(
            session_id,
            broadcast,
            "failed",
            pause_queue=True,
            error_text="turn intent could not be refreshed",
        )
        raise _TurnStartError(
            "turn intent could not be refreshed",
            queue_claim_settled=queue_settled,
        )
    # The durable write may finish after Stop cancelled (or tried to cancel)
    # this owner. Do not create a pump after the user-visible cancellation.
    if broadcast.cancelled:
        return await _finish_cancelled_startup(session_id, broadcast)

    _interrupted_at_startup.pop(session_id, None)

    async def _pump_gen_to_broadcast():
        nonlocal memory_outcome_scheduled
        turn_errored = False
        turn_error_text = ""
        terminal_published = False
        queue_settled = False

        async def _durable_error_event(
            error: Any,
            source_payload: dict[str, Any] | None = None,
        ) -> dict:
            """Persist and enrich a terminal error before subscribers see it."""
            nonlocal turn_error_text
            turn_error_text = str(error or turn_error_text or "turn stream failed")
            has_recovery = isinstance(
                (source_payload or {}).get("recovered_session"), dict)
            if (_classify_stream_error(turn_error_text).get("kind") == "context_window"
                    and not has_recovery):
                # A process that crossed the hard window can remain poisoned
                # even after this request ends. Evict it at the safe turn
                # boundary so the user's next attempt gets the current compact
                # settings instead of reusing the same dead runtime forever.
                _pending_runtime_rebuilds.add(session_id)
            terminal_at_ms = int(time.time() * 1000)
            duration_ms = max(
                0, terminal_at_ms - int(float(broadcast.started_at or 0) * 1000))
            recall = mem0.pop_recall_trace(session_id)
            receipt = _persistable_memory_recall(recall)
            snapshot_ready = False
            try:
                snapshot_ready = await asyncio.to_thread(
                    _persist_failed_turn_snapshot,
                    broadcast,
                    turn_error_text,
                    terminal_at_ms=terminal_at_ms,
                    elapsed_s=round(duration_ms / 1000, 1),
                    memory_recall=receipt,
                    canonical_terminal_published=False,
                )
            except Exception as exc:
                _safe_secondary_diagnostic(
                    "terminal_snapshot", session_id, exc)
            event = _error_event(turn_error_text)
            try:
                data = json.loads(event.get("data") or "{}")
            except (TypeError, ValueError):
                data = {"error": turn_error_text}
            data.update({
                "completed_at_ms": terminal_at_ms,
                "duration_ms": duration_ms,
                "model": model_to_use,
                "assistant_uuid": str(broadcast.last_assistant_uuid or ""),
                "memory_recall": recall,
                "snapshot_ready": snapshot_ready,
                "activity_source": broadcast.activity_source,
            })
            source_payload = source_payload or {}
            for field in (
                "result_error", "terminal_reason", "status", "origin",
                "model_usage",
            ):
                value = source_payload.get(field)
                if value not in (None, "", {}, []):
                    data[field] = value
            recovered = source_payload.get("recovered_session")
            if isinstance(recovered, dict):
                recovered_id = str(
                    recovered.get("id") or recovered.get("session_id") or "")
                try:
                    valid_recovered_id = str(uuid.UUID(recovered_id))
                except (ValueError, AttributeError):
                    valid_recovered_id = ""
                if valid_recovered_id:
                    # The metadata was produced by sess.register_session. Copy
                    # it as a public object, while refusing any accidental
                    # content-bearing fields should that internal contract
                    # expand later.
                    safe_recovered = {
                        key: value for key, value in recovered.items()
                        if key not in {
                            "summary", "prompt", "messages", "content",
                            "source_path", "path",
                        }
                    }
                    safe_recovered["id"] = valid_recovered_id
                    safe_recovered["session_id"] = valid_recovered_id
                    data["recovered_session"] = safe_recovered
                    stats = source_payload.get("recovery_stats")
                    if isinstance(stats, dict):
                        data["recovery_stats"] = {
                            key: int(value)
                            for key, value in stats.items()
                            if key in {
                                "included_messages", "omitted_messages",
                                "truncated_messages", "estimated_post_tokens",
                            } and isinstance(value, (int, float))
                        }
            event["data"] = json.dumps(data, ensure_ascii=False)
            return event

        try:
            # None → asyncio.timeout is a no-op wrapper (no deadline armed).
            async with asyncio.timeout(BG_TIMEOUT_S or None):
                async for ev in event_gen():
                    # Track in-band errors too (merge_q "error" → an SSE error
                    # event flows through event_gen without raising). The queue
                    # must pause on these exactly like an exception-path error.
                    if isinstance(ev, dict) and ev.get("event") == "error":
                        turn_errored = True
                        payload: dict[str, Any] = {}
                        try:
                            payload = json.loads(ev.get("data") or "{}")
                            if not isinstance(payload, dict):
                                payload = {}
                            turn_error_text = str(
                                payload.get("error") or payload.get("message") or "")
                        except (ValueError, TypeError):
                            pass
                        ev = await _durable_error_event(
                            turn_error_text, payload)
                    # SDK-level failures arrive as a NORMAL done event with
                    # is_error=True (max turns / budget / permission denied /
                    # API error — see _handle_result_message). Treat them as
                    # errors too so the queue pauses instead of headlessly
                    # cascading the next item onto a failing session.
                    elif isinstance(ev, dict) and ev.get("event") == "done":
                        done_data: dict[str, Any] = {}
                        try:
                            parsed_done = json.loads(ev.get("data") or "{}")
                            if isinstance(parsed_done, dict):
                                done_data = parsed_done
                            if done_data.get("is_error"):
                                turn_errored = True
                        except (ValueError, TypeError):
                            pass
                        # ResultMessage is the canonical transcript boundary.
                        # Even a graceful done(cancelled=true) must not get a
                        # duplicate display snapshot layered over JSONL.
                        broadcast.canonical_terminal_published = True
                        if broadcast.queue_item_id:
                            if done_data.get("is_error") or done_data.get("cancelled"):
                                queue_settled = await _release_queue_claim_owned(
                                    session_id,
                                    broadcast.queue_item_id,
                                    turn_id=broadcast.turn_id,
                                    pause=True,
                                )
                            else:
                                queue_settled = await _ack_queue_message_owned(
                                    session_id,
                                    broadcast.queue_item_id,
                                    broadcast.turn_id,
                                )
                                if queue_settled:
                                    await obs.to_thread_io(
                                        "chat.queue_attachment_finish",
                                        session_id,
                                        _durable_attachment_store.finish_queue_item,
                                        session_id,
                                        broadcast.queue_item_id,
                                        consume=True,
                                        owned=True,
                                    )
                            if not queue_settled:
                                raise RuntimeError("queue terminal ownership mismatch")
                        if (not done_data.get("is_error")
                                and not done_data.get("cancelled")
                                and int(done_data.get(
                                    "background_tasks_pending") or 0) > 0):
                            # The Result handler persisted the exact assistant
                            # boundary before yielding this event. Start fork +
                            # cold-client setup now, in parallel with the slow
                            # post-turn transcript/context bookkeeping below.
                            _schedule_detached_successor_prewarm(session_id)
                    if (isinstance(ev, dict)
                            and ev.get("event") == "cancelled"
                            and broadcast.cancelled):
                        snapshot_ready = await asyncio.to_thread(
                            _persist_cancelled_turn_snapshot, broadcast)
                        try:
                            cancelled_data = json.loads(ev.get("data") or "{}")
                        except (TypeError, ValueError):
                            cancelled_data = {}
                        if not isinstance(cancelled_data, dict):
                            cancelled_data = {}
                        cancelled_data["snapshot_ready"] = snapshot_ready
                        ev = {
                            **ev,
                            "data": json.dumps(cancelled_data),
                        }
                    broadcast.publish(ev)
                    if isinstance(ev, dict) and ev.get("event") == "done":
                        terminal_published = True
        except asyncio.TimeoutError:
            if terminal_published:
                sys.stderr.write(
                    f"[chat] post-turn bookkeeping timed out "
                    f"sid={session_id[:8]}\n")
                return
            turn_errored = True
            turn_error_text = f"turn exceeded {BG_TIMEOUT_S}s"
            sys.stderr.write(
                f"[chat] turn exceeded MUSELAB_TURN_TIMEOUT_S={BG_TIMEOUT_S}s, "
                f"aborting sid={session_id[:8]}\n")
            sys.stderr.flush()
            try:
                broadcast.publish(await _durable_error_event(turn_error_text))
            except Exception as exc:
                _safe_secondary_diagnostic(
                    "terminal_replay", session_id, exc)
        except Exception as e:
            if terminal_published:
                sys.stderr.write(
                    f"[chat] post-turn bookkeeping failed "
                    f"sid={session_id[:8]} exc={type(e).__name__}\n")
                return
            primary_already_recorded = turn_errored and bool(turn_error_text)
            turn_errored = True
            if not primary_already_recorded:
                turn_error_text = f"{type(e).__name__}: {e}"
            error_kind = _classify_stream_error(
                turn_error_text).get("kind", "unknown")
            # The authenticated SSE frame below is the user-facing error
            # surface. Server logs keep only safe diagnostics: SDK/Gateway
            # exception strings and tracebacks can contain prompts, paths,
            # credentials, or upstream protocol payloads.
            sys.stderr.write(
                f"[chat] background turn crashed sid={session_id[:8]} "
                f"exc={type(e).__name__} kind={error_kind}\n")
            sys.stderr.flush()
            if primary_already_recorded:
                _safe_secondary_diagnostic(
                    "terminal_replay", session_id, e)
            else:
                try:
                    broadcast.publish(
                        await _durable_error_event(turn_error_text))
                except Exception as exc:
                    _safe_secondary_diagnostic(
                        "terminal_replay", session_id, exc)
        finally:
            # A force-stop takeover owns all terminal bookkeeping. A pump that
            # eventually escapes hung SDK code must only join that owner; doing
            # Activity/queue/sidecar settlement again reintroduces the race the
            # watchdog was meant to close.
            if broadcast._startup_terminal_cleanup_task is not None:
                await _cancel_outstanding_steering_commands(broadcast)
                await _finish_cancelled_startup(session_id, broadcast)
                return

            if turn_errored or broadcast.cancelled:
                # Call even when the map is empty: the helper closes admission
                # synchronously before its first possible await, preventing a
                # late enqueue during the rest of terminal bookkeeping.
                await _cancel_outstanding_steering_commands(broadcast)

            if broadcast.cancelled:
                broadcast.perf_status = "cancelled"
                broadcast.perf_error_kind = "cancelled"
            elif turn_errored:
                broadcast.perf_status = "failed"
                if broadcast.perf_error_kind in {"", "none"}:
                    broadcast.perf_error_kind = str(
                        _classify_stream_error(turn_error_text).get("kind")
                        or "unknown"
                    )
            elif broadcast.perf_status == "unknown":
                broadcast.perf_status = "completed"
            broadcast.perf_background_count = max(
                broadcast.perf_background_count,
                len(_sessions_with_inflight_tasks.get(session_id, ())),
            )
            # Successful ResultMessage already popped its trace into the done
            # payload. Every other terminal path (transport error, timeout,
            # forced stop) must clear it here to isolate the next turn.
            if not terminal_published:
                mem0.pop_recall_trace(session_id)
            # A force-stop can terminate receive_response() before the SDK
            # emits ResultMessage. Keep that turn as trajectory evidence, but
            # never let partial output become a fact/Skill candidate.
            deleting_session = sess.session_is_deleting(session_id)
            if not memory_outcome_scheduled and not deleting_session:
                if broadcast.cancelled:
                    memory_outcome_scheduled = mem0.schedule_cancelled(
                        session_id, prompt, broadcast.turn_id)
                elif turn_errored:
                    memory_outcome_scheduled = mem0.schedule_failed(
                        session_id, model_to_use, prompt,
                        "".join(assistant_acc),
                        turn_error_text or "turn stream failed",
                        broadcast.turn_id)
            if (broadcast.cancelled
                    and not broadcast.canonical_terminal_published
                    and not broadcast.cancelled_snapshot_persisted):
                await asyncio.to_thread(
                    _persist_cancelled_turn_snapshot, broadcast)
            # ResultMessage settles (or deliberately defers) Activity before
            # publishing ``done``.  Finishing it again here races the browser's
            # read ACK and turns the current, already-viewed result unread; for
            # detached tasks it also collapses the intended running state too
            # early.  Only exceptional paths that never published a terminal
            # Result still need the cleanup fallback.
            if not terminal_published:
                _activity_status = (
                    "cancelled" if broadcast.cancelled
                    else "failed" if turn_errored
                    else "completed"
                )
                await _finish_activity(
                    session_id, broadcast, _activity_status)
            if broadcast.queue_item_id and not queue_settled:
                try:
                    if turn_errored or broadcast.cancelled:
                        if not await _release_queue_claim_owned(
                            session_id,
                            broadcast.queue_item_id,
                            turn_id=broadcast.turn_id,
                            pause=True,
                        ):
                            raise RuntimeError("queue terminal ownership mismatch")
                    else:
                        queue_settled = await _ack_queue_message_owned(
                            session_id,
                            broadcast.queue_item_id,
                            broadcast.turn_id,
                        )
                        if not queue_settled:
                            raise RuntimeError("queue terminal ownership mismatch")
                        await obs.to_thread_io(
                            "chat.queue_attachment_finish",
                            session_id,
                            _durable_attachment_store.finish_queue_item,
                            session_id,
                            broadcast.queue_item_id,
                            consume=True,
                            owned=True,
                        )
                except Exception as e:
                    # Never duplicate a turn by guessing. The durable inflight
                    # record remains for restart recovery if acknowledgement fails.
                    sys.stderr.write(
                        f"[chat] queue terminal ack failed sid={session_id[:8]} "
                        f"item={broadcast.queue_item_id[:8]} "
                        f"exc={type(e).__name__}\n")
            # Keep the completed canonical turn's reservation while committing
            # any hidden-owner reply that became READY during it.  Releasing the
            # slot first lets the queue reserve its next turn and can postpone
            # this Agent bubble until the whole queue drains.
            if not deleting_session:
                try:
                    await _flush_runtime_continuations_at_turn_boundary(
                        session_id,
                        expected_active=broadcast,
                    )
                except Exception as exc:
                    # The outbox remains durable. Queue drain below fails closed
                    # while a READY record exists and delivery re-kicks it after
                    # a later successful presentation commit.
                    sys.stderr.write(
                        f"[chat] runtime continuation boundary flush failed "
                        f"sid={session_id[:8]} exc={type(exc).__name__}\n"
                    )
                    sys.stderr.flush()
            broadcast.finish()
            _active_turns.pop(session_id, None)
            if (not deleting_session
                    and session_id in _pending_runtime_rebuilds):
                await _rebuild_session_runtime(session_id)
            # Grace-keep: a fast (esp. server-drained) turn can finish + get
            # popped here BEFORE the browser's reconnect SSE attaches. Stash the
            # finished broadcast so a slightly-late reconnect still replays it
            # (full events + done sentinel) instead of seeing "no active turn"
            # and silently dropping the rendered content until a manual refresh.
            if not deleting_session:
                _remember_recent_turn(session_id, broadcast)
            # Clear only after another durable owner exists. A failure whose
            # snapshot write failed deliberately leaves the pending intent for
            # restart recovery instead of acknowledging data we did not retain.
            if ((not turn_errored and not broadcast.cancelled)
                    or broadcast.failed_snapshot_persisted
                    or broadcast.cancelled_snapshot_persisted):
                await _settle_active_turn_sidecar_owned(
                    session_id, release=True)
            # Server-side queue drain (Option B). Now that _active_turns no
            # longer holds this sid, advance the queue:
            #   - errored → pause the queue (don't cascade failures headlessly;
            #     user resumes manually, which re-kicks the drain) + push.
            #   - clean   → pop the next queued item and start its turn. That
            #     turn's own cleanup re-enters here, keeping the chain going
            #     until the queue empties — all with no browser attached.
            try:
                if deleting_session:
                    pass
                elif turn_errored:
                    # Only pause + notify if items are actually waiting —
                    # a lone failed turn with an empty queue is just a normal
                    # error the user already saw in-stream; no need to buzz.
                    q = await obs.to_thread_io(
                        "chat.queue_read", session_id, sess.get_queue, session_id)
                    if q.get("items"):
                        await obs.to_thread_io(
                            "chat.queue_pause",
                            session_id,
                            sess.set_queue_paused,
                            session_id,
                            True,
                            owned=True,
                        )
                        _notify_queue_paused_on_error(session_id)
                elif broadcast.cancelled:
                    # User explicitly stopped this turn — pause the queue so
                    # the remaining items don't auto-fire. They resume manually.
                    # Keep the empty-queue invariant atomic.  A late terminal
                    # cleanup used to create ``{items: [], paused: true}``;
                    # the next message then entered a queue that could never
                    # drain, which looked like an intermittent send failure.
                    await obs.to_thread_io(
                        "chat.queue_pause_nonempty",
                        session_id,
                        sess.pause_queue_if_nonempty,
                        session_id,
                        owned=True,
                    )
                else:
                    await _maybe_drain_queue(session_id)
            except Exception as e:
                sys.stderr.write(
                    f"[chat] queue drain trigger failed sid={session_id[:8]} "
                    f"exc={type(e).__name__}\n")

    # Keep a handle to the detached pump so the force-stop watchdog can cancel
    # it if an interrupt + client teardown ever fails to unblock receive_response.
    broadcast.task = asyncio.create_task(_pump_gen_to_broadcast())
    broadcast.startup_owner_task = None

    return broadcast


def _notify_turn_done(session_id: str, *, session_name: str = "") -> None:
    """Best-effort turn notification that never owns the active-turn slot."""
    async def _go():
        try:
            from . import presence as _presence
            if _presence.recently_active():
                return
            from . import push as _push
            display_name = session_name
            try:
                latest = await asyncio.to_thread(sess.get_session, session_id)
                display_name = str((latest or {}).get("name") or display_name)
            except Exception:
                pass
            # Never put reply content on a lock screen. The actual response is
            # one tap away in the authenticated application.
            await asyncio.to_thread(
                _push.send_to_all,
                title=display_name or "muselab",
                body="Muse 已回复",
                url=f"/?session={session_id}",
                tag=f"turn-{session_id}",
                context=f"turn-done {session_id[:8]}",
            )
        except Exception as e:
            sys.stderr.write(
                f"[chat] turn push failed sid={session_id[:8]} "
                f"exc={type(e).__name__}\n")
            sys.stderr.flush()

    try:
        task = asyncio.get_running_loop().create_task(_go())
        _retain_detached_cleanup(task)
    except RuntimeError:
        pass  # no running loop (shouldn't happen in request context)


def _notify_queue_paused_on_error(session_id: str) -> None:
    """Push 'Muse 暂停了队列（出错）' when the headless drain pauses the queue
    after a turn errored. Best-effort + presence-gated (don't buzz a user
    who's already at a screen). Fire-and-forget so it never blocks cleanup."""
    async def _go():
        try:
            from . import presence as _presence
            if _presence.recently_active():
                return
            from . import push as _push
            sname = ""
            try:
                sessions = await asyncio.to_thread(sess.list_sessions)
                for s in sessions:
                    if s.get("id") == session_id:
                        sname = s.get("name", "")
                        break
            except Exception:
                pass
            await asyncio.to_thread(
                _push.send_to_all,
                title=sname or "muselab",
                body="队列已暂停（上一条出错），点开查看",
                url=f"/?session={session_id}",
                tag=f"queue-paused-{session_id}",
                context=f"queue-paused {session_id[:8]}",
            )
        except Exception as e:
            sys.stderr.write(
                f"[chat] queue-paused push failed "
                f"sid={session_id[:8]} exc={type(e).__name__}\n")
    try:
        task = asyncio.get_running_loop().create_task(_go())
        _retain_detached_cleanup(task)
    except RuntimeError:
        pass  # no running loop (shouldn't happen in request context)


def _schedule_queue_drain_retry(session_id: str, delay_s: float = 1.0) -> None:
    """Retain one delayed retry for a recoverable runtime handoff conflict."""
    existing = _queue_drain_retry_tasks.get(session_id)
    if existing is not None and not existing.done():
        return

    async def _retry() -> None:
        await asyncio.sleep(delay_s)
        queue = await obs.to_thread_io(
            "chat.queue_retry_read",
            session_id,
            sess.get_queue,
            session_id,
        )
        if queue.get("paused") or not (
            queue.get("items") or queue.get("inflight")
        ):
            return
        _schedule_queue_drain(session_id)

    task = asyncio.create_task(_retry())
    _queue_drain_retry_tasks[session_id] = task
    _maintenance_tasks.add(task)

    def _done(done: asyncio.Task) -> None:
        _maintenance_tasks.discard(done)
        if _queue_drain_retry_tasks.get(session_id) is done:
            _queue_drain_retry_tasks.pop(session_id, None)
        if done.cancelled():
            return
        try:
            exc = done.exception()
        except asyncio.CancelledError:
            return
        if exc is not None:
            sys.stderr.write(
                f"[chat] delayed queue drain failed "
                f"sid={session_id[:8]} exc={type(exc).__name__}\n"
            )

    task.add_done_callback(_done)


async def _maybe_drain_queue(session_id: str) -> None:
    """Drain trigger: if no turn is running for this session and the queue
    has a non-paused head item, pop it and start the next turn headlessly.

    Called from (a) a just-finished turn's cleanup (the chain that keeps the
    queue advancing with no browser attached) and (b) manual resume. Respects
    the per-sid _active_turns mutex — if a turn is somehow still in flight,
    do nothing; that turn's own completion re-triggers the drain.

    On a lost race (_TurnBusy) or start failure (_TurnStartError), the popped
    item is restored to the queue head so nothing is dropped. A start failure
    additionally pauses the queue (mirrors the turn-errored policy)."""
    async with _queue_drain_lock_for(session_id):
        runtime_meta = await obs.to_thread_io(
            "chat.queue_session_read",
            session_id,
            sess.get_session_meta,
            session_id,
        ) or {}
        successor_sid = str(runtime_meta.get("runtime_successor") or "")
        if runtime_meta.get("runtime_shadow"):
            if successor_sid:
                try:
                    moved = await obs.to_thread_io(
                        "chat.queue_migrate",
                        session_id,
                        sess.migrate_queue,
                        session_id,
                        successor_sid,
                        owned=True,
                    )
                except ValueError:
                    return
                try:
                    _durable_attachment_store.migrate_queue_items(
                        session_id,
                        [str(row.get("id") or "")
                         for row in moved["target"].get("items", [])],
                        successor_sid,
                    )
                except Exception as exc:
                    sys.stderr.write(
                        f"[chat] attachment ref migration deferred "
                        f"exc={type(exc).__name__}\n")
                if moved["target"].get("items"):
                    _schedule_queue_drain(successor_sid)
            return
        if session_id in _active_turns and not _active_turns[session_id].done:
            return
        # The watcher owns a logical response boundary even though no ordinary
        # TurnBroadcast is open. It will trigger this drain again after its
        # final continuation releases the single SDK pump.
        if (_sessions_with_inflight_tasks.get(session_id)
                or _session_has_live_watcher(session_id)
                or _session_has_scheduled_delivery(session_id)):
            queued = await obs.to_thread_io(
                "chat.queue_read", session_id, sess.get_queue, session_id)
            if queued.get("items") or queued.get("inflight"):
                try:
                    successor = await _continue_detached_runtime(session_id)
                    successor_sid = str(successor.get("session_id") or "")
                    if successor_sid and successor_sid != session_id:
                        _schedule_queue_drain(successor_sid)
                except HTTPException as exc:
                    sys.stderr.write(
                        f"[chat] queued runtime rollover deferred "
                        f"sid={session_id[:8]} status={exc.status_code}\n"
                    )
                    if exc.status_code in {404, 409}:
                        _schedule_queue_drain_retry(session_id)
            return
        # READY presentation events are part of the visible chronology. Commit
        # them before claiming the next FIFO item; if local I/O cannot do so,
        # leave the queue untouched and let the retained delivery owner re-kick.
        try:
            await _flush_runtime_continuations_at_turn_boundary(session_id)
        except Exception as exc:
            sys.stderr.write(
                f"[chat] queued runtime continuation flush failed "
                f"sid={session_id[:8]} exc={type(exc).__name__}\n"
            )
            sys.stderr.flush()
            return
        if await obs.to_thread_io(
            "chat.runtime_continuation_ready_read",
            session_id,
            _runtime_lineage_has_ready_continuation,
            session_id,
        ):
            return
        item = await obs.to_thread_io(
            "chat.queue_claim",
            session_id,
            sess.claim_queue_message,
            session_id,
            owned=True,
        )
        if item is None:
            return
        item_id = str(item.get("id") or "")
        try:
            _queue_snapshot = await obs.to_thread_io(
                "chat.queue_read", session_id, sess.get_queue, session_id)
            _queue_depth = (
                len(_queue_snapshot.get("items") or [])
                + int(bool(_queue_snapshot.get("inflight")))
            )
        except Exception:
            _queue_depth = -1
        try:
            _enqueued_at_ms = int(item.get("enqueued_at") or 0)
        except (TypeError, ValueError):
            _enqueued_at_ms = 0
        _perf_event(
            "queue.claim",
            sid8=obs.short_id(session_id),
            item8=obs.short_id(item_id),
            wait_ms=(max(0, int(time.time() * 1000) - _enqueued_at_ms)
                     if _enqueued_at_ms else -1),
            depth=_queue_depth,
        )
        # Queue files survive a process restart; staged uploads deliberately do
        # not, and they can also expire while waiting behind a long turn.  Never
        # degrade an attachment-bearing message into a text-only turn.  Check
        # the complete set before `_start_turn` consumes even one upload, then
        # atomically restore + pause the exact claim.  GET /queue resolves the
        # now-missing ids as `available: false`, which keeps the prompt visible
        # and gives the browser an explicit edit/reattach recovery path.
        attachment_ids = [
            aid.strip()
            for aid in str(item.get("image_ids") or "").split(",")
            if aid.strip()
        ]
        if attachment_ids:
            try:
                if any(
                    not _valid_staged_attachment_id(aid)
                    for aid in attachment_ids
                ):
                    unavailable_count = len(attachment_ids)
                else:
                    await asyncio.to_thread(_gc_images)
                    with _image_store_lock:
                        cached_ids = set(_image_store).intersection(
                            attachment_ids)
                    durable_ids = await asyncio.to_thread(
                        _durable_attachment_store.existing_ids,
                        attachment_ids,
                    )
                    unavailable_count = sum(
                        1 for aid in attachment_ids
                        if aid not in cached_ids and aid not in durable_ids)
            except (DurableAttachmentError, OSError, sqlite3.Error,
                    UnsafePrivatePath) as exc:
                restored = await _release_queue_claim_owned(
                    session_id, item_id, pause=True)
                sys.stderr.write(
                    f"[chat] queued attachment precheck failed "
                    f"sid={session_id[:8]} item={item_id[:8]} "
                    f"restored={restored} exc={type(exc).__name__}\n"
                )
                sys.stderr.flush()
                return
            if unavailable_count:
                restored = await _release_queue_claim_owned(
                    session_id, item_id, pause=True)
                sys.stderr.write(
                    f"[chat] queued attachments unavailable "
                    f"sid={session_id[:8]} item={item_id[:8]} "
                    f"count={unavailable_count} restored={restored}\n")
                sys.stderr.flush()
                return
        # Replay under the permission mode snapshotted at enqueue time. Items
        # from before the snapshot existed (or enqueued without one) fail CLOSED
        # to "default" — requiring tool approval is the safe direction; the old
        # behavior (falling through to bypassPermissions) let queued messages
        # skip approval the user's UI said was required.
        perm = (item.get("permission") or "").strip() or "default"
        if perm not in _VALID_PERMISSION_MODES:
            # Headless context — can't 400. An unknown persisted value (pre-
            # validation enqueue, hand-edited queue file) fails CLOSED to
            # "default" rather than crashing the drain or reaching the SDK.
            perm = "default"
        try:
            await _start_turn(
                session_id,
                item.get("text", ""),
                permission=perm,
                plan_return_permission=item.get("plan_return_permission", ""),
                image_ids=item.get("image_ids", ""),
                persist_permission=False,
                queue_item_id=item_id,
            )
        except asyncio.CancelledError:
            # If startup never bound the claim, it is safe to restore. Once a
            # turn id is bound, its detached pump owns acknowledgement; requeueing
            # here would create the executed+queued duplicate seen in production.
            q = await obs.to_thread_io(
                "chat.queue_read", session_id, sess.get_queue, session_id)
            inflight = q.get("inflight") or {}
            bound_turn_id = str(inflight.get("turn_id") or "")
            active = _active_turns.get(session_id)
            active_owns_claim = bool(
                bound_turn_id and active is not None and not active.done
                and active.turn_id == bound_turn_id
                and active.queue_item_id == item_id
            )
            session_exists = await obs.to_thread_io(
                "chat.session_read", session_id, sess.get_session, session_id)
            if session_exists is not None and not active_owns_claim:
                await _release_queue_claim_owned(
                    session_id, item_id, turn_id=bound_turn_id)
            raise
        except _TurnBusy:
            await _release_queue_claim_owned(session_id, item_id)
        except _TurnStartError as exc:
            # If startup failed before binding, this releases the unbound claim.
            # Bound startup failures settle with their exact turn id in _start_turn.
            if not exc.queue_claim_settled:
                await _release_queue_claim_owned(session_id, item_id, pause=True)
            _notify_queue_paused_on_error(session_id)
        except Exception as e:
            await _release_queue_claim_owned(session_id, item_id, pause=True)
            sys.stderr.write(
                f"[chat] queue drain crashed sid={session_id[:8]} "
                f"exc={type(e).__name__}\n")
            _notify_queue_paused_on_error(session_id)


def _schedule_queue_drain(session_id: str) -> None:
    """Kick one retained, coalesced drain task without delaying enqueue HTTP."""
    existing = _queue_drain_tasks.get(session_id)
    if existing is not None and not existing.done():
        # Coalescing used to discard this wakeup. During a long runtime fork,
        # that could strand a later accepted prompt on the hidden source queue.
        _queue_drain_rekicks.add(session_id)
        return
    task = asyncio.create_task(_maybe_drain_queue(session_id))
    _queue_drain_tasks[session_id] = task
    _maintenance_tasks.add(task)

    def _done(done: asyncio.Task) -> None:
        _maintenance_tasks.discard(done)
        if _queue_drain_tasks.get(session_id) is done:
            _queue_drain_tasks.pop(session_id, None)
        rekick = session_id in _queue_drain_rekicks
        _queue_drain_rekicks.discard(session_id)
        if done.cancelled():
            return
        try:
            exc = done.exception()
        except asyncio.CancelledError:
            return
        if exc is not None:
            sys.stderr.write(
                f"[chat] scheduled queue drain failed "
                f"sid={session_id[:8]} exc={type(exc).__name__}\n")
        if rekick:
            # Defer creation to the next loop turn so the completed task is no
            # longer observable as the owner. A later enqueue can set the bit
            # again while this follow-up pass runs.
            asyncio.get_running_loop().call_soon(
                _schedule_queue_drain, session_id)

    task.add_done_callback(_done)


def _mux_wrap_event(session_id: str, event: dict) -> dict:
    wrapped = dict(event)
    raw_data = event.get("data", "")
    try:
        payload = json.loads(raw_data)
    except (TypeError, ValueError):
        payload = {"data": raw_data}
    if not isinstance(payload, dict):
        payload = {"data": payload}
    payload["session_id"] = session_id
    wrapped["data"] = json.dumps(payload, ensure_ascii=False)
    return wrapped


def _mux_session_state_fingerprint(state: dict) -> str:
    stable = {
        key: value for key, value in state.items()
        if key not in {"events_so_far", "runtime_ui_revision"}
    }
    return json.dumps(stable, sort_keys=True, ensure_ascii=False)


async def _subscribe_multiplex(
    checkpoints: dict[str, dict],
    *,
    mobile: bool = False,
):
    """Merge attachable broadcasts while periodically discovering new ones."""
    # Flush the SSE response headers before discovery. Starlette's global
    # GZipMiddleware deliberately skips compressing ``text/event-stream``, but
    # its responder still buffers ``http.response.start`` until it sees the
    # first body frame. An idle mux (no checkpoints and no active turns) has no
    # session event to yield, so without this handshake frame the browser waits
    # for sse-starlette's 15-second heartbeat and our 5-second EventSource open
    # timeout fires first. ``ping`` is already part of the mux protocol and is
    # ignored by panes that have no active channel.
    yield {"event": "ping", "data": ""}

    # Bound the aggregate handoff so a stalled HTTP client leaves each child
    # parked on its disk-backed subscriber cursor instead of growing RAM.
    output: asyncio.Queue[dict] = asyncio.Queue(maxsize=256)
    children: dict[tuple[str, str], asyncio.Task] = {}
    completed: set[tuple[str, str]] = set()
    state_fingerprints: dict[str, str] = {}
    state_payloads: dict[str, dict] = {}
    # Checkpoints are one-shot reconnect intents for this root SSE handshake.
    # Copy them so consumption is private to the subscription lifecycle.
    pending_checkpoints = {
        session_id: dict(checkpoint)
        for session_id, checkpoint in checkpoints.items()
    }

    async def _pump_child(
        session_id: str,
        broadcast: TurnBroadcast,
        last_event_seq: int,
    ) -> None:
        try:
            async for event in _subscribe_broadcast(
                broadcast,
                mobile=mobile,
                last_event_seq=last_event_seq,
            ):
                await output.put(_mux_wrap_event(session_id, event))
        except asyncio.CancelledError:
            raise
        except Exception:
            await output.put({
                "event": "resync",
                "data": json.dumps({
                    "reason": "child_stream_error",
                    "fallback": "canonical_history",
                    "retryable": False,
                    "session_id": session_id,
                    "turn_id": broadcast.turn_id,
                }),
            })

    def _start_child(
        session_id: str,
        broadcast: TurnBroadcast,
        last_event_seq: int,
    ) -> None:
        key = (session_id, broadcast.turn_id)
        if key in children or key in completed:
            return
        children[key] = asyncio.create_task(
            _pump_child(session_id, broadcast, last_event_seq))

    async def _reconcile() -> None:
        for key, task in list(children.items()):
            if task.done():
                children.pop(key, None)
                completed.add(key)
                with suppress(asyncio.CancelledError, Exception):
                    task.result()

        candidate_ids = set(_active_turns)
        candidate_ids.update(_sessions_with_inflight_tasks)
        candidate_ids.update(
            sid for sid, watcher in _task_watchers.items()
            if watcher is not None and not watcher.done()
        )
        candidate_ids.update(
            sid for sid, broadcast in _recent_turns.items()
            if (
                getattr(broadcast, "is_continuation", False)
                and not getattr(broadcast, "continuation_consumed", False)
            ) or (
                getattr(broadcast, "is_scheduled_delivery", False)
                and not getattr(
                    broadcast, "scheduled_delivery_consumed", False)
            )
        )
        candidate_ids.update(pending_checkpoints)

        active_ids: set[str] = set()
        for session_id in sorted(candidate_ids):
            state = session_active_status(session_id)
            checkpoint = pending_checkpoints.get(session_id)
            checkpoint_recent = None
            if checkpoint is not None and checkpoint.get("turn_id"):
                recent = _get_recent_turn(session_id)
                if (recent is not None
                        and recent.turn_id == checkpoint.get("turn_id")):
                    checkpoint_recent = recent
            if state.get("active"):
                active_ids.add(session_id)
                state_payload = dict(state)
                state_payload["session_id"] = session_id
                fingerprint = _mux_session_state_fingerprint(state_payload)
                state_payloads[session_id] = state_payload
                state_event = None
                if state_fingerprints.get(session_id) != fingerprint:
                    state_fingerprints[session_id] = fingerprint
                    state_event = {
                        "event": "session_state",
                        "data": json.dumps(state_payload, ensure_ascii=False),
                    }

                current_start = None
                checkpoint_consumed = False
                if state.get("attachable"):
                    current_turn_id = str(state.get("turn_id") or "")
                    broadcast = _active_turns.get(session_id)
                    if (broadcast is None
                            or broadcast.turn_id != current_turn_id):
                        recent = _get_recent_turn(session_id)
                        broadcast = (
                            recent if recent is not None
                            and recent.turn_id == current_turn_id else None
                        )
                    if broadcast is not None:
                        resume_seq = 0
                        if checkpoint is not None:
                            requested_turn_id = str(
                                checkpoint.get("turn_id") or "")
                            if (requested_turn_id
                                    and requested_turn_id != current_turn_id
                                    and checkpoint_recent is None):
                                # Recover only the stale owner.  This targeted
                                # frame must precede the new turn's state so the
                                # frontend cannot apply it to the successor.
                                await output.put({
                                    "event": "resync",
                                    "data": json.dumps({
                                        "reason": "turn_changed",
                                        "fallback": "canonical_history",
                                        "retryable": False,
                                        "requested_turn_id": requested_turn_id,
                                        "current_turn_id": current_turn_id,
                                        "session_id": session_id,
                                        "turn_id": requested_turn_id,
                                    }),
                                })
                                checkpoint_consumed = True
                            elif requested_turn_id == current_turn_id:
                                resume_seq = int(
                                    checkpoint.get("last_event_seq", 0) or 0)
                                checkpoint_consumed = True
                        current_start = (broadcast, resume_seq)

                if state_event is not None:
                    await output.put(state_event)
                if current_start is not None:
                    _start_child(session_id, *current_start)
                    if checkpoint_consumed:
                        pending_checkpoints.pop(session_id, None)

            if checkpoint_recent is not None:
                _start_child(
                    session_id,
                    checkpoint_recent,
                    int(checkpoint.get("last_event_seq", 0) or 0),
                )
                pending_checkpoints.pop(session_id, None)
            elif checkpoint is not None and not state.get("active"):
                # The requested turn is outside the bounded recent window.
                # Canonicalize it once, report inactivity, then stop polling.
                requested_turn_id = str(checkpoint.get("turn_id") or "")
                await output.put({
                    "event": "resync",
                    "data": json.dumps({
                        "reason": "checkpoint_unavailable",
                        "fallback": "canonical_history",
                        "retryable": False,
                        "session_id": session_id,
                        "turn_id": requested_turn_id,
                    }),
                })
                pending_checkpoints.pop(session_id, None)
                if session_id not in state_fingerprints:
                    await output.put({
                        "event": "session_state",
                        "data": json.dumps({
                            "session_id": session_id,
                            "turn_id": requested_turn_id,
                            "active": False,
                            "stopping": False,
                            "attachable": False,
                        }),
                    })

        retained_completed = set(children)
        retained_completed.update(
            (session_id, broadcast.turn_id)
            for session_id, broadcast in _active_turns.items()
        )
        completed.intersection_update(retained_completed)

        for session_id in list(state_fingerprints):
            if session_id in active_ids:
                continue
            # The child pump owns terminal delivery. Do not publish inactive
            # while it can still enqueue done/cancelled/error frames; once the
            # task is done, every child frame is already ahead of this state
            # transition in the same FIFO output queue.
            if any(key[0] == session_id for key in children):
                continue
            previous = state_payloads.pop(session_id, {"session_id": session_id})
            state_fingerprints.pop(session_id, None)
            inactive = {
                **previous,
                "session_id": session_id,
                "active": False,
                "stopping": False,
                "attachable": False,
            }
            await output.put({
                "event": "session_state",
                "data": json.dumps(inactive, ensure_ascii=False),
            })

    try:
        while True:
            await _reconcile()
            try:
                event = await asyncio.wait_for(
                    output.get(), timeout=_MUX_RECONCILE_INTERVAL_S)
            except asyncio.TimeoutError:
                continue
            yield event
    finally:
        for task in children.values():
            task.cancel()
        if children:
            await asyncio.gather(*children.values(), return_exceptions=True)


async def _subscribe_broadcast(
    broadcast: TurnBroadcast,
    *,
    mobile: bool = False,
    last_event_seq: int = 0,
):
    """Yield full or incremental replay followed by its live tail.

    Sequence zero reads the complete disk-backed spool. Positive checkpoints
    use the broadcast's bounded exact-event window and receive one explicit
    ``resync`` event if the missing range is unavailable. Every attached reader
    then tails the same append-only spool, so stalled HTTP connections retain a
    file cursor rather than an unbounded Python queue.
    """
    subscriber = broadcast.subscribe(
        mobile=mobile, last_event_seq=last_event_seq)
    # A real subscriber is now attached. For a CONTINUATION broadcast this is
    # the one-and-only reconnect that replays the finished task's card flip +
    # reaction. A replay-gap fallback is not an attachment and must not consume
    # the continuation advertisement before canonical reconciliation starts.
    if (getattr(broadcast, "is_continuation", False)
            and not subscriber._resync_payload):
        broadcast.continuation_consumed = True
    if (getattr(broadcast, "is_scheduled_delivery", False)
            and not subscriber._resync_payload):
        broadcast.scheduled_delivery_consumed = True
    try:
        while True:
            ev = await subscriber.get()
            if ev is None:
                break
            yield ev
    finally:
        broadcast.unsubscribe(subscriber)


@router.get("/sessions/{sid}/active", dependencies=[Depends(require_token)])
def session_active_status(sid: str) -> dict:
    """Tell the frontend whether `sid` has an in-progress background
    turn. Used on session load to decide between "render JSONL history"
    and "open a reconnect SSE stream to follow the live tail."""
    runtime_task_ids = set(_sessions_with_inflight_tasks.get(sid, ()))
    runtime_task_ids.update(
        task_id
        for task_id, overlay in sess.get_runtime_task_overlays(sid).items()
        if overlay.get("state") == "running"
    )
    runtime_background_pending = len(runtime_task_ids)
    runtime_continuation_pending, runtime_ui_revision = (
        _runtime_continuation_projection_state(sid)
    )
    scheduled_state = _sdk_scheduled_snapshot(sid)
    b = _active_turns.get(sid)
    if b is not None and getattr(b, "is_continuation", False) \
            and getattr(b, "continuation_consumed", False):
        # A continuation already handed to a reconnect subscriber — don't
        # re-advertise even if it briefly lingers in _active_turns (reaction
        # still streaming). Prevents a second poll within the same window from
        # firing a duplicate reconnect. Falls through to the recent-fallback
        # (also consumed-gated) which returns inactive.
        b = None
    if not b:
        # Grace-keep fallback for HEADLESS CONTINUATION turns. The bg-task
        # watcher's continuation broadcast is short-lived in _active_turns —
        # it only sits there while its auto-continue reaction streams (~2s),
        # then _close_continuation pops it into _recent_turns. The frontend's
        # 8s poller almost always misses that ~2s window, so the card never
        # flips live. Surface a still-fresh continuation from _recent_turns
        # (within its TTL) so the poller catches it and reconnects in
        # continuation mode; GET /stream then grace-replays the buffered
        # events (task_notification flips the running card → ✅done, plus the
        # reaction bubble). Only continuations are surfaced — a plain
        # just-finished turn must NOT report active (it would trigger spurious
        # reconnects + duplicate replays). The frontend poller self-stops once
        # the card flips, so the 60s-TTL window yields exactly one replay.
        recent = _get_recent_turn(sid)
        if (recent is not None
                and (
                    getattr(recent, "is_continuation", False)
                    and not getattr(
                        recent, "continuation_consumed", False)
                    or getattr(recent, "is_scheduled_delivery", False)
                    and not getattr(
                        recent, "scheduled_delivery_consumed", False)
                )):
            b = recent
        else:
            background_pending = len(
                _sessions_with_inflight_tasks.get(sid, ()))
            if background_pending:
                # Busy but not attachable: the watcher is waiting for a task
                # notification and there is no TurnBroadcast to replay yet.
                # Frontends must keep the footer alive + queue new input, but
                # must not open an empty-prompt reconnect (that would receive
                # "no active turn"). A later continuation flips attachable.
                return {
                    "active": True,
                    "stopping": False,
                    "attachable": False,
                    "background": True,
                    "continuation": False,
                    "activity_source": "background",
                    "turn_id": _background_origin_turn_id.get(sid, ""),
                    "parent_turn_id": "",
                    "started_at": _background_turn_started_at.get(sid, 0),
                    "events_so_far": 0,
                    "background_tasks_pending": background_pending,
                    "runtime_background_tasks_pending": runtime_background_pending,
                    "runtime_continuation_pending": runtime_continuation_pending,
                    "runtime_ui_revision": runtime_ui_revision,
                    **scheduled_state,
                    "scheduled": False,
                    "user_text": "",
                    "user_images": [],
                    "user_docs": [],
                }
            return {
                "active": False,
                "stopping": False,
                "background_tasks_pending": 0,
                "runtime_background_tasks_pending": runtime_background_pending,
                "runtime_continuation_pending": runtime_continuation_pending,
                "runtime_ui_revision": runtime_ui_revision,
                **scheduled_state,
                "scheduled": False,
                # A just-finished direct or queued turn can disappear from the
                # live registry before the reconnect probe lands. Preserve its
                # delivery class through the grace broadcast so the browser
                # never guesses that a queued completion was foreground work.
                "activity_source": (
                    recent.activity_source if recent is not None else ""
                ),
            }
    _hydrate_staged_attachment_display(b)
    return {
        "active": True,
        "stopping": bool(b.cancelled),
        "attachable": True,
        "background": False,
        "background_tasks_pending": len(
            _sessions_with_inflight_tasks.get(sid, ())),
        "runtime_background_tasks_pending": runtime_background_pending,
        "runtime_continuation_pending": runtime_continuation_pending,
        "runtime_ui_revision": runtime_ui_revision,
        **scheduled_state,
        "turn_id": b.turn_id,
        "parent_turn_id": b.parent_turn_id,
        "model": b.model,
        "started_at": b.started_at,
        "events_so_far": b.replay_count(),
        # True when this is a HEADLESS CONTINUATION turn opened by the bg-task
        # watcher (no user prompt). The frontend attaches in "continuation"
        # mode — same reconnect SSE, but it must NOT truncate the in-flight
        # portion (the launching tool_use card lives there; the replayed
        # task_notification flips it to ✅done).
        "continuation": getattr(b, "is_continuation", False),
        "scheduled": getattr(b, "is_scheduled_delivery", False),
        "activity_source": b.activity_source,
        # The turn's user prompt + attachments. The browser needs these to
        # render the user bubble when it LIVE-reconnects to a turn the server
        # drained from the queue headlessly (the browser never "sent" it, so
        # the bubble isn't in `messages`). Same fields _broadcast_to_ui_messages
        # injects on a reload-rebuild — keeps the two reconnect paths in sync.
        "user_text": b.user_text or "",
        "user_images": b.user_images or [],
        "user_docs": b.user_docs or [],
    }


# ====== interrupted turns (process-crash recovery) ======

@router.get("/interrupted-turns", dependencies=[Depends(require_token)])
def list_interrupted_turns() -> dict:
    """Returns turns that were in-flight when the previous muselab process
    died. Empty list on clean restart. Frontend reads this once per session
    boot and toasts the user — does NOT auto-resume (user decides whether
    the conversation is worth retrying)."""
    items = []
    for sid, data in _interrupted_at_startup.items():
        # Materialize the pending user intent before the browser can acknowledge
        # the notification. The toast is optional UX; the history row is durable.
        _recover_interrupted_turn_snapshot(sid)
        items.append({
            "sid": sid,
            "preview": data.get("user_text_preview") or "",
            "model": data.get("model") or "",
            "started_at": data.get("started_at") or 0,
        })
    # Most recent first — usually what the user remembers best
    items.sort(key=lambda x: x["started_at"], reverse=True)
    return {"turns": items}


@router.post("/interrupted-turns/{sid}/dismiss",
             dependencies=[Depends(require_token)])
def dismiss_interrupted_turn(sid: str) -> dict:
    """Acknowledge the warning only after its user intent is durable history."""
    if sid in _interrupted_at_startup and not _recover_interrupted_turn_snapshot(sid):
        raise HTTPException(
            503, "interrupted turn could not be persisted; acknowledgement deferred")
    _interrupted_at_startup.pop(sid, None)
    return {"ok": True}


# ====== ask_user_question answer endpoint ======

class AnswerReq(BaseModel):
    answers: dict[str, Any]  # question_text -> chosen label (str) or labels (list[str])


@router.post("/answer/{session_id}/{question_id}",
              dependencies=[Depends(require_token)])
async def submit_answer_api(session_id: str, question_id: str, req: AnswerReq) -> dict:
    """Frontend POSTs the user's button click here. Resolves the Future the
    ask_user_question tool handler is await-ing; the tool then returns a
    text result and the model continues."""
    if not submit_answer(session_id, question_id, req.answers):
        raise HTTPException(404, "no pending question with that id "
                                  "(may have timed out or been answered already)")
    return {"ok": True}


# ====== permission request decision endpoint ======

class PermissionDecisionReq(BaseModel):
    decision: str           # "allow" | "deny" | "always"
    message: str | None = None
    # ExitPlanMode only: one of the SDK-provided setMode suggestions.
    mode: str | None = None


@router.post("/permission/{session_id}/{request_id}",
              dependencies=[Depends(require_token)])
async def submit_permission_decision_api(
    session_id: str, request_id: str, req: PermissionDecisionReq
) -> dict:
    """Frontend POSTs Allow / Deny / Always-allow click here."""
    if not perm.submit_decision(
        session_id, request_id, req.decision, req.message, mode=req.mode,
    ):
        raise HTTPException(404, "no pending permission request with that id "
                                  "(may have timed out or been answered already)")
    return {"ok": True}




@router.get("/providers", dependencies=[Depends(require_token)])
async def providers_list() -> dict:
    """Available model groups based on which provider API keys are configured."""
    groups = endpoints.available_groups()
    codex_capabilities = await _detect_gateway_context_capabilities(
        i["model"]
        for group in groups
        for i in group["items"]
        if _is_codex_gateway_model(i["model"])
    )
    # Flatten to the {group, label, model} shape the frontend expects.
    # supports_thinking / supports_effort are provider-level (see
    # available_groups) — the FE uses them to show/hide per-session controls so
    # models on vendors that reject or ignore the knobs don't get no-op switches.
    flat: list[dict[str, Any]] = []
    for g in groups:
        for i in g["items"]:
            model = i["model"]
            capability = codex_capabilities.get(model)
            controls = _model_control_capability(model, capability)
            is_ducc = endpoints.is_ducc_model(model)
            ducc_claude = is_ducc and endpoints.ducc_is_claude_model(model)
            flat.append({
                "group": g["group"],
                "label": i["label"],
                "model": model,
                "supports_thinking": (
                    ducc_claude if is_ducc
                    else g.get("supports_thinking", True)
                ),
                "supports_effort": (
                    ducc_claude if is_ducc
                    else g.get("supports_effort", False)
                ),
                "effort_levels": controls["effort_levels"],
                "service_tiers": controls["service_tiers"],
                "supports_fast": controls["supports_fast"],
            })
    # default_model: the configured "new-session default" (MUSELAB_MODEL),
    # already narrowed to a reachable model by _resolve_default_model. The
    # frontend seeds each new chat from this instead of the currently-viewed
    # session's locked model — without it, a new session inherited whatever
    # old tab you happened to be on.
    try:
        default_permission = _validate_permission(
            os.environ.get("MUSELAB_DEFAULT_PERMISSION", "bypassPermissions"))
    except HTTPException:
        # A stale hand-edited env value must not leak invalid state into the
        # selector or the SDK launch contract.
        default_permission = "bypassPermissions"
    busy_send_mode = os.environ.get(
        "MUSELAB_BUSY_SEND_MODE", "adjust").strip().lower()
    if busy_send_mode not in {"adjust", "queue"}:
        busy_send_mode = "adjust"
    return {
        "models": flat,
        "default_model": _resolve_default_model(""),
        "default_permission": default_permission,
        "busy_send_mode": busy_send_mode,
    }

def recover_durable_queue_attachments_at_startup(
    session_store=sess,
) -> dict[str, int]:
    """Reconcile durable refs against the complete, recovered queue set.

    This runs only after sessions.recover_queue_inflight has parsed and
    durably paused every queue. Consequently, an absent item is authoritative
    and its orphan ref can be released; a duplicate id fails startup closed.
    """
    owners: dict[str, str] = {}
    for sid in session_store.list_queue_session_ids():
        queue = session_store.get_queue(sid)
        rows = list(queue.get("items") or [])
        inflight = queue.get("inflight") or {}
        if inflight.get("item"):
            rows.append(inflight["item"])
        for row in rows:
            item_id = str(row.get("id") or "")
            if not item_id:
                continue
            previous = owners.get(item_id)
            if previous is not None and previous != sid:
                raise RuntimeError("duplicate durable queue item id")
            owners[item_id] = sid
    return _durable_attachment_store.reconcile_queue_refs(
        owners,
        ttl=_IMAGE_TTL_S,
    )



# Dynamic bridge for SDK client/runtime lifecycle. Every callback resolves the
# chat facade at call time, preserving monkeypatch behavior while the extracted
# module owns the exact shared pool, disconnect, and stream-pump containers.
def _invalidate_hook_setting_runtimes(
    scope: hook_settings.HookScope,
    workspace_root: Path,
) -> None:
    """Force affected pooled SDK clients to reload standard Hook settings."""
    try:
        session_ids = sess.indexed_session_ids()
    except Exception:
        return
    target = Path(workspace_root).resolve(strict=False)
    for session_id in session_ids:
        if scope != "user":
            try:
                session_root = Path(
                    sess.session_workspace(session_id)
                ).resolve(strict=False)
            except Exception:
                continue
            if session_root != target:
                continue
        _pending_runtime_rebuilds.add(session_id)


hook_settings.configure_runtime_invalidator(
    _invalidate_hook_setting_runtimes)


def _sdk_scheduled_snapshot(session_id: str) -> dict[str, Any]:
    """Return the privacy-safe, thread-safe native schedule projection."""
    with _sdk_cron_state_lock:
        count = len(_sdk_cron_jobs.get(session_id, {}))
    return {"scheduled_active": count > 0, "scheduled_count": count}


def _session_has_scheduled_tasks(session_id: str) -> bool:
    return bool(_sdk_scheduled_snapshot(session_id)["scheduled_active"])


def _scheduled_state_carrier(
    key: chat_runtime.ClientKey,
) -> TurnBroadcast | None:
    delivery = _sdk_scheduled_deliveries.get(key)
    if delivery is not None and not delivery.broadcast.done:
        return delivery.broadcast
    broadcast = _active_turns.get(key[0])
    return broadcast if broadcast is not None and not broadcast.done else None


def _publish_sdk_scheduled_state(
    key: chat_runtime.ClientKey,
    *,
    broadcast: TurnBroadcast | None = None,
) -> None:
    carrier = broadcast or _scheduled_state_carrier(key)
    if carrier is None:
        return
    carrier.publish({
        "event": "scheduled_tasks",
        "data": json.dumps(_sdk_scheduled_snapshot(key[0])),
    })


def _safe_sdk_cron_call(name: str, raw_input: Any) -> dict[str, Any]:
    data = raw_input if isinstance(raw_input, dict) else {}
    call: dict[str, Any] = {"name": name}
    if name == "CronCreate":
        cadence = str(data.get("cron") or "")
        if cadence and len(cadence) <= 128 and all(ch.isprintable() for ch in cadence):
            call["cron"] = cadence
        call["recurring"] = bool(data.get("recurring"))
        call["durable"] = bool(data.get("durable"))
    elif name == "CronDelete":
        job_id = str(data.get("id") or "")
        if _SDK_CRON_JOB_ID.fullmatch(job_id):
            call["id"] = job_id
    return call


def _sdk_tool_result_text(block: ToolResultBlock) -> str:
    content = getattr(block, "content", None)
    if isinstance(content, str):
        return content
    parts: list[str] = []
    if isinstance(content, list):
        for item in content:
            if isinstance(item, TextBlock):
                parts.append(str(getattr(item, "text", "") or ""))
            elif isinstance(item, dict) and item.get("type") == "text":
                parts.append(str(item.get("text") or ""))
    return "\n".join(part for part in parts if part)


def _observe_sdk_cron_message(
    key: chat_runtime.ClientKey,
    message: Any,
) -> None:
    """Track native Cron tool state without retaining its prompt or output."""
    content = getattr(message, "content", None)
    if not isinstance(content, list):
        return
    calls = _sdk_cron_tool_calls.setdefault(key, {})
    changed = False
    observed_result = False
    for block in content:
        if isinstance(block, ToolUseBlock) and block.name in {
            "CronCreate", "CronDelete", "CronList",
        }:
            tool_id = str(block.id or "")
            if tool_id:
                calls[tool_id] = _safe_sdk_cron_call(
                    str(block.name), getattr(block, "input", None))
            continue
        if not isinstance(block, ToolResultBlock):
            continue
        tool_id = str(getattr(block, "tool_use_id", "") or "")
        call = calls.pop(tool_id, None)
        if call is None or bool(getattr(block, "is_error", False)):
            continue
        observed_result = True
        text = _sdk_tool_result_text(block)
        name = call["name"]
        with _sdk_cron_state_lock:
            before = dict(_sdk_cron_jobs.get(key[0], {}))
            jobs = dict(before)
            if name == "CronCreate":
                match = _SDK_CRON_CREATE_RESULT.search(text)
                if match:
                    job_id = match.group(1)
                    jobs[job_id] = {
                        field: call[field]
                        for field in ("cron", "recurring", "durable")
                        if field in call
                    }
            elif name == "CronDelete":
                match = _SDK_CRON_DELETE_RESULT.search(text)
                if match:
                    jobs.pop(match.group(1), None)
            elif name == "CronList":
                if "no scheduled jobs" in text.casefold():
                    jobs = {}
                else:
                    listed = {
                        match.group(1)
                        for match in _SDK_CRON_LIST_LINE.finditer(text)
                    }
                    if listed:
                        jobs = {
                            job_id: dict(before.get(job_id, {}))
                            for job_id in listed
                        }
            if jobs:
                _sdk_cron_jobs[key[0]] = jobs
            else:
                _sdk_cron_jobs.pop(key[0], None)
            changed = jobs != before
    if not calls:
        _sdk_cron_tool_calls.pop(key, None)
    if changed or observed_result:
        _publish_sdk_scheduled_state(key)


def _scheduled_trigger_text(message: UserMessage) -> str:
    content = getattr(message, "content", None)
    if isinstance(content, str):
        return content
    parts: list[str] = []
    if isinstance(content, list):
        for block in content:
            if isinstance(block, TextBlock):
                parts.append(str(getattr(block, "text", "") or ""))
            elif isinstance(block, dict) and block.get("type") == "text":
                parts.append(str(block.get("text") or ""))
    return "\n".join(part for part in parts if part)


def _is_sdk_scheduled_trigger(message: Any) -> bool:
    if not isinstance(message, UserMessage):
        return False
    origin = sdk_lifecycle.normalize_origin(getattr(message, "origin", None))
    return bool(
        origin
        and origin["kind"] == "task-notification"
        and origin.get("subkind") == "scheduled-trigger"
    )


async def _register_scheduled_delivery(
    delivery: _ScheduledDelivery,
) -> bool:
    """Claim the visible slot after the preceding turn releases it."""
    session_id = delivery.key[0]
    deadline = time.monotonic() + 10.0
    while time.monotonic() < deadline:
        if sess.session_is_deleting(session_id):
            return False
        async with _lock:
            current = _active_turns.get(session_id)
            if current is delivery.broadcast:
                return True
            if current is None or current.done:
                if delivery.broadcast.done:
                    _remember_recent_turn(session_id, delivery.broadcast)
                    return False
                _active_turns[session_id] = delivery.broadcast
                return True
        await asyncio.sleep(0.02)
    return False


def _retain_maintenance_task(task: asyncio.Task) -> None:
    _maintenance_tasks.add(task)
    task.add_done_callback(_maintenance_tasks.discard)


async def _begin_scheduled_delivery(
    key: chat_runtime.ClientKey,
    message: UserMessage,
) -> _ScheduledDelivery:
    prompt = _scheduled_trigger_text(message)
    broadcast = TurnBroadcast(session_id=key[0], model=key[1] or MODEL)
    broadcast.user_text = prompt
    broadcast.is_scheduled_delivery = True
    broadcast.perf_client = "warm"
    delivery = _ScheduledDelivery(
        key=key,
        broadcast=broadcast,
        render_state={
            "tool_use_names": {},
            "streamed": [],
            "assistant_uuid": "",
        },
        subagent_mux=chat_subagents.SubagentStreamMux(key[0]),
    )
    _sdk_scheduled_deliveries[key] = delivery
    registration = asyncio.create_task(_register_scheduled_delivery(delivery))
    delivery.registration_task = registration
    _retain_maintenance_task(registration)
    broadcast.publish_startup("accepted")
    try:
        await _start_activity_early(key[0], broadcast, prompt or "定时任务")
    except asyncio.CancelledError:
        raise
    except Exception:
        pass
    return delivery


async def _observe_scheduled_task_lifecycle(
    delivery: _ScheduledDelivery,
    message: Any,
) -> bool:
    """Handle task lifecycle frames owned by an autonomous Cron turn."""
    session_id = delivery.key[0]
    broadcast = delivery.broadcast
    pending = delivery.pending_tasks
    if isinstance(message, TaskStartedMessage):
        task_id = str(getattr(message, "task_id", "") or "")
        description = getattr(message, "description", None)
        accepted = bool(task_id) and await _record_background_task_launch_owned(
            session_id,
            task_id,
            tool_use_id=getattr(message, "tool_use_id", None),
            description=description,
        )
        if accepted:
            pending[task_id] = {
                "tool_use_id": getattr(message, "tool_use_id", None),
                "description": description,
            }
            _pin_background_task(session_id, task_id)
            broadcast.publish({"event": "task_started", "data": json.dumps({
                "task_id": task_id,
                "tool_use_id": getattr(message, "tool_use_id", None),
                "description": description,
                "task_type": getattr(message, "task_type", None),
            })})
        return True
    if isinstance(message, TaskProgressMessage):
        broadcast.publish({"event": "task_progress", "data": json.dumps({
            "task_id": getattr(message, "task_id", "") or "",
            "tool_use_id": getattr(message, "tool_use_id", None),
            "last_tool_name": getattr(message, "last_tool_name", None),
            "usage": dict(getattr(message, "usage", None) or {}),
        })})
        return True

    terminal: dict[str, Any] | None = None
    if isinstance(message, TaskNotificationMessage):
        terminal = {
            "task_id": getattr(message, "task_id", "") or "",
            "tool_use_id": getattr(message, "tool_use_id", None),
            "status": getattr(message, "status", None),
            "summary": getattr(message, "summary", None),
            "output_file": getattr(message, "output_file", None),
            "usage": dict(getattr(message, "usage", None) or {}),
        }
    elif isinstance(message, TaskUpdatedMessage):
        terminal = _terminal_task_update(message)
    if terminal is None:
        return False
    task_id = str(terminal.get("task_id") or "")
    outcome = await _on_task_settled_owned(
        session_id,
        task_id,
        status=terminal.get("status") or None,
        tool_use_id=terminal.get("tool_use_id") or None,
        summary=terminal.get("summary") or None,
        output_file=terminal.get("output_file") or None,
        usage=dict(terminal.get("usage") or {}),
    )
    if outcome is not None:
        pending.pop(task_id, None)
    if outcome:
        terminal["background_tasks_pending"] = len(pending)
        broadcast.publish({
            "event": "task_notification",
            "data": json.dumps(terminal),
        })
    return True


async def _refresh_scheduled_session_summary(
    session_id: str,
    model: str,
) -> None:
    try:
        messages = await asyncio.to_thread(_get_session_msgs, session_id, model)
        await obs.to_thread_io(
            "chat.scheduled_session_index_write",
            session_id,
            sess.bump_session,
            session_id,
            message_count=len(messages),
            turn_count=sum(1 for msg in messages if _is_real_user_prompt(msg)),
            file_path=sess.INDEX,
            owned=True,
        )
    except Exception as exc:
        sys.stderr.write(
            f"[chat] scheduled session index refresh failed "
            f"sid={session_id[:8]} exc={type(exc).__name__}\n"
        )


async def _finish_scheduled_delivery(
    delivery: _ScheduledDelivery,
    result: ResultMessage,
) -> None:
    key = delivery.key
    session_id = key[0]
    broadcast = delivery.broadcast
    if _sdk_scheduled_deliveries.get(key) is delivery:
        _sdk_scheduled_deliveries.pop(key, None)
    registration = delivery.registration_task

    # Some compatible providers expose final prose only on ResultMessage.
    result_text = str(getattr(result, "result", None) or "")
    if result_text and not "".join(delivery.render_state["streamed"]).strip():
        broadcast.publish({
            "event": "text", "data": json.dumps({"text": result_text}),
        })

    origin = sdk_lifecycle.normalize_origin(getattr(result, "origin", None))
    terminal_reason = sdk_lifecycle.normalize_terminal_reason(
        getattr(result, "terminal_reason", None))
    status = sdk_lifecycle.terminal_status(
        terminal_reason,
        is_error=bool(getattr(result, "is_error", False)),
        cancelled=bool(broadcast.cancelled),
    )
    completed_at_ms = int(time.time() * 1000)
    assistant_uuid = str(delivery.render_state.get("assistant_uuid") or "")
    elapsed_s = round(max(0.0, time.time() - broadcast.started_at), 1)
    if assistant_uuid and not sess.session_is_deleting(session_id):
        try:
            await obs.to_thread_io(
                "chat.scheduled_footer_write",
                session_id,
                sess.set_message_annotation,
                session_id,
                assistant_uuid,
                model=broadcast.model,
                ts=completed_at_ms,
                turn_status=status,
                turn_id=broadcast.turn_id,
                elapsed_s=elapsed_s if elapsed_s >= 1 else None,
                file_path=sess._sidecar_path(session_id),
            )
        except Exception as exc:
            sys.stderr.write(
                f"[chat] scheduled footer annotation failed "
                f"sid={session_id[:8]} exc={type(exc).__name__}\n"
            )

    errors = [str(item) for item in (getattr(result, "errors", None) or [])]
    done_payload: dict[str, Any] = {
        "cancelled": status == "cancelled",
        "is_error": bool(getattr(result, "is_error", False)),
        "error": "\n".join(errors) or (result_text if status == "failed" else ""),
        "model": broadcast.model,
        "scheduled": True,
        "continuation": False,
        "activity_source": broadcast.activity_source,
        "duration_ms": getattr(result, "duration_ms", None),
        "assistant_uuid": assistant_uuid,
        "completed_at_ms": completed_at_ms,
        "status": status,
        "terminal_reason": terminal_reason,
        "origin": origin,
        "model_usage": sdk_lifecycle.normalize_model_usage(
            getattr(result, "model_usage", None)),
        "background_tasks_pending": len(delivery.pending_tasks),
    }
    broadcast.perf_status = status
    broadcast.perf_error_kind = "scheduled_turn" if status == "failed" else "none"
    broadcast.publish({"event": "done", "data": json.dumps(done_payload)})
    await _finish_activity(session_id, broadcast, status)
    broadcast.finish()
    registered_here = False
    async with _lock:
        if _active_turns.get(session_id) is broadcast:
            _active_turns.pop(session_id, None)
            registered_here = True
    # If an earlier user turn still owns the visible slot, its own completion
    # must be allowed to advance first. The retained registration owner will
    # publish this already-finished Cron replay as soon as that slot clears.
    # Avoid awaiting it here: the earlier turn's Result may itself be queued
    # behind this scheduled Result on the same sole stream.
    registration_finished = registration is None or registration.done()
    if (registered_here or registration_finished) \
            and not sess.session_is_deleting(session_id):
        _remember_recent_turn(session_id, broadcast)

    if delivery.pending_tasks:
        client = _clients.get(key)
        if client is not None:
            _spawn_task_watcher(
                session_id,
                client,
                delivery.pending_tasks,
                started_at=broadcast.started_at,
                origin_turn_id=broadcast.turn_id,
            )
    refresh = asyncio.create_task(
        _refresh_scheduled_session_summary(session_id, broadcast.model))
    _retain_maintenance_task(refresh)


async def _observe_sdk_scheduled_delivery(
    key: chat_runtime.ClientKey,
    message: Any,
) -> bool:
    delivery = _sdk_scheduled_deliveries.get(key)
    if delivery is None:
        if not _is_sdk_scheduled_trigger(message):
            return False
        delivery = await _begin_scheduled_delivery(key, message)
        return True

    if await _observe_scheduled_task_lifecycle(delivery, message):
        return True
    if isinstance(message, ResultMessage):
        await _finish_scheduled_delivery(delivery, message)
        return True
    if chat_subagents.is_subagent_message(message):
        for record in delivery.subagent_mux.feed(message):
            delivery.broadcast.publish({
                "event": record["event"],
                "data": json.dumps(record["data"]),
            })
        return True
    for event in _render_continuation_message(message, delivery.render_state):
        delivery.broadcast.publish(event)
    return True


def _on_sdk_runtime_disconnected(session_id: str) -> None:
    """Drop runtime-owned schedules and close any interrupted Cron delivery."""
    with _sdk_cron_state_lock:
        _sdk_cron_jobs.pop(session_id, None)
        for key in [key for key in _sdk_cron_tool_calls if key[0] == session_id]:
            _sdk_cron_tool_calls.pop(key, None)
    for key in [key for key in _sdk_scheduled_deliveries if key[0] == session_id]:
        delivery = _sdk_scheduled_deliveries.pop(key)
        broadcast = delivery.broadcast
        if delivery.registration_task is not None:
            delivery.registration_task.cancel()
        broadcast.cancelled = True
        if not broadcast.done:
            broadcast.publish({
                "event": "error",
                "data": json.dumps({
                    "error": "定时任务运行环境已断开",
                    "kind": "sdk",
                    "retryable": False,
                    "activity_source": "scheduled",
                }),
            })
            broadcast.finish()
        if _active_turns.get(session_id) is broadcast:
            _active_turns.pop(session_id, None)
        try:
            _remember_recent_turn(session_id, broadcast)
        except RuntimeError:
            broadcast.close()
        if broadcast.activity_started:
            try:
                task = asyncio.create_task(
                    _finish_activity(session_id, broadcast, "cancelled"))
                _retain_maintenance_task(task)
            except RuntimeError:
                pass
    # A foreground turn may be the only live carrier for the dot removal.
    for key in [key for key in _clients if key[0] == session_id]:
        _publish_sdk_scheduled_state(key)


async def _observe_sdk_stream_message(
    key: chat_runtime.ClientKey,
    message: Any,
) -> bool:
    """Observe safe lifecycle state and consume autonomous scheduled turns."""
    _observe_sdk_cron_message(key, message)
    consumed = await _observe_sdk_scheduled_delivery(key, message)
    if not hook_traces.is_hook_message(message):
        return consumed
    session_id = key[0]
    broadcast = _scheduled_state_carrier(key) or _active_turns.get(session_id)
    live = broadcast is not None and not broadcast.done
    turn_id = (
        str(broadcast.turn_id or "")
        if live
        else str(_background_origin_turn_id.get(session_id, "") or "")
    )
    origin = (
        "background"
        if (live and broadcast.activity_source in {"background", "scheduled"})
        or (not live and session_id in _sessions_with_inflight_tasks)
        else "foreground"
    )
    trace = await obs.to_thread_io(
        "chat.hook_trace_write",
        session_id,
        hook_traces.observe,
        session_id,
        message,
        turn_id=turn_id,
        origin=origin,
    )
    if trace is not None and live:
        broadcast.publish({
            "event": "hook_trace",
            "data": json.dumps(trace),
        })
    return consumed


chat_runtime.configure_hooks(chat_runtime.RuntimeHooks(
    sessions=sess,
    normalize_effort=lambda *a, **k: _normalize_effort(*a, **k),
    valid_efforts=_VALID_EFFORT,
    valid_service_tiers=_VALID_SERVICE_TIERS,
    normalize_plan_return_permission=lambda *a, **k: _normalize_plan_return_permission(*a, **k),
    build_and_connect_client=lambda *a, **k: _build_and_connect_client(*a, **k),
    has_enabled_external_mcp=lambda: _has_enabled_external_mcp(),
    await_mcp_ready=lambda *a, **k: _await_mcp_ready(*a, **k),
    active_turns=_active_turns,
    sessions_with_inflight_tasks=_sessions_with_inflight_tasks,
    session_has_live_watcher=lambda *a, **k: _session_has_live_watcher(*a, **k),
    session_has_scheduled_tasks=lambda *a, **k: _session_has_scheduled_tasks(*a, **k),
    session_runtime_disconnected=lambda *a, **k: _on_sdk_runtime_disconnected(*a, **k),
    pending_runtime_rebuilds=_pending_runtime_rebuilds,
    client_pool_cap=lambda: _CLIENT_POOL_CAP,
    disconnect_unpooled_client=lambda *a, **k: _disconnect_unpooled_client(*a, **k),
    disconnect_client=lambda *a, **k: disconnect_client(*a, **k),
    get_client=lambda *a, **k: get_client(*a, **k),
    ensure_session_stream=lambda *a, **k: _ensure_session_stream(*a, **k),
    join_session_disconnects=lambda *a, **k: _join_session_disconnects(*a, **k),
    evict_failed_session_stream=lambda *a, **k: _evict_failed_session_stream(*a, **k),
    retain_detached_cleanup=lambda *a, **k: _retain_detached_cleanup(*a, **k),
    observe_stream_message=lambda *a, **k: _observe_sdk_stream_message(*a, **k),
))


# Dynamic runtime bridge for transcript-fork/successor lifecycle. Every callback
# resolves the chat facade at call time, preserving the historical monkeypatch
# surface while keeping the extracted module independent of chat and the SDK.
from .activity import activity as _successor_activity  # noqa: E402

chat_successor.configure_hooks(chat_successor.SuccessorHooks(
    sessions=sess,
    activity=_successor_activity,
    model_default=MODEL,
    root=ROOT,
    is_chinese_locale=lambda: is_chinese_locale(),
    normalize_effort=lambda *a, **k: _normalize_effort(*a, **k),
    validate_permission=lambda *a, **k: _validate_permission(*a, **k),
    normalize_plan_return_permission=lambda *a, **k: _normalize_plan_return_permission(*a, **k),
    session_config_dir=lambda *a, **k: _session_config_dir(*a, **k),
    sdk_fork_session=lambda *a, **k: sdk_fork_session(*a, **k),
    sdk_delete_session=lambda *a, **k: sdk_delete_session(*a, **k),
    sdk_rename_session=lambda *a, **k: sdk_rename_session(*a, **k),
    find_session_jsonl=lambda *a, **k: _find_session_jsonl(*a, **k),
    jsonl_path_cache=_JSONL_PATH_CACHE,
    purge_single_session_storage=lambda *a, **k: _purge_single_session_storage(*a, **k),
    copy_runtime_continuation_snapshots=lambda *a, **k: _copy_runtime_continuation_snapshots(*a, **k),
    shaped_ui_messages=lambda *a, **k: _shaped_ui_messages(*a, **k),
    parse_bg_launch=lambda *a, **k: _parse_bg_launch(*a, **k),
    record_background_task_launch=lambda *a, **k: _record_background_task_launch(*a, **k),
    sessions_with_inflight_tasks=_sessions_with_inflight_tasks,
    bg_task_tool_use_ids=_bg_task_tool_use_ids,
    bg_task_descriptions=_bg_task_descriptions,
    active_turns=_active_turns,
    session_has_live_watcher=lambda *a, **k: _session_has_live_watcher(*a, **k),
    schedule_queue_drain=lambda *a, **k: _schedule_queue_drain(*a, **k),
    get_client=lambda *a, **k: get_client(*a, **k),
    session_runtime_lock_for=lambda *a, **k: _session_runtime_lock_for(*a, **k),
    maintenance_tasks=_maintenance_tasks,
    runtime_fork_uuid_mapping=lambda *a, **k: _runtime_fork_uuid_mapping(*a, **k),
    sync_runtime_successor_postlude=lambda *a, **k: _sync_runtime_successor_postlude(*a, **k),
    runtime_fork_boundary=lambda *a, **k: _runtime_fork_boundary(*a, **k),
    backfill_runtime_task_overlays=lambda *a, **k: _backfill_runtime_task_overlays(*a, **k),
    commit_fork_lifecycle=lambda *a, **k: _commit_fork_lifecycle(*a, **k),
    continue_detached_runtime_locked=lambda *a, **k: _continue_detached_runtime_locked(*a, **k),
    continue_detached_runtime=lambda *a, **k: _continue_detached_runtime(*a, **k),
    prepare_detached_successor_runtime=lambda *a, **k: _prepare_detached_successor_runtime(*a, **k),
    runtime_rollover_lock_for=lambda *a, **k: _runtime_rollover_lock_for(*a, **k),
    session_title_lock=lambda *a, **k: _session_title_lock(*a, **k),
))


# Dynamic runtime bridge for display-only overlays.  Every callback resolves the
# chat facade at call time, preserving the long-standing monkeypatch surface.
chat_overlays.configure_hooks(chat_overlays.OverlayHooks(
    broadcast_to_ui_messages=lambda *a, **k: _broadcast_to_ui_messages(*a, **k),
    ensure_transcript_index=lambda *a, **k: _ensure_transcript_index(*a, **k),
    turn_transcript_boundary=lambda *a, **k: _turn_transcript_boundary(*a, **k),
    transcript_ts_ms=lambda *a, **k: _transcript_ts_ms(*a, **k),
    classify_stream_error=lambda *a, **k: _classify_stream_error(*a, **k),
    indexed_ui_records=lambda *a, **k: _indexed_ui_records(*a, **k),
    turn_uuids_from_boundary=lambda *a, **k: _turn_uuids_from_boundary(*a, **k),
    delete_active_turn_sidecar=lambda *a, **k: _delete_active_turn_sidecar(*a, **k),
    turn_broadcast_factory=lambda *a, **k: TurnBroadcast(*a, **k),
    resolve_staged_attachment_display=lambda *a, **k: (
        _resolve_staged_attachment_display(*a, **k)),
    interrupted_at_startup=_interrupted_at_startup,
    persist_failed_turn_snapshot=lambda *a, **k: _persist_failed_turn_snapshot(*a, **k),
    load_cancelled_turn_snapshots=lambda *a, **k: _load_cancelled_turn_snapshots(*a, **k),
    cancelled_snapshot_canonical_span=lambda *a, **k: _cancelled_snapshot_canonical_span(*a, **k),
    persist_runtime_continuation_snapshot=lambda *a, **k: _persist_runtime_continuation_snapshot(*a, **k),
    runtime_rollover_lock_for=lambda *a, **k: _runtime_rollover_lock_for(*a, **k),
    session_runtime_lock_for=lambda *a, **k: _session_runtime_lock_for(*a, **k),
    session_has_live_watcher=lambda *a, **k: _session_has_live_watcher(*a, **k),
    schedule_queue_drain=lambda *a, **k: _schedule_queue_drain(*a, **k),
    runtime_prewarm_tasks=_runtime_prewarm_tasks,
    active_turns=_active_turns,
    maintenance_tasks=_maintenance_tasks,
))
